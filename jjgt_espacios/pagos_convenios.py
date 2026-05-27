# ══════════════════════════════════════════════════════════════════════════════
#  SUITE SALITRE · Espacios de Descanso Personal — Terminal de Transportes
#  MÓDULO DE PAGOS + CONVENIOS · Kiosco Táctil 24/7
# ══════════════════════════════════════════════════════════════════════════════
#
#  Instalación:
#    pip install streamlit pandas qrcode pillow reportlab requests
#               pytz openpyxl gspread google-auth google-auth-oauthlib
#
#  ── EJECUCIÓN LOCAL ────────────────────────────────────────────────────────
#    1. Coloca credentials.json de tu Service Account en la carpeta del proyecto
#    2. Crea .streamlit/secrets.toml con:
#
#       [sheetsemp]
#       credentials_sheet = '''{ ... pega aquí el JSON de credentials.json ... }'''
#       spreadsheet_id = "ID_DEL_SPREADSHEET_jjgt_pagos"
#
#    3. Ejecutar:  streamlit run pagos.py
#
#  ── EJECUCIÓN EN LA WEB (Streamlit Cloud) ─────────────────────────────────
#    1. Sube el repositorio a GitHub (sin credentials.json ni secrets.toml)
#    2. En Streamlit Cloud → App Settings → Secrets, pega:
#
#       [sheetsemp]
#       credentials_sheet = '''{ ... JSON de Service Account ... }'''
#       spreadsheet_id = "ID_DEL_SPREADSHEET_jjgt_pagos"
#
#    3. La BD SQLite se recrea en cada deploy (usar Google Sheets como fuente)
#
#  ── NOTAS ─────────────────────────────────────────────────────────────────
#    · La app SIEMPRE inicia en la pantalla de login de operador
#    · Cada nuevo registro (cliente, reserva, pago, factura) se escribe en
#      jjgt_pagos de Google Sheets EN EL MISMO MOMENTO que en SQLite,
#      con reintentos automáticos ante errores 429 (MAX_RETRIES=4, backoff exp.)
#    · Backup automático al cierre de turno del operador en carpeta backups/
#    · Convenios: reservas solo en jjgt_convenios, Cubiculos_Estado en ambos
#    · Config de convenios por empresa en Configuracion_Pagos de jjgt_convenios
# ══════════════════════════════════════════════════════════════════════════════

import streamlit as st
# sqlite3 importado solo como fallback en get_db() (stub de compatibilidad)
import json
import hashlib
import io
import base64
import os
import time
import threading
import queue
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import pytz

def fmt_cop(valor) -> str:
    """Formatea un número como pesos colombianos: 1234567 → '$ 1.234.567'"""
    try:
        valor = int(round(float(valor)))
        return f"$ {valor:,}".replace(",", ".")
    except (TypeError, ValueError):
        return "$ 0"

def fmt_tiempo(mins) -> str:
    """Formatea minutos restantes como string legible, ej: 45m o 1h 30m."""
    if mins is None:
        return "--"
    mins = int(mins)
    if mins <= 0:
        return "0m"
    if mins < 60:
        return f"{mins}m"
    horas, minutos = divmod(mins, 60)
    return f"{horas}h {minutos}m" if minutos else f"{horas}h"

# ── Logo de la empresa ────────────────────────────────────────────────────────
def _cargar_logo_b64() -> str:
    """
    Busca el logo, lo redimensiona a máx 300×120 px y lo comprime a JPEG
    de calidad 72 para que pese < 50 KB en base64 (evita WebSocketClosedError
    por payloads gigantes en Tornado/Streamlit).

    Rutas buscadas (en orden):
      1. logoSuite_Salitre.jpg junto al script
      2. assets/logoSuite_Salitre.jpg
      3. static/logoSuite_Salitre.jpg
    """
    _logo_names = [
        "logoSuite_Salitre.jpg",
        "assets/logoSuite_Salitre.jpg",
        "static/logoSuite_Salitre.jpg",
    ]
    _script_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else "."
    for _name in _logo_names:
        _path = os.path.join(_script_dir, _name)
        if os.path.exists(_path):
            try:
                # Intentar comprimir con Pillow (disponible si qrcode o reportlab está instalado)
                from PIL import Image as _PilImg
                _PilImg.MAX_IMAGE_PIXELS = None  # imagen muy grande — suprimir DecompressionBombWarning
                with _PilImg.open(_path) as _img:
                    # Convertir a RGB por si tiene canal alfa (PNG con transparencia)
                    if _img.mode in ("RGBA", "P", "LA"):
                        _bg = _PilImg.new("RGB", _img.size, (5, 11, 26))  # fondo = --bg-deep
                        _bg.paste(_img, mask=_img.split()[-1] if _img.mode in ("RGBA","LA") else None)
                        _img = _bg
                    elif _img.mode != "RGB":
                        _img = _img.convert("RGB")
                    # Redimensionar manteniendo proporción: máx 300×120 px
                    _img.thumbnail((300, 120), _PilImg.LANCZOS)
                    _buf = io.BytesIO()
                    _img.save(_buf, format="JPEG", quality=72, optimize=True)
                    return base64.b64encode(_buf.getvalue()).decode()
            except ImportError:
                # Pillow no disponible: leer raw (solo si ya es pequeño, < 200 KB)
                try:
                    _raw = open(_path, "rb").read()
                    if len(_raw) <= 200_000:
                        return base64.b64encode(_raw).decode()
                    # Archivo demasiado grande sin Pillow → no incluir (evitar crash WS)
                except Exception:
                    pass
            except Exception:
                pass
    return ""

_LOGO_B64: str = _cargar_logo_b64()
_LOGO_HTML: str = (
    f'<img src="data:image/jpeg;base64,{_LOGO_B64}" '
    'style="max-height: 250px;max-width:250px;object-fit:contain;'
    'filter:drop-shadow(0 0 12px rgba(0,212,255,0.35));margin-bottom:4px">'
    if _LOGO_B64 else ""
)

# ── Reintentos para escrituras en Google Sheets ───────────────────────────────
MAX_RETRIES         = 4          # intentos ante error 429
INITIAL_RETRY_DELAY = 2         # segundos base del backoff exponencial

# Imports opcionales — no bloquean si no están instalados
try:
    import qrcode
    from PIL import Image
    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
# [migrated to PG]     import gspread
# [migrated to PG]     from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

try:
# [migrated to PG]     import toml as toml_lib
    TOML_AVAILABLE = True
except ImportError:
    TOML_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Módulo de Facturación Mensual y Control de Cartera ────────────────────────
try:
    import facturacion_cartera as _fc_mod
    FC_AVAILABLE = True
except ImportError:
    _fc_mod      = None
    FC_AVAILABLE = False

try:
    import contabilidad as _cont_mod
    #_cont_mod.set_context(globals())
    CONT_AVAILABLE = True
except ImportError:
    _cont_mod = None
    CONT_AVAILABLE = False

try:
    import factura_electronica_dian as _fe_mod
    #_fe_mod.set_context(globals())
    FE_AVAILABLE = True
except ImportError:
    _fe_mod      = None
    FE_AVAILABLE = False

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES GLOBALES
# ──────────────────────────────────────────────────────────────────────────────
# En Streamlit Cloud, /tmp/ persiste durante la sesión pero no entre deploys.
# Para producción real usar una BD externa (Supabase, Railway, etc.)
_IS_CLOUD = bool(
    os.environ.get("STREAMLIT_SHARING_MODE") or          # señal clásica
    os.environ.get("STREAMLIT_SERVER_HEADLESS") or       # Cloud moderno
    os.environ.get("STCLOUD") or                         # alias manual
    os.environ.get("HOME", "").startswith("/home/appuser") or  # versiones antiguas
    os.environ.get("HOME", "").startswith("/home/user") or     # versiones recientes
    not os.path.exists(".")                              # fallback: sin escritura local
)
DB_PATH = "/tmp/terminal_descanso.db" if _IS_CLOUD else "terminal_descanso.db"
NEGOCIO      = "SUITE SALITRE · Espacios de Descanso"
TAGLINE      = "Tu espacio de descanso en la terminal"
DIRECCION    = "Terminal de Transportes · Modulo 3 Local 230"
TELEFONO     = "3219714969"
NIT          = "902.047.871-3"
TZ_COL       = pytz.timezone("America/Bogota")
NEQUI_NUM    = "3219714969"
DAVIPLATA_NUM= "3219714969"
CUENTA_BANCO = "Banco Caja Social · Cta Ahorros · 123-456789-12"
MP_LINK      = "https://mpago.la/XXXXXXX"
WHATSAPP_OP  = "573219714969"
DRIVE_FILE   = "jjgt_pagos"
DRIVE_FILE_CONVENIOS = "jjgt_convenios"
EMAIL        = "suitesalitre@gmail.com"

# ID del Spreadsheet de Convenios (se configura igual que el principal)
SPREADSHEET_ID_CONVENIOS = ""  # Se llena desde Configuracion_Pagos clave "convenios_spreadsheet_id"

METODOS_PAGO_CONVENIO = ["Convenio"]  # Método exclusivo para convenios (pago queda pendiente)

ESTADOS_CUBICULO = {
    "libre":        {"label": "LIBRE",        "color": "#00ff88", "bg": "rgba(0,255,136,0.12)"},
    "ocupado":      {"label": "OCUPADO",      "color": "#ff4757", "bg": "rgba(255,71,87,0.12)"},
    "por_liberar":  {"label": "POR LIBERAR",  "color": "#ffd32a", "bg": "rgba(255,211,42,0.12)"},
    "mantenimiento":{"label": "MANTENIM.",    "color": "#a29bfe", "bg": "rgba(162,155,254,0.12)"},
    "reservado":    {"label": "RESERVADO",    "color": "#74b9ff", "bg": "rgba(116,185,255,0.12)"},
}

METODOS_PAGO = ["Nequi", "Daviplata", "Efectivo", "PSE", "MercadoPago", "Transferencia", "Tarjeta", "Convenio"]
IVA_PCT      = 0.0

# ──────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title=f"{NEGOCIO} · Pagos",
    page_icon="💤",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Ocultar elementos de la interfaz de Streamlit usando CSS personalizado
hide_streamlit_style = """
            <style>
            #MainMenu { display: none !important; }
            footer    { display: none !important; }
            .stDeployButton { display: none !important; }
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────────────────
# CSS — MODO KIOSCO TÁCTIL
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Google Fonts ─────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=Inconsolata:wght@400;600;700&display=swap');

/* ── Variables ────────────────────────────────────────────── */
:root {
  --bg-deep:    #050b1a;
  --bg-card:    #0d1f3c;
  --bg-card2:   #0a1628;
  --cyan:       #00d4ff;
  --green:      #00ff88;
  --red:        #ff4757;
  --yellow:     #ffd32a;
  --purple:     #a29bfe;
  --text:       #e2e8f0;
  --text-dim:   #94a3b8;
  --border:     rgba(0,212,255,0.2);
  --radius:     16px;
}

/* ── Base ─────────────────────────────────────────────────── */
html, body, .stApp {
  background: linear-gradient(135deg, var(--bg-deep) 0%, #091428 50%, var(--bg-deep) 100%) !important;
  color: var(--text) !important;
  font-family: 'Syne', sans-serif !important;
}

/* Estrellas animadas */
.stApp::before {
  content: '';
  position: fixed; top: 0; left: 0; width: 100%; height: 100%;
  background-image:
    radial-gradient(1px 1px at 10% 20%, rgba(0,212,255,0.6) 0%, transparent 100%),
    radial-gradient(1px 1px at 30% 60%, rgba(0,255,136,0.4) 0%, transparent 100%),
    radial-gradient(1px 1px at 60% 10%, rgba(0,212,255,0.5) 0%, transparent 100%),
    radial-gradient(1px 1px at 80% 40%, rgba(255,255,255,0.3) 0%, transparent 100%),
    radial-gradient(1px 1px at 45% 80%, rgba(0,212,255,0.4) 0%, transparent 100%),
    radial-gradient(2px 2px at 20% 90%, rgba(0,255,136,0.3) 0%, transparent 100%),
    radial-gradient(1px 1px at 70% 70%, rgba(162,155,254,0.4) 0%, transparent 100%),
    radial-gradient(1px 1px at 90% 85%, rgba(0,212,255,0.3) 0%, transparent 100%);
  animation: twinkle 8s ease-in-out infinite alternate;
  pointer-events: none; z-index: 0;
}
@keyframes twinkle {
  0%   { opacity: 0.4; }
  50%  { opacity: 1;   }
  100% { opacity: 0.6; }
}

/* ── Ocultar UI innecesaria de Streamlit ──────────────────── */
#MainMenu { display: none !important; }
footer    { display: none !important; }
.stDeployButton { display: none !important; }

/* ── Header: transparente y sin altura para no ocupar espacio ─ */
/* NO usar visibility:hidden — se hereda a hijos (incluye botón sidebar) */
header[data-testid="stHeader"] {
  background: transparent !important;
  border-bottom: none !important;
  box-shadow: none !important;
}

/* ── Botón sidebar: siempre visible y accesible ────────────── */
[data-testid="collapsedControl"] {
  display: flex !important;
  visibility: visible !important;
  opacity: 1 !important;
  pointer-events: auto !important;
  z-index: 99999 !important;
}
[data-testid="collapsedControl"] svg {
  display: block !important;
  visibility: visible !important;
  opacity: 1 !important;
}

/* ── Sidebar operador ─────────────────────────────────────── */
section[data-testid="stSidebar"] {
  background: rgba(5,11,26,0.97) !important;
  border-right: 1px solid var(--border);
}

/* ── Sidebar — textos visibles sobre fondo oscuro ─────────── */
/* Fuerza todos los textos del sidebar a color claro */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] div,
section[data-testid="stSidebar"] small,
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
  color: #e2e8f0 !important;
}
/* Radio — opciones del menú */
section[data-testid="stSidebar"] .stRadio label,
section[data-testid="stSidebar"] .stRadio p,
section[data-testid="stSidebar"] .stRadio span,
section[data-testid="stSidebar"] [role="radiogroup"] label,
section[data-testid="stSidebar"] [role="radiogroup"] p {
  color: #e2e8f0 !important;
  font-size: 15px !important;
  font-weight: 600 !important;
  letter-spacing: 0.3px !important;
}
/* Radio — opción seleccionada resaltada en cyan */
section[data-testid="stSidebar"] .stRadio [aria-checked="true"] ~ div p,
section[data-testid="stSidebar"] .stRadio [aria-checked="true"] ~ div span,
section[data-testid="stSidebar"] [role="radiogroup"] [data-testid*="radioOption"]:has(input:checked) p,
section[data-testid="stSidebar"] [role="radiogroup"] [data-testid*="radioOption"]:has(input:checked) span {
  color: var(--cyan) !important;
}
/* Radio — bullet activo */
section[data-testid="stSidebar"] .stRadio [aria-checked="true"],
section[data-testid="stSidebar"] [role="radiogroup"] input[type="radio"]:checked {
  accent-color: var(--cyan) !important;
  border-color: var(--cyan) !important;
}
/* Caption / texto gris secundario */
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p,
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] span,
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] small {
  color: #94a3b8 !important;
  font-size: 12px !important;
}
/* Botones en el sidebar */
section[data-testid="stSidebar"] div.stButton > button {
  color: #e2e8f0 !important;
  border-color: rgba(226,232,240,0.25) !important;
  min-height: 52px !important;
  font-size: 15px !important;
}
section[data-testid="stSidebar"] div.stButton > button:hover {
  color: var(--cyan) !important;
  border-color: var(--cyan) !important;
  box-shadow: 0 0 16px rgba(0,212,255,0.25) !important;
}
/* Divider sidebar */
section[data-testid="stSidebar"] hr {
  border-color: rgba(226,232,240,0.12) !important;
}
/* Radio container — fondo suave al hover */
section[data-testid="stSidebar"] .stRadio > div > label:hover {
  background: rgba(0,212,255,0.06) !important;
  border-radius: 8px !important;
}

/* ── Todos los contenedores ───────────────────────────────── */
.block-container { padding: 1rem 2rem !important; }

/* ── Botones táctiles grandes ─────────────────────────────── */
div.stButton > button {
  min-height: 72px !important;
  border-radius: var(--radius) !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 700 !important;
  font-size: 18px !important;
  border: 1.5px solid var(--border) !important;
  background: linear-gradient(135deg, var(--bg-card), var(--bg-card2)) !important;
  color: var(--text) !important;
  transition: all 0.25s ease !important;
  text-transform: uppercase;
  letter-spacing: 1px;
}
div.stButton > button:hover {
  border-color: var(--cyan) !important;
  box-shadow: 0 0 24px rgba(0,212,255,0.35) !important;
  transform: translateY(-2px) !important;
  color: var(--cyan) !important;
}
div.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, #0a3d62, #1a5276) !important;
  border-color: var(--cyan) !important;
  color: var(--cyan) !important;
  box-shadow: 0 0 20px rgba(0,212,255,0.2) !important;
  font-size: 20px !important;
  min-height: 88px !important;
}
div.stButton > button[kind="primary"]:hover {
  box-shadow: 0 0 40px rgba(0,212,255,0.55) !important;
  transform: translateY(-3px) !important;
}

/* ── Inputs ───────────────────────────────────────────────── */
.stTextInput input, .stNumberInput input, .stSelectbox select,
.stTextArea textarea {
  background: rgba(13,31,60,0.8) !important;
  border: 1.5px solid var(--border) !important;
  border-radius: 12px !important;
  color: var(--text) !important;
  font-family: 'Syne', sans-serif !important;
  font-size: 18px !important; 
  padding: 14px 18px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
  border-color: var(--cyan) !important;
  box-shadow: 0 0 16px rgba(0,212,255,0.25) !important;
}
label, .stTextInput label, .stNumberInput label, .stSelectbox label {
  font-family: 'Syne', sans-serif !important;
  font-size: 15px !important;
  font-weight: 600 !important;
  color: var(--text-white) !important;
  text-transform: uppercase;
  letter-spacing: 1px;
}

/* ── Tarjetas generales ───────────────────────────────────── */
.card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 24px;
  margin-bottom: 16px;
}

/* ── Tarjeta cubículo ─────────────────────────────────────── */
.cubiculo-card {
  border-radius: var(--radius);
  padding: 20px 16px;
  text-align: center;
  cursor: pointer;
  transition: all 0.25s;
  min-height: 200px;
  display: flex;
  flex-direction: column;
  align-items: center;
  justify-content: center;
  gap: 8px;
  border: 2px solid transparent;
  position: relative;
  overflow: hidden;
}
.cubiculo-libre {
  background: rgba(0,255,136,0.08);
  border-color: rgba(0,255,136,0.4);
}
.cubiculo-libre:hover {
  background: rgba(0,255,136,0.18);
  border-color: #00ff88;
  box-shadow: 0 0 30px rgba(0,255,136,0.4);
  transform: scale(1.03);
}
.cubiculo-ocupado {
  background: rgba(255,71,87,0.08);
  border-color: rgba(255,71,87,0.3);
  opacity: 0.7;
  cursor: not-allowed;
}
.cubiculo-mantenimiento {
  background: rgba(162,155,254,0.08);
  border-color: rgba(162,155,254,0.3);
  opacity: 0.6;
  cursor: not-allowed;
}
.cubiculo-selected {
  border-color: var(--cyan) !important;
  box-shadow: 0 0 40px rgba(0,212,255,0.5) !important;
  background: rgba(0,212,255,0.12) !important;
  animation: pulse-selected 1.5s ease-in-out infinite;
}
@keyframes pulse-selected {
  0%, 100% { box-shadow: 0 0 30px rgba(0,212,255,0.4); }
  50%       { box-shadow: 0 0 60px rgba(0,212,255,0.7); }
}
.cubiculo-num {
  font-family: 'Inconsolata', monospace;
  font-size: 42px;
  font-weight: 700;
  line-height: 1;
}
.estado-badge {
  font-size: 11px;
  font-weight: 700;
  letter-spacing: 2px;
  padding: 4px 12px;
  border-radius: 20px;
  text-transform: uppercase;
}
.timer-display {
  font-family: 'Inconsolata', monospace;
  font-size: 18px;
  font-weight: 600;
}
.servicios-icons {
  font-size: 16px;
  opacity: 0.8;
  margin-top: 4px;
}

/* ── Temporizador grande ──────────────────────────────────── */
.clock-big {
  font-family: 'Inconsolata', monospace;
  font-size: 72px;
  font-weight: 700;
  color: var(--green);
  text-shadow: 0 0 30px rgba(0,255,136,0.6);
  text-align: center;
  line-height: 1;
}
.clock-label {
  font-family: 'Syne', sans-serif;
  font-size: 13px;
  font-weight: 600;
  color: var(--text-dim);
  text-align: center;
  letter-spacing: 3px;
  text-transform: uppercase;
  margin-top: 4px;
}

/* ── QR container ─────────────────────────────────────────── */
.qr-container {
  display: flex; flex-direction: column; align-items: center;
  padding: 24px;
  background: var(--bg-card);
  border-radius: var(--radius);
  border: 2px solid var(--cyan);
  animation: qr-pulse 2s ease-in-out infinite;
  margin: 16px auto;
  max-width: 400px;
}
@keyframes qr-pulse {
  0%, 100% { box-shadow: 0 0 20px rgba(0,212,255,0.3); }
  50%       { box-shadow: 0 0 50px rgba(0,212,255,0.6); }
}
.qr-container img {
  border-radius: 12px;
  margin-bottom: 12px;
}
.qr-instruccion {
  font-size: 15px;
  color: var(--text-dim);
  text-align: center;
  margin-top: 8px;
}

/* ── Método de pago cards ─────────────────────────────────── */
.pago-card {
  border-radius: var(--radius);
  padding: 20px 24px;
  min-height: 120px;
  cursor: pointer;
  transition: all 0.2s;
  border: 2px solid transparent;
  display: flex;
  align-items: center;
  gap: 16px;
  font-size: 20px;
  font-weight: 700;
}
.pago-nequi {
  background: linear-gradient(135deg, rgba(0,100,40,0.3), rgba(0,180,80,0.15));
  border-color: rgba(0,255,100,0.3);
}
.pago-nequi:hover { border-color: #00c853; box-shadow: 0 0 24px rgba(0,200,83,0.4); }
.pago-daviplata {
  background: linear-gradient(135deg, rgba(10,40,100,0.4), rgba(20,70,180,0.2));
  border-color: rgba(50,130,246,0.3);
}
.pago-daviplata:hover { border-color: #3b82f6; box-shadow: 0 0 24px rgba(59,130,246,0.4); }
.pago-efectivo {
  background: linear-gradient(135deg, rgba(20,60,20,0.4), rgba(40,120,40,0.2));
  border-color: rgba(100,200,80,0.3);
}
.pago-efectivo:hover { border-color: #84cc16; box-shadow: 0 0 24px rgba(132,204,22,0.4); }

/* ── Métricas KPI ─────────────────────────────────────────── */
.kpi-card {
  background: var(--bg-card);
  border: 1px solid var(--border);
  border-top: 3px solid var(--cyan);
  border-radius: var(--radius);
  padding: 20px;
  text-align: center;
}
.kpi-value {
  font-family: 'Inconsolata', monospace;
  font-size: 36px;
  font-weight: 700;
  color: var(--cyan);
  line-height: 1;
}
.kpi-label {
  font-size: 12px;
  font-weight: 600;
  color: var(--text-dim);
  letter-spacing: 2px;
  text-transform: uppercase;
  margin-top: 6px;
}

/* ── Voucher ──────────────────────────────────────────────── */
.voucher-box {
  background: linear-gradient(135deg, var(--bg-card), rgba(0,212,255,0.05));
  border: 1.5px solid var(--cyan);
  border-radius: var(--radius);
  padding: 32px;
  max-width: 580px;
  margin: 0 auto;
}
.voucher-codigo {
  font-family: 'Inconsolata', monospace;
  font-size: 80px;
  font-weight: 700;
  color: var(--green);
  text-shadow: 0 0 40px rgba(0,255,136,0.6);
  text-align: center;
  letter-spacing: 16px;
  line-height: 1;
}
.voucher-label {
  font-size: 12px;
  font-weight: 700;
  color: var(--text-dim);
  text-align: center;
  letter-spacing: 3px;
  text-transform: uppercase;
  margin-top: 8px;
}

/* ── Confirmación ─────────────────────────────────────────── */
.confirm-ok {
  background: linear-gradient(135deg, rgba(0,100,50,0.4), rgba(0,200,100,0.15));
  border: 2px solid var(--green);
  border-radius: var(--radius);
  padding: 40px;
  text-align: center;
  animation: confetti-bg 0.5s ease;
}
.confirm-fail {
  background: linear-gradient(135deg, rgba(100,20,20,0.4), rgba(200,50,50,0.15));
  border: 2px solid var(--red);
  border-radius: var(--radius);
  padding: 40px;
  text-align: center;
}

/* ── Header principal ─────────────────────────────────────── */
.main-header {
  text-align: center;
  padding: 24px 0 8px;
}
.main-logo {
  font-family: 'Syne', sans-serif;
  font-size: 48px;
  font-weight: 800;
  background: linear-gradient(90deg, var(--cyan), var(--green));
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
  background-clip: text;
  letter-spacing: -1px;
  line-height: 1;
}
.main-tagline {
  font-size: 16px;
  color: var(--text-dim);
  letter-spacing: 3px;
  text-transform: uppercase;
  margin-top: 6px;
}

/* ── Stepper ──────────────────────────────────────────────── */
.stepper {
  display: flex;
  align-items: center;
  justify-content: center;
  gap: 0;
  margin: 16px 0 24px;
  flex-wrap: wrap;
}
.step {
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 8px 16px;
  border-radius: 24px;
  font-size: 13px;
  font-weight: 700;
  letter-spacing: 1px;
  text-transform: uppercase;
  color: var(--text-dim);
  border: 1.5px solid rgba(148,163,184,0.2);
  background: transparent;
  white-space: nowrap;
}
.step.active {
  color: var(--cyan);
  border-color: var(--cyan);
  background: rgba(0,212,255,0.1);
  box-shadow: 0 0 16px rgba(0,212,255,0.25);
}
.step.done {
  color: var(--green);
  border-color: rgba(0,255,136,0.4);
  background: rgba(0,255,136,0.08);
}
.step-arrow {
  width: 32px;
  height: 2px;
  background: rgba(148,163,184,0.2);
  flex-shrink: 0;
}

/* ── Alerta pulsante ──────────────────────────────────────── */
.alerta-roja {
  animation: pulse-red 1s ease-in-out infinite;
  background: rgba(255,71,87,0.2);
  border: 2px solid var(--red);
  border-radius: 12px;
  padding: 16px;
  text-align: center;
  color: var(--red);
  font-weight: 700;
}
@keyframes pulse-red {
  0%, 100% { opacity: 1;   box-shadow: 0 0 10px rgba(255,71,87,0.4); }
  50%       { opacity: 0.7; box-shadow: 0 0 30px rgba(255,71,87,0.8); }
}

/* ── Live timer elements ──────────────────────────────────── */
.jjgt-timer {
  font-family: 'Inconsolata', monospace !important;
  font-size: 13px;
  font-weight: 700;
  transition: color 0.5s ease;
}
.jjgt-clock-hms {
  font-variant-numeric: tabular-nums;
  letter-spacing: 2px;
}
.jjgt-qr-countdown {
  font-family: 'Inconsolata', monospace;
  font-size: 12px;
  font-weight: 700;
  margin-top: 4px;
  transition: color 0.5s ease;
}
.jjgt-consumed {
  font-family: 'Inconsolata', monospace;
  font-size: 12px;
}

/* ── Countdown bar ────────────────────────────────────────── */
.countdown-bar {
  width: 100%;
  height: 8px;
  border-radius: 4px;
  background: rgba(255,255,255,0.1);
  overflow: hidden;
  margin-top: 8px;
}
.countdown-fill {
  height: 100%;
  border-radius: 4px;
  background: linear-gradient(90deg, var(--green), var(--cyan));
  transition: width 1s linear;
}

/* ── Tabs ─────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
  background: var(--bg-card) !important;
  border-radius: 12px !important;
  gap: 4px !important;
  padding: 4px !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  color: var(--text-dim) !important;
  border-radius: 8px !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 600 !important;
}
.stTabs [aria-selected="true"] {
  background: rgba(0,212,255,0.15) !important;
  color: var(--cyan) !important;
}

/* ── Divider ──────────────────────────────────────────────── */
hr { border-color: var(--border) !important; }

/* ── Dataframe ────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
  border: 1px solid var(--border) !important;
  border-radius: 12px !important;
  overflow: hidden;
}

/* ── Toggle / checkbox ────────────────────────────────────── */
.stCheckbox label {
  font-size: 16px !important;
  font-weight: 600 !important;
  text-transform: none !important;
  letter-spacing: 0 !important;
  color: var(--text) !important;
}

/* ── Metric ───────────────────────────────────────────────── */
[data-testid="stMetric"] {
  background: var(--bg-card) !important;
  border: 1px solid var(--border) !important;
  border-top: 3px solid var(--cyan) !important;
  border-radius: var(--radius) !important;
  padding: 16px !important;
}
[data-testid="stMetricValue"] {
  font-family: 'Inconsolata', monospace !important;
  color: var(--cyan) !important;
}
[data-testid="stMetricDelta"] { font-family: 'Inconsolata', monospace !important; }

/* ── Print ────────────────────────────────────────────────── */
@media print {
  .stApp::before, section[data-testid="stSidebar"], .stButton,
  [data-testid="stToolbar"] { display: none !important; }
  .voucher-box { border: 2px solid black !important; color: black !important; }
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# RELOJ Y TIMERS EN VIVO — JavaScript puro (sin recargar el servidor)
# ══════════════════════════════════════════════════════════════════════════════

def inject_live_clock():
    """
    Inyecta el motor de reloj/timers en vivo usando st.components.v1.html().
    Esto crea un <iframe> persistente que NO es destruido por los reruns de Streamlit.
    El script se comunica con el DOM padre via window.parent para actualizar
    los elementos .jjgt-clock-hms, .jjgt-timer[data-fin], etc.
    """
    import streamlit.components.v1 as components
    components.html("""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;overflow:hidden;background:transparent;">
<script>
(function() {
  var DAYS_ES   = ['Domingo','Lunes','Martes','Miércoles','Jueves','Viernes','Sábado'];
  var MONTHS_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                   'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];

  function pad2(n){ return n < 10 ? '0'+n : ''+n; }

  // Colombia = UTC-5 fijo (sin horario de verano)
  function getColombiaTime() {
    var now = new Date();
    var utc = now.getTime() + now.getTimezoneOffset() * 60000;
    return new Date(utc - 5 * 3600000);
  }

  // Acceder al DOM del padre (la página principal de Streamlit)
  function getParentDoc() {
    try { return window.parent.document; } catch(e) { return document; }
  }

  var alarmaDisparada = {};

  function updateClock() {
    var doc = getParentDoc();
    var t  = getColombiaTime();
    var hh = pad2(t.getHours());
    var mm = pad2(t.getMinutes());
    var ss = pad2(t.getSeconds());
    var hms = hh + ':' + mm + ':' + ss;
    var fecha = DAYS_ES[t.getDay()] + ', ' + pad2(t.getDate()) + ' DE ' +
                MONTHS_ES[t.getMonth()].toUpperCase() + ' DE ' + t.getFullYear();

    doc.querySelectorAll('.jjgt-clock-hms').forEach(function(el) {
      el.textContent = hms;
    });
    doc.querySelectorAll('.jjgt-clock-fecha').forEach(function(el) {
      el.textContent = fecha;
    });
    doc.querySelectorAll('.jjgt-clock-compact').forEach(function(el) {
      el.textContent = hh + ':' + mm;
    });
  }

  function updateTimers() {
    var doc = getParentDoc();
    var nowMs = getColombiaTime().getTime();
    doc.querySelectorAll('.jjgt-timer[data-fin]').forEach(function(el) {
      var fin = el.getAttribute('data-fin');
      if (!fin) return;
      var finMs  = new Date(fin).getTime();
      var diffMs = finMs - nowMs;
      var key    = el.getAttribute('data-key') || fin;

      if (diffMs <= 0) {
        el.textContent = '⏰ TIEMPO VENCIDO';
        el.style.color = '#ff4757';
        el.style.fontWeight = '700';
        el.style.animation = 'none';
        if (!alarmaDisparada[key]) {
          alarmaDisparada[key] = true;
          playAlarm();
          showBanner(doc);
        }
      } else {
        var totalSec = Math.floor(diffMs / 1000);
        var h = Math.floor(totalSec / 3600);
        var m = Math.floor((totalSec % 3600) / 60);
        var s = totalSec % 60;
        var label = (h > 0 ? pad2(h) + ':' : '') + pad2(m) + ':' + pad2(s);
        el.textContent = '⏱ ' + label;
        var diffMin = Math.floor(diffMs / 60000);
        if (diffMin <= 5) {
          el.style.color = '#ff4757';
          el.style.animation = 'pulse-red 1s ease-in-out infinite';
        } else if (diffMin <= 15) {
          el.style.color = '#ffd32a';
          el.style.animation = '';
        } else {
          el.style.color = '#00ff88';
          el.style.animation = '';
        }
      }
    });
  }

  function updateConsumed() {
    var doc = getParentDoc();
    var nowMs = getColombiaTime().getTime();
    doc.querySelectorAll('.jjgt-consumed[data-inicio]').forEach(function(el) {
      var inicio = el.getAttribute('data-inicio');
      if (!inicio) return;
      var diffMs  = nowMs - new Date(inicio).getTime();
      var totalSec = Math.max(0, Math.floor(diffMs / 1000));
      var h = Math.floor(totalSec / 3600);
      var m = Math.floor((totalSec % 3600) / 60);
      var s = totalSec % 60;
      el.textContent = (h > 0 ? h + 'h ' : '') + pad2(m) + ':' + pad2(s) + ' consumidos';
    });
  }

  function updateQrCountdown() {
    var doc = getParentDoc();
    var nowMs = getColombiaTime().getTime();
    doc.querySelectorAll('.jjgt-qr-countdown[data-exp]').forEach(function(el) {
      var exp = el.getAttribute('data-exp');
      if (!exp) return;
      var diffS = Math.floor((new Date(exp).getTime() - nowMs) / 1000);
      if (diffS <= 0) {
        el.textContent = '🔄 QR vencido — recargando...';
        el.style.color = '#ff4757';
      } else {
        el.textContent = '🔒 QR válido por ' + diffS + 's';
        el.style.color = diffS < 30 ? '#ffd32a' : '#00ff88';
      }
    });
  }

  function playAlarm() {
    try {
      var ctx = new (window.AudioContext || window.webkitAudioContext)();
      function beep(f, s, d) {
        var o = ctx.createOscillator(), g = ctx.createGain();
        o.connect(g); g.connect(ctx.destination);
        o.frequency.value = f; o.type = 'square';
        g.gain.setValueAtTime(0.35, ctx.currentTime + s);
        g.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime + s + d);
        o.start(ctx.currentTime + s);
        o.stop(ctx.currentTime + s + d + 0.05);
      }
      for (var i = 0; i < 6; i++) { beep(880, i*0.4, 0.3); beep(660, i*0.4+0.18, 0.15); }
    } catch(e) {}
  }

  function showBanner(doc) {
    try {
      var banner = doc.getElementById('jjgt-alarm-banner');
      if (!banner) {
        banner = doc.createElement('div');
        banner.id = 'jjgt-alarm-banner';
        banner.style.cssText =
          'position:fixed;top:0;left:0;right:0;z-index:99999;' +
          'background:rgba(255,71,87,0.97);color:#fff;font-weight:700;font-size:18px;' +
          'text-align:center;padding:14px;letter-spacing:1px;';
        doc.body.appendChild(banner);
      }
      banner.style.display = 'block';
      banner.textContent = '🔔 ¡TIEMPO VENCIDO! Cubículo listo para liberar — Notificar al operador';
      setTimeout(function() { banner.style.display = 'none'; }, 20000);
    } catch(e) {}
  }

  function tick() {
    updateClock();
    updateTimers();
    updateConsumed();
    updateQrCountdown();
  }

  // Arrancar inmediatamente y cada segundo
  tick();
  setInterval(tick, 1000);
})();
</script>
</body>
</html>
""", height=0, scrolling=False)


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE SHEETS — CREDENCIALES, CONEXIÓN Y ESCRITURA
# Patrón basado en el ejemplo de referencia proporcionado.
# ══════════════════════════════════════════════════════════════════════════════
# CAPA DE DATOS — PostgreSQL (PAGOS + CONVENIOS)
# Reemplaza completamente Google Sheets (jjgt_pagos y jjgt_convenios) como
# fuente primaria. Una sola base de datos PostgreSQL gestiona todo.
#
# Conexión: configura el bloque [postgres] en .streamlit/secrets.toml:
#
#   [postgres]
#   host     = "localhost"
#   port     = 5432
#   user     = "postgres"
#   password = "123456"
#   dbname   = "reservas"
#   sslmode  = "disable"          # "require" para Supabase/Railway
#
#   # O bien usa URL directa:
#   url = "postgresql://user:pass@host:5432/dbname?sslmode=require"
#
# También acepta DATABASE_URL como variable de entorno (formato psycopg2).
# ══════════════════════════════════════════════════════════════════════════════

# ── Dependencias eliminadas (Google Sheets) ───────────────────────────────────
# Las siguientes variables se conservan como stubs para no romper referencias
# residuales en el código de UI.
GSPREAD_AVAILABLE   = False
TOML_AVAILABLE      = False
MAX_RETRIES         = 1
INITIAL_RETRY_DELAY = 0

# ── Conexión PostgreSQL ───────────────────────────────────────────────────────

def _read_pg_secrets():
    import urllib.parse
    import os
    import streamlit as st

    def _build_url(host, port, user, password, dbname, sslmode="require"):
        safe_user = urllib.parse.quote(str(user), safe="")
        safe_pass = urllib.parse.quote(str(password), safe="")
        safe_host = str(host).strip().lstrip("@")
        return (
            f"postgresql://{safe_user}:{safe_pass}"
            f"@{safe_host}:{port}/{dbname}"
            f"?sslmode={sslmode}&connect_timeout=15"
        )

    # 1. Streamlit secrets — URL directa
    try:
        if "postgres" in st.secrets:
            pg = st.secrets["postgres"]
            if "url" in pg:
                url = str(pg["url"]).strip()
                if url:
                    return url, True
    except Exception as e:
        print("Error leyendo st.secrets:", e)

    # 2. DATABASE_URL de entorno
    env_url = os.environ.get("DATABASE_URL", "").strip()
    if env_url:
        return env_url, True

    # 3. Variables separadas en secrets
    try:
        if "postgres" in st.secrets:
            pg = st.secrets["postgres"]
            host = pg.get("host")
            if host:
                return _build_url(
                    host,
                    int(pg.get("port", 5432)),
                    pg.get("user"),
                    pg.get("password"),
                    pg.get("dbname", "postgres"),
                    pg.get("sslmode", "require"),
                ), True
    except Exception as e:
        print("Error leyendo variables separadas:", e)

    # 4. Fallback local
    print("⚠️ Usando fallback localhost")
    return _build_url("localhost", 5433, "postgres", "123456", "reservas", "disable"), False


_pg_conn_url, _PG_IS_REMOTE = _read_pg_secrets()

PG_HOST = "aws-1-us-west-1.pooler.supabase.com"
PG_PORT = 5432
PG_USER = "postgres.rjskwwllgzgmiiqzjudb"
PG_PASS = ""
PG_DB   = "postgres"

try:
    import psycopg2
    import psycopg2.extras
    PSYCOPG2_AVAILABLE = True
except ImportError:
    try:
        import psycopg2cffi as psycopg2          # type: ignore
        import psycopg2cffi.extras as _extras    # type: ignore
        psycopg2.extras = _extras
        PSYCOPG2_AVAILABLE = True
    except ImportError:
        PSYCOPG2_AVAILABLE = False


def _resolve_ipv4(hostname: str) -> str:
    import socket
    try:
        results = socket.getaddrinfo(hostname, None, socket.AF_INET)
        if results:
            return results[0][4][0]
    except Exception:
        pass
    return hostname


@st.cache_resource
def _get_pg_engine():
    """Retorna la URL de conexión cacheada."""
    return {"url": _pg_conn_url}


def get_pg_conn():
    """Retorna una conexión activa, reconectando si es necesario."""
    engine = _get_pg_engine()
    url    = engine.get("url")
    conn   = engine.get("conn")
    if conn is not None:
        try:
            conn.cursor().execute("SELECT 1")
            return conn
        except Exception:
            pass
    new_conn = psycopg2.connect(url)
    engine["conn"] = new_conn
    return new_conn


def _pg_exec(sql: str, params=None, fetch: str = None):
    conn = get_pg_conn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            conn.commit()
            if fetch == "all":
                return [dict(r) for r in cur.fetchall()]
            if fetch == "one":
                row = cur.fetchone()
                return dict(row) if row else None
            return None
    except Exception:
        conn.rollback()
        raise


# ══════════════════════════════════════════════════════════════════════════════
# DEFINICIÓN DE TABLAS (DRIVE_SHEETS + DRIVE_SHEETS_CONVENIOS → PostgreSQL)
# ══════════════════════════════════════════════════════════════════════════════

# Estructura de columnas de jjgt_pagos (tablas estándar)
DRIVE_SHEETS = {
    "Reservas": [
        "ID_Reserva","Numero_Reserva","Creado_En","Cubiculo_Num",
        "Cliente_Nombre","Documento","Telefono","Email",
        "Horas_Contratadas","Hora_Inicio","Hora_Fin_Prog","Hora_Fin_Real",
        "Precio_Hora","Subtotal","IVA","Total_COP",
        "Metodo_Pago","Estado_Pago","Codigo_Acceso",
        "WiFi_SSID","WiFi_Pass","Num_Factura","Referencia_Pago",
        "Operador","Notas","Acepto_Datos",
    ],
    "Pagos": [
        "ID_Pago","ID_Reserva","Num_Reserva","Fecha_Pago",
        "Monto_COP","Metodo","Referencia_Externa","Estado",
        "Confirmado_Por","Notas",
    ],
    "Clientes": [
        "ID_Cliente","Nombre","Tipo_Doc","Num_Doc",
        "Telefono","Email","Ciudad","Regimen","Tipo_Persona",
        "Razon_Social","NIT_Empresa","Activo",
        "Total_Reservas","Total_Gastado","Creado_En",
    ],
    "Cubiculos_Estado": [
        "Cubiculo_ID","Numero","Nombre","Estado",
        "Cliente_Actual","Hora_Inicio","Hora_Fin_Prog",
        "Tiempo_Rest_Min","Codigo_Acceso","WiFi_SSID","WiFi_Pass",
        "Precio_Hora_Base",
    ],
    "Facturas": [
        "ID_Factura","Num_Factura","Tipo","Fecha_Emision","Fecha_Vencimiento",
        "Cliente_Nombre","Cliente_Doc","Cliente_Email","Razon_Social","NIT_Empresa",
        "Descripcion","Subtotal","Descuento","Base_Gravable","IVA","Retenciones",
        "Total_COP","Metodo_Pago","Estado","Moneda",
        "Num_Reserva","Cubiculo","Creado_En","Actualizado_En",
    ],
    "Factura_items": [
        "ID_Item","ID_Factura","Codigo","Descripcion",
        "Cantidad","Unidad","Precio_Unitario","Descuento_Pct","IVA_Pct","Subtotal",
    ],
    "Operadores": [
        "ID_Operador","Nombre","Rol","Turno",
        "Hora_Ini_Turno","Hora_Fin_Turno","Permisos","Activo",
    ],
    "Configuracion_Pagos": ["Clave","Valor"],
    "Tarifas_Config": [
        "ID","Nombre","Descripcion","Precio_Hora_COP",
        "Desc_3h_Pct","Desc_6h_Pct","Hora_Ini_Espec","Hora_Fin_Espec",
        "Aplica_Festivos","Activo","Horas_A_Reservar",
    ],
    "Dashboard_Diario": [
        "Fecha","Total_Reservas","Completadas","Canceladas",
        "Ingresos_Brutos","IVA_Recaudado","Ingresos_Netos",
        "Nequi_COP","Daviplata_COP","Efectivo_COP","PSE_COP",
        "MP_COP","Otros_COP","Ocupacion_Pct","Hora_Pico",
        "Tiempo_Prom_Min","Clientes_Nuevos","Clientes_Recur",
        "Fact_Min","Fact_Max","Ticket_Prom_COP",
    ],
    "Log_Operaciones": [
        "Timestamp","Tipo_Op","Reserva_ID","Cubiculo",
        "Operador","Descripcion","Valor_Ant","Valor_Nuevo",
        "IP","Estado","Notas",
    ],
}

# Estructura de columnas de jjgt_convenios (tablas con campos de empresa)
DRIVE_SHEETS_CONVENIOS = {
    "Empresa": [
        "Id_Empresa","Nombre_Empresa","Nit_Empresa","Tipo_Empresa",
        "Email_Empresa","Telefono_Empresa","Tipo_Convenio",
        "Observaciones","Activo","Creado_En",
    ],
    "Reservas": [
        "ID_Reserva","Numero_Reserva","Creado_En","Cubiculo_Num",
        "Cliente_Nombre","Documento","Telefono","Email",
        "Horas_Contratadas","Hora_Inicio","Hora_Fin_Prog","Hora_Fin_Real",
        "Precio_Hora","Subtotal","IVA","Total_COP",
        "Metodo_Pago","Estado_Pago","Codigo_Acceso",
        "WiFi_SSID","WiFi_Pass","Num_Factura","Referencia_Pago",
        "Operador","Notas","Id_Empresa","Nombre_Empresa","Acepto_Datos",
    ],
    "Clientes": [
        "ID_Cliente","Nombre","Tipo_Doc","Num_Doc","Telefono","Email",
        "Ciudad","Regimen","Tipo_Persona","Razon_Social","NIT_Empresa",
        "Activo","Total_Reservas","Total_Gastado","Creado_En",
        "Id_Empresa","Nombre_Empresa",
    ],
    "Pagos": [
        "ID_Pago","ID_Reserva","Num_Reserva","Fecha_Pago",
        "Monto_COP","Metodo","Referencia_Externa","Estado",
        "Confirmado_Por","Notas","Id_Empresa","Nombre_Empresa",
    ],
    "Facturas": [
        "ID_Factura","Num_Factura","Tipo","Fecha_Emision","Fecha_Vencimiento",
        "Cliente_Nombre","Cliente_Doc","Cliente_Email","Razon_Social","NIT_Empresa",
        "Descripcion","Subtotal","Descuento","Base_Gravable","IVA","Retenciones",
        "Total_COP","Metodo_Pago","Estado","Moneda","Num_Reserva","Cubiculo",
        "Creado_En","Actualizado_En","Id_Empresa","Nombre_Empresa",
    ],
    "Factura_items": [
        "ID_Item","ID_Factura","Codigo","Descripcion","Cantidad","Unidad",
        "Precio_Unitario","Descuento_Pct","IVA_Pct","Subtotal","Id_Empresa",
    ],
    "Dashboard_Diario": [
        "Fecha","Total_Reservas","Completadas","Canceladas",
        "Ingresos_Brutos","IVA_Recaudado","Ingresos_Netos",
        "Nequi_COP","Daviplata_COP","Efectivo_COP","PSE_COP",
        "MP_COP","Otros_COP","Convenio_COP","Ocupacion_Pct","Hora_Pico",
        "Tiempo_Prom_Min","Clientes_Nuevos","Clientes_Recur",
        "Fact_Min","Fact_Max","Ticket_Prom_COP",
    ],
    "Cubiculos_Estado": [
        "Cubiculo_ID","Numero","Nombre","Estado","Cliente_Actual",
        "Hora_Inicio","Hora_Fin_Prog","Tiempo_Rest_Min","Codigo_Acceso",
        "WiFi_SSID","WiFi_Pass","Precio_Hora_Base",
    ],
    "Tarifas_Config": [
        "ID","Id_Empresa","Empresa_Nombre","Descripcion","Precio_Hora_COP",
        "Desc_3h_Pct","Desc_6h_Pct","Hora_Ini_Espec","Hora_Fin_Espec",
        "Aplica_Festivos","Activo",
    ],
    "Operadores": [
        "ID_Operador","Nombre","Rol","Turno",
        "Hora_Ini_Turno","Hora_Fin_Turno","Permisos","Activo",
    ],
    "Configuracion_Pagos": [
        "Id_Empresa","Nombre_Empresa","Clave","Valor","Actualizado_En",
    ],
    "Log_Operaciones": [
        "Timestamp","Tipo_Op","Reserva_ID","Cubiculo",
        "Operador","Descripcion","Valor_Ant","Valor_Nuevo",
        "IP","Estado","Notas",
    ],
    "Comprobantes_Contables": [
        "Fecha","Numero","Tipo","Evento","Tercero","Descripcion",
        "Cuenta_Debito","Nombre_Debito","Cuenta_Credito","Nombre_Credito",
        "Valor_COP","Medio_Pago","Soporte","Operador","Observaciones",
    ],
}

# ── SQL de creación de tablas ─────────────────────────────────────────────────
_CREATE_TABLES_SQL = """
CREATE TABLE IF NOT EXISTS reservas (
    id_reserva          TEXT PRIMARY KEY,
    numero_reserva      TEXT UNIQUE,
    creado_en           TEXT,
    cubiculo_num        TEXT,
    cliente_nombre      TEXT,
    documento           TEXT,
    telefono            TEXT,
    email               TEXT,
    horas_contratadas   TEXT,
    hora_inicio         TEXT,
    hora_fin_prog       TEXT,
    hora_fin_real       TEXT,
    precio_hora         TEXT,
    subtotal            TEXT,
    iva                 TEXT,
    total_cop           TEXT,
    metodo_pago         TEXT,
    estado_pago         TEXT DEFAULT 'confirmado',
    codigo_acceso       TEXT,
    wifi_ssid           TEXT,
    wifi_pass           TEXT,
    num_factura         TEXT,
    referencia_pago     TEXT,
    operador            TEXT,
    notas               TEXT,
    acepto_datos        TEXT
);

CREATE TABLE IF NOT EXISTS pagos (
    id_pago             TEXT PRIMARY KEY,
    id_reserva          TEXT,
    num_reserva         TEXT,
    fecha_pago          TEXT,
    monto_cop           TEXT,
    metodo              TEXT,
    referencia_externa  TEXT,
    estado              TEXT DEFAULT 'confirmado',
    confirmado_por      TEXT,
    notas               TEXT
);

CREATE TABLE IF NOT EXISTS clientes (
    id_cliente          TEXT,
    nombre              TEXT,
    tipo_doc            TEXT,
    num_doc             TEXT PRIMARY KEY,
    telefono            TEXT,
    email               TEXT,
    ciudad              TEXT,
    regimen             TEXT,
    tipo_persona        TEXT,
    razon_social        TEXT,
    nit_empresa         TEXT,
    activo              TEXT DEFAULT '1',
    total_reservas      TEXT,
    total_gastado       TEXT,
    creado_en           TEXT
);

CREATE TABLE IF NOT EXISTS cubiculos_estado (
    cubiculo_id         TEXT PRIMARY KEY,
    numero              TEXT UNIQUE,
    nombre              TEXT,
    estado              TEXT DEFAULT 'libre',
    cliente_actual      TEXT,
    hora_inicio         TEXT,
    hora_fin_prog       TEXT,
    tiempo_rest_min     TEXT,
    codigo_acceso       TEXT,
    wifi_ssid           TEXT,
    wifi_pass           TEXT,
    precio_hora_base    TEXT DEFAULT '15000'
);

CREATE TABLE IF NOT EXISTS facturas (
    id_factura          TEXT,
    num_factura         TEXT PRIMARY KEY,
    tipo                TEXT,
    fecha_emision       TEXT,
    fecha_vencimiento   TEXT,
    cliente_nombre      TEXT,
    cliente_doc         TEXT,
    cliente_email       TEXT,
    razon_social        TEXT,
    nit_empresa         TEXT,
    descripcion         TEXT,
    subtotal            TEXT,
    descuento           TEXT,
    base_gravable       TEXT,
    iva                 TEXT,
    retenciones         TEXT,
    total_cop           TEXT,
    metodo_pago         TEXT,
    estado              TEXT DEFAULT 'pagada',
    moneda              TEXT DEFAULT 'COP',
    num_reserva         TEXT,
    cubiculo            TEXT,
    creado_en           TEXT,
    actualizado_en      TEXT
);

CREATE TABLE IF NOT EXISTS factura_items (
    id_item             TEXT,
    id_factura          TEXT,
    codigo              TEXT,
    descripcion         TEXT,
    cantidad            TEXT,
    unidad              TEXT,
    precio_unitario     TEXT,
    descuento_pct       TEXT,
    iva_pct             TEXT,
    subtotal            TEXT
);

CREATE TABLE IF NOT EXISTS operadores (
    id_operador         TEXT PRIMARY KEY,
    nombre              TEXT UNIQUE,
    rol                 TEXT DEFAULT 'cajero',
    turno               TEXT,
    hora_ini_turno      TEXT,
    hora_fin_turno      TEXT,
    permisos            TEXT DEFAULT 'reservas,pagos,voucher',
    activo              TEXT DEFAULT '1'
);

CREATE TABLE IF NOT EXISTS configuracion_pagos (
    clave   TEXT PRIMARY KEY,
    valor   TEXT
);

CREATE TABLE IF NOT EXISTS tarifas_config (
    id                  TEXT PRIMARY KEY,
    nombre              TEXT,
    descripcion         TEXT,
    precio_hora_cop     TEXT,
    desc_3h_pct         TEXT,
    desc_6h_pct         TEXT,
    hora_ini_espec      TEXT,
    hora_fin_espec      TEXT,
    aplica_festivos     TEXT DEFAULT '0',
    activo              TEXT DEFAULT '1',
    horas_a_reservar    TEXT
);

CREATE TABLE IF NOT EXISTS dashboard_diario (
    fecha               TEXT PRIMARY KEY,
    total_reservas      TEXT,
    completadas         TEXT,
    canceladas          TEXT,
    ingresos_brutos     TEXT,
    iva_recaudado       TEXT,
    ingresos_netos      TEXT,
    nequi_cop           TEXT,
    daviplata_cop       TEXT,
    efectivo_cop        TEXT,
    pse_cop             TEXT,
    mp_cop              TEXT,
    otros_cop           TEXT,
    ocupacion_pct       TEXT,
    hora_pico           TEXT,
    tiempo_prom_min     TEXT,
    clientes_nuevos     TEXT,
    clientes_recur      TEXT,
    fact_min            TEXT,
    fact_max            TEXT,
    ticket_prom_cop     TEXT
);

CREATE TABLE IF NOT EXISTS log_operaciones (
    id                  SERIAL PRIMARY KEY,
    timestamp           TEXT,
    tipo_op             TEXT,
    reserva_id          TEXT,
    cubiculo            TEXT,
    operador            TEXT,
    descripcion         TEXT,
    valor_ant           TEXT,
    valor_nuevo         TEXT,
    ip                  TEXT,
    estado              TEXT,
    notas               TEXT
);

-- ── Tablas exclusivas de CONVENIOS ────────────────────────────────────────────

CREATE TABLE IF NOT EXISTS empresas_convenio (
    id_empresa          TEXT PRIMARY KEY,
    nombre_empresa      TEXT,
    nit_empresa         TEXT,
    tipo_empresa        TEXT DEFAULT 'Privada',
    email_empresa       TEXT,
    telefono_empresa    TEXT,
    tipo_convenio       TEXT DEFAULT 'Crédito',
    observaciones       TEXT,
    activo              TEXT DEFAULT '1',
    creado_en           TEXT
);

-- Reservas de convenio (estado pendiente, con referencia a empresa)
CREATE TABLE IF NOT EXISTS reservas_convenio (
    id_reserva          TEXT PRIMARY KEY,
    numero_reserva      TEXT UNIQUE,
    creado_en           TEXT,
    cubiculo_num        TEXT,
    cliente_nombre      TEXT,
    documento           TEXT,
    telefono            TEXT,
    email               TEXT,
    horas_contratadas   TEXT,
    hora_inicio         TEXT,
    hora_fin_prog       TEXT,
    hora_fin_real       TEXT,
    precio_hora         TEXT,
    subtotal            TEXT,
    iva                 TEXT,
    total_cop           TEXT,
    metodo_pago         TEXT DEFAULT 'Convenio',
    estado_pago         TEXT DEFAULT 'pendiente',
    codigo_acceso       TEXT,
    wifi_ssid           TEXT,
    wifi_pass           TEXT,
    num_factura         TEXT,
    referencia_pago     TEXT,
    operador            TEXT,
    notas               TEXT,
    id_empresa          TEXT,
    nombre_empresa      TEXT,
    acepto_datos        TEXT
);

CREATE TABLE IF NOT EXISTS pagos_convenio (
    id_pago             TEXT PRIMARY KEY,
    id_reserva          TEXT,
    num_reserva         TEXT,
    fecha_pago          TEXT,
    monto_cop           TEXT,
    metodo              TEXT DEFAULT 'Convenio',
    referencia_externa  TEXT,
    estado              TEXT DEFAULT 'pendiente',
    confirmado_por      TEXT,
    notas               TEXT,
    id_empresa          TEXT,
    nombre_empresa      TEXT
);

CREATE TABLE IF NOT EXISTS clientes_convenio (
    id_cliente          TEXT,
    nombre              TEXT,
    tipo_doc            TEXT,
    num_doc             TEXT,
    telefono            TEXT,
    email               TEXT,
    ciudad              TEXT,
    regimen             TEXT,
    tipo_persona        TEXT,
    razon_social        TEXT,
    nit_empresa         TEXT,
    activo              TEXT DEFAULT '1',
    total_reservas      TEXT,
    total_gastado       TEXT,
    creado_en           TEXT,
    id_empresa          TEXT,
    nombre_empresa      TEXT,
    PRIMARY KEY (num_doc, id_empresa)
);

CREATE TABLE IF NOT EXISTS facturas_convenio (
    id_factura          TEXT,
    num_factura         TEXT PRIMARY KEY,
    tipo                TEXT,
    fecha_emision       TEXT,
    fecha_vencimiento   TEXT,
    cliente_nombre      TEXT,
    cliente_doc         TEXT,
    cliente_email       TEXT,
    razon_social        TEXT,
    nit_empresa         TEXT,
    descripcion         TEXT,
    subtotal            TEXT,
    descuento           TEXT,
    base_gravable       TEXT,
    iva                 TEXT,
    retenciones         TEXT,
    total_cop           TEXT,
    metodo_pago         TEXT DEFAULT 'Convenio',
    estado              TEXT DEFAULT 'pendiente',
    moneda              TEXT DEFAULT 'COP',
    num_reserva         TEXT,
    cubiculo            TEXT,
    creado_en           TEXT,
    actualizado_en      TEXT,
    id_empresa          TEXT,
    nombre_empresa      TEXT
);

CREATE TABLE IF NOT EXISTS factura_items_convenio (
    id_item             TEXT,
    id_factura          TEXT,
    codigo              TEXT,
    descripcion         TEXT,
    cantidad            TEXT,
    unidad              TEXT,
    precio_unitario     TEXT,
    descuento_pct       TEXT,
    iva_pct             TEXT,
    subtotal            TEXT,
    id_empresa          TEXT
);

CREATE TABLE IF NOT EXISTS dashboard_diario_convenio (
    fecha               TEXT,
    id_empresa          TEXT,
    total_reservas      TEXT,
    completadas         TEXT,
    canceladas          TEXT,
    ingresos_brutos     TEXT,
    iva_recaudado       TEXT,
    ingresos_netos      TEXT,
    nequi_cop           TEXT,
    daviplata_cop       TEXT,
    efectivo_cop        TEXT,
    pse_cop             TEXT,
    mp_cop              TEXT,
    otros_cop           TEXT,
    convenio_cop        TEXT,
    ocupacion_pct       TEXT,
    hora_pico           TEXT,
    tiempo_prom_min     TEXT,
    clientes_nuevos     TEXT,
    clientes_recur      TEXT,
    fact_min            TEXT,
    fact_max            TEXT,
    ticket_prom_cop     TEXT,
    PRIMARY KEY (fecha, id_empresa)
);

-- Configuración por empresa (equivalente a Configuracion_Pagos en jjgt_convenios)
CREATE TABLE IF NOT EXISTS configuracion_empresa_convenio (
    id_empresa          TEXT,
    nombre_empresa      TEXT,
    clave               TEXT,
    valor               TEXT,
    actualizado_en      TEXT,
    PRIMARY KEY (id_empresa, clave)
);

-- Tarifas específicas por empresa
CREATE TABLE IF NOT EXISTS tarifas_config_convenio (
    id                  TEXT PRIMARY KEY,
    id_empresa          TEXT,
    empresa_nombre      TEXT,
    descripcion         TEXT,
    precio_hora_cop     TEXT,
    desc_3h_pct         TEXT,
    desc_6h_pct         TEXT,
    hora_ini_espec      TEXT,
    hora_fin_espec      TEXT,
    aplica_festivos     TEXT DEFAULT '0',
    activo              TEXT DEFAULT '1'
);
"""

_SEED_SQL = """
INSERT INTO tarifas_config (id, nombre, descripcion, precio_hora_cop, desc_3h_pct,
    desc_6h_pct, hora_ini_espec, hora_fin_espec, aplica_festivos, activo, horas_a_reservar)
VALUES
    ('1','Estándar','Tarifa estándar diurna','15000','0','0','','','0','1',''),
    ('2','Madrugada','Tarifa madrugada 00-06h','15000','0','0','00:00','06:00','0','1',''),
    ('3','Noche Completa','Noche completa','40000','0','0','22:00','06:00','0','1','8')
ON CONFLICT (id) DO NOTHING;

INSERT INTO configuracion_pagos (clave, valor) VALUES
    ('factura_prefijo','FACT'),
    ('factura_contador','0'),
    ('nequi_numero','3219714969'),
    ('daviplata_numero','3219714969'),
    ('mp_link','https://mpago.la/XXXXXXX'),
    ('cuenta_bancaria','Bancolombia · Cta Ahorros · 123-456789-12'),
    ('negocio_nit','902.047.871-3'),
    ('convenios_dias_credito','30'),
    ('convenios_horas_min','1.0'),
    ('convenios_descuento_pct','0'),
    ('convenios_max_pendientes','10')
ON CONFLICT (clave) DO NOTHING;
"""


def init_db():
    """
    Crea todas las tablas en PostgreSQL si no existen y siembra datos iniciales.
    Llamar una vez al inicio de la aplicación.
    """
    if not PSYCOPG2_AVAILABLE:
        st.error(
            "❌ **psycopg2 no está instalado.**\n\n"
            "Agrega esto a tu `requirements.txt`:\n```\npsycopg2-binary\n```\n"
            "O instala localmente: `pip install psycopg2-binary`"
        )
        st.stop()
    try:
        conn = get_pg_conn()
        with conn.cursor() as cur:
            cur.execute(_CREATE_TABLES_SQL)
            cur.execute(_SEED_SQL)
        conn.commit()
    except Exception as e:
        st.error(f"❌ Error inicializando base de datos PostgreSQL: {e}")


def get_db():
    """Compatibilidad: retorna una conexión PostgreSQL."""
    return get_pg_conn()


# ══════════════════════════════════════════════════════════════════════════════
# MAPEO: nombre de hoja → tabla PG y columnas
# ══════════════════════════════════════════════════════════════════════════════

_TABLA_MAP = {
    "Reservas":            "reservas",
    "Pagos":               "pagos",
    "Clientes":            "clientes",
    "Cubiculos_Estado":    "cubiculos_estado",
    "Facturas":            "facturas",
    "Factura_items":       "factura_items",
    "Operadores":          "operadores",
    "Configuracion_Pagos": "configuracion_pagos",
    "Tarifas_Config":      "tarifas_config",
    "Dashboard_Diario":    "dashboard_diario",
    "Log_Operaciones":     "log_operaciones",
}

_TABLA_MAP_CONV = {
    "Empresa":             "empresas_convenio",
    "Reservas":            "reservas_convenio",
    "Pagos":               "pagos_convenio",
    "Clientes":            "clientes_convenio",
    "Facturas":            "facturas_convenio",
    "Factura_items":       "factura_items_convenio",
    "Dashboard_Diario":    "dashboard_diario_convenio",
    "Cubiculos_Estado":    "cubiculos_estado",          # compartida con jjgt_pagos
    "Tarifas_Config":      "tarifas_config_convenio",
    "Operadores":          "operadores",               # compartida
    "Configuracion_Pagos": "configuracion_empresa_convenio",
    "Log_Operaciones":     "log_operaciones",          # compartida
}

_PK_MAP = {
    "reservas":            "numero_reserva",
    "pagos":               "id_pago",
    "clientes":            "num_doc",
    "cubiculos_estado":    "cubiculo_id",
    "facturas":            "num_factura",
    "factura_items":       None,
    "operadores":          "nombre",
    "configuracion_pagos": "clave",
    "tarifas_config":      "id",
    "dashboard_diario":    "fecha",
    "log_operaciones":     None,
    # convenio
    "empresas_convenio":           "id_empresa",
    "reservas_convenio":           "numero_reserva",
    "pagos_convenio":              "id_pago",
    "clientes_convenio":           None,   # PK compuesta (num_doc, id_empresa)
    "facturas_convenio":           "num_factura",
    "factura_items_convenio":      None,
    "dashboard_diario_convenio":   None,   # PK compuesta (fecha, id_empresa)
    "configuracion_empresa_convenio": None, # PK compuesta (id_empresa, clave)
    "tarifas_config_convenio":     "id",
}

_COL_MAP = {
    "reservas": [
        "id_reserva","numero_reserva","creado_en","cubiculo_num",
        "cliente_nombre","documento","telefono","email",
        "horas_contratadas","hora_inicio","hora_fin_prog","hora_fin_real",
        "precio_hora","subtotal","iva","total_cop",
        "metodo_pago","estado_pago","codigo_acceso",
        "wifi_ssid","wifi_pass","num_factura","referencia_pago",
        "operador","notas","acepto_datos",
    ],
    "pagos": [
        "id_pago","id_reserva","num_reserva","fecha_pago",
        "monto_cop","metodo","referencia_externa","estado",
        "confirmado_por","notas",
    ],
    "clientes": [
        "id_cliente","nombre","tipo_doc","num_doc",
        "telefono","email","ciudad","regimen","tipo_persona",
        "razon_social","nit_empresa","activo",
        "total_reservas","total_gastado","creado_en",
    ],
    "cubiculos_estado": [
        "cubiculo_id","numero","nombre","estado",
        "cliente_actual","hora_inicio","hora_fin_prog",
        "tiempo_rest_min","codigo_acceso","wifi_ssid","wifi_pass",
        "precio_hora_base",
    ],
    "facturas": [
        "id_factura","num_factura","tipo","fecha_emision","fecha_vencimiento",
        "cliente_nombre","cliente_doc","cliente_email","razon_social","nit_empresa",
        "descripcion","subtotal","descuento","base_gravable","iva","retenciones",
        "total_cop","metodo_pago","estado","moneda",
        "num_reserva","cubiculo","creado_en","actualizado_en",
    ],
    "factura_items": [
        "id_item","id_factura","codigo","descripcion",
        "cantidad","unidad","precio_unitario","descuento_pct","iva_pct","subtotal",
    ],
    "operadores": [
        "id_operador","nombre","rol","turno",
        "hora_ini_turno","hora_fin_turno","permisos","activo",
    ],
    "configuracion_pagos": ["clave","valor"],
    "tarifas_config": [
        "id","nombre","descripcion","precio_hora_cop",
        "desc_3h_pct","desc_6h_pct","hora_ini_espec","hora_fin_espec",
        "aplica_festivos","activo","horas_a_reservar",
    ],
    "dashboard_diario": [
        "fecha","total_reservas","completadas","canceladas",
        "ingresos_brutos","iva_recaudado","ingresos_netos",
        "nequi_cop","daviplata_cop","efectivo_cop","pse_cop",
        "mp_cop","otros_cop","ocupacion_pct","hora_pico",
        "tiempo_prom_min","clientes_nuevos","clientes_recur",
        "fact_min","fact_max","ticket_prom_cop",
    ],
    "log_operaciones": [
        "timestamp","tipo_op","reserva_id","cubiculo",
        "operador","descripcion","valor_ant","valor_nuevo",
        "ip","estado","notas",
    ],
    # convenio
    "empresas_convenio": [
        "id_empresa","nombre_empresa","nit_empresa","tipo_empresa",
        "email_empresa","telefono_empresa","tipo_convenio",
        "observaciones","activo","creado_en",
    ],
    "reservas_convenio": [
        "id_reserva","numero_reserva","creado_en","cubiculo_num",
        "cliente_nombre","documento","telefono","email",
        "horas_contratadas","hora_inicio","hora_fin_prog","hora_fin_real",
        "precio_hora","subtotal","iva","total_cop",
        "metodo_pago","estado_pago","codigo_acceso",
        "wifi_ssid","wifi_pass","num_factura","referencia_pago",
        "operador","notas","id_empresa","nombre_empresa","acepto_datos",
    ],
    "pagos_convenio": [
        "id_pago","id_reserva","num_reserva","fecha_pago",
        "monto_cop","metodo","referencia_externa","estado",
        "confirmado_por","notas","id_empresa","nombre_empresa",
    ],
    "clientes_convenio": [
        "id_cliente","nombre","tipo_doc","num_doc","telefono","email",
        "ciudad","regimen","tipo_persona","razon_social","nit_empresa",
        "activo","total_reservas","total_gastado","creado_en",
        "id_empresa","nombre_empresa",
    ],
    "facturas_convenio": [
        "id_factura","num_factura","tipo","fecha_emision","fecha_vencimiento",
        "cliente_nombre","cliente_doc","cliente_email","razon_social","nit_empresa",
        "descripcion","subtotal","descuento","base_gravable","iva","retenciones",
        "total_cop","metodo_pago","estado","moneda","num_reserva","cubiculo",
        "creado_en","actualizado_en","id_empresa","nombre_empresa",
    ],
    "factura_items_convenio": [
        "id_item","id_factura","codigo","descripcion","cantidad","unidad",
        "precio_unitario","descuento_pct","iva_pct","subtotal","id_empresa",
    ],
    "dashboard_diario_convenio": [
        "fecha","id_empresa","total_reservas","completadas","canceladas",
        "ingresos_brutos","iva_recaudado","ingresos_netos",
        "nequi_cop","daviplata_cop","efectivo_cop","pse_cop",
        "mp_cop","otros_cop","convenio_cop","ocupacion_pct","hora_pico",
        "tiempo_prom_min","clientes_nuevos","clientes_recur",
        "fact_min","fact_max","ticket_prom_cop",
    ],
    "configuracion_empresa_convenio": [
        "id_empresa","nombre_empresa","clave","valor","actualizado_en",
    ],
    "tarifas_config_convenio": [
        "id","id_empresa","empresa_nombre","descripcion","precio_hora_cop",
        "desc_3h_pct","desc_6h_pct","hora_ini_espec","hora_fin_espec",
        "aplica_festivos","activo",
    ],
}


def _col_pg(hoja_sheets: str, col_sheets: str) -> str:
    """Convierte nombre de columna Sheets a nombre PG (lowercase)."""
    return col_sheets.lower()


def _tabla_pg(hoja: str) -> str:
    return _TABLA_MAP.get(hoja, hoja.lower())


def _tabla_pg_conv(hoja: str) -> str:
    return _TABLA_MAP_CONV.get(hoja, hoja.lower())


# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES INTERNAS DE LECTURA / ESCRITURA — tablas jjgt_pagos
# ══════════════════════════════════════════════════════════════════════════════

def _pg_read_table(hoja: str) -> list:
    """Lee registros de tabla PG y retorna lista de dicts con claves Sheets."""
    if not PSYCOPG2_AVAILABLE:
        return []
    tabla = _tabla_pg(hoja)
    col_pg_list = _COL_MAP.get(tabla, [])
    col_sh_list = DRIVE_SHEETS.get(hoja, [])
    try:
        rows = _pg_exec(f"SELECT * FROM {tabla} ORDER BY 1", fetch="all") or []
    except Exception as e:
        st.warning(f"⚠️ Error leyendo tabla '{tabla}': {e}")
        return []
    result = []
    for row in rows:
        rec = {}
        for pg_col, sh_col in zip(col_pg_list, col_sh_list):
            val = row.get(pg_col, row.get(pg_col.lower(), ""))
            rec[sh_col] = "" if val is None else str(val)
        result.append(rec)
    return result


def _pg_read_table_conv(hoja: str) -> list:
    """Lee registros de tabla de convenios PG, retorna dicts con claves Sheets."""
    if not PSYCOPG2_AVAILABLE:
        return []
    tabla = _tabla_pg_conv(hoja)
    col_pg_list = _COL_MAP.get(tabla, [])
    col_sh_list = DRIVE_SHEETS_CONVENIOS.get(hoja, [])
    try:
        rows = _pg_exec(f"SELECT * FROM {tabla} ORDER BY 1", fetch="all") or []
    except Exception as e:
        st.warning(f"⚠️ Error leyendo tabla convenio '{tabla}': {e}")
        return []
    result = []
    for row in rows:
        rec = {}
        for pg_col, sh_col in zip(col_pg_list, col_sh_list):
            val = row.get(pg_col, row.get(pg_col.lower(), ""))
            rec[sh_col] = "" if val is None else str(val)
        result.append(rec)
    return result


def _pg_upsert(hoja: str, col_clave_sh: str, valor_clave: str, fila: list) -> bool:
    """INSERT ... ON CONFLICT DO UPDATE para tablas de jjgt_pagos."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla   = _tabla_pg(hoja)
    cols_pg = _COL_MAP.get(tabla, [])
    if not cols_pg:
        return False
    fila_str = [str(v) if v is not None else "" for v in fila]
    fila_str = (fila_str + [""] * len(cols_pg))[:len(cols_pg)]
    pk_col = _PK_MAP.get(tabla)
    if not pk_col:
        return _pg_append(hoja, fila)
    cols_sql     = ", ".join(f'"{c}"' for c in cols_pg)
    placeholders = ", ".join(["%s"] * len(cols_pg))
    update_sql   = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in cols_pg if c != pk_col)
    sql = (
        f'INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders}) '
        f'ON CONFLICT ("{pk_col}") DO UPDATE SET {update_sql}'
    )
    try:
        _pg_exec(sql, fila_str)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error upsert '{tabla}': {e}")
        return False


def _pg_upsert_conv(hoja: str, col_clave_sh: str, valor_clave: str, fila: list) -> bool:
    """INSERT ... ON CONFLICT DO UPDATE para tablas de jjgt_convenios."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla   = _tabla_pg_conv(hoja)
    cols_pg = _COL_MAP.get(tabla, [])
    if not cols_pg:
        return False
    fila_str = [str(v) if v is not None else "" for v in fila]
    fila_str = (fila_str + [""] * len(cols_pg))[:len(cols_pg)]
    pk_col = _PK_MAP.get(tabla)
    if not pk_col:
        return _pg_append_conv(hoja, fila)
    cols_sql     = ", ".join(f'"{c}"' for c in cols_pg)
    placeholders = ", ".join(["%s"] * len(cols_pg))
    update_sql   = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in cols_pg if c != pk_col)
    sql = (
        f'INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders}) '
        f'ON CONFLICT ("{pk_col}") DO UPDATE SET {update_sql}'
    )
    try:
        _pg_exec(sql, fila_str)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error upsert convenio '{tabla}': {e}")
        return False


def _pg_append(hoja: str, fila: list) -> bool:
    """INSERT de una nueva fila en tablas de jjgt_pagos."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla   = _tabla_pg(hoja)
    cols_pg = _COL_MAP.get(tabla, [])
    if not cols_pg:
        return False
    fila_str = [str(v) if v is not None else "" for v in fila]
    fila_str = (fila_str + [""] * len(cols_pg))[:len(cols_pg)]
    cols_sql     = ", ".join(f'"{c}"' for c in cols_pg)
    placeholders = ", ".join(["%s"] * len(cols_pg))
    sql = f'INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders})'
    try:
        _pg_exec(sql, fila_str)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error insert '{tabla}': {e}")
        return False


def _pg_append_conv(hoja: str, fila: list) -> bool:
    """INSERT de una nueva fila en tablas de jjgt_convenios."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla   = _tabla_pg_conv(hoja)
    cols_pg = _COL_MAP.get(tabla, [])
    if not cols_pg:
        return False
    fila_str = [str(v) if v is not None else "" for v in fila]
    fila_str = (fila_str + [""] * len(cols_pg))[:len(cols_pg)]
    cols_sql     = ", ".join(f'"{c}"' for c in cols_pg)
    placeholders = ", ".join(["%s"] * len(cols_pg))
    sql = f'INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders})'
    try:
        _pg_exec(sql, fila_str)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error insert convenio '{tabla}': {e}")
        return False


def _pg_update_col(hoja: str, where_col_sh: str, where_val: str,
                   set_col_sh: str, new_val: str) -> bool:
    """UPDATE tabla SET col = val WHERE where_col = where_val (jjgt_pagos)."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla     = _tabla_pg(hoja)
    where_col = where_col_sh.lower()
    set_col   = set_col_sh.lower()
    sql = f'UPDATE {tabla} SET "{set_col}" = %s WHERE "{where_col}" = %s'
    try:
        _pg_exec(sql, [new_val, where_val])
        return True
    except Exception as e:
        st.warning(f"⚠️ Error update '{tabla}': {e}")
        return False


def _pg_update_col_conv(hoja: str, where_col_sh: str, where_val: str,
                         set_col_sh: str, new_val: str) -> bool:
    """UPDATE tabla SET col = val WHERE where_col = where_val (jjgt_convenios)."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla     = _tabla_pg_conv(hoja)
    where_col = where_col_sh.lower()
    set_col   = set_col_sh.lower()
    sql = f'UPDATE {tabla} SET "{set_col}" = %s WHERE "{where_col}" = %s'
    try:
        _pg_exec(sql, [new_val, where_val])
        return True
    except Exception as e:
        st.warning(f"⚠️ Error update convenio '{tabla}': {e}")
        return False


def _pg_delete_row(hoja: str, col_clave_sh: str, valor_clave: str) -> bool:
    """DELETE FROM tabla WHERE col = valor (jjgt_pagos)."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla  = _tabla_pg(hoja)
    col_pg = col_clave_sh.lower()
    sql    = f'DELETE FROM {tabla} WHERE "{col_pg}" = %s'
    try:
        _pg_exec(sql, [valor_clave])
        return True
    except Exception as e:
        st.warning(f"⚠️ Error delete '{tabla}': {e}")
        return False


def _pg_delete_row_conv(hoja: str, col_clave_sh: str, valor_clave: str) -> bool:
    """DELETE FROM tabla WHERE col = valor (jjgt_convenios)."""
    if not PSYCOPG2_AVAILABLE:
        return False
    tabla  = _tabla_pg_conv(hoja)
    col_pg = col_clave_sh.lower()
    sql    = f'DELETE FROM {tabla} WHERE "{col_pg}" = %s'
    try:
        _pg_exec(sql, [valor_clave])
        return True
    except Exception as e:
        st.warning(f"⚠️ Error delete convenio '{tabla}': {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# API PÚBLICA — mismas firmas que los _gs_* originales
# ══════════════════════════════════════════════════════════════════════════════

def _gs_read_sheet(hoja: str, force: bool = False) -> list:
    """Lee todos los registros de una tabla PG (jjgt_pagos)."""
    return _pg_read_table(hoja)


def _gs_read_sheet_conv(hoja: str, force: bool = False) -> list:
    """Lee todos los registros de una tabla PG de convenios (jjgt_convenios)."""
    return _pg_read_table_conv(hoja)


def _gs_invalidate_cache(*hojas):
    """No-op: PostgreSQL no tiene caché de sesión."""
    pass


def _gs_invalidate_cache_conv(*hojas):
    """No-op: PostgreSQL no tiene caché de sesión."""
    pass


def _gs_val(row: dict, key: str, default="") -> str:
    v = row.get(key, default)
    return str(v) if v != "" else str(default)


def _gs_float(row: dict, key: str, default: float = 0.0) -> float:
    try:
        return float(_gs_val(row, key, default))
    except Exception:
        return float(default)


def _gs_int(row: dict, key: str, default: int = 0) -> int:
    try:
        return int(float(_gs_val(row, key, default)))
    except Exception:
        return int(default)


def _gs_fecha_ymd(row: dict, key: str) -> str:
    raw = _gs_val(row, key, "").strip()
    if not raw:
        return ""
    if len(raw) >= 10 and raw[4:5] == "-":
        return raw[:10]
    if "/" in raw:
        parte  = raw.split(" ")[0]
        partes = parte.split("/")
        if len(partes) == 3:
            a, b, c = partes
            if len(c) == 4:
                if int(a) > 12:
                    return f"{c}-{b.zfill(2)}-{a.zfill(2)}"
                else:
                    return f"{c}-{b.zfill(2)}-{a.zfill(2)}"
            elif len(a) == 4:
                return f"{a}-{b.zfill(2)}-{c.zfill(2)}"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y", "%d-%m-%Y %H:%M"):
        try:
            return datetime.strptime(raw[:len(fmt)], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw[:10]


def _gs_append(sh, hoja: str, fila: list) -> bool:
    """Equivalente al _gs_append original: inserta fila en PG (sh ignorado)."""
    return _pg_append(hoja, fila)


def _gs_upsert(sh, hoja: str, col_clave: str, valor_clave: str, fila: list) -> bool:
    """Equivalente al _gs_upsert original (sh ignorado)."""
    return _pg_upsert(hoja, col_clave, valor_clave, fila)


def _gs_with_retry(func, *args, operacion: str = "", **kwargs):
    """Sin reintentos en PG — ejecuta directo."""
    return func(*args, **kwargs)


def _gs_get_or_create_ws(sh, hoja: str):
    """Stub de compatibilidad — en PG no hay worksheets."""
    return None


def _gs_update_row(ws, row_num: int, padded: list):
    """No-op: actualizaciones en PG se hacen por clave."""
    pass


def _gs_update_range(ws, range_name: str, data: list):
    """No-op."""
    pass


# ══════════════════════════════════════════════════════════════════════════════
# STUBS DE CONEXIÓN — reemplazan get_active_client / get_active_client_convenios
# ══════════════════════════════════════════════════════════════════════════════

class _PGStub:
    """Sustituye al objeto spreadsheet de gspread en llamadas heredadas."""
    title = f"PostgreSQL · {PG_DB}"

    def worksheet(self, name):           return self
    def get_all_values(self):            return []
    def get_all_records(self, **kw):     return []
    def append_row(self, row):           pass
    def update_cell(self, row, col, val):pass
    def delete_rows(self, row):          pass
    def update(self, values, range_name):pass
    def add_worksheet(self, title, rows, cols): return self
    def del_worksheet(self, ws):         pass
    def worksheets(self):                return []
    def fetch_sheet_metadata(self):      return {"sheets": []}


_PG_STUB = _PGStub()


def get_active_client():
    """
    Reemplaza flujo credenciales→cliente→spreadsheet.
    Retorna (None, _PGStub) para mantener el patrón `_, sh = get_active_client()`.
    """
    init_db()
    return None, _PG_STUB


def get_active_client_convenios():
    """
    Reemplaza get_active_client_convenios de Google Sheets.
    Retorna (None, _PGStub) — las tablas convenio ya existen en PG.
    """
    init_db()
    return None, _PG_STUB


def _get_module_level_client():
    return get_active_client()


def _reset_gs_cache():
    pass


def load_credentials_from_toml():
    return None, None


def get_google_sheets_connection(_creds):
    return None


def get_or_create_spreadsheet(client):
    return _PG_STUB


# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA POR ENTIDAD — tablas jjgt_pagos (mismas firmas que antes)
# ══════════════════════════════════════════════════════════════════════════════

def gs_escribir_cliente(sh, datos: dict) -> bool:
    num_doc = str(datos.get("numero_documento", ""))
    total_res = total_gas = "0"
    try:
        reservas_pg = _pg_read_table("Reservas")
        res_cliente = [r for r in reservas_pg
                       if str(r.get("Documento", "")) == num_doc
                       and str(r.get("Estado_Pago", "")) == "confirmado"]
        total_res = str(len(res_cliente))
        total_gas = str(round(sum(_gs_float(r, "Total_COP") for r in res_cliente), 2))
    except Exception:
        pass
    fila = [
        str(datos.get("id", "")),
        str(datos.get("nombre", "")),
        str(datos.get("tipo_documento", "CC")),
        num_doc,
        str(datos.get("telefono", "")),
        str(datos.get("email", "")),
        str(datos.get("ciudad", "")),
        str(datos.get("regimen", "Simplificado")),
        str(datos.get("tipo_persona", "Natural")),
        str(datos.get("razon_social", "")),
        str(datos.get("nit_empresa", "")),
        str(datos.get("activo", "1")),
        total_res,
        total_gas,
        str(datos.get("creado_en", "")),
    ]
    return _pg_upsert("Clientes", "Num_Doc", num_doc, fila)


def gs_escribir_factura(sh, datos: dict) -> bool:
    subtotal      = float(datos.get("subtotal", 0) or 0)
    descuento     = float(datos.get("descuento", 0) or 0)
    base_gravable = round(subtotal - descuento, 2)
    fila = [
        str(datos.get("id", "")),
        str(datos.get("numero", "")),
        str(datos.get("tipo", "Factura de Venta")),
        str(datos.get("fecha_emision", "")),
        str(datos.get("fecha_vencimiento", "")),
        str(datos.get("cliente_nombre", "")),
        str(datos.get("cliente_doc", "")),
        str(datos.get("cliente_email", "")),
        str(datos.get("razon_social", "")),
        str(datos.get("nit_empresa", "")),
        str(datos.get("descripcion", "")),
        str(subtotal),
        str(descuento),
        str(base_gravable),
        str(datos.get("iva", "")),
        str(datos.get("retenciones", "0")),
        str(datos.get("total", "")),
        str(datos.get("metodo_pago", "")),
        str(datos.get("estado", "pagada")),
        str(datos.get("moneda", "COP")),
        str(datos.get("num_reserva", "")),
        str(datos.get("cubiculo", "")),
        str(datos.get("creado_en", "")),
        str(datos.get("actualizado_en", "")),
    ]
    return _pg_upsert("Facturas", "Num_Factura", str(datos.get("numero", "")), fila)


def gs_escribir_reserva(sh, datos: dict) -> bool:
    fila = [
        str(datos.get("id", "")),
        str(datos.get("numero_reserva", "")),
        str(datos.get("creado_en", "")),
        str(datos.get("cubiculo_num", "")),
        str(datos.get("cliente_nombre", "")),
        str(datos.get("cliente_doc", "")),
        str(datos.get("cliente_tel", "")),
        str(datos.get("cliente_email", "")),
        str(datos.get("horas", "")),
        str(datos.get("hora_inicio", "")),
        str(datos.get("hora_fin_prog", "")),
        str(datos.get("hora_fin_real", "")),
        str(datos.get("precio_hora", "")),
        str(datos.get("subtotal", "")),
        str(datos.get("iva", "")),
        str(datos.get("total", "")),
        str(datos.get("metodo_pago", "")),
        str(datos.get("estado_pago", "confirmado")),
        str(datos.get("codigo_acceso", "")),
        str(datos.get("wifi_ssid", "")),
        str(datos.get("wifi_pass", "")),
        str(datos.get("num_factura", "")),
        str(datos.get("referencia_pago", "")),
        str(datos.get("operador", "sistema")),
        str(datos.get("notas", "")),
        "Sí" if datos.get("acepto_datos") else "No",
    ]
    return _pg_upsert("Reservas", "Numero_Reserva",
                      str(datos.get("numero_reserva", "")), fila)


def gs_escribir_pago(sh, datos: dict) -> bool:
    fila = [
        str(datos.get("id", "")),
        str(datos.get("reserva_id", "")),
        str(datos.get("num_reserva", "")),
        str(datos.get("fecha_pago", "")),
        str(datos.get("monto", "")),
        str(datos.get("metodo", "")),
        str(datos.get("referencia_externa", "")),
        str(datos.get("estado", "confirmado")),
        str(datos.get("confirmado_por", "sistema")),
        str(datos.get("notas", "")),
    ]
    return _pg_upsert("Pagos", "ID_Pago", str(datos.get("id", "")), fila)


def gs_escribir_log(sh, tipo_op, reserva_id, cubiculo, operador, descripcion, estado="exito"):
    _pg_append("Log_Operaciones", [
        ahora_col().isoformat(), tipo_op, str(reserva_id), str(cubiculo),
        operador, descripcion, "", "", "", estado, ""
    ])


# ══════════════════════════════════════════════════════════════════════════════
# ESCRITURA POR ENTIDAD — tablas jjgt_convenios
# ══════════════════════════════════════════════════════════════════════════════

def gs_escribir_empresa(sh_conv, datos: dict) -> bool:
    """Escribe o actualiza una empresa en la tabla empresas_convenio."""
    fila = [
        str(datos.get("id_empresa", "")),
        str(datos.get("nombre_empresa", "")),
        str(datos.get("nit_empresa", "")),
        str(datos.get("tipo_empresa", "Privada")),
        str(datos.get("email_empresa", "")),
        str(datos.get("telefono_empresa", "")),
        str(datos.get("tipo_convenio", "Crédito")),
        str(datos.get("observaciones", "")),
        str(datos.get("activo", "1")),
        str(datos.get("creado_en", ahora_col().isoformat())),
    ]
    return _pg_upsert_conv("Empresa", "Id_Empresa",
                            str(datos.get("id_empresa", "")), fila)


def gs_escribir_reserva_convenio(sh_conv, datos: dict) -> bool:
    """Escribe una reserva de convenio en reservas_convenio."""
    fila = [
        str(datos.get("id", "")),
        str(datos.get("numero_reserva", "")),
        str(datos.get("creado_en", "")),
        str(datos.get("cubiculo_num", "")),
        str(datos.get("cliente_nombre", "")),
        str(datos.get("cliente_doc", "")),
        str(datos.get("cliente_tel", "")),
        str(datos.get("cliente_email", "")),
        str(datos.get("horas", "")),
        str(datos.get("hora_inicio", "")),
        str(datos.get("hora_fin_prog", "")),
        str(datos.get("hora_fin_real", "")),
        str(datos.get("precio_hora", "")),
        str(datos.get("subtotal", "")),
        str(datos.get("iva", "")),
        str(datos.get("total", "")),
        str(datos.get("metodo_pago", "Convenio")),
        str(datos.get("estado_pago", "pendiente")),
        str(datos.get("codigo_acceso", "")),
        str(datos.get("wifi_ssid", "")),
        str(datos.get("wifi_pass", "")),
        str(datos.get("num_factura", "")),
        str(datos.get("referencia_pago", "")),
        str(datos.get("operador", "sistema")),
        str(datos.get("notas", "")),
        str(datos.get("id_empresa", "")),
        str(datos.get("nombre_empresa", "")),
        "Sí" if datos.get("acepto_datos") else "No",
    ]
    return _pg_upsert_conv("Reservas", "Numero_Reserva",
                            str(datos.get("numero_reserva", "")), fila)


def gs_escribir_pago_convenio(sh_conv, datos: dict) -> bool:
    """Escribe un pago de convenio (pendiente) en pagos_convenio."""
    fila = [
        str(datos.get("id", "")),
        str(datos.get("reserva_id", "")),
        str(datos.get("num_reserva", "")),
        str(datos.get("fecha_pago", "")),
        str(datos.get("monto", "")),
        str(datos.get("metodo", "Convenio")),
        str(datos.get("referencia_externa", "")),
        str(datos.get("estado", "pendiente")),
        str(datos.get("confirmado_por", "sistema")),
        str(datos.get("notas", "")),
        str(datos.get("id_empresa", "")),
        str(datos.get("nombre_empresa", "")),
    ]
    return _pg_upsert_conv("Pagos", "Num_Reserva",
                            str(datos.get("num_reserva", "")), fila)


def gs_escribir_cliente_convenio(sh_conv, datos: dict, id_empresa: str, nombre_empresa: str) -> bool:
    """Escribe un cliente de convenio en clientes_convenio."""
    num_doc = str(datos.get("numero_documento", ""))
    fila = [
        str(datos.get("id", "")),
        str(datos.get("nombre", "")),
        str(datos.get("tipo_documento", "CC")),
        num_doc,
        str(datos.get("telefono", "")),
        str(datos.get("email", "")),
        str(datos.get("ciudad", "")),
        str(datos.get("regimen", "Simplificado")),
        str(datos.get("tipo_persona", "Natural")),
        str(datos.get("razon_social", "")),
        str(datos.get("nit_empresa", "")),
        str(datos.get("activo", "1")),
        "", "",
        str(datos.get("creado_en", ahora_col().isoformat())),
        id_empresa,
        nombre_empresa,
    ]
    # Upsert por num_doc + id_empresa usando SQL directo (PK compuesta)
    tabla = "clientes_convenio"
    cols_pg = _COL_MAP.get(tabla, [])
    fila_str = [str(v) if v is not None else "" for v in fila]
    fila_str = (fila_str + [""] * len(cols_pg))[:len(cols_pg)]
    cols_sql     = ", ".join(f'"{c}"' for c in cols_pg)
    placeholders = ", ".join(["%s"] * len(cols_pg))
    update_cols  = [c for c in cols_pg if c not in ("num_doc","id_empresa")]
    update_sql   = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in update_cols)
    sql = (
        f'INSERT INTO {tabla} ({cols_sql}) VALUES ({placeholders}) '
        f'ON CONFLICT ("num_doc","id_empresa") DO UPDATE SET {update_sql}'
    )
    try:
        _pg_exec(sql, fila_str)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error upsert clientes_convenio: {e}")
        return False


def gs_escribir_factura_convenio(sh_conv, datos: dict, id_empresa: str, nombre_empresa: str) -> bool:
    """Escribe una factura de convenio en facturas_convenio."""
    subtotal      = float(datos.get("subtotal", 0) or 0)
    descuento     = float(datos.get("descuento", 0) or 0)
    base_gravable = round(subtotal - descuento, 2)
    fila = [
        str(datos.get("id", "")),
        str(datos.get("numero", "")),
        str(datos.get("tipo", "Factura de Venta Convenio")),
        str(datos.get("fecha_emision", "")),
        str(datos.get("fecha_vencimiento", "")),
        str(datos.get("cliente_nombre", "")),
        str(datos.get("cliente_doc", "")),
        str(datos.get("cliente_email", "")),
        str(datos.get("razon_social", "")),
        str(datos.get("nit_empresa", "")),
        str(datos.get("descripcion", "")),
        str(subtotal), str(descuento), str(base_gravable),
        str(datos.get("iva", "")),
        str(datos.get("retenciones", "0")),
        str(datos.get("total", "")),
        str(datos.get("metodo_pago", "Convenio")),
        str(datos.get("estado", "pendiente")),
        str(datos.get("moneda", "COP")),
        str(datos.get("num_reserva", "")),
        str(datos.get("cubiculo", "")),
        str(datos.get("creado_en", "")),
        str(datos.get("actualizado_en", "")),
        id_empresa,
        nombre_empresa,
    ]
    return _pg_upsert_conv("Facturas", "Num_Factura",
                            str(datos.get("numero", "")), fila)


# ══════════════════════════════════════════════════════════════════════════════
# SYNC DE CUBÍCULOS / DASHBOARD — equivalentes PG
# ══════════════════════════════════════════════════════════════════════════════

def gs_sync_cubiculos(sh):
    """Actualiza estado de cubículos en PG desde la tabla reservas."""
    if not PSYCOPG2_AVAILABLE:
        return False
    try:
        now = ahora_col()
        reservas_pg  = _pg_read_table("Reservas")
        cubiculos_pg = _pg_read_table("Cubiculos_Estado")

        res_activas = {}
        for r in reservas_pg:
            if (_gs_val(r, "Estado_Pago") in ("confirmado", "pendiente") and
                    not _gs_val(r, "Hora_Fin_Real")):
                cub_num = _gs_val(r, "Cubiculo_Num")
                if cub_num:
                    res_activas[cub_num] = r

        for cub in cubiculos_pg:
            num = _gs_val(cub, "Numero")
            if not num:
                continue
            if num in res_activas:
                res = res_activas[num]
                hora_fin    = _gs_val(res, "Hora_Fin_Prog")
                min_rest    = ""
                estado_calc = "ocupado"
                try:
                    fin = datetime.fromisoformat(hora_fin)
                    if fin.tzinfo is None:
                        fin = TZ_COL.localize(fin)
                    diff = (fin - now).total_seconds() / 60
                    min_rest    = str(max(0, int(diff)))
                    estado_calc = "por_liberar" if 0 < diff <= 5 else ("libre" if diff <= 0 else "ocupado")
                except Exception:
                    pass
                conn = get_pg_conn()
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE cubiculos_estado
                        SET estado=%s, cliente_actual=%s, hora_inicio=%s,
                            hora_fin_prog=%s, codigo_acceso=%s, tiempo_rest_min=%s
                        WHERE numero=%s
                    """, [
                        estado_calc,
                        _gs_val(res, "Cliente_Nombre"),
                        _gs_val(res, "Hora_Inicio"),
                        hora_fin,
                        _gs_val(res, "Codigo_Acceso"),
                        min_rest,
                        num,
                    ])
                conn.commit()
        return True
    except Exception as e:
        st.warning(f"⚠️ Error sync cubículos: {e}")
        return False


def gs_sync_dashboard(sh):
    """Actualiza/inserta fila de hoy en dashboard_diario (incluye convenio)."""
    if not PSYCOPG2_AVAILABLE:
        return False
    try:
        today       = ahora_col().strftime("%Y-%m-%d")
        reservas_pg = _pg_read_table("Reservas")
        hoy_res = [r for r in reservas_pg
                   if _gs_fecha_ymd(r, "Creado_En") == today
                   and _gs_val(r, "Estado_Pago") == "confirmado"]

        total_res = len(hoy_res)
        brutos    = sum(_gs_float(r, "Total_COP") for r in hoy_res)
        iva       = sum(_gs_float(r, "IVA")       for r in hoy_res)
        nequi     = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "Nequi")
        daviplata = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "Daviplata")
        efectivo  = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "Efectivo")
        pse       = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "PSE")
        mp        = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "MercadoPago")
        convenio  = sum(_gs_float(r, "Total_COP") for r in hoy_res if _gs_val(r, "Metodo_Pago") == "Convenio")
        otros     = brutos - (nequi + daviplata + efectivo + pse + mp + convenio)
        avg_horas = (sum(_gs_float(r, "Horas_Contratadas") for r in hoy_res) / total_res) if total_res > 0 else 0
        ticket    = round(brutos / total_res, 0) if total_res > 0 else 0

        fila = [
            today, str(total_res), str(total_res), "0",
            str(round(brutos, 2)), str(round(iva, 2)), str(round(brutos - iva, 2)),
            str(round(nequi, 2)), str(round(daviplata, 2)), str(round(efectivo, 2)),
            str(round(pse, 2)), str(round(mp, 2)), str(round(convenio, 2)),
            "0", "", str(round(avg_horas * 60)), "0", "0", "0", "0",
            str(round(ticket, 0)),
        ]
        return _pg_upsert("Dashboard_Diario", "Fecha", today, fila)
    except Exception as e:
        st.session_state["_dashboard_sync_pendiente"] = True
        st.session_state["_dashboard_sync_error"]     = str(e)
        return False


def _gs_sync_dashboard_convenios(sh_conv, nueva_reserva: dict = None):
    """
    Actualiza dashboard_diario_convenio.
    Si se pasa nueva_reserva: acumulación incremental por empresa.
    Si no: recalcula desde reservas_convenio completo.
    """
    if not PSYCOPG2_AVAILABLE:
        return False
    try:
        today      = ahora_col().strftime("%Y-%m-%d")
        id_empresa = (str(nueva_reserva.get("id_empresa", "")) if nueva_reserva else "")

        if nueva_reserva and id_empresa:
            # Leer fila existente
            row_act = _pg_exec(
                'SELECT * FROM dashboard_diario_convenio WHERE fecha=%s AND id_empresa=%s',
                [today, id_empresa], fetch="one"
            ) or {}

            def _fa(k, d=0.0):
                try: return float(row_act.get(k) or d)
                except: return float(d)
            def _ia(k, d=0):
                try: return int(float(row_act.get(k) or d))
                except: return int(d)

            total  = float(nueva_reserva.get("total",  0) or 0)
            iva    = float(nueva_reserva.get("iva",    0) or 0)
            horas  = float(nueva_reserva.get("horas",  0) or 0)
            estado = nueva_reserva.get("estado_pago", "pendiente")

            n_res  = _ia("total_reservas", 0) + 1
            n_conf = _ia("completadas",    0) + (1 if estado == "confirmado" else 0)
            n_pend = _ia("canceladas",     0) + (1 if estado == "pendiente"  else 0)
            brutos = _fa("ingresos_brutos", 0) + total
            iva_t  = _fa("iva_recaudado",   0) + iva
            conv_c = _fa("convenio_cop",    0) + total
            avg_h  = ((_fa("tiempo_prom_min",0)/60 * max(_ia("total_reservas",1)-1,0)) + horas) / n_res if n_res > 0 else horas
            ticket = round(brutos / n_res, 0) if n_res > 0 else total
            min_f  = min(_fa("fact_min", total), total) if _fa("fact_min",0)>0 else total
            max_f  = max(_fa("fact_max", total), total)

            fila = [
                today, id_empresa, str(n_res), str(n_conf), str(n_pend),
                str(round(brutos,2)), str(round(iva_t,2)), str(round(brutos-iva_t,2)),
                "0","0","0","0","0","0", str(round(conv_c,2)),
                "0","", str(round(avg_h*60)), str(n_res), "0",
                str(round(min_f,2)), str(round(max_f,2)), str(int(ticket)),
            ]
        else:
            # Recálculo completo desde reservas_convenio
            rows  = _pg_read_table_conv("Reservas")
            today_rows = [r for r in rows if _gs_fecha_ymd(r,"Creado_En") == today]
            by_empresa: dict = {}
            for r in today_rows:
                eid = _gs_val(r,"Id_Empresa","")
                by_empresa.setdefault(eid, []).append(r)

            for eid, rlist in by_empresa.items():
                n_res  = len(rlist)
                brutos = sum(_gs_float(r,"Total_COP") for r in rlist)
                iva_t  = sum(_gs_float(r,"IVA") for r in rlist)
                conv_c = brutos
                avg_h  = (sum(_gs_float(r,"Horas_Contratadas") for r in rlist)/n_res) if n_res>0 else 0
                ticket = round(brutos/n_res,0) if n_res>0 else 0
                min_f  = min((_gs_float(r,"Total_COP") for r in rlist), default=0)
                max_f  = max((_gs_float(r,"Total_COP") for r in rlist), default=0)
                fila = [
                    today, eid, str(n_res), str(n_res), "0",
                    str(round(brutos,2)), str(round(iva_t,2)), str(round(brutos-iva_t,2)),
                    "0","0","0","0","0","0", str(round(conv_c,2)),
                    "0","", str(round(avg_h*60)), str(n_res), "0",
                    str(round(min_f,2)), str(round(max_f,2)), str(int(ticket)),
                ]
                tabla = "dashboard_diario_convenio"
                cols_pg = _COL_MAP.get(tabla,[])
                fila_str = [str(v) if v is not None else "" for v in fila]
                fila_str = (fila_str + [""]*len(cols_pg))[:len(cols_pg)]
                cols_sql = ", ".join(f'"{c}"' for c in cols_pg)
                ph       = ", ".join(["%s"]*len(cols_pg))
                up       = ", ".join(f'"{c}"=EXCLUDED."{c}"' for c in cols_pg if c not in ("fecha","id_empresa"))
                _pg_exec(
                    f'INSERT INTO {tabla} ({cols_sql}) VALUES ({ph}) ON CONFLICT ("fecha","id_empresa") DO UPDATE SET {up}',
                    fila_str
                )
            return True

        # Insert/update para acumulación incremental
        tabla   = "dashboard_diario_convenio"
        cols_pg = _COL_MAP.get(tabla, [])
        fila_str = [str(v) if v is not None else "" for v in fila]
        fila_str = (fila_str + [""]*len(cols_pg))[:len(cols_pg)]
        cols_sql = ", ".join(f'"{c}"' for c in cols_pg)
        ph       = ", ".join(["%s"]*len(cols_pg))
        up       = ", ".join(f'"{c}"=EXCLUDED."{c}"' for c in cols_pg if c not in ("fecha","id_empresa"))
        _pg_exec(
            f'INSERT INTO {tabla} ({cols_sql}) VALUES ({ph}) ON CONFLICT ("fecha","id_empresa") DO UPDATE SET {up}',
            fila_str
        )
        return True
    except Exception as e:
        st.warning(f"⚠️ Error sync dashboard convenios: {e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN (get_config / set_config — tablas PG)
# ══════════════════════════════════════════════════════════════════════════════

_config_cache: dict = {}
_config_cache_lock = threading.Lock()


def get_config(clave: str, default: str = "") -> str:
    with _config_cache_lock:
        if clave in _config_cache:
            return _config_cache[clave]
    valor = None
    try:
        row = _pg_exec(
            'SELECT valor FROM configuracion_pagos WHERE clave = %s',
            [clave], fetch="one"
        )
        if row:
            valor = str(row.get("valor", default))
    except Exception:
        pass
    if valor is None:
        valor = default
    with _config_cache_lock:
        _config_cache[clave] = valor
    return valor


def set_config(clave: str, valor: str):
    try:
        _pg_upsert("Configuracion_Pagos", "Clave", clave, [clave, valor])
        _gs_invalidate_cache("Configuracion_Pagos")
    except Exception:
        pass
    with _config_cache_lock:
        _config_cache[clave] = valor


# ── Configuración por empresa (equivale a Configuracion_Pagos de jjgt_convenios) ─

def get_config_empresa_convenio(sh_conv, id_empresa: str) -> dict:
    """
    Lee la configuración de pagos/tarifas para una empresa específica desde
    configuracion_empresa_convenio. Fallback a configuracion_pagos global.
    """
    defaults = {
        "convenios_dias_credito":   float(get_config("convenios_dias_credito",  "30")),
        "convenios_horas_min":      float(get_config("convenios_horas_min",     "1.0")),
        "convenios_descuento_pct":  float(get_config("convenios_descuento_pct", "0")),
        "convenios_max_pendientes": int(  get_config("convenios_max_pendientes","10")),
    }
    if not id_empresa:
        return defaults
    try:
        rows = _pg_exec(
            'SELECT clave, valor FROM configuracion_empresa_convenio WHERE id_empresa = %s',
            [id_empresa], fetch="all"
        ) or []
        result = dict(defaults)
        for row in rows:
            clave = str(row.get("clave","")).strip()
            valor = str(row.get("valor","")).strip()
            if clave in result:
                try:
                    if clave == "convenios_max_pendientes":
                        result[clave] = int(float(valor))
                    else:
                        result[clave] = float(valor)
                except (ValueError, TypeError):
                    pass
        return result
    except Exception:
        return defaults


def upsert_config_empresa_convenio(sh_conv, id_empresa: str, nombre_empresa: str,
                                    configs: dict) -> bool:
    """
    Guarda/actualiza la configuración de una empresa en
    configuracion_empresa_convenio. Una fila por clave por empresa.
    configs = {"convenios_dias_credito": 30, "convenios_horas_min": 1.0, ...}
    """
    if not id_empresa:
        return False
    now_iso = ahora_col().isoformat()
    ok = True
    for clave, valor in configs.items():
        fila = [id_empresa, nombre_empresa, str(clave), str(valor), now_iso]
        tabla    = "configuracion_empresa_convenio"
        cols_pg  = _COL_MAP.get(tabla, [])
        fila_str = [str(v) if v is not None else "" for v in fila]
        fila_str = (fila_str + [""]*len(cols_pg))[:len(cols_pg)]
        cols_sql = ", ".join(f'"{c}"' for c in cols_pg)
        ph       = ", ".join(["%s"]*len(cols_pg))
        up       = ", ".join(f'"{c}"=EXCLUDED."{c}"' for c in cols_pg if c not in ("id_empresa","clave"))
        try:
            _pg_exec(
                f'INSERT INTO {tabla} ({cols_sql}) VALUES ({ph}) '
                f'ON CONFLICT ("id_empresa","clave") DO UPDATE SET {up}',
                fila_str
            )
        except Exception as e:
            st.warning(f"⚠️ Error config empresa convenio: {e}")
            ok = False
    return ok


# ══════════════════════════════════════════════════════════════════════════════
# OPERADORES (lectura y autenticación)
# ══════════════════════════════════════════════════════════════════════════════

def _hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()


def verificar_pin(pin: str, rol: str = None) -> bool:
    return get_operador_por_pin(pin, rol) is not None


def get_operador_por_pin(pin: str, rol: str = None) -> dict:
    """
    Autentica un operador por PIN.
    PINs almacenados como hashes SHA-256 en session_state o secrets.
    Si la tabla operadores está vacía, acepta PINs de bootstrap.
    """
    pin_hash = _hash_pin(pin)

    PINS_SEED = {
        "Admin JJGT": "6789",
        "Op. Mañana": "0001",
        "Op. Tarde":  "0002",
        "Op. Noche":  "0003",
    }

    def _build_result(r, nombre_op):
        op_rol = _gs_val(r, "Rol", "cajero")
        if rol and op_rol != rol:
            return None
        return {
            "id":                _gs_val(r, "ID_Operador", "0"),
            "nombre":            nombre_op,
            "rol":               op_rol,
            "turno":             _gs_val(r, "Turno", "diurno"),
            "hora_inicio_turno": _gs_val(r, "Hora_Ini_Turno", "06:00"),
            "hora_fin_turno":    _gs_val(r, "Hora_Fin_Turno", "14:00"),
            "permisos":          _gs_val(r, "Permisos", "reservas,pagos,voucher").split(","),
        }

    rows = _pg_read_table("Operadores")
    rows_activos = [r for r in rows
                    if _gs_val(r, "Activo", "1") in ("1", "True", "true", "1.0")]

    pin_hashes = st.session_state.get("_op_pin_hashes", {})
    for r in rows_activos:
        nombre = _gs_val(r, "Nombre")
        if pin_hashes.get(nombre) == pin_hash:
            result = _build_result(r, nombre)
            if result:
                return result

    try:
        pins_secrets = st.secrets.get("operadores_pins", {})
        for nombre_s, pin_s in pins_secrets.items():
            if _hash_pin(str(pin_s)) == pin_hash:
                pin_hashes[nombre_s] = pin_hash
                st.session_state["_op_pin_hashes"] = pin_hashes
                for r in rows_activos:
                    if _gs_val(r, "Nombre") == nombre_s:
                        result = _build_result(r, nombre_s)
                        if result:
                            return result
                if not rol or rol == "admin":
                    return {
                        "id": "0", "nombre": nombre_s, "rol": "admin",
                        "turno": "diurno", "hora_inicio_turno": "06:00",
                        "hora_fin_turno": "22:00",
                        "permisos": ["admin","reservas","pagos","voucher","reportes","configuracion"],
                    }
    except Exception:
        pass

    for nombre_seed, pin_seed in PINS_SEED.items():
        if _hash_pin(pin_seed) == pin_hash:
            for r in rows_activos:
                if _gs_val(r, "Nombre") == nombre_seed:
                    result = _build_result(r, nombre_seed)
                    if result:
                        pin_hashes[nombre_seed] = pin_hash
                        st.session_state["_op_pin_hashes"] = pin_hashes
                        return result
            if nombre_seed == "Admin JJGT" and (not rol or rol == "admin"):
                pin_hashes[nombre_seed] = pin_hash
                st.session_state["_op_pin_hashes"] = pin_hashes
                return {
                    "id": "1", "nombre": "Admin JJGT", "rol": "admin",
                    "turno": "diurno", "hora_inicio_turno": "06:00",
                    "hora_fin_turno": "22:00",
                    "permisos": ["admin","reservas","pagos","voucher","reportes","configuracion"],
                }

    if not rows_activos:
        if pin in ("admin", "6789", "0000"):
            if not rol or rol == "admin":
                return {
                    "id": "0", "nombre": "Admin (emergencia)", "rol": "admin",
                    "turno": "diurno", "hora_inicio_turno": "00:00",
                    "hora_fin_turno": "23:59",
                    "permisos": ["admin","reservas","pagos","voucher","reportes","configuracion"],
                }

    return None


# ══════════════════════════════════════════════════════════════════════════════
# EMPRESAS DE CONVENIO — lectura/búsqueda
# ══════════════════════════════════════════════════════════════════════════════

def get_empresas_convenio() -> list:
    """Lee empresas activas de la tabla empresas_convenio en PostgreSQL."""
    try:
        rows = _pg_read_table_conv("Empresa")
        return [r for r in rows if _gs_val(r, "Activo", "1") in ("1","","True","true","SI","si")]
    except Exception:
        return []


# ══════════════════════════════════════════════════════════════════════════════
# STUBS DE COMPATIBILIDAD
# ══════════════════════════════════════════════════════════════════════════════

def _restore_from_sheets() -> bool:
    return True

def _seed_data(cur):
    pass

def encolar_sync(event: dict):
    pass

def _ensure_sync_thread():
    pass

def _marcar_sync_pendiente(tabla: str, registro_id: int):
    pass

def sheets_append_row(worksheet_name: str, row_data: list) -> bool:
    return _pg_append(worksheet_name, row_data)

def sheets_upsert_row(worksheet_name, col_key, key_value, row_data):
    return _pg_upsert(worksheet_name, col_key, key_value, row_data)

def sheets_update_row(worksheet_name, col_key, key_value, update_data):
    try:
        for col_name, new_val in update_data.items():
            _pg_update_col(worksheet_name, col_key, key_value, col_name, str(new_val))
        return True
    except Exception:
        return False

def load_data_from_sheet(client, sheet_name: str, worksheet_name: str) -> pd.DataFrame:
    data = _pg_read_table(worksheet_name)
    return pd.DataFrame(data) if data else pd.DataFrame()

def gs_sync_factura_items(sh):
    return True

def gs_sync_operadores(sh):
    return True

def gs_sync_configuracion_pagos(sh):
    with _config_cache_lock:
        _config_cache.clear()
    return True


def sincronizacion_completa() -> dict:
    """Verifica conexión PG y retorna conteo de registros por tabla (pagos + convenios)."""
    with _config_cache_lock:
        _config_cache.clear()
    conteos = {}
    for hoja in ["Clientes","Reservas","Pagos","Facturas",
                 "Factura_items","Cubiculos_Estado","Tarifas_Config",
                 "Operadores","Configuracion_Pagos"]:
        try:
            data = _pg_read_table(hoja)
            conteos[hoja] = len(data)
        except Exception as e:
            conteos[hoja] = f"Error: {e}"
    for hoja_c in ["Empresa","Reservas","Pagos","Clientes","Facturas"]:
        try:
            data_c = _pg_read_table_conv(hoja_c)
            conteos[f"Conv_{hoja_c}"] = len(data_c)
        except Exception as e:
            conteos[f"Conv_{hoja_c}"] = f"Error: {e}"
    _, sh = get_active_client()
    gs_sync_dashboard(sh)
    gs_escribir_log(sh, "sync_completa", "", "", "sistema",
                    f"Sync completa PG (pagos+convenios): {conteos}", "exito")
    return conteos


# Variables globales stub para compatibilidad con referencias residuales
_gs_module_client      = None
_gs_module_spreadsheet = None
_gs_cached_creds       = None

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE NEGOCIO
# ══════════════════════════════════════════════════════════════════════════════

def ahora_col() -> datetime:
    return datetime.now(TZ_COL)


def calcular_precio(horas: float, tarifa_nombre: str = None,
                    precio_override: Optional[float] = None) -> dict:
    """Calcula precio con tarifa vigente desde PostgreSQL.

    LÓGICA:
    - 1 hora  = precio completo (ej: 15.000)
    - ≥2 horas = 10.000 por hora
    """
    if tarifa_nombre is None:
        hora_actual = ahora_col().hour
        tarifa_nombre = "Madrugada" if 0 <= hora_actual < 6 else "Estándar"

    rows = _gs_read_sheet("Tarifas_Config")

    precio_hora    = 15000
    hora_ini_espec = ""
    hora_fin_espec = ""
    tarifa_row     = None

    for r in rows:
        if (
            _gs_val(r, "Nombre") == tarifa_nombre
            and _gs_val(r, "Activo", "1") in ("1", "True", "true")
        ):
            tarifa_row = r
            precio_hora    = _gs_float(r, "Precio_Hora_COP", 15000)
            hora_ini_espec = _gs_val(r, "Hora_Ini_Espec", "")
            hora_fin_espec = _gs_val(r, "Hora_Fin_Espec", "")
            break

    if precio_override is not None and precio_override > 0:
        precio_hora = precio_override

    es_noche_completa = (tarifa_nombre == "Noche Completa")

    if es_noche_completa:
        if hora_ini_espec and hora_fin_espec:
            try:
                now = ahora_col()
                fmt = "%H:%M"
                ini_t = datetime.strptime(hora_ini_espec, fmt)
                fin_t = datetime.strptime(hora_fin_espec, fmt)
                ini_dt = now.replace(hour=ini_t.hour, minute=ini_t.minute, second=0, microsecond=0)
                fin_dt = now.replace(hour=fin_t.hour, minute=fin_t.minute, second=0, microsecond=0)
                if fin_dt <= ini_dt:
                    fin_dt += timedelta(days=1)
                horas = round((fin_dt - ini_dt).total_seconds() / 3600, 2)
            except Exception:
                pass
        subtotal_bruto = round(precio_hora, 0)
        subtotal = subtotal_bruto
        iva      = round(subtotal * IVA_PCT, 0)
        total    = round(subtotal + iva, 0)
        return {
            "precio_hora":       precio_hora,
            "horas":             horas,
            "tarifa":            tarifa_nombre,
            "descuento_pct":     0,
            "descuento_val":     0,
            "subtotal":          subtotal,
            "iva":               iva,
            "total":             round(total, 0),
            "hora_ini_espec":    hora_ini_espec,
            "hora_fin_espec":    hora_fin_espec,
            "es_noche_completa": True,
        }

    horas    = int(horas)
    subtotal = precio_hora if horas == 1 else (horas * 10000 if horas > 1 else 0)
    subtotal = round(subtotal, 0)
    iva      = round(subtotal * IVA_PCT, 0)
    total    = round(subtotal + iva, 0)
    return {
        "precio_hora":       precio_hora,
        "horas":             horas,
        "tarifa":            tarifa_nombre,
        "descuento_pct":     0,
        "descuento_val":     0,
        "subtotal":          subtotal,
        "iva":               iva,
        "total":             total,
        "hora_ini_espec":    hora_ini_espec,
        "hora_fin_espec":    hora_fin_espec,
        "es_noche_completa": False,
    }


def calcular_precio_convenio(horas: float, cfg_empresa: dict,
                              id_empresa: str = "") -> dict:
    """
    Calcula precio para reserva de convenio con 3 niveles de tarifa:
    1. Tarifa específica de la empresa en tarifas_config_convenio
    2. Tarifa genérica [Convenio] en tarifas_config
    3. Fallback: Estándar/Madrugada de tarifas_config
    Aplica descuento por volumen + descuento adicional del convenio en cascada.
    """
    hora_actual       = ahora_col().hour
    tarifa_base       = "Madrugada" if 0 <= hora_actual < 6 else "Estándar"
    precio_hora       = 15000
    desc_3h           = 0
    desc_6h           = 0
    tarifa_nombre_log = tarifa_base
    tarifa_encontrada = False

    # ── Nivel 1: tarifa específica de la empresa en tarifas_config_convenio ───
    if id_empresa:
        try:
            rows_tar = _pg_read_table_conv("Tarifas_Config")
            for d_tar in rows_tar:
                activo    = str(d_tar.get("Activo","1")).strip() in ("1","True","true")
                emp_match = str(d_tar.get("Id_Empresa","")).strip() == id_empresa.strip()
                if not activo or not emp_match:
                    continue
                h_ini = str(d_tar.get("Hora_Ini_Espec","")).strip()
                h_fin = str(d_tar.get("Hora_Fin_Espec","")).strip()
                en_horario = True
                if h_ini and h_fin:
                    try:
                        hi = int(h_ini.split(":")[0])
                        hf = int(h_fin.split(":")[0])
                        en_horario = (hora_actual >= hi or hora_actual < hf) if hi >= hf \
                                      else (hi <= hora_actual < hf)
                    except Exception:
                        en_horario = True
                if not en_horario:
                    continue
                precio_hora       = float(d_tar.get("Precio_Hora_COP") or precio_hora)
                desc_3h           = float(d_tar.get("Desc_3h_Pct") or 0)
                desc_6h           = float(d_tar.get("Desc_6h_Pct") or 0)
                tarifa_nombre_log = str(d_tar.get("Empresa_Nombre") or f"Empresa {id_empresa}")
                tarifa_encontrada = True
                break
        except Exception:
            pass

    # ── Nivel 2: tarifa genérica [Convenio] en tarifas_config ─────────────────
    if not tarifa_encontrada:
        for r in _gs_read_sheet("Tarifas_Config"):
            nombre = _gs_val(r, "Nombre", "")
            activo = _gs_val(r, "Activo", "1") in ("1","True","true")
            if "[Convenio]" in nombre and activo:
                precio_hora       = _gs_float(r, "Precio_Hora_COP", precio_hora)
                desc_3h           = _gs_float(r, "Desc_3h_Pct", 0)
                desc_6h           = _gs_float(r, "Desc_6h_Pct", 0)
                tarifa_nombre_log = nombre
                tarifa_encontrada = True
                break

    # ── Nivel 3: tarifa base Estándar/Madrugada ────────────────────────────────
    if not tarifa_encontrada:
        for r in _gs_read_sheet("Tarifas_Config"):
            if _gs_val(r,"Nombre") == tarifa_base and _gs_val(r,"Activo","1") in ("1","True","true"):
                precio_hora       = _gs_float(r,"Precio_Hora_COP", 15000)
                desc_3h           = _gs_float(r,"Desc_3h_Pct", 0)
                desc_6h           = _gs_float(r,"Desc_6h_Pct", 0)
                tarifa_nombre_log = tarifa_base
                break

    desc_vol_pct        = desc_6h if horas >= 6 else (desc_3h if horas >= 2 else 0.0)
    desc_conv_pct       = float(cfg_empresa.get("convenios_descuento_pct", 0))
    subtotal_bruto      = round(precio_hora * horas, 2)
    desc_vol_val        = round(subtotal_bruto * (desc_vol_pct  / 100), 2)
    after_vol           = subtotal_bruto - desc_vol_val
    desc_conv_val       = round(after_vol   * (desc_conv_pct / 100), 2)
    subtotal            = round(after_vol   - desc_conv_val, 2)
    descuento_total_val = desc_vol_val + desc_conv_val
    descuento_total_pct = round(descuento_total_val / subtotal_bruto * 100, 2) if subtotal_bruto > 0 else 0
    iva                 = round(subtotal * IVA_PCT, 2)
    total               = round(subtotal + iva, 2)

    return {
        "precio_hora":        precio_hora,
        "horas":              horas,
        "tarifa":             tarifa_nombre_log,
        "descuento_pct":      descuento_total_pct,
        "descuento_vol_pct":  desc_vol_pct,
        "descuento_conv_pct": desc_conv_pct,
        "descuento_val":      descuento_total_val,
        "subtotal":           subtotal,
        "iva":                iva,
        "total":              total,
    }


def get_cubiculos() -> list:
    """Lee cubículos desde PostgreSQL (cubiculos_estado)."""
    rows = _gs_read_sheet("Cubiculos_Estado")
    now  = ahora_col()
    result = []
    for i, row in enumerate(rows):
        numero = _gs_val(row, "Numero")
        if not numero:
            continue
        estado      = _gs_val(row, "Estado", "libre")
        hora_fin    = _gs_val(row, "Hora_Fin_Prog")
        hora_inicio = _gs_val(row, "Hora_Inicio")
        wifi_ssid   = _gs_val(row, "WiFi_SSID")
        wifi_pass   = _gs_val(row, "WiFi_Pass")
        precio_base = _gs_float(row, "Precio_Hora_Base", 15000)
        cub_id      = _gs_val(row, "Cubiculo_ID") or str(i + 1)

        cub = {
            "id":                cub_id,
            "numero":            numero,
            "nombre":            _gs_val(row, "Nombre", f"Cubículo {numero}"),
            "estado":            estado,
            "reserva_activa_id": None,
            "hora_disponible":   hora_fin,
            "precio_hora_base":  precio_base,
            "wifi_ssid":         wifi_ssid,
            "wifi_password":     wifi_pass,
            "servicios":         ["Baño", "WiFi", "Carga USB", "Carga 110V"],
            "hora_fin":          hora_fin,
            "hora_inicio":       hora_inicio,
            "minutos_restantes": None,
            "cliente_actual":    _gs_val(row, "Cliente_Actual"),
        }

        if hora_fin and estado in ("ocupado", "por_liberar"):
            try:
                fin = datetime.fromisoformat(hora_fin)
                if fin.tzinfo is None:
                    fin = TZ_COL.localize(fin)
                diff = (fin - now).total_seconds() / 60
                cub["minutos_restantes"] = max(0, int(diff))
                if 0 < diff <= 5:
                    cub["estado"] = "por_liberar"
                elif diff <= 0:
                    cub["estado"] = "por_liberar"
                    cub["minutos_restantes"] = 0
            except Exception:
                pass

        result.append(cub)

    result.sort(key=lambda c: c["numero"])
    return result


def get_cubiculos_libres() -> list:
    return [c for c in get_cubiculos() if c["estado"] == "libre"]


def generar_numero_reserva() -> str:
    now    = ahora_col()
    prefix = f"CR-{now.strftime('%Y%m%d')}-"
    try:
        rows = _gs_read_sheet("Reservas")
        n = sum(1 for r in rows if _gs_val(r, "Numero_Reserva", "").startswith(prefix))
        # También contar en reservas de convenio
        rows_c = _pg_read_table_conv("Reservas")
        n += sum(1 for r in rows_c if _gs_val(r, "Numero_Reserva", "").startswith(prefix))
    except Exception:
        n = 0
    return f"{prefix}{(n + 1):04d}"


def generar_numero_factura() -> str:
    """Genera número de factura incrementando el contador en configuracion_pagos."""
    prefix = get_config("factura_prefijo", "FACT")
    year   = ahora_col().year
    n = 1
    try:
        row = _pg_exec(
            "SELECT valor FROM configuracion_pagos WHERE clave = 'factura_contador'",
            fetch="one"
        )
        if row:
            n = int(float(row.get("valor", 0))) + 1
    except Exception:
        pass
    try:
        _pg_upsert("Configuracion_Pagos", "Clave", "factura_contador",
                   ["factura_contador", str(n)])
        _gs_invalidate_cache("Configuracion_Pagos")
    except Exception:
        pass
    with _config_cache_lock:
        _config_cache["factura_contador"] = str(n)
    return f"{prefix}-{year}-{n:04d}"


def generar_codigo_acceso() -> str:
    import random
    return str(random.randint(1000, 9999))


def activar_cubiculo(cubiculo_id, reserva_id, hora_fin: str):
    """Activa cubículo actualizando estado en PostgreSQL."""
    try:
        _, _sh = get_active_client()
        gs_sync_cubiculos(_sh)
        _gs_invalidate_cache("Cubiculos_Estado")
    except Exception:
        pass


def liberar_cubiculo(cubiculo_id):
    """
    Libera el cubículo actualizando PostgreSQL como fuente primaria.
    Actualiza tanto la tabla reservas (y reservas_convenio) como cubiculos_estado.
    """
    hora_fin_real = ahora_col().isoformat()

    # Obtener número del cubículo
    cubiculos_pg = _pg_read_table("Cubiculos_Estado")
    num_cub = ""
    for r in cubiculos_pg:
        if _gs_val(r, "Cubiculo_ID") == str(cubiculo_id):
            num_cub = _gs_val(r, "Numero")
            break

    # Buscar reserva activa en jjgt_pagos (método normal)
    num_reserva = ""
    reservas_pg = _pg_read_table("Reservas")
    for r in reservas_pg:
        if (_gs_val(r, "Cubiculo_Num") == num_cub and
                _gs_val(r, "Estado_Pago") in ("confirmado", "pendiente") and
                not _gs_val(r, "Hora_Fin_Real")):
            num_reserva = _gs_val(r, "Numero_Reserva")
            break

    # Buscar también en reservas_convenio
    num_reserva_conv = ""
    if not num_reserva:
        reservas_conv = _pg_read_table_conv("Reservas")
        for r in reservas_conv:
            if (_gs_val(r, "Cubiculo_Num") == num_cub and
                    _gs_val(r, "Estado_Pago") in ("confirmado", "pendiente") and
                    not _gs_val(r, "Hora_Fin_Real")):
                num_reserva_conv = _gs_val(r, "Numero_Reserva")
                break

    try:
        conn = get_pg_conn()
        with conn.cursor() as cur:
            if num_reserva:
                cur.execute(
                    'UPDATE reservas SET hora_fin_real=%s WHERE numero_reserva=%s',
                    [hora_fin_real, num_reserva]
                )
            if num_reserva_conv:
                cur.execute(
                    'UPDATE reservas_convenio SET hora_fin_real=%s WHERE numero_reserva=%s',
                    [hora_fin_real, num_reserva_conv]
                )
            if num_cub:
                cur.execute("""
                    UPDATE cubiculos_estado
                    SET estado='libre', cliente_actual='', hora_inicio='',
                        hora_fin_prog='', codigo_acceso='', tiempo_rest_min=''
                    WHERE numero=%s
                """, [num_cub])
        conn.commit()
    except Exception as e:
        st.warning(f"⚠️ Error liberando cubículo: {e}")

    _, sh = get_active_client()
    gs_sync_dashboard(sh)


def registrar_en_facturacion(reserva: dict, cliente: dict) -> tuple:
    """
    Escribe cliente, factura e ítems directamente en PostgreSQL.
    Retorna (numero_factura, num_factura_str).
    """
    now     = ahora_col()
    numf    = generar_numero_factura()
    num_doc = cliente.get("numero_documento", "")
    descripcion_item = (f"Espacio de descanso {reserva['cubiculo_num']} · "
                        f"WiFi · Baño · Carga · {reserva['horas']}h")

    _, sh = get_active_client()

    gs_escribir_cliente(sh, {
        "id":               "",
        "nombre":           cliente["nombre"],
        "tipo_documento":   cliente.get("tipo_doc", "CC"),
        "numero_documento": num_doc,
        "telefono":         cliente.get("telefono", ""),
        "email":            cliente.get("email", ""),
        "ciudad":           cliente.get("ciudad", ""),
        "regimen":          cliente.get("regimen", "Simplificado"),
        "tipo_persona":     "Jurídica" if cliente.get("razon_social") else "Natural",
        "razon_social":     cliente.get("razon_social", ""),
        "nit_empresa":      cliente.get("nit_empresa", ""),
        "activo":           "1",
        "creado_en":        now.isoformat(),
    })
    gs_escribir_factura(sh, {
        "id":               numf,
        "numero":           numf,
        "tipo":             "Factura de Venta",
        "fecha_emision":    now.strftime("%Y-%m-%d"),
        "fecha_vencimiento":now.strftime("%Y-%m-%d"),
        "cliente_nombre":   cliente["nombre"],
        "cliente_doc":      num_doc,
        "cliente_email":    cliente.get("email", ""),
        "razon_social":     cliente.get("razon_social", ""),
        "nit_empresa":      cliente.get("nit_empresa", ""),
        "descripcion":      descripcion_item,
        "subtotal":         reserva["subtotal"],
        "descuento":        reserva.get("descuento_val", 0),
        "iva":              reserva["iva"],
        "retenciones":      "0",
        "total":            reserva["total"],
        "metodo_pago":      reserva["metodo_pago"],
        "estado":           "pagada",
        "moneda":           "COP",
        "num_reserva":      "",
        "cubiculo":         reserva["cubiculo_num"],
        "creado_en":        now.isoformat(),
        "actualizado_en":   now.isoformat(),
    })
    _pg_append("Factura_items", [
        numf, numf,
        "DESCANSO-H",
        descripcion_item,
        str(reserva["horas"]),
        "hora",
        str(reserva["precio_hora"]),
        "0", "19.0",
        str(reserva["subtotal"]),
    ])
    _gs_invalidate_cache("Facturas", "Clientes", "Factura_items")
    return numf, numf


def crear_reserva_completa(cubiculo: dict, cliente: dict, calc: dict, metodo: str,
                           acepto_datos: bool = False) -> dict:
    """
    Crea reserva, pago, factura y activa el cubículo.
    Todo se escribe directamente en PostgreSQL.
    """
    if (st.session_state.get("voucher") and
            st.session_state.get("pago_confirmado") and
            st.session_state["voucher"].get("cubiculo") == cubiculo["numero"]):
        return st.session_state["voucher"]

    now     = ahora_col()
    hoy_str = now.strftime("%Y-%m-%d")
    num_doc = cliente.get("numero_documento", "")

    # ── Verificación 1: cubículo sigue libre ──────────────────────────────────
    cubiculos_now = get_cubiculos()
    cub_actual = next((c for c in cubiculos_now if c["numero"] == cubiculo["numero"]), None)
    if cub_actual and cub_actual["estado"] != "libre":
        raise Exception(f"El cubículo {cubiculo['numero']} ya no está disponible.")

    # ── Verificación 2: cliente sin reserva activa hoy ────────────────────────
    if num_doc:
        reservas_pg = _pg_read_table("Reservas")
        for r in reservas_pg:
            if (str(r.get("Documento", "")) == num_doc and
                    _gs_val(r, "Estado_Pago") == "confirmado" and
                    not _gs_val(r, "Hora_Fin_Real") and
                    _gs_fecha_ymd(r, "Hora_Inicio") == hoy_str):
                raise Exception(
                    f"El cliente con documento {num_doc} ya tiene la reserva activa "
                    f"{_gs_val(r, 'Numero_Reserva')} en el cubículo {_gs_val(r, 'Cubiculo_Num')}. "
                    f"No se puede crear una nueva reserva hasta que termine la actual."
                )

    # ── Hora inicio / fin ─────────────────────────────────────────────────────
    if calc.get("es_noche_completa") and calc.get("hora_ini_espec") and calc.get("hora_fin_espec"):
        try:
            fmt   = "%H:%M"
            ini_t = datetime.strptime(calc["hora_ini_espec"], fmt)
            fin_t = datetime.strptime(calc["hora_fin_espec"], fmt)
            now_nc  = now.replace(hour=ini_t.hour, minute=ini_t.minute, second=0, microsecond=0)
            hora_fin = now_nc.replace(hour=fin_t.hour, minute=fin_t.minute, second=0, microsecond=0)
            if hora_fin <= now_nc:
                hora_fin += timedelta(days=1)
            now = now_nc
        except Exception:
            hora_fin = now + timedelta(hours=calc["horas"])
    else:
        hora_fin = now + timedelta(hours=calc["horas"])

    num_res = generar_numero_reserva()
    codigo  = generar_codigo_acceso()

    reserva_data = {
        "cubiculo_num":  cubiculo["numero"],
        "horas":         calc["horas"],
        "precio_hora":   calc["precio_hora"],
        "subtotal":      calc["subtotal"],
        "descuento_val": calc["descuento_val"],
        "iva":           calc["iva"],
        "total":         calc["total"],
        "metodo_pago":   metodo,
    }

    num_fact, _ = registrar_en_facturacion(reserva_data, cliente)

    _, sh = get_active_client()
    gs_escribir_reserva(sh, {
        "id":              num_res,
        "numero_reserva":  num_res,
        "creado_en":       now.isoformat(),
        "cubiculo_num":    cubiculo["numero"],
        "cliente_nombre":  cliente["nombre"],
        "cliente_doc":     num_doc,
        "cliente_tel":     cliente.get("telefono", ""),
        "cliente_email":   cliente.get("email", ""),
        "horas":           calc["horas"],
        "hora_inicio":     now.isoformat(),
        "hora_fin_prog":   hora_fin.isoformat(),
        "hora_fin_real":   "",
        "precio_hora":     calc["precio_hora"],
        "subtotal":        calc["subtotal"],
        "iva":             calc["iva"],
        "total":           calc["total"],
        "metodo_pago":     metodo,
        "estado_pago":     "confirmado",
        "codigo_acceso":   codigo,
        "wifi_ssid":       cubiculo.get("wifi_ssid", ""),
        "wifi_pass":       cubiculo.get("wifi_password", ""),
        "num_factura":     num_fact,
        "referencia_pago": "",
        "operador":        st.session_state.get("operador_info", {}).get("nombre", "sistema"),
        "notas":           "",
        "acepto_datos":    acepto_datos,
    })
    gs_escribir_pago(sh, {
        "id":                 num_res,
        "reserva_id":         num_res,
        "num_reserva":        num_res,
        "fecha_pago":         now.isoformat(),
        "monto":              calc["total"],
        "metodo":             metodo,
        "referencia_externa": "",
        "estado":             "confirmado",
        "confirmado_por":     "sistema",
        "notas":              "",
    })
    gs_escribir_log(sh, "nueva_reserva", num_res, cubiculo["numero"],
                    st.session_state.get("operador_info", {}).get("nombre", "sistema"),
                    f"Reserva {num_res} | Factura {num_fact} | {metodo} | ${calc['total']:,.0f}")

    # Actualizar estado del cubículo a ocupado en PG
    try:
        conn = get_pg_conn()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE cubiculos_estado
                SET estado='ocupado', cliente_actual=%s,
                    hora_inicio=%s, hora_fin_prog=%s, codigo_acceso=%s
                WHERE numero=%s
            """, [
                cliente["nombre"],
                now.isoformat(),
                hora_fin.isoformat(),
                codigo,
                cubiculo["numero"],
            ])
        conn.commit()
    except Exception as e:
        st.warning(f"⚠️ Error actualizando cubículo: {e}")

    _gs_invalidate_cache("Reservas", "Cubiculos_Estado", "Pagos")
    gs_sync_dashboard(sh)
    activar_cubiculo(cubiculo["id"], num_res, hora_fin.isoformat())

    if CONT_AVAILABLE and _cont_mod is not None:
        try:
            _cont_mod.on_reserva_creada(
                voucher={"numero_reserva": num_res, "numero_factura": num_fact, "cubiculo": cubiculo["numero"]},
                calc=calc, cliente=cliente, metodo=metodo,
            )
        except Exception:
            pass

    return {
        "numero_reserva": num_res,
        "numero_factura": num_fact,
        "cubiculo":       cubiculo["numero"],
        "codigo_acceso":  codigo,
        "wifi_ssid":      cubiculo["wifi_ssid"],
        "wifi_password":  cubiculo["wifi_password"],
        "hora_inicio":    now.strftime("%H:%M"),
        "hora_fin":       hora_fin.strftime("%H:%M"),
        "horas":          calc["horas"],
        "metodo_pago":    metodo,
        "subtotal":       calc["subtotal"],
        "iva":            calc["iva"],
        "total":          calc["total"],
        "cliente_nombre": cliente["nombre"],
    }


def crear_reserva_convenio(cubiculo: dict, cliente: dict, calc: dict,
                            id_empresa: str, nombre_empresa: str, operador: str = "sistema") -> dict:
    """
    Crea reserva de convenio.
    ► Escribe en reservas_convenio, pagos_convenio, clientes_convenio,
      facturas_convenio, factura_items_convenio (PostgreSQL).
    ► Actualiza cubiculos_estado (compartida entre pagos y convenios).
    ► El pago queda en estado PENDIENTE — se cobra a la empresa.
    """
    now         = ahora_col()
    num_res     = generar_numero_reserva()
    num_fac     = generar_numero_factura()
    cod_acc     = generar_codigo_acceso()
    hora_inicio = now
    hora_fin    = now + timedelta(hours=calc["horas"])
    diff_min    = int((hora_fin - now).total_seconds() / 60)

    # Verificar disponibilidad
    cubiculos_now = get_cubiculos()
    cub_actual = next((c for c in cubiculos_now if c["numero"] == cubiculo["numero"]), None)
    if cub_actual and cub_actual["estado"] != "libre":
        raise Exception(f"El cubículo {cubiculo['numero']} ya no está disponible.")

    descripcion_item = (f"Convenio {nombre_empresa} · Cubículo {cubiculo['numero']} · "
                        f"WiFi · Baño · Carga · {calc['horas']}h")

    reserva_data = {
        "id":              num_res,
        "numero_reserva":  num_res,
        "creado_en":       now.isoformat(),
        "cubiculo_num":    cubiculo["numero"],
        "cliente_nombre":  cliente["nombre"],
        "cliente_doc":     cliente.get("numero_documento", ""),
        "cliente_tel":     cliente.get("telefono", ""),
        "cliente_email":   cliente.get("email", ""),
        "horas":           calc["horas"],
        "hora_inicio":     hora_inicio.isoformat(),
        "hora_fin_prog":   hora_fin.isoformat(),
        "hora_fin_real":   "",
        "precio_hora":     calc["precio_hora"],
        "subtotal":        calc["subtotal"],
        "iva":             calc["iva"],
        "total":           calc["total"],
        "metodo_pago":     "Convenio",
        "estado_pago":     "pendiente",
        "codigo_acceso":   cod_acc,
        "wifi_ssid":       cubiculo.get("wifi_ssid", ""),
        "wifi_pass":       cubiculo.get("wifi_password", ""),
        "num_factura":     num_fac,
        "referencia_pago": f"CONV-{id_empresa}-{num_res}",
        "operador":        operador,
        "notas":           f"Convenio: {nombre_empresa} | Desc: {calc.get('descuento_pct',0):.1f}%",
        "id_empresa":      id_empresa,
        "nombre_empresa":  nombre_empresa,
    }

    pago_data = {
        "id":                 f"CONV-{int(time.time())}",
        "reserva_id":         num_res,
        "num_reserva":        num_res,
        "fecha_pago":         now.isoformat(),
        "monto":              calc["total"],
        "metodo":             "Convenio",
        "referencia_externa": f"CONV-{id_empresa}",
        "estado":             "pendiente",
        "confirmado_por":     "convenio",
        "notas":              f"Empresa: {nombre_empresa}",
        "id_empresa":         id_empresa,
        "nombre_empresa":     nombre_empresa,
    }

    if CONT_AVAILABLE and _cont_mod is not None:
        try:
            _voucher_conv = {
                "numero_reserva": num_res,
                "numero_factura": num_fac,
                "cubiculo":       cubiculo["numero"],
            }
            _cont_mod.on_reserva_convenio_creada(_voucher_conv, calc, cliente, "Convenio")
        except Exception as e:
            print(f"[contabilidad] ERROR on_reserva_convenio_creada: {e}")

    _, sh_conv = get_active_client_convenios()

    # ── 1. Escribir en tablas de convenio (PostgreSQL) ────────────────────────
    gs_escribir_reserva_convenio(sh_conv, reserva_data)
    gs_escribir_pago_convenio(sh_conv, pago_data)
    gs_escribir_cliente_convenio(sh_conv, {
        "nombre":           cliente["nombre"],
        "tipo_documento":   cliente.get("tipo_doc", "CC"),
        "numero_documento": cliente.get("numero_documento", ""),
        "telefono":         cliente.get("telefono", ""),
        "email":            cliente.get("email", ""),
    }, id_empresa, nombre_empresa)

    gs_escribir_factura_convenio(sh_conv, {
        "id":               num_fac,
        "numero":           num_fac,
        "tipo":             "Factura de Venta Convenio",
        "fecha_emision":    now.strftime("%Y-%m-%d"),
        "fecha_vencimiento":(now + timedelta(days=30)).strftime("%Y-%m-%d"),
        "cliente_nombre":   cliente["nombre"],
        "cliente_doc":      cliente.get("numero_documento", ""),
        "cliente_email":    cliente.get("email", ""),
        "descripcion":      descripcion_item,
        "subtotal":         calc["subtotal"],
        "descuento":        calc.get("descuento_val", 0),
        "iva":              calc["iva"],
        "total":            calc["total"],
        "metodo_pago":      "Convenio",
        "estado":           "pendiente",
        "moneda":           "COP",
        "num_reserva":      num_res,
        "cubiculo":         cubiculo["numero"],
        "creado_en":        now.isoformat(),
        "actualizado_en":   now.isoformat(),
    }, id_empresa, nombre_empresa)

    # Ítems de factura convenio
    _pg_append_conv("Factura_items", [
        num_fac, num_fac,
        "CONVENIO-H",
        descripcion_item,
        str(calc["horas"]),
        "hora",
        str(calc["precio_hora"]),
        str(round(calc.get("descuento_pct", 0), 2)),
        "19.0",
        str(calc["subtotal"]),
        id_empresa,
    ])

    # ── 2. Actualizar cubiculos_estado (tabla compartida) ─────────────────────
    try:
        conn = get_pg_conn()
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE cubiculos_estado
                SET estado='ocupado', cliente_actual=%s,
                    hora_inicio=%s, hora_fin_prog=%s,
                    codigo_acceso=%s, tiempo_rest_min=%s
                WHERE numero=%s
            """, [
                cliente["nombre"],
                hora_inicio.isoformat(),
                hora_fin.isoformat(),
                cod_acc,
                str(diff_min),
                cubiculo["numero"],
            ])
        conn.commit()
    except Exception as e:
        st.warning(f"⚠️ Error actualizando cubículo (convenio): {e}")

    # Dashboard convenio acumulación incremental
    try:
        _gs_sync_dashboard_convenios(sh_conv, nueva_reserva=reserva_data)
    except Exception:
        pass

    _gs_invalidate_cache("Cubiculos_Estado", "Reservas")
    _gs_invalidate_cache_conv("Reservas", "Pagos", "Clientes", "Facturas", "Factura_items")

    _, sh = get_active_client()
    gs_escribir_log(sh, "nueva_reserva_conv", num_res, cubiculo["numero"],
                    st.session_state.get("operador_info", {}).get("nombre", "sistema"),
                    f"Conv {num_res} | Empresa {nombre_empresa} | "
                    f"Factura {num_fac} | Convenio | ${calc['total']:,.0f}")

    return {
        "numero_reserva":  num_res,
        "numero_factura":  num_fac,
        "cubiculo":        cubiculo["numero"],
        "hora_inicio":     hora_inicio.strftime("%H:%M"),
        "hora_fin":        hora_fin.strftime("%H:%M"),
        "horas":           calc["horas"],
        "precio_hora":     calc["precio_hora"],
        "descuento_pct":   calc.get("descuento_pct", 0),
        "descuento_val":   calc.get("descuento_val", 0),
        "subtotal":        calc["subtotal"],
        "iva":             calc["iva"],
        "total":           calc["total"],
        "metodo_pago":     "Convenio",
        "estado_pago":     "pendiente",
        "codigo_acceso":   cod_acc,
        "wifi_ssid":       cubiculo.get("wifi_ssid", ""),
        "wifi_password":   cubiculo.get("wifi_password", ""),
        "id_empresa":      id_empresa,
        "nombre_empresa":  nombre_empresa,
    }

# ── Helpers de QR de pago ─────────────────────────────────────────────────────

def qr_data_para_metodo(metodo: str, monto: int, ref: str) -> str:
    """Devuelve el texto/URL que se codificará en el QR según el método de pago."""
    if metodo == "Nequi":
        num = get_config("nequi_numero", NEQUI_NUM)
        return f"nequi://{num}?amount={monto}&ref={ref}"
    elif metodo == "Daviplata":
        num = get_config("daviplata_numero", DAVIPLATA_NUM)
        return f"daviplata://{num}?amount={monto}&ref={ref}"
    elif metodo == "MercadoPago":
        link = get_config("mp_link", MP_LINK)
        return f"{link}?amount={monto}&ref={ref}"
    elif metodo in ("PSE", "Transferencia"):
        cuenta = get_config("cuenta_bancaria", CUENTA_BANCO)
        return f"Pago {metodo} · {cuenta} · Monto: {monto} COP · Ref: {ref}"
    else:
        return f"Pago {metodo} · Monto: {monto} COP · Ref: {ref}"


def generar_qr_b64(data: str, fill: str = "#000000", back: str = "#ffffff",
                   size: int = 200) -> str:
    """
    Genera un QR con el texto ``data`` y lo devuelve como data-URI base64
    (formato ``data:image/png;base64,...``).
    Retorna cadena vacía si qrcode/Pillow no están instalados.
    """
    if not QR_AVAILABLE:
        return ""
    try:
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=max(4, size // 40),
            border=3,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color=fill or "#000000",
                            back_color=back or "#ffffff")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/png;base64,{b64}"
    except Exception:
        return ""


def mostrar_qr(metodo: str, monto: int, ref: str, size: int = 220):
    """
    Renderiza el QR de cobro en Streamlit para el método indicado.
    Devuelve (data_uri_str, b64_solo_str) para uso opcional en PDF.
    Si qrcode no está instalado muestra un aviso amigable en su lugar.
    """
    data    = qr_data_para_metodo(metodo, monto, ref)
    b64_uri = generar_qr_b64(data, size=size)
    if not b64_uri:
        st.info(
            f"📲 **Pago {metodo}** · Referencia: `{ref}` · "
            f"Monto: **{fmt_cop(monto)} COP**  \n"
            "_(Instala `qrcode` y `Pillow` para mostrar el QR)_"
        )
        return ("", "")
    b64_only = b64_uri.split(",")[1] if "," in b64_uri else b64_uri
    st.markdown(
        f'<div style="display:flex;justify-content:center;margin:8px 0">'
        f'<img src="{b64_uri}" width="{size}" height="{size}" '
        f'style="border-radius:12px;border:2px solid rgba(0,212,255,0.4)">'
        f'</div>',
        unsafe_allow_html=True,
    )
    return (b64_uri, b64_only)


def generar_ticket_pdf(voucher: dict) -> Optional[bytes]:
    if not REPORTLAB_AVAILABLE:
        return None
    buf = io.BytesIO()
    W   = 80 * mm
    doc = SimpleDocTemplate(buf, pagesize=(W, 200*mm),
                            rightMargin=4*mm, leftMargin=4*mm,
                            topMargin=4*mm, bottomMargin=4*mm)
    styles = getSampleStyleSheet()
    c_style = ParagraphStyle("Center", alignment=TA_CENTER,
                              fontName="Helvetica", fontSize=8, leading=10)
    h_style = ParagraphStyle("Header", alignment=TA_CENTER,
                              fontName="Helvetica-Bold", fontSize=11, leading=13)
    big_style= ParagraphStyle("Big", alignment=TA_CENTER,
                               fontName="Helvetica-Bold", fontSize=22, leading=26)
    elems = []

    def hr(): return HRFlowable(width="100%", thickness=0.5, color=colors.black, spaceAfter=3, spaceBefore=3)
    def p(text, style=c_style): return Paragraph(text, style)

    # ── Logo en el ticket PDF ────────────────────────────────────────────────
    if _LOGO_B64:
        try:
            from reportlab.platypus import Image as RLImage
            _logo_buf_ticket = io.BytesIO(base64.b64decode(_LOGO_B64))
            _logo_img_ticket = RLImage(_logo_buf_ticket, width=28*mm, height=38*mm)
            _logo_img_ticket.hAlign = "CENTER"
            elems.append(_logo_img_ticket)
            elems.append(Spacer(1, 2))
        except Exception:
            pass

    elems += [
        p(NEGOCIO, h_style),
        p(TAGLINE, c_style),
        p(DIRECCION, c_style),
        p(f"Tel: {TELEFONO} · {NIT}", c_style),
        hr(),
        p(f"Reserva: <b>{voucher['numero_reserva']}</b>", c_style),
        p(f"Fecha: {ahora_col().strftime('%d/%m/%Y %H:%M')}", c_style),
        hr(),
        Spacer(1, 3),
        p(f"CUBÍCULO: <b>{voucher['cubiculo']}</b>", h_style),
        p(f"Entrada: {voucher['hora_inicio']}  →  Salida: {voucher['hora_fin']}", c_style),
        p(f"Tiempo: <b>{voucher['horas']}h</b>", c_style),
        hr(),
        p("SERVICIOS INCLUIDOS:", c_style),
        p("✓ Baño  ✓ WiFi  ✓ Carga USB/110V", c_style),
        hr(),
        p("CÓDIGO DE ACCESO:", c_style),
        Spacer(1, 4),
        p(f"<b>{voucher['codigo_acceso']}</b>", big_style),
        Spacer(1, 4),
        p(f"WiFi: <b>{voucher['wifi_ssid']}</b>", c_style),
        p(f"Clave WiFi: <b>{voucher['wifi_password']}</b>", c_style),
        hr(),
        p(f"Subtotal:    {fmt_cop(voucher['subtotal'])}", c_style),
        p(f"IVA (19%):   {fmt_cop(voucher['iva'])}", c_style),
    ]

    # Total en tabla destacada
    tabla_total = Table([[f"TOTAL: {fmt_cop(voucher['total'])} COP"]],
                        colWidths=[W - 8*mm])
    tabla_total.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#0a3d62")),
        ("TEXTCOLOR",     (0,0), (-1,-1), colors.HexColor("#00d4ff")),
        ("FONTNAME",      (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 13),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("ROUNDEDCORNERS",(0,0), (-1,-1), [4,4,4,4]),
    ]))
    elems.append(Spacer(1, 4))
    elems.append(tabla_total)
    elems += [
        Spacer(1, 4),
        p(f"Método: <b>{voucher['metodo_pago']}</b>", c_style),
        p(f"Factura N°: <b>{voucher['numero_factura']}</b>", c_style),
        hr(),
    ]

    # QR
    qr_data = qr_data_para_metodo("Nequi", 0, voucher["numero_reserva"])
    b64_qr  = generar_qr_b64(qr_data, fill="#000000", back="#ffffff")
    if b64_qr and QR_AVAILABLE:
        try:
            qr_bytes = base64.b64decode(b64_qr.split(",")[1])
            img_buf  = io.BytesIO(qr_bytes)
            from reportlab.platypus import Image as RLImage
            rl_img = RLImage(img_buf, width=30*mm, height=40*mm)
            elems.append(rl_img)
        except Exception:
            pass

    elems += [
        hr(),
        p("¡Gracias por tu visita!", h_style),
        p("¡Buen viaje!", c_style),
        p(f"{NEGOCIO} · Terminal de Transportes", c_style),
    ]

    doc.build(elems)
    return buf.getvalue()


def voucher_html(v: dict) -> str:
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>
body {{font-family:monospace;max-width:320px;margin:20px auto;
      background:#050b1a;color:#e2e8f0;padding:20px;border:2px solid #00d4ff;border-radius:12px;}}
h1{{color:#00d4ff;text-align:center;font-size:18px;}}
.cod{{font-size:56px;font-weight:700;text-align:center;color:#00ff88;
     letter-spacing:10px;text-shadow:0 0 20px rgba(0,255,136,0.6);margin:16px 0;}}
.row{{display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.1);}}
.total{{font-size:20px;font-weight:700;color:#00d4ff;text-align:right;margin-top:12px;}}
</style></head><body>
<h1>💤 {NEGOCIO}</h1>
<p style="text-align:center;color:#94a3b8;font-size:12px">{TAGLINE}</p>
<hr style="border-color:rgba(0,212,255,0.3)">
<div class="row"><span>Reserva</span><b>{v['numero_reserva']}</b></div>
<div class="row"><span>Cubículo</span><b>{v['cubiculo']}</b></div>
<div class="row"><span>Entrada</span><b>{v['hora_inicio']}</b></div>
<div class="row"><span>Salida</span><b>{v['hora_fin']}</b></div>
<div class="row"><span>Tiempo</span><b>{v['horas']}h</b></div>
<div class="row"><span>WiFi</span><b>{v['wifi_ssid']}</b></div>
<div class="row"><span>Clave WiFi</span><b>{v['wifi_password']}</b></div>
<hr style="border-color:rgba(0,212,255,0.3)">
<p style="text-align:center;color:#94a3b8;font-size:12px">CÓDIGO DE ACCESO</p>
<div class="cod">{v['codigo_acceso']}</div>
<hr style="border-color:rgba(0,212,255,0.3)">
<div class="row"><span>✓ Baño</span><span>✓ WiFi</span><span>✓ Carga</span></div>
<div class="row"><span>Subtotal</span><span>{fmt_cop(v['subtotal'])}</span></div>
<div class="row"><span>IVA 19%</span><span>{fmt_cop(v['iva'])}</span></div>
<div class="total">TOTAL: {fmt_cop(v['total'])} COP</div>
<div class="row"><span>Pago</span><b>{v['metodo_pago']}</b></div>
<div class="row"><span>Factura</span><b>{v['numero_factura']}</b></div>
<hr style="border-color:rgba(0,212,255,0.3)">
<p style="text-align:center;color:#94a3b8;font-size:11px">¡Gracias por tu visita! · ¡Buen viaje!</p>
</body></html>"""


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════════════════════

def init_state():
    defaults = {
        "pantalla":              "operador_login",
        "cubiculo_sel":          None,
        "horas_sel":             1.0,
        "calc":                  None,
        "cliente":               {},
        "metodo_pago":           None,
        "voucher":               None,
        "pago_confirmado":       False,
        "operador_ok":           False,
        "operador_info":         {},
        "pin_intentos":          0,
        "modulo_op":             "dashboard",
        "_backup_fecha":         "",
        "_procesando_reserva":   False,
        # ── Convenios ───────────────────────────────
        "_conv_cfg_pendiente":   {},   # config pendiente de guardar por empresa
        "_conv_emp_sel":         None, # empresa seleccionada en formulario de convenio
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _auto_backup_diario():
    """
    Genera backup al inicio del proceso si no se hizo hoy.
    Mantiene compatibilidad — el backup principal ahora ocurre al cerrar turno.
    """
    today_str = ahora_col().strftime("%Y-%m-%d")
    if st.session_state.get("_backup_fecha") == today_str:
        return
    try:
        os.makedirs("backups", exist_ok=True)
        data, fname, _ = generar_backup_diario()
        backup_path = os.path.join("backups", fname)
        if not os.path.exists(backup_path):
            with open(backup_path, "wb") as bf:
                bf.write(data)
        st.session_state["_backup_fecha"] = today_str
    except Exception:
        pass


def _backup_cierre_turno(operador_info: dict):
    """
    Genera un backup al momento de cerrar el turno del operador.

    - En entorno LOCAL: guarda el archivo en la carpeta backups/ del proyecto.
    - En Streamlit Cloud: almacena el backup en session_state["_backup_cierre_bytes"]
      para que el sidebar pueda ofrecer el botón de descarga antes de cerrar sesión.
      (En Cloud /tmp no persiste entre sesiones, así que no se escribe a disco.)
    """
    try:
        ahora  = ahora_col()
        turno  = operador_info.get("turno", "turno")
        nombre = operador_info.get("nombre", "op").replace(" ", "_").lower()
        ts_str = ahora.strftime("%Y-%m-%d_%H%M")

        data, fname_base, mime = generar_backup_diario()
        # Nombre del archivo de cierre diferente al backup genérico
        fname = f"cierre_{turno}_{nombre}_{ts_str}.xlsx" if mime.endswith("sheet") else                 f"cierre_{turno}_{nombre}_{ts_str}.zip"

        if _IS_CLOUD:
            # En Cloud: guardar en session_state para descarga inmediata en el sidebar
            st.session_state["_backup_cierre_bytes"] = data
            st.session_state["_backup_cierre_fname"] = fname
            st.session_state["_backup_cierre_mime"]  = mime
        else:
            # En local: guardar en carpeta backups/
            os.makedirs("backups", exist_ok=True)
            backup_path = os.path.join("backups", fname)
            with open(backup_path, "wb") as bf:
                bf.write(data)

        # Registrar en Google Sheets
        try:
            _, sh_cierre = get_active_client()
            if sh_cierre:
                gs_sync_dashboard(sh_cierre)
            sheets_append_row("Log_Operaciones", [
                ahora.isoformat(), "cierre_turno", "", "",
                operador_info.get("nombre", ""),
                f"Backup cierre de turno '{turno}' — {fname}",
                "", "", "", "exito", ""
            ])
        except Exception:
            pass
    except Exception:
        pass


init_state()


def ir_a(pantalla: str):
    # Al navegar fuera de confirmación, liberar el lock de procesamiento
    if pantalla != "confirmacion":
        st.session_state["_procesando_reserva"] = False
    st.session_state.pantalla = pantalla
    st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# COMPONENTES REUTILIZABLES
# ══════════════════════════════════════════════════════════════════════════════

def render_header(subtitulo: str = ""):
    st.markdown(f"""
    <div class="main-header">
      <div class="main-logo">💤 {NEGOCIO}</div>
      <div class="main-tagline">{subtitulo or TAGLINE}</div>
    </div>""", unsafe_allow_html=True)


def render_stepper(paso_actual: int):
    pasos = ["Cubículo", "Datos", "Pago", "Confirmación", "Voucher"]
    html  = '<div class="stepper">'
    for i, nombre in enumerate(pasos):
        cls = "active" if i == paso_actual else ("done" if i < paso_actual else "")
        ico = ["🛏️","👤","💳","✅","🎫"][i]
        html += f'<div class="step {cls}">{ico} {nombre}</div>'
        if i < len(pasos)-1:
            html += '<div class="step-arrow"></div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def render_reloj():
    """
    Reloj con hora y fecha actualizados en tiempo real por JavaScript.
    Los elementos .jjgt-clock-hms y .jjgt-clock-fecha son actualizados cada segundo
    por el script inyectado en inject_live_clock() — sin recarga del servidor.
    """
    now = ahora_col()
    st.markdown(f"""
    <div style="text-align:center;margin-bottom:12px">
      <div class="clock-big jjgt-clock-hms">{now.strftime('%H:%M:%S')}</div>
      <div class="clock-label jjgt-clock-fecha">{now.strftime('%A, %d de %B de %Y').upper()}</div>
    </div>
    """, unsafe_allow_html=True)


def render_cubiculo_card(cub: dict, seleccionado: bool = False) -> bool:
    """
    Renderiza tarjeta de cubículo con timer en vivo (actualizado por JS cada segundo).
    Retorna True si fue clickeado.
    """
    estado = cub["estado"]
    info   = ESTADOS_CUBICULO.get(estado, ESTADOS_CUBICULO["libre"])
    libre  = estado == "libre"
    cls    = f"cubiculo-{estado.replace('_','-')}"
    if seleccionado:
        cls += " cubiculo-selected"

    # Timer HTML: usa clase .jjgt-timer con data-fin para actualización JS en tiempo real
    timer_html = ""
    if cub.get("hora_fin") and not libre:
        mins  = cub.get("minutos_restantes", 999)
        color = "#ff4757" if mins <= 5 else ("#ffd32a" if mins <= 15 else "#00ff88")
        timer_html = (
            f'<div class="timer-display jjgt-timer" ' 
            f'data-fin="{cub["hora_fin"]}" data-key="cub-{cub["id"]}" ' 
            f'style="color:{color}">⏱ {fmt_tiempo(mins)}</div>'
        )
    elif cub.get("minutos_restantes") is not None and not libre:
        mins = cub["minutos_restantes"]
        color = "#ffd32a" if mins <= 15 else "#ff4757"
        timer_html = f'<div class="timer-display" style="color:{color}">⏱ {fmt_tiempo(mins)}</div>'

    st.markdown(f"""
    <div class="cubiculo-card {cls}" style="background:{info['bg']};border-color:{info['color']}40">
      <div class="cubiculo-num" style="color:{info['color']}">{cub['numero']}</div>
      <div class="estado-badge" style="background:{info['color']}22;color:{info['color']};
           border:1px solid {info['color']}55">{info['label']}</div>
      {timer_html}
      <div class="servicios-icons">🚿 🌐 🔌</div>
    </div>
    """, unsafe_allow_html=True)

    if libre:
        label = f"✅ Seleccionar {cub['numero']}"
        return st.button(label, key=f"btn_cub_{cub['id']}", use_container_width=True)
    return False


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 0 — BIENVENIDA
# ══════════════════════════════════════════════════════════════════════════════

def show_bienvenida():
    # Seguridad: si no hay operador autenticado, redirigir al login
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header()
    inject_live_clock()
    render_reloj()

    # ── Auto-rerun del servidor cada 60s para refrescar estado (primer run excluido) ─
    now_ts = int(time.time())
    last_rerun_b = st.session_state.get("_bienvenida_last_rerun", 0)
    if last_rerun_b == 0:
        # Primera carga: solo registrar timestamp, no rerunear
        st.session_state["_bienvenida_last_rerun"] = now_ts
    elif now_ts - last_rerun_b >= 60:
        st.session_state["_bienvenida_last_rerun"] = now_ts
        st.rerun()

    # ── Auto-liberar cubículos vencidos en background ─────────────────────────
    cubiculos_raw = get_cubiculos()
    alarma_vencidos = []
    for cub in cubiculos_raw:
        if cub["estado"] in ("ocupado", "por_liberar"):
            mins = cub.get("minutos_restantes")
            if mins is not None and mins <= 0:
                liberar_cubiculo(cub["id"])
                alarma_vencidos.append(cub["numero"])
    if alarma_vencidos:
        st.markdown("""
        <script>
        (function(){
          try {
            var ctx = new (window.AudioContext || window.webkitAudioContext)();
            function beep(freq, start, dur) {
              var o = ctx.createOscillator();
              var g = ctx.createGain();
              o.connect(g); g.connect(ctx.destination);
              o.frequency.value = freq; o.type = 'square';
              g.gain.setValueAtTime(0.3, ctx.currentTime + start);
              g.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime + start + dur);
              o.start(ctx.currentTime + start);
              o.stop(ctx.currentTime + start + dur + 0.05);
            }
            for(var i=0; i<3; i++){beep(880,i*0.4,0.3); beep(660,i*0.4+0.15,0.15);}
          } catch(e){}
        })();
        </script>
        """, unsafe_allow_html=True)

    cubiculos  = get_cubiculos()  # Re-consultar después de liberaciones
    libres     = sum(1 for c in cubiculos if c["estado"] == "libre")
    total      = len(cubiculos)
    tarifa_act = "Madrugada" if 0 <= ahora_col().hour < 6 else "Estándar"
    precio_min = 15000 if tarifa_act == "Madrugada" else 15000

    # Disponibilidad
    color_disp = "#00ff88" if libres > 3 else ("#ffd32a" if libres > 0 else "#ff4757")
    icono_disp = "🟢" if libres > 3 else ("🟡" if libres > 0 else "🔴")
    st.markdown(f"""
    <div style="text-align:center;margin:12px 0 24px;
                font-size:22px;font-weight:700;color:{color_disp}">
      {icono_disp} {libres} de {total} cubículos disponibles ahora
    </div>
    """, unsafe_allow_html=True)

    # Tarifas
    hora_col = ahora_col().hour
    tarifa_badge = ""
    if 0 <= hora_col < 6:
        tarifa_badge = '<span style="background:rgba(0,212,255,0.2);color:#00d4ff;padding:3px 10px;border-radius:12px;font-size:13px;font-weight:700;letter-spacing:1px">🌙 TARIFA MADRUGADA ACTIVA</span>'

    st.markdown(f"""
    <div style="text-align:center;margin-bottom:24px">
      <div style="font-size:18px;color:#94a3b8;margin-bottom:8px">
        Desde <b style="color:#00ff88">{fmt_cop(precio_min)}/hora</b> &nbsp;·&nbsp;
        <span>🚿 Baño</span> &nbsp;·&nbsp;
        <span>🌐 WiFi</span> &nbsp;·&nbsp;
        <span>🔌 Carga</span>
      </div>
      {tarifa_badge}
    </div>
    """, unsafe_allow_html=True)

    # Botón principal + info del operador activo
    op_bienvenida = st.session_state.get("operador_info", {})
    op_nombre_b   = op_bienvenida.get("nombre", "")
    if op_nombre_b:
        st.markdown(f"""
        <div style="text-align:center;font-size:13px;color:#94a3b8;margin-bottom:8px">
          👤 Atendido por: <b style="color:#00d4ff">{op_nombre_b}</b>
          · Turno {op_bienvenida.get("turno","").capitalize()}
        </div>
        """, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 2, 1])
    with col:
        if libres > 0:
            if st.button("🛏️  RESERVAR MI ESPACIO", type="primary", use_container_width=True):
                ir_a("seleccion")
        else:
            st.markdown('<div class="alerta-roja">❌ No hay cubículos disponibles en este momento.<br>Por favor espera o consulta al operador.</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Mini grid de estado (timers actualizados por JS cada segundo)
    st.markdown("---")
    st.markdown("#### Estado en tiempo real")
    cols = st.columns(6)
    for i, cub in enumerate(cubiculos):
        with cols[i % 6]:
            estado   = cub["estado"]
            info     = ESTADOS_CUBICULO.get(estado, ESTADOS_CUBICULO["libre"])
            hora_fin = cub.get("hora_fin") or ""
            mins     = cub.get("minutos_restantes")
            if hora_fin and estado not in ("libre", "mantenimiento"):
                timer_html = (
                    f'<div class="jjgt-timer" data-fin="{hora_fin}" ' +
                    f'data-key="bien-{cub["id"]}" ' +
                    f'style="font-size:11px;color:#ffd32a">⏱{fmt_tiempo(mins)}</div>'
                )
            else:
                timer_html = ""
            st.markdown(f"""
            <div style="background:{info['bg']};border:1.5px solid {info['color']}44;
                        border-radius:10px;padding:10px;text-align:center;margin-bottom:8px">
              <div style="font-family:'Inconsolata',monospace;font-size:22px;
                          font-weight:700;color:{info['color']}">{cub['numero']}</div>
              <div style="font-size:10px;font-weight:700;color:{info['color']};
                          letter-spacing:1px">{info['label']}</div>
              {timer_html}
            </div>
            """, unsafe_allow_html=True)

    # Botones de acción del operador
    st.markdown("<br>", unsafe_allow_html=True)
    _, col_back, _ = st.columns([2, 1, 2])
    with col_back:
        if st.button("⬅️ Volver al Panel", use_container_width=True):
            ir_a("operador")


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 1 — SELECCIÓN DE CUBÍCULO Y TIEMPO
# ══════════════════════════════════════════════════════════════════════════════

def show_seleccion():
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header("Elige tu cubículo y tiempo de descanso")
    render_stepper(0)
    inject_live_clock()

    col_left, col_right = st.columns([1, 1])

    with col_left:
        st.markdown("#### ⏱️ ¿Cuánto tiempo necesitas?")
        tiempos = {"30 min": 0.3, "1 hora": 1.0, "2 horas": 2.0,
                   "3 horas": 3.0, "4 horas": 4.0, "⚙️ Personalizar": -1}
        cols_t  = st.columns(3)
        for i, (label, val) in enumerate(tiempos.items()):
            with cols_t[i % 3]:
                sel = st.session_state.horas_sel
                activo = (sel == val) or (val == -1 and sel not in [0.3,1,2,3,4])
                if st.button(label, key=f"btn_t_{i}",
                             type="primary" if activo else "secondary",
                             use_container_width=True):
                    if val == -1:
                        st.session_state.horas_sel = 1.5
                    else:
                        st.session_state.horas_sel = val
                    st.rerun()

        horas = st.session_state.horas_sel
        if horas not in [0.5, 1.0, 2.0, 3.0, 4.0]:
            horas = st.number_input("Horas personalizadas", min_value=0.5, max_value=18.0,
                                     value=float(horas), step=0.5)
            st.session_state.horas_sel = horas

        # Cálculo en tiempo real
        calc = calcular_precio(st.session_state.horas_sel)
        st.session_state.calc = calc
        st.markdown("<br>", unsafe_allow_html=True)

        # Resumen de precio
        tarifa_badge = ""
        if calc["tarifa"] == "Madrugada":
            tarifa_badge = "🌙 Tarifa madrugada"
        elif calc["descuento_pct"] > 0:
            tarifa_badge = f"🏷️ Descuento {calc['descuento_pct']:.0f}% aplicado"

        st.markdown(f"""
        <div style="background:var(--bg-card);border:1.5px solid rgba(0,212,255,0.3);
                    border-radius:12px;padding:20px">
          <div style="font-size:13px;color:#94a3b8;letter-spacing:2px;text-transform:uppercase">
            Tarifa: {calc['tarifa']} {tarifa_badge}
          </div>
          <div style="font-size:13px;color:#94a3b8;margin-top:4px">
            {fmt_cop(calc['precio_hora'])}/hora × {calc['horas']}h
            {"  ·  -" + fmt_cop(calc['descuento_val']) if calc['descuento_val'] > 0 else ""}
          </div>
          <div style="font-size:13px;color:#94a3b8">Subtotal: {fmt_cop(calc['subtotal'])}</div>
          <div style="font-size:13px;color:#94a3b8">IVA 19%: {fmt_cop(calc['iva'])}</div>
          <div style="font-family:'Inconsolata',monospace;font-size:32px;
                      font-weight:700;color:#00d4ff;margin-top:8px">
            {fmt_cop(calc['total'])} COP
          </div>
        </div>
        """, unsafe_allow_html=True)

    with col_right:
        st.markdown("#### 🛏️ Selecciona tu cubículo")
        cubiculos = get_cubiculos()
        cols_c    = st.columns(4)
        for i, cub in enumerate(cubiculos):
            with cols_c[i % 4]:
                sel = st.session_state.cubiculo_sel
                clicked = render_cubiculo_card(
                    cub, seleccionado=(sel and sel["id"] == cub["id"]))
                if clicked and cub["estado"] == "libre":
                    st.session_state.cubiculo_sel = cub
                    st.rerun()

    st.markdown("---")
    col_v1, col_v2, col_c = st.columns([1, 1, 2])
    with col_v1:
        if st.button("← Volver", use_container_width=True):
            ir_a("bienvenida")
    with col_v2:
        if st.button("✖ Cancelar al panel", use_container_width=True):
            ir_a("operador")
    with col_c:
        if st.session_state.cubiculo_sel and st.session_state.calc:
            cub_sel = st.session_state.cubiculo_sel
            cval    = st.session_state.calc
            st.markdown(f"""
            <div style="background:rgba(0,212,255,0.1);border:1px solid rgba(0,212,255,0.4);
                        border-radius:10px;padding:12px 16px;text-align:center;
                        font-weight:700;color:#00d4ff;margin-bottom:8px">
              {cub_sel['numero']} · {cval['horas']}h · {fmt_cop(cval['total'])} COP
            </div>
            """, unsafe_allow_html=True)
            if st.button("CONTINUAR →", type="primary", use_container_width=True):
                ir_a("datos")
        else:
            st.info("👆 Selecciona un cubículo libre para continuar")
            # Verificar si hay cubículos ocupados por convenio
            all_cubs = get_cubiculos()
            conv_ocup = [c for c in all_cubs if c["estado"] == "ocupado"
                        and c.get("cliente_actual","")]
            if conv_ocup:
                st.caption(f"ℹ️ {len(conv_ocup)} cubículo(s) ocupado(s) — "
                           "puede incluir reservas de convenio")


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 2 — DATOS DEL CLIENTE
# ══════════════════════════════════════════════════════════════════════════════

def show_datos():
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header("Tus datos para la reserva")
    inject_live_clock()
    render_stepper(1)

    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        nombre = st.text_input("👤 Nombre completo *", key="datos_nombre",
                               placeholder="Ej: María García López",
                               value=st.session_state.cliente.get("nombre",""))
        c1, c2 = st.columns(2)
        with c1:
            tipo_doc = st.selectbox("Tipo documento", ["CC", "Pasaporte", "CE", "NIT"], key="datos_tipo_doc")
        with c2:
            num_doc = st.text_input("N° de documento *", key="datos_num_doc",
                                    placeholder="1234567890",
                                    value=st.session_state.cliente.get("numero_documento",""))
        telefono = st.text_input("📱 Teléfono celular *", key="datos_telefono",
                                 placeholder="310 555 0000",
                                 value=st.session_state.cliente.get("telefono",""))
        email = st.text_input("📧 Email (opcional — para factura digital)", key="datos_email",
                              placeholder="tu@email.com",
                              value=st.session_state.cliente.get("email",""))

        factura_emp = st.checkbox("🏢 Requiero factura a nombre de empresa", key="datos_req_fact")
        razon_social = ""
        nit_emp = ""
        if factura_emp:
            razon_social = st.text_input("Razón social", key="datos_razon_social", value=st.session_state.cliente.get("razon_social",""))
            nit_emp      = st.text_input("NIT de la empresa", key="datos_nit_empresa", value=st.session_state.cliente.get("nit_empresa",""))

        st.markdown("</div>", unsafe_allow_html=True)

        # Autorización tratamiento de datos personales
        st.markdown("<br>", unsafe_allow_html=True)
        acepto_datos_kiosco = st.checkbox(
            "✅ Autorizo el tratamiento de mis datos personales conforme a la "
            "Ley 1581 de 2012 y la política de privacidad de SUITE SALITRE.",
            key="kiosco_acepto_datos",
        )
        st.markdown(
            "<div style='font-size:12px;color:#94a3b8;margin-top:-8px;margin-bottom:8px'>"
            "La información suministrada será usada exclusivamente para la gestión de su reserva y facturación."
            "</div>",
            unsafe_allow_html=True,
        )

        # Validación
        errores = []
        if not nombre.strip():
            errores.append("El nombre es obligatorio")
        if not num_doc.strip():
            errores.append("El número de documento es obligatorio")
        if not telefono.strip():
            errores.append("El teléfono es obligatorio")
        if not acepto_datos_kiosco:
            errores.append("Debes aceptar la autorización de tratamiento de datos personales")

        btn_v, btn_p, btn_c = st.columns([1, 1, 2])
        with btn_v:
            if st.button("← Volver", use_container_width=True):
                ir_a("seleccion")
        with btn_p:
            if st.button("✖ Panel", use_container_width=True):
                ir_a("operador")
        with btn_c:
            if st.button("CONTINUAR AL PAGO →", type="primary", use_container_width=True):
                if errores:
                    for e in errores:
                        st.error(e)
                else:
                    st.session_state.cliente = {
                        "nombre": nombre.strip(),
                        "tipo_doc": tipo_doc,
                        "numero_documento": num_doc.strip(),
                        "telefono": telefono.strip(),
                        "email": email.strip(),
                        "razon_social": razon_social,
                        "nit_empresa": nit_emp,
                        "acepto_datos": acepto_datos_kiosco,
                    }
                    ir_a("pago")


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 3 — PAGO
# ══════════════════════════════════════════════════════════════════════════════

def show_pago():
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header("Elige cómo pagar")
    render_stepper(2)
    inject_live_clock()

    calc   = st.session_state.calc
    cub    = st.session_state.cubiculo_sel
    # Referencia única por sesión + timestamp para garantizar QR fresco
    ref    = f"CR-{ahora_col().strftime('%Y%m%d%H%M')}-{cub['numero'].replace('#','')}"
    monto  = int(calc["total"])

    # Auto-refresh del QR cada 5 min: rerun de Streamlit = QR nuevo desde Python
    now_ts_pago = int(time.time())
    last_qr_rerun = st.session_state.get("_pago_qr_rerun_ts", 0)
    if now_ts_pago - last_qr_rerun >= 300:  # 5 minutos
        st.session_state["_pago_qr_rerun_ts"] = now_ts_pago
        st.rerun()

    # Resumen
    st.markdown(f"""
    <div style="background:rgba(0,212,255,0.08);border:1px solid rgba(0,212,255,0.3);
                border-radius:12px;padding:14px 20px;display:flex;
                justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:12px">
      <div>
        <span style="color:#94a3b8;font-size:13px">Cubículo</span><br>
        <b style="font-size:20px;color:#00d4ff">{cub['numero']}</b>
      </div>
      <div>
        <span style="color:#94a3b8;font-size:13px">Tiempo</span><br>
        <b style="font-size:20px">{calc['horas']}h</b>
      </div>
      <div>
        <span style="color:#94a3b8;font-size:13px">Total a pagar</span><br>
        <b style="font-family:'Inconsolata';font-size:28px;color:#00ff88">{fmt_cop(monto)} COP</b>
      </div>
    </div>
    """, unsafe_allow_html=True)

    tab_nequi, tab_davi, tab_efect, tab_otros = st.tabs([
        "💚 Nequi", "💙 Daviplata", "💵 Efectivo", "📲 Otros métodos"])

    with tab_nequi:
        _pago_nequi(monto, ref)
    with tab_davi:
        _pago_daviplata(monto, ref)
    with tab_efect:
        _pago_efectivo(monto, ref)
    with tab_otros:
        _pago_otros(monto, ref)

    st.markdown("<br>", unsafe_allow_html=True)
    col_pb1, col_pb2 = st.columns(2)
    with col_pb1:
        if st.button("← Volver a mis datos", use_container_width=True):
            ir_a("datos")
    with col_pb2:
        if st.button("✖ Cancelar al panel", use_container_width=True):
            ir_a("operador")


def _pago_nequi(monto: int, ref: str):
    """
    Pantalla de pago con Nequi.
    Muestra el QR de cobro personal con instrucciones paso a paso para que el
    cliente abra Nequi, acceda a 'Tu QR personal' y escanee para pagar.
    El QR codifica el deep-link nequi:// con monto y referencia exactos.
    """
    nequi_num = get_config("nequi_numero", NEQUI_NUM)

    # ── Encabezado ────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,rgba(0,150,60,0.25),rgba(0,220,90,0.10));
                border:2px solid rgba(0,220,90,0.45);border-radius:16px;
                padding:20px 24px;margin-bottom:18px">
      <div style="display:flex;align-items:center;gap:12px;margin-bottom:6px">
        <span style="font-size:32px">💚</span>
        <div>
          <div style="font-size:22px;font-weight:800;color:#00c853;letter-spacing:-0.5px">
            Paga con Nequi
          </div>
          <div style="font-size:13px;color:#94a3b8;margin-top:2px">
            Número de cuenta: <b style="font-family:'Inconsolata';color:#00ff88">{nequi_num}</b>
          </div>
        </div>
        <div style="margin-left:auto;text-align:right">
          <div style="font-size:13px;color:#94a3b8">Total a pagar</div>
          <div style="font-family:'Inconsolata';font-size:32px;font-weight:800;
                      color:#00ff88;text-shadow:0 0 20px rgba(0,255,136,0.4)">
            {fmt_cop(monto)} COP
          </div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_qr, col_steps = st.columns([1, 1], gap="large")

    with col_qr:
        # ── QR de cobro ───────────────────────────────────────────────────────
        st.markdown("""
        <div style="text-align:center;font-size:13px;font-weight:700;
                    color:#00c853;letter-spacing:2px;text-transform:uppercase;
                    margin-bottom:8px">
          📷 Escanea este QR con Nequi
        </div>
        """, unsafe_allow_html=True)

        _, qr_b64 = mostrar_qr("Nequi", monto, ref, size=280)

        # Datos del pago bajo el QR
        st.markdown(f"""
        <div style="background:rgba(0,0,0,0.3);border:1px solid rgba(0,220,90,0.25);
                    border-radius:10px;padding:12px;margin-top:10px;text-align:center">
          <div style="font-size:11px;color:#94a3b8;letter-spacing:1px;
                      text-transform:uppercase;margin-bottom:4px">Referencia de pago</div>
          <div style="font-family:'Inconsolata';font-size:14px;font-weight:700;
                      color:#e2e8f0;word-break:break-all">{ref}</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:6px">
            Número Nequi del comercio: <b style="color:#00c853">{nequi_num}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)

    with col_steps:
        # ── Instrucciones paso a paso ─────────────────────────────────────────
        st.markdown("""
        <div style="font-size:14px;font-weight:700;color:#00c853;letter-spacing:1px;
                    text-transform:uppercase;margin-bottom:14px">
          📱 ¿Cómo pagar?
        </div>
        """, unsafe_allow_html=True)

        pasos = [
            ("1", "Abre la app <b>Nequi</b> en tu celular",
             "No necesitas iniciar sesión si ya estás autenticado"),
            ("2", 'Toca el ícono <b style="color:#00ff88">$</b> en la pantalla principal',
             "Es el botón de cobros / pagos rápidos"),
            ("3", 'Selecciona <b>"Usa tu QR"</b>',
             "Aparece en el menú de opciones de pago"),
            ("4", 'Elige <b>"Tu QR personal"</b>',
             "Para cobrar desde tu cuenta personal"),
            ("5", 'Toca <b>"Generar tu QR"</b> y muéstraselo al operador',
             "O usa el QR de la pantalla para transferir directamente"),
            ("6", f'Envía <b style="color:#00ff88">{fmt_cop(monto)} COP</b> al número <b style="color:#00ff88">{nequi_num}</b>',
             "Confirma el monto antes de aprobar"),
        ]

        for num, titulo, subtitulo in pasos:
            st.markdown(f"""
            <div style="display:flex;gap:12px;align-items:flex-start;
                        margin-bottom:12px;padding:10px 12px;
                        background:rgba(0,100,40,0.12);
                        border-left:3px solid rgba(0,220,90,0.5);
                        border-radius:0 8px 8px 0">
              <div style="min-width:28px;height:28px;border-radius:50%;
                          background:#00c853;color:#000;font-weight:800;
                          font-size:13px;display:flex;align-items:center;
                          justify-content:center;flex-shrink:0">{num}</div>
              <div>
                <div style="font-size:14px;color:#e2e8f0">{titulo}</div>
                <div style="font-size:11px;color:#94a3b8;margin-top:2px">{subtitulo}</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        # Alerta de monto
        st.markdown(f"""
        <div style="background:rgba(0,200,83,0.15);border:1px solid rgba(0,200,83,0.4);
                    border-radius:10px;padding:12px;text-align:center;margin-top:4px">
          <div style="font-size:12px;color:#94a3b8;margin-bottom:4px">
            ⚠️ Verifica siempre el monto antes de confirmar
          </div>
          <div style="font-family:'Inconsolata';font-size:26px;font-weight:800;color:#00ff88">
            {fmt_cop(monto)} COP
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Botón de confirmación ─────────────────────────────────────────────────
    if st.button("✅ Ya pagué con Nequi — Confirmar pago", type="primary",
                 use_container_width=True, key="confirm_nequi"):
        _confirmar_pago("Nequi")


def _pago_daviplata(monto: int, ref: str):
    davi_num = get_config("daviplata_numero", DAVIPLATA_NUM)
    st.markdown(f"""
    <div style="background:rgba(10,40,100,0.3);border:2px solid rgba(59,130,246,0.35);
                border-radius:12px;padding:20px;margin-bottom:16px">
      <div style="font-size:20px;font-weight:700;color:#60a5fa;margin-bottom:8px">
        💙 Pago con Daviplata
      </div>
      <div style="font-size:16px;color:#e2e8f0">
        Abre <b>Daviplata</b> → <b>Pagar</b> → Escanear QR<br>
        o envía a <b style="font-family:'Inconsolata';font-size:22px;color:#60a5fa">{davi_num}</b><br>
        <b style="color:#60a5fa">Monto: {fmt_cop(monto)} COP</b>
      </div>
    </div>
    """, unsafe_allow_html=True)
    mostrar_qr("Daviplata", monto, ref)
    if st.button("✅ Ya pagué con Daviplata — Confirmar pago", type="primary",
                 use_container_width=True, key="confirm_davi"):
        _confirmar_pago("Daviplata")


def _pago_efectivo(monto: int, ref: str):
    tel_op = get_config("whatsapp_op", WHATSAPP_OP)
    cambios = {
        10000: monto - 10000, 20000: monto - 20000,
        50000: 50000 - monto, 100000: 100000 - monto,
    }
    st.markdown(f"""
    <div style="background:rgba(20,60,20,0.3);border:2px solid rgba(132,204,22,0.35);
                border-radius:12px;padding:20px;margin-bottom:16px">
      <div style="font-size:20px;font-weight:700;color:#84cc16;margin-bottom:8px">
        💵 Pago en Efectivo
      </div>
      <div style="font-size:16px;color:#e2e8f0;margin-bottom:12px">
        Dirígete a la caja del operador<br>
        o llama al <b style="color:#84cc16">{tel_op}</b>
      </div>
      <div style="font-size:28px;font-weight:700;color:#00ff88;text-align:center">
        Total: {fmt_cop(monto)} COP
      </div>
    </div>
    """, unsafe_allow_html=True)

    if st.button(f"🔔 LLAMAR AL OPERADOR — {tel_op}", use_container_width=True, key="btn_llamar"):
        st.markdown(f"""
        <audio autoplay>
          <source src="data:audio/mp3;base64," type="audio/mp3">
        </audio>
        <script>
          var audio = new AudioContext();
          var osc = audio.createOscillator();
          osc.connect(audio.destination);
          osc.frequency.value = 880;
          osc.start();
          setTimeout(() => osc.stop(), 500);
        </script>
        """, unsafe_allow_html=True)
        st.success(f"🔔 Operador notificado · Tel: {tel_op}")

    st.markdown("**Sugerencia de denominaciones:**")
    c1, c2, c3 = st.columns(3)
    for col, bill in zip([c1, c2, c3], [20000, 50000, 100000]):
        cambio = bill - monto
        if cambio >= 0:
            col.markdown(f"""
            <div style="background:rgba(0,0,0,0.3);border:1px solid rgba(132,204,22,0.3);
                        border-radius:8px;padding:12px;text-align:center">
              <div style="font-size:18px;font-weight:700;color:#84cc16">{fmt_cop(bill)}</div>
              <div style="font-size:12px;color:#94a3b8">Cambio: {fmt_cop(cambio)}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>**Confirmación del operador (requiere PIN):**")
    pin_op = st.text_input("PIN operador", type="password", max_chars=8,
                           key="pin_efectivo", placeholder="••••")
    if st.button("✅ CONFIRMAR PAGO EN EFECTIVO", type="primary",
                 use_container_width=True, key="confirm_efect"):
        if verificar_pin(pin_op):
            _confirmar_pago("Efectivo")
        else:
            st.error("PIN incorrecto")


def _pago_otros(monto: int, ref: str):
    cuenta = get_config("cuenta_bancaria", CUENTA_BANCO)
    mp_link = get_config("mp_link", MP_LINK)
    metodo_sel = st.radio("Selecciona plataforma:",
                          ["PSE", "MercadoPago", "Transferencia bancaria", "Tarjeta (Datáfono)"],
                          horizontal=True, key="pago_otros_metodo")

    if metodo_sel == "PSE":
        st.info(f"**PSE · Referencia:** `{ref}` · **Monto:** {fmt_cop(monto)}")
        mostrar_qr("PSE", monto, ref, size=250)
        if st.button("✅ Confirmar pago PSE", type="primary",
                     use_container_width=True, key="confirm_pse"):
            _confirmar_pago("PSE")

    elif metodo_sel == "MercadoPago":
        st.markdown(f"""
        <div style="background:rgba(0,158,243,0.15);border:1px solid rgba(0,158,243,0.4);
                    border-radius:12px;padding:16px;margin-bottom:12px">
          <div style="font-size:17px;font-weight:700;color:#009ef3">🔵 MercadoPago Colombia</div>
          <div style="font-size:15px;margin-top:8px">
            Link de pago: <a href="{mp_link}" style="color:#009ef3">{mp_link}</a><br>
            Referencia: <b>{ref}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)
        mostrar_qr("MercadoPago", monto, ref, size=250)
        if st.button("✅ Confirmar pago MercadoPago", type="primary",
                     use_container_width=True, key="confirm_mp"):
            _confirmar_pago("MercadoPago")

    elif metodo_sel == "Transferencia bancaria":
        st.markdown(f"""
        <div style="background:rgba(100,100,0,0.15);border:1px solid rgba(200,200,0,0.3);
                    border-radius:12px;padding:16px;margin-bottom:12px">
          <div style="font-size:17px;font-weight:700;color:#eab308">🏦 Transferencia Bancaria</div>
          <div style="font-size:15px;margin-top:8px">
            <b>{cuenta}</b><br>
            NIT: {NIT}<br>
            Monto exacto: <b style="color:#00ff88">{fmt_cop(monto)} COP</b><br>
            Referencia: <b style="font-family:'Inconsolata'">{ref}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)
        mostrar_qr("Transferencia", monto, ref, size=250)
        pin_op2 = st.text_input("PIN operador para confirmar", type="password",
                                 max_chars=8, key="pin_transf")
        if st.button("✅ Confirmar transferencia recibida", type="primary",
                     use_container_width=True, key="confirm_transf"):
            if verificar_pin(pin_op2):
                _confirmar_pago("Transferencia")
            else:
                st.error("PIN incorrecto")

    else:
        tel_op = get_config("whatsapp_op", WHATSAPP_OP)
        st.markdown(f"""
        <div style="background:rgba(50,50,50,0.4);border:1px solid rgba(200,200,200,0.2);
                    border-radius:12px;padding:20px;text-align:center">
          <div style="font-size:60px">💳</div>
          <div style="font-size:18px;font-weight:700;margin-top:8px">Pago con Datáfono</div>
          <div style="font-size:16px;color:#94a3b8;margin-top:8px">
            El datáfono está disponible en caja.<br>
            Solicita al operador: <b>{tel_op}</b>
          </div>
        </div>
        """, unsafe_allow_html=True)


def _confirmar_pago(metodo: str):
    st.session_state.metodo_pago = metodo
    ir_a("confirmacion")


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 4 — CONFIRMACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def show_confirmacion():
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header("Procesando tu pago")
    render_stepper(3)

    if CONT_AVAILABLE and _cont_mod is not None:
        try:
            _cont_mod.on_pago_convenio(
                empresa    = nombre_empresa,
                valor      = monto_total,
                num_reserva= num_reserva,
                num_factura= num_factura,
            )
        except Exception:
            pass

    _, col, _ = st.columns([1, 2, 1])
    with col:
        # ── Guardia anti-doble-clic ────────────────────────────────────────────
        # Si ya hay un proceso en curso (mismo cubiculo + cliente), mostrar spinner
        # y no volver a ejecutar crear_reserva_completa.
        if st.session_state.get("_procesando_reserva"):
            st.info("⏳ Ya se está procesando una reserva, por favor espera...")
            return

        # Si ya se confirmó exitosamente en esta sesión, mostrar resultado directo
        if st.session_state.get("pago_confirmado") and st.session_state.get("voucher"):
            voucher = st.session_state["voucher"]
            st.markdown("""
            <div class="confirm-ok">
              <div style="font-size:80px">✅</div>
              <div style="font-size:32px;font-weight:800;color:#00ff88;margin:12px 0">
                ¡PAGO CONFIRMADO!
              </div>
              <div style="font-size:18px;color:#e2e8f0">
                Tu cubículo está listo. ¡Disfruta tu descanso!
              </div>
            </div>
            """, unsafe_allow_html=True)
            time.sleep(0.5)
            st.markdown(f"""
            <div style="text-align:center;margin:24px 0">
              <div style="font-size:16px;color:#94a3b8;letter-spacing:3px;
                          text-transform:uppercase">Tu cubículo</div>
              <div style="font-family:'Inconsolata';font-size:80px;font-weight:700;
                          color:#00d4ff;line-height:1">{voucher['cubiculo']}</div>
              <div style="font-size:16px;color:#94a3b8;letter-spacing:3px;
                          text-transform:uppercase;margin-top:8px">Código de acceso</div>
              <div style="font-family:'Inconsolata';font-size:72px;font-weight:700;
                          color:#00ff88;letter-spacing:16px;
                          text-shadow:0 0 40px rgba(0,255,136,0.6)">{voucher['codigo_acceso']}</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("📋 VER MI VOUCHER COMPLETO", type="primary", use_container_width=True):
                ir_a("voucher")
            return

        # ── Verificar reserva duplicada: mismo cliente con reserva activa hoy ──
        cliente_sel = st.session_state.get("cliente", {})
        num_doc     = cliente_sel.get("numero_documento", "")

        if not num_doc:
            st.session_state["_procesando_reserva"] = False
            st.error("❌ No se encontró documento del cliente en la sesión. Vuelve a ingresar los datos.")
            if st.button("← Volver a datos", use_container_width=True):
                ir_a("datos")
            return

        # Consulta de duplicado desde Google Sheets
        reserva_dup = None
        error_dup   = None
        try:
            hoy_str = ahora_col().strftime("%Y-%m-%d")
            reservas_gs = _gs_read_sheet("Reservas")
            for r in reservas_gs:
                if (str(r.get("Documento", "")) == num_doc and
                        _gs_val(r, "Estado_Pago") == "confirmado" and
                        not _gs_val(r, "Hora_Fin_Real") and
                        _gs_fecha_ymd(r, "Hora_Inicio") == hoy_str):
                    reserva_dup = (
                        _gs_val(r, "Numero_Reserva"),
                        _gs_val(r, "Hora_Inicio"),
                        _gs_val(r, "Hora_Fin_Prog"),
                        _gs_val(r, "Cubiculo_Num"),
                    )
                    break
        except Exception as e_dup:
            error_dup = str(e_dup)

        if error_dup:
            st.error(f"❌ Error al verificar reservas existentes: {error_dup}. No se puede continuar.")
            if st.button("← Volver al pago", use_container_width=True, key="btn_dup_err"):
                ir_a("pago")
            return

        if reserva_dup:
            num_res_dup, hi_dup, hf_dup, cub_num_dup = reserva_dup
            st.markdown(f"""
            <div class="confirm-fail">
              <div style="font-size:50px">⚠️</div>
              <div style="font-size:22px;font-weight:700;color:#ffd32a;margin:12px 0">
                Reserva duplicada detectada
              </div>
              <div style="font-size:15px;color:#e2e8f0">
                El cliente con documento <b>{num_doc}</b> ya tiene una reserva activa hoy:<br>
                <b>Reserva {num_res_dup}</b> · Cubículo <b>{cub_num_dup or '—'}</b><br>
                {hi_dup[:16] if hi_dup else ''} → {hf_dup[:16] if hf_dup else ''}
              </div>
              <div style="font-size:13px;color:#94a3b8;margin-top:12px">
                No se puede crear una nueva reserva para un cliente
                que ya tiene una reserva activa que aún no ha terminado.
              </div>
            </div>
            """, unsafe_allow_html=True)
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                if st.button("← Volver al pago", use_container_width=True, key="btn_dup_back"):
                    ir_a("pago")
            with col_d2:
                if st.button("✖ Ir al panel", use_container_width=True, key="btn_dup_panel"):
                    ir_a("operador")
            return  # ← SIEMPRE retorna aquí si hay duplicado

        # ── Marcar proceso en curso ANTES de ejecutar ──────────────────────────
        st.session_state["_procesando_reserva"] = True

        with st.spinner("⚙️ Activando tu cubículo..."):
            time.sleep(1.2)

        try:
            voucher = crear_reserva_completa(
                st.session_state.cubiculo_sel,
                st.session_state.cliente,
                st.session_state.calc,
                st.session_state.metodo_pago,
                acepto_datos=st.session_state.cliente.get("acepto_datos", False),
            )
            st.session_state.voucher        = voucher
            st.session_state.pago_confirmado = True
            st.session_state["_procesando_reserva"] = False

            st.markdown("""
            <div class="confirm-ok">
              <div style="font-size:80px">✅</div>
              <div style="font-size:32px;font-weight:800;color:#00ff88;margin:12px 0">
                ¡PAGO CONFIRMADO!
              </div>
              <div style="font-size:18px;color:#e2e8f0">
                Tu cubículo está listo. ¡Disfruta tu descanso!
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Código de acceso grande
            time.sleep(0.5)
            st.markdown(f"""
            <div style="text-align:center;margin:24px 0">
              <div style="font-size:16px;color:#94a3b8;letter-spacing:3px;
                          text-transform:uppercase">Tu cubículo</div>
              <div style="font-family:'Inconsolata';font-size:80px;font-weight:700;
                          color:#00d4ff;line-height:1">{voucher['cubiculo']}</div>
              <div style="font-size:16px;color:#94a3b8;letter-spacing:3px;
                          text-transform:uppercase;margin-top:8px">Código de acceso</div>
              <div style="font-family:'Inconsolata';font-size:72px;font-weight:700;
                          color:#00ff88;letter-spacing:16px;
                          text-shadow:0 0 40px rgba(0,255,136,0.6)">{voucher['codigo_acceso']}</div>
            </div>
            """, unsafe_allow_html=True)

            time.sleep(0.8)
            if st.button("📋 VER MI VOUCHER COMPLETO", type="primary", use_container_width=True):
                ir_a("voucher")

        except Exception as e:
            st.session_state["_procesando_reserva"] = False
            st.markdown(f"""
            <div class="confirm-fail">
              <div style="font-size:60px">❌</div>
              <div style="font-size:24px;font-weight:700;color:#ff4757;margin:12px 0">
                Error al procesar el pago
              </div>
              <div style="font-size:15px;color:#94a3b8">{str(e)}</div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("🔄 Intentar de nuevo", use_container_width=True):
                ir_a("pago")


# ══════════════════════════════════════════════════════════════════════════════
# PANTALLA 5 — VOUCHER
# ══════════════════════════════════════════════════════════════════════════════

def show_voucher():
    if not st.session_state.get("operador_ok"):
        ir_a("operador_login")
        return
    render_header("Tu reserva confirmada")
    render_stepper(4)
    inject_live_clock()

    v = st.session_state.voucher
    if not v:
        ir_a("bienvenida")
        return

    # Obtener hora_inicio real de Sheets para el timer de tiempo consumido
    hora_inicio_iso = ""
    hora_fin_iso = ""
    try:
        num_res_v = v.get("numero_reserva", "")
        reservas_gs = _gs_read_sheet("Reservas")
        for r_v in reservas_gs:
            if _gs_val(r_v, "Numero_Reserva") == num_res_v:
                hora_inicio_iso = _gs_val(r_v, "Hora_Inicio", "")
                hora_fin_iso    = _gs_val(r_v, "Hora_Fin_Prog", "")
                break
    except Exception:
        pass

    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown(f"""
        <div class="voucher-box">
          <div style="text-align:center;margin-bottom:16px">
            <div style="font-size:22px;font-weight:800;color:#00d4ff">{NEGOCIO}</div>
            <div style="font-size:13px;color:#94a3b8">{TAGLINE}</div>
          </div>
          <div style="background:rgba(0,0,0,0.3);border-radius:8px;padding:12px;margin-bottom:12px">
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Reserva</span>
              <b style="font-family:'Inconsolata'">{v['numero_reserva']}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Factura</span>
              <b style="font-family:'Inconsolata'">{v['numero_factura']}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Cubículo</span>
              <b style="color:#00d4ff;font-size:20px">{v['cubiculo']}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Entrada → Salida</span>
              <b>{v['hora_inicio']} → {v['hora_fin']}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Tiempo</span>
              <b>{v['horas']}h</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Tiempo restante</span>
              <b class="jjgt-timer" data-fin="{hora_fin_iso}" data-key="voucher-fin"
                 style="color:#00ff88;font-family:'Inconsolata'">calculando...</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">Consumido</span>
              <span class="jjgt-consumed" data-inicio="{hora_inicio_iso}"
                    style="color:#94a3b8;font-family:'Inconsolata'">calculando...</span>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0;
                        border-bottom:1px solid rgba(255,255,255,0.1)">
              <span style="color:#94a3b8">WiFi</span>
              <b>{v['wifi_ssid']}</b>
            </div>
            <div style="display:flex;justify-content:space-between;padding:4px 0">
              <span style="color:#94a3b8">Clave WiFi</span>
              <b style="font-family:'Inconsolata'">{v['wifi_password']}</b>
            </div>
          </div>
          <div style="text-align:center;margin:4px 0 12px">
            <div style="font-size:12px;color:#94a3b8;letter-spacing:2px;text-transform:uppercase">
              Código de acceso al cubículo
            </div>
            <div class="voucher-codigo">{v['codigo_acceso']}</div>
          </div>
          <div style="background:rgba(0,0,0,0.3);border-radius:8px;padding:12px">
            <div style="display:flex;justify-content:space-between;padding:3px 0">
              <span style="color:#94a3b8">Subtotal</span><span>{fmt_cop(v['subtotal'])}</span>
            </div>
            <div style="display:flex;justify-content:space-between;padding:3px 0">
              <span style="color:#94a3b8">IVA 19%</span><span>{fmt_cop(v['iva'])}</span>
            </div>
            <div style="display:flex;justify-content:space-between;padding:6px 0;
                        font-family:'Inconsolata';font-size:22px;font-weight:700;color:#00ff88;
                        border-top:1px solid rgba(255,255,255,0.15);margin-top:4px">
              <span>TOTAL</span><span>{fmt_cop(v['total'])} COP</span>
            </div>
            <div style="display:flex;justify-content:space-between;padding:3px 0">
              <span style="color:#94a3b8">Pago</span><b>{v['metodo_pago']}</b>
            </div>
          </div>
          <div style="display:flex;gap:8px;margin-top:16px;flex-wrap:wrap;justify-content:center">
            <span style="background:rgba(0,255,136,0.15);color:#00ff88;padding:6px 14px;
                         border-radius:20px;font-size:13px;font-weight:700">✓ Baño</span>
            <span style="background:rgba(0,255,136,0.15);color:#00ff88;padding:6px 14px;
                         border-radius:20px;font-size:13px;font-weight:700">✓ WiFi</span>
            <span style="background:rgba(0,255,136,0.15);color:#00ff88;padding:6px 14px;
                         border-radius:20px;font-size:13px;font-weight:700">✓ Carga</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Botones de acción — 4 columnas al mismo nivel (sin anidamiento)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        wa_msg = (f"JJGT Reserva {v['numero_reserva']} · Cubículo {v['cubiculo']} · "
                  f"Código: {v['codigo_acceso']} · WiFi: {v['wifi_ssid']} Clave: {v['wifi_password']} · "
                  f"Total: {fmt_cop(v['total'])} COP · Factura: {v['numero_factura']}")
        wa_url = f"https://wa.me/?text={wa_msg.replace(' ', '%20')}"
        st.markdown(f'<a href="{wa_url}" target="_blank"><button style="width:100%;padding:14px;'
                    f'background:rgba(37,211,102,0.2);border:2px solid rgba(37,211,102,0.5);'
                    f'border-radius:12px;color:#25d366;font-weight:700;font-size:16px;cursor:pointer">'
                    f'📱 WhatsApp</button></a>', unsafe_allow_html=True)

    with c2:
        if REPORTLAB_AVAILABLE:
            pdf_bytes = generar_ticket_pdf(v)
            if pdf_bytes:
                st.download_button("📄 Descargar PDF", data=pdf_bytes,
                                   file_name=f"ticket_{v['numero_reserva']}.pdf",
                                   mime="application/pdf", use_container_width=True)
        else:
            html_v = voucher_html(v)
            st.download_button("📄 Voucher HTML", data=html_v.encode(),
                               file_name=f"voucher_{v['numero_reserva']}.html",
                               mime="text/html", use_container_width=True)

    with c3:
        if st.button("✅ IR A MI CUBÍCULO", type="primary", use_container_width=True):
            for k in ["cubiculo_sel","calc","metodo_pago","voucher","pago_confirmado","cliente"]:
                if k == "horas_sel":
                    st.session_state[k] = 1.0
                elif k in ["cubiculo_sel","calc","metodo_pago","voucher"]:
                    st.session_state[k] = None
                elif k == "pago_confirmado":
                    st.session_state[k] = False
                elif k == "cliente":
                    st.session_state[k] = {}
            st.session_state["pantalla"] = "bienvenida"
            st.rerun()

    with c4:
        if st.button("🔧 Volver al Panel", use_container_width=True):
            for k in ["cubiculo_sel","calc","metodo_pago","voucher","pago_confirmado","cliente"]:
                st.session_state.pop(k, None)
            st.session_state["cubiculo_sel"]    = None
            st.session_state["calc"]            = None
            st.session_state["metodo_pago"]     = None
            st.session_state["voucher"]         = None
            st.session_state["pago_confirmado"] = False
            st.session_state["cliente"]         = {}
            ir_a("operador")


# ══════════════════════════════════════════════════════════════════════════════
# PANEL OPERADOR — LOGIN
# ══════════════════════════════════════════════════════════════════════════════

def show_operador_login():
    """
    Pantalla de login de operador — PANTALLA INICIAL de la app.
    Si el operador ya está autenticado (sesión activa), va directo al panel.
    Siempre muestra título y reloj.
    """
    # Si ya está autenticado, ir directamente al panel
    if st.session_state.get("operador_ok"):
        ir_a("operador")
        return

    render_header("Acceso de Operadores")
    inject_live_clock()
    render_reloj()

    st.markdown("<br>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1, 1])
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🔐 Identificación de Operador")
        st.caption("Ingresa tu PIN para iniciar tu turno")

        # Selector de turno visual
        ahora_h = ahora_col().hour
        turno_actual = "mañana" if 6 <= ahora_h < 14 else ("tarde" if 14 <= ahora_h < 22 else "noche")
        turno_label  = {"mañana": "☀️ Turno Mañana (06:00–14:00)",
                        "tarde":  "🌤️ Turno Tarde (14:00–22:00)",
                        "noche":  "🌙 Turno Noche (22:00–06:00)"}
        st.info(f"**Turno activo:** {turno_label.get(turno_actual, turno_actual)}")

        pin = st.text_input("PIN de operador", type="password", max_chars=8,
                            placeholder="••••", key="pin_login_input")

        if st.button("✅ INGRESAR AL SISTEMA", type="primary",
                     use_container_width=True, key="btn_login"):
            op = get_operador_por_pin(pin)
            if op:
                st.session_state.operador_ok      = True
                st.session_state.operador_info    = op
                st.session_state.pin_intentos     = 0
                st.session_state.pantalla         = "operador"
                st.rerun()
            else:
                st.session_state.pin_intentos += 1
                intentos = st.session_state.pin_intentos
                if intentos >= 3:
                    st.error("🔒 Demasiados intentos incorrectos. Contacta al administrador.")
                else:
                    st.error(f"❌ PIN incorrecto ({intentos}/3)")
                    if intentos >= 3:
                        if st.button("🔄 Reiniciar intentos", use_container_width=True, key="btn_reset_intentos"):
                            st.session_state.pin_intentos = 0
                            st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

    # ── Info de PINs de acceso inicial — SIEMPRE visible ─────────────────────
    ops_gs = _gs_read_sheet("Operadores")
    _, sh_status = _get_module_level_client()

    with st.expander("ℹ️ PINs de acceso inicial y configuración", expanded=not sh_status):
        #if not sh_status:
        #    st.error("⚠️ **Sin conexión a Google Sheets** — Revisa las credenciales en secrets.toml")
        #    st.markdown("""
#**Para configurar la conexión:**
#1. Crea el archivo `.streamlit/secrets.toml` con:
#```toml
#[sheetsemp]
#credentials_sheet = '''{ ... JSON de tu Service Account ... }'''
#spreadsheet_id = "ID_DEL_SPREADSHEET"
#```
#2. O en Streamlit Cloud → **App Settings → Secrets**
#            """)
#            st.divider()

#        st.markdown("**🔑 PINs por defecto (primer acceso):**")
#        st.markdown("""
#| Operador | PIN | Rol | Turno |
#|---|---|---|---|
#| **Admin JJGT** | `1234` | admin | Todos los módulos |
#| Op. Mañana | `1111` | cajero | 06:00–14:00 |
#| Op. Tarde | `2222` | cajero | 14:00–22:00 |
#| Op. Noche | `3333` | cajero | 22:00–06:00 |
#        """)
#        st.warning("⚠️ **Cambia estos PINs** en ⚙️ Configuración → Operadores tras el primer acceso.")

        if not ops_gs:
            st.info(
                "📋 **Hoja Operadores vacía** — Los PINs por defecto de arriba funcionarán "
                "para el primer acceso. Después crea los operadores reales en ⚙️ Configuración."
            )
        else:
            st.success(f"✅ {len(ops_gs)} operador(es) encontrado(s) en Google Sheets")
            st.caption("Para operadores creados manualmente en Sheets, el PIN es el configurado al crearlos.")

        if st.button("🔄 Reiniciar contador de intentos", key="btn_rst_global"):
            st.session_state.pin_intentos = 0
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PANEL OPERADOR — MÓDULOS
# ══════════════════════════════════════════════════════════════════════════════

def show_operador():
    if not st.session_state.operador_ok:
        ir_a("operador_login")
        return

    # ── Inyectar contexto a módulos externos en cada render ───────────────────
    if CONT_AVAILABLE and _cont_mod is not None:
        _cont_mod.set_context(globals())
    if FC_AVAILABLE and _fc_mod is not None:
        _fc_mod.set_context(globals())
    if FE_AVAILABLE and _fe_mod is not None:
        _fe_mod.set_context(globals())
    # ─────────────────────────────────────────────────────────────────────────

    op = st.session_state.get("operador_info", {})
    permisos = op.get("permisos", ["reservas","pagos","voucher"])
    es_admin = "admin" in permisos or op.get("rol") == "admin"

    # ── Reintento silencioso de sync de dashboard pendiente ───────────────────
    # Si una reserva anterior no pudo sincronizar el Dashboard_Diario por caída
    # de conexión (Nequi sin comprobante, timeout, etc.), se reintenta aquí de
    # forma silenciosa cada vez que el operador abre su panel.
    if st.session_state.get("_dashboard_sync_pendiente"):
        try:
            _, _sh_retry = get_active_client()
            if _sh_retry:
                _gs_invalidate_cache("Reservas")
                exito = gs_sync_dashboard(_sh_retry)
                if exito:
                    err_prev = st.session_state.pop("_dashboard_sync_error", "")
                    st.toast("✅ Dashboard sincronizado (reintento exitoso)", icon="✅")
        except Exception:
            pass  # Se volverá a intentar en el próximo render

    # Auto-rerun: cada 30s en Dashboard, cada 60s en Cubículos
    now_ts = int(time.time())
    last_rerun = st.session_state.get("_op_last_rerun", 0)
    current_mod = st.session_state.get("modulo_op_radio", "🏠 Dashboard")
    rerun_interval = 60 if current_mod == "🏠 Dashboard" else 120
    if last_rerun == 0:
        st.session_state["_op_last_rerun"] = now_ts
    elif now_ts - last_rerun >= rerun_interval:
        if current_mod in ("🏠 Dashboard", "🛏️ Cubículos"):
            st.session_state["_op_last_rerun"] = now_ts
            st.rerun()

    # Sidebar
    with st.sidebar:
        if _LOGO_B64:
            st.markdown(
                f'<div style="text-align:center;padding:8px 0 4px">'
                f'{_LOGO_HTML}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(f"## 💤 {NEGOCIO}")
        #st.divider()

        # Reloj en el sidebar (actualizado por JS iframe)
        inject_live_clock()
        ahora_sb = ahora_col()
        st.markdown(f"""
        <div style="text-align:center;margin-bottom:8px">
          <div class="jjgt-clock-hms" style="font-family:'Inconsolata',monospace;
               font-size:28px;font-weight:700;color:#00d4ff;
               text-shadow:0 0 12px rgba(0,212,255,0.5)">
            {ahora_sb.strftime('%H:%M:%S')}
          </div>
          <div class="jjgt-clock-fecha" style="font-size:11px;color:#94a3b8;
               letter-spacing:1px;text-transform:uppercase">
            {ahora_sb.strftime('%a %d/%m/%Y').upper()}
          </div>
        </div>
        """, unsafe_allow_html=True)
        #st.divider()

        # Info del operador activo
        turno_icons = {"mañana":"☀️","tarde":"🌤️","noche":"🌙","diurno":"☀️","admin":"👑"}
        t_icon = turno_icons.get(op.get("turno",""), "👤")
        st.markdown(f"""
        <div style="background:rgba(0,212,255,0.1);border:1px solid rgba(0,212,255,0.3);
                    border-radius:10px;padding:12px;margin-bottom:8px">
          <div style="font-weight:700;color:#00d4ff">{t_icon} {op.get("nombre","Operador")}</div>
          <div style="font-size:12px;color:#94a3b8">Turno: {op.get("turno","").capitalize()}</div>
          <div style="font-size:12px;color:#94a3b8">{op.get("hora_inicio_turno","--")} → {op.get("hora_fin_turno","--")}</div>
          <div style="font-size:11px;color:#94a3b8;margin-top:4px">Rol: {op.get("rol","cajero").capitalize()}</div>
        </div>
        """, unsafe_allow_html=True)

        # Módulos filtrados por permisos
        todos_modulos = [
            ("🏠 Dashboard",             True),
            ("➕ Nueva Reserva",          "reservas" in permisos or es_admin),
            ("🤝 Convenios",              "reservas" in permisos or es_admin),
            ("🛏️ Cubículos",             "reservas" in permisos or es_admin),
            ("⏳ Pagos Pendientes",       "pagos"    in permisos or es_admin),
            ("💳 Facturación & Cartera",  es_admin and FC_AVAILABLE),
            ("⚡ Factura Electrónica",    es_admin and FE_AVAILABLE),
            ("📒 Contabilidad",           es_admin and CONT_AVAILABLE),
            ("📊 Reportes",              "reportes"  in permisos or es_admin),
            ("🗑️ Gestión de Datos",      es_admin),
            ("☁️ Google Drive",          es_admin),
            ("⚙️ Configuración",         "configuracion" in permisos or es_admin),
        ]
        modulos_visibles = [m for m, visible in todos_modulos if visible]

        modulo = st.radio("Módulo", modulos_visibles, key="modulo_op_radio")
        st.divider()

        #drive_ok = GSPREAD_AVAILABLE and bool(get_config("drive_spreadsheet_id",""))
        #conv_ok  = GSPREAD_AVAILABLE and bool(get_config("convenios_spreadsheet_id",""))
        #st.caption("🟢 Sync Drive activa" if drive_ok else "🔴 Drive sin configurar")
        #st.caption("🤝 Convenios activo" if conv_ok else "⚪ Convenios sin configurar")

        # ── Backup pendiente de descarga (Cloud) ─────────────────────────
        if st.session_state.get("_backup_cierre_bytes"):
            st.download_button(
                "📥 Descargar backup de cierre",
                data=st.session_state["_backup_cierre_bytes"],
                file_name=st.session_state.get("_backup_cierre_fname", "backup_cierre.xlsx"),
                mime=st.session_state.get("_backup_cierre_mime",
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                use_container_width=True,
                key="btn_dl_backup_cierre",
            )
            if st.button("✅ Ya descargué — Cerrar sesión", use_container_width=True,
                         key="btn_logout_post_backup"):
                for k in ["_backup_cierre_bytes","_backup_cierre_fname","_backup_cierre_mime"]:
                    st.session_state.pop(k, None)
                st.session_state.operador_ok   = False
                st.session_state.operador_info = {}
                ir_a("operador_login")
        else:
            if st.button("🚪 Cerrar sesión", use_container_width=True):
                # Genera backup (local → escribe a disco / Cloud → guarda en session_state)
                _backup_cierre_turno(op)
                if _IS_CLOUD and st.session_state.get("_backup_cierre_bytes"):
                    # En Cloud: mantener sesión abierta para que el operador pueda descargar
                    st.rerun()
                else:
                    st.session_state.operador_ok   = False
                    st.session_state.operador_info = {}
                    ir_a("operador_login")

    render_header(f"Panel — {op.get('nombre','Operador')}")

    # Mapa de módulos — claves exactas del radio button
    mod_map = {
        "🏠 Dashboard":            _op_dashboard,
        "➕ Nueva Reserva":        _op_nueva_reserva,
        "🤝 Convenios":            _op_convenios,
        "🛏️ Cubículos":            _op_cubiculos,
        "⏳ Pagos Pendientes":      _op_pagos_pendientes,
        "💳 Facturación & Cartera": _fc_mod.show_facturacion_cartera if FC_AVAILABLE else _op_dashboard,
        "⚡ Factura Electrónica":   _fe_mod.render_panel_fe if FE_AVAILABLE else _op_dashboard,
        "📒 Contabilidad":          _cont_mod.render_panel_contabilidad if CONT_AVAILABLE else _op_dashboard,
        "📊 Reportes":             _op_reportes,
        "🗑️ Gestión de Datos":     _op_gestion_datos,
        "☁️ Google Drive":         _op_google_drive,
        "⚙️ Configuración":        _op_configuracion,
    }

    # Router robusto: busca por substring de la parte de texto (sin emoji)
    # para tolerar variantes de selector de emoji entre plataformas (Windows, Mac, Linux)
    _txt_map = {
    "dashboard":              _op_dashboard,
    "nueva reserva":          _op_nueva_reserva,
    "convenios":              _op_convenios,
    "cubículos":              _op_cubiculos,
    "pagos pendientes":       _op_pagos_pendientes,
    "facturación":            _fc_mod.show_facturacion_cartera if FC_AVAILABLE else _op_dashboard,
    "factura electrónica":    _fe_mod.render_panel_fe if FE_AVAILABLE else _op_dashboard,
    "electronica":            _fe_mod.render_panel_fe if FE_AVAILABLE else _op_dashboard,
    "contabilidad":           _cont_mod.render_panel_contabilidad if CONT_AVAILABLE else _op_dashboard,
    "cartera":                _fc_mod.show_facturacion_cartera if FC_AVAILABLE else _op_dashboard,
    "reportes":               _op_reportes,
    "gestion de datos":       _op_gestion_datos,
    "google drive":           _op_google_drive,
    "configuración":          _op_configuracion,
    }

    func = mod_map.get(modulo)
    if func is None:
        modulo_lower = modulo.lower()
        for txt, fn in _txt_map.items():
            if txt in modulo_lower:
                func = fn
                break

    (func or _op_dashboard)()


def _op_nueva_reserva():
    """
    Módulo del panel de operador para crear una reserva.
    Dos modos:
    - Rápido (inline): formulario completo en el panel — ideal para operadores
    - Kiosk: activa el flujo del kiosco público paso a paso
    """
    st.markdown("### ➕ Nueva Reserva")
    op = st.session_state.get("operador_info", {})
    st.caption(f"Operador: **{op.get('nombre','—')}** · Turno: {op.get('turno','—')}")

    tab_rapido, tab_noche, tab_kiosk = st.tabs(["⚡ Formulario Rápido", "🌙 Noche Completa", "🖥️ Flujo Kiosco"]) # , "🤝 Convenio Empresarial"

    # ── TAB KIOSCO ────────────────────────────────────────────────────────────
    with tab_kiosk:
        st.markdown("Activa el flujo completo del kiosco táctil paso a paso.")
        st.info("El cliente selecciona cubículo, ingresa sus datos y elige método de pago.")
        cubiculos_k = get_cubiculos_libres()
        if cubiculos_k:
            if st.button("🛏️ INICIAR FLUJO KIOSCO", type="primary",
                         use_container_width=True, key="btn_kiosco_flow"):
                st.session_state["cubiculo_sel"]    = None
                st.session_state["horas_sel"]       = 1.0
                st.session_state["calc"]            = None
                st.session_state["cliente"]         = {}
                st.session_state["metodo_pago"]     = None
                st.session_state["voucher"]         = None
                st.session_state["pago_confirmado"] = False
                ir_a("seleccion")
        else:
            st.error("❌ No hay cubículos disponibles en este momento.")

    # ── TAB NOCHE COMPLETA ────────────────────────────────────────────────────
    with tab_noche:
        st.markdown("**Reserva de Noche Completa — horario y precio fijo desde configuración:**")

        # Leer tarifa Noche Completa desde Sheets
        _rows_nc = _gs_read_sheet("Tarifas_Config")
        _tarifa_nc = None
        for _r_nc in _rows_nc:
            if _gs_val(_r_nc, "Nombre") == "Noche Completa" and \
               _gs_val(_r_nc, "Activo", "1") in ("1", "True", "true"):
                _tarifa_nc = _r_nc
                break

        # ── Leer Horas_A_Reservar y Precio_Hora_COP desde fila "Tarifa Unica" ──
        _horas_config_tu  = 0.0
        _precio_config_tu = 0.0
        _desc_tarifa_unica_encontrada = False
        for _r_tu in _rows_nc:
            if _gs_val(_r_tu, "Descripcion", "").strip().lower() == "tarifa unica" and \
               _gs_val(_r_tu, "Activo", "1") in ("1", "True", "true"):
                _horas_raw_tu = _gs_val(_r_tu, "Horas_A_Reservar", "")
                try:
                    _horas_config_tu = float(str(_horas_raw_tu).replace(",", ".").strip())
                except (ValueError, TypeError):
                    _horas_config_tu = 0.0
                _precio_config_tu = _gs_float(_r_tu, "Precio_Hora_COP", 0.0)
                _desc_tarifa_unica_encontrada = True
                break

        if not _tarifa_nc:
            st.error(
                "❌ No se encontró la tarifa **'Noche Completa'** activa en la hoja "
                "**Tarifas_Config**. Verifica que exista una fila con Nombre='Noche Completa' "
                "y Activo='1'."
            )
        else:
            _nc_precio  = _gs_float(_tarifa_nc, "Precio_Hora_COP", 0)
            _nc_ini_str = _gs_val(_tarifa_nc, "Hora_Ini_Espec", "")
            _nc_fin_str = _gs_val(_tarifa_nc, "Hora_Fin_Espec", "")
            _nc_desc    = _gs_val(_tarifa_nc, "Descripcion", "Noche Completa")

            # ── Precio COP: prioridad → Tarifa Unica > fila Noche Completa ────────
            # 1. Si existe "Tarifa Unica" con Precio_Hora_COP > 0, usar ese.
            # 2. Si no, usar el Precio_Hora_COP de la fila Noche Completa (_nc_precio).
            _nc_precio_base = _precio_config_tu if _precio_config_tu > 0 else _nc_precio
            _nc_precio_fuente = "Tarifa Unica" if _precio_config_tu > 0 else "Noche Completa"

            # ── Horas A Reservar: prioridad → Tarifa Unica > cálculo por horario ─
            # 1. Si existe configuración "Tarifa Unica" con Horas_A_Reservar > 0, usarla.
            # 2. Si no, calcular a partir de Hora_Ini_Espec → Hora_Fin_Espec.
            _nc_horas = 0.0
            _nc_horas_fuente = "calculada"  # para mostrar al operador

            if _horas_config_tu > 0:
                _nc_horas = _horas_config_tu
                _nc_horas_fuente = "Tarifa Unica"
            elif _nc_ini_str and _nc_fin_str:
                try:
                    _t_ini = datetime.strptime(_nc_ini_str, "%H:%M")
                    _t_fin = datetime.strptime(_nc_fin_str, "%H:%M")
                    _dt_ini = ahora_col().replace(hour=_t_ini.hour, minute=_t_ini.minute,
                                                  second=0, microsecond=0)
                    _dt_fin = ahora_col().replace(hour=_t_fin.hour, minute=_t_fin.minute,
                                                  second=0, microsecond=0)
                    if _dt_fin <= _dt_ini:
                        _dt_fin += timedelta(days=1)
                    _nc_horas = round((_dt_fin - _dt_ini).total_seconds() / 3600, 2)
                    _nc_horas_fuente = "calculada"
                except Exception:
                    _nc_horas = 0.0

            # ── Campos editables: Precio COP y Horas A Reservar ──────────────────
            _col_hr1, _col_hr2, _col_hr3 = st.columns([2, 2, 1])

            with _col_hr1:
                _nc_precio_input = st.number_input(
                    "💰 Precio Tarifa (COP)",
                    min_value=0,
                    max_value=9_999_999,
                    value=int(_nc_precio_base) if _nc_precio_base > 0 else 0,
                    step=500,
                    format="%d",
                    key="nc_precio_reservar",
                    help=(
                        f"Cargado desde Tarifas_Config · Descripcion='Tarifa Unica' "
                        f"(Precio_Hora_COP = {fmt_cop(_precio_config_tu)} COP). "
                        f"Puedes ajustarlo manualmente; este valor se usará como precio total."
                        if _desc_tarifa_unica_encontrada and _precio_config_tu > 0
                        else (
                            f"Cargado desde fila 'Noche Completa' "
                            f"(Precio_Hora_COP = {fmt_cop(_nc_precio)} COP). "
                            f"Puedes ajustarlo manualmente."
                            if _nc_precio > 0
                            else "Ingresa el precio total de la reserva."
                        )
                    ),
                )

            with _col_hr2:
                _nc_horas_input = st.number_input(
                    "🕐 Horas A Reservar",
                    min_value=0.5,
                    max_value=24.0,
                    value=float(_nc_horas) if _nc_horas > 0 else 1.0,
                    step=0.5,
                    format="%.1f",
                    key="nc_horas_reservar",
                    help=(
                        f"Valor cargado desde Tarifas_Config · Descripcion='Tarifa Unica' "
                        f"({_nc_horas_fuente}). Puedes ajustarlo manualmente."
                        if _desc_tarifa_unica_encontrada
                        else "Ingresa manualmente las horas a reservar."
                    ),
                )

            with _col_hr3:
                if _desc_tarifa_unica_encontrada:
                    _badge_precio_txt = fmt_cop(_precio_config_tu) if _precio_config_tu > 0 else "—"
                    _badge_horas_txt  = f"{_horas_config_tu:.1f} h" if _horas_config_tu > 0 else "—"
                    st.markdown(
                        f"<div style='padding-top:8px;font-size:12px;color:#a29bfe;line-height:1.7'>"
                        f"📋 <b>Tarifa Unica</b><br>"
                        f"<span style='color:#00ff88'>💰 {_badge_precio_txt}</span><br>"
                        f"<span style='color:#a29bfe'>🕐 {_badge_horas_txt}</span><br>"
                        f"<span style='color:#94a3b8;font-size:10px'>Sheets · Config</span></div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        "<div style='padding-top:8px;font-size:11px;color:#ffd32a;line-height:1.6'>"
                        "⚠️ Sin fila<br>'Tarifa Unica'<br>en Tarifas_Config</div>",
                        unsafe_allow_html=True,
                    )

            # Usar los valores ingresados/ajustados por el operador
            _nc_horas       = _nc_horas_input
            _nc_precio_final = float(_nc_precio_input)

            # Banner informativo de la tarifa
            st.markdown(f"""
            <div style="background:rgba(162,155,254,0.10);border:1.5px solid rgba(162,155,254,0.4);
                        border-radius:12px;padding:16px 20px;margin-bottom:16px">
              <div style="font-size:15px;font-weight:700;color:#a29bfe;letter-spacing:1px">
                🌙 {_nc_desc}
                <span style="font-size:11px;font-weight:400;color:#94a3b8;margin-left:8px">
                  · precio desde: {_nc_precio_fuente}
                </span>
              </div>
              <div style="margin-top:8px;font-size:14px;color:#e2e8f0">
                <b>Horario fijo:</b> {_nc_ini_str or '—'} → {_nc_fin_str or '—'}
                &nbsp;·&nbsp; <b>Horas a reservar:</b>
                <span style="color:#a29bfe;font-weight:700">{_nc_horas:.1f} h</span>
                &nbsp;·&nbsp; <b>Precio total:</b>
                <span style="color:#00ff88;font-weight:700">{fmt_cop(_nc_precio_final)} COP</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

            if not _nc_ini_str or not _nc_fin_str:
                st.warning(
                    "⚠️ La tarifa 'Noche Completa' no tiene **Hora_Ini_Espec** y/o "
                    "**Hora_Fin_Espec** configuradas. Completa esos campos en Tarifas_Config."
                )
            else:
                # Calcular precio completo usando horas y precio confirmados por el operador
                _calc_nc = calcular_precio(_nc_horas, "Noche Completa",
                                           precio_override=_nc_precio_final)

                # Paso 1: Cubículo
                st.markdown("#### 🛏️ Paso 1 — Cubículo")
                _cubiculos_nc = get_cubiculos_libres()
                if not _cubiculos_nc:
                    st.error("❌ No hay cubículos disponibles.")
                else:
                    _opc_nc = {f"{c['numero']} — {c['nombre']}": c for c in _cubiculos_nc}
                    _sel_nc = st.selectbox("Cubículo libre", list(_opc_nc.keys()),
                                           key="nc_cubiculo")
                    _cub_nc = _opc_nc[_sel_nc]

                    # Resumen de precio
                    st.markdown(f"""
                    <div style="background:rgba(0,212,255,0.08);border:1px solid rgba(0,212,255,0.2);
                                border-radius:8px;padding:10px;margin-top:4px;font-size:14px">
                      Tarifa: <b>Noche Completa</b> ·
                      Horario: <b>{_nc_ini_str} → {_nc_fin_str}</b> ·
                      Horas: <b style="color:#a29bfe">{_nc_horas:.1f} h</b><br>
                      Precio unitario aplicado:
                      <b style="color:#ffd32a">{fmt_cop(_nc_precio_final)} COP</b>
                      &nbsp;·&nbsp;
                      <b style="color:#00ff88;font-size:16px">Total: {fmt_cop(_calc_nc['total'])} COP</b>
                    </div>
                    """, unsafe_allow_html=True)

                    st.divider()

                    # Paso 2: Datos del cliente
                    st.markdown("#### 👤 Paso 2 — Datos del cliente")
                    _col_nc1, _col_nc2 = st.columns(2)
                    with _col_nc1:
                        _nc_nombre   = st.text_input("Nombre completo *", key="nc_nombre",
                                                     placeholder="Nombre del pasajero")
                        _nc_tel      = st.text_input("Teléfono", key="nc_tel",
                                                     placeholder="300 000 0000")
                        _nc_tipo_doc = st.selectbox("Tipo documento",
                                                    ["CC","CE","Pasaporte","NIT","TI"],
                                                    key="nc_tipodoc")
                    with _col_nc2:
                        _nc_doc      = st.text_input("Número de documento", key="nc_doc")
                        _nc_email    = st.text_input("Email (opcional)", key="nc_email")
                        _nc_req_fact = st.checkbox("¿Requiere factura empresarial?",
                                                   key="nc_req_fact")

                    _nc_razon = ""
                    _nc_nit   = ""
                    if _nc_req_fact:
                        _col_nf1, _col_nf2 = st.columns(2)
                        with _col_nf1:
                            _nc_razon = st.text_input("Razón social", key="nc_razon")
                        with _col_nf2:
                            _nc_nit   = st.text_input("NIT empresa",  key="nc_nit")

                    st.divider()

                    # Paso 3: Método de pago
                    st.markdown("#### 💳 Paso 3 — Método de pago")
                    _nc_metodo = st.selectbox("Método de pago", METODOS_PAGO, key="nc_metodo")
                    _nc_monto  = int(_calc_nc["total"])
                    _nc_ref_pago = (f"NC-{ahora_col().strftime('%Y%m%d%H%M')}-"
                                    f"{_cub_nc['numero'].replace('#','')}")

                    METODOS_CON_QR_NC = ["Nequi", "Daviplata", "PSE", "MercadoPago", "Transferencia"]
                    if _nc_metodo in METODOS_CON_QR_NC:
                        st.markdown(f"**Monto a cobrar:** `{fmt_cop(_nc_monto)} COP`")
                        mostrar_qr(_nc_metodo, _nc_monto, _nc_ref_pago, size=220)
                        st.caption(f"Referencia: `{_nc_ref_pago}`")
                    elif _nc_metodo == "Efectivo":
                        st.markdown(f"""
                        <div style="background:rgba(132,204,22,0.1);border:1px solid rgba(132,204,22,0.3);
                                    border-radius:10px;padding:16px;text-align:center">
                          <div style="font-size:13px;color:#94a3b8;margin-bottom:4px">TOTAL EN EFECTIVO</div>
                          <div style="font-size:36px;font-weight:700;color:#84cc16">{fmt_cop(_nc_monto)} COP</div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info(f"💳 Procesar {fmt_cop(_nc_monto)} COP en el datáfono antes de confirmar.")

                    _nc_referencia = st.text_input(
                        "Referencia / número de comprobante de pago",
                        key="nc_ref",
                        placeholder="Ingresa el número de transacción o comprobante"
                    )

                    st.divider()

                    # Paso 4: Autorización tratamiento de datos personales
                    st.markdown("#### 🔒 Paso 4 — Autorización de datos personales")
                    _nc_acepto_datos = st.checkbox(
                        "✅ Autorizo el tratamiento de mis datos personales conforme a la "
                        "Ley 1581 de 2012 y la política de privacidad de SUITE SALITRE.",
                        key="nc_acepto_datos",
                    )
                    st.markdown(
                        "<div style='font-size:12px;color:#94a3b8;margin-top:-8px;margin-bottom:8px'>"
                        "La información suministrada será usada exclusivamente para la gestión de su reserva y facturación."
                        "</div>",
                        unsafe_allow_html=True,
                    )

                    st.divider()

                    # Paso 5: Confirmar
                    _col_nc_btn1, _col_nc_btn2 = st.columns(2)
                    with _col_nc_btn1:
                        if st.button("✅ CONFIRMAR PAGO Y CREAR RESERVA", type="primary",
                                     use_container_width=True, key="nc_confirmar"):
                            if not _nc_nombre.strip():
                                st.error("⚠️ El nombre del cliente es obligatorio.")
                            elif not _nc_acepto_datos:
                                st.error("⚠️ Debes aceptar la autorización de tratamiento de datos personales para continuar.")
                            else:
                                if _nc_metodo in METODOS_CON_QR_NC and not _nc_referencia.strip():
                                    st.toast("⚠️ Sin referencia de pago — se registrará sin comprobante.",
                                             icon="⚠️")
                                _nc_cliente = {
                                    "nombre":           _nc_nombre.strip(),
                                    "tipo_doc":         _nc_tipo_doc,
                                    "numero_documento": _nc_doc.strip(),
                                    "telefono":         _nc_tel.strip(),
                                    "email":            _nc_email.strip(),
                                    "razon_social":     _nc_razon.strip(),
                                    "nit_empresa":      _nc_nit.strip(),
                                }
                                try:
                                    _nc_voucher = crear_reserva_completa(
                                        _cub_nc, _nc_cliente, _calc_nc, _nc_metodo,
                                        acepto_datos=_nc_acepto_datos
                                    )
                                    # Actualizar referencia de pago si fue ingresada
                                    if _nc_referencia.strip():
                                        try:
                                            _pg_update_col("Reservas","Numero_Reserva",
                                                           _nc_voucher["numero_reserva"],
                                                           "Referencia_Pago",_nc_referencia.strip())
                                        except Exception:
                                            pass
                                    st.session_state["_nc_last_voucher"] = _nc_voucher
                                    st.session_state["_nc_voucher_ref"]  = _nc_referencia.strip() or _nc_ref_pago
                                    st.success(
                                        f"✅ Reserva **{_nc_voucher['numero_reserva']}** creada — "
                                        f"Cubículo **{_nc_voucher['cubiculo']}** activado. "
                                        f"Horario: {_nc_ini_str} → {_nc_fin_str} · "
                                        f"Horas: {_nc_horas:.1f} h"
                                    )
                                    st.rerun()
                                except Exception as _e_nc:
                                    st.error(f"❌ Error al crear reserva: {_e_nc}")

                    with _col_nc_btn2:
                        if st.button("🔄 Limpiar formulario", use_container_width=True,
                                     key="nc_limpiar"):
                            for _k_nc in ["nc_nombre", "nc_tel", "nc_doc", "nc_email",
                                          "_nc_last_voucher", "_nc_voucher_ref"]:
                                st.session_state.pop(_k_nc, None)
                            st.rerun()

                    # ── Voucher de última reserva Noche Completa ───────────────────────────
                    _last_nc = st.session_state.get("_nc_last_voucher")
                    if _last_nc:
                        st.divider()
                        st.markdown("#### 🎫 Voucher — Última reserva Noche Completa creada")
                        _col_ncv1, _col_ncv2 = st.columns([2, 1])
                        with _col_ncv1:
                            _ref_nc_op = st.session_state.get("_nc_voucher_ref", "")
                            st.markdown(f"""
                            <div class="voucher-box" style="max-width:100%">
                              <div style="text-align:center;margin-bottom:12px">
                                <div style="font-size:18px;font-weight:800;color:#a29bfe">{NEGOCIO}</div>
                                <div style="font-size:12px;color:#94a3b8">🌙 Noche Completa · Reserva por operador</div>
                              </div>
                              <table style="width:100%;font-size:14px;border-collapse:collapse">
                                <tr><td style="color:#94a3b8;padding:3px 0">Reserva</td>
                                    <td style="text-align:right;font-family:'Inconsolata';font-weight:700">{_last_nc["numero_reserva"]}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Factura</td>
                                    <td style="text-align:right;font-family:'Inconsolata'">{_last_nc["numero_factura"]}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Cliente</td>
                                    <td style="text-align:right;font-weight:700">{_last_nc["cliente_nombre"]}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Cubículo</td>
                                    <td style="text-align:right;font-size:18px;color:#a29bfe;font-weight:700">{_last_nc["cubiculo"]}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Horario</td>
                                    <td style="text-align:right">{_nc_ini_str} → {_nc_fin_str}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Horas Reservadas</td>
                                    <td style="text-align:right;color:#a29bfe;font-weight:700">{_nc_horas:.1f} h</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Precio Aplicado</td>
                                    <td style="text-align:right;color:#ffd32a">{fmt_cop(_nc_precio_final)} COP</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">WiFi</td>
                                    <td style="text-align:right;font-family:'Inconsolata'">{_last_nc["wifi_ssid"]} / {_last_nc["wifi_password"]}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Método</td>
                                    <td style="text-align:right">{_last_nc["metodo_pago"]}{(" · " + _ref_nc_op) if _ref_nc_op else ""}</td></tr>
                                <tr><td style="color:#94a3b8;padding:3px 0">Total</td>
                                    <td style="text-align:right;font-size:18px;font-weight:700;color:#00ff88">{fmt_cop(_last_nc["total"])} COP</td></tr>
                              </table>
                              <div style="text-align:center;margin-top:16px">
                                <div style="font-size:12px;color:#94a3b8;letter-spacing:2px;text-transform:uppercase">CÓDIGO DE ACCESO</div>
                                <div style="font-family:'Inconsolata';font-size:56px;font-weight:700;
                                            color:#00ff88;letter-spacing:12px;
                                            text-shadow:0 0 30px rgba(0,255,136,0.6)">{_last_nc["codigo_acceso"]}</div>
                              </div>
                            </div>
                            """, unsafe_allow_html=True)

                        with _col_ncv2:
                            if REPORTLAB_AVAILABLE:
                                _pdf_nc = generar_ticket_pdf(_last_nc)
                                if _pdf_nc:
                                    st.download_button("📄 Imprimir PDF",
                                                       data=_pdf_nc,
                                                       file_name=f"ticket_{_last_nc['numero_reserva']}.pdf",
                                                       mime="application/pdf",
                                                       use_container_width=True)
                            _html_nc = voucher_html(_last_nc)
                            st.download_button("🌐 Voucher HTML",
                                               data=_html_nc.encode(),
                                               file_name=f"voucher_{_last_nc['numero_reserva']}.html",
                                               mime="text/html",
                                               use_container_width=True)
                            _wa_nc = (f"JJGT Reserva {_last_nc['numero_reserva']} | "
                                      f"🌙 Noche Completa {_nc_ini_str}→{_nc_fin_str} | "
                                      f"Horas: {_nc_horas:.1f} h | "
                                      f"Precio: {fmt_cop(_nc_precio_final)} COP | "
                                      f"Cubículo {_last_nc['cubiculo']} | "
                                      f"Código: {_last_nc['codigo_acceso']} | "
                                      f"WiFi: {_last_nc['wifi_ssid']} Clave: {_last_nc['wifi_password']} | "
                                      f"Total: {fmt_cop(_last_nc['total'])} COP")
                            _wa_url_nc = f"https://wa.me/?text={_wa_nc.replace(' ','%20')}"
                            st.markdown(
                                f'''<a href="{_wa_url_nc}" target="_blank">
                                <button style="width:100%;padding:12px;margin-top:8px;
                                  background:rgba(37,211,102,0.2);border:2px solid rgba(37,211,102,0.5);
                                  border-radius:12px;color:#25d366;font-weight:700;font-size:15px;
                                  cursor:pointer">📱 Enviar WhatsApp</button></a>''',
                                unsafe_allow_html=True)
                            st.markdown("<br>", unsafe_allow_html=True)
                            if st.button("🆕 Nueva reserva", use_container_width=True,
                                         key="nc_nueva", type="primary"):
                                for _k2_nc in ["_nc_last_voucher", "_nc_voucher_ref"]:
                                    st.session_state.pop(_k2_nc, None)
                                st.rerun()

    # ── TAB FORMULARIO RÁPIDO ─────────────────────────────────────────────────
    with tab_rapido:
        st.markdown("**Formulario rápido de asignación directa:**")

        # Paso 1: Cubículo y tiempo
        st.markdown("#### 🛏️ Paso 1 — Cubículo y tiempo")
        cubiculos_libres = get_cubiculos_libres()
        if not cubiculos_libres:
            st.error("❌ No hay cubículos disponibles.")
            return

        col_cub, col_hrs = st.columns(2)
        with col_cub:
            opciones_cub = {f"{c['numero']} — {c['nombre']}": c for c in cubiculos_libres}
            sel_label    = st.selectbox("Cubículo libre", list(opciones_cub.keys()), key="seleccion_kiosk_cub",
                                        index=0)
            cubiculo_sel = opciones_cub[sel_label]

        with col_hrs:
            horas_sel = st.number_input("Horas a reservar",
                                        min_value=0.5, max_value=18.0,
                                        value=1.0, step=0.3, key="op_res_horas")
            calc = calcular_precio(horas_sel)
            st.markdown(f"""
            <div style="background:rgba(0,212,255,0.08);border:1px solid rgba(0,212,255,0.2);
                        border-radius:8px;padding:10px;margin-top:8px;font-size:14px">
              Subtotal: <b>{fmt_cop(calc['subtotal'])}</b> ·
              IVA 19%: <b>{fmt_cop(calc['iva'])}</b><br>
              <b style="color:#00ff88;font-size:16px">Total: {fmt_cop(calc['total'])} COP</b>
            </div>
            """, unsafe_allow_html=True)

        st.divider()

        # Paso 2: Datos del cliente
        st.markdown("#### 👤 Paso 2 — Datos del cliente")
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            c_nombre   = st.text_input("Nombre completo *", key="op_res_nombre",
                                       placeholder="Nombre del pasajero")
            c_tel      = st.text_input("Teléfono", key="op_res_tel",
                                       placeholder="300 000 0000")
            c_tipo_doc = st.selectbox("Tipo documento",
                                      ["CC","CE","Pasaporte","NIT","TI"],
                                      key="op_res_tipodoc")
        with col_c2:
            c_doc   = st.text_input("Número de documento", key="op_res_doc")
            c_email = st.text_input("Email (opcional)", key="op_res_email")
            req_fact = st.checkbox("¿Requiere factura empresarial?",
                                   key="op_res_req_fact")

        razon_social = ""
        nit_empresa  = ""
        if req_fact:
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                razon_social = st.text_input("Razón social", key="op_res_razon")
            with col_f2:
                nit_empresa  = st.text_input("NIT empresa",  key="op_res_nit")

        st.divider()

        # Paso 3: Método de pago con QR
        st.markdown("#### 💳 Paso 3 — Método de pago")
        metodo_pago = st.selectbox("Método de pago",
                                   [m for m in METODOS_PAGO if m != "Convenio"],
                                   key="op_res_metodo")

        monto_total = int(calc["total"])
        # Referencia única para este intento de pago
        ref_pago = f"OP-{ahora_col().strftime('%Y%m%d%H%M')}-{cubiculo_sel['numero'].replace('#','')}"

        # Mostrar QR según método
        METODOS_CON_QR = ["Nequi", "Daviplata", "PSE", "MercadoPago", "Transferencia"]
        if metodo_pago in METODOS_CON_QR:
            st.markdown(f"**Monto a cobrar:** `{fmt_cop(monto_total)} COP`")
            mostrar_qr(metodo_pago, monto_total, ref_pago, size=220)
            st.caption(f"Referencia: `{ref_pago}`")
        elif metodo_pago == "Efectivo":
            st.markdown(f"""
            <div style="background:rgba(132,204,22,0.1);border:1px solid rgba(132,204,22,0.3);
                        border-radius:10px;padding:16px;text-align:center">
              <div style="font-size:13px;color:#94a3b8;margin-bottom:4px">TOTAL EN EFECTIVO</div>
              <div style="font-size:36px;font-weight:700;color:#84cc16">{fmt_cop(monto_total)} COP</div>
            </div>
            """, unsafe_allow_html=True)
        else:  # Tarjeta
            st.info(f"💳 Procesar {fmt_cop(monto_total)} COP en el datáfono antes de confirmar.")

        referencia = st.text_input("Referencia / número de comprobante de pago",
                                   key="op_res_ref",
                                   placeholder="Ingresa el número de transacción o comprobante")

        st.divider()

        # Paso 4: Autorización tratamiento de datos personales
        st.markdown("#### 🔒 Paso 4 — Autorización de datos personales")
        acepto_datos_rapido = st.checkbox(
            "✅ Autorizo el tratamiento de mis datos personales conforme a la "
            "Ley 1581 de 2012 y la política de privacidad de SUITE SALITRE.",
            key="op_res_acepto_datos",
        )
        st.markdown(
            "<div style='font-size:12px;color:#94a3b8;margin-top:-8px;margin-bottom:8px'>"
            "La información suministrada será usada exclusivamente para la gestión de su reserva y facturación."
            "</div>",
            unsafe_allow_html=True,
        )

        st.divider()

        # Paso 5: Confirmar
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("✅ CONFIRMAR PAGO Y CREAR RESERVA", type="primary",
                         use_container_width=True, key="op_res_confirmar"):
                if not c_nombre.strip():
                    st.error("⚠️ El nombre del cliente es obligatorio.")
                elif not acepto_datos_rapido:
                    st.error("⚠️ Debes aceptar la autorización de tratamiento de datos personales para continuar.")
                else:
                    # Advertir si falta referencia para métodos digitales, pero no bloquear
                    if metodo_pago in METODOS_CON_QR and not referencia.strip():
                        st.toast("⚠️ Sin referencia de pago — se registrará sin comprobante.",
                                 icon="⚠️")

                    cliente_data = {
                        "nombre":           c_nombre.strip(),
                        "tipo_doc":         c_tipo_doc,
                        "numero_documento": c_doc.strip(),
                        "telefono":         c_tel.strip(),
                        "email":            c_email.strip(),
                        "razon_social":     razon_social.strip(),
                        "nit_empresa":      nit_empresa.strip(),
                    }
                    with st.spinner("Creando reserva y sincronizando con Google Sheets..."):
                        try:
                            voucher = crear_reserva_completa(
                                cubiculo_sel, cliente_data, calc, metodo_pago,
                                acepto_datos=acepto_datos_rapido)
                            # Guardar referencia de pago en PostgreSQL si se proporcionó
                            if referencia.strip() and voucher:
                                try:
                                    _pg_update_col("Reservas","Numero_Reserva",
                                                   voucher["numero_reserva"],
                                                   "Referencia_Pago", referencia.strip())
                                except Exception:
                                    pass
                            st.session_state["_op_last_voucher"] = voucher
                            st.session_state["_op_voucher_ref"]  = referencia.strip() or ref_pago
                            st.success(
                                f"✅ Reserva **{voucher['numero_reserva']}** creada — "
                                f"Cubículo **{voucher['cubiculo']}** activado.")
                            st.rerun()

                            #try:
                            #    import contabilidad
                            #    contabilidad.on_pago_convenio(
                            #        empresa    = nombre_empresa,
                            #        valor      = monto_total,
                            #        num_reserva= num_reserva,
                            #        num_factura= num_factura,
                            #    )
                            #except Exception:
                            #    pass

                        except Exception as e:
                            st.error(f"❌ Error al crear reserva: {e}")
        with col_btn2:
            if st.button("🔄 Limpiar formulario", use_container_width=True,
                         key="op_res_limpiar"):
                for k in ["op_res_nombre","op_res_tel","op_res_doc","op_res_email",
                          "_op_last_voucher","_op_voucher_ref"]:
                    st.session_state.pop(k, None)
                st.rerun()

        # ── Voucher de última reserva ──────────────────────────────────────────
        last_v = st.session_state.get("_op_last_voucher")
        if last_v:
            st.divider()
            st.markdown("#### 🎫 Voucher — Última reserva creada")
            col_v1, col_v2 = st.columns([2, 1])
            with col_v1:
                ref_op = st.session_state.get("_op_voucher_ref","")
                st.markdown(f"""
                <div class="voucher-box" style="max-width:100%">
                  <div style="text-align:center;margin-bottom:12px">
                    <div style="font-size:18px;font-weight:800;color:#00d4ff">{NEGOCIO}</div>
                    <div style="font-size:12px;color:#94a3b8">Reserva creada por operador</div>
                  </div>
                  <table style="width:100%;font-size:14px;border-collapse:collapse">
                    <tr><td style="color:#94a3b8;padding:3px 0">Reserva</td>
                        <td style="text-align:right;font-family:'Inconsolata';font-weight:700">{last_v["numero_reserva"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Factura</td>
                        <td style="text-align:right;font-family:'Inconsolata'">{last_v["numero_factura"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Cliente</td>
                        <td style="text-align:right;font-weight:700">{last_v["cliente_nombre"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Cubículo</td>
                        <td style="text-align:right;font-size:18px;color:#00d4ff;font-weight:700">{last_v["cubiculo"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Horario</td>
                        <td style="text-align:right">{last_v["hora_inicio"]} → {last_v["hora_fin"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">WiFi</td>
                        <td style="text-align:right;font-family:'Inconsolata'">{last_v["wifi_ssid"]} / {last_v["wifi_password"]}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Método</td>
                        <td style="text-align:right">{last_v["metodo_pago"]}{(" · " + ref_op) if ref_op else ""}</td></tr>
                    <tr><td style="color:#94a3b8;padding:3px 0">Total</td>
                        <td style="text-align:right;font-size:18px;font-weight:700;color:#00ff88">{fmt_cop(last_v["total"])} COP</td></tr>
                  </table>
                  <div style="text-align:center;margin-top:16px">
                    <div style="font-size:12px;color:#94a3b8;letter-spacing:2px;text-transform:uppercase">CÓDIGO DE ACCESO</div>
                    <div style="font-family:'Inconsolata';font-size:56px;font-weight:700;
                                color:#00ff88;letter-spacing:12px;
                                text-shadow:0 0 30px rgba(0,255,136,0.6)">{last_v["codigo_acceso"]}</div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

            with col_v2:
                if REPORTLAB_AVAILABLE:
                    pdf_b = generar_ticket_pdf(last_v)
                    if pdf_b:
                        st.download_button("📄 Imprimir PDF",
                                           data=pdf_b,
                                           file_name=f"ticket_{last_v['numero_reserva']}.pdf",
                                           mime="application/pdf",
                                           use_container_width=True)
                html_v_str = voucher_html(last_v)
                st.download_button("🌐 Voucher HTML",
                                   data=html_v_str.encode(),
                                   file_name=f"voucher_{last_v['numero_reserva']}.html",
                                   mime="text/html",
                                   use_container_width=True)
                wa_msg = (f"JJGT Reserva {last_v['numero_reserva']} | "
                          f"Cubículo {last_v['cubiculo']} | "
                          f"Código: {last_v['codigo_acceso']} | "
                          f"WiFi: {last_v['wifi_ssid']} Clave: {last_v['wifi_password']} | "
                          f"Total: {fmt_cop(last_v['total'])} COP")
                wa_url = f"https://wa.me/?text={wa_msg.replace(' ','%20')}"
                st.markdown(
                    f'''<a href="{wa_url}" target="_blank">
                    <button style="width:100%;padding:12px;margin-top:8px;
                      background:rgba(37,211,102,0.2);border:2px solid rgba(37,211,102,0.5);
                      border-radius:12px;color:#25d366;font-weight:700;font-size:15px;
                      cursor:pointer">📱 Enviar WhatsApp</button></a>''',
                    unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🆕 Nueva reserva", use_container_width=True,
                             key="op_res_nueva", type="primary"):
                    for k in ["_op_last_voucher","_op_voucher_ref"]:
                        st.session_state.pop(k, None)
                    st.rerun()

    # ── TAB CONVENIO ──────────────────────────────────────────────────────────
    #with tab_conv:
    #    _tab_convenio_reserva()


def _op_dashboard():
    st.markdown("### 🏠 Dashboard Operacional")
    inject_live_clock()
    # El rerun automático del dashboard se maneja en show_operador() cada 30s
    cubiculos = get_cubiculos()
    libres    = sum(1 for c in cubiculos if c["estado"] == "libre")
    ocupados  = sum(1 for c in cubiculos if c["estado"] == "ocupado")
    manten    = sum(1 for c in cubiculos if c["estado"] == "mantenimiento")

    # ── Auto-liberar cubículos con tiempo vencido + alarma ────────────────────
    vencidos = []
    for cub in cubiculos:
        if cub["estado"] in ("ocupado", "por_liberar"):
            mins = cub.get("minutos_restantes")
            if mins is not None and mins <= 0:
                liberar_cubiculo(cub["id"])
                vencidos.append(cub["numero"])

    if vencidos:
        # Sonido de alarma (beeps repetidos via Web Audio API)
        st.markdown(f"""
        <div class="alerta-roja">
          🔔 ¡TIEMPO VENCIDO! Cubículos liberados automáticamente: <b>{', '.join(vencidos)}</b>
        </div>
        <script>
        (function(){{
          try {{
            var ctx = new (window.AudioContext || window.webkitAudioContext)();
            function beep(freq, start, dur) {{
              var o = ctx.createOscillator();
              var g = ctx.createGain();
              o.connect(g); g.connect(ctx.destination);
              o.frequency.value = freq;
              o.type = 'square';
              g.gain.setValueAtTime(0.3, ctx.currentTime + start);
              g.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime + start + dur);
              o.start(ctx.currentTime + start);
              o.stop(ctx.currentTime + start + dur + 0.05);
            }}
            for(var i=0; i<5; i++) {{ beep(880, i*0.4, 0.3); beep(660, i*0.4+0.15, 0.15); }}
          }} catch(e) {{}}
        }})();
        </script>
        """, unsafe_allow_html=True)

    # ── KPIs ──────────────────────────────────────────────────────────────────
    today = ahora_col().strftime("%Y-%m-%d")
    reservas_gs = _gs_read_sheet("Reservas")
    pagos_gs    = _gs_read_sheet("Pagos")

    ingresos_hoy = sum(
        _gs_float(r, "Total_COP") for r in reservas_gs
        if _gs_val(r, "Estado_Pago") == "confirmado" and
           _gs_fecha_ymd(r, "Creado_En") == today
    )
    reservas_hoy = sum(
        1 for r in reservas_gs
        if _gs_val(r, "Estado_Pago") == "confirmado" and
           _gs_fecha_ymd(r, "Creado_En") == today
    )
    pend = sum(1 for r in pagos_gs if _gs_val(r, "Estado") == "pendiente")

    c1, c2, c3, c4, c5 = st.columns(5)
    # Convenios pendientes del día
    conv_hoy  = sum(1 for r in reservas_gs
                    if _gs_val(r, "Metodo_Pago") == "Convenio"
                    and _gs_fecha_ymd(r, "Creado_En") == today)
    conv_monto = sum(_gs_float(r, "Total_COP") for r in reservas_gs
                     if _gs_val(r, "Metodo_Pago") == "Convenio"
                     and _gs_fecha_ymd(r, "Creado_En") == today)

    c1.metric("🟢 Libres",    libres)
    c2.metric("🔴 Ocupados",  ocupados)
    c3.metric("🔧 Mantenim.", manten)
    c4.metric("💰 Ingresos hoy", fmt_cop(ingresos_hoy))
    c5.metric("📋 Reservas hoy", reservas_hoy)

    if conv_hoy > 0:
        st.info(f"🤝 **{conv_hoy} reserva(s) de convenio hoy** · "
                f"Monto: {fmt_cop(conv_monto)} COP (pendiente de cobro)")
    if pend > 0:
        st.warning(f"⚠️ {pend} pago(s) pendientes de confirmación")

    st.markdown("---")
    st.markdown("#### Estado en tiempo real")
    inject_live_clock()
    # Recargar cubículos frescos post-liberaciones
    cubiculos = get_cubiculos()
    cols = st.columns(4)
    for i, cub in enumerate(cubiculos):
        with cols[i % 4]:
            estado    = cub["estado"]
            info      = ESTADOS_CUBICULO.get(estado, ESTADOS_CUBICULO["libre"])
            mins      = cub.get("minutos_restantes")
            hora_fin  = cub.get("hora_fin") or ""
            alerta_prefix = "⚠️ " if (mins is not None and 0 < mins <= 5) else ""
            # Timer HTML con data-fin para actualización JS en tiempo real
            if hora_fin and estado not in ("libre", "mantenimiento"):
                timer_html = (
                    f'<div class="jjgt-timer" data-fin="{hora_fin}" ' +
                    f'data-key="dash-{cub["id"]}" ' +
                    f'style="font-size:13px;color:{"#ff4757" if (mins or 999)<=5 else "#ffd32a" if (mins or 999)<=15 else "#00ff88"};">' +
                    f'⏱ {fmt_tiempo(mins)}</div>'
                )
            else:
                timer_html = ""
            st.markdown(f"""
            <div style="background:{info['bg']};border:1.5px solid {info['color']}55;
                        border-radius:10px;padding:12px;text-align:center;margin-bottom:8px">
              <div style="font-family:'Inconsolata';font-size:28px;
                          font-weight:700;color:{info['color']}">{alerta_prefix}{cub['numero']}</div>
              <div style="font-size:10px;font-weight:700;color:{info['color']};
                          letter-spacing:1px">{info['label']}</div>
              {timer_html}
            </div>
            """, unsafe_allow_html=True)


def _op_cubiculos():
    st.markdown("### 🛏️ Gestión de Cubículos")
    cubiculos = get_cubiculos()

    # Auto-liberar vencidos con alarma
    for cub in cubiculos:
        if cub["estado"] in ("ocupado", "por_liberar"):
            mins = cub.get("minutos_restantes")
            if mins is not None and mins <= 0:
                liberar_cubiculo(cub["id"])
                st.markdown(f"""
                <div class="alerta-roja">🔔 Cubículo {cub["numero"]} liberado automáticamente (tiempo vencido)</div>
                <script>
                (function(){{try{{var ctx=new(window.AudioContext||window.webkitAudioContext)();
                function b(f,s,d){{var o=ctx.createOscillator(),g=ctx.createGain();
                o.connect(g);g.connect(ctx.destination);o.frequency.value=f;o.type="square";
                g.gain.setValueAtTime(0.3,ctx.currentTime+s);
                g.gain.exponentialRampToValueAtTime(0.001,ctx.currentTime+s+d);
                o.start(ctx.currentTime+s);o.stop(ctx.currentTime+s+d+0.05);}}
                for(var i=0;i<5;i++){{b(880,i*0.4,0.3);b(660,i*0.4+0.15,0.15);}}
                }}catch(e){{}}}})()</script>
                """, unsafe_allow_html=True)
    cubiculos = get_cubiculos()  # refrescar después de liberaciones

    # Precargar reservas de convenio activas (una sola llamada a Sheets)
    reservas_todas = _gs_read_sheet("Reservas")
    conv_activas = {
        _gs_val(r, "Cubiculo_Num"): _gs_val(r, "Nombre_Empresa", "Convenio")
        for r in reservas_todas
        if _gs_val(r, "Metodo_Pago") == "Convenio"
        and _gs_val(r, "Estado_Pago") == "pendiente"
        and not _gs_val(r, "Hora_Fin_Real")
    }

    for cub in cubiculos:
        estado = cub["estado"]
        info   = ESTADOS_CUBICULO.get(estado, ESTADOS_CUBICULO["libre"])
        mins   = cub.get("minutos_restantes")
        conv_badge = ""
        if cub["numero"] in conv_activas:
            conv_badge = f" 🤝 {conv_activas[cub['numero']]}"

        label_exp = f"{cub['numero']} — {info['label']} | WiFi: {cub['wifi_ssid']}{conv_badge}"
        if mins is not None and mins <= 10 and estado != "libre":
            label_exp += f" ⚠️ {fmt_tiempo(mins)}"
        with st.expander(label_exp):
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                if estado in ("ocupado", "por_liberar") and st.button(f"🔓 Liberar {cub['numero']}",
                                                     key=f"lib_{cub['id']}"):
                    liberar_cubiculo(cub["id"])
                    st.success(f"Cubículo {cub['numero']} liberado")
                    st.rerun()
            with c2:
                if estado != "mantenimiento":
                    if st.button(f"🔧 Mantenimiento {cub['numero']}",
                                 key=f"mant_{cub['id']}"):
                        try:
                            _pg_exec("UPDATE cubiculos_estado SET estado='mantenimiento' WHERE numero=%s", [cub["numero"]])
                        except Exception: pass
                        _gs_invalidate_cache("Cubiculos_Estado")
                        st.rerun()
                else:
                    if st.button(f"✅ Listo {cub['numero']}", key=f"ok_{cub['id']}"):
                        try:
                            _pg_exec("UPDATE cubiculos_estado SET estado='libre' WHERE numero=%s", [cub["numero"]])
                        except Exception: pass
                        _gs_invalidate_cache("Cubiculos_Estado")
                        st.rerun()
            with c3:
                new_wifi_pw = st.text_input("Nueva clave WiFi",
                                            value=cub["wifi_password"],
                                            key=f"wifi_{cub['id']}")
            with c4:
                if st.button(f"💾 Guardar WiFi", key=f"swifi_{cub['id']}"):
                    try:
                        _pg_exec("UPDATE cubiculos_estado SET wifi_pass=%s WHERE numero=%s", [new_wifi_pw, cub["numero"]])
                        _gs_invalidate_cache("Cubiculos_Estado")
                        st.success("WiFi actualizado")
                    except Exception as _ew: st.warning(f"Error: {_ew}")
                    st.rerun()
            if cub.get("minutos_restantes") is not None:
                st.info(f"⏱ Tiempo restante: {fmt_tiempo(cub['minutos_restantes'])}")


def _op_pagos_pendientes():
    st.markdown("### ⏳ Pagos Pendientes de Confirmación")

    # ── jjgt_pagos: pagos normales pendientes ────────────────────────────────
    pagos_gs    = _gs_read_sheet("Pagos")
    reservas_gs = _gs_read_sheet("Reservas")
    res_map = {_gs_val(r, "Numero_Reserva"): r for r in reservas_gs}
    pendientes_pagos = [r for r in pagos_gs if _gs_val(r, "Estado") == "pendiente"]

    # ── jjgt_convenios: pagos de convenio pendientes ─────────────────────────
    # ── Pagos de convenio pendientes desde PostgreSQL ────────────────────────
    pendientes_conv = []
    res_map_conv    = {}
    try:
        pagos_conv_all = _pg_read_table_conv("Pagos")
        for d in pagos_conv_all:
            if _gs_val(d, "Estado") == "pendiente":
                d["_fuente"] = "convenios"
                pendientes_conv.append(d)
        reservas_conv_all = _pg_read_table_conv("Reservas")
        for d2 in reservas_conv_all:
            num = _gs_val(d2, "Numero_Reserva")
            if num:
                res_map_conv[num] = d2
    except Exception:
        pass

    # Combinar: marcar origen
    for p in pendientes_pagos:
        p["_fuente"] = "pagos"

    pendientes = pendientes_pagos + pendientes_conv

    if not pendientes:
        st.success("✅ No hay pagos pendientes en este momento.")
        return

    n_conv = len(pendientes_conv)
    n_norm = len(pendientes_pagos)
    if n_conv > 0:
        st.info(f"📋 {n_norm} pago(s) normales + 🤝 {n_conv} pago(s) de convenio pendientes")

    for pag in pendientes:
        pago_id  = _gs_val(pag, "ID_Pago")
        num_res  = _gs_val(pag, "Num_Reserva") or _gs_val(pag, "ID_Reserva")
        monto    = _gs_float(pag, "Monto_COP")
        metodo   = _gs_val(pag, "Metodo")
        fecha    = _gs_val(pag, "Fecha_Pago")
        fuente   = pag.get("_fuente", "pagos")
        # Buscar datos de reserva en el mapa correcto según el origen del pago
        if fuente == "convenios":
            res_data = res_map_conv.get(num_res, {})
        else:
            res_data = res_map.get(num_res, {})
        cliente  = res_data.get("Cliente_Nombre", _gs_val(res_data, "Cliente_Nombre", "—"))
        horas    = res_data.get("Horas_Contratadas", _gs_val(res_data, "Horas_Contratadas", "—"))
        cub_num  = res_data.get("Cubiculo_Num", _gs_val(res_data, "Cubiculo_Num", "—"))

        with st.container():
            # Identificar si es convenio
            es_convenio  = (metodo == "Convenio")
            empresa_pend = _gs_val(res_data, "Nombre_Empresa", "") if es_convenio else ""
            color_borde  = "rgba(162,155,254,0.4)" if es_convenio else "rgba(255,211,42,0.4)"
            color_num    = "#a29bfe" if es_convenio else "#ffd32a"
            badge        = f" 🤝 <b style='color:#a29bfe'>CONVENIO: {empresa_pend}</b>" if es_convenio else ""

            st.markdown(f"""
            <div style="background:rgba(255,211,42,0.08);border:1px solid {color_borde};
                        border-radius:12px;padding:16px;margin-bottom:12px">
              <div style="display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px">
                <div>
                  <b style="color:{color_num}">{num_res}</b> · {cliente}{badge}<br>
                  <span style="color:#94a3b8">{metodo} · {fmt_cop(monto)} COP · {horas}h · Cubículo {cub_num}</span>
                  {'<br><span style="color:#a29bfe;font-size:12px">Registrado en jjgt_convenios</span>' if es_convenio else ''}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                pin_conf = st.text_input("PIN para confirmar", type="password",
                                         key=f"pin_conf_{pago_id}", max_chars=8)
                if st.button(f"✅ CONFIRMAR PAGO", key=f"conf_{pago_id}",
                             type="primary", use_container_width=True):
                    if verificar_pin(pin_conf):
                        _, sh_conf      = get_active_client()
                        _, sh_conv_conf = get_active_client_convenios()

                        def _confirmar_en_sh(sh_target, id_pago_val, num_res_val):
                            """Confirma pago y reserva en PostgreSQL (jjgt_pagos o jjgt_convenios)."""
                            # sh_target es _PGStub — se ignora; la fuente determina la tabla
                            pass  # La confirmación se maneja por _fuente abajo

                        if fuente == "convenios":
                            # Pago de convenio: actualizar en tablas PG de convenio
                            try:
                                _pg_update_col_conv("Pagos",    "ID_Pago",         pago_id, "Estado",      "confirmado")
                                _pg_update_col_conv("Reservas", "Numero_Reserva",  num_res, "Estado_Pago", "confirmado")
                                _pg_update_col_conv("Facturas", "Num_Reserva",     num_res, "Estado",      "confirmado")
                            except Exception as _e_conv:
                                st.warning(f"⚠️ Error confirmando convenio: {_e_conv}")
                            try:
                                _, _sh_conv_dash = get_active_client_convenios()
                                _gs_sync_dashboard_convenios(_sh_conv_dash)
                            except Exception: pass
                            if CONT_AVAILABLE and _cont_mod is not None:
                                try:
                                    _cont_mod.on_pago_convenio(
                                        empresa    = nombre_empresa,
                                        valor      = monto_total,
                                        num_reserva= num_reserva,
                                        num_factura= num_factura,
                                    )
                                except Exception:
                                    pass
                            # Pago normal: actualizar en jjgt_pagos
                            _confirmar_en_sh(sh_conf, pago_id, num_res)

                            if CONT_AVAILABLE and _cont_mod is not None:
                                try:
                                    _cont_mod.on_pago_convenio(
                                        empresa    = nombre_empresa,
                                        valor      = monto_total,
                                        num_reserva= num_reserva,
                                        num_factura= num_factura,
                                    )
                                except Exception:
                                    pass

                        if sh_conf:
                            gs_sync_cubiculos(sh_conf)
                            gs_sync_dashboard(sh_conf)
                            _gs_invalidate_cache("Pagos", "Reservas", "Cubiculos_Estado",
                                                 "Facturas", "Dashboard_Diario")
                            # Si era un convenio en jjgt_convenios (fuente anterior),
                            # ya se confirmó arriba directamente
                            if metodo == "Convenio" and fuente != "convenios":
                                # Convenio confirmado desde panel normal → actualizar en PG convenios
                                try:
                                    _pg_update_col_conv("Pagos",    "Num_Reserva",    num_res, "Estado",      "confirmado")
                                    _pg_update_col_conv("Reservas", "Numero_Reserva", num_res, "Estado_Pago", "confirmado")
                                    _pg_update_col_conv("Facturas", "Num_Reserva",    num_res, "Estado",      "confirmado")
                                    _, _sh_dc2 = get_active_client_convenios()
                                    _gs_sync_dashboard_convenios(_sh_dc2)
                                except Exception:
                                    pass
                        st.success(f"✅ Pago {num_res} confirmado")
                        st.rerun()

                    else:
                        st.error("PIN incorrecto")
            with c2:
                if st.button(f"❌ Rechazar", key=f"rej_{pago_id}", use_container_width=True):
                    try:
                        if fuente == "convenios":
                            _pg_update_col_conv("Pagos",    "ID_Pago",        pago_id, "Estado",      "rechazado")
                            _pg_update_col_conv("Reservas", "Numero_Reserva", num_res, "Estado_Pago", "rechazado")
                        else:
                            _pg_update_col("Pagos",    "ID_Pago",        pago_id, "Estado",      "rechazado")
                            _pg_update_col("Reservas", "Numero_Reserva", num_res, "Estado_Pago", "rechazado")
                    except Exception:
                        pass
                    if cub_num:
                        cubiculos_now = get_cubiculos()
                        cub_to_free = next((c for c in cubiculos_now if c["numero"] == cub_num), None)
                        if cub_to_free:
                            liberar_cubiculo(cub_to_free["id"])
                    _gs_invalidate_cache("Pagos", "Reservas", "Cubiculos_Estado")
                    st.warning("Pago rechazado y cubículo liberado")
                    st.rerun()


def generar_backup_diario():
    """
    Genera backup completo de TODAS las tablas PostgreSQL:
    - jjgt_pagos    (DRIVE_SHEETS)          via _pg_read_table
    - jjgt_convenios (DRIVE_SHEETS_CONVENIOS) via _pg_read_table_conv
    Con openpyxl: un .xlsx con pestanas agrupadas por base de datos + hoja RESUMEN.
    Sin openpyxl: un .zip con carpetas/prefijos por base de datos + RESUMEN.txt.
    Retorna (bytes, filename, mimetype).
    """
    HOJAS_PAGOS = [
        "Reservas", "Pagos", "Clientes", "Facturas", "Factura_items",
        "Cubiculos_Estado", "Tarifas_Config", "Operadores",
        "Configuracion_Pagos", "Dashboard_Diario", "Log_Operaciones",
    ]
    HOJAS_CONV = [
        "Empresa", "Reservas", "Pagos", "Clientes", "Facturas",
        "Factura_items", "Dashboard_Diario", "Cubiculos_Estado",
        "Tarifas_Config", "Operadores", "Configuracion_Pagos",
        "Log_Operaciones",
    ]

    ahora     = ahora_col()
    now_str   = ahora.strftime("%Y-%m-%d_%H%M")
    now_human = ahora.strftime("%d/%m/%Y %H:%M")

    # ── Leer todas las tablas de ambas bases de datos ─────────────────────────
    res_pagos: dict = {}
    res_conv:  dict = {}
    errores:   list = []

    for hoja in HOJAS_PAGOS:
        try:
            filas = _pg_read_table(hoja)
            cols  = DRIVE_SHEETS.get(hoja, [])
            res_pagos[hoja] = pd.DataFrame(filas) if filas else pd.DataFrame(columns=cols)
        except Exception as exc:
            errores.append(f"[pagos] {hoja}: {exc}")
            cols = DRIVE_SHEETS.get(hoja, [])
            res_pagos[hoja] = pd.DataFrame(columns=cols)

    for hoja in HOJAS_CONV:
        try:
            filas = _pg_read_table_conv(hoja)
            cols  = DRIVE_SHEETS_CONVENIOS.get(hoja, [])
            res_conv[hoja] = pd.DataFrame(filas) if filas else pd.DataFrame(columns=cols)
        except Exception as exc:
            errores.append(f"[convenios] {hoja}: {exc}")
            cols = DRIVE_SHEETS_CONVENIOS.get(hoja, [])
            res_conv[hoja] = pd.DataFrame(columns=cols)

    # ── XLSX ─────────────────────────────────────────────────────────────────
    if OPENPYXL_AVAILABLE:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        def _fmt_hdr(ws, df, fill_hex):
            """Aplica formato a la fila de encabezados y ajusta anchos."""
            fill = PatternFill("solid", fgColor=fill_hex)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")
            for col_idx, col in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if len(df) else 0,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:

            # -- Hoja RESUMEN ------------------------------------------------
            filas_res = []
            for hoja in HOJAS_PAGOS:
                df_h = res_pagos.get(hoja, pd.DataFrame())
                filas_res.append({
                    "Base_Datos": "jjgt_pagos",
                    "Tabla":      hoja,
                    "Registros":  len(df_h),
                    "Columnas":   len(df_h.columns),
                    "Estado":     "ERROR" if any(f"[pagos] {hoja}:" in e for e in errores) else "OK",
                })
            for hoja in HOJAS_CONV:
                df_h = res_conv.get(hoja, pd.DataFrame())
                filas_res.append({
                    "Base_Datos": "jjgt_convenios",
                    "Tabla":      hoja,
                    "Registros":  len(df_h),
                    "Columnas":   len(df_h.columns),
                    "Estado":     "ERROR" if any(f"[convenios] {hoja}:" in e for e in errores) else "OK",
                })
            df_res = pd.DataFrame(filas_res)
            total  = df_res["Registros"].sum()
            df_res = pd.concat([df_res, pd.DataFrame([{
                "Base_Datos": "TOTAL",
                "Tabla":      f"{len(HOJAS_PAGOS) + len(HOJAS_CONV)} tablas",
                "Registros":  total,
                "Columnas":   "",
                "Estado":     f"Backup: {now_human}",
            }])], ignore_index=True)

            df_res.to_excel(writer, sheet_name="RESUMEN", index=False)
            ws_r = writer.sheets["RESUMEN"]
            # Encabezado gris oscuro
            for cell in ws_r[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="2D3748")
                cell.alignment = Alignment(horizontal="center")
            # Fila de totales en negrita
            for cell in ws_r[ws_r.max_row]:
                cell.font = Font(bold=True)
            for col_idx, col in enumerate(df_res.columns, 1):
                max_len = max(len(str(col)),
                              df_res[col].astype(str).str.len().max())
                ws_r.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

            # -- Pestanas jjgt_pagos (encabezado azul) -----------------------
            for hoja in HOJAS_PAGOS:
                df_h = res_pagos[hoja]
                df_h.to_excel(writer, sheet_name=hoja[:31], index=False)
                _fmt_hdr(writer.sheets[hoja[:31]], df_h, "1E3A5F")

            # -- Pestanas jjgt_convenios (prefijo C_, encabezado verde) ------
            for hoja in HOJAS_CONV:
                sheet_name = f"C_{hoja}"[:31]
                df_h = res_conv[hoja]
                df_h.to_excel(writer, sheet_name=sheet_name, index=False)
                _fmt_hdr(writer.sheets[sheet_name], df_h, "1A4731")

        return (
            buf.getvalue(),
            f"backup_jjgt_{now_str}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ── ZIP fallback ─────────────────────────────────────────────────────────
    import zipfile

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        total_reg = 0

        # CSVs jjgt_pagos
        for hoja in HOJAS_PAGOS:
            df_h = res_pagos[hoja]
            total_reg += len(df_h)
            zf.writestr(f"pagos/{hoja}.csv", df_h.to_csv(index=False, encoding="utf-8"))

        # CSVs jjgt_convenios
        for hoja in HOJAS_CONV:
            df_h = res_conv[hoja]
            total_reg += len(df_h)
            zf.writestr(f"convenios/{hoja}.csv", df_h.to_csv(index=False, encoding="utf-8"))

        # RESUMEN.txt
        lineas = [
            f"BACKUP JJGT ESPACIOS — {now_human}",
            "=" * 56,
            "",
            "[ jjgt_pagos ]",
        ]
        for hoja in HOJAS_PAGOS:
            lineas.append(f"  {hoja:<30} {len(res_pagos[hoja]):>6} registros")
        lineas += ["", "[ jjgt_convenios ]"]
        for hoja in HOJAS_CONV:
            lineas.append(f"  {hoja:<30} {len(res_conv[hoja]):>6} registros")
        lineas += [
            "",
            "=" * 56,
            f"  TOTAL REGISTROS:                   {total_reg:>6}",
        ]
        if errores:
            lineas += ["", "ERRORES:"] + [f"  {e}" for e in errores]
        zf.writestr("RESUMEN.txt", "\n".join(lineas))

    return (
        buf.getvalue(),
        f"backup_jjgt_{now_str}.zip",
        "application/zip",
    )


def generar_reporte_pdf(rows: list, desde: str, hasta: str,
                        total_ing: float, total_res: int, completadas: int) -> bytes:
    """
    Genera un PDF de reporte de reservas para el período indicado.
    Retorna los bytes del PDF listos para descarga.
    """
    if not REPORTLAB_AVAILABLE:
        return b""

    buf = io.BytesIO()
    doc_pdf = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm,  bottomMargin=18*mm,
        title=f"Reporte Reservas {desde} – {hasta}",
    )

    styles = getSampleStyleSheet()
    # ── Estilos personalizados ────────────────────────────────────────────────
    titulo_style = ParagraphStyle(
        "TituloReporte",
        parent=styles["Normal"],
        fontSize=18, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0A1628"),
        spaceAfter=4, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "SubReporte",
        parent=styles["Normal"],
        fontSize=10, fontName="Helvetica",
        textColor=colors.HexColor("#64748B"),
        spaceAfter=2, alignment=TA_CENTER,
    )
    label_style = ParagraphStyle(
        "LabelKPI",
        parent=styles["Normal"],
        fontSize=8, fontName="Helvetica",
        textColor=colors.HexColor("#64748B"),
        alignment=TA_CENTER, spaceAfter=0,
    )
    valor_style = ParagraphStyle(
        "ValorKPI",
        parent=styles["Normal"],
        fontSize=16, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0E7490"),
        alignment=TA_CENTER, spaceAfter=0,
    )
    sec_style = ParagraphStyle(
        "SeccionReporte",
        parent=styles["Normal"],
        fontSize=11, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#0E7490"),
        spaceBefore=10, spaceAfter=4,
    )
    pie_style = ParagraphStyle(
        "PieReporte",
        parent=styles["Normal"],
        fontSize=7, fontName="Helvetica",
        textColor=colors.HexColor("#94A3B8"),
        alignment=TA_CENTER, spaceBefore=6,
    )

    def hr_line(color="#CBD5E1", thickness=0.5):
        return HRFlowable(width="100%", thickness=thickness,
                          color=colors.HexColor(color), spaceAfter=6, spaceBefore=2)

    ahora_str = ahora_col().strftime("%d/%m/%Y %H:%M")
    periodo_str = f"{desde} – {hasta}" if desde != hasta else desde

    story = []

    # ── Encabezado ────────────────────────────────────────────────────────────
    if _LOGO_B64:
        try:
            from reportlab.platypus import Image as RLImage
            _logo_buf_rep = io.BytesIO(base64.b64decode(_LOGO_B64))
            _logo_img_rep = RLImage(_logo_buf_rep, width=28*mm, height=38*mm)
            _logo_img_rep.hAlign = "LEFT"
            _header_logo_col = [_logo_img_rep]
        except Exception:
            _header_logo_col = [Paragraph("💤", titulo_style)]
    else:
        _header_logo_col = [Paragraph("💤", titulo_style)]

    _header_text_col = [
        Paragraph(NEGOCIO, titulo_style),
        Paragraph("REPORTE DE RESERVAS", sub_style),
        Paragraph(f"Período: {periodo_str}  ·  Generado: {ahora_str}", sub_style),
    ]
    _header_table = Table(
        [[_header_logo_col, _header_text_col]],
        colWidths=[32*mm, doc_pdf.width - 32*mm],
    )
    _header_table.setStyle([
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ])
    story.append(_header_table)
    story.append(hr_line("#0E7490", 1.5))
    story.append(Spacer(1, 4*mm))

    # ── KPIs ──────────────────────────────────────────────────────────────────
    col_w = (doc_pdf.width - 8*mm) / 3
    kpi_table = Table(
        [[
            [Paragraph("TOTAL RESERVAS", label_style), Paragraph(str(total_res), valor_style)],
            [Paragraph("COMPLETADAS", label_style),    Paragraph(str(completadas), valor_style)],
            [Paragraph("INGRESOS", label_style),       Paragraph(fmt_cop(total_ing), valor_style)],
        ]],
        colWidths=[col_w, col_w, col_w],
    )
    kpi_table.setStyle([
        ("BACKGROUND",  (0,0), (-1,-1), colors.HexColor("#E0F7FA")),
        ("BOX",         (0,0), (-1,-1), 0.5, colors.HexColor("#0E7490")),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, colors.HexColor("#BAE6FD")),
        ("TOPPADDING",  (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING",(0,0), (-1,-1), 6),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
    ])
    story.append(kpi_table)
    story.append(Spacer(1, 5*mm))

    # ── Resumen por método de pago ────────────────────────────────────────────
    if rows:
        metodos_totales = {}
        for r in rows:
            if r.get("Estado") == "confirmado":
                m = r.get("Método", "—")
                metodos_totales[m] = metodos_totales.get(m, 0) + r.get("Total", 0)

        if metodos_totales:
            story.append(Paragraph("Desglose por método de pago", sec_style))
            story.append(hr_line())
            met_data = [["Método de Pago", "Ingresos COP", "% del Total"]]
            for met, val in sorted(metodos_totales.items(), key=lambda x: -x[1]):
                pct = f"{val/total_ing*100:.1f}%" if total_ing > 0 else "—"
                met_data.append([met, fmt_cop(val), pct])
            met_data.append(["TOTAL", fmt_cop(total_ing), "100%"])

            met_table = Table(met_data, colWidths=[doc_pdf.width*0.45, doc_pdf.width*0.30, doc_pdf.width*0.25])
            met_table.setStyle([
                ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#0E7490")),
                ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 9),
                ("FONTNAME",     (0,-1),(-1,-1), "Helvetica-Bold"),
                ("BACKGROUND",   (0,-1),(-1,-1), colors.HexColor("#E0F7FA")),
                ("ROWBACKGROUNDS",(0,1),(-1,-2),
                    [colors.white, colors.HexColor("#F8FAFC")]),
                ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0), (-1,-1), 5),
                ("LEFTPADDING",  (0,0), (-1,-1), 8),
                ("RIGHTPADDING", (0,0), (-1,-1), 8),
                ("ALIGN",        (1,0), (-1,-1), "RIGHT"),
            ])
            story.append(met_table)
            story.append(Spacer(1, 5*mm))

    # ── Detalle de reservas ───────────────────────────────────────────────────
    story.append(Paragraph("Detalle de reservas del período", sec_style))
    story.append(hr_line())

    if rows:
        col_widths = [
            doc_pdf.width * 0.22,  # Reserva
            doc_pdf.width * 0.22,  # Cliente
            doc_pdf.width * 0.08,  # Cub
            doc_pdf.width * 0.08,  # Horas
            doc_pdf.width * 0.14,  # Total
            doc_pdf.width * 0.14,  # Método
            doc_pdf.width * 0.12,  # Estado
        ]
        det_data = [["Reserva", "Cliente", "Cub.", "Horas", "Total COP", "Método", "Estado"]]
        for r in rows:
            det_data.append([
                r.get("Reserva", ""),
                (r.get("Cliente","") or "")[:20],
                r.get("Cubículo",""),
                str(r.get("Horas","")),
                fmt_cop(r.get("Total",0)),
                r.get("Método",""),
                r.get("Estado",""),
            ])

        det_table = Table(det_data, colWidths=col_widths, repeatRows=1)
        # Color de fila Estado
        row_bg = []
        for i, r in enumerate(rows, start=1):
            bg = colors.HexColor("#F0FFF4") if r.get("Estado") == "confirmado"                  else colors.HexColor("#FEF2F2")
            row_bg.append(("BACKGROUND", (0, i), (-1, i), bg))

        det_table.setStyle([
            ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#0E7490")),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 7.5),
            ("GRID",         (0,0), (-1,-1), 0.25, colors.HexColor("#CBD5E1")),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("LEFTPADDING",  (0,0), (-1,-1), 5),
            ("RIGHTPADDING", (0,0), (-1,-1), 5),
            ("ALIGN",        (3,0), (4,-1), "RIGHT"),
        ] + row_bg)
        story.append(det_table)
    else:
        story.append(Paragraph("No hay reservas para el período seleccionado.", styles["Normal"]))

    # ── Pie de página ─────────────────────────────────────────────────────────
    story.append(Spacer(1, 6*mm))
    story.append(hr_line("#CBD5E1"))
    story.append(Paragraph(
        f"{NEGOCIO}  ·  {DIRECCION}  ·  NIT {NIT}  ·  {TELEFONO}",
        pie_style
    ))
    story.append(Paragraph(f"Reporte generado el {ahora_str}", pie_style))

    doc_pdf.build(story)
    return buf.getvalue()


def _op_reportes():
    st.markdown("### 📊 Reportes")
    tab_rep_normal, tab_rep_conv = st.tabs(["📋 Reservas Generales", "🤝 Convenios"])

    with tab_rep_normal:
        _op_reportes_core()
    with tab_rep_conv:
        _op_reportes_convenio()


def _op_reportes_core():
    """Reporte de reservas (todas excepto convenio puro)."""
    periodo = st.selectbox("Período", ["Hoy", "Esta semana", "Este mes", "Histórico"],
                           key="rep_core_periodo")
    today   = ahora_col()

    if periodo == "Hoy":
        desde = today.strftime("%Y-%m-%d")
        hasta = desde
    elif periodo == "Esta semana":
        desde = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
        hasta = today.strftime("%Y-%m-%d")
    elif periodo == "Este mes":
        desde = today.strftime("%Y-%m-01")
        hasta = today.strftime("%Y-%m-%d")
    else:
        desde = "2020-01-01"
        hasta = today.strftime("%Y-%m-%d")

    reservas_gs = _gs_read_sheet("Reservas")

    # Filtrar por período
    rows = []
    for r in reservas_gs:
        fecha_r = _gs_fecha_ymd(r, "Creado_En")
        if desde <= fecha_r <= hasta:
            rows.append({
                "Reserva":  _gs_val(r, "Numero_Reserva"),
                "Cliente":  _gs_val(r, "Cliente_Nombre"),
                "Cubículo": _gs_val(r, "Cubiculo_Num"),
                "Horas":    _gs_float(r, "Horas_Contratadas"),
                "Total":    _gs_float(r, "Total_COP"),
                "Método":   _gs_val(r, "Metodo_Pago"),
                "Estado":   _gs_val(r, "Estado_Pago"),
                "Fecha":    fecha_r,
            })

    total_ing    = sum(r["Total"] for r in rows if r["Estado"] == "confirmado")
    total_res    = len(rows)
    completadas  = sum(1 for r in rows if r["Estado"] == "confirmado")

    c1, c2, c3 = st.columns(3)
    conv_rep  = sum(1 for r in rows if r["Método"] == "Convenio")
    conv_mont = sum(r["Total"] for r in rows if r["Método"] == "Convenio")

    c1.metric("Total reservas",      total_res)
    c2.metric("Completadas",         completadas)
    c3.metric("Ingresos confirmados",fmt_cop(total_ing))
    if conv_rep > 0:
        st.info(f"🤝 **{conv_rep} reserva(s) de convenio** en el período · "
                f"Monto pendiente: **{fmt_cop(conv_mont)} COP** (no incluido en ingresos)")

    if rows:
        df = pd.DataFrame(rows)
        df["Total"] = df["Total"].apply(fmt_cop)
        for _c in ["Reserva","Cliente","Cubículo","Método","Estado","Fecha"]:
            if _c in df.columns: df[_c] = df[_c].astype(str)
        st.dataframe(df, use_container_width=True, hide_index=True)

        col_csv, col_pdf = st.columns(2)
        with col_csv:
            csv = pd.DataFrame(rows).to_csv(index=False).encode()
            st.download_button(
                "📥 Exportar CSV",
                data=csv,
                file_name=f"reservas_{desde}_{hasta}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with col_pdf:
            if REPORTLAB_AVAILABLE:
                with st.spinner("Generando PDF..."):
                    pdf_bytes = generar_reporte_pdf(
                        rows, desde, hasta, total_ing, total_res, completadas
                    )
                if pdf_bytes:
                    st.download_button(
                        "📄 Exportar PDF",
                        data=pdf_bytes,
                        file_name=f"reporte_{desde}_{hasta}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
            else:
                st.caption("*(instala reportlab para PDF)*")

        try:
            import plotly.express as px
            df_plot = pd.DataFrame(rows)
            df_plot = df_plot[df_plot["Estado"] == "confirmado"]
            if not df_plot.empty:
                fig = px.pie(df_plot, names="Método", values="Total",
                             title="Distribución por método de pago",
                             template="plotly_dark",
                             color_discrete_sequence=["#00d4ff","#3b82f6","#00ff88",
                                                      "#ffd32a","#a29bfe","#f43f5e"])
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)",
                                  plot_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            st.markdown("*(Instala plotly para ver gráficas: `pip install plotly`)*")
    else:
        st.info("No hay reservas para el período seleccionado.")

    # ── Descarga completa desde Google Sheets ─────────────────────────────────
    st.divider()
    st.markdown("#### 💾 Exportar datos desde Google Sheets")
    col_b1, col_b2 = st.columns(2)
    with col_b1:
        st.markdown("**Exportar todas las reservas** (histórico completo)")
        if st.button("📊 Exportar historial completo CSV", use_container_width=True):
            all_rows = _gs_read_sheet("Reservas")
            if all_rows:
                df_all = pd.DataFrame(all_rows)
                csv_all = df_all.to_csv(index=False).encode()
                st.download_button(
                    f"📥 Descargar historial_{ahora_col().strftime('%Y%m%d')}.csv",
                    data=csv_all,
                    file_name=f"historial_jjgt_{ahora_col().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            else:
                st.warning("No hay datos en Google Sheets para exportar.")
    with col_b2:
        st.markdown("**Backup Excel** (todas las hojas desde Sheets)")
        if st.button("🗂️ Generar backup Excel ahora", use_container_width=True):
            with st.spinner("Descargando datos desde Google Sheets..."):
                if OPENPYXL_AVAILABLE:
                    buf = io.BytesIO()
                    hojas = ["Reservas","Pagos","Clientes","Facturas",
                             "Factura_items","Cubiculos_Estado","Tarifas_Config",
                             "Operadores","Configuracion_Pagos"]
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        for hoja in hojas:
                            try:
                                data_h = _gs_read_sheet(hoja)
                                if data_h:
                                    pd.DataFrame(data_h).to_excel(writer, sheet_name=hoja[:31], index=False)
                            except Exception:
                                pass
                    now_str = ahora_col().strftime("%Y-%m-%d_%H%M")
                    st.download_button(
                        f"📥 Descargar backup_{now_str}.xlsx",
                        data=buf.getvalue(),
                        file_name=f"backup_jjgt_{now_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.warning("Instala openpyxl para generar Excel: `pip install openpyxl`")




def _op_reportes_convenio():
    """Reporte exclusivo de reservas de convenio desde jjgt_pagos (Metodo_Pago == Convenio)."""
    st.markdown("#### 🤝 Reporte de Convenios")

    periodo_c = st.selectbox("Período", ["Hoy","Esta semana","Este mes","Histórico"],
                              key="rep_conv_periodo")
    today_c   = ahora_col()
    if   periodo_c == "Hoy":
        desde_c = today_c.strftime("%Y-%m-%d"); hasta_c = desde_c
    elif periodo_c == "Esta semana":
        desde_c = (today_c - timedelta(days=today_c.weekday())).strftime("%Y-%m-%d")
        hasta_c = today_c.strftime("%Y-%m-%d")
    elif periodo_c == "Este mes":
        desde_c = today_c.strftime("%Y-%m-01"); hasta_c = today_c.strftime("%Y-%m-%d")
    else:
        desde_c = "2020-01-01"; hasta_c = today_c.strftime("%Y-%m-%d")

    # Leer de jjgt_convenios → hoja Reservas (fuente correcta de datos de convenio)
    _, sh_rep_c = get_active_client_convenios()
    if not sh_rep_c:
        st.warning("⚠️ jjgt_convenios no configurado. Ve a ⚙️ Configuración → Convenios.")
        return
    rows_c = []
    try:
        reservas_conv = _gs_read_sheet_conv("Reservas", force=False)
        for r in reservas_conv:
            fecha_r = _gs_fecha_ymd(r, "Creado_En")
            if desde_c <= fecha_r <= hasta_c:
                rows_c.append({
                    "Reserva":   _gs_val(r,"Numero_Reserva"),
                    "Empresa":   _gs_val(r,"Nombre_Empresa","—"),
                    "Id_Empresa":_gs_val(r,"Id_Empresa",""),
                    "Cliente":   _gs_val(r,"Cliente_Nombre"),
                    "Cubículo":  _gs_val(r,"Cubiculo_Num"),
                    "Horas":     _gs_float(r,"Horas_Contratadas"),
                    "Total":     _gs_float(r,"Total_COP"),
                    "Estado":    _gs_val(r,"Estado_Pago"),
                    "Fecha":     fecha_r,
                })
    except Exception as e_rep:
        st.error(f"Error leyendo jjgt_convenios → Reservas: {e_rep}")
        return

    total_c   = sum(r["Total"] for r in rows_c)
    pend_c    = sum(r["Total"] for r in rows_c if r["Estado"] == "pendiente")
    conf_c    = sum(r["Total"] for r in rows_c if r["Estado"] == "confirmado")
    n_emp_c   = len({r["Empresa"] for r in rows_c})

    mc1,mc2,mc3,mc4,mc5 = st.columns(5)
    mc1.metric("Reservas conv.",    len(rows_c))
    mc2.metric("Empresas",          n_emp_c)
    mc3.metric("Total facturado",   fmt_cop(total_c))
    mc4.metric("Pendiente cobro",   fmt_cop(pend_c))
    mc5.metric("Confirmado",        fmt_cop(conf_c))

    if rows_c:
        df_rc = pd.DataFrame(rows_c)
        df_rc_disp = df_rc.copy()
        df_rc_disp["Total"]    = df_rc_disp["Total"].apply(fmt_cop)
        if "Descuento" in df_rc_disp.columns:
            df_rc_disp["Descuento"] = df_rc_disp["Descuento"].apply(fmt_cop)
        for _c in ["Reserva","Empresa","Id_Empresa","Cliente","Cubículo","Estado","Fecha"]:
            if _c in df_rc_disp.columns: df_rc_disp[_c] = df_rc_disp[_c].astype(str)
        st.dataframe(df_rc_disp, use_container_width=True, hide_index=True)

        # Resumen por empresa
        st.divider()
        st.markdown("**Resumen por empresa:**")
        emp_res = {}
        for r in rows_c:
            emp = r["Empresa"]
            if emp not in emp_res:
                emp_res[emp] = {"Empresa":emp,"Reservas":0,"Total":0.0,"Pendiente":0.0}
            emp_res[emp]["Reservas"]  += 1
            emp_res[emp]["Total"]     += r["Total"]
            if r["Estado"] == "pendiente":
                emp_res[emp]["Pendiente"] += r["Total"]
        df_emp = pd.DataFrame(list(emp_res.values()))
        df_emp["Total"]     = df_emp["Total"].apply(fmt_cop)
        df_emp["Pendiente"] = df_emp["Pendiente"].apply(fmt_cop)
        for _c in ["Empresa"]:
            if _c in df_emp.columns: df_emp[_c] = df_emp[_c].astype(str)
        st.dataframe(df_emp, use_container_width=True, hide_index=True)

        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            st.download_button(
                "📥 Exportar CSV Convenios",
                data=pd.DataFrame(rows_c).to_csv(index=False).encode(),
                file_name=f"convenios_{desde_c}_{hasta_c}.csv",
                mime="text/csv", use_container_width=True)
        with col_exp2:
            if OPENPYXL_AVAILABLE:
                buf_rc = io.BytesIO()
                with pd.ExcelWriter(buf_rc, engine="openpyxl") as writer:
                    pd.DataFrame(rows_c).to_excel(writer,
                        sheet_name="Reservas_Convenios", index=False)
                    pd.DataFrame(list({e["Empresa"]:e for e in list(emp_res.values())}.values())).to_excel(
                        writer, sheet_name="Resumen_Empresa", index=False)
                st.download_button(
                    "📊 Exportar Excel Convenios",
                    data=buf_rc.getvalue(),
                    file_name=f"reporte_convenios_{desde_c}_{hasta_c}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

        # También leer de jjgt_convenios si está configurado
        _, sh_rc = get_active_client_convenios()
        if sh_rc:
            with st.expander("📂 Ver datos completos en jjgt_convenios"):
                try:
                    vals_rc_raw = _pg_read_table_conv("Reservas")
                    # Convert to list-of-lists format expected by caller
                    if vals_rc_raw:
                        hdr_rc = list(vals_rc_raw[0].keys())
                        vals_rc = [hdr_rc] + [[r.get(k,"") for k in hdr_rc] for r in vals_rc_raw]
                    else:
                        vals_rc = []
                    if vals_rc and len(vals_rc) > 1:
                        df_jjgt_c = pd.DataFrame(vals_rc[1:], columns=vals_rc[0])
                        st.success(f"✅ {len(df_jjgt_c)} registros en jjgt_convenios → Reservas")
                        st.dataframe(df_jjgt_c, use_container_width=True, hide_index=True)
                        st.download_button(
                            "📥 Descargar Reservas jjgt_convenios",
                            data=df_jjgt_c.to_csv(index=False).encode(),
                            file_name=f"jjgt_conv_reservas_{today_c.strftime('%Y%m%d')}.csv",
                            mime="text/csv", use_container_width=True)
                    else:
                        st.info("jjgt_convenios → Reservas vacía o sin datos.")
                except Exception as e:
                    st.error(f"Error leyendo jjgt_convenios: {e}")
    else:
        st.info("No hay reservas de convenio para el período seleccionado.")

def _op_gestion_datos():
    """
    Módulo de gestión de datos — permite al admin eliminar registros
    de reservas, pagos, clientes y facturas. Lee y escribe en Google Sheets.
    """
    st.markdown("### 🗑️ Gestión de Datos")
    st.warning("⚠️ **Zona de administración** — Las eliminaciones son permanentes e irreversibles.")

    op = st.session_state.get("operador_info", {})
    if op.get("rol") != "admin" and "admin" not in op.get("permisos", []):
        st.error("❌ Solo los administradores pueden acceder a esta sección.")
        return

    tab_res, tab_cli, tab_pag, tab_fact, tab_gd_conv = st.tabs([
        "📋 Reservas", "👤 Clientes", "💰 Pagos", "🧾 Facturas", "🤝 Convenios"])

    def _gs_delete_row(sh, hoja, col_clave, valor_clave):
        """Elimina una fila en PostgreSQL buscando por col_clave == valor_clave.
        sh se ignora — fuente única es PostgreSQL.
        Para tablas de convenio usa _pg_delete_row_conv si el prefijo es CONV_.
        """
        _pg_delete_row(hoja, col_clave, valor_clave)

    def _gs_update_cell_by_key(sh, hoja, col_clave, valor_clave, col_dest, nuevo_valor):
        """Actualiza col_dest en PostgreSQL donde col_clave == valor_clave."""
        _pg_update_col(hoja, col_clave, valor_clave, col_dest, str(nuevo_valor))

    # ── RESERVAS ──────────────────────────────────────────────────────────────
    with tab_res:
        st.markdown("#### Eliminar reservas")
        reservas_gs = _gs_read_sheet("Reservas")
        if not reservas_gs:
            st.info("No hay reservas registradas en Google Sheets.")
        else:
            df_r = pd.DataFrame([{
                "Número":   _gs_val(r, "Numero_Reserva"),
                "Cliente":  _gs_val(r, "Cliente_Nombre"),
                "Cubículo": _gs_val(r, "Cubiculo_Num"),
                "Inicio":   _gs_val(r, "Hora_Inicio", "")[:16],
                "Total":    fmt_cop(_gs_float(r, "Total_COP")),
                "Estado":   _gs_val(r, "Estado_Pago"),
                "Creada":   _gs_fecha_ymd(r, "Creado_En"),
            } for r in reservas_gs])
            for _c in ["Número","Cliente","Cubículo","Inicio","Total","Estado","Creada"]:
                if _c in df_r.columns: df_r[_c] = df_r[_c].astype(str)
            st.dataframe(df_r, use_container_width=True, hide_index=True)

            col_r1, col_r2 = st.columns([2,1])
            with col_r1:
                nums_r = [_gs_val(r, "Numero_Reserva") for r in reservas_gs if _gs_val(r, "Numero_Reserva")]
                sel_num_r = st.selectbox("Seleccionar reserva", nums_r, key="del_res_sel")
            with col_r2:
                pin_del_r = st.text_input("PIN admin para confirmar", type="password", key="del_res_pin")
            if st.button("🗑️ ELIMINAR RESERVA SELECCIONADA", type="primary",
                         use_container_width=True, key="btn_del_res"):
                if not verificar_pin(pin_del_r, "admin"):
                    st.error("❌ PIN admin incorrecto")
                else:
                    _, sh_del = get_active_client()
                    if sh_del:
                        _gs_delete_row(sh_del, "Reservas", "Numero_Reserva", sel_num_r)
                        _gs_delete_row(sh_del, "Pagos", "Num_Reserva", sel_num_r)
                        # Liberar cubículo si estaba ocupado
                        res_data = next((r for r in reservas_gs if _gs_val(r, "Numero_Reserva") == sel_num_r), {})
                        cub_num = _gs_val(res_data, "Cubiculo_Num")
                        if cub_num:
                            cubiculos_now = get_cubiculos()
                            cub_obj = next((c for c in cubiculos_now if c["numero"] == cub_num), None)
                            if cub_obj and cub_obj["estado"] != "libre":
                                liberar_cubiculo(cub_obj["id"])
                        gs_sync_cubiculos(sh_del)
                        _gs_invalidate_cache("Reservas", "Pagos", "Cubiculos_Estado")
                    st.success(f"✅ Reserva {sel_num_r} eliminada")
                    st.rerun()

    # ── CLIENTES ──────────────────────────────────────────────────────────────
    with tab_cli:
        st.markdown("#### Eliminar clientes")
        clientes_gs = _gs_read_sheet("Clientes")
        if not clientes_gs:
            st.info("No hay clientes registrados en Google Sheets.")
        else:
            df_c = pd.DataFrame([{
                "Nombre":    _gs_val(r, "Nombre"),
                "Tipo Doc":  _gs_val(r, "Tipo_Doc"),
                "Documento": _gs_val(r, "Num_Doc"),
                "Teléfono":  _gs_val(r, "Telefono"),
                "Creado":    _gs_fecha_ymd(r, "Creado_En"),
            } for r in clientes_gs])
            for _c in ["Nombre","Tipo Doc","Documento","Teléfono","Creado"]:
                if _c in df_c.columns: df_c[_c] = df_c[_c].astype(str)
            st.dataframe(df_c, use_container_width=True, hide_index=True)

            col_c1, col_c2 = st.columns([2,1])
            with col_c1:
                cli_opts = {f"{_gs_val(r,'Nombre')} ({_gs_val(r,'Num_Doc') or 'sin doc'})": _gs_val(r,"Num_Doc")
                            for r in clientes_gs if _gs_val(r, "Nombre")}
                sel_cli = st.selectbox("Seleccionar cliente", list(cli_opts.keys()), key="gd_pagos_cli_sel", index=0)
                cli_doc_str = cli_opts[sel_cli]
            with col_c2:
                pin_del_c = st.text_input("PIN admin", type="password", key="del_cli_pin")
            st.warning("⚠️ Eliminar un cliente también eliminará sus reservas y pagos asociados.")
            if st.button("🗑️ ELIMINAR CLIENTE", type="primary",
                         use_container_width=True, key="btn_del_cli"):
                if not verificar_pin(pin_del_c, "admin"):
                    st.error("❌ PIN admin incorrecto")
                else:
                    _, sh_del2 = get_active_client()
                    if sh_del2:
                        # Eliminar reservas y pagos del cliente
                        reservas_gs2 = _gs_read_sheet("Reservas")
                        for r in reservas_gs2:
                            if _gs_val(r, "Documento") == cli_doc_str:
                                nr = _gs_val(r, "Numero_Reserva")
                                _gs_delete_row(sh_del2, "Reservas", "Numero_Reserva", nr)
                                _gs_delete_row(sh_del2, "Pagos", "Num_Reserva", nr)
                        _gs_delete_row(sh_del2, "Clientes", "Num_Doc", cli_doc_str)
                        gs_sync_cubiculos(sh_del2)
                        _gs_invalidate_cache("Clientes", "Reservas", "Pagos", "Cubiculos_Estado")
                    st.success(f"✅ Cliente eliminado junto a sus reservas")
                    st.rerun()

    # ── PAGOS ──────────────────────────────────────────────────────────────────
    with tab_pag:
        st.markdown("#### Eliminar / anular pagos")
        pagos_gs = _gs_read_sheet("Pagos")
        if not pagos_gs:
            st.info("No hay pagos registrados en Google Sheets.")
        else:
            df_p = pd.DataFrame([{
                "ID Pago":  _gs_val(r, "ID_Pago"),
                "Reserva":  _gs_val(r, "Num_Reserva"),
                "Monto":    fmt_cop(_gs_float(r, "Monto_COP")),
                "Método":   _gs_val(r, "Metodo"),
                "Estado":   _gs_val(r, "Estado"),
                "Fecha":    _gs_val(r, "Fecha_Pago", "")[:16],
            } for r in pagos_gs])
            for _c in ["ID Pago","Reserva","Monto","Método","Estado","Fecha"]:
                if _c in df_p.columns: df_p[_c] = df_p[_c].astype(str)
            st.dataframe(df_p, use_container_width=True, hide_index=True)

            col_p1, col_p2 = st.columns([2,1])
            with col_p1:
                pago_opts = {f"#{_gs_val(r,'ID_Pago')} — {_gs_val(r,'Num_Reserva')} — {fmt_cop(_gs_float(r,'Monto_COP'))}":
                             _gs_val(r, "ID_Pago") for r in pagos_gs if _gs_val(r, "ID_Pago")}
                sel_pago_id = st.selectbox("Seleccionar pago", list(pago_opts.keys()), key="gd_pagos_pag_sel", index=0)
                sel_pago_key = pago_opts[sel_pago_id]
            with col_p2:
                pin_del_p = st.text_input("PIN admin", type="password", key="del_pago_pin")
            accion_p = st.radio("Acción", ["Anular (marcar como anulado)", "Eliminar definitivamente"],
                                key="del_pago_accion", horizontal=True)
            if st.button("✅ APLICAR ACCIÓN", type="primary",
                         use_container_width=True, key="btn_del_pago"):
                if not verificar_pin(pin_del_p, "admin"):
                    st.error("❌ PIN admin incorrecto")
                else:
                    _, sh_pag = get_active_client()
                    if sh_pag:
                        if "Anular" in accion_p:
                            _gs_update_cell_by_key(sh_pag, "Pagos", "ID_Pago", sel_pago_key, "Estado", "anulado")
                            _gs_invalidate_cache("Pagos")
                            st.success(f"✅ Pago anulado")
                        else:
                            _gs_delete_row(sh_pag, "Pagos", "ID_Pago", sel_pago_key)
                            _gs_invalidate_cache("Pagos")
                            st.success(f"✅ Pago eliminado")
                    st.rerun()

    # ── FACTURAS ──────────────────────────────────────────────────────────────
    with tab_fact:
        st.markdown("#### Eliminar / anular facturas")
        facturas_gs = _gs_read_sheet("Facturas")
        if not facturas_gs:
            st.info("No hay facturas registradas en Google Sheets.")
        else:
            df_f = pd.DataFrame([{
                "Número":  _gs_val(r, "Num_Factura"),
                "Cliente": _gs_val(r, "Cliente__id"),
                "Total":   fmt_cop(_gs_float(r, "Total_COP")),
                "Estado":  _gs_val(r, "Estado"),
                "Fecha":   _gs_fecha_ymd(r, "Fecha_Emision"),
            } for r in facturas_gs])
            for _c in ["Número","Cliente","Total","Estado","Fecha"]:
                if _c in df_f.columns: df_f[_c] = df_f[_c].astype(str)
            st.dataframe(df_f, use_container_width=True, hide_index=True)

            col_f1, col_f2 = st.columns([2,1])
            with col_f1:
                fact_opts = {f"{_gs_val(r,'Num_Factura')} — {_gs_val(r,'Cliente__id')}":
                             _gs_val(r, "Num_Factura") for r in facturas_gs if _gs_val(r, "Num_Factura")}
                sel_fact_num = st.selectbox("Seleccionar factura", list(fact_opts.keys()), key="gd_pagos_fac_sel",index=0)
                sel_fact_key = fact_opts[sel_fact_num]
            with col_f2:
                pin_del_f = st.text_input("PIN admin", type="password", key="del_fact_pin")
            accion_f = st.radio("Acción", ["Anular (marcar como anulada)", "Eliminar definitivamente"],
                                key="del_fact_accion", horizontal=True)
            if st.button("✅ APLICAR ACCIÓN", type="primary",
                         use_container_width=True, key="btn_del_fact"):
                if not verificar_pin(pin_del_f, "admin"):
                    st.error("❌ PIN admin incorrecto")
                else:
                    _, sh_fact = get_active_client()
                    if sh_fact:
                        if "Anular" in accion_f:
                            _gs_update_cell_by_key(sh_fact, "Facturas", "Num_Factura", sel_fact_key, "Estado", "anulada")
                            st.success("✅ Factura anulada")
                        else:
                            _gs_delete_row(sh_fact, "Facturas", "Num_Factura", sel_fact_key)
                            _gs_delete_row(sh_fact, "Factura_items", "ID_Factura", sel_fact_key)
                            st.success("✅ Factura eliminada definitivamente")
                        _gs_invalidate_cache("Facturas", "Factura_items")
                    st.rerun()

    # ── TAB CONVENIOS ─────────────────────────────────────────────────────────
    with tab_gd_conv:
        st.markdown("#### 🤝 Gestión de datos de Convenios")
        st.caption("Operaciones sobre jjgt_convenios: Reservas, Clientes, Pagos, Facturas y Empresas.")

        _, sh_gd_conv = get_active_client_convenios()
        if not sh_gd_conv:
            st.warning("⚠️ jjgt_convenios no configurado. Ve a ⚙️ Configuración → Convenios.")
        else:
            sub_gd_res, sub_gd_cli, sub_gd_pag, sub_gd_fact, sub_gd_emp = st.tabs([
                "📋 Reservas Conv.", "👤 Clientes Conv.",
                "💰 Pagos Conv.", "🧾 Facturas Conv.", "🏢 Empresas"])

            # ── Sub-tab: Reservas de convenio ─────────────────────────────────
            with sub_gd_res:
                st.markdown("**Eliminar reservas de jjgt_convenios**")
                if st.button("🔄 Actualizar lista", key="btn_reload_gd_res"):
                    _gs_invalidate_cache_conv("Reservas")
                try:
                    res_conv_gd = _gs_read_sheet_conv("Reservas", force=False)
                    if not res_conv_gd:
                        st.info("No hay reservas en jjgt_convenios.")
                    else:
                        _df_gd_r = pd.DataFrame([{
                            "Reserva":  str(r.get("Numero_Reserva","")),
                            "Empresa":  str(r.get("Nombre_Empresa","—")),
                            "Cliente":  str(r.get("Cliente_Nombre","")),
                            "Estado":   str(r.get("Estado_Pago","")),
                            "Total":    fmt_cop(float(r.get("Total_COP","0") or 0)),
                            "Fecha":    str(r.get("Creado_En","")[:10]),
                        } for r in res_conv_gd])
                        st.dataframe(_df_gd_r, use_container_width=True, hide_index=True)
                        col_gr1, col_gr2 = st.columns([2,1])
                        with col_gr1:
                            opts_gr = {f"{r.get('Numero_Reserva','')} — {r.get('Nombre_Empresa','—')} — {r.get('Cliente_Nombre','')}":
                                       r.get("Numero_Reserva","") for r in res_conv_gd if r.get("Numero_Reserva")}
                            sel_gr = st.selectbox("Reserva a eliminar", list(opts_gr.keys()), key="gd_conv_res_eliminar_sel")
                            sel_gr_k = opts_gr.get(sel_gr,"")
                        with col_gr2:
                            pin_gr = st.text_input("PIN admin", type="password", key="gd_conv_res_pin")
                        if st.button("🗑️ Eliminar reserva de jjgt_convenios", type="primary",
                                     use_container_width=True, key="btn_gd_conv_res"):
                            if not verificar_pin(pin_gr, "admin"):
                                st.error("PIN admin incorrecto")
                            elif sel_gr_k:
                                try:
                                    _pg_delete_row_conv("Reservas", "Numero_Reserva", sel_gr_k)
                                    _pg_delete_row_conv("Pagos",    "Num_Reserva",    sel_gr_k)
                                    st.success(f"✅ Reserva {sel_gr_k} eliminada de jjgt_convenios")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {e}")
                except Exception as e:
                    st.error(f"Error accediendo a jjgt_convenios: {e}")

            # ── Sub-tab: Clientes de convenio ─────────────────────────────────
            with sub_gd_cli:
                st.markdown("**Eliminar clientes de jjgt_convenios**")
                if st.button("🔄 Actualizar lista", key="btn_reload_gd_cli"):
                    _gs_invalidate_cache_conv("Clientes")
                try:
                    cli_conv_gd = _gs_read_sheet_conv("Clientes", force=False)
                    if not cli_conv_gd:
                        st.info("No hay clientes en jjgt_convenios.")
                    else:
                        _df_gd_c = pd.DataFrame([{
                            "Nombre":  str(c.get("Nombre","")),
                            "Doc":     str(c.get("Num_Doc","")),
                            "Empresa": str(c.get("Nombre_Empresa","—")),
                        } for c in cli_conv_gd])
                        st.dataframe(_df_gd_c, use_container_width=True, hide_index=True)
                        col_gc1, col_gc2 = st.columns([2,1])
                        with col_gc1:
                            opts_gc = {f"{c.get('Nombre','')} — {c.get('Num_Doc','')}":
                                       c.get("Num_Doc","") for c in cli_conv_gd if c.get("Num_Doc")}
                            sel_gc = st.selectbox("Cliente a eliminar", list(opts_gc.keys()), key="gd_conv_cli_sel")
                            sel_gc_k = opts_gc.get(sel_gc,"")
                        with col_gc2:
                            pin_gc = st.text_input("PIN admin", type="password", key="gd_conv_cli_pin")
                        if st.button("🗑️ Eliminar cliente de jjgt_convenios", type="primary",
                                     use_container_width=True, key="btn_gd_conv_cli"):
                            if not verificar_pin(pin_gc, "admin"):
                                st.error("PIN admin incorrecto")
                            elif sel_gc_k:
                                _pg_delete_row_conv("Clientes", "Num_Doc", sel_gc_k)
                                st.success(f"✅ Cliente eliminado de jjgt_convenios")
                                st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

            # ── Sub-tab: Pagos de convenio ────────────────────────────────────
            with sub_gd_pag:
                st.markdown("**Anular / eliminar pagos de jjgt_convenios**")
                if st.button("🔄 Actualizar lista", key="btn_reload_gd_pag"):
                    _gs_invalidate_cache_conv("Pagos")
                try:
                    pag_conv_gd = _gs_read_sheet_conv("Pagos", force=False)
                    if not pag_conv_gd:
                        st.info("No hay pagos en jjgt_convenios.")
                    else:
                        _df_gd_p = pd.DataFrame([{
                            "ID":      str(p.get("ID_Pago","")),
                            "Reserva": str(p.get("Num_Reserva","")),
                            "Empresa": str(p.get("Nombre_Empresa","—")),
                            "Monto":   fmt_cop(float(p.get("Monto_COP","0") or 0)),
                            "Estado":  str(p.get("Estado","")),
                        } for p in pag_conv_gd])
                        st.dataframe(_df_gd_p, use_container_width=True, hide_index=True)
                        col_gp1, col_gp2 = st.columns([2,1])
                        with col_gp1:
                            opts_gp = {f"#{p.get('ID_Pago','')} — {p.get('Num_Reserva','')} — {p.get('Nombre_Empresa','—')}":
                                       p.get("ID_Pago","") for p in pag_conv_gd if p.get("ID_Pago")}
                            sel_gp = st.selectbox("Pago a gestionar", list(opts_gp.keys()), key="gd_conv_pag_sel")
                            sel_gp_k = opts_gp.get(sel_gp,"")
                        with col_gp2:
                            pin_gp = st.text_input("PIN admin", type="password", key="gd_conv_pag_pin")
                        accion_gp = st.radio("Acción", ["Anular (marcar como anulado)", "Eliminar definitivamente"],
                                             key="gd_conv_pag_accion", horizontal=True)
                        if st.button("✅ Aplicar acción al pago", type="primary",
                                     use_container_width=True, key="btn_gd_conv_pag"):
                            if not verificar_pin(pin_gp, "admin"):
                                st.error("PIN admin incorrecto")
                            elif sel_gp_k:
                                if "Anular" in accion_gp:
                                    try:
                                        try:
                                            _pg_update_col_conv("Pagos","ID_Pago",sel_gp_k,"Estado","anulado")
                                        except Exception: pass
                                        st.success("✅ Pago anulado en jjgt_convenios")
                                    except Exception as e:
                                        st.error(f"Error: {e}")
                                else:
                                    _pg_delete_row_conv("Pagos", "ID_Pago", sel_gp_k)
                                    st.success("✅ Pago eliminado de jjgt_convenios")
                                st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

            # ── Sub-tab: Facturas de convenio ─────────────────────────────────
            with sub_gd_fact:
                st.markdown("**Anular / eliminar facturas de jjgt_convenios**")
                try:
                    fact_conv_gd = _pg_read_table_conv("Facturas")
                    if not fact_conv_gd:
                        st.info("No hay facturas en jjgt_convenios.")
                    else:
                        _df_gd_f = pd.DataFrame([{
                            "Número":  str(f.get("Num_Factura","")),
                            "Empresa": str(f.get("Nombre_Empresa","—")),
                            "Cliente": str(f.get("Cliente_Nombre","")),
                            "Total":   fmt_cop(float(f.get("Total_COP","0") or 0)),
                            "Estado":  str(f.get("Estado","")),
                        } for f in fact_conv_gd])
                        st.dataframe(_df_gd_f, use_container_width=True, hide_index=True)
                        col_gf1, col_gf2 = st.columns([2,1])
                        with col_gf1:
                            opts_gf = {f"{f.get('Num_Factura','')} — {f.get('Nombre_Empresa','—')}":
                                       f.get("Num_Factura","") for f in fact_conv_gd if f.get("Num_Factura")}
                            sel_gf = st.selectbox("Factura a gestionar", list(opts_gf.keys()), key="gd_conv_fact_sel")
                            sel_gf_k = opts_gf.get(sel_gf,"")
                        with col_gf2:
                            pin_gf = st.text_input("PIN admin", type="password", key="gd_conv_fact_pin")
                        accion_gf = st.radio("Acción", ["Anular (marcar como anulada)", "Eliminar definitivamente"],
                                             key="gd_conv_fact_accion", horizontal=True)
                        if st.button("✅ Aplicar acción a factura", type="primary",
                                     use_container_width=True, key="btn_gd_conv_fact"):
                            if not verificar_pin(pin_gf, "admin"):
                                st.error("PIN admin incorrecto")
                            elif sel_gf_k:
                                if "Anular" in accion_gf:
                                    try:
                                        try:
                                            _pg_update_col_conv("Facturas","Num_Factura",sel_gf_k,"Estado","anulada")
                                        except Exception: pass
                                        st.success("✅ Factura anulada en jjgt_convenios")
                                    except Exception as e:
                                        st.error(f"Error: {e}")
                                else:
                                    _pg_delete_row_conv("Facturas",      "Num_Factura", sel_gf_k)
                                    _pg_delete_row_conv("Factura_items", "ID_Factura",  sel_gf_k)
                                    st.success("✅ Factura eliminada de jjgt_convenios")
                                st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

            # ── Sub-tab: Empresas ─────────────────────────────────────────────
            with sub_gd_emp:
                st.markdown("**Gestionar empresas del convenio en jjgt_convenios**")
                empresas_gd = get_empresas_convenio()
                if not empresas_gd:
                    st.info("No hay empresas registradas en jjgt_convenios.")
                else:
                    _df_gd_e = pd.DataFrame([{
                        "ID":      str(e.get("Id_Empresa","")),
                        "Empresa": str(e.get("Nombre_Empresa","")),
                        "NIT":     str(e.get("Nit_Empresa","")),
                        "Tipo":    str(e.get("Tipo_Convenio","")),
                        "Activo":  str(e.get("Activo","1")),
                    } for e in empresas_gd])
                    st.dataframe(_df_gd_e, use_container_width=True, hide_index=True)

                    col_ge1, col_ge2 = st.columns([2,1])
                    with col_ge1:
                        opts_ge = {f"{e.get('Nombre_Empresa','')} ({e.get('Nit_Empresa','')})":
                                   e.get("Id_Empresa","") for e in empresas_gd}
                        sel_ge = st.selectbox("Empresa a gestionar", list(opts_ge.keys()), key="gd_emp_sel")
                        sel_ge_k = opts_ge.get(sel_ge,"")
                    with col_ge2:
                        pin_ge = st.text_input("PIN admin", type="password", key="gd_emp_pin")

                    accion_ge = st.radio("Acción", ["Desactivar", "Activar", "Eliminar definitivamente"],
                                         key="gd_emp_accion", horizontal=True)
                    if st.button("✅ Aplicar acción a empresa", type="primary",
                                 use_container_width=True, key="btn_gd_emp"):
                        if not verificar_pin(pin_ge, "admin"):
                            st.error("PIN admin incorrecto")
                        elif sel_ge_k:
                            if accion_ge == "Eliminar definitivamente":
                                _pg_delete_row_conv("Empresa", "Id_Empresa", sel_ge_k)
                                st.success("✅ Empresa eliminada de jjgt_convenios")
                            else:
                                new_val = "0" if accion_ge == "Desactivar" else "1"
                                try:
                                    _pg_update_col_conv("Empresa","Id_Empresa",sel_ge_k,"Activo",new_val)
                                    st.success(f"✅ Empresa {'desactivada' if new_val=='0' else 'activada'}")
                                except Exception as e:
                                    st.error(f"Error: {e}")
                            st.rerun()

def _op_google_drive():
    """Panel de datos PostgreSQL — consulta, resumen y descarga de todas las tablas."""
    st.markdown("### 🗄️ Base de Datos PostgreSQL")
    st.info(
        "Toda la información se almacena y lee directamente desde **PostgreSQL**. "
        "Selecciona una tabla para consultarla o descárgala como CSV."
    )

    # ── Resumen de todas las tablas ───────────────────────────────────────────
    st.divider()
    st.markdown("#### 📊 Resumen de tablas")

    col_r1, col_r2 = st.columns(2)
    with col_r1:
        if st.button("🔎 Ver conteo de registros — jjgt_pagos",
                     use_container_width=True, key="btn_resumen_pagos"):
            with st.spinner("Consultando jjgt_pagos..."):
                filas_res = []
                for hoja in DRIVE_SHEETS.keys():
                    try:
                        n = len(_pg_read_table(hoja))
                        estado = "✅"
                    except Exception as e:
                        n = 0
                        estado = f"❌ {e}"
                    filas_res.append({"Tabla": hoja, "Registros": n, "Estado": estado})
                df_res = pd.DataFrame(filas_res)
                st.dataframe(df_res, use_container_width=True, hide_index=True)
                st.caption(f"Total registros: **{df_res['Registros'].sum()}**")

    with col_r2:
        if st.button("🔎 Ver conteo de registros — jjgt_convenios",
                     use_container_width=True, key="btn_resumen_conv"):
            with st.spinner("Consultando jjgt_convenios..."):
                filas_res_c = []
                for hoja in DRIVE_SHEETS_CONVENIOS.keys():
                    try:
                        n = len(_pg_read_table_conv(hoja))
                        estado = "✅"
                    except Exception as e:
                        n = 0
                        estado = f"❌ {e}"
                    filas_res_c.append({"Tabla": hoja, "Registros": n, "Estado": estado})
                df_res_c = pd.DataFrame(filas_res_c)
                st.dataframe(df_res_c, use_container_width=True, hide_index=True)
                st.caption(f"Total registros: **{df_res_c['Registros'].sum()}**")

    # ── Recargar caché de cubículos y config ──────────────────────────────────
    st.divider()
    if st.button("🔄 Recargar todos los datos desde PostgreSQL",
                 type="primary", use_container_width=True, key="btn_reload_pg"):
        with st.spinner("Recargando datos..."):
            _gs_invalidate_cache(*list(DRIVE_SHEETS.keys()))
            _gs_invalidate_cache_conv(*list(DRIVE_SHEETS_CONVENIOS.keys()))
            with _config_cache_lock:
                _config_cache.clear()
            cubiculos_now = get_cubiculos()
            st.session_state.pop("_conv_cfg_pendiente", None)
            st.success(f"✅ Datos recargados — {len(cubiculos_now)} cubículos activos")
            st.rerun()

    # ── Consultar tabla jjgt_pagos ────────────────────────────────────────────
    st.divider()
    st.markdown("#### 🗃️ Consultar tabla — jjgt_pagos")
    hoja_sel = st.selectbox(
        "Selecciona la tabla:",
        list(DRIVE_SHEETS.keys()),
        key="gd_drive_hoja_sel",
    )
    col_p1, col_p2 = st.columns([3, 1])
    with col_p1:
        limite_p = st.number_input(
            "Máximo de filas a mostrar (0 = todos)",
            min_value=0, max_value=100000, value=500, step=100,
            key="gd_limite_pagos",
        )
    with col_p2:
        st.write("")
        st.write("")
        cargar_p = st.button("🔄 Cargar datos", use_container_width=True, key="btn_cargar_pagos")

    if cargar_p:
        with st.spinner(f"Leyendo '{hoja_sel}' desde PostgreSQL..."):
            try:
                filas_p = _pg_read_table(hoja_sel)
                if not filas_p:
                    st.info(f"La tabla '{hoja_sel}' está vacía.")
                else:
                    df_p = pd.DataFrame(filas_p)
                    total_p = len(df_p)
                    if limite_p and limite_p < total_p:
                        df_p = df_p.head(limite_p)
                        st.warning(f"Mostrando {limite_p} de {total_p} registros.")
                    else:
                        st.success(f"✅ {total_p} registros en '{hoja_sel}'")
                    st.dataframe(df_p, use_container_width=True, hide_index=True)
                    # Descargar CSV completo (sin límite)
                    df_full_p = pd.DataFrame(_pg_read_table(hoja_sel))
                    st.download_button(
                        f"📥 Descargar {hoja_sel}.csv (completo)",
                        data=df_full_p.to_csv(index=False).encode("utf-8"),
                        file_name=f"jjgt_{hoja_sel.lower()}_{ahora_col().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="dl_pagos_csv",
                    )
            except Exception as e:
                st.error(f"❌ Error leyendo '{hoja_sel}': {e}")

    # ── Consultar tabla jjgt_convenios ────────────────────────────────────────
    st.divider()
    st.markdown("#### 🤝 Consultar tabla — jjgt_convenios")
    hoja_conv_drive = st.selectbox(
        "Selecciona la tabla de convenios:",
        list(DRIVE_SHEETS_CONVENIOS.keys()),
        key="gs_conv_drive_hoja",
    )
    col_c1, col_c2 = st.columns([3, 1])
    with col_c1:
        limite_c = st.number_input(
            "Máximo de filas a mostrar (0 = todos)",
            min_value=0, max_value=100000, value=500, step=100,
            key="gd_limite_conv",
        )
    with col_c2:
        st.write("")
        st.write("")
        cargar_c = st.button("🔄 Cargar datos", use_container_width=True, key="btn_cargar_conv")

    if cargar_c:
        with st.spinner(f"Leyendo '{hoja_conv_drive}' desde PostgreSQL..."):
            try:
                filas_c = _pg_read_table_conv(hoja_conv_drive)
                if not filas_c:
                    st.info(f"La tabla '{hoja_conv_drive}' está vacía.")
                else:
                    df_c = pd.DataFrame(filas_c)
                    total_c = len(df_c)
                    if limite_c and limite_c < total_c:
                        df_c = df_c.head(limite_c)
                        st.warning(f"Mostrando {limite_c} de {total_c} registros.")
                    else:
                        st.success(f"✅ {total_c} registros en '{hoja_conv_drive}'")
                    st.dataframe(df_c, use_container_width=True, hide_index=True)
                    df_full_c = pd.DataFrame(_pg_read_table_conv(hoja_conv_drive))
                    st.download_button(
                        f"📥 Descargar {hoja_conv_drive}.csv (completo)",
                        data=df_full_c.to_csv(index=False).encode("utf-8"),
                        file_name=f"conv_{hoja_conv_drive.lower()}_{ahora_col().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="dl_conv_csv",
                    )
            except Exception as e:
                st.error(f"❌ Error leyendo '{hoja_conv_drive}': {e}")

    # ── Índice de tablas ──────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### 📋 Índice de tablas disponibles")
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.markdown("**jjgt_pagos**")
        for nombre, cols in DRIVE_SHEETS.items():
            st.markdown(f"- **{nombre}** · {len(cols)} columnas")
    with col_info2:
        st.markdown("**jjgt_convenios**")
        for nombre, cols in DRIVE_SHEETS_CONVENIOS.items():
            st.markdown(f"- **{nombre}** · {len(cols)} columnas")


def _op_configuracion():
    st.markdown("### ⚙️ Configuración del Sistema")
    tabs = st.tabs(["🏢 Negocio", "💰 Pagos", "🛏️ Tarifas", "👤 Operadores", "🔗 Google Sheets", "🤝 Convenios", "🗂️ Config Pagos Convenios"])
    # Constantes compartidas entre tabs
    _TURNOS            = ["mañana","tarde","noche","diurno","admin"]
    _TURNO_HORAS       = {
        "mañana": ("06:00","14:00"),
        "tarde":  ("14:00","22:00"),
        "noche":  ("22:00","06:00"),
        "diurno": ("06:00","14:00"),
        "admin":  ("00:00","23:59"),
    }
    _PERMISOS_OPCIONES = ["reservas","pagos","voucher","reportes","configuracion","admin"]

    with tabs[0]:
        st.markdown("**Datos del negocio**")
        col_ng1, col_ng2 = st.columns(2)
        with col_ng1:
            nombre_n = st.text_input("Nombre del negocio",
                                      value=get_config("negocio_nombre", NEGOCIO))
            nit_n    = st.text_input("NIT / RUT",
                                      value=get_config("negocio_nit", NIT))
            dir_n    = st.text_input("Dirección en terminal",
                                      value=get_config("negocio_direccion", DIRECCION))
        with col_ng2:
            tel_n    = st.text_input("Teléfono",
                                      value=get_config("negocio_telefono", TELEFONO))
            email_n  = st.text_input("Email del negocio",
                                      value=get_config("negocio_email", ""),
                                      placeholder="contacto@jjgt.com.co")
            web_n    = st.text_input("Sitio web (opcional)",
                                      value=get_config("negocio_web", ""),
                                      placeholder="https://www.jjgt.com.co")
        if st.button("💾 Guardar datos negocio", type="primary"):
            for k, v in [
                ("negocio_nombre",   nombre_n),
                ("negocio_nit",      nit_n),
                ("negocio_direccion",dir_n),
                ("negocio_telefono", tel_n),
                ("negocio_email",    email_n),
                ("negocio_web",      web_n),
            ]:
                set_config(k, v)
            st.success("✅ Datos del negocio actualizados")
            _, sh_cfg = get_active_client()
            if sh_cfg:
                gs_sync_configuracion_pagos(sh_cfg)

    with tabs[1]:
        st.markdown("**Datos de plataformas de pago**")
        nequi_c  = st.text_input("Número Nequi", value=get_config("nequi_numero", NEQUI_NUM))
        davi_c   = st.text_input("Número Daviplata", value=get_config("daviplata_numero", DAVIPLATA_NUM))
        banco_c  = st.text_input("Cuenta bancaria", value=get_config("cuenta_bancaria", CUENTA_BANCO))
        mp_c     = st.text_input("Link MercadoPago", value=get_config("mp_link", MP_LINK))
        wa_c     = st.text_input("WhatsApp operador", value=get_config("whatsapp_op", WHATSAPP_OP))
        if st.button("💾 Guardar datos de pago", type="primary"):
            for k, v in [("nequi_numero",nequi_c),("daviplata_numero",davi_c),
                          ("cuenta_bancaria",banco_c),("mp_link",mp_c),("whatsapp_op",wa_c)]:
                set_config(k, v)
            st.success("✅ Datos de pago actualizados")
            _, sh_cfg2 = get_active_client()
            if sh_cfg2:
                gs_sync_configuracion_pagos(sh_cfg2)

    with tabs[2]:
        st.markdown("### 🛏️ Gestión de Tarifas")
        st.caption("Administra tarifas en **jjgt_pagos** (generales) y **jjgt_convenios** (con empresa). "
                   "Los cambios se sincronizan en ambas hojas cuando corresponde.")

        _tar_subtab_pagos, _tar_subtab_conv = st.tabs(["📋 Tarifas jjgt_pagos", "🤝 Tarifas jjgt_convenios"])

        # ── Sub-tab jjgt_pagos ────────────────────────────────────────────────
        with _tar_subtab_pagos:
            st.markdown("#### Tarifas en `jjgt_pagos` → hoja `Tarifas_Config`")
            tfs_gs = _gs_read_sheet("Tarifas_Config")

            # ── Tabla de tarifas actuales ──────────────────────────────────
            if tfs_gs:
                df_t = pd.DataFrame([{
                    "ID":            _gs_val(r, "ID_Tarifa"),
                    "Nombre":        _gs_val(r, "Nombre"),
                    "Empresa":       _gs_val(r, "Empresa_Nombre"),
                    "Descripción":   _gs_val(r, "Descripcion"),
                    "Precio/hora":   fmt_cop(_gs_float(r, "Precio_Hora_COP", 15000)),
                    "Desc 3h%":      _gs_val(r, "Desc_3h_Pct"),
                    "Desc 6h%":      _gs_val(r, "Desc_6h_Pct"),
                    "H.Ini.Espec":   _gs_val(r, "Hora_Ini_Espec"),
                    "H.Fin.Espec":   _gs_val(r, "Hora_Fin_Espec"),
                    "Festivo":       _gs_val(r, "Aplica_Festivo"),
                    "Activo":        _gs_val(r, "Activo"),
                } for r in tfs_gs])
                for _c in df_t.columns:
                    df_t[_c] = df_t[_c].astype(str)
                st.dataframe(df_t, use_container_width=True, hide_index=True)
            else:
                st.info("No hay tarifas en jjgt_pagos todavía.")

            st.divider()

            # ── CREAR nueva tarifa en jjgt_pagos ──────────────────────────
            with st.expander("➕ Crear nueva tarifa en jjgt_pagos", expanded=False):
                with st.form("form_tar_pagos_crear", clear_on_submit=True):
                    _tc1, _tc2 = st.columns(2)
                    with _tc1:
                        tp_nombre = st.text_input("Nombre tarifa *", placeholder="Tarifa Estándar")
                        tp_desc   = st.text_input("Descripción", placeholder="Tarifa por hora estándar")
                        tp_precio = st.number_input("Precio/hora COP", min_value=0, step=500, value=15000)
                        tp_d3     = st.number_input("Descuento 3h+ (%)", min_value=0.0, max_value=50.0, step=1.0)
                    with _tc2:
                        tp_d6     = st.number_input("Descuento 6h+ (%)", min_value=0.0, max_value=50.0, step=1.0)
                        tp_hini   = st.text_input("Hora inicio especial (opcional)", placeholder="22:00")
                        tp_hfin   = st.text_input("Hora fin especial (opcional)", placeholder="06:00")
                        tp_fest   = st.checkbox("Aplica en festivos")
                        tp_activo = st.checkbox("Activo", value=True)
                    if st.form_submit_button("💾 Crear tarifa en jjgt_pagos", type="primary", use_container_width=True):
                        if not tp_nombre.strip():
                            st.error("El nombre de la tarifa es obligatorio")
                        else:
                            _, _sh_tp = get_active_client()
                            if not _sh_tp:
                                st.error("Sin conexión a jjgt_pagos. Verifica Google Sheets.")
                            else:
                                _tar_id_p = f"TAR-{int(time.time())}"
                                _fila_tp = [
                                    _tar_id_p, tp_nombre.strip(), tp_desc.strip(),
                                    str(tp_precio), str(tp_d3), str(tp_d6),
                                    tp_hini.strip(), tp_hfin.strip(),
                                    "1" if tp_fest else "0",
                                    "1" if tp_activo else "0",
                                ]
                                _gs_append(_sh_tp, "Tarifas_Config", _fila_tp)
                                _gs_invalidate_cache("Tarifas_Config")
                                st.success(f"✅ Tarifa '{tp_nombre}' creada en jjgt_pagos · ID: {_tar_id_p}")
                                st.rerun()

            # ── MODIFICAR tarifa en jjgt_pagos ────────────────────────────
            with st.expander("✏️ Modificar tarifa en jjgt_pagos", expanded=False):
                tfs_mod = _gs_read_sheet("Tarifas_Config")
                if not tfs_mod:
                    st.info("No hay tarifas para modificar.")
                else:
                    _opts_mod_tp = {
                        f"{_gs_val(r,'ID_Tarifa')} — {_gs_val(r,'Nombre')}": r
                        for r in tfs_mod if _gs_val(r, "ID_Tarifa")
                    }
                    if not _opts_mod_tp:
                        st.info("No hay tarifas con ID válido para modificar.")
                    else:
                        _sel_mod_tp_label = st.selectbox(
                            "Selecciona tarifa a modificar", list(_opts_mod_tp.keys()), key="mod_tar_pagos_sel"
                        )
                        _row_mod_tp = _opts_mod_tp.get(_sel_mod_tp_label) if _sel_mod_tp_label else None
                        if _row_mod_tp is None:
                            st.info("Selecciona una tarifa válida.")
                        else:
                            _tar_id_mod_tp = _gs_val(_row_mod_tp, "ID_Tarifa")
                            with st.form("form_tar_pagos_mod", clear_on_submit=False):
                                _mc1, _mc2 = st.columns(2)
                                with _mc1:
                                    tp_m_nombre = st.text_input("Nombre tarifa *",
                                        value=_gs_val(_row_mod_tp, "Nombre"))
                                    tp_m_desc   = st.text_input("Descripción",
                                        value=_gs_val(_row_mod_tp, "Descripcion"))
                                    tp_m_precio = st.number_input("Precio/hora COP", min_value=0, step=500,
                                        value=int(_gs_float(_row_mod_tp, "Precio_Hora_COP", 15000)))
                                    tp_m_d3     = st.number_input("Descuento 3h+ (%)", min_value=0.0,
                                        max_value=50.0, step=1.0,
                                        value=float(_gs_float(_row_mod_tp, "Desc_3h_Pct", 0)))
                                with _mc2:
                                    tp_m_d6     = st.number_input("Descuento 6h+ (%)", min_value=0.0,
                                        max_value=50.0, step=1.0,
                                        value=float(_gs_float(_row_mod_tp, "Desc_6h_Pct", 0)))
                                    tp_m_hini   = st.text_input("Hora inicio especial",
                                        value=_gs_val(_row_mod_tp, "Hora_Ini_Espec"))
                                    tp_m_hfin   = st.text_input("Hora fin especial",
                                        value=_gs_val(_row_mod_tp, "Hora_Fin_Espec"))
                                    tp_m_fest   = st.checkbox("Aplica en festivos",
                                        value=_gs_val(_row_mod_tp, "Aplica_Festivo") == "1")
                                    tp_m_activo = st.checkbox("Activo",
                                        value=_gs_val(_row_mod_tp, "Activo") in ("1","True","true"))
                                if st.form_submit_button("💾 Guardar cambios en jjgt_pagos", type="primary", use_container_width=True):
                                    if not tp_m_nombre.strip():
                                        st.error("El nombre es obligatorio")
                                    else:
                                        _, _sh_tp_m = get_active_client()
                                        if not _sh_tp_m:
                                            st.error("Sin conexión a jjgt_pagos.")
                                        else:
                                            try:
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "nombre", tp_m_nombre.strip())
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "descripcion", tp_m_desc.strip())
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "precio_hora_cop", str(tp_m_precio))
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "desc_3h_pct", str(tp_m_d3))
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "desc_6h_pct", str(tp_m_d6))
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "hora_ini_espec", tp_m_hini.strip())
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "hora_fin_espec", tp_m_hfin.strip())
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "aplica_festivo", "1" if tp_m_fest else "0")
                                                _pg_update_col("tarifas_config", "id_tarifa", _tar_id_mod_tp,
                                                               "activo", "1" if tp_m_activo else "0")
                                            except Exception:
                                                pass
                                            _gs_invalidate_cache("Tarifas_Config")
                                            st.success(f"✅ Tarifa '{tp_m_nombre}' actualizada en jjgt_pagos.")
                                            st.rerun()

            # ── ELIMINAR tarifa en jjgt_pagos ─────────────────────────────
            with st.expander("🗑️ Eliminar tarifa en jjgt_pagos", expanded=False):
                tfs_del = _gs_read_sheet("Tarifas_Config")
                if not tfs_del:
                    st.info("No hay tarifas para eliminar.")
                else:
                    _opts_del_tp = {
                        f"{_gs_val(r,'ID_Tarifa')} — {_gs_val(r,'Nombre')}": _gs_val(r,"ID_Tarifa")
                        for r in tfs_del if _gs_val(r,"ID_Tarifa")
                    }
                    if not _opts_del_tp:
                        st.info("No hay tarifas con ID válido para eliminar.")
                    else:
                        _col_del1, _col_del2 = st.columns([2,1])
                        with _col_del1:
                            _sel_del_tp = st.selectbox("Tarifa a eliminar", list(_opts_del_tp.keys()), key="del_tar_pagos_sel")
                        with _col_del2:
                            _pin_del_tp = st.text_input("PIN admin", type="password", key="del_tar_pagos_pin")
                        _id_del_tp = _opts_del_tp.get(_sel_del_tp) if _sel_del_tp else None
                        if _id_del_tp is None:
                            st.info("Selecciona una tarifa válida.")
                        else:
                            if st.button("🗑️ Eliminar tarifa de jjgt_pagos", type="primary", use_container_width=True, key="btn_del_tar_pagos"):
                                if not verificar_pin(_pin_del_tp, "admin"):
                                    st.error("PIN admin incorrecto")
                                else:
                                    try:
                                        _pg_delete_row("tarifas_config", "id_tarifa", _id_del_tp)
                                    except Exception:
                                        pass
                                    _gs_invalidate_cache("Tarifas_Config")
                                    st.success(f"✅ Tarifa '{_sel_del_tp}' eliminada de jjgt_pagos.")
                                    st.rerun()

        # ── Sub-tab jjgt_convenios ────────────────────────────────────────────
        with _tar_subtab_conv:
            st.markdown("#### Tarifas en `jjgt_convenios` → hoja `Tarifas_Config`")
            tfs_cv = _gs_read_sheet_conv("Tarifas_Config", force=False)

            # ── Tabla de tarifas convenio actuales ────────────────────────
            if tfs_cv:
                df_tc = pd.DataFrame([{
                    "ID":          _gs_val(r, "ID_Tarifa"),
                    "Nombre":      _gs_val(r, "Nombre"),
                    "Empresa":     _gs_val(r, "Empresa_Nombre"),
                    "Descripción": _gs_val(r, "Descripcion"),
                    "Precio/hora": fmt_cop(_gs_float(r, "Precio_Hora_COP", 12000)),
                    "Desc 3h%":    _gs_val(r, "Desc_3h_Pct"),
                    "Desc 6h%":    _gs_val(r, "Desc_6h_Pct"),
                    "Festivo":     _gs_val(r, "Aplica_Festivo"),
                    "Activo":      _gs_val(r, "Activo"),
                } for r in tfs_cv])
                for _c in df_tc.columns:
                    df_tc[_c] = df_tc[_c].astype(str)
                st.dataframe(df_tc, use_container_width=True, hide_index=True)
            else:
                st.info("No hay tarifas en jjgt_convenios todavía.")

            st.divider()

            # Selector de empresa para crear/modificar tarifas de convenio
            _emps_tc = get_empresas_convenio()
            _opts_tc_emp = {"(Genérica — sin empresa específica)": ("", "")}
            for _e in _emps_tc:
                _opts_tc_emp[f"{_e.get('Nombre_Empresa','—')} ({_e.get('Nit_Empresa','')})"] = (
                    _e.get("Id_Empresa", ""), _e.get("Nombre_Empresa", ""))

            # ── CREAR tarifa en jjgt_convenios ────────────────────────────
            with st.expander("➕ Crear nueva tarifa en jjgt_convenios", expanded=False):
                _sel_tc_emp_crear = st.selectbox(
                    "Empresa a la que aplica", list(_opts_tc_emp.keys()), key="tar_conv_emp_crear"
                )
                _tc_emp_id_c, _tc_emp_nom_c = _opts_tc_emp.get(_sel_tc_emp_crear, ("", "")) if _sel_tc_emp_crear else ("", "")
                with st.form("form_tar_conv_crear", clear_on_submit=True):
                    _tcc1, _tcc2 = st.columns(2)
                    with _tcc1:
                        tc_nombre = st.text_input("Nombre tarifa *", placeholder="Corporativo Noche")
                        tc_desc   = st.text_input("Descripción", placeholder="Tarifa nocturna empresarial")
                        tc_precio = st.number_input("Precio/hora COP", min_value=0, step=500, value=12000)
                        tc_d3     = st.number_input("Descuento 3h+ (%)", min_value=0.0, max_value=50.0, step=1.0)
                    with _tcc2:
                        tc_d6     = st.number_input("Descuento 6h+ (%)", min_value=0.0, max_value=50.0, step=1.0)
                        tc_hini   = st.text_input("Hora inicio especial", placeholder="22:00")
                        tc_hfin   = st.text_input("Hora fin especial", placeholder="06:00")
                        tc_fest   = st.checkbox("Aplica en festivos")
                        tc_activo = st.checkbox("Activo", value=True)
                    if st.form_submit_button("💾 Crear tarifa en jjgt_convenios", type="primary", use_container_width=True):
                        if not tc_nombre.strip():
                            st.error("El nombre es obligatorio")
                        else:
                            _, _sh_tc_c = get_active_client_convenios()
                            if not _sh_tc_c:
                                st.error("Sin conexión a jjgt_convenios. Verifica el ID del spreadsheet.")
                            else:
                                _tar_id_tc = f"CONV-{int(time.time())}"
                                _nombre_tc = f"[Convenio] {tc_nombre.strip()}"
                                _fila_tc = [
                                    _tar_id_tc, _tc_emp_id_c, _tc_emp_nom_c,
                                    _nombre_tc, tc_desc.strip(), str(tc_precio),
                                    str(tc_d3), str(tc_d6),
                                    tc_hini.strip(), tc_hfin.strip(),
                                    "1" if tc_fest else "0",
                                    "1" if tc_activo else "0",
                                ]
                                _gs_append(_sh_tc_c, "Tarifas_Config", _fila_tc)
                                _gs_invalidate_cache_conv("Tarifas_Config")
                                _emp_lbl_tc = f" para **{_tc_emp_nom_c}**" if _tc_emp_id_c else " (genérica)"
                                st.success(f"✅ Tarifa '{_nombre_tc}'{_emp_lbl_tc} creada en jjgt_convenios.")
                                st.rerun()

            # ── MODIFICAR tarifa en jjgt_convenios ────────────────────────
            with st.expander("✏️ Modificar tarifa en jjgt_convenios", expanded=False):
                tfs_cv_mod = _gs_read_sheet_conv("Tarifas_Config", force=False)
                if not tfs_cv_mod:
                    st.info("No hay tarifas de convenio para modificar.")
                else:
                    _opts_mod_tc = {
                        f"{_gs_val(r,'ID_Tarifa')} — {_gs_val(r,'Nombre')} [{_gs_val(r,'Empresa_Nombre') or 'Genérica'}]": r
                        for r in tfs_cv_mod if _gs_val(r, "ID_Tarifa")
                    }
                    if not _opts_mod_tc:
                        st.info("No hay tarifas con ID válido para modificar.")
                    else:
                        _sel_mod_tc_label = st.selectbox(
                            "Tarifa a modificar", list(_opts_mod_tc.keys()), key="mod_tar_conv_sel"
                        )
                        _row_mod_tc = _opts_mod_tc.get(_sel_mod_tc_label) if _sel_mod_tc_label else None
                        if _row_mod_tc is None:
                            st.info("Selecciona una tarifa válida.")
                        else:
                            _tar_id_mod_tc = _gs_val(_row_mod_tc, "ID_Tarifa")
                            with st.form("form_tar_conv_mod", clear_on_submit=False):
                                _mcc1, _mcc2 = st.columns(2)
                                with _mcc1:
                                    tc_m_nombre = st.text_input("Nombre tarifa *",
                                        value=_gs_val(_row_mod_tc, "Nombre"))
                                    tc_m_desc   = st.text_input("Descripción",
                                        value=_gs_val(_row_mod_tc, "Descripcion"))
                                    tc_m_precio = st.number_input("Precio/hora COP", min_value=0, step=500,
                                        value=int(_gs_float(_row_mod_tc, "Precio_Hora_COP", 12000)))
                                    tc_m_d3     = st.number_input("Descuento 3h+ (%)", min_value=0.0,
                                        max_value=50.0, step=1.0,
                                        value=float(_gs_float(_row_mod_tc, "Desc_3h_Pct", 0)))
                                with _mcc2:
                                    tc_m_d6     = st.number_input("Descuento 6h+ (%)", min_value=0.0,
                                        max_value=50.0, step=1.0,
                                        value=float(_gs_float(_row_mod_tc, "Desc_6h_Pct", 0)))
                                    tc_m_hini   = st.text_input("Hora inicio especial",
                                        value=_gs_val(_row_mod_tc, "Hora_Ini_Espec"))
                                    tc_m_hfin   = st.text_input("Hora fin especial",
                                        value=_gs_val(_row_mod_tc, "Hora_Fin_Espec"))
                                    tc_m_fest   = st.checkbox("Aplica en festivos",
                                        value=_gs_val(_row_mod_tc, "Aplica_Festivo") == "1")
                                    tc_m_activo = st.checkbox("Activo",
                                        value=_gs_val(_row_mod_tc, "Activo") in ("1","True","true"))
                                if st.form_submit_button("💾 Guardar cambios en jjgt_convenios", type="primary", use_container_width=True):
                                    if not tc_m_nombre.strip():
                                        st.error("El nombre es obligatorio")
                                    else:
                                        try:
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Nombre", tc_m_nombre.strip())
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Descripcion", tc_m_desc.strip())
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Precio_Hora_COP", str(tc_m_precio))
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Desc_3h_Pct", str(tc_m_d3))
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Desc_6h_Pct", str(tc_m_d6))
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Hora_Ini_Espec", tc_m_hini.strip())
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Hora_Fin_Espec", tc_m_hfin.strip())
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Aplica_Festivo", "1" if tc_m_fest else "0")
                                            _pg_update_col_conv("Tarifas_Config", "ID_Tarifa", _tar_id_mod_tc,
                                                                "Activo", "1" if tc_m_activo else "0")
                                        except Exception:
                                            pass
                                        _gs_invalidate_cache_conv("Tarifas_Config")
                                        st.success(f"✅ Tarifa '{tc_m_nombre}' actualizada en jjgt_convenios.")
                                        st.rerun()

            # ── ELIMINAR tarifa en jjgt_convenios ─────────────────────────
            with st.expander("🗑️ Eliminar tarifa en jjgt_convenios", expanded=False):
                tfs_cv_del = _gs_read_sheet_conv("Tarifas_Config", force=False)
                if not tfs_cv_del:
                    st.info("No hay tarifas de convenio para eliminar.")
                else:
                    _opts_del_tc = {
                        f"{_gs_val(r,'ID_Tarifa')} — {_gs_val(r,'Nombre')} [{_gs_val(r,'Empresa_Nombre') or 'Genérica'}]":
                            _gs_val(r, "ID_Tarifa")
                        for r in tfs_cv_del if _gs_val(r, "ID_Tarifa")
                    }
                    if not _opts_del_tc:
                        st.info("No hay tarifas con ID válido para eliminar.")
                    else:
                        _col_del_tc1, _col_del_tc2 = st.columns([2, 1])
                        with _col_del_tc1:
                            _sel_del_tc = st.selectbox("Tarifa a eliminar", list(_opts_del_tc.keys()), key="del_tar_conv_sel")
                        with _col_del_tc2:
                            _pin_del_tc = st.text_input("PIN admin", type="password", key="del_tar_conv_pin")
                        _id_del_tc = _opts_del_tc.get(_sel_del_tc) if _sel_del_tc else None
                        if _id_del_tc is None:
                            st.info("Selecciona una tarifa válida.")
                        else:
                            if st.button("🗑️ Eliminar tarifa de jjgt_convenios", type="primary", use_container_width=True, key="btn_del_tar_conv"):
                                if not verificar_pin(_pin_del_tc, "admin"):
                                    st.error("PIN admin incorrecto")
                                else:
                                    try:
                                        _pg_delete_row_conv("Tarifas_Config", "ID_Tarifa", _id_del_tc)
                                    except Exception:
                                        pass
                                    _gs_invalidate_cache_conv("Tarifas_Config")
                                    st.success(f"✅ Tarifa '{_sel_del_tc}' eliminada de jjgt_convenios.")
                                    st.rerun()

    with tabs[3]:
        st.markdown("#### 👥 Gestión de Operadores")
        st.info("ℹ️ Los PINs se gestionan en `st.secrets['operadores_pins']` o al crear el operador.")

        ops_gs = _gs_read_sheet("Operadores")

        TURNOS            = _TURNOS
        TURNO_HORAS       = _TURNO_HORAS
        PERMISOS_OPCIONES = _PERMISOS_OPCIONES

        if ops_gs:
            st.markdown("**Operadores registrados:**")
            for op_row in ops_gs:
                op_id     = _gs_val(op_row, "ID_Operador", "0")
                op_nombre = _gs_val(op_row, "Nombre")
                op_rol    = _gs_val(op_row, "Rol", "cajero")
                op_turno  = _gs_val(op_row, "Turno", "diurno")
                op_hi     = _gs_val(op_row, "Hora_Ini_Turno", "06:00")
                op_hf     = _gs_val(op_row, "Hora_Fin_Turno", "14:00")
                op_perms  = _gs_val(op_row, "Permisos", "reservas,pagos,voucher")
                op_activo = _gs_val(op_row, "Activo", "1") in ("1","True","true")

                if not op_nombre:
                    continue

                estado_badge = "🟢" if op_activo else "🔴"
                with st.expander(f"{estado_badge} {op_nombre} · {op_rol.capitalize()} · Turno {op_turno or '-'}"):
                    col_a, col_b = st.columns(2)
                    with col_a:
                        new_nombre = st.text_input("Nombre", value=op_nombre, key=f"op_nom_{op_id}")
                        new_rol    = st.selectbox("Rol", ["admin","cajero","supervisor"],
                                                  index=["admin","cajero","supervisor"].index(op_rol)
                                                  if op_rol in ["admin","cajero","supervisor"] else 1,
                                                  key=f"op_rol_{op_id}")
                        new_turno  = st.selectbox("Turno", TURNOS,
                                                   index=TURNOS.index(op_turno) if op_turno in TURNOS else 0,
                                                   key=f"op_trn_{op_id}")
                        hi_def, hf_def = TURNO_HORAS.get(new_turno, ("06:00","14:00"))
                        new_hi = st.text_input("Hora inicio turno", value=op_hi or hi_def, key=f"op_hi_{op_id}")
                        new_hf = st.text_input("Hora fin turno",    value=op_hf or hf_def, key=f"op_hf_{op_id}")
                    with col_b:
                        perms_list = [p.strip() for p in op_perms.split(",")]
                        new_perms  = st.multiselect("Permisos", PERMISOS_OPCIONES,
                                                     default=[p for p in perms_list if p in PERMISOS_OPCIONES],
                                                     key=f"op_per_{op_id}")
                        new_activo = st.checkbox("Activo", value=op_activo, key=f"op_act_{op_id}")
                        nuevo_pin  = st.text_input("Nuevo PIN (dejar vacío = no cambiar)",
                                                    type="password", max_chars=8, key=f"op_pin_{op_id}")

                    cc1, cc2 = st.columns(2)
                    with cc1:
                        if st.button("💾 Guardar cambios", key=f"op_save_{op_id}", type="primary",
                                      use_container_width=True):
                            perms_str = ",".join(new_perms) if new_perms else "reservas,pagos,voucher"
                            if nuevo_pin and len(nuevo_pin) < 4:
                                st.error("El PIN debe tener al menos 4 dígitos")
                            else:
                                try:
                                    for col_name, col_val in [
                                        ("Nombre", new_nombre), ("Rol", new_rol),
                                        ("Turno", new_turno), ("Hora_Ini_Turno", new_hi),
                                        ("Hora_Fin_Turno", new_hf), ("Permisos", perms_str),
                                        ("Activo", str(int(new_activo))),
                                    ]:
                                        _pg_update_col("Operadores","ID_Operador",str(op_id),col_name,col_val)
                                except Exception:
                                    pass
                                _gs_invalidate_cache("Operadores")
                                # Actualizar hash en session_state si cambió PIN
                                if nuevo_pin and len(nuevo_pin) >= 4:
                                    pin_hashes = st.session_state.get("_op_pin_hashes", {})
                                    pin_hashes[new_nombre] = _hash_pin(nuevo_pin)
                                    st.session_state["_op_pin_hashes"] = pin_hashes
                                    st.success("✅ Operador actualizado con nuevo PIN")
                                else:
                                    st.success("✅ Operador actualizado")
                                _gs_invalidate_cache("Operadores")
                                st.rerun()
                    with cc2:
                        if op_id != "1" and op_nombre != "Admin JJGT":
                            if st.button("🗑️ Eliminar", key=f"op_del_{op_id}",
                                          use_container_width=True):
                                try:
                                    _pg_delete_row("Operadores","Nombre",op_nombre)
                                    _gs_invalidate_cache("Operadores")
                                except Exception: pass
                                st.warning(f"Operador {op_nombre} eliminado")
                                st.rerun()

        st.divider()
        st.markdown("#### ➕ Crear nuevo operador")
        with st.form("form_nuevo_op", clear_on_submit=True):
            fcol1, fcol2 = st.columns(2)
            with fcol1:
                f_nombre = st.text_input("Nombre del operador *")
                f_pin    = st.text_input("PIN (4-8 dígitos) *", type="password", max_chars=8)
                f_pin2   = st.text_input("Confirmar PIN *",     type="password", max_chars=8)
                f_rol    = st.selectbox("Rol", ["cajero","supervisor","admin"])
            with fcol2:
                f_turno  = st.selectbox("Turno", TURNOS)
                hi_def2, hf_def2 = TURNO_HORAS.get(f_turno, ("06:00","14:00"))
                f_hi     = st.text_input("Hora inicio", value=hi_def2)
                f_hf     = st.text_input("Hora fin",    value=hf_def2)
                f_perms  = st.multiselect("Permisos", PERMISOS_OPCIONES,
                                           default=["reservas","pagos","voucher"])
            submitted = st.form_submit_button("➕ Crear operador", type="primary",
                                               use_container_width=True)
            if submitted:
                if not f_nombre:
                    st.error("El nombre es obligatorio")
                elif len(f_pin) < 4:
                    st.error("El PIN debe tener al menos 4 dígitos")
                elif f_pin != f_pin2:
                    st.error("Los PINs no coinciden")
                else:
                    perms_str2 = ",".join(f_perms) if f_perms else "reservas,pagos,voucher"
                    new_op_id  = str(int(time.time()))  # ID único basado en timestamp
                    _, sh_opnew = get_active_client()
                    if sh_opnew:
                        _gs_append(sh_opnew, "Operadores", [
                            new_op_id, f_nombre, f_rol, f_turno,
                            f_hi, f_hf, perms_str2, "1"
                        ])
                        _gs_invalidate_cache("Operadores")
                    # Guardar hash del PIN en session_state
                    pin_hashes = st.session_state.get("_op_pin_hashes", {})
                    pin_hashes[f_nombre] = _hash_pin(f_pin)
                    st.session_state["_op_pin_hashes"] = pin_hashes
                    st.success(f"✅ Operador **{f_nombre}** creado en turno **{f_turno}**")
                    st.rerun()


    with tabs[4]:
        st.markdown("#### 🔗 Credenciales Google Sheets")
        st.markdown("""
        La aplicación se conecta al archivo `jjgt_pagos` en Google Sheets para sincronización automática.
        """)

        # ── Entorno de ejecución ───────────────────────────────────────────
        # _IS_CLOUD se definió a nivel de módulo al inicio
        entorno = "☁️ Streamlit Cloud" if _IS_CLOUD else "💻 Ejecución local"
        st.info(f"**Entorno detectado:** {entorno}")

        # ── Estado actual de la conexión ───────────────────────────────────
        _, sh_status = _get_module_level_client()
        if sh_status:
            sid_actual = get_config("drive_spreadsheet_id","")
            drive_url_conf = f"https://docs.google.com/spreadsheets/d/{sid_actual}/edit"
            st.success(f"✅ Conectado a **{sh_status.title}**")
            #st.markdown(f'🔗 <a href="{drive_url_conf}" target="_blank">Abrir {sh_status.title} en Google Sheets</a>',
            #            unsafe_allow_html=True)
        else:
            st.warning("⚠️ Sin conexión activa a Google Sheets")

        st.divider()

        # ── Configurar Spreadsheet ID ──────────────────────────────────────
        st.markdown("**ID del Spreadsheet `jjgt_pagos`**")
        sid_input = st.text_input("Spreadsheet ID",
                                   value=get_config("drive_spreadsheet_id",""),
                                   placeholder="1abc...xyz",
                                   help="El ID aparece en la URL del archivo: docs.google.com/spreadsheets/d/**ID**/edit")
        if st.button("💾 Guardar Spreadsheet ID", use_container_width=True):
            if sid_input.strip():
                set_config("drive_spreadsheet_id", sid_input.strip())
                global _gs_module_spreadsheet  # noqa
                _gs_module_spreadsheet = None
                st.success("✅ Spreadsheet ID guardado. Reconectando...")
                st.rerun()
            else:
                st.error("El ID no puede estar vacío")

        st.divider()

        # ── Instrucciones según entorno ────────────────────────────────────
        if _IS_CLOUD:
            st.markdown("""
            **📋 Configuración para Streamlit Cloud:**

            En la app de Streamlit Cloud → **Settings → Secrets**, agrega:
            ```toml
            [sheetsemp]
            credentials_sheet = '''
            {
              "type": "service_account",
              "project_id": "...",
              "private_key_id": "...",
              "private_key": "-----BEGIN RSA PRIVATE KEY-----\n...\n-----END RSA PRIVATE KEY-----\n",
              "client_email": "...@....iam.gserviceaccount.com",
              ...
            }
            '''
            spreadsheet_id = "ID_DE_TU_SPREADSHEET"
            ```
            """)
        else:
            st.markdown("""
            **📋 Configuración para ejecución local:**

            Crea el archivo `.streamlit/secrets.toml` con:
            ```toml
            [sheetsemp]
            credentials_sheet = '''{ ... JSON de tu Service Account ... }'''
            spreadsheet_id = "ID_DE_TU_SPREADSHEET"
            ```

            O simplemente coloca `credentials.json` en el directorio raíz del proyecto.
            """)

            # Subir credentials.json localmente
            st.markdown("**O sube tu archivo credentials.json directamente:**")
            creds_file = st.file_uploader("credentials.json", type=["json"])
            if creds_file:
                creds_path = "credentials.json"
                with open(creds_path, "wb") as f:
                    f.write(creds_file.read())
                set_config("drive_credentials_path", creds_path)
                global _gs_module_client  # noqa
                _gs_module_client = None
                _gs_module_spreadsheet = None  # noqa
                try:
                    st.session_state._gs_client = None
                except Exception:
                    pass
                st.success("✅ credentials.json guardado. Reconectando...")
                st.rerun()



# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO CONVENIOS — funciones principales
# ══════════════════════════════════════════════════════════════════════════════


    # ── Tab 5: CONVENIOS ───────────────────────────────────────────────────────
    with tabs[5]:
        #st.markdown("#### 🤝 Configuración de Convenios — `jjgt_convenios`")
        #st.info("Configura el archivo Google Sheets `jjgt_convenios` y gestiona las empresas con convenio.")

        #st.markdown("**ID del Spreadsheet `jjgt_convenios`**")
        sid_conv_cfg = get_config("convenios_spreadsheet_id", "")
        #sid_conv_input = st.text_input("Spreadsheet ID de jjgt_convenios", value=sid_conv_cfg,
        #    placeholder="1abc...xyz", key="cfg_sid_convenios",
        #    help="ID del archivo jjgt_convenios en Google Sheets")
        #if st.button("Guardar ID jjgt_convenios", type="primary", use_container_width=True, key="btn_save_conv_id"):
        #    if sid_conv_input.strip():
        #        set_config("convenios_spreadsheet_id", sid_conv_input.strip())
        #        st.success("ID de jjgt_convenios guardado.")
        #        st.rerun()
        #    else:
        #        st.error("El ID no puede estar vacío")

        st.divider()
        st.markdown("#### Empresas con convenio")
        empresas_cfg2 = get_empresas_convenio()
        if empresas_cfg2:
            _df_emp_cfg = pd.DataFrame([{
                "ID":        str(e.get("Id_Empresa","")),
                "Empresa":   str(e.get("Nombre_Empresa","")),
                "NIT":       str(e.get("Nit_Empresa","")),
                "Tipo Conv.":str(e.get("Tipo_Convenio","")),
            } for e in empresas_cfg2])
            st.dataframe(_df_emp_cfg, use_container_width=True, hide_index=True)
            col_de1c, col_de2c = st.columns([2,1])
            with col_de1c:
                emp_del_opts2 = {f"{e.get('Nombre_Empresa','')} ({e.get('Nit_Empresa','')})": e.get("Id_Empresa","") for e in empresas_cfg2}
                emp_del_sel2 = st.selectbox("Empresa a eliminar", list(emp_del_opts2.keys()), key="del_emp_cfg_sel")
                emp_del_id2 = emp_del_opts2[emp_del_sel2]
            with col_de2c:
                pin_del_emp2 = st.text_input("PIN admin", type="password", key="del_emp_cfg_pin")
            if st.button("Eliminar empresa", use_container_width=True, key="btn_del_emp_cfg"):
                if not verificar_pin(pin_del_emp2, "admin"):
                    st.error("PIN incorrecto")
                else:
                    try:
                        _pg_delete_row_conv("Empresa","Id_Empresa",emp_del_id2)
                    except Exception as e2: st.error(f"Error: {e2}")
                    st.success("Empresa eliminada"); st.rerun()
        else:
            st.info("No hay empresas registradas.")

        st.divider()
        st.markdown("#### Crear empresa con convenio")
        with st.form("form_empresa_conv_cfg", clear_on_submit=True):
            ec1c, ec2c = st.columns(2)
            with ec1c:
                e_nomb = st.text_input("Nombre empresa *")
                e_nitc = st.text_input("NIT *", placeholder="900.123.456-7")
                e_tipo_ec = st.selectbox("Tipo empresa", ["Privada","Publica","Mixta","ONG"])
                e_emailc = st.text_input("Email empresa")
            with ec2c:
                e_telc = st.text_input("Telefono empresa")
                e_tipo_cc = st.selectbox("Tipo convenio", ["Credito","Descuento","Prepago","Corporativo"])
                e_obsc = st.text_area("Observaciones", height=80)
            if st.form_submit_button("Guardar empresa", type="primary", use_container_width=True):
                if not e_nomb.strip() or not e_nitc.strip():
                    st.error("Nombre y NIT son obligatorios")
                else:
                        _, sh_conv_save2 = get_active_client_convenios()
                    #if not sh_conv_save2:
                    #    st.error("Configura primero el ID de jjgt_convenios")
                    #else:
                        emp_id2 = f"EMP-{int(time.time())}"
                        ok2 = gs_escribir_empresa(sh_conv_save2, {
                            "id_empresa": emp_id2, "nombre_empresa": e_nomb.strip(),
                            "nit_empresa": e_nitc.strip(), "tipo_empresa": e_tipo_ec,
                            "email_empresa": e_emailc.strip(), "telefono_empresa": e_telc.strip(),
                            "tipo_convenio": e_tipo_cc, "observaciones": e_obsc.strip(),
                            "activo": "1", "creado_en": ahora_col().isoformat(),
                        })
                        if ok2: st.success(f"Empresa '{e_nomb}' guardada · ID: {emp_id2}")
                        else: st.error("Error al guardar.")

        st.divider()
        st.markdown("#### 💰 Configuración de pagos por empresa")
        st.caption("Selecciona la empresa y define sus parámetros de convenio. "
                   "Se guarda en Configuracion_Pagos de jjgt_convenios con una fila por clave.")

        _, sh_cfg_cv = get_active_client_convenios()
        empresas_cfg_disp = get_empresas_convenio()

        if not empresas_cfg_disp:
            st.info("Primero crea al menos una empresa en la sección anterior.")
        else:
            # Selector de empresa (fuera del form para cargar sus valores actuales)
            opts_emp_cfg = {
                f"{e.get('Nombre_Empresa','—')} | NIT: {e.get('Nit_Empresa','')}": e
                for e in empresas_cfg_disp
            }
            emp_cfg_label = st.selectbox(
                "Empresa a configurar", list(opts_emp_cfg.keys()), key="cfg_emp_cfg_sel"
            )
            emp_cfg_sel = opts_emp_cfg[emp_cfg_label]
            emp_cfg_id   = emp_cfg_sel.get("Id_Empresa", "")
            emp_cfg_nomb = emp_cfg_sel.get("Nombre_Empresa", "—")

            # Cargar valores actuales de esa empresa
            cfg_actual = get_config_empresa_convenio(sh_cfg_cv, emp_cfg_id)

            with st.form("form_cfg_pagos_conv", clear_on_submit=False):
                st.markdown(f"**Configurando:** {emp_cfg_nomb}")
                col_cv1c, col_cv2c = st.columns(2)
                with col_cv1c:
                    conv_dias = st.number_input(
                        "Días de crédito (plazo cobro)", min_value=1, max_value=90,
                        value=int(cfg_actual["convenios_dias_credito"]),
                        key="cfg_cv_dias",
                        help="Días que tiene la empresa para pagar desde la reserva")
                    conv_hmin = st.number_input(
                        "Horas mínimas por reserva", min_value=0.5, max_value=18.0, step=0.5,
                        value=float(cfg_actual["convenios_horas_min"]),
                        key="cfg_cv_hmin",
                        help="Mínimo de horas que debe reservar un empleado de esta empresa")
                with col_cv2c:
                    conv_desc = st.number_input(
                        "Descuento adicional (%)", min_value=0.0, max_value=50.0, step=1.0,
                        value=float(cfg_actual["convenios_descuento_pct"]),
                        key="cfg_cv_desc",
                        help="% de descuento adicional sobre el precio base (se acumula con desc. por horas)")
                    conv_maxp = st.number_input(
                        "Máx. reservas pendientes", min_value=1, max_value=100,
                        value=int(cfg_actual["convenios_max_pendientes"]),
                        key="cfg_cv_maxp",
                        help="Máximo de reservas pendientes de cobro que puede tener la empresa")
                if st.form_submit_button("💾 Guardar configuración de esta empresa",
                                         type="primary", use_container_width=True):
                    # 1. Guardar valores globales en jjgt_pagos
                    for k, v in [
                        ("convenios_dias_credito",  str(conv_dias)),
                        ("convenios_horas_min",      str(conv_hmin)),
                        ("convenios_descuento_pct",  str(conv_desc)),
                        ("convenios_max_pendientes", str(conv_maxp)),
                    ]:
                        set_config(k, v)
                    # 2. Guardar por empresa en Configuracion_Pagos de jjgt_convenios
                    if sh_cfg_cv:
                        ok_cfg = upsert_config_empresa_convenio(
                            sh_cfg_cv, emp_cfg_id, emp_cfg_nomb,
                            {
                                "convenios_dias_credito":   conv_dias,
                                "convenios_horas_min":      conv_hmin,
                                "convenios_descuento_pct":  conv_desc,
                                "convenios_max_pendientes": conv_maxp,
                            }
                        )
                        if ok_cfg:
                            st.success(
                                f"✅ Configuración guardada para **{emp_cfg_nomb}** "
                                f"en jjgt_convenios · Configuracion_Pagos")
                        else:
                            st.error("❌ Error guardando en jjgt_convenios. "
                                     "Verifica la conexión.")
                    else:
                        st.warning("⚠️ jjgt_convenios no configurado — "
                                   "solo se guardó el valor global.")

            # Vista consolidada de configuración de todas las empresas
            with st.expander("🔍 Ver configuración de todas las empresas", expanded=False):
                _, sh_vcfg = get_active_client_convenios()
                rows_ver = []
                for emp_v in empresas_cfg_disp:
                    cfg_v = get_config_empresa_convenio(sh_vcfg, emp_v.get("Id_Empresa",""))
                    rows_ver.append({
                        "Empresa":      emp_v.get("Nombre_Empresa","—"),
                        "Dias credito": int(cfg_v["convenios_dias_credito"]),
                        "Horas min":    cfg_v["convenios_horas_min"],
                        "Descuento %":  cfg_v["convenios_descuento_pct"],
                        "Max pendient": int(cfg_v["convenios_max_pendientes"]),
                    })
                _df_rows_ver = pd.DataFrame(rows_ver)
                for _c in ["Empresa","Horas min","Descuento %"]:
                    if _c in _df_rows_ver.columns: _df_rows_ver[_c] = _df_rows_ver[_c].astype(str)
                st.dataframe(_df_rows_ver, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("#### 🛏️ Tarifa especial para convenios")
        st.caption("Cada tarifa se asocia a una empresa. Se guarda en Tarifas_Config de "
                   "jjgt_pagos (sin empresa) y en jjgt_convenios (con Id_Empresa).")
        # Selector empresa para la tarifa — fuera del form para ser legible al submit
        _emps_tar  = get_empresas_convenio()
        _opts_tar  = {"(Genérica — aplica a todas las empresas)": ("","")}
        for _e in _emps_tar:
            _opts_tar[f"{_e.get('Nombre_Empresa','—')} ({_e.get('Nit_Empresa','')})"] = (
                _e.get("Id_Empresa",""), _e.get("Nombre_Empresa",""))
        _sel_tar = st.selectbox("Empresa a la que aplica la tarifa",
                                list(_opts_tar.keys()), key="cfg_tar_conv_emp_sel")
        _tar_emp_id, _tar_emp_nom = _opts_tar[_sel_tar]

        with st.form("form_tarifa_conv_cfg", clear_on_submit=True):
            tc1c, tc2c = st.columns(2)
            with tc1c:
                tconv_nomb = st.text_input("Nombre tarifa convenio *", placeholder="Corporativo Noche")
                tconv_prec = st.number_input("Precio/hora COP", min_value=0, step=500, value=12000)
                tconv_d3   = st.number_input("Descuento 3h+ (%)", min_value=0.0, max_value=50.0)
                tconv_fest = st.checkbox("Aplica en festivos")
            with tc2c:
                tconv_d6   = st.number_input("Descuento 6h+ (%)", min_value=0.0, max_value=50.0)
                tconv_hini = st.text_input("Hora inicio especial", placeholder="22:00")
                tconv_hfin = st.text_input("Hora fin especial",    placeholder="06:00")
                tconv_desc_txt = st.text_input("Descripción", placeholder="Tarifa nocturna empresarial")
            if st.form_submit_button("💾 Guardar tarifa convenio", type="primary",
                                     use_container_width=True):
                if not tconv_nomb.strip():
                    st.error("El nombre de la tarifa es obligatorio")
                else:
                    tar_id_conv = f"CONV-{int(time.time())}"
                    nombre_tar  = f"[Convenio] {tconv_nomb.strip()}"
                    desc_tar    = tconv_desc_txt.strip() or "Tarifa convenio"
                    # jjgt_pagos: 10 campos estándar (sin empresa)
                    tar_fila_p = [
                        tar_id_conv, nombre_tar, desc_tar, str(tconv_prec),
                        str(tconv_d3), str(tconv_d6),
                        tconv_hini.strip(), tconv_hfin.strip(),
                        "1" if tconv_fest else "0", "1",
                    ]
                    # jjgt_convenios: 12 campos (con Id_Empresa y Nombre_Empresa)
                    tar_fila_c = [
                        tar_id_conv, _tar_emp_id, _tar_emp_nom,
                        nombre_tar, desc_tar, str(tconv_prec),
                        str(tconv_d3), str(tconv_d6),
                        tconv_hini.strip(), tconv_hfin.strip(),
                        "1" if tconv_fest else "0", "1",
                    ]
                    # Guardar en jjgt_pagos
                    _, sh_tar_p = get_active_client()
                    if sh_tar_p:
                        _gs_append(sh_tar_p, "Tarifas_Config", tar_fila_p)
                        _gs_invalidate_cache("Tarifas_Config")
                    # Guardar en jjgt_convenios
                    _, sh_tar_cv = get_active_client_convenios()
                    if sh_tar_cv:
                        _gs_append(sh_tar_cv, "Tarifas_Config", tar_fila_c)
                        _gs_invalidate_cache_conv("Tarifas_Config")
                    destinos = []
                    if sh_tar_p:  destinos.append("jjgt_pagos")
                    if sh_tar_cv: destinos.append("jjgt_convenios")
                    emp_lbl = f" para **{_tar_emp_nom}**" if _tar_emp_id else " (genérica)"
                    if destinos:
                        st.success(f"✅ Tarifa '{nombre_tar}'{emp_lbl} "
                                   f"guardada en: {', '.join(destinos)}")
                    else:
                        st.error("❌ Sin conexión a Google Sheets.")

        st.divider()
        st.markdown("#### Operadores habilitados para convenios")
        ops_cv = _gs_read_sheet("Operadores")
        if ops_cv:
            _df_ops_cv = pd.DataFrame([{
                "Nombre": str(_gs_val(r,"Nombre")),
                "Rol":    str(_gs_val(r,"Rol")),
                "Turno":  str(_gs_val(r,"Turno")),
            } for r in ops_cv if _gs_val(r,"Activo")=="1"])
            st.dataframe(_df_ops_cv, use_container_width=True, hide_index=True)
        st.info("Para modificar operadores ve a la pestana Operadores.")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 6 — 🗂️ Configuración de Pagos de Convenios (Configuracion_Pagos)
    # ══════════════════════════════════════════════════════════════════════════
    with tabs[6]:
        st.markdown("### 🗂️ Configuración de Pagos de Convenios")
        st.caption(
            "Administra directamente los registros de la tabla/hoja **Configuracion_Pagos** "
            "de `jjgt_convenios`. Cada registro es una clave-valor asociada a una empresa."
        )

        _, _sh_cpc = get_active_client_convenios()

        # ── Leer registros actuales ────────────────────────────────────────
        def _leer_config_pagos_conv():
            try:
                return _pg_read_table_conv("Configuracion_Pagos") or []
            except Exception:
                return []

        _cpc_rows = _leer_config_pagos_conv()

        # ── Tabla de registros actuales ────────────────────────────────────
        st.markdown("#### 📋 Registros actuales en `Configuracion_Pagos`")
        if _cpc_rows:
            _df_cpc = pd.DataFrame([{
                "ID Empresa":   str(r.get("Id_Empresa", r.get("id_empresa", ""))),
                "Empresa":      str(r.get("Nombre_Empresa", r.get("nombre_empresa", "—"))),
                "Clave":        str(r.get("Clave", r.get("clave", ""))),
                "Valor":        str(r.get("Valor", r.get("valor", ""))),
                "Actualizado":  str(r.get("Actualizado_En", r.get("actualizado_en", ""))),
            } for r in _cpc_rows])
            st.dataframe(_df_cpc, use_container_width=True, hide_index=True)
        else:
            st.info("No hay registros en Configuracion_Pagos todavía, o jjgt_convenios no está configurado.")

        st.divider()

        # ── CREAR nuevo registro ───────────────────────────────────────────
        with st.expander("➕ Crear nuevo registro en Configuracion_Pagos", expanded=False):
            _emps_cpc_crear = get_empresas_convenio()
            if not _emps_cpc_crear:
                st.info("Primero crea empresas de convenio en la pestaña **Convenios**.")
            else:
                _opts_cpc_crear = {
                    f"{e.get('Nombre_Empresa','—')} | NIT: {e.get('Nit_Empresa','')}": e
                    for e in _emps_cpc_crear
                }
                _sel_cpc_crear_label = st.selectbox(
                    "Empresa", list(_opts_cpc_crear.keys()), key="cpc_crear_emp"
                )
                _emp_cpc_crear = _opts_cpc_crear[_sel_cpc_crear_label]
                _emp_cpc_crear_id   = _emp_cpc_crear.get("Id_Empresa", "")
                _emp_cpc_crear_nomb = _emp_cpc_crear.get("Nombre_Empresa", "—")

                with st.form("form_cpc_crear", clear_on_submit=True):
                    _cpc_c1, _cpc_c2 = st.columns(2)
                    with _cpc_c1:
                        cpc_clave = st.text_input(
                            "Clave *",
                            placeholder="convenios_dias_credito",
                            help="Nombre único del parámetro de configuración para esta empresa"
                        )
                    with _cpc_c2:
                        cpc_valor = st.text_input(
                            "Valor *",
                            placeholder="30",
                            help="Valor del parámetro (texto, número o JSON simple)"
                        )
                    if st.form_submit_button("💾 Crear registro", type="primary", use_container_width=True):
                        if not cpc_clave.strip() or not cpc_valor.strip():
                            st.error("Clave y Valor son obligatorios")
                        elif not _sh_cpc:
                            st.error("Sin conexión a jjgt_convenios. Verifica el ID del spreadsheet en pestaña Convenios.")
                        else:
                            try:
                                ok_cpc = upsert_config_empresa_convenio(
                                    _sh_cpc,
                                    _emp_cpc_crear_id,
                                    _emp_cpc_crear_nomb,
                                    {cpc_clave.strip(): cpc_valor.strip()}
                                )
                                if ok_cpc:
                                    st.success(
                                        f"✅ Registro **{cpc_clave.strip()}** = `{cpc_valor.strip()}` "
                                        f"creado para **{_emp_cpc_crear_nomb}**."
                                    )
                                    st.rerun()
                                else:
                                    st.error("Error al crear el registro. Verifica la conexión a jjgt_convenios.")
                            except Exception as _ecpc:
                                st.error(f"Error: {_ecpc}")

        # ── MODIFICAR registro ─────────────────────────────────────────────
        with st.expander("✏️ Modificar registro en Configuracion_Pagos", expanded=False):
            _cpc_rows_mod = _leer_config_pagos_conv()
            if not _cpc_rows_mod:
                st.info("No hay registros para modificar.")
            else:
                # Construir opciones con empresa + clave
                _opts_mod_cpc = {
                    f"{r.get('Nombre_Empresa', r.get('nombre_empresa','—'))} · {r.get('Clave', r.get('clave',''))}": r
                    for r in _cpc_rows_mod
                    if r.get("Clave") or r.get("clave")
                }
                if not _opts_mod_cpc:
                    st.info("No hay registros con clave válida para modificar.")
                else:
                    _sel_mod_cpc_label = st.selectbox(
                        "Registro a modificar", list(_opts_mod_cpc.keys()), key="cpc_mod_sel"
                    )
                    _row_mod_cpc = _opts_mod_cpc.get(_sel_mod_cpc_label) if _sel_mod_cpc_label else None
                    if _row_mod_cpc is None:
                        st.info("Selecciona un registro válido.")
                    else:
                        _cpc_mod_id   = _row_mod_cpc.get("Id_Empresa", _row_mod_cpc.get("id_empresa", ""))
                        _cpc_mod_nomb = _row_mod_cpc.get("Nombre_Empresa", _row_mod_cpc.get("nombre_empresa", "—"))
                        _cpc_mod_clav = _row_mod_cpc.get("Clave", _row_mod_cpc.get("clave", ""))
                        _cpc_mod_val  = _row_mod_cpc.get("Valor", _row_mod_cpc.get("valor", ""))

                        st.info(f"Empresa: **{_cpc_mod_nomb}** · Clave: `{_cpc_mod_clav}`")

                        with st.form("form_cpc_mod", clear_on_submit=False):
                            cpc_m_valor = st.text_input(
                                "Nuevo valor *", value=str(_cpc_mod_val),
                                help="Modifica el valor de esta clave para la empresa seleccionada"
                            )
                            if st.form_submit_button("💾 Guardar cambio", type="primary", use_container_width=True):
                                if not cpc_m_valor.strip():
                                    st.error("El valor no puede estar vacío")
                                elif not _sh_cpc:
                                    st.error("Sin conexión a jjgt_convenios.")
                                else:
                                    try:
                                        ok_mod_cpc = upsert_config_empresa_convenio(
                                            _sh_cpc, _cpc_mod_id, _cpc_mod_nomb,
                                            {_cpc_mod_clav: cpc_m_valor.strip()}
                                        )
                                        if ok_mod_cpc:
                                            st.success(
                                                f"✅ Registro **{_cpc_mod_clav}** actualizado a "
                                                f"`{cpc_m_valor.strip()}` para **{_cpc_mod_nomb}**."
                                            )
                                            st.rerun()
                                        else:
                                            st.error("Error al actualizar. Verifica la conexión.")
                                    except Exception as _ecpc_m:
                                        st.error(f"Error: {_ecpc_m}")

        # ── ELIMINAR registro ──────────────────────────────────────────────
        with st.expander("🗑️ Eliminar registro de Configuracion_Pagos", expanded=False):
            _cpc_rows_del = _leer_config_pagos_conv()
            if not _cpc_rows_del:
                st.info("No hay registros para eliminar.")
            else:
                _opts_del_cpc = {
                    f"{r.get('Nombre_Empresa', r.get('nombre_empresa','—'))} · {r.get('Clave', r.get('clave',''))}": r
                    for r in _cpc_rows_del
                    if r.get("Clave") or r.get("clave")
                }
                if not _opts_del_cpc:
                    st.info("No hay registros con clave válida para eliminar.")
                else:
                    _col_del_cpc1, _col_del_cpc2 = st.columns([2, 1])
                    with _col_del_cpc1:
                        _sel_del_cpc_label = st.selectbox(
                            "Registro a eliminar", list(_opts_del_cpc.keys()), key="cpc_del_sel"
                        )
                    with _col_del_cpc2:
                        _pin_del_cpc = st.text_input("PIN admin", type="password", key="cpc_del_pin")
                    _row_del_cpc = _opts_del_cpc.get(_sel_del_cpc_label) if _sel_del_cpc_label else None
                    if _row_del_cpc is None:
                        st.info("Selecciona un registro válido.")
                    else:
                        _cpc_del_id   = _row_del_cpc.get("Id_Empresa", _row_del_cpc.get("id_empresa", ""))
                        _cpc_del_clav = _row_del_cpc.get("Clave", _row_del_cpc.get("clave", ""))
                        _cpc_del_nomb = _row_del_cpc.get("Nombre_Empresa", _row_del_cpc.get("nombre_empresa", "—"))

                        st.warning(
                            f"⚠️ Se eliminará la clave **{_cpc_del_clav}** de la empresa **{_cpc_del_nomb}**. "
                            "Esta acción no se puede deshacer."
                        )
                        if st.button("🗑️ Eliminar registro", type="primary", use_container_width=True, key="btn_del_cpc"):
                            if not verificar_pin(_pin_del_cpc, "admin"):
                                st.error("PIN admin incorrecto")
                            elif not _sh_cpc:
                                st.error("Sin conexión a jjgt_convenios.")
                            else:
                                try:
                                    _pg_delete_row_conv("Configuracion_Pagos", "Clave", _cpc_del_clav)
                                    st.success(
                                        f"✅ Registro **{_cpc_del_clav}** eliminado de "
                                        f"Configuracion_Pagos para **{_cpc_del_nomb}**."
                                    )
                                    st.rerun()
                                except Exception as _edel_cpc:
                                    st.error(f"Error al eliminar: {_edel_cpc}")

        # ── Vista consolidada de todas las claves ──────────────────────────
        with st.expander("🔍 Vista consolidada por empresa", expanded=False):
            _emps_vista = get_empresas_convenio()
            if not _emps_vista:
                st.info("No hay empresas de convenio registradas.")
            else:
                _rows_vista_cpc = []
                for _ev in _emps_vista:
                    _cfg_ev = get_config_empresa_convenio(_sh_cpc, _ev.get("Id_Empresa", ""))
                    for _kv, _vv in _cfg_ev.items():
                        _rows_vista_cpc.append({
                            "Empresa": _ev.get("Nombre_Empresa", "—"),
                            "Clave":   _kv,
                            "Valor":   str(_vv),
                        })
                if _rows_vista_cpc:
                    _df_vista_cpc = pd.DataFrame(_rows_vista_cpc)
                    st.dataframe(_df_vista_cpc, use_container_width=True, hide_index=True)
                else:
                    st.info("Sin configuraciones registradas para las empresas existentes.")


def _tab_convenio_reserva():
    """Tab de formulario rápido para reservas bajo convenio empresarial."""
    st.markdown("**Formulario de asignación para convenio empresarial:**")
    _, sh_conv = get_active_client_convenios()
    empresas = get_empresas_convenio()
    if not empresas:
        st.info("No hay empresas con convenio activo. Créalas en ⚙️ Configuración → Convenios.")
        return

    op = st.session_state.get("operador_info", {})

    # ── Paso 1: Empresa ───────────────────────────────────────────────────────
    st.markdown("#### 🏢 Paso 1 — Empresa del convenio")
    emp_opts = {f"{e.get('Nombre_Empresa','—')} ({e.get('Tipo_Convenio','—')})": e for e in empresas}
    emp_label   = st.selectbox("Empresa / Convenio *", list(emp_opts.keys()), key="conv_res_empresa")
    empresa_sel = emp_opts[emp_label]

    # Cargar configuración de la empresa (descuentos, horas mínimas, etc.)
    cfg_emp = get_config_empresa_convenio(sh_conv, empresa_sel.get("Id_Empresa", ""))
    horas_min    = float(cfg_emp.get("convenios_horas_min", 1.0))
    desc_conv    = float(cfg_emp.get("convenios_descuento_pct", 0.0))
    max_pend     = int(cfg_emp.get("convenios_max_pendientes", 10))

    col_emp_info1, col_emp_info2, col_emp_info3 = st.columns(3)
    nit_emp_str   = empresa_sel.get('Nit_Empresa', '—')
    tipo_conv_str = empresa_sel.get('Tipo_Convenio', '—')
    dias_cred_val = int(cfg_emp.get('convenios_dias_credito', 30))
    col_emp_info1.info(f'NIT: {nit_emp_str} | Tipo convenio: {tipo_conv_str}')
    col_emp_info2.info(f'Desc. adicional: {desc_conv:.1f}% | Horas min.: {horas_min}h')
    col_emp_info3.info(f'Max. pendientes: {max_pend} | Dias credito: {dias_cred_val}')

    st.divider()

    # ── Paso 2: Cubículo y tiempo ─────────────────────────────────────────────
    st.markdown("#### 🛏️ Paso 2 — Cubículo y tiempo")

    # Refrescar estado de cubículos desde Sheets
    cubiculos_libres = get_cubiculos_libres()
    if not cubiculos_libres:
        st.error("❌ No hay cubículos disponibles en este momento.")
        return

    col_cub, col_hrs = st.columns(2)
    with col_cub:
        opciones_cub = {f"{c['numero']} — {c['nombre']} | WiFi: {c['wifi_ssid']}": c
                        for c in cubiculos_libres}
        sel_label    = st.selectbox("Cubículo libre *", list(opciones_cub.keys()),
                                    key="conv_res_cubiculo")
        cubiculo_sel = opciones_cub[sel_label]
        st.caption(f"Estado actual: 🟢 **LIBRE** · Precio base: {fmt_cop(cubiculo_sel.get('precio_hora_base',15000))}/h")

    with col_hrs:
        horas_sel = st.number_input(
            f"Horas a reservar (mín. {horas_min}h)",
            min_value=horas_min, max_value=18.0,
            value=max(horas_min, 1.0), step=0.5,
            key="conv_res_horas")

        # Calcular precio con descuentos del convenio
        calc = calcular_precio_convenio(horas_sel, cfg_emp, id_empresa=empresa_sel.get("Id_Empresa",""))

        desc_info = ""
        if calc.get("descuento_vol_pct", 0) > 0:
            desc_info += f"Desc. volumen: {calc['descuento_vol_pct']:.0f}%  "
        if calc.get("descuento_conv_pct", 0) > 0:
            desc_info += f"Desc. convenio: {calc['descuento_conv_pct']:.0f}%"

        st.markdown(f"""
        <div style="background:rgba(162,155,254,0.12);border:1px solid rgba(162,155,254,0.4);
                    border-radius:10px;padding:12px;margin-top:8px;font-size:14px">
          <div style="color:#94a3b8;font-size:12px;margin-bottom:4px">PRECIO CONVENIO</div>
          <div>Precio/hora: <b>{fmt_cop(calc['precio_hora'])}</b>
               × {calc['horas']}h = {fmt_cop(round(calc['precio_hora']*calc['horas']))}</div>
          {'<div style="color:#ffd32a">🏷️ ' + desc_info + f" → Ahorro: {fmt_cop(calc['descuento_val'])}</div>" if calc['descuento_val'] > 0 else ''}
          <div>Subtotal: <b>{fmt_cop(calc['subtotal'])}</b>
               · IVA 19%: <b>{fmt_cop(calc['iva'])}</b></div>
          <div style="font-family:'Inconsolata';font-size:22px;font-weight:700;
                      color:#a29bfe;margin-top:6px">
            TOTAL: {fmt_cop(calc['total'])} COP
          </div>
          <div style="color:#ffd32a;font-size:12px;margin-top:4px">
            ⚠️ Pago: PENDIENTE — se cobra a {empresa_sel.get('Nombre_Empresa','la empresa')}
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # ── Paso 3: Datos del cliente ─────────────────────────────────────────────
    st.markdown("#### 👤 Paso 3 — Datos del cliente")
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        c_nombre   = st.text_input("Nombre completo *", key="conv_res_nombre",
                                   placeholder="Nombre del empleado / pasajero")
        c_tel      = st.text_input("Teléfono", key="conv_res_tel", placeholder="300 000 0000")
        c_tipo_doc = st.selectbox("Tipo documento", ["CC","CE","Pasaporte","NIT","TI"],
                                  key="conv_res_tipodoc")
    with col_c2:
        c_doc   = st.text_input("Número de documento", key="conv_res_doc")
        c_email = st.text_input("Email (opcional)", key="conv_res_email")
        st.markdown(f"""
        <div style="background:rgba(255,211,42,0.08);border:1px solid rgba(255,211,42,0.3);
                    border-radius:8px;padding:10px;margin-top:24px;font-size:13px">
          <b style="color:#ffd32a">💼 Convenio:</b> {empresa_sel.get('Nombre_Empresa','—')}<br>
          <b style="color:#ffd32a">📋 Reserva:</b> {empresa_sel.get('Tipo_Convenio','—')}<br>
          <span style="color:#94a3b8;font-size:11px">El pago queda PENDIENTE·
          Se registra solo en jjgt_convenios</span>
        </div>
        """, unsafe_allow_html=True)

    st.divider()

    # ── Paso 4: PIN y confirmar ───────────────────────────────────────────────
    st.markdown("#### 🔐 Paso 4 — Confirmación")
    col_pin, col_resumen, col_btn = st.columns([1, 2, 2])
    with col_pin:
        pin_conv = st.text_input("PIN operador *", type="password", max_chars=8, key="conv_res_pin")
    with col_resumen:
        st.markdown(f"""
        <div style="background:rgba(0,0,0,0.2);border-radius:8px;padding:10px;font-size:13px">
          Empresa: <b>{empresa_sel.get('Nombre_Empresa','—')}</b><br>
          Cubículo: <b>{cubiculo_sel['numero']}</b> · Horas: <b>{horas_sel}h</b><br>
          Total: <b style="color:#a29bfe">{fmt_cop(calc['total'])} COP</b>
          {'· Desc: ' + str(round(calc.get("descuento_pct",0),1)) + '%' if calc.get('descuento_pct',0) > 0 else ''}
        </div>
        """, unsafe_allow_html=True)
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🤝 CREAR RESERVA CONVENIO", type="primary",
                     use_container_width=True, key="btn_crear_conv"):
            errores = []
            if not c_nombre.strip():
                errores.append("El nombre del cliente es obligatorio")
            if not verificar_pin(pin_conv):
                errores.append("PIN de operador incorrecto")
            if errores:
                for er in errores:
                    st.error(er)
            else:
                # Doble verificación de disponibilidad antes de crear
                cub_check = next((c for c in get_cubiculos_libres()
                                  if c["numero"] == cubiculo_sel["numero"]), None)
                if not cub_check:
                    st.error(f"❌ El cubículo {cubiculo_sel['numero']} ya no está disponible. "
                             "Selecciona otro.")
                else:
                    try:
                        with st.spinner("Creando reserva de convenio en jjgt_convenios..."):
                            voucher_conv = crear_reserva_convenio(
                                cubiculo_sel,
                                {"nombre":            c_nombre.strip(),
                                 "tipo_doc":          c_tipo_doc,
                                 "numero_documento":  c_doc.strip(),
                                 "telefono":          c_tel.strip(),
                                 "email":             c_email.strip()},
                                calc,
                                empresa_sel.get("Id_Empresa", ""),
                                empresa_sel.get("Nombre_Empresa", ""),
                                operador=op.get("nombre", "sistema"),
                            )
                        _gs_invalidate_cache("Cubiculos_Estado")
                        # Guardar voucher y hacer rerun para que el dashboard
                        # recargue los datos frescos de PostgreSQL
                        st.session_state["_conv_voucher_ultimo"] = voucher_conv
                        st.rerun()
                    except Exception as ex:
                        st.error(f"❌ Error al crear reserva de convenio: {ex}")

    # ── Mostrar voucher si acaba de crearse una reserva (post-rerun) ──────────
    voucher_conv = st.session_state.pop("_conv_voucher_ultimo", None)
    if voucher_conv:
        st.balloons()
        vc = voucher_conv
        desc_str = ("· Desc: " + str(round(vc.get("descuento_pct", 0), 1)) + "%"
                    if vc.get("descuento_val", 0) > 0 else "")
        st.markdown(
            f'''<div style="background:rgba(162,155,254,0.15);border:2px solid #a29bfe;
                        border-radius:14px;padding:24px;text-align:center;margin-top:12px">
              <div style="font-size:50px">🤝</div>
              <div style="font-size:22px;font-weight:800;color:#a29bfe;margin:8px 0">
                ¡RESERVA CONVENIO CREADA!
              </div>
              <div style="font-family:'Inconsolata';font-size:40px;font-weight:700;color:#00d4ff">
                {vc["cubiculo"]}</div>
              <div style="color:#e2e8f0;margin-top:8px;font-size:15px">
                Reserva: <b>{vc["numero_reserva"]}</b><br>
                Empresa: <b>{vc["nombre_empresa"]}</b><br>
                Horario: <b>{vc["hora_inicio"]} → {vc["hora_fin"]}</b> ({vc["horas"]}h)<br>
                Total: <b style="color:#a29bfe">{fmt_cop(vc["total"])} COP</b> {desc_str}
              </div>
              <div style="margin-top:12px">
                <span style="font-size:13px;color:#94a3b8">Código de acceso</span><br>
                <span style="font-family:'Inconsolata';font-size:44px;font-weight:700;
                             color:#00ff88;letter-spacing:12px">{vc["codigo_acceso"]}</span>
              </div>
              <div style="margin-top:8px;color:#ffd32a;font-size:13px">
                ⚠️ Pago PENDIENTE — se cobra a la empresa
              </div>
              <div style="color:#94a3b8;font-size:11px;margin-top:4px">
                Registrado en jjgt_convenios · Cubículo marcado como OCUPADO
              </div>
            </div>''',
            unsafe_allow_html=True,
        )


def _op_convenios():
    """Módulo completo de gestión de convenios con empresas."""
    st.markdown("### 🤝 Módulo de Convenios Empresariales")
    tab_dash, tab_reserva, tab_reporte, tab_gestion = st.tabs([
        "📊 Dashboard Convenios", "➕ Nueva Reserva Convenio",
        "📋 Reporte Convenios",   "🗑️ Gestión Datos Convenios",
    ])

    # Dashboard Convenios
    with tab_dash:
        st.markdown("#### Dashboard Operacional — Convenios")
        inject_live_clock()
        today = ahora_col().strftime("%Y-%m-%d")
        empresas = get_empresas_convenio()
        # Leer directamente desde PostgreSQL — sin depender de convenios_spreadsheet_id
        col_ref, _ = st.columns([1, 5])
        with col_ref:
            if st.button("🔄 Actualizar", key="btn_refresh_conv_dash"):
                st.rerun()
        reservas_gs  = _pg_read_table_conv("Reservas")
        pagos_gs     = _pg_read_table_conv("Pagos")
        res_conv     = reservas_gs
        res_conv_hoy = [r for r in res_conv if _gs_fecha_ymd(r, "Creado_En") == today]
        pend_conv    = [r for r in pagos_gs if _gs_val(r, "Estado") == "pendiente"]
        total_pend_cop = sum(_gs_float(r, "Monto_COP") for r in pend_conv)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Empresas activas",    len(empresas))
        c2.metric("Reservas hoy (Conv)", len(res_conv_hoy))
        c3.metric("Pagos pendientes",    len(pend_conv))
        c4.metric("Monto pendiente",     fmt_cop(total_pend_cop))
        st.divider()
        if res_conv_hoy:
            st.markdown("**Reservas de convenio del día:**")
            df_rch = pd.DataFrame([{
                "Reserva":   _gs_val(r, "Numero_Reserva"),
                "Empresa":   _gs_val(r, "Nombre_Empresa", "—"),
                "Cliente":   _gs_val(r, "Cliente_Nombre"),
                "Cubículo":  _gs_val(r, "Cubiculo_Num"),
                "Horas":     _gs_val(r, "Horas_Contratadas"),
                "Total COP": fmt_cop(_gs_float(r, "Total_COP")),
                "Estado":    _gs_val(r, "Estado_Pago"),
            } for r in res_conv_hoy])
            for _c in ["Reserva","Empresa","Cliente","Cubículo","Horas","Estado"]:
                if _c in df_rch.columns: df_rch[_c] = df_rch[_c].astype(str)
            st.dataframe(df_rch, use_container_width=True, hide_index=True)
        else:
            st.info("No hay reservas de convenio registradas hoy.")
        st.divider()
        # Config resumida por empresa
        _, sh_dash_conv = get_active_client_convenios()
        if sh_dash_conv and empresas:
            with st.expander("⚙️ Config por empresa (descuentos / crédito)", expanded=False):
                rows_cfg_dash = []
                for emp_d in empresas:
                    cfg_d = get_config_empresa_convenio(sh_dash_conv, emp_d.get("Id_Empresa",""))
                    rows_cfg_dash.append({
                        "Empresa":      emp_d.get("Nombre_Empresa","—"),
                        "Descuento %":  cfg_d["convenios_descuento_pct"],
                        "Horas min":    cfg_d["convenios_horas_min"],
                        "Dias credito": int(cfg_d["convenios_dias_credito"]),
                        "Max pend.":    int(cfg_d["convenios_max_pendientes"]),
                    })
                _df_cfg_dash = pd.DataFrame(rows_cfg_dash)
                for _c in ["Empresa","Descuento %","Horas min"]:
                    if _c in _df_cfg_dash.columns: _df_cfg_dash[_c] = _df_cfg_dash[_c].astype(str)
                st.dataframe(_df_cfg_dash, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("#### Estado en tiempo real (cubículos)")
        cubiculos = get_cubiculos()
        cols_rt = st.columns(4)
        for i, cub in enumerate(cubiculos):
            with cols_rt[i % 4]:
                estado = cub["estado"]
                info   = ESTADOS_CUBICULO.get(estado, ESTADOS_CUBICULO["libre"])
                mins   = cub.get("minutos_restantes")
                hora_fin = cub.get("hora_fin") or ""
                if hora_fin and estado not in ("libre","mantenimiento"):
                    timer_html = (f'<div class="jjgt-timer" data-fin="{hora_fin}" '
                                  f'data-key="cdash-{cub["id"]}" '
                                  f'style="font-size:12px;color:{"#ff4757" if (mins or 999)<=5 else "#ffd32a" if (mins or 999)<=15 else "#00ff88"};">'
                                  f'⏱ {fmt_tiempo(mins)}</div>')
                else:
                    timer_html = ""
                st.markdown(f"""
                <div style="background:{info['bg']};border:1.5px solid {info['color']}55;
                            border-radius:10px;padding:10px;text-align:center;margin-bottom:8px">
                  <div style="font-family:'Inconsolata';font-size:24px;font-weight:700;color:{info['color']}">{cub['numero']}</div>
                  <div style="font-size:10px;font-weight:700;color:{info['color']};letter-spacing:1px">{info['label']}</div>
                  {timer_html}
                </div>
                """, unsafe_allow_html=True)

    # Nueva Reserva Convenio
    with tab_reserva:
        _tab_convenio_reserva()

    # Reporte Convenios
    with tab_reporte:
        st.markdown("#### Reporte de Convenios")
        periodo_c = st.selectbox("Período", ["Hoy","Esta semana","Este mes","Histórico"], key="conv_rep_periodo")
        today_c = ahora_col()
        if periodo_c == "Hoy":
            desde_c = today_c.strftime("%Y-%m-%d"); hasta_c = desde_c
        elif periodo_c == "Esta semana":
            desde_c = (today_c - timedelta(days=today_c.weekday())).strftime("%Y-%m-%d"); hasta_c = today_c.strftime("%Y-%m-%d")
        elif periodo_c == "Este mes":
            desde_c = today_c.strftime("%Y-%m-01"); hasta_c = today_c.strftime("%Y-%m-%d")
        else:
            desde_c = "2020-01-01"; hasta_c = today_c.strftime("%Y-%m-%d")
        # Leer de jjgt_convenios (no de jjgt_pagos)
        rows_c = []
        for r in _gs_read_sheet_conv("Reservas", force=False):
            fecha_r = _gs_fecha_ymd(r, "Creado_En")
            if desde_c <= fecha_r <= hasta_c:
                rows_c.append({
                    "Reserva":   _gs_val(r,"Numero_Reserva"),
                    "Empresa":   _gs_val(r,"Nombre_Empresa","—"),
                    "Cliente":   _gs_val(r,"Cliente_Nombre"),
                    "Cubículo":  _gs_val(r,"Cubiculo_Num"),
                    "Horas":     _gs_float(r,"Horas_Contratadas"),
                    "Total":     _gs_float(r,"Total_COP"),
                    "Estado":    _gs_val(r,"Estado_Pago"),
                    "Fecha":     fecha_r,
                })
        total_c   = sum(r["Total"] for r in rows_c)
        pend_c    = sum(r["Total"] for r in rows_c if r["Estado"] == "pendiente")
        confirm_c = sum(r["Total"] for r in rows_c if r["Estado"] == "confirmado")
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Reservas conv.",   len(rows_c))
        m2.metric("Total facturado",  fmt_cop(total_c))
        m3.metric("Pendiente cobro",  fmt_cop(pend_c))
        m4.metric("Confirmado",       fmt_cop(confirm_c))
        if rows_c:
            df_rc = pd.DataFrame(rows_c)
            df_rc_disp = df_rc.copy()
            df_rc_disp["Total"] = df_rc_disp["Total"].apply(fmt_cop)
            for _c in ["Reserva","Empresa","Cliente","Cubículo","Estado","Fecha"]:
                if _c in df_rc_disp.columns: df_rc_disp[_c] = df_rc_disp[_c].astype(str)
            st.dataframe(df_rc_disp, use_container_width=True, hide_index=True)
            col_csv_c, col_xls_c = st.columns(2)
            with col_csv_c:
                st.download_button("📥 Exportar CSV Convenios",
                    data=df_rc.to_csv(index=False).encode(),
                    file_name=f"convenios_{desde_c}_{hasta_c}.csv", mime="text/csv",
                    use_container_width=True)
            with col_xls_c:
                if OPENPYXL_AVAILABLE:
                    buf_c = io.BytesIO()
                    with pd.ExcelWriter(buf_c, engine="openpyxl") as writer:
                        df_rc.to_excel(writer, sheet_name="Reservas_Convenios", index=False)
                        emp_res = {}
                        for r in rows_c:
                            emp = r["Empresa"]
                            if emp not in emp_res:
                                emp_res[emp] = {"Empresa":emp,"Reservas":0,"Total":0,"Pendiente":0}
                            emp_res[emp]["Reservas"] += 1
                            emp_res[emp]["Total"] += r["Total"]
                            if r["Estado"] == "pendiente": emp_res[emp]["Pendiente"] += r["Total"]
                        pd.DataFrame(list(emp_res.values())).to_excel(writer, sheet_name="Resumen_Empresa", index=False)
                    st.download_button("📊 Exportar Excel Convenios",
                        data=buf_c.getvalue(),
                        file_name=f"reporte_convenios_{desde_c}_{hasta_c}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
        else:
            st.info("No hay reservas de convenio para el período seleccionado.")

    # Gestión Datos Convenios
    with tab_gestion:
        st.markdown("#### Gestión de Datos — Convenios")
        st.warning("Las eliminaciones son permanentes e irreversibles.")
        op_g = st.session_state.get("operador_info", {})
        if op_g.get("rol") != "admin" and "admin" not in op_g.get("permisos", []):
            st.error("Solo administradores pueden acceder.")
            return

        def _del_row_both(hoja, col_clave, valor_clave):
            """Elimina en jjgt_pagos y en jjgt_convenios (si la tabla existe en ambos)."""
            _pg_delete_row(hoja, col_clave, valor_clave)
            try:
                _pg_delete_row_conv(hoja, col_clave, valor_clave)
            except Exception:
                pass

        stabs = st.tabs(["Reservas","Clientes","Pagos","Facturas"])
        with stabs[0]:
            res_cv = _gs_read_sheet_conv("Reservas", force=False)
            if not res_cv:
                st.info("No hay reservas de convenio.")
            else:
                _df_cv_res = pd.DataFrame([{
                    "Reserva": str(_gs_val(r,"Numero_Reserva")),
                    "Empresa": str(_gs_val(r,"Nombre_Empresa","—")),
                    "Cliente": str(_gs_val(r,"Cliente_Nombre")),
                    "Estado":  str(_gs_val(r,"Estado_Pago")),
                    "Total":   fmt_cop(_gs_float(r,"Total_COP")),
                } for r in res_cv])
                st.dataframe(_df_cv_res, use_container_width=True, hide_index=True)
                col_d1,col_d2 = st.columns([2,1])
                with col_d1:
                    opts = {f"{_gs_val(r,'Numero_Reserva')} — {_gs_val(r,'Nombre_Empresa','—')}":
                            _gs_val(r,"Numero_Reserva") for r in res_cv}
                    sel = st.selectbox("Reserva a eliminar", list(opts.keys()), key="del_cv_res")
                    sel_key = opts[sel]
                with col_d2:
                    pin_d = st.text_input("PIN admin", type="password", key="del_cv_res_pin")
                if st.button("Eliminar reserva convenio", type="primary", use_container_width=True, key="btn_del_cv_res"):
                    if not verificar_pin(pin_d, "admin"):
                        st.error("PIN admin incorrecto")
                    else:
                        _del_row_both("Reservas","Numero_Reserva", sel_key)
                        _del_row_both("Pagos","Num_Reserva", sel_key)
                        _gs_invalidate_cache("Reservas","Pagos")
                        st.success(f"Reserva {sel_key} eliminada.")
                        st.rerun()
        with stabs[1]:
            _, sh_cv_cli = get_active_client_convenios()
            if False:  # PG always available
                st.info("Configure jjgt_convenios para ver clientes.")
            else:
                try:
                    clientes_cv = _pg_read_table_conv("Clientes")
                    if not clientes_cv:
                        st.info("No hay clientes de convenio.")
                    else:
                        _df_cv_cli = pd.DataFrame([{
                            "Nombre":  str(r.get("Nombre","")),
                            "Doc":     str(r.get("Num_Doc","")),
                            "Empresa": str(r.get("Nombre_Empresa","—")),
                        } for r in clientes_cv])
                        st.dataframe(_df_cv_cli, use_container_width=True, hide_index=True)
                        col_cc1,col_cc2=st.columns([2,1])
                        with col_cc1:
                            opts_cc={f"{r.get('Nombre','')} — {r.get('Num_Doc','')}":r.get("Num_Doc","") for r in clientes_cv if r.get("Num_Doc")}
                            sel_cc=st.selectbox("Cliente",list(opts_cc.keys()),key="del_cv_cli")
                            sel_cc_k=opts_cc[sel_cc]
                        with col_cc2:
                            pin_cc=st.text_input("PIN admin",type="password",key="del_cv_cli_pin")
                        if st.button("Eliminar cliente",type="primary",use_container_width=True,key="btn_del_cv_cli"):
                            if not verificar_pin(pin_cc,"admin"):
                                st.error("PIN incorrecto")
                            else:
                                try:
                                    _pg_delete_row_conv("Clientes","Num_Doc",sel_cc_k)
                                except Exception: pass
                                st.success("Cliente eliminado de PostgreSQL (convenios).")
                                st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")
        with stabs[2]:
            pag_cv = _gs_read_sheet_conv("Pagos", force=False)
            if not pag_cv:
                st.info("No hay pagos de convenio.")
            else:
                _df_cv_pag = pd.DataFrame([{
                    "ID":      str(_gs_val(r,"ID_Pago")),
                    "Reserva": str(_gs_val(r,"Num_Reserva")),
                    "Empresa": str(_gs_val(r,"Nombre_Empresa","—")),
                    "Monto":   fmt_cop(_gs_float(r,"Monto_COP")),
                    "Estado":  str(_gs_val(r,"Estado")),
                } for r in pag_cv])
                st.dataframe(_df_cv_pag, use_container_width=True, hide_index=True)
                col_p1,col_p2=st.columns([2,1])
                with col_p1:
                    opts_p={f"#{_gs_val(r,'ID_Pago')}—{_gs_val(r,'Num_Reserva')}":_gs_val(r,"ID_Pago") for r in pag_cv if _gs_val(r,"ID_Pago")}
                    sel_p=st.selectbox("Pago",list(opts_p.keys()),key="del_cv_pag"); sel_p_k=opts_p[sel_p]
                with col_p2:
                    pin_p=st.text_input("PIN admin",type="password",key="del_cv_pag_pin")
                accion_p=st.radio("Acción",["Anular","Eliminar definitivamente"],key="del_cv_pag_ac",horizontal=True)
                if st.button("Aplicar acción pago",type="primary",use_container_width=True,key="btn_del_cv_pag"):
                    if not verificar_pin(pin_p,"admin"):
                        st.error("PIN incorrecto")
                    else:
                        try:
                            if "Anular" in accion_p:
                                _pg_update_col_conv("Pagos","ID_Pago",sel_p_k,"Estado","anulado")
                            else:
                                _pg_delete_row_conv("Pagos","ID_Pago",sel_p_k)
                        except Exception: pass
                        _gs_invalidate_cache("Pagos")
                        st.success("Acción aplicada.")
                        st.rerun()
        with stabs[3]:
            fact_cv = _pg_read_table_conv("Facturas")
            if not fact_cv:
                st.info("No hay facturas de convenio en jjgt_convenios.")
            else:
                _df_cv_fac = pd.DataFrame([{
                    "Número":  str(r.get("Num_Factura","")),
                    "Empresa": str(r.get("Nombre_Empresa","—")),
                    "Total":   fmt_cop(float(r.get("Total_COP","0") or 0)),
                    "Estado":  str(r.get("Estado","")),
                } for r in fact_cv])
                st.dataframe(_df_cv_fac, use_container_width=True, hide_index=True)
                col_f1,col_f2=st.columns([2,1])
                with col_f1:
                    opts_f={f"{r.get('Num_Factura','')}—{r.get('Nombre_Empresa','—')}":r.get("Num_Factura","") for r in fact_cv if r.get("Num_Factura")}
                    sel_f=st.selectbox("Factura",list(opts_f.keys()),key="del_cv_fac"); sel_f_k=opts_f[sel_f]
                with col_f2:
                    pin_f=st.text_input("PIN admin",type="password",key="del_cv_fac_pin")
                accion_f=st.radio("Acción",["Anular","Eliminar definitivamente"],key="del_cv_fac_ac",horizontal=True)
                if st.button("Aplicar acción factura",type="primary",use_container_width=True,key="btn_del_cv_fac"):
                    if not verificar_pin(pin_f,"admin"):
                        st.error("PIN incorrecto")
                    else:
                        try:
                            if "Anular" in accion_f:
                                _pg_update_col_conv("Facturas","Num_Factura",sel_f_k,"Estado","anulada")
                            else:
                                _pg_delete_row_conv("Facturas","Num_Factura",sel_f_k)
                        except Exception: pass
                        st.success("Acción aplicada en PostgreSQL (convenios).")
                        st.rerun()



# ══════════════════════════════════════════════════════════════════════════════
# MAIN — ROUTER
# ══════════════════════════════════════════════════════════════════════════════

def main():

    # ── Contexto para módulos externos ────────────────────────────────────────
    if FC_AVAILABLE and _fc_mod is not None:
        _fc_mod.set_context(globals())
    # ─────────────────────────────────────────────────────────────────────────

    # ── Contexto para módulos contable y factura electrónica ─────────────────
    if CONT_AVAILABLE and _cont_mod is not None:
        _cont_mod.set_context(globals())
    if FE_AVAILABLE and _fe_mod is not None:
        _fe_mod.set_context(globals())
    # ─────────────────────────────────────────────────────────────────────────

    init_state()
    # Verificar conexión a Google Sheets al arrancar
    if GSPREAD_AVAILABLE:
        try:
            _, sh_check = get_active_client()
            if not sh_check:
                st.warning(
                    "⚠️ **Sin conexión a Google Sheets.** "
                    "Configura las credenciales en ⚙️ Configuración → Google Sheets. "
                    "Sin esta conexión la app no podrá mostrar ni guardar datos."
                )
        except Exception as _conn_err:
            # Error de red transitorio al arrancar — no bloquear la UI
            st.toast(
                "⚠️ Conexión a Google Sheets inestable al arrancar. "
                "La app reintentará en la siguiente operación.",
                icon="🔄"
            )
    _ensure_sync_thread()  # stub de compatibilidad

    pantalla = st.session_state.pantalla
    router = {
        "operador_login": show_operador_login,
        "operador":       show_operador,
        "bienvenida":     show_bienvenida,
        "seleccion":      show_seleccion,
        "datos":          show_datos,
        "pago":           show_pago,
        "confirmacion":   show_confirmacion,
        "voucher":        show_voucher,
    }
    router.get(pantalla, show_operador_login)()


if __name__ == "__main__":
    main()
