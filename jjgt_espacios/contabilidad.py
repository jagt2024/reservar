# ══════════════════════════════════════════════════════════════════════════════
#  SUITE SALITRE · Espacios de Descanso Personal — Terminal de Transportes
#  MÓDULO CONTABLE — contabilidad.py
#  Motor de comprobantes contables automáticos (Colombia · PUC)
#
#  DOBLE SPREADSHEET:
#    origen="pagos"     → escribe/lee en jjgt_pagos     (get_active_client)
#    origen="convenios" → escribe/lee en jjgt_convenios  (get_active_client_convenios)
#
#  Todas las funciones públicas aceptan el parámetro: origen: str = "pagos"
#  Los hooks on_reserva_creada y on_pago_convenio lo pasan automáticamente.
# ══════════════════════════════════════════════════════════════════════════════

from __future__ import annotations

import io
import threading
from datetime import datetime
from typing import Optional

import pytz

# ──────────────────────────────────────────────────────────────────────────────
# DEPENDENCIAS OPCIONALES
# ──────────────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────
TZ_COL = pytz.timezone("America/Bogota")

# Nombre de la hoja dentro de cada spreadsheet
GS_HOJA_COMP = "Comprobantes_Contables"

# Orígenes válidos
ORIGEN_PAGOS     = "pagos"
ORIGEN_CONVENIOS = "convenios"

# Prefijos de numeración
_PREFIJOS = {
    "ingreso":  "CI",
    "egreso":   "CE",
    "diario":   "CD",
    "nomina":   "CN",
    "factura":  "CF",
}

# Columnas del libro de comprobantes
_COLUMNAS = [
    "Fecha", "Numero", "Tipo", "Evento", "Tercero", "Descripcion",
    "Cuenta_Debito", "Nombre_Debito", "Cuenta_Credito", "Nombre_Credito",
    "Valor_COP", "Medio_Pago", "Soporte", "Operador", "Observaciones",
]

# Catálogo PUC Colombia
PLAN_CUENTAS: dict[str, str] = {
    "110505": "Caja General",
    "111005": "Bancos – Cja Social",
    "130505": "Clientes Nacionales",
    "220505": "Proveedores Nacionales",
    "236505": "IVA Generado x Pagar",
    "236805": "Retención Fte x Pagar",
    "236540": "ICA x Pagar",
    "410510": "Ingresos – Alojamiento",
    "511010": "Gasto Personal – Sueldos",
    "511505": "Gasto Personal – Prestaciones",
    "519515": "Otros Gastos Operac.",
    "516010": "Depreciación PPE",
    "159210": "Deprec. Acumulada PPE",
    "282005": "Nómina por Pagar",
}

# Reglas contables por evento
REGLAS_CONTABLES: dict[str, tuple[str, str]] = {
    "nueva_reserva":    ("130505", "410510"),
    "pago_efectivo":    ("110505", "130505"),
    "pago_banco":       ("111005", "130505"),
    "pago_digital":     ("111005", "130505"),
    "pago_convenio":    ("130505", "410510"),
    "pago_proveedor":   ("220505", "111005"),
    "iva_ventas":       ("410510", "236505"),
    "nomina_causacion": ("511010", "282005"),
    "nomina_pago":      ("282005", "111005"),
    "depreciacion":     ("516010", "159210"),
    "ajuste_diario":    ("519515", "111005"),
}

# Contexto inyectado desde pagos_convenios.py
_ctx: dict = {}
_ctx_lock = threading.Lock()


# ══════════════════════════════════════════════════════════════════════════════
# API DE CONTEXTO
# ══════════════════════════════════════════════════════════════════════════════

def set_context(globals_dict: dict) -> None:
    """
    Inyecta el contexto de pagos_convenios.py.
    Llamar desde main() de pagos_convenios.py:
        import contabilidad
        contabilidad.set_context(globals())
    """
    with _ctx_lock:
        _ctx.update(globals_dict)


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES INTERNAS
# ══════════════════════════════════════════════════════════════════════════════

def _ahora() -> datetime:
    return datetime.now(TZ_COL)


def _fecha_str() -> str:
    return _ahora().strftime("%Y-%m-%d")


def _nombre_cuenta(codigo: str) -> str:
    return PLAN_CUENTAS.get(codigo, codigo)


def _get_spreadsheet(origen: str):
    """
    Retorna el spreadsheet correcto según el origen.
    origen="pagos"     → get_active_client()           → jjgt_pagos
    origen="convenios" → get_active_client_convenios()  → jjgt_convenios
    Retorna None si no hay cliente disponible.
    """
    if origen == ORIGEN_CONVENIOS:
        fn = _ctx.get("get_active_client_convenios")
    else:
        fn = _ctx.get("get_active_client")

    if not fn:
        print(f"[contabilidad] WARN: get_active_client{'_convenios' if origen == ORIGEN_CONVENIOS else ''} no está en contexto.")
        return None

    try:
        _, sh = fn()
        return sh
    except Exception as e:
        print(f"[contabilidad] ERROR obteniendo spreadsheet ({origen}): {e}")
        return None


def _get_read_fn(origen: str):
    """
    Retorna la función de lectura correcta según el origen.
    origen="pagos"     → _gs_read_sheet
    origen="convenios" → _gs_read_sheet_conv
    """
    if origen == ORIGEN_CONVENIOS:
        return _ctx.get("_gs_read_sheet_conv")
    return _ctx.get("_gs_read_sheet")


def _consecutivo(tipo: str, origen: str = ORIGEN_PAGOS) -> str:
    """Genera el número correlativo leyendo del spreadsheet correcto."""
    prefijo = _PREFIJOS.get(tipo, "CO")
    n = 1
    try:
        fn_read = _get_read_fn(origen)
        fn_val  = _ctx.get("_gs_val")
        if fn_read and fn_val:
            rows = fn_read(GS_HOJA_COMP, force=True)
            mismo_tipo = [r for r in rows if fn_val(r, "Tipo", "") == tipo]
            n = len(mismo_tipo) + 1
    except Exception:
        pass
    return f"{prefijo}-{n:04d}"


def _escribir_comprobante(fila: list, origen: str = ORIGEN_PAGOS) -> None:
    """
    Escribe una fila en Comprobantes_Contables del spreadsheet indicado.
    origen="pagos"     → jjgt_pagos
    origen="convenios" → jjgt_convenios
    """
    try:
        fn_append = _ctx.get("_gs_append")
        if not fn_append:
            print(f"[contabilidad] WARN: _gs_append no está en contexto — comprobante no guardado.")
            return

        sh = _get_spreadsheet(origen)
        if not sh:
            print(f"[contabilidad] WARN: sin spreadsheet ({origen}) — comprobante no guardado: {fila[1] if len(fila) > 1 else '?'}")
            return

        fn_append(sh, GS_HOJA_COMP, fila)

        fn_inv = _ctx.get("_gs_invalidate_cache")
        if fn_inv:
            # Invalidar caché del spreadsheet correcto
            if origen == ORIGEN_CONVENIOS:
                # _gs_invalidate_cache no distingue spreadsheet; invalidar manualmente
                try:
                    import streamlit as st
                    st.session_state[f"_gs_conv_ts_{GS_HOJA_COMP}"] = 0
                except Exception:
                    pass
            else:
                fn_inv(GS_HOJA_COMP)

    except Exception as e:
        print(f"[contabilidad] ERROR al escribir comprobante ({origen}): {e}")
        try:
            import streamlit as st
            st.error(f"⚠️ Error al guardar comprobante contable ({origen}): {e}")
        except Exception:
            pass


def _operador_actual() -> str:
    try:
        import streamlit as st
        return st.session_state.get("operador_info", {}).get("nombre", "sistema")
    except Exception:
        return "sistema"


def _medio_a_evento(metodo_pago: str) -> str:
    m = (metodo_pago or "").lower()
    if m == "efectivo":
        return "pago_efectivo"
    if m == "convenio":
        return "pago_convenio"
    return "pago_banco"


def _cuentas_para_evento(evento: str) -> tuple[str, str]:
    debito, credito = REGLAS_CONTABLES.get(evento, ("519515", "111005"))
    return debito, credito


# ══════════════════════════════════════════════════════════════════════════════
# COMPROBANTES PÚBLICOS
# Todos aceptan origen: str = "pagos" | "convenios"
# ══════════════════════════════════════════════════════════════════════════════

def comp_ingreso_pago(
    tercero: str,
    valor: float,
    metodo_pago: str,
    soporte: str,
    descripcion: str = "",
    observaciones: str = "",
    origen: str = ORIGEN_PAGOS,
) -> str:
    """Registra el ingreso de caja/banco cuando el cliente paga."""
    tipo   = "ingreso"
    evento = _medio_a_evento(metodo_pago)
    numero = _consecutivo(tipo, origen)
    deb, cred = _cuentas_para_evento(evento)

    desc = descripcion or f"Ingreso {metodo_pago} — reserva {soporte}"
    fila = [
        _fecha_str(), numero, tipo, evento,
        tercero, desc,
        deb, _nombre_cuenta(deb),
        cred, _nombre_cuenta(cred),
        round(valor, 2), metodo_pago, soporte,
        _operador_actual(), observaciones,
    ]
    _escribir_comprobante(fila, origen)
    return numero


def comp_ingreso_factura(
    tercero: str,
    valor_subtotal: float,
    valor_iva: float,
    num_factura: str,
    descripcion: str = "",
    observaciones: str = "",
    origen: str = ORIGEN_PAGOS,
) -> str:
    """Causación contable al emitir una factura de venta (2 líneas: subtotal + IVA)."""
    tipo   = "factura"
    numero = _consecutivo(tipo, origen)
    desc   = descripcion or f"Causación factura {num_factura}"

    # Línea 1: causación del ingreso
    deb, cred = _cuentas_para_evento("nueva_reserva")
    fila1 = [
        _fecha_str(), numero, tipo, "nueva_reserva",
        tercero, desc,
        deb, _nombre_cuenta(deb),
        cred, _nombre_cuenta(cred),
        round(valor_subtotal, 2), "—", num_factura,
        _operador_actual(), observaciones,
    ]
    _escribir_comprobante(fila1, origen)

    # Línea 2: IVA
    if valor_iva and valor_iva > 0:
        deb_iva, cred_iva = _cuentas_para_evento("iva_ventas")
        numero_iva = f"{numero}-IVA"
        fila2 = [
            _fecha_str(), numero_iva, tipo, "iva_ventas",
            tercero, f"IVA factura {num_factura}",
            deb_iva, _nombre_cuenta(deb_iva),
            cred_iva, _nombre_cuenta(cred_iva),
            round(valor_iva, 2), "—", num_factura,
            _operador_actual(), observaciones,
        ]
        _escribir_comprobante(fila2, origen)

    return numero


def comp_egreso_proveedor(
    beneficiario: str,
    valor: float,
    descripcion: str,
    medio_pago: str,
    soporte: str,
    cuenta_gasto: str = "519515",
    observaciones: str = "",
    origen: str = ORIGEN_PAGOS,
) -> str:
    """Registra un pago a proveedor / gasto operacional."""
    tipo   = "egreso"
    numero = _consecutivo(tipo, origen)
    deb, cred = cuenta_gasto, "111005"

    fila = [
        _fecha_str(), numero, tipo, "pago_proveedor",
        beneficiario, descripcion,
        deb, _nombre_cuenta(deb),
        cred, _nombre_cuenta(cred),
        round(valor, 2), medio_pago, soporte,
        _operador_actual(), observaciones,
    ]
    _escribir_comprobante(fila, origen)
    return numero


def comp_diario_ajuste(
    descripcion: str,
    cuenta_debito: str,
    cuenta_credito: str,
    valor: float,
    soporte: str = "Ajuste contable",
    observaciones: str = "",
    origen: str = ORIGEN_PAGOS,
) -> str:
    """Comprobante diario para ajustes manuales."""
    tipo   = "diario"
    numero = _consecutivo(tipo, origen)

    fila = [
        _fecha_str(), numero, tipo, "ajuste_diario",
        "SUITE SALITRE", descripcion,
        cuenta_debito,  _nombre_cuenta(cuenta_debito),
        cuenta_credito, _nombre_cuenta(cuenta_credito),
        round(valor, 2), "—", soporte,
        _operador_actual(), observaciones,
    ]
    _escribir_comprobante(fila, origen)
    return numero


def comp_nomina(
    periodo: str,
    total_sueldos: float,
    total_prestaciones: float,
    total_pagado: float,
    observaciones: str = "",
    origen: str = ORIGEN_PAGOS,
) -> str:
    """Comprobante de nómina en tres pasos: sueldos, prestaciones y pago."""
    tipo   = "nomina"
    numero = _consecutivo(tipo, origen)

    if total_sueldos > 0:
        fila1 = [
            _fecha_str(), numero, tipo, "nomina_causacion",
            "Personal Nómina", f"Sueldos nómina {periodo}",
            "511010", _nombre_cuenta("511010"),
            "282005", _nombre_cuenta("282005"),
            round(total_sueldos, 2), "—", f"Nómina {periodo}",
            _operador_actual(), observaciones,
        ]
        _escribir_comprobante(fila1, origen)

    if total_prestaciones > 0:
        num2 = f"{numero}-PS"
        fila2 = [
            _fecha_str(), num2, tipo, "nomina_causacion",
            "Personal Nómina", f"Prestaciones nómina {periodo}",
            "511505", _nombre_cuenta("511505"),
            "282005", _nombre_cuenta("282005"),
            round(total_prestaciones, 2), "—", f"Nómina {periodo}",
            _operador_actual(), observaciones,
        ]
        _escribir_comprobante(fila2, origen)

    if total_pagado > 0:
        num3 = f"{numero}-PAGO"
        fila3 = [
            _fecha_str(), num3, tipo, "nomina_pago",
            "Personal Nómina", f"Pago nómina {periodo}",
            "282005", _nombre_cuenta("282005"),
            "111005", _nombre_cuenta("111005"),
            round(total_pagado, 2), "Transferencia", f"Nómina {periodo}",
            _operador_actual(), observaciones,
        ]
        _escribir_comprobante(fila3, origen)

    return numero


# ══════════════════════════════════════════════════════════════════════════════
# HOOKS DE INTEGRACIÓN CON pagos_convenios.py
# ══════════════════════════════════════════════════════════════════════════════

def on_reserva_creada(voucher: dict, calc: dict, cliente: dict, metodo: str) -> None:
    """
    Hook para reservas normales (jjgt_pagos).
    Llamar al final de crear_reserva_completa() en pagos_convenios.py:

        if CONT_AVAILABLE and _cont_mod:
            _cont_mod.on_reserva_creada(voucher, calc, cliente, metodo)
    """
    try:
        num_res  = voucher.get("numero_reserva", "")
        num_fact = voucher.get("numero_factura", "")
        nombre   = cliente.get("nombre", "Cliente")

        comp_ingreso_factura(
            tercero        = nombre,
            valor_subtotal = calc.get("subtotal", 0),
            valor_iva      = calc.get("iva", 0),
            num_factura    = num_fact,
            descripcion    = (
                f"Factura {num_fact} · Reserva {num_res} · "
                f"Cubículo {voucher.get('cubiculo','')} · "
                f"{calc.get('horas',0)}h"
            ),
            origen = ORIGEN_PAGOS,
        )

        comp_ingreso_pago(
            tercero     = nombre,
            valor       = calc.get("total", 0),
            metodo_pago = metodo,
            soporte     = num_res,
            descripcion = f"Pago {metodo} · Reserva {num_res} · Factura {num_fact}",
            origen      = ORIGEN_PAGOS,
        )
    except Exception as e:
        print(f"[contabilidad] ERROR on_reserva_creada: {e}")


def on_reserva_convenio_creada(voucher: dict, calc: dict, cliente: dict, metodo: str) -> None:
    """
    Hook para reservas bajo convenio empresarial (jjgt_convenios).
    Llamar al final de crear_reserva_convenio() en pagos_convenios.py:

        if CONT_AVAILABLE and _cont_mod:
            _cont_mod.on_reserva_convenio_creada(voucher, calc, cliente, metodo)
    """
    try:
        num_res  = voucher.get("numero_reserva", "")
        num_fact = voucher.get("numero_factura", "")
        nombre   = cliente.get("nombre", "Cliente")

        comp_ingreso_factura(
            tercero        = nombre,
            valor_subtotal = calc.get("subtotal", 0),
            valor_iva      = calc.get("iva", 0),
            num_factura    = num_fact,
            descripcion    = (
                f"Factura {num_fact} · Reserva {num_res} · "
                f"Cubículo {voucher.get('cubiculo','')} · "
                f"{calc.get('horas',0)}h [Convenio]"
            ),
            origen = ORIGEN_CONVENIOS,
        )

        comp_ingreso_pago(
            tercero     = nombre,
            valor       = calc.get("total", 0),
            metodo_pago = metodo,
            soporte     = num_res,
            descripcion = f"Pago {metodo} · Reserva {num_res} · Factura {num_fact} [Convenio]",
            origen      = ORIGEN_CONVENIOS,
        )
    except Exception as e:
        print(f"[contabilidad] ERROR on_reserva_convenio_creada: {e}")


def on_pago_convenio(
    empresa: str,
    valor: float,
    num_reserva: str,
    num_factura: str = "",
) -> None:
    """
    Hook para pagos bajo convenio empresarial (jjgt_convenios).
    Llamar desde crear_reserva_convenio() en pagos_convenios.py:

        if CONT_AVAILABLE and _cont_mod:
            _cont_mod.on_pago_convenio(empresa, valor, num_reserva, num_factura)
    """
    try:
        comp_ingreso_pago(
            tercero       = empresa,
            valor         = valor,
            metodo_pago   = "Convenio",
            soporte       = num_reserva,
            descripcion   = f"Convenio {empresa} · Reserva {num_reserva}",
            observaciones = f"Factura convenio: {num_factura}",
            origen        = ORIGEN_CONVENIOS,
        )
    except Exception as e:
        print(f"[contabilidad] ERROR on_pago_convenio: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def exportar_libro_contable(
    fecha_ini: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    origen: str = ORIGEN_PAGOS,
) -> Optional[bytes]:
    """
    Lee comprobantes del spreadsheet indicado y genera un libro Excel descargable.
    origen="pagos"     → lee de jjgt_pagos
    origen="convenios" → lee de jjgt_convenios
    """
    if not OPENPYXL_OK:
        return None

    filas: list[list] = []
    try:
        fn_read  = _get_read_fn(origen)
        fn_val   = _ctx.get("_gs_val")
        fn_float = _ctx.get("_gs_float")
        if fn_read and fn_val:
            rows = fn_read(GS_HOJA_COMP, force=True)
            for r in rows:
                fecha = fn_val(r, "Fecha", "")
                if fecha_ini and fecha < fecha_ini:
                    continue
                if fecha_fin and fecha > fecha_fin:
                    continue
                filas.append([
                    fecha,
                    fn_val(r, "Numero", ""),
                    fn_val(r, "Tipo", ""),
                    fn_val(r, "Evento", ""),
                    fn_val(r, "Tercero", ""),
                    fn_val(r, "Descripcion", ""),
                    fn_val(r, "Cuenta_Debito", ""),
                    fn_val(r, "Nombre_Debito", ""),
                    fn_val(r, "Cuenta_Credito", ""),
                    fn_val(r, "Nombre_Credito", ""),
                    fn_float(r, "Valor_COP", 0) if fn_float else float(fn_val(r, "Valor_COP", 0) or 0),
                    fn_val(r, "Medio_Pago", ""),
                    fn_val(r, "Soporte", ""),
                    fn_val(r, "Operador", ""),
                    fn_val(r, "Observaciones", ""),
                ])
    except Exception as e:
        print(f"[contabilidad] ERROR leyendo comprobantes ({origen}): {e}")

    if not filas:
        return None

    wb    = Workbook()
    label = "Convenios" if origen == ORIGEN_CONVENIOS else "Pagos"

    C_AZUL_OSC  = "050B1A"
    C_AZUL_CARD = "0D1F3C"
    C_CYAN      = "00D4FF"
    C_VERDE     = "00FF88"
    C_BLANCO    = "E2E8F0"
    C_GRIS      = "94A3B8"
    C_AMARILLO  = "FFD32A"

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=C_BLANCO, size=10) -> Font:
        return Font(name="Arial", bold=bold, color=color, size=size)

    def _border() -> Border:
        thin = Side(style="thin", color="1E3A5F")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _money_fmt() -> str:
        return '#,##0.00'

    TIPO_COLORES = {
        "ingreso":  C_VERDE,
        "egreso":   "FF4757",
        "diario":   C_AMARILLO,
        "nomina":   "A29BFE",
        "factura":  "74B9FF",
    }

    # ── Hoja 1: Comprobantes ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Comprobantes"

    ws1.merge_cells("A1:O1")
    ws1["A1"] = f"SUITE SALITRE · Libro de Comprobantes Contables — {label}"
    ws1["A1"].font      = _font(bold=True, color=C_CYAN, size=13)
    ws1["A1"].fill      = _fill(C_AZUL_OSC)
    ws1["A1"].alignment = _center()
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:O2")
    rango_txt = (
        f"Período: {fecha_ini or '—'}  a  {fecha_fin or '—'}    "
        f"Generado: {_ahora().strftime('%Y-%m-%d %H:%M')} (hora Colombia)"
    )
    ws1["A2"] = rango_txt
    ws1["A2"].font      = _font(color=C_GRIS, size=9)
    ws1["A2"].fill      = _fill(C_AZUL_CARD)
    ws1["A2"].alignment = _center()
    ws1.row_dimensions[2].height = 18

    encabezados = [
        "Fecha", "Número", "Tipo", "Evento", "Tercero", "Descripción",
        "Cta. Débito", "Nombre Débito", "Cta. Crédito", "Nombre Crédito",
        "Valor COP", "Medio Pago", "Soporte", "Operador", "Observaciones",
    ]
    anchos = [12, 10, 10, 18, 24, 36, 11, 24, 11, 24, 16, 14, 18, 18, 28]

    for col, (hdr, ancho) in enumerate(zip(encabezados, anchos), start=1):
        cell = ws1.cell(row=3, column=col, value=hdr)
        cell.font      = _font(bold=True, color=C_AZUL_OSC, size=10)
        cell.fill      = _fill(C_CYAN)
        cell.alignment = _center()
        cell.border    = _border()
        ws1.column_dimensions[get_column_letter(col)].width = ancho
    ws1.row_dimensions[3].height = 22

    for i, fila in enumerate(filas, start=4):
        tipo_fila = fila[2] if len(fila) > 2 else ""
        bg        = TIPO_COLORES.get(tipo_fila, C_AZUL_CARD)
        fill_row  = _fill(C_AZUL_CARD) if i % 2 == 0 else _fill("0A1628")
        for col, val in enumerate(fila, start=1):
            cell = ws1.cell(row=i, column=col, value=val)
            if col == 3:
                cell.font = _font(bold=True, color=bg, size=9)
                cell.fill = fill_row
            elif col == 11:
                cell.font          = _font(bold=True, color=C_VERDE, size=10)
                cell.fill          = fill_row
                cell.number_format = _money_fmt()
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            else:
                cell.font      = _font(size=9)
                cell.fill      = fill_row
                cell.alignment = _left()
            cell.border = _border()
        ws1.row_dimensions[i].height = 16

    fila_tot = len(filas) + 4
    ws1.merge_cells(f"A{fila_tot}:J{fila_tot}")
    ws1[f"A{fila_tot}"]           = "TOTAL COMPROBANTES"
    ws1[f"A{fila_tot}"].font      = _font(bold=True, color=C_AMARILLO, size=10)
    ws1[f"A{fila_tot}"].fill      = _fill(C_AZUL_OSC)
    ws1[f"A{fila_tot}"].alignment = _center()
    ws1[f"K{fila_tot}"]           = f"=SUM(K4:K{fila_tot-1})"
    ws1[f"K{fila_tot}"].font      = _font(bold=True, color=C_AMARILLO, size=11)
    ws1[f"K{fila_tot}"].fill      = _fill(C_AZUL_OSC)
    ws1[f"K{fila_tot}"].number_format = _money_fmt()
    ws1[f"K{fila_tot}"].alignment = Alignment(horizontal="right", vertical="center")
    ws1[f"K{fila_tot}"].border    = _border()
    ws1.freeze_panes = "A4"

    # ── Hoja 2: Resumen ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    ws2["A1"]           = f"Resumen por Tipo — {label}"
    ws2["A1"].font      = _font(bold=True, color=C_CYAN, size=12)
    ws2["A1"].fill      = _fill(C_AZUL_OSC)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 26

    for col, h in enumerate(["Tipo", "Cantidad", "Total COP", "% del Total"], start=1):
        c            = ws2.cell(row=2, column=col, value=h)
        c.font       = _font(bold=True, color=C_AZUL_OSC)
        c.fill       = _fill(C_CYAN)
        c.alignment  = _center()
        c.border     = _border()
    for col, w in zip("ABCD", [14, 12, 18, 14]):
        ws2.column_dimensions[col].width = w

    totales_tipo: dict[str, list] = {}
    for fila in filas:
        t = fila[2] if len(fila) > 2 else "—"
        v = fila[10] if len(fila) > 10 else 0
        if t not in totales_tipo:
            totales_tipo[t] = [0, 0.0]
        totales_tipo[t][0] += 1
        totales_tipo[t][1] += float(v or 0)

    gran_total = sum(v[1] for v in totales_tipo.values()) or 1
    for i, (tipo, (cnt, tot)) in enumerate(sorted(totales_tipo.items()), start=3):
        bg = TIPO_COLORES.get(tipo, C_AZUL_CARD)
        ws2.cell(row=i, column=1, value=tipo.upper()).font = _font(bold=True, color=bg)
        ws2.cell(row=i, column=2, value=cnt).font = _font()
        c_tot = ws2.cell(row=i, column=3, value=tot)
        c_tot.font = _font(color=C_VERDE, bold=True)
        c_tot.number_format = _money_fmt()
        pct = ws2.cell(row=i, column=4, value=round(tot / gran_total * 100, 2))
        pct.number_format = "0.00%"
        pct.font = _font()
        for col in range(1, 5):
            ws2.cell(row=i, column=col).fill      = _fill(C_AZUL_CARD)
            ws2.cell(row=i, column=col).border    = _border()
            ws2.cell(row=i, column=col).alignment = _center()

    fila_gt = len(totales_tipo) + 3
    ws2.cell(row=fila_gt, column=1, value="TOTAL").font = _font(bold=True, color=C_AMARILLO)
    ws2.cell(row=fila_gt, column=2, value=len(filas)).font = _font(bold=True, color=C_AMARILLO)
    c_gt = ws2.cell(row=fila_gt, column=3, value=gran_total)
    c_gt.font = _font(bold=True, color=C_AMARILLO)
    c_gt.number_format = _money_fmt()
    ws2.cell(row=fila_gt, column=4, value=1.0).number_format = "0.00%"
    for col in range(1, 5):
        ws2.cell(row=fila_gt, column=col).fill      = _fill(C_AZUL_OSC)
        ws2.cell(row=fila_gt, column=col).border    = _border()
        ws2.cell(row=fila_gt, column=col).alignment = _center()

    # ── Hoja 3: Balance de cuentas ───────────────────────────────────────────
    ws3 = wb.create_sheet("Balance_Cuentas")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    ws3["A1"]           = f"Balance de Cuentas del Período — {label}"
    ws3["A1"].font      = _font(bold=True, color=C_CYAN, size=12)
    ws3["A1"].fill      = _fill(C_AZUL_OSC)
    ws3["A1"].alignment = _center()
    ws3.row_dimensions[1].height = 26

    for col, h in enumerate(["Cuenta", "Nombre", "Total Débitos COP", "Total Créditos COP", "Saldo COP"], start=1):
        c            = ws3.cell(row=2, column=col, value=h)
        c.font       = _font(bold=True, color=C_AZUL_OSC)
        c.fill       = _fill(C_CYAN)
        c.alignment  = _center()
        c.border     = _border()
    for col, w in zip("ABCDE", [12, 28, 20, 20, 18]):
        ws3.column_dimensions[col].width = w

    balance: dict[str, dict[str, float]] = {}
    for fila in filas:
        cta_deb  = str(fila[6])  if len(fila) > 6  else ""
        cta_cred = str(fila[8])  if len(fila) > 8  else ""
        valor    = float(fila[10]) if len(fila) > 10 else 0.0
        for cta in set([cta_deb, cta_cred]):
            if not cta or cta == "—":
                continue
            if cta not in balance:
                balance[cta] = {"debito": 0.0, "credito": 0.0}
        if cta_deb and cta_deb != "—":
            balance[cta_deb]["debito"] += valor
        if cta_cred and cta_cred != "—":
            balance[cta_cred]["credito"] += valor

    for i, (cta, vals) in enumerate(sorted(balance.items()), start=3):
        saldo      = vals["debito"] - vals["credito"]
        nombre_cta = PLAN_CUENTAS.get(cta, "—")
        fill_row   = _fill(C_AZUL_CARD) if i % 2 == 0 else _fill("0A1628")
        color_saldo = C_VERDE if saldo >= 0 else "FF4757"

        ws3.cell(row=i, column=1, value=cta).font = _font(bold=True, color=C_CYAN)
        ws3.cell(row=i, column=2, value=nombre_cta).font = _font()
        c_deb = ws3.cell(row=i, column=3, value=vals["debito"])
        c_deb.number_format = _money_fmt()
        c_deb.font = _font()
        c_cred = ws3.cell(row=i, column=4, value=vals["credito"])
        c_cred.number_format = _money_fmt()
        c_cred.font = _font()
        c_sal = ws3.cell(row=i, column=5, value=saldo)
        c_sal.number_format = _money_fmt()
        c_sal.font = _font(bold=True, color=color_saldo)
        for col in range(1, 6):
            ws3.cell(row=i, column=col).fill      = fill_row
            ws3.cell(row=i, column=col).border    = _border()
            ws3.cell(row=i, column=col).alignment = _center()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
# PANEL STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════

def render_panel_contabilidad() -> None:
    """
    Renderiza el panel de contabilidad con soporte dual (jjgt_pagos / jjgt_convenios).
    Agregar en show_operador() de pagos_convenios.py:

        elif menu == "📒 Contabilidad":
            from contabilidad import render_panel_contabilidad
            render_panel_contabilidad()
    """
    try:
        import streamlit as st
    except ImportError:
        return

    fn_val   = _ctx.get("_gs_val")
    fn_float = _ctx.get("_gs_float")
    fn_fmt   = _ctx.get("fmt_cop", lambda v: f"${v:,.0f}".replace(",", "."))
    fn_ahora = _ctx.get("ahora_col", _ahora)

    st.markdown("## 📒 Módulo Contable — Comprobantes")
    st.caption(
        "Los movimientos generan comprobantes automáticamente en cada spreadsheet. "
        "Selecciona el origen para ver o exportar los registros correspondientes."
    )

    # ── Selector de origen ───────────────────────────────────────────────────
    origen_label = st.radio(
        "📂 Origen de datos:",
        ["🏠 Pagos (jjgt_pagos)", "🤝 Convenios (jjgt_convenios)"],
        horizontal=True,
        key="cont_origen_sel",
    )
    origen = ORIGEN_CONVENIOS if "Convenios" in origen_label else ORIGEN_PAGOS
    fn_read = _get_read_fn(origen)

    st.divider()

    tab1, tab2, tab3 = st.tabs(
        ["📋 Ver comprobantes", "➕ Comprobante manual", "📥 Exportar libro"]
    )

    # ── TAB 1: Ver comprobantes ──────────────────────────────────────────────
    with tab1:
        col_reload, _ = st.columns([1, 5])
        with col_reload:
            if st.button("🔄 Recargar", key="btn_reload_comp"):
                try:
                    if origen == ORIGEN_CONVENIOS:
                        st.session_state[f"_gs_conv_ts_{GS_HOJA_COMP}"] = 0
                    else:
                        fn_inv = _ctx.get("_gs_invalidate_cache")
                        if fn_inv:
                            fn_inv(GS_HOJA_COMP)
                except Exception:
                    pass
                st.rerun()

        try:
            if fn_read and fn_val and fn_float:
                rows_comp = fn_read(GS_HOJA_COMP, force=True)
                if not rows_comp:
                    st.info(
                        f"No hay comprobantes registrados en **{origen_label}** aún. "
                        "Se generarán automáticamente al procesar reservas y pagos."
                    )
                else:
                    import pandas as pd
                    data = []
                    for r in rows_comp:
                        data.append({
                            "Fecha":       fn_val(r, "Fecha", ""),
                            "Número":      fn_val(r, "Numero", ""),
                            "Tipo":        fn_val(r, "Tipo", ""),
                            "Tercero":     fn_val(r, "Tercero", ""),
                            "Descripción": fn_val(r, "Descripcion", ""),
                            "Cta. Deb.":   fn_val(r, "Cuenta_Debito", ""),
                            "Cta. Cred.":  fn_val(r, "Cuenta_Credito", ""),
                            "Valor COP":   fn_float(r, "Valor_COP", 0),
                            "Soporte":     fn_val(r, "Soporte", ""),
                            "Operador":    fn_val(r, "Operador", ""),
                        })
                    df = pd.DataFrame(data)
                    st.dataframe(df, use_container_width=True, hide_index=True)

                    total = df["Valor COP"].sum() if not df.empty else 0
                    st.markdown(
                        f"<div style='text-align:right;font-weight:700;color:#00FF88;font-size:16px'>"
                        f"Total registrado: {fn_fmt(total)} COP &nbsp;·&nbsp; "
                        f"{len(df)} comprobantes</div>",
                        unsafe_allow_html=True,
                    )
            else:
                st.warning(
                    f"Función de lectura no disponible para **{origen_label}**. "
                    "Verifica que `contabilidad.set_context(globals())` esté en `main()`."
                )
        except Exception as e:
            st.error(f"Error al leer comprobantes ({origen}): {e}")

    # ── TAB 2: Comprobante manual ────────────────────────────────────────────
    with tab2:
        st.markdown(
            f"Registra un comprobante contable manual en **{origen_label}**."
        )

        tipo_man = st.selectbox(
            "Tipo de comprobante",
            ["Egreso / Pago proveedor", "Ajuste / Diario", "Nómina"],
            key="cont_tipo_man",
        )

        if tipo_man == "Egreso / Pago proveedor":
            col1, col2 = st.columns(2)
            with col1:
                ben    = st.text_input("Beneficiario / Proveedor", key="cont_ben")
                desc_e = st.text_input("Descripción del gasto", key="cont_desc_e")
                sop_e  = st.text_input("Soporte (N° factura proveedor)", key="cont_sop_e")
            with col2:
                val_e   = st.number_input("Valor COP", min_value=0, step=1000, key="cont_val_e")
                medio_e = st.selectbox("Medio de pago", ["Transferencia", "Efectivo", "Tarjeta"], key="cont_medio_e")
                cta_g   = st.text_input("Cuenta de gasto (PUC)", value="519515", key="cont_cta_g",
                                        help="Ej: 519515 = Otros gastos.")
            if st.button("💾 Registrar egreso", key="btn_cont_egreso", type="primary"):
                if not ben or not val_e:
                    st.error("Completa beneficiario y valor.")
                else:
                    num = comp_egreso_proveedor(ben, float(val_e), desc_e, medio_e, sop_e, cta_g, origen=origen)
                    st.success(f"✅ Comprobante de egreso registrado en **{origen_label}**: **{num}**")

        elif tipo_man == "Ajuste / Diario":
            col1, col2 = st.columns(2)
            with col1:
                desc_d  = st.text_input("Descripción", key="cont_desc_d")
                cta_deb = st.text_input("Cuenta Débito (PUC)", value="519515", key="cont_cta_deb")
                cta_cre = st.text_input("Cuenta Crédito (PUC)", value="111005", key="cont_cta_cre")
            with col2:
                val_d = st.number_input("Valor COP", min_value=0, step=1000, key="cont_val_d")
                sop_d = st.text_input("Soporte / referencia", key="cont_sop_d")
            st.caption(
                f"Débito → **{_nombre_cuenta(cta_deb)}** &nbsp;|&nbsp; "
                f"Crédito → **{_nombre_cuenta(cta_cre)}**"
            )
            if st.button("💾 Registrar ajuste", key="btn_cont_diario", type="primary"):
                if not desc_d or not val_d:
                    st.error("Completa descripción y valor.")
                else:
                    num = comp_diario_ajuste(desc_d, cta_deb, cta_cre, float(val_d), sop_d, origen=origen)
                    st.success(f"✅ Comprobante diario registrado en **{origen_label}**: **{num}**")

        elif tipo_man == "Nómina":
            col1, col2 = st.columns(2)
            with col1:
                periodo = st.text_input("Período de nómina", placeholder="Mayo 2025", key="cont_periodo")
                sueldos = st.number_input("Total sueldos COP", min_value=0, step=10000, key="cont_sueldos")
            with col2:
                presta = st.number_input("Total prestaciones COP", min_value=0, step=10000, key="cont_presta")
                pagado = st.number_input("Total pagado (transferencia) COP", min_value=0, step=10000, key="cont_pagado")
            if st.button("💾 Registrar nómina", key="btn_cont_nomina", type="primary"):
                if not periodo:
                    st.error("Ingresa el período.")
                else:
                    num = comp_nomina(periodo, float(sueldos), float(presta), float(pagado), origen=origen)
                    st.success(f"✅ Nómina registrada en **{origen_label}**: **{num}**")

        with st.expander("📖 Plan de Cuentas PUC (referencia rápida)"):
            import pandas as pd
            df_puc = pd.DataFrame(
                [{"Código": k, "Nombre": v} for k, v in PLAN_CUENTAS.items()]
            )
            st.dataframe(df_puc, use_container_width=True, hide_index=True)

    # ── TAB 3: Exportar libro ────────────────────────────────────────────────
    with tab3:
        st.markdown(f"Exporta el libro contable de **{origen_label}** a Excel (.xlsx).")
        now_col = fn_ahora()
        col_a, col_b = st.columns(2)
        with col_a:
            fi = st.date_input(
                "Fecha inicio",
                value=now_col.replace(day=1).date(),
                key="cont_fi",
            )
        with col_b:
            ff = st.date_input(
                "Fecha fin",
                value=now_col.date(),
                key="cont_ff",
            )

        if st.button("⚙️ Generar libro contable", key="btn_cont_export",
                     type="primary", use_container_width=True):
            with st.spinner("Generando libro Excel…"):
                xlsx_bytes = exportar_libro_contable(str(fi), str(ff), origen=origen)
            if xlsx_bytes:
                sufijo = "convenios" if origen == ORIGEN_CONVENIOS else "pagos"
                nombre_arch = (
                    f"libro_contable_SuiteSalitre_{sufijo}_"
                    f"{fi.strftime('%Y%m%d')}_{ff.strftime('%Y%m%d')}.xlsx"
                )
                st.download_button(
                    label="📥 Descargar libro contable Excel",
                    data=xlsx_bytes,
                    file_name=nombre_arch,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success(
                    f"✅ Libro generado: **{nombre_arch}** — "
                    f"Hojas: Comprobantes, Resumen, Balance_Cuentas"
                )
            else:
                st.warning(
                    f"No se encontraron comprobantes en **{origen_label}** "
                    "para el período seleccionado."
                )
