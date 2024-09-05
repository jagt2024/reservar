import streamlit as st
from streamlit_option_menu import option_menu
import numpy as np
from inicio_emp import InicioEmp
from crear_reserva_emp import CrearReservaEmp
from modificar_reserva_emp import ModificarReservaEmp
from eliminar_reserva_emp import EliminarReservaEmp
from servicios_emp import ServiciosEmp
from informacion_emp import InformacionEmp
from generar_excel_emp import GenerarExcelEmp
from generaQR.generar_qr import GenerarQr
from consulta_st_excel import ConsultarAgenda
from descargar_agenda_emp import download_and_process_data
from user_management import user_management_system, logout
import datetime as dt
from openpyxl import load_workbook
import os
from datetime import date
import datetime
import time
import pytz
import toml

# Cargar configuraciones desde config.toml
with open("./.streamlit/config.toml", "r") as f:
    config = toml.load(f)

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "5"
os.environ["REQUESTS_READ_TIMEOUT"] = "5"

datos_book_emp = load_workbook("archivos/parametros_empresa.xlsx", read_only=False)

def dataBook_emp(hoja):
    ws1 = datos_book_emp[hoja]
    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
      data.append(_row)
      #print(f'data {data}')
    return data

fecha_hasta = int('202401130')
#print(f'fecha hasta: {feha_hasta}')

fecha = dt.datetime.now()
fecha_hoy = int(fecha.strftime("%Y%m%d"))
#print(f'fecha_hoy: {fecha_hoy}')

empresa = dataBook_emp("sw")
result_emp = np.setdiff1d(empresa,'')
#print(f'sw_empresa {result_emp}')

sw_empresa = result_emp

#ws = result_emp
   #ws = result_emp["A2"].value
#print(ws)
#if sw_empresa == ["False"]:
    #empresa.cell(column=1, row=2).value = ["True"])
#    sw_empresa = ["True"]
#    print(sw_empresa)
  
page_title = 'Agenda Actividad' 
page_icon= "assets/barberia.png" 
title="Resevas"
layout = 'centered'

st.set_page_config(page_title=page_title, page_icon=page_icon,layout=layout)

class Model:
  
  menuTitle = "Reserve y Agende en Linea"
  option1 = 'Inicio'
  option2 = 'Crear Reserva'
  option10 = 'Descargar Agenda'
  option9  = 'Consultar Agenda'
  option3  = 'Modificar Reserva'
  option4  = 'Eliminar Reserva'
  option5  = 'Nuestros Servicios'
  option6  = 'Mas Informacion'
  option7  = 'Generar Archivos'
  option8  = 'Generar Codigo QR'
    
  #def __init__(self):
  #  self.apps=[]
    
  def add_app(self,title, function):
      self.apps.append({
         "title":title,
         "function":function
       })
      
  def css_load(css_file):
    
    #file_path = r"C:/Users/hp  pc/Desktop/Programas practica Python/App - Reservas/style/main.css"
    #print(file_path)

    #try:
    #  if os.path.exists(file_path):
    #    pass
        #os.chmod(file_path, 0o666)
        #print("File permissions modified successfully!")
    #  else:
    #    print("File not found:", file_path)
    #except PermissionError:
    #  print("Permission denied: You don't have the necessary permissions to change the permissions of this file.")
    
    try:
      with open(css_file, "r") as file:
        st.markdown(f"<style>{file.read()}</style>",unsafe_allow_html=True)
    except IOError  as e:
      if "not readable" in str(e):
        print("Error no readable check mode")
      else:
        print(f"IOError {e}")
      
  css_load(r"style/main.css")

if fecha_hasta < fecha_hoy:
  
   sw_empresa == ['False']
      
   st.warning('Ha caducado el tiempo autorizado para su uso favor comuniquese con el administrador')

else:  

  def view(model):
    try:
  
      with st.sidebar:
    
        app = option_menu(model.menuTitle,
                         [model.option1, model.option2,model.option10,model.option9,model.option3,model.option4,model.option5,model.option6,model.option7,model.option8],
                         icons=['bi bi-app-indicator',
                                'bi bi-calendar2-date', 
                                'bi bi-calendar2-date',
                                'bi bi-calendar2-date',
                                'bi bi-award',
                                'bi bi-clipboard-minus-fill'
                               ],
                         default_index=0,
                         styles= {
                           "container":{ "reservas": "5!important",
                           "background-color":'#acbee9'},
                           "icon":{"color":"white","font-size":"23px"},
                           "nav-lik":{"color":"white","font-size":"20px","text-aling":"left", "margin":"0px"},
                           "nav-lik-selected":{"backgroud-color":"#02ab21"},})
                       #orientation='horizontal')
      st.markdown(f"""
      <style>
            @import url('https://fonts.googleapis.com/css2?family={config['fonts']['clock_font']}:wght@400;500&display=swap');
            .clock {{
              font-family: '{config['fonts']['clock_font']}', sans-serif;
              font-size: {config['font_sizes']['clock_font_size']};
              color: {config['colors']['clock_color']};
              text-shadow: 0 0 10px rgba(255,255,0,0.7);
              margin-top: {config['layout']['top_margin']};
              margin-bottom: {config['layout']['bottom_margin']};
            }}
            .calendar {{
              font-family: '{config['fonts']['calendar_font']}', sans-serif;
              font-size: {config['font_sizes']['calendar_font_size']};
              color: {config['colors']['calendar_color']};
              margin-top: {config['layout']['top_margin']};
              margin-bottom: {config['layout']['bottom_margin']};
            }}
      </style>
      """, unsafe_allow_html=True)

      # Crear columnas para el diseño
      col1, col2 = st.columns(2)

      # Columna 1: Reloj Digital
      with col1:
        #st.header("Reloj Digital")
        clock_placeholder = st.empty()

        # Columna 2: Calendario
      with col2:
        #st.header("Calendario")
        calendar_placeholder = st.empty()
          
      def update_clock_and_calendar():
         while True:
              now = datetime.datetime.now(pytz.timezone('America/Bogota'))
              clock_placeholder.markdown(f'<p class="clock">Hora: {now.strftime("%H:%M:%S")}</p>', unsafe_allow_html=True)

              today = date.today()
              meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
              mes_es = meses_es[today.month - 1]
              dias_es = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
              dia_es = dias_es[today.weekday()]
        
              calendar_placeholder.markdown(f'<p class="calendar">Día: {dia_es.capitalize()} {today.day} de {mes_es} de {today.year}<br></p>', unsafe_allow_html=True)
        
              time.sleep(5)
              #st.experimental_rerun()      

      with st.sidebar:
        st.markdown("---")
        st.text("Version: 0.0.1")
        st.text("Ano: 2024")
        st.text("Autor: JAGT")
        st.markdown("---")
    
      if sw_empresa == ['True']:
      
        #st.title('***BARBERIA STYLOS***')
      
        if app == model.option1:
          InicioEmp().view(InicioEmp.Model())
        if app == model.option2:
          CrearReservaEmp().view(CrearReservaEmp.Model())
        if app == model.option3:
          ModificarReservaEmp().view(ModificarReservaEmp.Model())
        if app == model.option4:
          EliminarReservaEmp().view(EliminarReservaEmp.Model())
        if app == model.option5:
          ServiciosEmp().view(ServiciosEmp.Model())
        if app == model.option6:
          InformacionEmp().view(InformacionEmp.Model())
        if app == model.option7:
          GenerarExcelEmp().view(GenerarExcelEmp.Model())
        if app == model.option8:
          GenerarQr().view(GenerarQr.Model())
        if app == model.option9:
             ConsultarAgenda().view(ConsultarAgenda.Model())
        if app == model.option10:
           if user_management_system():
             download_and_process_data('./.streamlit/secrets.toml')
             logout()
         
    except SystemError as err:
      raise Exception(f'A ocurrido un error en main.py: {err}')
    
    
    update_clock_and_calendar()

  view(Model())