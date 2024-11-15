import streamlit as st
from google_sheets_emp import GoogleSheet
from google_calendar_emp import GoogleCalendar
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
import datetime
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import os 
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
import sys
import logging
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json
#import pywhatkit
#import pyautogui, webbrowser
#from time import sleep

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "5"
os.environ["REQUESTS_READ_TIMEOUT"] = "5"

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='crear_reserva_emp.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

#servers = ["co.pool.ntp.org","south-america.pool.ntp.org"] 
#c= ntplib.NTPClient()
#response = c.request(servers[0], version = 3)
#print(f"version, {response.version}")
#print(ctime(response.tx_time))

#with open('datos.txt', 'a+') as f:
#  f.write(str(ctime(response.tx_time)) + '\n')
#f.close()

datos_book = load_workbook("archivos/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       idcalendar = _row[2]
       break
  return idcalendar

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       precio = _row[1]
       #print(f'su correo es {_row[1]}')
  return precio
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       emailenc = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return emailenc
 
#encargados = dataBookEncEmail("encargado", "Mario Vargas")
#print(encargados)

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheets']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def consultar_reserva(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-emp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Normalizar el formato de fecha y hora para la búsqueda
        #try:
        #    fecha_busqueda = datetime.strptime(fecha, '%d-%m-%Y').strftime#('%d-%m-%Y')
        #    hora_busqueda = datetime.strptime(hora, '%H:%M').strftime('%H:%M')
        #except ValueError:
        #    st.error("Formato de fecha u hora inválido")
        #    return False, None
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False, None
      
class CrearReservaEmp:
 try:
  class Model:
    pageTitle = "***Generar Reserva***"
  
  def view(self,model):
    st.title(model.pageTitle)

    with st.form(key='myform1',clear_on_submit=True):
  
      horas = dataBook("horario")
      
      servicio = dataBook("servicio")
      result_serv = np.setdiff1d(servicio,'')
      #print(f'servicio y result_serv {servicio}, {result_serv}')
     
      servicioprecio = dataBookServicio("servicio")
      muestra = (f'servicio precio; {servicioprecio}')
      #print(f'(muestra= {muestra})')
      #result_servpre = np.setdiff1d(servicioprecio,'')
    
      encargado = dataBook("encargado")
      result_estil = np.setdiff1d(encargado,'X') 
                            
      document='gestion-reservas-emp'
      sheet = 'reservas'
      credentials = st.secrets['sheetsemp']['credentials_sheet']
      time_zone = 'GMT-05:00' # 'South America'
            
      c1, c2 = st.columns(2)
      nombre = c1.text_input('Nombre o Empresa*: ', placeholder='Nombre') # label_visibility='hidden')
      email  = c2.text_input('Email:', placeholder='Email')
      fecha  = c1.date_input('Fecha*: ')
      servicios = c1.selectbox('Servicios*: ', result_serv) 
      st.text(muestra)
      notas = c1.text_area('Nota o Mensaje(Opcional)')
      precio = dataBookPrecio("servicio", servicios)
      #result_precio = np.setdiff1d(precio,'')
      #print(f'Precio = {precio}')
      
      idcalendarserv = dataBookServicioId("servicio", servicios)
      #print(f'idcalendarserv = {idcalendarserv}')
      result_id = np.setdiff1d(idcalendarserv,'')
  
      #idcalendar = "josegarjagt@gmail.com"
                 
      if fecha:
        id = ""
        if servicios == servicio:
          id = result_id
        
      #calendar = GoogleCalendar(id) #credentials, idcalendar
        
      encargado = c2.selectbox('Encargado:',result_estil)
      #hora = c2.selectbox('Hora: ',horas)
            
      emailencargado = dataBookEncEmail("encargado",encargado)
      result_email = np.setdiff1d(emailencargado,'X') 
      #print(f'Emailencargado y result_email {emailencargado}, {result_email}')
              
      #hours_blocked = calendar.list_upcoming_events()
      #result_hours = np.setdiff1d(horas, hours_blocked) 
      hora = c2.selectbox('Hora: ', horas)
    
      whatsapp = c2.checkbox('Envio a WhatsApp Si/No (Opcional)')
      telefono = c2.text_input('Nro. Telefono')

      enviar = st.form_submit_button(" Reservar ")

      #Backend
      if enviar:
        with st.spinner('Cargando...'):
          if not nombre or not servicio or not encargado:
            st.warning('Se Require completar los campos con * son obligatorios')
        
          elif not validate_email(email):
            st.warning('El email no es valido')
        
          else:
                      
            parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
            hours1 = parsed_time.hour
            minutes1 =  parsed_time.minute
                    
            end_hours = add_hour_and_half(hora)
          
            parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
            hours2 = parsed_time2.hour
            minutes2 =  parsed_time2.minute

            start_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours1,minutes1).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
                             
            end_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours2,minutes2).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
            gs = GoogleSheet(credentials, document, sheet)
            existe = False
            
            if len(gs.sheet.get_all_values()) +1 > 2:
              #last_row = len(gs.sheet.get_all_values()) +1
              #print(f'last_row {last_row}')
              #data = gs.sheet.get_values()
              #print(f'data {data}')
              #data2 = data[1:]
              #print(f'data2 {data2}')
              #range_start = f"A{last_row}"
              #print(f'range_start {range_start}')
              #range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
              #print(f'range_end {range_end}')
          
              #for row in data2:
        
              #  nom = [row[0]]
              #  serv = [row[4]]
              #  fech = str(row[2])
              #  hora2 = str(row[3])
              #  uid1 = str(row[8])

              #  if nom != ['DATA']:
              
               
               #   fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
               #   fech1 = int(fech2.strftime("%Y%m%d"))
               #   fechacalendarint = int(fecha.strftime("%Y%m%d"))
               #   hora3 = datetime.datetime.strptime(hora2,'%H:%M')
               #   fechahora_ini = int(hora3.strftime('%H%M'))
               #   horacalendar = datetime.datetime.strptime(hora,'%H:%M')
               #   horacalendarint = int(horacalendar.strftime('%H%M'))
              
                  #print(f'Fechas y horas {fech1}, {fechacalendarint}, {fechahora_ini}, { #horacalendarint}')
              
                #  if nom == [nombre] and fech1 == fechacalendarint and fechahora_ini == horacalendarint:
                #    existe = True
                #    st.warning("El cliente ya tiene agenda para esa fecha y hora")
                #    break
              
             existe_db = consultar_reserva(nombre, str(fecha), hora)

             if existe_db:
               existe = True
               st.warning("Ciente Ya tiene agenda para esa fecha y hora")
             else:
               #gs = GoogleSheet(credentials, document, sheet)
               existe = False  
            
            if existe == False: 
                       
                hora_actual = dt.datetime.utcnow()
                #hours1 = hora_actual.hour
                #horaweb = hours1 - 5
                #minutes1 =   hora_actual.minute
                #hora_actual_int = int(horaweb)
                hora_actual_int = int(hora_actual.strftime("%H%M"))
                #print(f'hora_actual = {hora_actual_int}')
                
                hora_calendar = datetime.datetime.strptime(hora,'%H:%M')
                hora_calendar_int = int(hora_calendar.strftime('%H%M'))
                #st.warning(f'hora_actual = {hora_actual_int}, hora_calendar = {hora_calendar_int}')
                
                hoy = dt.datetime.now()
                fechoy = int(hoy.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))
          
                if fechacalendarint >= fechoy:
                  
                  #if fechacalendarint == fechoy and  hora_calendar_int < hora_actual_int:
            
                  #  st.warning('La hora seleccionda es invalida para hoy')
                  #  print('La hora seleccionda es invalida para hoy')
                    #break
                   
                  #else:
                  whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Resserva se realizo con exito para el dia: {fecha} a las: {hora} con el encargado: {encargado} para el servicio de : {servicios}")
                  
                  boton = '=ArrayFormula(SI(M3=VERDADERO;HIPERVINCULO(O3;"Enviar");"No Enviar"))'
                  
                  uid = generate_uid()
                  values = [(nombre,email,str(fecha),hora, servicios, precio, encargado, notas, uid, whatsapp,str(57)+telefono, whatsappweb, boton)]
                  
                  gs = GoogleSheet(credentials, document, sheet)
          
                  range = gs.get_last_row_range()
                  gs.write_data(range,values)
                     
                  #calendar.create_event(servicios+". "+nombre, start_time, end_time, time_zone, attendees=result_email)

                  st.success('Su solicitud ha sido reservada de forrma exitosa')
                  send_email2(email, nombre, fecha, hora, servicios, precio, encargado,  notas)
                  send_email_emp(email, nombre, fecha, hora, servicios, precio, encargado, notas)                    

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): Proceso {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el abogado encargado: {encargado} para el servicio de : {servicios} para realizar {acciones}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                  #  sendMessage(contact, message)
                  #  sendMessage(str(57)+str(telefonoencargado), message)
                  
 except Exception as e:
    logging.error(f"Error crítico en la aplicación: {str(e)}")
    st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")

sys.excepthook = global_exception_handler