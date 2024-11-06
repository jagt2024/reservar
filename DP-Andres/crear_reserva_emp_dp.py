import streamlit as st
from google_sheets_emp import GoogleSheet
from google_calendar_emp import GoogleCalendar
#from validacion_reservas_dp import validar_existencia_reserva, formatear_fecha, formatear_hora
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import sqlite3
from sqlite3 import Error
import os 
import sys
import logging
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='crear_reserva_emp_dp.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Cargar el archivo Excel una sola vez
datos_book = load_workbook("archivos-dp/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       data = _row[2]
       break
  return data

def dataBookZonaEnc(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[5]
       break
  return data

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data

def get_conductores_por_zona(zona):
    
    if not zona:
        return []
    
    # Mapeo de zonas a nombres de hojas
    mapping = {
        'Norte': 'encargado_norte',
        'Sur': 'encargado_sur',
        'Oriente': 'encargado_oriente',
        'Occidente': 'encargado_occidente'
    }
    
    try:
        hoja = mapping.get(zona)
        if hoja:
            encargado = dataBook(hoja)
            return [c for c in encargado if c != 'X' and c is not None]
        return []
    except Exception as e:
        st.error(f"Error al obtener conductores: {str(e)}")
        return []

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=30)).time()
    return new_time.strftime("%H:%M")

def add_hour_and_half2(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")
  
def calcular_diferencia_tiempo_futuro(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Calcular la fecha y hora futura (1 hora y 30 minutos después)
        tiempo_futuro = fecha_hora + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos
        diferencia = tiempo_futuro - fecha_hora
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def calcular_diferencia_tiempo(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora_parametro = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Obtener fecha y hora actual
        fecha_hora_actual = datetime.now()
        
        # Calcular la fecha y hora futura sumando 1:30h a la fecha del parámetro
        tiempo_futuro = fecha_hora_parametro + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos entre el tiempo futuro y la hora actual
        diferencia = tiempo_futuro - fecha_hora_actual
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

#class CrearReservaEmp:
  
#  class Model:
#    pageTitle = "***Generar Reserva***"
  
 # def view(self,model):
 #   st.title(model.pageTitle)

def create_connection():
    """Create a database connection to the SQLite database"""
    conn = None
    try:
        conn = sqlite3.connect('reservas_dp.db')
        return conn
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def create_tables(conn):
    """Create the necessary tables in the database"""
    try:
        cursor = conn.cursor()
        
        # Create reservas table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS reservas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                email TEXT,
                fecha DATE NOT NULL,
                hora TEXT NOT NULL,
                servicio TEXT NOT NULL,
                precio TEXT,
                encargado TEXT NOT NULL,
                email_encargado TEXT,
                zona TEXT,
                direccion TEXT,
                notas TEXT,
                uid TEXT UNIQUE,
                whatsapp BOOLEAN,
                telefono TEXT,
                whatsapp_web TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
    except Error as e:
        print(f"Error creating tables: {e}")

def insert_reserva(conn, reserva_data):
    
    sql = '''INSERT INTO reservas(
                nombre, email, fecha, hora, servicio, precio, 
                encargado, email_encargado, zona, direccion, 
                notas, uid, whatsapp, telefono, whatsapp_web)
             VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, reserva_data)
        conn.commit()
        return cursor.lastrowid
    except Error as e:
        print(f"Error inserting reserva: {e}")
        return None

def check_existing_reserva(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing reserva: {e}")
        return False 

def check_existing_encargado(conn, encargado, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE encargado = ? AND fecha = ? AND hora = ?
            '''
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (encargado, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing encargado: {e}")
        return False 

def limpiar_campos_formulario():
    
    try:
        # Lista de campos a limpiar
         valores_default = {
            'nombre': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
            #'fecha',
            #'hora',
            #'servicio_selector'
         }
        
         # Actualizar el session state con los valores por defecto
         for campo, valor in valores_default.items():
            if campo in st.session_state:
                # Eliminar la entrada actual del session state
                del st.session_state[campo]
        
         # Forzar la recarga de la página para reiniciar los widgets
         st.rerun()
        
         return True
        
    except Exception as e:
        st.error(f"Error al limpiar los campos del formulario: {str(e)}")
        logging.error(f"Error en limpiar_campos_formulario: {str(e)}")
        return False

def inicializar_valores_default():
    
    valores_default = {
            'nombre': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
    }
    
    for campo, valor in valores_default.items():
        if campo not in st.session_state:
            st.session_state[campo] = valor

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheetsemp']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def consultar_reserva(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-dp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Normalizar el formato de fecha y hora para la búsqueda
        #try:
        #    fecha_busqueda = datetime.strptime(fecha, '%d-%m-%Y').strftime#('%d-%m-%Y')
        #    hora_busqueda = datetime.strptime(hora, '%H:%M').strftime('%H:%M')
        #except ValueError:
        #    st.error("Formato de fecha u hora inválido")
        #    return False, None
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False, None

def consultar_encargado(encargado, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-dp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Normalizar el formato de fecha y hora para la búsqueda
        #try:
        #    fecha_busqueda = datetime.strptime(fecha, '%d/%m/%Y').strftime('%d/%#m/%Y')
        #    hora_busqueda = datetime.strptime(hora, '%H:%M').strftime('%H:%M')
        #except ValueError:
        #    st.error("Formato de fecha u hora inválido")
        #    return False, None
        
        # Realizar la búsqueda
        encargado_registro = df[
            (df['ENCARGADO'].str.lower() == encargado.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not encargado_registro.empty:
            # Si encuentra el encargado, devuelve True y los detalles
            #detalles_encargado = encargado_registro.iloc[0].to_dict()
            return True #, detalles_encargado
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar encargado: {str(e)}")
        return False #, None


def crea_reserva():
    
  try:
     
    st.title('Generar Reserva del Servicio')
    
    st.write("---")
       
    try:
                
        horas = dataBook("horario")
        #print(f'horas {horas}')

        #zonas = dataBook("zonas")
        #result_zonas = np.setdiff1d(zonas,'')
        #print(f'zona {zona} {result_zonas}')
      
        #servicio = dataBook("servicio")
        #result_serv = np.setdiff1d(servicio,'')
        
        servicios_precio = dataBook("servicio")
        result_serv2 = np.setdiff1d(servicios_precio,'')

        #zona_enc = dataBookZonaEnc("zonas", encargado)
        #result_zonaencc = np.setdiff1d(zona_enc,'')
        #print(f'servicio2 {servicio2}')
            
        #servicio2 = dataBookServicio2("servicio", 'Hacia el Aeropuerto')
        #result_serv2 = np.setdiff1d(servicio2,'')
        #print(f'servicio2 {servicio2}')
           
        servicioprecio = dataBookServicio("servicio")
        #print(f'servicio Precio {servicioprecio}')
        muestra = (f'servicio precio; {servicioprecio}')
        #print(f'(muestra= {muestra})')
        #result_servpre = np.setdiff1d(servicioprecio,'')
    
        #print(f'encargado {encargado}, {result_estil}') 
                            
        document='gestion-reservas-dp'
        sheet = 'reservas'
        credentials = st.secrets['sheetsemp']['credentials_sheet']
        time_zone = 'America/Bogota' #'GMT-05:00' # 'South America'
        
        #calendar = GoogleCalendar() #credentials, idcalendar
                       
        # Inicializar valores por defecto
        inicializar_valores_default()
        
        # Crear columnas para organizar la interfaz
        col1, col2 = st.columns([1, 1])
        
        with col1:
            
            nombre = st.text_input('Nombre Solicitante*: ', placeholder='Nombre', key='nombre',value=st.session_state.nombre
                    ) # label_visibility='hidden')
                        
            # Lista de servicios disponibles
            servicios = ['Hacia el Aeropuerto', 'Desde el Aeropuerto ']
            precios = ['35.000' , '30.000']
            # Selector de servicio
            servicio_seleccionado = st.selectbox(
                'Seleccione el servicio:',
                servicios,
                key='servicio_selector'
            )
            
            # Si es hacia el aeropuerto, mostrar selector de zona
            if servicio_seleccionado == 'Hacia el Aeropuerto':
                precio_serv ='35.000'
                zonas = ['Norte', 'Sur', 'Oriente', 'Occidente']
                zona_seleccionada = st.selectbox(
                    'Seleccione la zona:',
                    zonas,
                    key='zona_selector'
                )
                
                # Obtener conductores según la zona
                encargado = get_conductores_por_zona(zona_seleccionada)
                            
            else:
                precio_serv ='30.000'
                              
                # Para otros servicios, mostrar lista general de conductores
                encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
                
            fecha  = st.date_input('Fecha Servicio*: ')
            notas = st.text_area('Nota o Mensaje(Opcional)', key='notas'
                                 ,value=st.session_state.notas
                                 )
        with col2:

            email  = st.text_input('Email Solicitante:', placeholder='Email', key='email',value=st.session_state.email
            )
            direccion = st.text_input('Direccion Ubicacion solicitante :', placeholder='Direccion', key='direccion',value=st.session_state.direccion)  
                        
            # Mostrar selector de conductor si hay conductores disponibles
            if encargado:
                conductor_seleccionado = st.selectbox(
                    'Conductor Encargado:',
                    encargado,
                    key='conductor_selector'
                )
                
            #hours_blocked = calendar.list_upcoming_events()
            #result_hours = np.setdiff1d(horas, '05:00')
            hora = st.selectbox('Hora Servicio: ', horas)
            #print(f'fecha: {fecha} hora : {hora}')

            resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')
            
            # Check if reservation already exists in database
            #existe_db2 = check_existing_encargado(conn, conductor_seleccionado, #str(fecha), hora)

            #existe_db2 = consultar_encargado(conductor_seleccionado, str(fecha), hora)

            existe_db2 = consultar_encargado(conductor_seleccionado, str(fecha), hora)

            if existe_db2:
               
               #print(f'resultado {resultado}')
               if resultado > 0 and resultado <= 90:
                  st.warning("Conductor se encuetra atendiedo un servicio")
               elif resultado >= 720:
                  st.warning("Conductor ya tiene agenda para esa fecha y hora")
               elif resultado < 0:
                  st.warning("No pude agendarse con una fecha y/o  hora vencida")
            elif resultado < 0:
                 st.warning("No sepuede agendar con una fecha y/o  hora vencida")
            else:
               st.success("La reserva está disponible")

            whatsapp = st.checkbox('Envio a WhatsApp Si/No (Opcional)')
            telefono = st.text_input('Nro. Telefono', key='telefono',value=st.session_state.telefono)

            # Mostrar resumen de la selección
            st.write("---")
            st.write("### Resumen de Solicitud:")
            
            info = {
                    "🚗 Conductor Encargado": conductor_seleccionado,   "🎯 Servicio": servicio_seleccionado, "Fecha": fecha, "Hora":  hora
                }
            
            if servicio_seleccionado == 'Hacia el Aeropuerto':
                info["📍 Zona"] = zona_seleccionada
                
                for key, value in info.items():
                    st.write(f"{key}: **{value}**")

            else:
               encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
               
               zona_enc = dataBookZonaEnc("encargado", conductor_seleccionado)
               
               info["📍 Zona"] = zona_enc
                
               for key, value in info.items():
                  st.write(f"{key}: **{value}**")

               #st.warning("No hay conductores disponibles para la selección actual.")
                 
    except Exception as e:
       st.error(f"Error en la aplicación: {str(e)}")
       st.error("Por favor, verifica que el archivo Excel y las hojas existan.")

    st.write("---")
    
    with st.form(key='myform0',clear_on_submit=True):
    
     enviar = st.form_submit_button(" Reservar ")

     #Backend
     if enviar:
        with st.spinner('Cargando...'):
         if not nombre or not servicio_seleccionado or not encargado or not email:
            st.warning('Se Require completar los campos con * son obligatorios')
        
         elif not validate_email(email):
            st.warning('El email no es valido')
        
         else:
            # Create database connection
            conn = create_connection()
            if conn is None:
                st.error("Error: No se pudo conectar a la base de datos")
                return
                
            # Create tables if they don't exist
            #create_tables(conn)
                
            parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
            hours1 = parsed_time.hour
            minutes1 = parsed_time.minute
            
            if servicios == 'Hacia el Aeropuerto':        
                end_hours = add_hour_and_half(hora)
            else:
                end_hours = add_hour_and_half2(hora)
          
            parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
            hours2 = parsed_time2.hour
            minutes2 =  parsed_time2.minute

            start_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours1,minutes1).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
                             
            end_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours2,minutes2).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
            # Check if reservation already exists in database
            existe_db = consultar_reserva(nombre, str(fecha), hora)

            if existe_db:
               st.warning("Usuario ya tiene agenda para esa fecha y hora")
            else:
               #gs = GoogleSheet(credentials, document, sheet)
               existe = False
            
            #if len(gs.sheet.get_all_values()) +1 > 2:
            #  last_row = len(gs.sheet.get_all_values()) +1
            #  #print(f'last_row {last_row}')
            #  data = gs.sheet.get_values()
            #  #print(f'data {data}')
            #  data2 = data[1:]
            #  #print(f'data2 {data2}')
            #  range_start = f"A{last_row}"
            #  #print(f'range_start {range_start}')
            #  range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
            #  #print(f'range_end {range_end}')
          
            #for row in existe_db:
        
                #nom = [row[0]]
                #serv = [row[4]]
                #fech = str(row[2])
                #hora2 = str(row[3])
                #zona_enc =str[8]
                #conductor_enc = str[6]
                #uid1 = str(row[11])

            if existe_db:
              
                  #fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
                  #fech1 = int(fech2.strftime("%Y%m%d"))
                  #fechacalendarint = int(fecha.strftime("%Y%m%d"))
                  #hora3 = datetime.datetime.strptime(hora2,'%H:%M')
                  #fechahora_ini = int(hora3.strftime('%H%M'))
                  #horacalendar = datetime.datetime.strptime(hora,'%H:%M')
                  #horacalendarint = int(horacalendar.strftime('%H%M'))
              
                  #print(f'Fechas y horas {fech1}, {fechacalendarint}, {fechahora_ini}, { #horacalendarint}')
                existe = True
                st.warning("Ciente Ya tiene agenda para esa fecha y hora")

              
            if existe == False:
                
                precio = dataBookPrecio("precios", servicio_seleccionado)
                result_precio = np.setdiff1d(precio,'')
                #print(f'Precio = {precio} result_precio = {result_precio}')
               
                emailencargado = dataBookEncEmail("encargado",conductor_seleccionado)
                #result_email = np.setdiff1d(emailencargado,'') 
                #print(f'Emailencargado = {emailencargado}, result_email ={result_email}')
      
                #st.text(muestra)
                       
                hora_actual = dt.datetime.utcnow()
                #hours1 = hora_actual.hour
                #horaweb = hours1 - 5
                #minutes1 =   hora_actual.minute
                #hora_actual_int = int(horaweb)
                hora_actual_int = int(hora_actual.strftime("%H%M"))
                #print(f'hora_actual = {hora_actual_int}')
                
                #hora_calendar = dt.datetime.strptime(hora,'%H:%M')
                ##hora_calendar_int = int(hora_calendar.strftime('%H%M'))
                #st.warning(f'hora_actual = {hora_actual_int}, hora_calendar = {hora_calendar_int}')
                
                hoy = dt.datetime.now()
                fechoy = int(hoy.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))
          
                # if fechacalendarint >= fechoy:
                  
                  #if fechacalendarint == fechoy and  hora_calendar_int < hora_actual_int:
            
                  #  st.warning('La hora seleccionda es invalida para hoy')
                  #  print('La hora seleccionda es invalida para hoy')
                    #break
                   
                  #else:
                whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Resserva se realizo con exito para el dia: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para el servicio de : {servicio_seleccionado}")
                
                if servicio_seleccionado == 'Hacia el Aeropuerto':
                    
                    uid = generate_uid()
                    values = [(nombre,email,str(fecha),hora, servicio_seleccionado, precio_serv, conductor_seleccionado, str(emailencargado), zona_seleccionada, direccion, notas, uid, whatsapp,str(57)+telefono, whatsappweb)]
                  
                    try:
                        reserva_data = (
                        nombre, email, fecha, hora, servicio_seleccionado,precio_serv, conductor_seleccionado, str(emailencargado), zona_seleccionada, direccion, notas, uid, whatsapp, str(57)+telefono, whatsappweb
                        )
                     
                        insert_reserva(conn, reserva_data)
                        
                        gs = GoogleSheet(credentials, document, sheet)
          
                        range = gs.get_last_row_range()
                        gs.write_data(range,values)

                        send_email2(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado,  notas)
                
                        send_email_emp(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado, notas, str(emailencargado))
                        
                        if limpiar_campos_formulario():
                            st.success('Campos limpiaddos exitosamente')
                     
                        st.success('Su solicitud ha sido reservada de forrma exitosa')
                        
                    except Exception as e:
                        st.error(f"Error al guardar en la base de datos: {str(e)}")
                    finally:
                        conn.close()
                
                else:
                    
                    uid = generate_uid()
                    
                    values = [(nombre,email,str(fecha),hora, servicio_seleccionado, precio_serv, conductor_seleccionado, str(emailencargado), str(zona_enc), direccion, notas, uid, whatsapp,str(57)+telefono, whatsappweb)]
                  
                    try:
                        reserva_data = (
                        nombre, email, fecha, hora, servicio_seleccionado,precio_serv,
                        conductor_seleccionado, str(emailencargado), str(zona_enc), direccion, notas, uid, whatsapp, str(57)+telefono, whatsappweb
                        )
                     
                        insert_reserva(conn, reserva_data)
                    
                        gs = GoogleSheet(credentials, document, sheet)
          
                        range = gs.get_last_row_range()
                        gs.write_data(range,values)
                
                        send_email2(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado,  notas)
                
                        send_email_emp(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado, notas, str(emailencargado)) 

                        if limpiar_campos_formulario():
                            st.success('Campos limpiaddos exitosamente')
                     
                        st.success('Su solicitud ha sido reservada de forrma exitosa')                     
                        
                    except Exception as e:
                        st.error(f"Error al guardar en la base de datos: {str(e)}")
                    finally:
                        conn.close()
                 
                  #calendar.create_event(servicios+". "+nombre, 
                  #start_time, end_time, time_zone, attendees=result_email)   

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): Proceso {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el abogado encargado: {encargado} para el servicio de : {servicios} para realizar {acciones}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                  #  sendMessage(contact, message)
                  #  sendMessage(str(57)+str(telefonoencargado), message)

  except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")

sys.excepthook = global_exception_handler

#if __name__ == "__main__":
#   crea_reserva()