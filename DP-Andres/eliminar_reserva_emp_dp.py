import streamlit as st
from google_sheets_emp import GoogleSheet
from google_calendar_emp import GoogleCalendar
#from validacion_reservas_dp import validar_existencia_reserva, formatear_fecha, formatear_hora
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import sqlite3
from sqlite3 import Error
import os 
import sys
import logging
from typing import List, Optional
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='eliminar_reserva_emp_dp.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Cargar el archivo Excel una sola vez
datos_book = load_workbook("archivos-dp/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       data = _row[2]
       break
  return data

def dataBookZonaEnc(hoja, zona):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[5] == zona:
       data = _row[0]
       break
  return data

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data

def get_conductores_por_zona(zona):
    
    if not zona:
        return []
    
    # Mapeo de zonas a nombres de hojas
    mapping = {
        'Norte': 'encargado_norte',
        'Sur': 'encargado_sur',
        'Oriente': 'encargado_oriente',
        'Occidente': 'encargado_occidente'
    }
    
    try:
        hoja = mapping.get(zona)
        if hoja:
            encargado = dataBook(hoja)
            return [c for c in encargado if c != 'X' and c is not None]
        return []
    except Exception as e:
        st.error(f"Error al obtener conductores: {str(e)}")
        return []

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=30)).time()
    return new_time.strftime("%H:%M")

def add_hour_and_half2(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")
  
def calcular_diferencia_tiempo(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora_parametro = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Obtener fecha y hora actual
        fecha_hora_actual = datetime.now()
        
        # Calcular la fecha y hora futura sumando 1:30h a la fecha del parámetro
        tiempo_futuro = fecha_hora_parametro + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos entre el tiempo futuro y la hora actual
        diferencia = tiempo_futuro - fecha_hora_actual
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

def generate_whatsapp_link(phone_number, message):
    encoded_message = message.replace(' ', '%20')
    return f"https://wa.me/{phone_number}?text={encoded_message}"

#class CrearReservaEmp:
  
#  class Model:
#    pageTitle = "***Generar Reserva***"
  
 # def view(self,model):
 #   st.title(model.pageTitle)

def create_connection():
    """Create a database connection to the SQLite database"""
    conn = None
    try:
        conn = sqlite3.connect('reservas_dp.db')
        return conn
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def actualizar_reserva(conn, nombre, fecha, hora, servicio, nuevos_datos):
    from sqlite3 import Error
    
    try:
        # Construir la consulta SQL dinámicamente basada en los campos a actualizar
        set_clauses = []
        params = []
        
        # Mapeo de campos permitidos para actualizar
        campos_permitidos = {
            'nombre': 'nombre',
            'fecha': 'fecha',
            'hora': 'hora',
            'servicio': 'servicio',
            'precio': 'precio',
            'notas': 'notas',
            'uid': 'uid',
            'whatsapp': 'whatsapp',
            'telefono': 'telefono',
            'whatsapp_web': 'whatsapp_web'
        }
        
        # Construir las cláusulas SET y parámetros
        for key, value in nuevos_datos.items():
            if key in campos_permitidos:
                set_clauses.append(f"{campos_permitidos[key]}=?")
                params.append(value)
        
        # Si no hay campos para actualizar, retornar
        if not set_clauses:
            print("No se proporcionaron campos válidos para actualizar")
            return 0
        
        # Construir la consulta SQL completa
        sql = f'''UPDATE reservas 
                SET {', '.join(set_clauses)}
                WHERE notas <> 'Reserva Cancelada'
                AND nombre=? 
                AND fecha=? 
                AND hora=? 
                AND servicio=?'''
        
        # Agregar los parámetros de búsqueda
        params.extend([nombre, fecha, hora, servicio])
        
        # Ejecutar la consulta
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()
        
        # Retornar el número de filas afectadas
        rows_affected = cursor.rowcount
        if rows_affected == 0:
            print("No se encontró ninguna reserva que coincida con los criterios de búsqueda")
        else:
            print(f"Se actualizó exitosamente la reserva en la BD.")
        return rows_affected
    
    except Error as e:
        print(f"Error actualizando reserva: {e}")
        return None

def check_existing_uuid(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT uid FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        uuid = cursor.fetchone()[0]
        return uuid
    except Error as e:
        print(f"Error checking existing uuid: {e}")
        return False 

def check_existing_otros(conn, nombre, fecha, hora):
   
    sql = '''SELECT encargado, zona, telefono, direccion, whatsapp FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        result = cursor.fetchone()
        
        if result is not None:
            return result  # Retorna la tupla completa con todos los valores
        
        # Si no se encuentra ningún resultado, retornar error específico
        return {
            'error': True,
            'code': 'NOT_FOUND',
            'message': f"No se encontró ninguna reserva para {nombre} en la fecha {fecha} a las {hora}",
            'details': {
                'nombre': nombre,
                'fecha': fecha,
                'hora': hora
            }
        }
        
    except Error as e:
        # En caso de error en la base de datos
        return {
            'error': True,
            'code': 'DATABASE_ERROR',
            'message': f"Error al consultar la base de datos: {str(e)}",
            'details': {
                'nombre': nombre,
                'fecha': fecha,
                'hora': hora,
                'error_details': str(e)
            }
        }

def check_existing_reserva(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing reserva: {e}")
        return False 

def check_existing_encargado(conn, encargado, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE encargado = ? AND fecha = ? AND hora = ?
            '''
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (encargado, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing encargado: {e}")
        return False 

def limpiar_campos_formulario():
    
    try:
        # Definir los valores por defecto para cada campo
        valores_default = {
            'nombre_ant': '',
            'email': ''
        }
        
        # Actualizar el session state con los valores por defecto
        for campo, valor in valores_default.items():
            if campo in st.session_state:
                # Eliminar la entrada actual del session state
                del st.session_state[campo]
        
        # Forzar la recarga de la página para reiniciar los widgets
        st.rerun()
        
        return True
        
    except Exception as e:
        st.error(f"Error al limpiar los campos del formulario: {str(e)}")
        logging.error(f"Error en limpiar_campos_formulario: {str(e)}")
        return False

def inicializar_valores_default():
    
    valores_default = {
        'nombre_ant': '',
        'email': ''
    }
    
    for campo, valor in valores_default.items():
        if campo not in st.session_state:
            st.session_state[campo] = valor

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheetsemp']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def consultar_otros(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-dp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            return False, None
            
        # Extraer los campos solicitados
        datos_reserva = {
            'ENCARGADO': reserva['ENCARGADO'].iloc[0],
            'ZONA': reserva['ZONA'].iloc[0],
            'TELEFONO': reserva['TELEFONO'].iloc[0],
            'DIRECCION': reserva['DIRECCION'].iloc[0],
            'WHATSAPP': reserva['WHATSAPP'].iloc[0]
        }
        
        return True, datos_reserva
        
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False,(f"Error al consultar la reserva: {str(e)}")

def consultar_reserva(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-dp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False #, None

def eliminar_reserva_sheet(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-dp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            st.warning("No se encontró la reserva a eliminar.")
            return False
        
        # Obtener el índice de la fila a eliminar
        row_index = int(reserva.index[0]) + 2 
                
        # Eliminar la fila usando delete_rows (inicio, fin)
        worksheet.delete_rows(row_index, row_index)
        
        st.success("Reserva eliminada exitosamente.")
        return True
        
    except Exception as e:
        st.error(f"Error al eliminar la reserva: {str(e)}")
        return False

def eliminar_reserva():
    
  try:
     
    st.title('Eliminar Reserva de Servicio')
       
    try:
                
        horas = dataBook("horario")
        #print(f'horas {horas}')

        zonas = dataBook("zonas")
        result_zonas = np.setdiff1d(zonas,'')
        #print(f'zona {zona} {result_zonas}')
      
        servicio = dataBook("servicio")
        result_serv = np.setdiff1d(servicio,'')
        
        servicios_precio = dataBook("servicio")
        result_serv2 = np.setdiff1d(servicios_precio,'')        
            
        #servicio2 = dataBookServicio2("servicio", 'Hacia el Aeropuerto')
        #result_serv2 = np.setdiff1d(servicio2,'')
        #print(f'servicio2 {servicio2}')
           
        servicioprecio = dataBookServicio("servicio")
        #print(f'servicio Precio {servicioprecio}')
        muestra = (f'servicio precio; {servicioprecio}')
        #print(f'(muestra= {muestra})')
        #result_servpre = np.setdiff1d(servicioprecio,'')
    
        #print(f'encargado {encargado}, {result_estil}') 
                            
        document='gestion-reservas-dp'
        sheet = 'reservas'
        credentials = st.secrets['sheetsemp']['credentials_sheet']
        time_zone = 'America/Bogota' #'GMT-05:00' # 'South America'
        
        #calendar = GoogleCalendar() #credentials, idcalendar
        
        # Inicializar valores por defecto
        inicializar_valores_default()

        st.write("---")
        st.subheader('Ingrese los datos de la Reserva Agendada')
        colum1, colum2 = st.columns([1, 1])
        
        with colum1:
        
            nombre_c = st.text_input('Nombre Solicitante*: ', placeholder='Nombre', key='nombre_nom', value=st.session_state.nombre_ant) 
                       
            # Lista de servicios disponibles
            servicios_c = ['Hacia el Aeropuerto', 'Desde el Aeropuerto ']
            servicio_seleccionado_c = st.selectbox(
                'Seleccione el servicio:',
                servicios_c,
                key='servicio_selector_ant'
            )

        with colum2:
            
            fecha_c  = st.date_input('Fecha Servicio*: ', key='date_ant')
            hora_c = st.selectbox('Hora Servicio: ', horas,  key='hora_del')
            email  = st.text_input('Email Solicitante:', placeholder='Email', key='email_del', value=st.session_state.email)
        
        if nombre_c and hora_c !=  dt.datetime.utcnow().strftime("%H%M"):
         
            #conn = create_connection()
               
            # Check if reservation already exists in database

            #existe_db2 = check_existing_reserva(conn, nombre_c, str(fecha_c), hora_c)
            
            valida, result = consultar_otros(nombre_c, str(fecha_c), hora_c)

            if valida:
                
               encargado = result['ENCARGADO']
               zona = result['ZONA']
               telefono = result['TELEFONO']
               direccion = result['DIRECCION']
               whatsapp = result['WHATSAPP']      

            else:
               # Si hay error, result será un diccionario
               print(f"Error: {result['message']}")

            existe_db2 = consultar_reserva(nombre_c, str(fecha_c), hora_c)

            if existe_db2:
                resultado = calcular_diferencia_tiempo(f'{fecha_c} {hora_c}')
                #print(f'resultado {resultado}')
                if resultado < 0:
                    st.warning("No sepuede eliminar un servicio ya vencido")
                    #conn.close()
                    #return
            
                # Mostrar resumen de la selección
                st.write("---")
                st.write("### Resumen de Solicitud a Eliminar:")
            
                info = {
                     "🚗 Conductor Encargado": encargado,"🎯 Servicio": servicio_seleccionado_c, "Fecha": fecha_c, "Hora":  hora_c
                 }
                 
                if servicio_seleccionado_c == 'Hacia el Aeropuerto':
                   info["📍 Zona"] = zona
                                   
                   for key, value in info.items():
                       st.write(f"{key}: **{value}**")
                else:
                    encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
                    #info["📍 Zona"] = zona
                
                    for key, value in info.items():
                        st.write(f"{key}: **{value}**")

                        #st.warning("No hay conductores disponibles para la selección actual.")
            else:    
              st.warning("El servicio No existe Favor verficar")

    except Exception as e:
       st.error(f"Error en la aplicación: {str(e)}")
       #st.error("Por favor, verifica que el archivo Excel y las hojas existan.")

    with st.form(key='myform6',clear_on_submit=True):
        
     eliminar = st.form_submit_button("Eliminar")
     
     #Backend
     if eliminar:
        with st.spinner('Cargando...'):
         if not nombre_c or not servicio_seleccionado_c:
            st.warning('Se Require completar los campos para cosulta y Modificcacion')

         elif not validate_email(email):
            st.warning('El email no es valido')

         else:
            # Create database connection
            conn = create_connection()
            if conn is None:
                st.error("Error: No se pudo conectar a la base de datos")
                return

            precio = dataBookPrecio("precios", servicio_seleccionado_c)
            result_precio = np.setdiff1d(precio,'')
            #print(f'Precio = {precio} result_precio = {result_precio}')
               
            emailencargado = dataBookEncEmail("encargado",encargado)
            uid = check_existing_uuid(conn, nombre_c, fecha_c, hora_c)
                   
            whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre_c} De acuerdo co  su  solicitud,  la Resserva fue Cancelada con exito para el servicio de movilizacion: {servicio_seleccionado_c}")
            
            resultado = calcular_diferencia_tiempo(f'{fecha_c} {hora_c}')
            #print(f'resultado {resultado}')
            if resultado < 0:
               st.warning("No sepuede eliminar un servicio ya vencido")
            else:
                      
             try:
                                          
                nuevos_datos = {
                             
                   'nombre': nombre_c,
                   'fecha': fecha_c,
                   'hora': hora_c,
                   'servicio': servicio_seleccionado_c,
                   'preecio': '0',
                   'notas': 'Reserva Cancelada'
                }
  
                #values = [(nombre_c,email,str(fecha_c),hora_c, #servicio_seleccionado_c, '0', encargado, str(emailencargado), #zona, direccion, 'Reserva Cancelada', uid, whatsapp,telefono, #whatsappweb)]
  
                #actualizar_reserva(conn, nombre_c, fecha_c, hora_c,#servicio_seleccionado_c, nuevos_datos)

                        
                eliminar_reserva_sheet(nombre_c, str(fecha_c), hora_c)
                    
                #gs = GoogleSheet(credentials, document, sheet)
                #range = gs.write_data_by_uid(uid, values)
                                                             
                send_email2(email, nombre_c, fecha_c, hora_c, servicio_seleccionado_c, "0", encargado,  notas='De acuerdo con su solicitud su reserva de movilizacion se cancelo. Gracias por su atencion.')
                     
                send_email_emp(email, nombre_c, fecha_c, hora_c, servicio_seleccionado_c, '0', encargado, 'Reserva Cancelada', str(emailencargado)) 

                st.success('Su solicitud ha sido cacelada de forrma exitosa, la confirmacion fue enviada al correo')

                if whatsapp == True:
                   contact = str(57)+telefono
                   message = f'Cordial saludo: Sr(a): {nombre_c} La Reserva se modifico con exito para el dia: {fecha_c} a las: {hora_c} con el encargado: {encargado} para realizar el servcio: {servicio_seleccionado_c}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                   phone_number = contact
                   mensaje = message 
                   whatsapp_link = generate_whatsapp_link(phone_number, mensaje)
                   st.markdown(f"Click si desea Enviar a su Whatsapp {whatsapp_link}")
                   time.sleep(5)
                        
             except Exception as e:
                st.error(f"Error al guardar en la base de datos: {str(e)}")
             #finally:
             #   conn.close()
             
             if limpiar_campos_formulario():               
                st.success('Los ccaampos fueron limpiados exitosamente')                                   
                  #calendar.create_event(servicios+". "+nombre, 
                  #start_time, end_time, time_zone, attendees=result_email)   

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): Proceso {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el abogado encargado: {encargado} para el servicio de : {servicios} para realizar {acciones}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                  #  sendMessage(contact, message)
                  #  sendMessage(str(57)+str(telefonoencargado), message)

  except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error(f"Error crítico en la aplicación. Por favor, contacte al administrador.{str(e)}")
        #print(f'Error crítico en la aplicación: {str(e)}')

sys.excepthook = global_exception_handler

#if __name__ == "__main__":
#   eliminar_reserva()