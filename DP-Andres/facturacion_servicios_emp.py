import streamlit as st
import pandas as pd
from datetime import datetime
import random
import qrcode
import numpy as np
from PIL import Image
import io
import base64
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
#from user_management import logout 
import os
import sqlite3
import json
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import logging

# Configurar logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

datos_book = load_workbook("archivos/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

# Función para cargar datos del emisor desde Excel
def cargar_datos_emisor():
    try:
        df = pd.read_excel('./archivos/parametros_empresa.xlsx', sheet_name='emisor')
        emisor = df.iloc[0]
        return {
            'nombre': emisor['NOMBRE'],
            'nit': emisor['NIT'],
            'direccion': emisor['DIRECCION'],
            'ciudad': emisor['CIUDAD']
        }
    except Exception as e:
        st.error(f"Error al cargar datos del emisor: {str(e)}")
        return None

# Función para inicializar la base de datos
def init_db():
    conn = sqlite3.connect('facturas.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS facturas
                 (numero_factura TEXT PRIMARY KEY, 
                  fecha_factura TEXT,
                  emisor_nombre TEXT,
                  emisor_nit TEXT,
                  emisor_ciudad TEXT,
                  cliente_nombre TEXT,
                  cliente_nit TEXT,
                  cliente_direccion TEXT,
                  cliente_email TEXT,
                  servicios TEXT,
                  subtotal REAL,
                  iva_total REAL,
                  total REAL)''')
    conn.commit()
    conn.close()

# Función para guardar la factura en la base de datos
def guardar_factura_en_db(numero_factura, fecha_factura, emisor_nombre, emisor_nit, emisor_ciudad, cliente_nombre, cliente_nit, cliente_direccion, cliente_email, servicios, subtotal, iva_total, total):
    conn = sqlite3.connect('facturas.db')
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO facturas VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (numero_factura, fecha_factura, emisor_nombre, emisor_nit, emisor_ciudad,
         cliente_nombre, cliente_nit, cliente_direccion, cliente_email, json.dumps(servicios), subtotal, iva_total, total))
    conn.commit()
    conn.close()
    
def enviar_factura_por_email(email_cliente, pdf_buffer, numero_factura):
    smtp_server = "smtp.gmail.com"
    port = 465
    sender_email = st.secrets['emails']['smtp_user']
    password = st.secrets['emails']['smtp_password']

    logger.info(f"Iniciando envío de correo a {email_cliente}")
    
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = email_cliente
    message["Subject"] = f"Factura {numero_factura}"

    body = f"Adjunto encontrará la factura {numero_factura}. Gracias por su preferencia."
    message.attach(MIMEText(body, "plain"))

    pdf_attachment = MIMEApplication(pdf_buffer.getvalue(), _subtype="pdf")
    pdf_attachment.add_header('Content-Disposition', f'attachment; filename=factura_{numero_factura}.pdf')
    message.attach(pdf_attachment)

    try:
        logger.debug("Intentando conexión SMTP_SSL")
        with smtplib.SMTP_SSL(smtp_server, port) as server:
            logger.debug("Conexión SMTP_SSL exitosa")
            logger.debug("Intentando login")
            server.login(sender_email, password)
            logger.debug("Login exitoso")
            logger.debug("Enviando mensaje")
            server.send_message(message)
            logger.info("Mensaje enviado exitosamente")
        return True
    except Exception as e:
        logger.error(f"Error al enviar el correo: {str(e)}", exc_info=True)
        return False

def generar_numero_factura():
    if 'numero_factura' not in st.session_state:
        st.session_state.numero_factura = f"FACT-{random.randint(1000, 9999)}"
    return st.session_state.numero_factura

def limpiar_campos():
    st.session_state.numero_factura = generar_numero_factura()
    if 'nombre_cliente' in st.session_state:
        del st.session_state.nombre_cliente
        st.session_state.nombre_cliente = ''
    if 'direccion_cliente' in st.session_state:
        del st.session_state.direccion_cliente
        st.session_state.direccion_cliente=''
    if 'email_cliente' in st.session_state:
        del st.session_state.email_cliente
        st.session_state.email_cliente=''
    if 'nit_cliente' in st.session_state:
        del st.session_state.nit_cliente
        st.session_state.nit_cliente = ''
    for i in range(10):
        if f"desc_{i}" in st.session_state:
            del st.session_state[f"desc_{i}"]
            st.session_state[f"desc_{i}"] = ''
        if f"cant_{i}" in st.session_state:
            del st.session_state[f"cant_{i}"]
            st.session_state[f"cant_{i}"] = 0
        if f"precio_{i}" in st.session_state:
            del st.session_state[f"precio_{i}"]
            st.session_state[f"precio_{i}"]= 0

def calcular_iva(precio):
    return precio * 0.19  # 19% IVA en Colombia

def calcular_precio_sin_iva(precio_con_iva):
    return precio_con_iva / 1.19

def generar_qr(datos):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(datos)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    return img_to_bytes(img)

def img_to_bytes(img):
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()

def generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, servicios, subtotal, iva_total, total, logo_bytes, qr_bytes, emisor_nombre, emisor_nit, emisor_ciudad):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    
    if logo_bytes:
        logo_image = ReportLabImage(io.BytesIO(logo_bytes), width=2*inch, height=1*inch)
        elements.append(logo_image)

    elements.append(Paragraph(f"Emisor: {emisor_nombre}", styles['Normal']))
    elements.append(Paragraph(f"NIT Emisor: {emisor_nit}", styles['Normal']))
    elements.append(Paragraph(f"Ciudad: {emisor_ciudad}", styles['Normal']))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Factura Nº: {numero_factura}", styles['Heading1']))
    elements.append(Paragraph(f"Fecha: {fecha_factura}", styles['Normal']))
    elements.append(Paragraph(f"Cliente: {nombre_cliente}", styles['Normal']))
    elements.append(Paragraph(f"NIT/Cédula: {nit_cliente}", styles['Normal']))
    elements.append(Paragraph(f"Dirección: {direccion_cliente}", styles['Normal']))
    elements.append(Paragraph(f"Email: {email_cliente}", styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Descripción', 'Cantidad', 'Precio Unitario (sin IVA)', 'IVA', 'Subtotal (con IVA)']]
    for servicio in servicios:
        data.append([
            servicio['descripcion'],
            str(servicio['cantidad']),
            f"${servicio['precio_unitario_sin_iva']:,.2f}",
            f"${servicio['iva']:,.2f}",
            f"${servicio['subtotal']:,.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Subtotal (sin IVA): ${subtotal:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"IVA Total: ${iva_total:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"Total: ${total:,.2f}", styles['Normal']))

    qr_image = ReportLabImage(io.BytesIO(qr_bytes), width=2*inch, height=2*inch)
    elements.append(qr_image)

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generar_factura():
    st.title("Generación de Factura")

    init_db()

    # Cargar datos del emisor
    emisor_data = cargar_datos_emisor()
    if not emisor_data:
        st.error("No se pudieron cargar los datos del emisor. Por favor, verifique el archivo Excel.")
        return

    logo_path = "./assets/dp_andres.png"
    logo_bytes = None
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_bytes = f.read()
        st.image(logo_bytes, width=200, caption="Logo de la Empresa")
    else:
        st.warning("No se encontró el logo en ./assets/dp_andres.png")

    st.subheader("Información del Emisor")
    st.write(f"**Nombre del Emisor:** {emisor_data['nombre']}")
    st.write(f"**NIT del Emisor:** {emisor_data['nit']}")
    st.write(f"**Dirección del Emisor:** {emisor_data['direccion']}")
    st.write(f"**Ciudad del Emisor:** {emisor_data['ciudad']}")

    st.subheader("Información del Cliente")
    nombre_cliente = st.text_input("Nombre del Cliente", key="nombre_cliente")
    direccion_cliente = st.text_input("Dirección del Cliente", key="direccion_cliente")
    nit_cliente = st.text_input("NIT/Cédula del Cliente", key="nit_cliente")
    email_cliente = st.text_input("Correo Electrónico del Cliente", key="email_cliente")

    st.subheader("Número de Factura")
    numero_factura = st.text_input("Número de Factura", value=generar_numero_factura(), key="numero_factura")
    
    servicio = dataBook("servicio")
    result_serv = np.setdiff1d(servicio,'')

    st.subheader("Servicios")
    servicios = []
        
    for i in range(10):  # Limit to 10 services for this example
        with st.expander(f"Servicio {i + 1}", expanded=i == 0):
            descripcion = st.selectbox('Descripcion del Servicio: ', result_serv, key=f"desc_{i}")
            #st.text_input(f"Descripción del Servicio", key=f"desc_{i}")
            cantidad = st.number_input(f"Cantidad", min_value=1, value=1, step=1, key=f"cant_{i}")
            precio_unitario_con_iva = st.number_input(f"Precio Unitario (con IVA)", min_value=0.0, value=0.0, step=1000.0, key=f"precio_{i}")
        
            if descripcion and cantidad and precio_unitario_con_iva > 0:
                precio_unitario_sin_iva = calcular_precio_sin_iva(precio_unitario_con_iva)
                iva = calcular_iva(precio_unitario_sin_iva)
                subtotal = cantidad * precio_unitario_con_iva
                servicios.append({
                    "descripcion": descripcion,
                    "cantidad": cantidad,
                    "precio_unitario_sin_iva": precio_unitario_sin_iva,
                    "iva": iva * cantidad,
                    "subtotal": subtotal
                })
    
    if servicios:
        subtotal = sum(servicio["precio_unitario_sin_iva"] * servicio["cantidad"] for servicio in servicios)
        iva_total = sum(servicio["iva"] for servicio in servicios)
        total = sum(servicio["subtotal"] for servicio in servicios)
     
        if st.button("Generar Factura", key="generar_factura"):
            fecha_factura = datetime.now().strftime('%Y-%m-%d')
            
            st.subheader("Factura Generada")
            
            if logo_bytes:
                st.image(logo_bytes, width=200)

            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**Emisor:** {emisor_data['nombre']}")
                st.write(f"**NIT Emisor:** {emisor_data['nit']}")
                st.write(f"**Dirección Emisor:** {emisor_data['direccion']}")
                st.write(f"**Ciudad Emisor:** {emisor_data['ciudad']}")
                st.write(f"**Número de Factura:** {numero_factura}")
                st.write(f"**Fecha:** {fecha_factura}")
            with col2:
                st.write(f"**Cliente:** {nombre_cliente}")
                st.write(f"**NIT/Cédula:** {nit_cliente}")
                st.write(f"**Dirección:** {direccion_cliente}")
                st.write(f"**Email:** {email_cliente}")
                
            df_servicios = pd.DataFrame(servicios)
            st.table(df_servicios)

            st.write(f"**Subtotal (sin IVA):** ${subtotal:,.2f} COP")
            st.write(f"**IVA Total:** ${iva_total:,.2f} COP")
            st.write(f"**Total:** ${total:,.2f} COP")

            datos_qr = f"Factura: {numero_factura}\nFecha: {fecha_factura}\nEmisor: {emisor_data['nombre']}\nCliente: {nombre_cliente}\nTotal: ${total:,.2f} COP"
            qr_bytes = generar_qr(datos_qr)
            
            st.image(qr_bytes, caption='Código QR de la Factura', width=300)
            
            try:
               
                pdf_buffer = generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, servicios, subtotal, iva_total, total, logo_bytes, qr_bytes, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'])

                st.download_button(
                label="Descargar Factura como PDF",
                data=pdf_buffer,
                file_name=f"factura_{numero_factura}.pdf",
                mime="application/pdf",
                )

                df_servicios = pd.DataFrame(servicios)
                csv = df_servicios.to_csv(index=False)
                st.download_button(
                label="Descargar Factura como CSV",
                data=csv,
                file_name=f"factura_{numero_factura}.csv",
                mime="text/csv",
                )

                st.download_button(
                label="Descargar Código QR",
                data=qr_bytes,
                file_name=f"codigo_qr_factura_{numero_factura}.png",
                mime="image/png",
                )

            except Exception as e:
                st.error(f"Error al generar el PDF o guardar en la base de datos: {str(e)}")
    else:
        st.warning("Por favor, agregue al menos un servicio para generar la factura.")

    #Opción para enviar la factura por correo electrónico
    if st.button("Enviar Factura por Correo Electrónico"):
       if email_cliente:
          fecha_factura = datetime.now().strftime('%Y-%m-%d')
          datos_qr = f"Factura: {numero_factura}\nFecha: {fecha_factura}\nEmisor: {emisor_data['nombre']}\nCliente: {nombre_cliente}\nTotal: ${total:,.2f} COP"
          qr_bytes = generar_qr(datos_qr)
          with st.spinner("Enviando factura por correo electrónico..."):
            logger.info(f"Generando PDF para factura {numero_factura}")
            pdf_buffer = generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, servicios, subtotal, iva_total, total, logo_bytes, qr_bytes, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'])
            logger.info("PDF generado correctamente")
                    
            logger.info(f"Intentando enviar correo a {email_cliente}")
            if enviar_factura_por_email(email_cliente, pdf_buffer, numero_factura):
               st.success(f"Factura enviada exitosamente a {email_cliente}")
               logger.info(f"Factura enviada exitosamente a {email_cliente}")
            else:
               st.error("No se pudo enviar la factura por correo electrónico. Por favor, revise los logs para más detalles.")
               logger.error("Fallo en el envío de la factura por correo electrónico")
       else:
          st.warning("Por favor, ingrese el correo electrónico del cliente para enviar la factura.")
          logger.warning("Intento de envío de factura sin dirección de correo electrónico")
    
    if st.button("Confirmar y Guardar Factura"):
       fecha_factura = datetime.now().strftime('%Y-%m-%d')
       guardar_factura_en_db(numero_factura, fecha_factura, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'], nombre_cliente, nit_cliente, direccion_cliente, email_cliente, servicios, subtotal, iva_total, total)
       st.success("Factura guardada en la base de datos exitosamente.")
    
    # Limpiar campos después de generar y descargar la factura
    st.button("Limpiar campos y generar nueva factura", on_click=limpiar_campos)

#if __name__ == "__main__":
#    generar_factura()