import streamlit as st
from google_sheets_emp import GoogleSheet
from google_calendar_emp import GoogleCalendar
#from validacion_reservas_dp import validar_existencia_reserva, formatear_fecha, formatear_hora
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import sqlite3
from sqlite3 import Error
import os 
import sys
import logging
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='modificar_reserva_emp_dp.log', filemode='w',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# Cargar el archivo Excel una sola vez
datos_book = load_workbook("archivos-dp/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       data = _row[2]
       break
  return data

def dataBookZonaEnc(hoja, zona):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[5] == zona:
       data = _row[0]
       break
  return data

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data

def get_conductores_por_zona(zona):
    
    if not zona:
        return []
    
    # Mapeo de zonas a nombres de hojas
    mapping = {
        'Norte': 'encargado_norte',
        'Sur': 'encargado_sur',
        'Oriente': 'encargado_oriente',
        'Occidente': 'encargado_occidente'
    }
    
    try:
        hoja = mapping.get(zona)
        if hoja:
            encargado = dataBook(hoja)
            return [c for c in encargado if c != 'X' and c is not None]
        return []
    except Exception as e:
        st.error(f"Error al obtener conductores: {str(e)}")
        return []

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=30)).time()
    return new_time.strftime("%H:%M")

def add_hour_and_half2(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")
  
def calcular_diferencia_tiempo(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora_parametro = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Obtener fecha y hora actual
        fecha_hora_actual = datetime.now()
        
        # Calcular la fecha y hora futura sumando 1:30h a la fecha del parámetro
        tiempo_futuro = fecha_hora_parametro + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos entre el tiempo futuro y la hora actual
        diferencia = tiempo_futuro - fecha_hora_actual
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

#class CrearReservaEmp:
  
#  class Model:
#    pageTitle = "***Generar Reserva***"
  
 # def view(self,model):
 #   st.title(model.pageTitle)

def create_connection():
    """Create a database connection to the SQLite database"""
    conn = None
    try:
        conn = sqlite3.connect('reservas_dp.db')
        return conn
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def actualizar_reserva(conn, nombre, fecha, hora, servicio, nuevos_datos):
    from sqlite3 import Error
    
    try:
        # Construir la consulta SQL dinámicamente basada en los campos a actualizar
        set_clauses = []
        params = []
        
        # Mapeo de campos permitidos para actualizar
        campos_permitidos = {
            'nombre': 'nombre',
            'email': 'email',
            'fecha': 'fecha',
            'hora': 'hora',
            'servicio': 'servicio',
            'precio': 'precio',
            'encargado': 'encargado',
            'email_encargado': 'email_encargado',
            'zona': 'zona',
            'direccion': 'direccion',
            'notas': 'notas',
            'uid': 'uid',
            'whatsapp': 'whatsapp',
            'telefono': 'telefono',
            'whatsapp_web': 'whatsapp_web'
        }
        
        # Construir las cláusulas SET y parámetros
        for key, value in nuevos_datos.items():
            if key in campos_permitidos:
                set_clauses.append(f"{campos_permitidos[key]}=?")
                params.append(value)
        
        # Si no hay campos para actualizar, retornar
        if not set_clauses:
            print("No se proporcionaron campos válidos para actualizar")
            return 0
        
        # Construir la consulta SQL completa
        sql = f'''UPDATE reservas 
                SET {', '.join(set_clauses)}
                WHERE nombre=? 
                AND fecha=? 
                AND hora=? 
                AND servicio=?'''
        
        # Agregar los parámetros de búsqueda
        params.extend([nombre, fecha, hora, servicio])
        
        # Ejecutar la consulta
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()
        
        # Retornar el número de filas afectadas
        rows_affected = cursor.rowcount
        if rows_affected == 0:
            print("No se encontró ninguna reserva que coincida con los criterios de búsqueda")
        else:
            print(f"Se actualizó exitosamente la reserva")
        return rows_affected
    
    except Error as e:
        print(f"Error actualizando reserva: {e}")
        return None

def check_existing_uuid(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT uid FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        uuid = cursor.fetchone()[0]
        return uuid
    except Error as e:
        print(f"Error checking existing uuid: {e}")
        return False 

def check_existing_reserva(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing reserva: {e}")
        return False 

def check_existing_encargado(conn, encargado, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE encargado = ? AND fecha = ? AND hora = ?
            '''
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (encargado, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing encargado: {e}")
        return False 
    
def limpiar_campos(widgets):

    for widget in widgets:
        widget.delete(0, 'end')

def modificar_reserva():
    
  try:
     
    st.title('Modificar Reserva de Servicio')
       
    try:
                
        horas = dataBook("horario")
        #print(f'horas {horas}')

        zonas = dataBook("zonas")
        result_zonas = np.setdiff1d(zonas,'')
        #print(f'zona {zona} {result_zonas}')
      
        servicio = dataBook("servicio")
        result_serv = np.setdiff1d(servicio,'')
        
        servicios_precio = dataBook("servicio")
        result_serv2 = np.setdiff1d(servicios_precio,'')        
            
        #servicio2 = dataBookServicio2("servicio", 'Hacia el Aeropuerto')
        #result_serv2 = np.setdiff1d(servicio2,'')
        #print(f'servicio2 {servicio2}')
           
        servicioprecio = dataBookServicio("servicio")
        #print(f'servicio Precio {servicioprecio}')
        muestra = (f'servicio precio; {servicioprecio}')
        #print(f'(muestra= {muestra})')
        #result_servpre = np.setdiff1d(servicioprecio,'')
    
        #print(f'encargado {encargado}, {result_estil}') 
                            
        document='gestion-reservas-dp'
        sheet = 'reservas'
        credentials = st.secrets['sheetsemp']['credentials_sheet']
        time_zone = 'America/Bogota' #'GMT-05:00' # 'South America'
        
        #calendar = GoogleCalendar() #credentials, idcalendar

        st.write("---")
        st.subheader('Ingrese los datos de la Reserva Agendada')
        colum1, colum2 = st.columns([1, 1])
        
        with colum1:
        
            nombre_c = st.text_input('Nombre Solicitante*: ', placeholder='Nombre', key='nombre_ant') 
                       
            # Lista de servicios disponibles
            servicios_c = ['Hacia el Aeropuerto', 'Desde el Aeropuerto ']
            servicio_seleccionado_c = st.selectbox(
                'Seleccione el servicio:',
                servicios_c,
                key='servicio_selector_c'
            )

        with colum2:
            
            fecha_c  = st.date_input('Fecha Servicio*: ', key='fecha_ant')
            hora_c = st.selectbox('Hora Servicio: ', horas, key='hora_ant')
        
        if nombre_c and hora_c !=  dt.datetime.utcnow().strftime("%H%M"):
         
            conn = create_connection()

            # Check if reservation already exists in database
            existe_db2 = check_existing_reserva(conn, nombre_c, str(fecha_c), hora_c)

            if existe_db2:
                resultado = calcular_diferencia_tiempo(f'{fecha_c} {hora_c}')
                #print(f'resultado {resultado}')
                if resultado < 0:
                    st.warning("No sepuede modificar un servicio ya vencido")
                    conn.close()
                    return
            
                st.write("---")
                st.subheader('Ingrese los datos de la Nueva Reserva')        
                # Crear columnas para organizar la interfaz
                col1, col2 = st.columns([1, 1])
        
                with col1:
            
                    nombre = st.text_input('Nombre Solicitante*: ', placeholder='Nombre')           
                    # Lista de servicios disponibles
                    servicios = ['Hacia el Aeropuerto', 'Desde el Aeropuerto ']
                    precios = ['35.000' , '30.000']
                    # Selector de servicio
                    servicio_seleccionado = st.selectbox(
                        'Seleccione el servicio:',
                        servicios,
                        key='servicio_selector'
                    )
            
                    # Si es hacia el aeropuerto, mostrar selector de zona
                    if servicio_seleccionado == 'Hacia el Aeropuerto':
                        precio_serv ='35.000'
                        zonas = ['Norte', 'Sur', 'Oriente', 'Occidente']
                        zona_seleccionada = st.selectbox(
                            'Seleccione la zona:',
                            zonas,
                            key='zona_selector'
                        )
                
                        # Obtener conductores según la zona
                        encargado = get_conductores_por_zona(zona_seleccionada)
                    else:
                        precio_serv ='30.000'
                        # Para otros servicios, mostrar lista general de conductores
                        encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
                
                        conn = create_connection()

                        existe_db2 = check_existing_encargado(conn, conductor_seleccionado, str(fecha), hora)

                        if existe_db2:
                            st.warning("Conductor ya tiene agenda para esa fecha y hora")
                            conn.close()
                            return
                        else:
                            st.success("La reserva está disponible")  

                    fecha  = st.date_input('Fecha Servicio*: ')
                    notas = st.text_area('Nota o Mensaje(Opcional)')
            
                with col2:

                    email  = st.text_input('Email Solicitante:', placeholder='Email')
                    direccion = st.text_input('Direccion Ubicacion solicitante :', placeholder='Direccion')  
                        
                    # Mostrar selector de conductor si hay conductores disponibles
                    if encargado:
                        conductor_seleccionado = st.selectbox(
                        'Conductor Encargado:',
                        encargado,
                        key='conductor_selector'
                        )
                
                    #hours_blocked = calendar.list_upcoming_events()
                    #result_hours = np.setdiff1d(horas, '05:00')
                    hora = st.selectbox('Hora Servicio: ', horas)
                    #print(f'fecha: {fecha} hora : {hora}')

                    conn = create_connection()
                    resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')       

                    # Check if reservation already exists in database
                    existe_db2 = check_existing_encargado(conn, conductor_seleccionado, str(fecha), hora)

                    if existe_db2:
                        resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')
                        #print(f'resultado {resultado}')
                        if resultado > 0 and resultado <= 90:
                            st.warning("Conductor se encuetra atendiedo un servicio")
                            conn.close()
                            return
                        elif resultado >= 720:
                            st.warning("Conductor ya tiene agenda para esa fecha y hora")
                        elif resultado < 0:
                            st.warning("No pude agendarse con una fecha y/o  hora vencida")
                    
                    elif resultado < 0:
                        st.warning("No sepuede modificar un servicio ya vencido")
                        conn.close()
                        return
                    else:
                        st.success("La reserva está disponible")

                    whatsapp = st.checkbox('Envio a WhatsApp Si/No (Opcional)')
                    telefono = st.text_input('Nro. Telefono')

                    # Mostrar resumen de la selección
                    st.write("---")
                    st.write("### Resumen de Solicitud a Modificar:")
            
                    info = {
                        "🚗 Conductor Encargado": conductor_seleccionado,   "🎯 Servicio": servicio_seleccionado, "Fecha": fecha, "Hora":  hora
                    }
            
                    if servicio_seleccionado == 'Hacia el Aeropuerto':
                        info["📍 Zona"] = zona_seleccionada
                
                        for key, value in info.items():
                            st.write(f"{key}: **{value}**")
                    else:
                        encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
                        #info["📍 Zona"] = zona_seleccionada
                
                        for key, value in info.items():
                            st.write(f"{key}: **{value}**")

                        #st.warning("No hay conductores disponibles para la selección actual.")
            else:    
              st.warning("El servicio No existe Favor verficar")
              conn.close()
              return

    except Exception as e:
       st.error(f"Error en la aplicación: {str(e)}")
       #st.error("Por favor, verifica que el archivo Excel y las hojas existan.")

    with st.form(key='myform1',clear_on_submit=True):
        
     actualizar = st.form_submit_button("Actualizar")
     
     #Backend
     if actualizar:
        with st.spinner('Cargando...'):
         if not nombre or not servicio or not encargado or not email:
            st.warning('Se Require completar los campos para cosulta y Modificcacion')
        
         elif not validate_email(email):
            st.warning('El email no es valido')
        
         else:
            # Create database connection
            conn = create_connection()
            if conn is None:
                st.error("Error: No se pudo conectar a la base de datos")
                return
            
            parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
            hours1 = parsed_time.hour
            minutes1 = parsed_time.minute
                        
            if servicio_seleccionado == 'Hacia el Aeropuerto':        
                end_hours = add_hour_and_half(hora)
                #print(f'end_hours1 : {end_hours}')
            else:
                end_hours = add_hour_and_half2(hora)
                #print(f'end_hours2 : {end_hours}')
          
            parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
            hours2 = parsed_time2.hour
            minutes2 =  parsed_time2.minute
          
            # Check if reservation already exists in database
            existe_db = check_existing_reserva(conn, nombre, str(fecha), hora)

            if existe_db:
               st.warning("Usuario ya tiene agenda para esa fecha y hora")
               conn.close()
               return
            else:
               #gs = GoogleSheet(credentials, document, sheet)
               existe = False
          
            if existe_db:            
                 
                existe = True
                st.warning("Ciente Ya tiene agenda para esa fecha y hora")

            if existe == False:
                
                precio = dataBookPrecio("precios", servicio_seleccionado)
                result_precio = np.setdiff1d(precio,'')
                #print(f'Precio = {precio} result_precio = {result_precio}')
               
                emailencargado = dataBookEncEmail("encargado",conductor_seleccionado)
                       
                hora_actual = dt.datetime.utcnow()
                #hours1 = hora_actual.hour
                #horaweb = hours1 - 5
                #minutes1 =   hora_actual.minute
                #hora_actual_int = int(horaweb)
                hora_actual_int = int(hora_actual.strftime("%H%M"))
                #print(f'hora_actual = {hora_actual_int}')
                
                #hora_calendar = dt.datetime.strptime(hora,'%H:%M')
                ##hora_calendar_int = int(hora_calendar.strftime('%H%M'))
                #st.warning(f'hora_actual = {hora_actual_int}, hora_calendar = {hora_calendar_int}')
                
                hoy = dt.datetime.now()
                fechoy = int(hoy.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))

                uid = check_existing_uuid(conn, nombre_c, fecha_c, hora_c)
                   
                whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Resserva se realizo con exito para el dia: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para el servicio de : {servicio_seleccionado}")
                  
                values = [(nombre,email,str(fecha),hora, servicio_seleccionado, precio_serv, conductor_seleccionado, str(emailencargado), zona_seleccionada, direccion, notas, uid, whatsapp,str(57)+telefono, whatsappweb)]

                try:
                                          
                     nuevos_datos = {
                        'nombre': nombre,
                        'email': email,
                        'fecha': fecha,
                        'hora': hora,
                        'servicio': servicio_seleccionado,
                        'precio':precio_serv,
                        'encargado': conductor_seleccionado,
                        'email_encargado': str(emailencargado),
                        'zona': zona_seleccionada,
                        'direccion': direccion,
                        'notas': notas,
                        'uid': uid,
                        'whatsapp': whatsapp,
                        'telefono': str(57)+telefono,
                        'whatsapp_web': whatsappweb
                     }
  
  
                     actualizar_reserva(conn, nombre_c, fecha_c, hora_c,servicio_seleccionado_c, nuevos_datos)
                                                             
                     send_email2(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado,  notas='De acuerdo con su solicitud su reserva se reprogramo. Gracias por su atencion.')
                     
                     send_email_emp(email, nombre, fecha, hora, servicio_seleccionado, precio_serv, conductor_seleccionado, notas, str(emailencargado)) 

                     #st.success('Su Solicitud se ha enviado a correo ingresado')
                        
                except Exception as e:
                    st.error(f"Error al guardar en la base de datos: {str(e)}")
                finally:
                    conn.close()
                
                gs = GoogleSheet(credentials, document, sheet)

                range = gs.write_data_by_uid(uid, values)
                     
                st.success('Su solicitud ha sido reservada de forrma exitosa')
                
                limpiar_campos([nombre_c, email, direccion, telefono])

                # Versión alternativa si prefieres limpiar campos individualmente
                #def limpiar_formulario(nombre, email, direccion_entry, telefono_entry):
    
                #Limpia todos los campos del formulario individualmente.
    
                nombre.delete(0, 'end')
                email.delete(0, 'end')
                direccion.delete(0, 'end')
                telefono.delete(0, 'end')

                
                  #calendar.create_event(servicios+". "+nombre, 
                  #start_time, end_time, time_zone, attendees=result_email)   

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): Proceso {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el abogado encargado: {encargado} para el servicio de : {servicios} para realizar {acciones}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                  #  sendMessage(contact, message)
                  #  sendMessage(str(57)+str(telefonoencargado), message)

  except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")

sys.excepthook = global_exception_handler

#if __name__ == "__main__":
#   modificar_reserva()