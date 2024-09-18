import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import os
import toml
import base64
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
import datetime
import time
import io
from openpyxl import load_workbook

datos_book = load_workbook("archivos/parametros.xlsx", read_only=False)

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "15"
os.environ["REQUESTS_READ_TIMEOUT"] = "15"

def dataBookSheetUrl(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row)
      url = _row[1]
    return url

def download_file(service, file_id, max_retries=5):
    for attempt in range(max_retries):
        try:
            request = service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            return fh.getvalue()
        except HttpError as error:
            if error.resp.status in [403, 500, 503]:
                time.sleep(5 * (attempt + 1))  # Espera exponencial
            else:
                raise
    raise Exception(f"Failed to download after {max_retries} attempts")

def filter_data_by_last_days(df, num_days=8):
    today = pd.Timestamp.today().date()
    start_date = today - pd.Timedelta(days=num_days-1)
    return df[df['FECHA'].dt.date >= start_date]

def load_credentials_from_toml(file_path):
    with open(file_path, 'r') as toml_file:
        config = toml.load(toml_file)
        credentials = config['sheets']['credentials_sheet']
    return credentials
    #config['credentials_sheet']

sheetUrl = dataBookSheetUrl("sw")

def get_google_sheet_data(creds):
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_info(creds, scopes=scope)
    client = gspread.authorize(credentials)

    sheet_url = str(sheetUrl)
    sheet = client.open_by_url(sheet_url)
    worksheet = sheet.worksheet('reservas')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    
    # Asegúrate de que la columna 'fecha' esté correctamente formateada
    df['FECHA'] = pd.to_datetime(df['FECHA'])
    return df

def get_binary_file_downloader_html(bin_file, file_label='File'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">Descargar {file_label}</a>'
    return href

def process_and_display_data(df):
    
    #st.write("Primeros 5 registros de la hoja:")
    #st.dataframe(df.head())

    #st.write("Últimos 5 registros de la hoja:")
    #st.dataframe(df.tail())
    df = filter_data_by_last_days(df)

    temp_file_path = "./archivos/temp_gestion-reservas.xlsx"
    df.to_excel(temp_file_path, index=False)
    
    #st.markdown(get_binary_file_downloader_html(temp_file_path, 'Excel'), #unsafe_allow_html=True)

    #st.write("Estadísticas básicas:")
    #st.write(df.describe())

    #if 'fecha' in df.columns and 'monto' in df.columns:
    #    st.write("Gráfico de Reservas por Fecha:")
    #    chart_data = df.groupby('fecha')['monto'].sum().reset_index()
    #    st.bar_chart(chart_data.set_index('fecha'))

def download_and_process_data(creds_path):
    try:
        creds = load_credentials_from_toml(creds_path)
        st.success("Credenciales cargadas correctamente")
    except Exception as e:
        st.error(f"Error al cargar las credenciales: {str(e)}")
        return

    try:
        with st.spinner('Descargando datos...'):
            df = get_google_sheet_data(creds)
        st.success('Datos descargados correctamente! en /archivos/temp_gestion-reservas.xlsx')
        process_and_display_data(df)
    except Exception as e:
        st.error(f"Error al procesar los datos: {str(e)}")

# Esta función main() puede ser eliminada si no se va a usar este script directamente
#def main():
#    st.title('Descarga de Google Sheets')
#    if st.button('Descargar Datos'):
#        download_and_process_data('path/to/your/credentials.toml')

#if __name__ == "__main__":
#    main()