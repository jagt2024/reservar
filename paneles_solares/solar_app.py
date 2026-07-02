# SolarCalc Pro v1.1 — build 202606010310
import streamlit as st
import sqlite3
import pandas as pd
import math
import webbrowser
import io
from datetime import datetime

# ── Excel / PDF exports ──────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import (Drawing, Rect, String, Line, Circle,
                                        Polygon, Group, PolyLine)
from reportlab.graphics import renderPDF
from reportlab.graphics.shapes import Drawing

# ─── PAGE CONFIG ────────────────────────────────────────────────────────────
import base64

def render_svg(svg_string: str, height: int = 700) -> None:
    """Render an SVG string reliably inside Streamlit using an HTML iframe."""
    b64 = base64.b64encode(svg_string.encode("utf-8")).decode("utf-8")
    html = (
        f'<div style="width:100%; border-radius:12px; overflow:hidden;">'
        f'<img src="data:image/svg+xml;base64,{b64}" '
        f'style="width:100%; height:{height}px; object-fit:contain; '
        f'background:#0A0E1A; border-radius:12px;" alt="Plano Solar"/>'
        f'</div>'
    )
    st.markdown(html, unsafe_allow_html=True)
st.set_page_config(
    page_title="SolarCalc Pro",
    page_icon="☀️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── CUSTOM CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@300;400;500;600;700&family=Share+Tech+Mono&family=Barlow:wght@300;400;600;700&display=swap');

:root {
    --sol: #FFB300;
    --sol-light: #FFD54F;
    --sol-dark: #E65100;
    --bg: #0A0E1A;
    --bg2: #0F1525;
    --bg3: #161D30;
    --card: #1A2235;
    --card2: #1E2A3F;
    --border: #2A3A55;
    --text: #E8EDF5;
    --text2: #8A9BBD;
    --green: #00E676;
    --cyan: #00BCD4;
    --red: #FF5252;
}

html, body, [class*="css"] {
    font-family: 'Barlow', sans-serif;
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

.stApp { background: var(--bg) !important; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: var(--bg2) !important;
    border-right: 1px solid var(--border);
}
[data-testid="stSidebar"] * { color: var(--text) !important; }

/* Header hero */
.hero-header {
    background: linear-gradient(135deg, #0A0E1A 0%, #1A2235 50%, #0D1520 100%);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.hero-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(255,179,0,0.08) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-title {
    font-family: 'Rajdhani', sans-serif;
    font-size: 2.8rem;
    font-weight: 700;
    color: var(--sol);
    margin: 0;
    line-height: 1;
    letter-spacing: 2px;
}
.hero-sub {
    font-size: 1rem;
    color: var(--text2);
    margin-top: 0.4rem;
    letter-spacing: 3px;
    text-transform: uppercase;
    font-weight: 300;
}

/* Cards */
.sol-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}
.sol-card-title {
    font-family: 'Rajdhani', sans-serif;
    font-size: 1.3rem;
    font-weight: 600;
    color: var(--sol);
    letter-spacing: 1px;
    margin-bottom: 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: center;
    gap: 0.5rem;
}

/* Metric boxes */
.metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 1rem;
    margin: 1rem 0;
}
.metric-box {
    background: var(--card2);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.metric-val {
    font-family: 'Share Tech Mono', monospace;
    font-size: 1.8rem;
    color: var(--sol);
    font-weight: 700;
    line-height: 1;
}
.metric-unit {
    font-size: 0.75rem;
    color: var(--sol-light);
    margin-top: 0.2rem;
}
.metric-label {
    font-size: 0.75rem;
    color: var(--text2);
    margin-top: 0.4rem;
    text-transform: uppercase;
    letter-spacing: 1px;
}

/* Result highlight */
.result-highlight {
    background: linear-gradient(135deg, rgba(255,179,0,0.1), rgba(255,179,0,0.05));
    border: 1px solid rgba(255,179,0,0.4);
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin: 0.8rem 0;
}
.result-highlight .val {
    font-family: 'Share Tech Mono', monospace;
    font-size: 2rem;
    color: var(--sol);
    font-weight: 700;
}
.result-ok {
    background: linear-gradient(135deg, rgba(0,230,118,0.1), rgba(0,230,118,0.03));
    border: 1px solid rgba(0,230,118,0.3);
    border-radius: 10px;
    padding: 1rem 1.5rem;
    margin: 0.5rem 0;
}
.result-ok .val {
    font-family: 'Share Tech Mono', monospace;
    font-size: 1.6rem;
    color: var(--green);
    font-weight: 600;
}

/* Step badge */
.step-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 28px;
    height: 28px;
    background: var(--sol);
    color: var(--bg);
    border-radius: 50%;
    font-family: 'Rajdhani', sans-serif;
    font-weight: 700;
    font-size: 0.9rem;
    margin-right: 0.5rem;
    flex-shrink: 0;
}

/* Tables */
.dataframe { font-family: 'Share Tech Mono', monospace !important; font-size: 0.8rem !important; }

/* Inputs */
.stTextInput input, .stNumberInput input, .stSelectbox select {
    background: var(--card2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: var(--sol) !important;
    box-shadow: 0 0 0 2px rgba(255,179,0,0.2) !important;
}

/* Buttons */
.stButton button {
    background: var(--sol) !important;
    color: var(--bg) !important;
    font-family: 'Rajdhani', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    letter-spacing: 1px !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.5rem 1.5rem !important;
    transition: all 0.2s !important;
}
.stButton button:hover {
    background: var(--sol-light) !important;
    transform: translateY(-1px) !important;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: var(--bg2) !important;
    border-radius: 10px !important;
    padding: 4px !important;
    gap: 4px !important;
    border: 1px solid var(--border) !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: var(--text2) !important;
    border-radius: 8px !important;
    font-family: 'Rajdhani', sans-serif !important;
    font-weight: 600 !important;
    letter-spacing: 1px !important;
    padding: 0.5rem 1.2rem !important;
}
.stTabs [aria-selected="true"] {
    background: var(--sol) !important;
    color: var(--bg) !important;
}

/* Formula box */
.formula-box {
    background: var(--bg3);
    border-left: 3px solid var(--sol);
    border-radius: 0 8px 8px 0;
    padding: 0.8rem 1rem;
    margin: 0.5rem 0;
    font-family: 'Share Tech Mono', monospace;
    font-size: 0.85rem;
    color: var(--sol-light);
}

/* Info note */
.info-note {
    background: rgba(0,188,212,0.08);
    border: 1px solid rgba(0,188,212,0.25);
    border-radius: 8px;
    padding: 0.7rem 1rem;
    font-size: 0.85rem;
    color: var(--cyan);
}

/* Separator */
.sep { border: none; border-top: 1px solid var(--border); margin: 1.5rem 0; }

/* Warning */
.warn-box {
    background: rgba(255,82,82,0.08);
    border: 1px solid rgba(255,82,82,0.3);
    border-radius: 8px;
    padding: 0.7rem 1rem;
    color: var(--red);
    font-size: 0.85rem;
}

div[data-testid="stDataFrameContainer"] table {
    background: var(--card) !important;
}
</style>
""", unsafe_allow_html=True)

# ─── DATABASE ────────────────────────────────────────────────────────────────
import pathlib
import os
import tempfile

# ─── RUTA PERSISTENTE DE LA BASE DE DATOS ────────────────────────────────────
# Prioridad:
#   1. st.secrets["db_path"]       → Streamlit Cloud con storage externo / servidor
#   2. SOLARCALC_DB_PATH env var   → Docker / systemd / CI
#   3. Misma carpeta que el script → Local Windows/Mac/Linux (si es escribible)
#   4. /tmp                        → Fallback Streamlit Cloud (efímero pero funcional)
def _resolve_db_path() -> str:
    # 1. Streamlit secrets
    try:
        import streamlit as _st
        if "db_path" in _st.secrets:
            p = _st.secrets["db_path"]
            # Asegurar que el directorio padre exista
            pathlib.Path(p).parent.mkdir(parents=True, exist_ok=True)
            return p
    except Exception:
        pass

    # 2. Variable de entorno
    env_path = os.environ.get("SOLARCALC_DB_PATH")
    if env_path:
        pathlib.Path(env_path).parent.mkdir(parents=True, exist_ok=True)
        return env_path

    # 3. Misma carpeta del script — solo si es escribible
    script_dir = pathlib.Path(__file__).parent.resolve()
    candidate  = script_dir / "solar_calc.db"
    try:
        # Prueba de escritura sin crear el archivo
        test_file = script_dir / ".write_test"
        test_file.touch()
        test_file.unlink()
        return str(candidate)
    except (OSError, PermissionError):
        pass

    # 4. /tmp — siempre escribible (Streamlit Cloud, contenedores sin volumen)
    tmp_path = pathlib.Path(tempfile.gettempdir()) / "solar_calc.db"
    return str(tmp_path)

DB_PATH = _resolve_db_path()

# ── CRÍTICO: fijar env var ANTES de importar módulos externos ─────────────────
os.environ["SOLARCALC_DB_PATH"] = DB_PATH

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    conn = get_conn()
    c = conn.cursor()
    # Proyecto
    c.execute("""
        CREATE TABLE IF NOT EXISTS proyectos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            municipio TEXT,
            hsp REAL,
            tension_dc INTEGER,
            creado TEXT DEFAULT (datetime('now'))
        )
    """)
    # Cargas
    c.execute("""
        CREATE TABLE IF NOT EXISTS cargas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER,
            electrodomestico TEXT NOT NULL,
            cantidad INTEGER DEFAULT 1,
            potencia_w REAL DEFAULT 0,
            horas_dia REAL DEFAULT 0,
            es_motor INTEGER DEFAULT 0,
            FOREIGN KEY(proyecto_id) REFERENCES proyectos(id)
        )
    """)
    # Panel solar
    c.execute("""
        CREATE TABLE IF NOT EXISTS paneles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER,
            modelo TEXT,
            potencia_wp REAL,
            voc REAL,
            isc REAL,
            FOREIGN KEY(proyecto_id) REFERENCES proyectos(id)
        )
    """)
    # Resultados guardados
    c.execute("""
        CREATE TABLE IF NOT EXISTS resultados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER,
            consumo_dia_wh REAL,
            consumo_con_fs REAL,
            tension_dc INTEGER,
            hsp REAL,
            potencia_instalada_w REAL,
            num_paneles INTEGER,
            capacidad_baterias_ah REAL,
            num_baterias INTEGER,
            corriente_mppt REAL,
            generado TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(proyecto_id) REFERENCES proyectos(id)
        )
    """)
    # Recibos de energía eléctrica
    c.execute("""
        CREATE TABLE IF NOT EXISTS recibos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id INTEGER,
            periodo TEXT,
            kwh_periodo REAL,
            dias_periodo INTEGER DEFAULT 30,
            estrato TEXT,
            tarifa_kwh REAL,
            valor_total REAL,
            observaciones TEXT,
            creado TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(proyecto_id) REFERENCES proyectos(id)
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ─── Módulo de seguridad ─────────────────────────────────────────────────────
from modulo_seguridad import (
    init_seguridad_db, mostrar_login, mostrar_usuario_sidebar,
    mostrar_gestion_usuarios, mostrar_cambio_password,
    tiene_permiso, usuario_activo, registrar_auditoria, ROLES,
)
init_seguridad_db()

# ─── Módulo de cableado ──────────────────────────────────────────────────────
from modulo_cableado import mostrar_cableado

# ─── Log ruta de BD (visible en consola al iniciar) ──────────────────────────
import sys as _sys
print(f"\n☀  SolarCalc Pro — Base de datos en: {DB_PATH}\n", file=_sys.stderr)

# ─── ELECTRODOMÉSTICOS MOTORIZADOS ──────────────────────────────────────────
MOTORIZADOS = ["lavadora", "nevera", "licuadora", "ventilador", "bomba", "aire acondicionado", "compresor"]

def es_motorizado(nombre: str) -> bool:
    return any(m in nombre.lower() for m in MOTORIZADOS)

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def tension_dc(consumo_wh: float) -> int:
    if consumo_wh < 2000:
        return 12
    elif consumo_wh < 4000:
        return 24
    else:
        return 48

def num_paneles(potencia_instalada: float, potencia_panel: float) -> int:
    if potencia_panel <= 0:
        return 0
    return math.ceil(potencia_instalada / potencia_panel)

def calcular_baterias(consumo_wh: float, vdc: int,
                      dod: float = 0.50,
                      cap: float = 100,
                      dias_autonomia: float = 1.0,
                      eficiencia: float = 0.90) -> dict:
    """
    Dimensionamiento de banco de baterías — 5 parámetros principales:
      1. consumo_wh     : Consumo diario de energía (Wh/día)
      2. dias_autonomia : Días de autonomía sin sol (default 1 día = 24 h)
      3. vdc            : Voltaje del banco de baterías (12 / 24 / 48 V)
      4. dod            : Profundidad máxima de descarga — fracción (0.0-1.0)
      5. eficiencia     : Eficiencia global del sistema — fracción (0.0-1.0)
                          (pérdidas en inversores, cables, autodescargas, etc.)

    Fórmula:
      E_total (Wh) = consumo_wh × días_autonomía / eficiencia
      Ah_req       = E_total / (V_banco × DoD)
      N° baterías  = ⌈ Ah_req / cap_bateria ⌉
    """
    e_total  = (consumo_wh * dias_autonomia) / max(eficiencia, 0.01)
    ah_bruto = e_total / max(vdc, 1)
    ah_req   = e_total / (max(vdc, 1) * max(dod, 0.01))
    n        = math.ceil(ah_req / max(cap, 1))
    # Ajustar a número par si es > 1 (facilita configuración serie/paralelo)
    if n > 1 and n % 2 != 0:
        n += 1
    cap_real_ah  = n * cap
    energia_real = cap_real_ah * vdc / 1000          # kWh brutos
    autonomia_real_h = (cap_real_ah * vdc * dod * eficiencia) / max(consumo_wh / 24, 0.001)
    return {
        "ah_bruto":          ah_bruto,
        "ah_dod":            ah_req,
        "ah_final":          ah_req,
        "num_baterias":      n,
        "cap_real_ah":       cap_real_ah,
        "energia_kwh":       energia_real,
        "energia_util_kwh":  energia_real * dod * eficiencia,
        "autonomia_real_h":  autonomia_real_h,
        "autonomia_real_d":  autonomia_real_h / 24,
        "e_total_wh":        e_total,
        "eficiencia":        eficiencia,
        "dod":               dod,
        "dias_autonomia":    dias_autonomia,
        "vdc":               vdc,
        "cap_bat_ah":        cap,
    }

def _bat_params(ss: dict = None) -> dict:
    """Devuelve los parámetros de baterías guardados en session_state (o defaults).
    Úsalo para pasar parámetros consistentes a calcular_baterias() en todo el app."""
    _ss = ss or {}
    return {
        "dod":            _ss.get("calc_dod_pct",       80) / 100,
        "cap":            _ss.get("calc_bat_cap_ah",    100),
        "dias_autonomia": _ss.get("calc_dias_autonomia", 1.0),
        "eficiencia":     _ss.get("calc_eff_pct",        90) / 100,
    }

# Tamaños comerciales de inversores (kW)
_KW_COMERCIALES = [0.5, 1, 1.5, 2, 3, 3.5, 4, 5, 6, 8, 10, 12, 15, 20, 25, 30, 40, 50]

def calcular_inversor(cargas_df: "pd.DataFrame",
                      fs: float = 0.80,
                      fm: float = 1.25,
                      vdc: int = 24) -> dict:
    """Dimensionamiento de inversor según metodología del documento técnico.

    Pasos:
      1. Potencia instalada = suma de (cantidad × potencia_w)
      2. Demanda simultánea = potencia instalada × FS (factor de simultaneidad 0.6-1.0)
      3. Potencia de arranque de motores = excedente sobre potencia nominal
         Factor arranque: nevera/AC/ventilador→3, bomba→4, compresor→5, otros motores→2
      4. Potencia requerida = P_simult + P_arranque
      5. Inversor mínimo = P_requerida × FM (factor de margen 1.20-1.30)
      6. Tamaño comercial = siguiente valor estándar superior (kW)
      7. Corriente DC = P_inv / V_sistema
    """
    import pandas as _pd

    # Factores de arranque por tipo de carga
    _FACTORES_ARRANQUE = {
        "nevera": 3, "refriger": 3, "congelador": 3,
        "aire acondicionado": 3, "aire": 3,
        "ventilador": 3, "bomba": 4,
        "compresor": 5,
        "lavadora": 2, "licuadora": 2, "procesadora": 2,
        "extractor": 2, "motor": 2,
    }

    if cargas_df is None or cargas_df.empty:
        return {
            "pot_instalada": 0, "pot_simultanea": 0, "pot_arranque": 0,
            "pot_requerida": 0, "pot_inv_minima": 0, "inv_kw": 1.0,
            "inv_w": 1000, "corr_dc": 0,
            "fs": fs, "fm": fm, "desglose": [],
        }

    df = cargas_df.copy()
    df["pot_total_w"] = df["cantidad"] * df["potencia_w"]
    pot_instalada = df["pot_total_w"].sum()

    # Demanda simultánea
    pot_simultanea = pot_instalada * fs

    # Arranque de motores: solo se suma el excedente del motor más exigente
    # (pico de arranque, no todos a la vez)
    pot_arranque = 0.0
    desglose_arranque = []
    for _, row in df.iterrows():
        nombre = str(row.get("electrodomestico", "")).lower()
        es_mot = int(row.get("es_motor", 0))
        pot_nom = float(row.get("pot_total_w", row["cantidad"] * row["potencia_w"]))
        if es_mot:
            factor_a = 2  # default
            for keyword, fa in _FACTORES_ARRANQUE.items():
                if keyword in nombre:
                    factor_a = fa
                    break
            pot_arranque_eq = pot_nom * (factor_a - 1)  # excedente sobre nominal
            desglose_arranque.append({
                "equipo": row.get("electrodomestico", "—"),
                "pot_nominal_w": pot_nom,
                "factor_arranque": factor_a,
                "excedente_w": pot_arranque_eq,
            })
        # Se toma solo el mayor pico de arranque (no suma simultánea de todos los arranques)
    if desglose_arranque:
        pot_arranque = max(d["excedente_w"] for d in desglose_arranque)

    pot_requerida = pot_simultanea + pot_arranque
    pot_inv_minima = pot_requerida * fm

    # Tamaño comercial
    inv_kw = float(next(
        (k for k in _KW_COMERCIALES if k * 1000 >= pot_inv_minima),
        math.ceil(pot_inv_minima / 1000)
    ))
    inv_w = inv_kw * 1000
    corr_dc = inv_w / vdc if vdc > 0 else 0

    return {
        "pot_instalada":   pot_instalada,
        "pot_simultanea":  pot_simultanea,
        "pot_arranque":    pot_arranque,
        "pot_requerida":   pot_requerida,
        "pot_inv_minima":  pot_inv_minima,
        "inv_kw":          inv_kw,
        "inv_w":           inv_w,
        "corr_dc":         corr_dc,
        "fs":              fs,
        "fm":              fm,
        "desglose_arranque": desglose_arranque,
    }

# ─── EXPORT: EXCEL ────────────────────────────────────────────────────────────
def generar_excel(proyecto_id: int, proyecto_info: tuple) -> bytes:
    conn = get_conn()
    cargas_df = pd.read_sql("SELECT * FROM cargas WHERE proyecto_id=?", conn, params=(proyecto_id,))
    panel_row = conn.execute("SELECT * FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                              (proyecto_id,)).fetchone()
    recibos_df = pd.read_sql("SELECT * FROM recibos WHERE proyecto_id=? ORDER BY id DESC", conn, params=(proyecto_id,))
    conn.close()

    wb = openpyxl.Workbook()

    # ── Paleta ────────────────────────────────────────────────────────────────
    C_DARK   = "0A0E1A"
    C_CARD   = "1A2235"
    C_HEADER = "FFB300"
    C_SUB    = "1E2A3F"
    C_TEXT   = "E8EDF5"
    C_GREEN  = "00E676"
    C_CYAN   = "00BCD4"
    C_WARN   = "FF5252"
    C_MONO   = "FFD54F"
    C_BORDER = "2A3A55"

    def hdr_fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def hdr_font(bold=True, color=C_TEXT, size=10): return Font(bold=bold, color=color, size=size, name="Calibri")
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    def border():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    # ══ HOJA 1 — CARGAS ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "1 · Análisis de Cargas"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 17
    ws1.column_dimensions["H"].width = 17
    ws1.column_dimensions["I"].width = 10

    # Título principal
    ws1.merge_cells("A1:I1")
    c = ws1["A1"]
    c.value = "☀  SOLARCALC PRO — ANÁLISIS DE CARGA ELÉCTRICA"
    c.fill = hdr_fill(C_DARK); c.font = Font(bold=True, color=C_HEADER, size=16, name="Calibri")
    c.alignment = center(); ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:I2")
    c = ws1["A2"]
    c.value = f"Proyecto: {proyecto_info[1]}  |  Municipio: {proyecto_info[2] or '—'}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill = hdr_fill(C_CARD); c.font = Font(color="8A9BBD", size=9, name="Calibri")
    c.alignment = center(); ws1.row_dimensions[2].height = 18

    ws1.row_dimensions[3].height = 6

    # Encabezados tabla
    headers = ["ID","Electrodoméstico","Cant.","Potencia (W)","Pot. Total (W)","Horas/día","Consumo día (Wh)","Motor"]
    for col_i, h in enumerate(headers, 1):
        c = ws1.cell(row=4, column=col_i, value=h)
        c.fill = hdr_fill(C_HEADER); c.font = Font(bold=True, color=C_DARK, size=9, name="Calibri")
        c.alignment = center(); c.border = border()
    ws1.row_dimensions[4].height = 22

    # Datos
    if not cargas_df.empty:
        cargas_df["pot_total_w"]   = cargas_df["cantidad"] * cargas_df["potencia_w"]
        cargas_df["consumo_wh"]    = cargas_df["pot_total_w"] * cargas_df["horas_dia"]

        _vdc_xl = proyecto_info[3] or 24
        _inv_xl = calcular_inversor(cargas_df, fs=0.80, fm=1.25, vdc=_vdc_xl)

        alt = False
        for _, row in cargas_df.iterrows():
            r = ws1.max_row + 1
            bg = "161D30" if alt else C_SUB; alt = not alt
            vals = [row["id"], row["electrodomestico"], row["cantidad"],
                    row["potencia_w"], row["pot_total_w"], row["horas_dia"],
                    row["consumo_wh"],
                    "⚡ Motor" if row["es_motor"] else "—"]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(row=r, column=ci, value=v)
                c.fill = hdr_fill(bg)
                fc = C_WARN if (ci == 8 and row["es_motor"]) else C_TEXT
                c.font = Font(color=fc, size=9, name="Calibri")
                c.alignment = center() if ci != 2 else left()
                c.border = border()
                if ci in (4,5,7): c.number_format = "#,##0.0"
            ws1.row_dimensions[r].height = 18

        # Totales
        consumo_total = cargas_df["consumo_wh"].sum()
        consumo_fs    = consumo_total * 1.20

        totales = [
            ("", "TOTAL BASE",    "", "", cargas_df["pot_total_w"].sum(), "", consumo_total, ""),
            ("", "TOTAL + 20% FS","", "", "", "",                              consumo_fs,   ""),
        ]
        for t_vals in totales:
            r = ws1.max_row + 1
            for ci, v in enumerate(t_vals, 1):
                c = ws1.cell(row=r, column=ci, value=v if v != "" else None)
                c.fill = hdr_fill("0F1525")
                c.font = Font(bold=True, color=C_HEADER, size=9, name="Calibri")
                c.alignment = center()
                c.border = border()
                if isinstance(v, float): c.number_format = "#,##0.0"
            ws1.row_dimensions[r].height = 20

        # Bloque resumen — inversor con metodología técnica
        r = ws1.max_row + 2
        summaries = [
            ("Tensión DC del sistema (V)",          proyecto_info[3] or tension_dc(consumo_fs),         C_MONO),
            ("HSP guardado (h)",                     proyecto_info[4] or "—",                            C_CYAN),
            ("Consumo base (Wh/día)",                round(consumo_total,1),                             C_TEXT),
            ("Consumo + 20% FS (Wh/día)",            round(consumo_fs,1),                                C_HEADER),
            ("— CÁLCULO INVERSOR —",                 "",                                                  C_WARN),
            ("Pot. instalada total (W)",             round(_inv_xl["pot_instalada"],1),                  C_TEXT),
            (f"Demanda simultánea (×FS {int(_inv_xl['fs']*100)}%)",  round(_inv_xl["pot_simultanea"],1), C_CYAN),
            ("Pot. arranque motores (excedente)",    round(_inv_xl["pot_arranque"],1),                   C_WARN),
            ("Pot. requerida (sim+arranque)",        round(_inv_xl["pot_requerida"],1),                  C_HEADER),
            (f"Pot. mínima inversor (×FM {int(_inv_xl['fm']*100)}%)",round(_inv_xl["pot_inv_minima"],1), C_HEADER),
            ("✅ Inversor recomendado (comercial)",  f"{_inv_xl['inv_kw']:.1f} kW / {_inv_xl['inv_w']:,.0f} W", C_GREEN),
            (f"Corriente DC @ {_vdc_xl}V",          f"{_inv_xl['corr_dc']:.1f} A",                      C_MONO),
        ]
        ws1.merge_cells(f"A{r}:C{r}")
        ws1.cell(row=r, column=1, value="PARÁMETRO").fill = hdr_fill(C_HEADER)
        ws1.cell(row=r, column=1).font = Font(bold=True, color=C_DARK, size=9, name="Calibri")
        ws1.cell(row=r, column=1).alignment = center()
        ws1.merge_cells(f"D{r}:F{r}")
        ws1.cell(row=r, column=4, value="VALOR").fill = hdr_fill(C_HEADER)
        ws1.cell(row=r, column=4).font = Font(bold=True, color=C_DARK, size=9, name="Calibri")
        ws1.cell(row=r, column=4).alignment = center()
        r += 1
        for param, val, fc in summaries:
            ws1.merge_cells(f"A{r}:C{r}")
            ws1.cell(row=r, column=1, value=param).fill = hdr_fill(C_CARD)
            ws1.cell(row=r, column=1).font = Font(color="8A9BBD", size=9, name="Calibri")
            ws1.cell(row=r, column=1).alignment = left(); ws1.cell(row=r,column=1).border = border()
            ws1.merge_cells(f"D{r}:F{r}")
            ws1.cell(row=r, column=4, value=val).fill = hdr_fill(C_SUB)
            ws1.cell(row=r, column=4).font = Font(bold=True, color=fc, size=11, name="Calibri")
            ws1.cell(row=r, column=4).alignment = center(); ws1.cell(row=r,column=4).border = border()
            ws1.row_dimensions[r].height = 20; r += 1

    # ══ HOJA 2 — SISTEMA COMPLETO ════════════════════════════════════════════
    ws2 = wb.create_sheet("2 · Sistema Completo")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 22

    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value = "☀  SOLARCALC PRO — DIMENSIONAMIENTO SISTEMA FV"
    c.fill = hdr_fill(C_DARK); c.font = Font(bold=True, color=C_HEADER, size=15, name="Calibri")
    c.alignment = center(); ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:C2")
    ws2["A2"].value = f"Proyecto: {proyecto_info[1]}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws2["A2"].fill = hdr_fill(C_CARD); ws2["A2"].font = Font(color="8A9BBD", size=9, name="Calibri")
    ws2["A2"].alignment = center(); ws2.row_dimensions[2].height = 18

    if not cargas_df.empty:
        consumo_total2 = cargas_df["consumo_wh"].sum()
        consumo_fs2    = consumo_total2 * 1.20   # FS 20%
        vdc2           = proyecto_info[3] or tension_dc(consumo_fs2)
        hsp2           = proyecto_info[4] or 4.2
        pp_wp          = panel_row[3] if panel_row else 550
        isc2           = panel_row[5] if panel_row else 8.02
        pot_inst2      = consumo_fs2 / hsp2
        n_pan2         = num_paneles(pot_inst2, pp_wp)
        bats2          = calcular_baterias(consumo_fs2, vdc2, **_bat_params(st.session_state))
        # 4. Controlador: Isc × N_paneles
        corr_mppt2     = isc2 * n_pan2

        if corr_mppt2 <= 40:   mppt2 = "MPPT 40A"
        elif corr_mppt2 <= 60: mppt2 = "MPPT 60A"
        elif corr_mppt2 <= 100:mppt2 = "MPPT 100A"
        else: mppt2 = f"MPPT {math.ceil(corr_mppt2/50)*50}A"
        # Inversor — metodología técnica
        _inv2 = calcular_inversor(cargas_df, fs=0.80, fm=1.25, vdc=vdc2)
        inv_kw2 = _inv2["inv_kw"]
        _db2 = _inv2.get("desglose_arranque", [])
        _arr_desc2 = " | ".join(f"{d['equipo']}(×{d['factor_arranque']})" for d in _db2) or "—"

        secciones = [
            ("⚡ CONSUMO ENERGÉTICO", [
                ("Consumo base diario (Wh/día)",       f"{consumo_total2:,.1f} Wh"),
                ("Factor de seguridad",                 "20%"),
                ("Consumo diario con FS (Wh/día)",      f"{consumo_fs2:,.1f} Wh"),
            ]),
            ("🔋 TENSIÓN DC DEL SISTEMA", [
                ("Tensión estándar seleccionada",       f"{vdc2} V DC"),
                ("Rango de consumo",                    "< 2kWh=12V | 2-4kWh=24V | ≥4kWh=48V"),
            ]),
            ("🌞 HORA SOLAR PICO", [
                ("Municipio analizado",                 proyecto_info[2] or "—"),
                ("HSP utilizado",                       f"{hsp2} h/día"),
            ]),
            ("🔆 CAMPO FOTOVOLTAICO", [
                ("Panel solar",                         panel_row[2] if panel_row else "—"),
                ("Potencia por panel",                  f"{pp_wp} Wp"),
                ("Tensión Voc",                         f"{panel_row[4] if panel_row else '—'} V"),
                ("Corriente Isc",                       f"{isc2} A"),
                ("Potencia mínima instalada (Wp)",      f"{pot_inst2:,.0f} Wp"),
                ("Número de paneles requeridos",        f"{n_pan2} paneles"),
                ("Potencia real instalada (Wp)",        f"{n_pan2 * pp_wp:,} Wp"),
            ]),
            ("🔋 BANCO DE BATERÍAS", [
                ("① Consumo diario con FS (Wh/día)",    f"{consumo_fs2:,.1f} Wh"),
                ("② Días de autonomía",                  f"{bats2['dias_autonomia']:.3f} días ({bats2['dias_autonomia']*24:.0f} h)"),
                ("③ Voltaje banco",                      f"{vdc2} V DC"),
                (f"④ DoD",                               f"{int(bats2['dod']*100)}%"),
                (f"⑤ Eficiencia sistema (η)",            f"{int(bats2['eficiencia']*100)}%"),
                ("E_total = consumo × días ÷ η",         f"{bats2['e_total_wh']:,.1f} Wh"),
                ("Ah = E_total ÷ (V × DoD)",             f"{bats2['ah_dod']:,.1f} Ah"),
                (f"Número de baterías {int(bats2['cap_bat_ah'])} Ah", f"{bats2['num_baterias']} unidades"),
                ("Capacidad real instalada",              f"{bats2['cap_real_ah']:,} Ah @ {vdc2}V"),
                ("Energía bruta almacenada",              f"{bats2['energia_kwh']:.2f} kWh"),
                ("✅ Energía utilizable",                 f"{bats2['energia_util_kwh']:.2f} kWh"),
                ("Autonomía real",                        f"{bats2['autonomia_real_h']:.1f} h ({bats2['autonomia_real_d']:.2f} días)"),
            ]),
            ("🎛 CONTROLADOR MPPT", [
                ("Tensión del sistema",                 f"{vdc2} V"),
                ("Isc panel × N° paneles",              f"{isc2} A × {n_pan2} = {corr_mppt2:.1f} A"),
                ("Controlador recomendado",             mppt2),
            ]),
            ("🔌 INVERSOR DC/AC — DIMENSIONAMIENTO TÉCNICO", [
                ("1. Pot. instalada total",             f"{_inv2['pot_instalada']:,.0f} W"),
                (f"2. Demanda simultánea (×FS {int(_inv2['fs']*100)}%)",
                                                        f"{_inv2['pot_simultanea']:,.0f} W"),
                (f"3. Pico de arranque motores",        f"{_inv2['pot_arranque']:,.0f} W — {_arr_desc2}"),
                ("4. Potencia requerida (2+3)",         f"{_inv2['pot_requerida']:,.0f} W"),
                (f"5. Pot. mínima inversor (×FM {int(_inv2['fm']*100)}%)",
                                                        f"{_inv2['pot_inv_minima']:,.0f} W"),
                ("✅ INVERSOR RECOMENDADO",             f"{inv_kw2:.1f} kW / {_inv2['inv_w']:,.0f} W"),
                (f"Corriente DC @ {vdc2}V",             f"{_inv2['corr_dc']:.1f} A"),
            ]),
        ]

        r2 = 4
        for titulo, filas in secciones:
            ws2.row_dimensions[r2].height = 8; r2 += 1
            ws2.merge_cells(f"A{r2}:C{r2}")
            ws2.cell(row=r2, column=1, value=titulo).fill = hdr_fill(C_HEADER)
            ws2.cell(row=r2, column=1).font = Font(bold=True, color=C_DARK, size=10, name="Calibri")
            ws2.cell(row=r2, column=1).alignment = left(); ws2.row_dimensions[r2].height = 22; r2 += 1
            for param, val in filas:
                ws2.cell(row=r2, column=1, value=param).fill = hdr_fill(C_CARD)
                ws2.cell(row=r2, column=1).font = Font(color="8A9BBD", size=9, name="Calibri")
                ws2.cell(row=r2, column=1).alignment = left(); ws2.cell(row=r2,column=1).border = border()
                ws2.merge_cells(f"B{r2}:C{r2}")
                ws2.cell(row=r2, column=2, value=val).fill = hdr_fill(C_SUB)
                ws2.cell(row=r2, column=2).font = Font(bold=True, color=C_MONO, size=10, name="Calibri")
                ws2.cell(row=r2, column=2).alignment = center(); ws2.cell(row=r2,column=2).border = border()
                ws2.row_dimensions[r2].height = 19; r2 += 1

    # ══ HOJA 3 — RECIBOS ═════════════════════════════════════════════════════
    if not recibos_df.empty:
        ws3 = wb.create_sheet("3 · Recibos de Energía")
        ws3.sheet_view.showGridLines = False
        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 18
        ws3.column_dimensions["D"].width = 14
        ws3.column_dimensions["E"].width = 12
        ws3.column_dimensions["F"].width = 13
        ws3.column_dimensions["G"].width = 13
        ws3.column_dimensions["H"].width = 16
        ws3.column_dimensions["I"].width = 16
        ws3.column_dimensions["J"].width = 16

        # Título
        ws3.merge_cells("A1:J1")
        c = ws3["A1"]
        c.value = "🧾  RECIBOS DE ENERGÍA ELÉCTRICA"
        c.fill = hdr_fill(C_DARK); c.font = Font(bold=True, color=C_HEADER, size=14, name="Calibri")
        c.alignment = center(); ws3.row_dimensions[1].height = 30

        ws3.merge_cells("A2:J2")
        ws3["A2"].value = f"Proyecto: {proyecto_info[1]}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws3["A2"].fill = hdr_fill(C_CARD); ws3["A2"].font = Font(color="8A9BBD", size=9, name="Calibri")
        ws3["A2"].alignment = center(); ws3.row_dimensions[2].height = 16

        hdrs3 = ["ID","Período","Estrato","kWh período","Días","kWh/día","Wh/día","Wh/día +20%FS","Tarifa ($/kWh)","Valor total ($)"]
        for ci, h in enumerate(hdrs3, 1):
            c = ws3.cell(row=4, column=ci, value=h)
            c.fill = hdr_fill(C_HEADER); c.font = Font(bold=True, color=C_DARK, size=9, name="Calibri")
            c.alignment = center(); c.border = border()
        ws3.row_dimensions[4].height = 22

        recibos_df["kwh_dia_x"]   = recibos_df["kwh_periodo"] / recibos_df["dias_periodo"]
        recibos_df["wh_dia_x"]    = recibos_df["kwh_dia_x"] * 1000
        recibos_df["wh_dia_fs_x"] = recibos_df["wh_dia_x"] * 1.20
        recibos_df["tarifa_x"]    = recibos_df.apply(
            lambda r: r["valor_total"]/r["kwh_periodo"] if r["valor_total"] and r["kwh_periodo"]>0 else (r["tarifa_kwh"] or ""), axis=1)

        alt3 = False
        for _, row in recibos_df.iterrows():
            r3 = ws3.max_row + 1
            bg3 = "161D30" if alt3 else C_SUB; alt3 = not alt3
            vals3 = [row["id"], row["periodo"], row["estrato"] or "—",
                     row["kwh_periodo"], row["dias_periodo"],
                     round(row["kwh_dia_x"],2), round(row["wh_dia_x"],0),
                     round(row["wh_dia_fs_x"],0),
                     round(row["tarifa_x"],1) if isinstance(row["tarifa_x"], float) else "—",
                     row["valor_total"] if row["valor_total"] else "—"]
            for ci, v in enumerate(vals3, 1):
                c = ws3.cell(row=r3, column=ci, value=v)
                c.fill = hdr_fill(bg3)
                c.font = Font(color=C_MONO if ci in (6,7,8) else C_TEXT, size=9, name="Calibri")
                c.alignment = center() if ci != 2 else left()
                c.border = border()
                if isinstance(v, float) and ci in (4,6,7,8,9,10): c.number_format = "#,##0.0"
            ws3.row_dimensions[r3].height = 18

        # Fila promedios
        if len(recibos_df) > 1:
            r3avg = ws3.max_row + 1
            ws3.merge_cells(f"A{r3avg}:C{r3avg}")
            ws3.cell(row=r3avg, column=1, value="PROMEDIO").fill = hdr_fill(C_HEADER)
            ws3.cell(row=r3avg, column=1).font = Font(bold=True, color=C_DARK, size=9, name="Calibri")
            ws3.cell(row=r3avg, column=1).alignment = center()
            for ci, col in [(4,"kwh_periodo"),(6,"kwh_dia_x"),(7,"wh_dia_x"),(8,"wh_dia_fs_x")]:
                c = ws3.cell(row=r3avg, column=ci, value=round(recibos_df[col].mean(),1))
                c.fill = hdr_fill("0F1525"); c.font = Font(bold=True, color=C_HEADER, size=9, name="Calibri")
                c.alignment = center(); c.border = border(); c.number_format = "#,##0.0"
            ws3.row_dimensions[r3avg].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── EXPORT: PDF ──────────────────────────────────────────────────────────────
def generar_pdf(proyecto_id: int, proyecto_info: tuple) -> bytes:
    conn = get_conn()
    cargas_df = pd.read_sql("SELECT * FROM cargas WHERE proyecto_id=?", conn, params=(proyecto_id,))
    panel_row = conn.execute("SELECT * FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                              (proyecto_id,)).fetchone()
    recibos_df = pd.read_sql("SELECT * FROM recibos WHERE proyecto_id=? ORDER BY id DESC", conn, params=(proyecto_id,))
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    # Colores PDF
    SOL    = colors.HexColor("#FFB300")
    DARK   = colors.HexColor("#0A0E1A")
    CARD   = colors.HexColor("#1A2235")
    CARD2  = colors.HexColor("#1E2A3F")
    TEXT   = colors.HexColor("#E8EDF5")
    TEXT2  = colors.HexColor("#8A9BBD")
    GREEN  = colors.HexColor("#00E676")
    CYAN   = colors.HexColor("#00BCD4")
    MONO   = colors.HexColor("#FFD54F")
    RED    = colors.HexColor("#FF5252")
    BORDER = colors.HexColor("#2A3A55")

    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=18,
                                textColor=SOL, alignment=TA_CENTER, spaceAfter=4)
    sub_st    = ParagraphStyle("sub",    fontName="Helvetica",      fontSize=9,
                                textColor=TEXT2, alignment=TA_CENTER, spaceAfter=8)
    sec_st    = ParagraphStyle("sec",    fontName="Helvetica-Bold", fontSize=11,
                                textColor=SOL, spaceBefore=12, spaceAfter=4)
    body_st   = ParagraphStyle("body",   fontName="Helvetica",      fontSize=9,
                                textColor=TEXT)

    story = []

    # ── Encabezado ────────────────────────────────────────────────────────────
    story.append(Paragraph("☀  SOLARCALC PRO — DIMENSIONAMIENTO FOTOVOLTAICO", titulo_st))
    story.append(Paragraph(
        f"Proyecto: <b>{proyecto_info[1]}</b>  |  Municipio: {proyecto_info[2] or '—'}  |  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sub_st))
    story.append(HRFlowable(width="100%", thickness=1, color=SOL, spaceAfter=10))

    if not cargas_df.empty:
        cargas_df["pot_total_w"]  = cargas_df["cantidad"] * cargas_df["potencia_w"]
        cargas_df["consumo_wh"]   = cargas_df["pot_total_w"] * cargas_df["horas_dia"]

        consumo_total = cargas_df["consumo_wh"].sum()
        consumo_fs    = consumo_total * 1.20
        vdc_p         = proyecto_info[3] or tension_dc(consumo_fs)
        hsp_p         = proyecto_info[4] or 4.2
        pp_wp_p       = panel_row[3] if panel_row else 550
        isc_p         = panel_row[5] if panel_row else 8.02
        pot_inst_p    = consumo_fs / hsp_p
        n_pan_p       = num_paneles(pot_inst_p, pp_wp_p)
        bats_p        = calcular_baterias(consumo_fs, vdc_p, **_bat_params(st.session_state))
        corr_p        = isc_p * n_pan_p
        if corr_p <= 40:   mppt_p = "MPPT 40A"
        elif corr_p <= 60: mppt_p = "MPPT 60A"
        elif corr_p <= 100:mppt_p = "MPPT 100A"
        else: mppt_p = f"MPPT {math.ceil(corr_p/50)*50}A"

        # Inversor con metodología técnica
        _inv_p = calcular_inversor(cargas_df, fs=0.80, fm=1.25, vdc=vdc_p)
        inv_kw_p  = _inv_p["inv_kw"]
        pot_inv_fs = _inv_p["inv_w"]

        # ── Tabla cargas ──────────────────────────────────────────────────────
        story.append(Paragraph("⚡  ANÁLISIS DE CARGA ELÉCTRICA", sec_st))
        tbl_hdr = [["ID","Electrodoméstico","Cant.","Potencia\n(W)","Pot. Total\n(W)",
                     "Horas/\ndía","Consumo día\n(Wh)","Motor"]]
        tbl_data = tbl_hdr[:]
        for _, row in cargas_df.iterrows():
            tbl_data.append([
                str(int(row["id"])),
                row["electrodomestico"],
                str(int(row["cantidad"])),
                f"{row['potencia_w']:,.0f}",
                f"{row['pot_total_w']:,.0f}",
                f"{row['horas_dia']:.1f}",
                f"{row['consumo_wh']:,.1f}",
                "⚡ Motor" if row["es_motor"] else "—",
            ])
        # fila totales
        tbl_data.append(["","TOTAL BASE","","",
                          f"{cargas_df['pot_total_w'].sum():,.0f}","",
                          f"{consumo_total:,.1f}",""])
        tbl_data.append(["","TOTAL + 20% FS","","","","",
                          f"{consumo_fs:,.1f}",""])

        col_w = [1.2*cm, 6.5*cm, 1.2*cm, 2.0*cm, 2.0*cm, 1.5*cm, 2.5*cm, 1.9*cm]
        t = Table(tbl_data, colWidths=col_w, repeatRows=1)
        n_data = len(tbl_data)
        t.setStyle(TableStyle([
            # Header
            ("BACKGROUND",  (0,0),  (-1,0),  SOL),
            ("TEXTCOLOR",   (0,0),  (-1,0),  DARK),
            ("FONTNAME",    (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0,0),  (-1,0),  8),
            ("ALIGN",       (0,0),  (-1,0),  "CENTER"),
            ("VALIGN",      (0,0),  (-1,-1), "MIDDLE"),
            # Body alt rows
            *[("BACKGROUND",(0,i),(-1,i), CARD if i%2==0 else CARD2) for i in range(1, n_data-2)],
            ("TEXTCOLOR",   (0,1),  (-1,n_data-3), TEXT),
            ("FONTNAME",    (0,1),  (-1,-1), "Helvetica"),
            ("FONTSIZE",    (0,1),  (-1,-1), 8),
            ("ALIGN",       (0,1),  (-1,-1), "CENTER"),
            ("ALIGN",       (1,1),  (1,-1),  "LEFT"),
            # Totals
            ("BACKGROUND",  (0,n_data-2), (-1,n_data-1), colors.HexColor("#0F1525")),
            ("TEXTCOLOR",   (0,n_data-2), (-1,n_data-1), SOL),
            ("FONTNAME",    (0,n_data-2), (-1,n_data-1), "Helvetica-Bold"),
            ("FONTSIZE",    (0,n_data-2), (-1,n_data-1), 8),
            # Grid
            ("GRID",        (0,0),  (-1,-1), 0.4, BORDER),
            ("ROWBACKGROUND",(0,0), (-1,0),  SOL),
            ("TOPPADDING",  (0,0),  (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6))

        # ── Recibos de energía ────────────────────────────────────────────────
        if not recibos_df.empty:
            story.append(Paragraph("🧾  RECIBOS DE ENERGÍA ELÉCTRICA", sec_st))
            recibos_df["kwh_dia_p"]   = recibos_df["kwh_periodo"] / recibos_df["dias_periodo"]
            recibos_df["wh_dia_p"]    = recibos_df["kwh_dia_p"] * 1000
            recibos_df["wh_dia_fs_p"] = recibos_df["wh_dia_p"] * 1.20
            recibos_df["tarifa_p"]    = recibos_df.apply(
                lambda r: r["valor_total"]/r["kwh_periodo"] if r["valor_total"] and r["kwh_periodo"]>0
                          else (r["tarifa_kwh"] or 0), axis=1)

            rec_hdr  = [["Período","Estrato","kWh período","Días","kWh/día","Wh/día","Wh/día +20%FS","Valor ($)"]]
            rec_rows = rec_hdr[:]
            for _, row in recibos_df.iterrows():
                rec_rows.append([
                    row["periodo"],
                    row["estrato"] or "—",
                    f"{row['kwh_periodo']:.1f}",
                    str(int(row["dias_periodo"])),
                    f"{row['kwh_dia_p']:.2f}",
                    f"{row['wh_dia_p']:,.0f}",
                    f"{row['wh_dia_fs_p']:,.0f}",
                    f"${row['valor_total']:,.0f}" if row["valor_total"] else "—",
                ])
            if len(recibos_df) > 1:
                rec_rows.append([
                    "PROMEDIO","","",str(int(recibos_df["dias_periodo"].mean())),
                    f"{recibos_df['kwh_dia_p'].mean():.2f}",
                    f"{recibos_df['wh_dia_p'].mean():,.0f}",
                    f"{recibos_df['wh_dia_fs_p'].mean():,.0f}",""])

            rw_rec = [4.5*cm,2.5*cm,2.5*cm,1.5*cm,2.2*cm,2.5*cm,2.8*cm,2.5*cm]
            tr_rec = Table(rec_rows, colWidths=rw_rec, repeatRows=1)
            nr_rec = len(rec_rows)
            tr_rec.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0),  SOL),
                ("TEXTCOLOR",   (0,0), (-1,0),  DARK),
                ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0),  8),
                ("ALIGN",       (0,0), (-1,0),  "CENTER"),
                *[("BACKGROUND",(0,i),(-1,i), CARD if i%2==0 else CARD2) for i in range(1,nr_rec)],
                ("TEXTCOLOR",   (0,1),(-1,-1),  TEXT),
                ("TEXTCOLOR",   (4,1),(6,nr_rec-1), MONO),
                ("FONTNAME",    (0,1),(-1,-1),  "Helvetica"),
                ("FONTSIZE",    (0,1),(-1,-1),  8),
                ("ALIGN",       (0,1),(-1,-1),  "CENTER"),
                ("BACKGROUND",  (0,nr_rec-1),(-1,nr_rec-1), colors.HexColor("#0F1525")) if len(recibos_df)>1 else ("BACKGROUND",(0,0),(-1,0),SOL),
                ("TEXTCOLOR",   (0,nr_rec-1),(-1,nr_rec-1), SOL) if len(recibos_df)>1 else ("TEXTCOLOR",(0,0),(-1,0),DARK),
                ("FONTNAME",    (0,nr_rec-1),(-1,nr_rec-1), "Helvetica-Bold") if len(recibos_df)>1 else ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("GRID",        (0,0),(-1,-1),  0.4, BORDER),
                ("VALIGN",      (0,0),(-1,-1),  "MIDDLE"),
                ("TOPPADDING",  (0,0),(-1,-1),  3),
                ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ]))
            story.append(tr_rec)
            story.append(Spacer(1, 0.4*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6))

        # ── Resumen del sistema ───────────────────────────────────────────────
        story.append(Paragraph("📊  RESUMEN EJECUTIVO DEL SISTEMA", sec_st))

        _db_arr = _inv_p.get("desglose_arranque", [])
        _arr_desc = ""
        if _db_arr:
            _arr_desc = " | ".join(f"{d['equipo']}(×{d['factor_arranque']})" for d in _db_arr)
        resumen_data = [
            ["MÓDULO","PARÁMETRO","VALOR"],
            ["⚡ Consumo", "Consumo base (Wh/día)",         f"{consumo_total:,.1f} Wh"],
            ["",          "Factor de seguridad 20%",         "× 1.20"],
            ["",          "Consumo con FS (Wh/día)",         f"{consumo_fs:,.1f} Wh"],
            ["🔋 Tensión DC","Tensión estándar del sistema", f"{vdc_p} V DC"],
            ["🌞 HSP",    "Hora Solar Pico",                  f"{hsp_p} h/día"],
            ["",          "Municipio",                        proyecto_info[2] or "—"],
            ["🔆 Paneles",f"Panel: {panel_row[2] if panel_row else '—'}  {pp_wp_p}Wp",
                          f"Voc={panel_row[4] if panel_row else '—'}V / Isc={isc_p}A"],
            ["",          "Potencia mínima instalada (Wp)",   f"{pot_inst_p:,.0f} Wp"],
            ["",          "Número de paneles",                 f"{n_pan_p} paneles"],
            ["",          "Potencia real instalada (Wp)",      f"{n_pan_p*pp_wp_p:,} Wp"],
            ["🔋 Baterías","① Consumo diario con FS (Wh/día)",    f"{consumo_fs:,.1f} Wh"],
            ["",          "② Días autonomía",                       f"{bats_p['dias_autonomia']:.3f} d ({bats_p['dias_autonomia']*24:.0f} h)"],
            ["",          "③ Voltaje banco",                        f"{vdc_p} V DC"],
            ["",          f"④ DoD",                                 f"{int(bats_p['dod']*100)}%"],
            ["",          f"⑤ Eficiencia sistema (η)",              f"{int(bats_p['eficiencia']*100)}%"],
            ["",          "E_total = consumo × días ÷ η",           f"{bats_p['e_total_wh']:,.1f} Wh"],
            ["",          "Ah = E_total ÷ (V × DoD)",               f"{bats_p['ah_dod']:,.1f} Ah"],
            ["",          f"Nº baterías {int(bats_p['cap_bat_ah'])} Ah", f"{bats_p['num_baterias']} u."],
            ["",          "Cap. real instalada",                    f"{bats_p['cap_real_ah']:,} Ah @ {vdc_p}V"],
            ["",          "Energía bruta",                          f"{bats_p['energia_kwh']:.2f} kWh"],
            ["",          "✅ Energía utilizable",                  f"{bats_p['energia_util_kwh']:.2f} kWh"],
            ["",          "Autonomía real",                         f"{bats_p['autonomia_real_h']:.1f} h ({bats_p['autonomia_real_d']:.2f} d)"],
            ["🎛 MPPT",   f"Isc ({isc_p}A) × {n_pan_p} paneles",
                          f"{corr_p:.1f} A"],
            ["",          "Controlador recomendado",           mppt_p],
            ["🔌 Inversor","Potencia instalada total",         f"{_inv_p['pot_instalada']:,.0f} W"],
            ["",          f"Demanda simultánea (×FS {int(_inv_p['fs']*100)}%)",
                                                               f"{_inv_p['pot_simultanea']:,.0f} W"],
            ["",          f"Pico arranque motores{': '+_arr_desc if _arr_desc else ''}",
                                                               f"{_inv_p['pot_arranque']:,.0f} W"],
            ["",          "Potencia requerida (sim+arranque)", f"{_inv_p['pot_requerida']:,.0f} W"],
            ["",          f"Pot. mínima (×FM {int(_inv_p['fm']*100)}%)",
                                                               f"{_inv_p['pot_inv_minima']:,.0f} W"],
            ["",          "✅ INVERSOR RECOMENDADO (comercial)",
                                                               f"{inv_kw_p:.1f} kW / {_inv_p['inv_w']:,.0f} W"],
            ["",          f"Corriente DC @ {vdc_p}V",          f"{_inv_p['corr_dc']:.1f} A"],
        ]

        rw = [1.8*cm, 8*cm, 8*cm]
        tr = Table(resumen_data, colWidths=rw, repeatRows=1)
        nr = len(resumen_data)
        tr.setStyle(TableStyle([
            ("BACKGROUND",  (0,0),  (-1,0),  SOL),
            ("TEXTCOLOR",   (0,0),  (-1,0),  DARK),
            ("FONTNAME",    (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0,0),  (-1,0),  9),
            ("ALIGN",       (0,0),  (-1,0),  "CENTER"),
            *[("BACKGROUND",(0,i),(-1,i), CARD if i%2==0 else CARD2) for i in range(1,nr)],
            ("TEXTCOLOR",   (0,1),  (0,-1),  SOL),
            ("TEXTCOLOR",   (1,1),  (1,-1),  TEXT2),
            ("TEXTCOLOR",   (2,1),  (2,-1),  MONO),
            ("FONTNAME",    (0,1),  (-1,-1), "Helvetica"),
            ("FONTNAME",    (2,1),  (2,-1),  "Helvetica-Bold"),
            ("FONTSIZE",    (0,1),  (-1,-1), 9),
            ("ALIGN",       (0,1),  (-1,-1), "LEFT"),
            ("ALIGN",       (2,1),  (2,-1),  "CENTER"),
            ("VALIGN",      (0,0),  (-1,-1), "MIDDLE"),
            ("GRID",        (0,0),  (-1,-1), 0.4, BORDER),
            ("TOPPADDING",  (0,0),  (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1),  4),
        ]))
        story.append(tr)

        # ── Sección: Protecciones y equipos ──────────────────────────────────
        if panel_row:
            isc_pdf  = float(panel_row[5]) if panel_row[5] else 14.0
            voc_pdf  = float(panel_row[4]) if panel_row[4] else 49.9
            fus_pdf  = isc_pdf * 1.25
            fus_std_pdf = next((f for f in [10,15,20,25,30,40,50] if f >= fus_pdf), 50)
            corr_ac_pdf = pot_inv_fs / 220 if pot_inv_fs > 0 else 20
            brk_ac_pdf  = corr_ac_pdf * 1.25
            brk_std_pdf = next((f for f in [16,20,25,32,40,50,63] if f >= brk_ac_pdf), 63)

            story.append(Spacer(1, 0.4*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6))
            story.append(Paragraph("🛡  PROTECCIONES ELÉCTRICAS", sec_st))
            prot_data = [
                ["COMPONENTE", "CALCULO", "SELECCION"],
                ["Fusible string DC", f"Isc {isc_pdf}A x 1.25 = {fus_pdf:.1f}A", f"{fus_std_pdf}A DC"],
                ["DPS DC", "Tipo II - 1000 VDC", "40 kA"],
                ["Breaker AC 2P", f"P/220V x 1.25 = {brk_ac_pdf:.1f}A", f"{brk_std_pdf}A AC"],
                ["DPS AC", "Tipo II - 275V", "40 kA"],
                ["Puesta a tierra", "Sistema TT", "Verde/Amarillo"],
            ]
            t_prot = Table(prot_data, colWidths=[5*cm, 8*cm, 5*cm], repeatRows=1)
            t_prot.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), SOL),
                ("TEXTCOLOR",   (0,0), (-1,0), DARK),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 8),
                ("ALIGN",       (0,0), (-1,0), "CENTER"),
                *[("BACKGROUND",(0,i),(-1,i), CARD if i%2==0 else CARD2) for i in range(1, len(prot_data))],
                ("TEXTCOLOR",   (0,1), (-1,-1), TEXT),
                ("TEXTCOLOR",   (2,1), (2,-1),  colors.HexColor("#00E676")),
                ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE",    (0,1), (-1,-1), 8),
                ("ALIGN",       (0,1), (-1,-1), "CENTER"),
                ("GRID",        (0,0), (-1,-1), 0.4, BORDER),
                ("TOPPADDING",  (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ]))
            story.append(t_prot)

            # Catálogo de equipos recomendados
            story.append(Spacer(1, 0.4*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER, spaceAfter=6))
            story.append(Paragraph("🏪  EQUIPOS RECOMENDADOS (REFERENCIA)", sec_st))
            eq_data = [
                ["TIPO", "FABRICANTE", "MODELO", "TIPO"],
                ["Panel Solar",   "JinkoSolar",    "Tiger Neo 550Wp",   "Monocristalino PERC"],
                ["Panel Solar",   "Canadian Solar", "HiKu6 550Wp",      "Monocristalino"],
                ["Inversor OFF",  "Victron Energy", "MultiPlus II 5kVA", "OFFGRID/HIBRIDO"],
                ["Inversor HYB",  "Deye",           "SUN-5K-SG04LP1",   "HIBRIDO"],
                ["Inversor ON",   "Growatt",        "MOD 5000TL3-X",    "ONGRID"],
                ["Bateria LiFe",  "Pylontech",      "US5000 4.8kWh",    "LiFePO4 48V"],
                ["Bateria LiFe",  "BYD",            "Battery Box 5.1kWh","LiFePO4 51.2V"],
            ]
            t_eq = Table(eq_data, colWidths=[3.5*cm, 4*cm, 5.5*cm, 4.5*cm], repeatRows=1)
            t_eq.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), SOL),
                ("TEXTCOLOR",   (0,0), (-1,0), DARK),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 8),
                ("ALIGN",       (0,0), (-1,0), "CENTER"),
                *[("BACKGROUND",(0,i),(-1,i), CARD if i%2==0 else CARD2) for i in range(1, len(eq_data))],
                ("TEXTCOLOR",   (0,1), (-1,-1), TEXT),
                ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
                ("FONTSIZE",    (0,1), (-1,-1), 8),
                ("ALIGN",       (0,1), (-1,-1), "LEFT"),
                ("GRID",        (0,0), (-1,-1), 0.4, BORDER),
                ("TOPPADDING",  (0,0), (-1,-1), 3),
                ("BOTTOMPADDING",(0,0),(-1,-1), 3),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ]))
            story.append(t_eq)

    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    story.append(Paragraph(
        f"<font color='#2A3A55'>SOLARCALC PRO · Dimensionamiento Fotovoltaico Aislado · {datetime.now().year}</font>",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, alignment=TA_CENTER, spaceBefore=4)))

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
# ─── CONTROL DE ACCESO ──────────────────────────────────────────────────────
if not usuario_activo():
    mostrar_login()   # muestra login y llama st.stop() hasta que inicie sesión

_u = usuario_activo()

with st.sidebar:
    # ── Logo ─────────────────────────────────────────────────────────────────
    st.markdown("""
    <div style='text-align:center; padding:0.8rem 0 0.5rem;'>
        <div style='font-family:Rajdhani,sans-serif; font-size:2rem; color:#FFB300;
                    font-weight:700; letter-spacing:3px; line-height:1.1;'>
            ☀ SOLAR<br>CALC
        </div>
        <div style='font-size:0.65rem; color:#8A9BBD; letter-spacing:3px;
                    margin-top:0.3rem; text-transform:uppercase;'>
            Dimensionamiento Fotovoltaico
        </div>
    </div>
    <hr style='border-color:#2A3A55; margin:0.6rem 0;'>
    """, unsafe_allow_html=True)

    # ── Usuario activo + logout ───────────────────────────────────────────────
    mostrar_usuario_sidebar()

    # ── 0. Selector de TIPO DE SISTEMA ───────────────────────────────────────
    st.markdown("""
    <div style='font-size:0.65rem;color:#8A9BBD;letter-spacing:2px;
     text-transform:uppercase;margin-bottom:0.5rem;'>⚙ TIPO DE SISTEMA</div>
    """, unsafe_allow_html=True)

    if "tipo_sistema" not in st.session_state:
        st.session_state["tipo_sistema"] = "OFF-GRID"

    _SISTEMAS = [
        ("OFF-GRID", "🔋", "Aislado",        "#FFB300", "rgba(255,179,0,0.15)",  "dimensionamiento"),
        ("ON-GRID",  "🔌", "Interconectado", "#FF6B35", "rgba(255,107,53,0.15)","ongrid"),
        ("HIBRIDO",  "⚡", "ON + Baterías",  "#F59E0B", "rgba(245,158,11,0.15)","hibrido"),
    ]
    for _tipo, _icon, _sub, _col, _bg, _mod in _SISTEMAS:
        _activo = st.session_state["tipo_sistema"] == _tipo
        _bord   = f"1px solid {_col}" if _activo else "1px solid #2A3A55"
        _bg_val = _bg if _activo else "#0F1525"
        _fw     = "700" if _activo else "400"
        _tcol   = _col if _activo else "#8A9BBD"
        st.markdown(
            f"<div style='background:{_bg_val};border:{_bord};border-radius:8px;"
            f"padding:0.4rem 0.5rem;text-align:center;font-family:Rajdhani,sans-serif;"
            f"font-size:0.83rem;color:{_tcol};font-weight:{_fw};margin-bottom:3px;'>"
            f"{_icon} {_tipo}<br>"
            f"<span style='font-size:0.62rem;font-weight:400;color:#8A9BBD;'>{_sub}</span></div>",
            unsafe_allow_html=True)
        if _activo:
            st.markdown(
                f"<div style='text-align:center;font-size:0.63rem;color:{_col};"
                f"margin-bottom:3px;letter-spacing:1px;'>▶ ACTIVO</div>",
                unsafe_allow_html=True)
        else:
            if st.button(f"Usar {_tipo}", key=f"btn_{_tipo.lower().replace('-','_')}",
                         use_container_width=True):
                st.session_state["tipo_sistema"]  = _tipo
                st.session_state["modulo_activo"] = _mod
                if usuario_activo():
                    registrar_auditoria(_u["id"], _u["username"],
                                        "CAMBIO_SISTEMA", f"→ {_tipo}", "app")
                st.rerun()

    tipo_sistema_activo = st.session_state["tipo_sistema"]

    st.markdown("<hr style='border-color:#2A3A55;margin:0.7rem 0;'>", unsafe_allow_html=True)

    # ── 1. Selector de proyecto activo ────────────────────────────────────────
    st.markdown("""
    <div style='font-size:0.65rem;color:#8A9BBD;letter-spacing:2px;
     text-transform:uppercase;margin-bottom:0.4rem;'>📁 PROYECTO ACTIVO</div>
    """, unsafe_allow_html=True)

    conn = get_conn()
    proyectos_df = pd.read_sql("SELECT id, nombre FROM proyectos ORDER BY id DESC", conn)
    conn.close()

    opciones = ["── Seleccionar proyecto ──"] + \
               [f"{r['id']} | {r['nombre']}" for _, r in proyectos_df.iterrows()]
    sel = st.selectbox("Proyecto:", opciones, label_visibility="collapsed", key="sel_proyecto")

    proyecto_id = None
    if sel != "── Seleccionar proyecto ──":
        proyecto_id = int(sel.split(" | ")[0])

    # Mostrar datos del proyecto activo
    if proyecto_id:
        conn = get_conn()
        p_sb = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
        conn.close()
        if p_sb:
            hsp_sb  = p_sb[4] if p_sb[4] else "—"
            vdc_sb  = p_sb[3] if p_sb[3] else "—"
            mun_sb  = p_sb[2] if p_sb[2] else "Sin municipio"
            mod_sb  = st.session_state.get("modulo_activo","dimensionamiento")
            st.markdown(f"""
            <div style='background:#0F1525;border:1px solid #2A3A55;border-radius:8px;
                        padding:0.7rem 0.9rem;margin-bottom:0.3rem;'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;
                            font-weight:700;font-size:0.95rem;'>
                    #{p_sb[0]} {p_sb[1]}</div>
                <div style='color:#8A9BBD;font-size:0.72rem;margin-top:0.2rem;'>
                    📍 {mun_sb}</div>
                <div style='display:flex;gap:0.8rem;margin-top:0.3rem;'>
                    <span style='font-size:0.72rem;color:#00BCD4;'>
                        ☀ HSP: {hsp_sb} h</span>
                    <span style='font-size:0.72rem;color:#FFD54F;'>
                        ⚡ VDC: {vdc_sb} V</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style='background:#0F1525;border:1px dashed #2A3A55;border-radius:8px;
                    padding:0.7rem;text-align:center;margin-bottom:0.3rem;'>
            <div style='color:#2A3A55;font-size:0.78rem;'>Sin proyecto seleccionado</div>
        </div>
        """, unsafe_allow_html=True)

    # ── 2. Crear nuevo proyecto ───────────────────────────────────────────────
    with st.expander("✦ Crear nuevo proyecto", expanded=False):
        nuevo_nombre    = st.text_input("Nombre",    placeholder="Ej: Finca La Esperanza", key="sb_nom")
        nuevo_municipio = st.text_input("Municipio", placeholder="Ej: Medellín, Colombia", key="sb_mun")
        if st.button("✦ Crear", use_container_width=True, key="sb_crear"):
            if nuevo_nombre.strip():
                conn = get_conn()
                conn.execute("INSERT INTO proyectos(nombre, municipio) VALUES(?,?)",
                             (nuevo_nombre.strip(), nuevo_municipio.strip()))
                conn.commit()
                nuevo_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                conn.close()
                st.success(f"✓ Proyecto #{nuevo_id} creado")
                st.rerun()
            else:
                st.error("Ingresa un nombre")

    # ── 2b. Eliminar proyecto ─────────────────────────────────────────────────
    if proyecto_id:
        with st.expander("🗑 Eliminar proyecto", expanded=False):
            conn = get_conn()
            p_del = conn.execute(
                "SELECT nombre FROM proyectos WHERE id=?", (proyecto_id,)
            ).fetchone()
            conn.close()
            nombre_del = p_del[0] if p_del else "—"

            st.markdown(f"""
            <div class='warn-box' style='margin-bottom:0.6rem;'>
                ⚠ Esta acción eliminará <b>permanentemente</b> el proyecto
                <b style='color:#FFB300;'>#{proyecto_id} — {nombre_del}</b>
                y todos sus datos asociados (cargas, paneles, baterías, recibos,
                resultados). <b>No se puede deshacer.</b>
            </div>
            """, unsafe_allow_html=True)

            confirmar_nombre = st.text_input(
                "Escribe el nombre del proyecto para confirmar:",
                placeholder=nombre_del,
                key="del_confirm_nombre",
            )

            if st.button("🗑 Eliminar definitivamente", use_container_width=True,
                         key="sb_eliminar"):
                if confirmar_nombre.strip() == nombre_del.strip():
                    conn = get_conn()
                    conn.execute("DELETE FROM cargas     WHERE proyecto_id=?", (proyecto_id,))
                    conn.execute("DELETE FROM paneles    WHERE proyecto_id=?", (proyecto_id,))
                    conn.execute("DELETE FROM resultados WHERE proyecto_id=?", (proyecto_id,))
                    conn.execute("DELETE FROM recibos    WHERE proyecto_id=?", (proyecto_id,))
                    conn.execute("DELETE FROM proyectos  WHERE id=?",          (proyecto_id,))
                    conn.commit()
                    conn.close()
                    if usuario_activo():
                        _u2 = usuario_activo()
                        registrar_auditoria(
                            _u2["id"], _u2["username"],
                            "ELIMINAR_PROYECTO",
                            f"Proyecto #{proyecto_id} '{nombre_del}' eliminado",
                            "app",
                        )
                    st.success(f"✓ Proyecto #{proyecto_id} '{nombre_del}' eliminado")
                    # Limpiar selección y volver al inicio
                    for _k in ("sel_proyecto", "del_confirm_nombre"):
                        if _k in st.session_state:
                            del st.session_state[_k]
                    st.rerun()
                else:
                    st.error("El nombre no coincide. Escríbelo exactamente para confirmar.")

    st.markdown("<hr style='border-color:#2A3A55;margin:0.6rem 0;'>", unsafe_allow_html=True)

    # ── 3. Navegación de módulos ──────────────────────────────────────────────
    st.markdown("""
    <div style='font-size:0.65rem;color:#8A9BBD;letter-spacing:2px;
     text-transform:uppercase;margin-bottom:0.4rem;'>🧭 MÓDULOS</div>
    """, unsafe_allow_html=True)

    if "modulo_activo" not in st.session_state:
        st.session_state["modulo_activo"] = "dimensionamiento"

    # Grupos de módulos — dependen del tipo de sistema activo
    _CATALOGOS = [
        ("🏪  Proveedores", "proveedores"),
        ("🔩  Materiales",  "materiales"),
        ("🔧  Equipos",     "equipos"),
        ("👷  Personal",    "personal"),
    ]
    if tipo_sistema_activo == "ON-GRID":
        GRUPOS_MODULOS = [
            ("PROYECTO ON-GRID", [
                ("🔌  Dimensionamiento ON-GRID", "ongrid"),
                ("🔬  Simulador ON-GRID",        "simulador"),
                ("💰  Cotización Comercial",     "cotizacion"),
                ("⏚   Puesta a Tierra",          "tierra"),
            ]),
            ("CATÁLOGOS", _CATALOGOS),
        ] + ([("ADMINISTRACIÓN", [
                ("🔐  Usuarios",  "usuarios"),
                ("👤  Mi perfil", "perfil"),
              ])] if tiene_permiso("ver_usuarios") else [("MI CUENTA", [("👤  Mi perfil","perfil")])])
    elif tipo_sistema_activo == "HIBRIDO":
        GRUPOS_MODULOS = [
            ("PROYECTO HÍBRIDO", [
                ("⚡  Dimensionamiento HÍBRIDO", "hibrido"),
                ("🔬  Simulador HÍBRIDO",        "simulador"),
                ("💰  Cotización Comercial",     "cotizacion"),
                ("⏚   Puesta a Tierra",          "tierra"),
            ]),
            ("CATÁLOGOS", _CATALOGOS),
        ] + ([("ADMINISTRACIÓN", [
                ("🔐  Usuarios",  "usuarios"),
                ("👤  Mi perfil", "perfil"),
              ])] if tiene_permiso("ver_usuarios") else [("MI CUENTA", [("👤  Mi perfil","perfil")])])
    else:
        GRUPOS_MODULOS = [
            ("PROYECTO OFF-GRID", [
                ("🔋  Dimensionamiento OFF-GRID", "dimensionamiento"),
                ("🔬  Simulador OFF-GRID",        "simulador"),
                ("💰  Cotización Comercial",      "cotizacion"),
                ("⏚   Puesta a Tierra",           "tierra"),
            ]),
            ("CATÁLOGOS", _CATALOGOS),
        ] + ([("ADMINISTRACIÓN", [
                ("🔐  Usuarios",  "usuarios"),
                ("👤  Mi perfil", "perfil"),
              ])] if tiene_permiso("ver_usuarios") else [("MI CUENTA", [("👤  Mi perfil","perfil")])])

    for grupo_label, modulos in GRUPOS_MODULOS:
        st.markdown(f"""
        <div style='font-size:0.6rem;color:#2A3A55;letter-spacing:2px;
         text-transform:uppercase;margin:0.5rem 0 0.2rem;padding-left:0.2rem;
         border-left:2px solid #2A3A55;'>
            {grupo_label}
        </div>""", unsafe_allow_html=True)
        for label, key in modulos:
            activo = st.session_state["modulo_activo"] == key
            # Color badge for active module
            if activo:
                st.markdown(f"""
                <div style='background:rgba(255,179,0,0.12);border:1px solid rgba(255,179,0,0.45);
                            border-radius:6px;padding:0.35rem 0.7rem;margin-bottom:3px;
                            font-family:Barlow,sans-serif;font-size:0.85rem;color:#FFB300;
                            font-weight:600;cursor:default;'>
                    {label}
                </div>""", unsafe_allow_html=True)
            if st.button(
                label if not activo else f"› {label.strip()}",
                use_container_width=True,
                key=f"nav_{key}",
                help=f"Ir a {label.strip()}",
                disabled=activo,
            ):
                st.session_state["modulo_activo"] = key
                st.rerun()

    st.markdown("<hr style='border-color:#2A3A55;margin:0.6rem 0;'>", unsafe_allow_html=True)

    # ── 4. Footer del sidebar ─────────────────────────────────────────────────
    _col_ts = {"ON-GRID":"#FF6B35","HIBRIDO":"#F59E0B"}.get(tipo_sistema_activo,"#FFB300")
    st.markdown(f"""
    <div style='text-align:center;padding:0.3rem 0;'>
        <div style='font-size:0.62rem;color:#2A3A55;letter-spacing:1px;'>
            SolarCalc Pro v1.2<br>
            SQLite3 + Streamlit<br>
            Sistema: <span style='color:{_col_ts};'>
                {tipo_sistema_activo}</span><br>
            Módulo: <span style='color:#8A9BBD;'>
                {st.session_state.get("modulo_activo","—")}</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ─── MAIN CONTENT ────────────────────────────────────────────────────────────
modulo_activo = st.session_state.get("modulo_activo", "dimensionamiento")
tipo_sistema_activo = st.session_state.get("tipo_sistema", "OFF-GRID")

# ── Guard: project required for most modules ─────────────────────────────────
if not proyecto_id and modulo_activo not in ("materiales", "equipos", "personal"):
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>☀ SOLARCALC PRO</div>
        <div class='hero-sub'>Sistema de Dimensionamiento de Paneles Fotovoltaicos</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class='sol-card' style='text-align:center; padding:3rem;'>
        <div style='font-size:3rem;'>⚡</div>
        <div style='font-family:Rajdhani,sans-serif; font-size:1.5rem; color:#FFB300; margin-top:1rem;'>
            SELECCIONA O CREA UN PROYECTO</div>
        <div style='color:#8A9BBD; margin-top:0.5rem;'>Usa el panel lateral para comenzar</div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

if modulo_activo == "proveedores":
    from modulo_proveedores import mostrar_proveedores
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>🏪 PROVEEDORES</div>
        <div class='hero-sub'>CONSULTA DE PRECIOS · TIENDAS EN LÍNEA</div>
    </div>""", unsafe_allow_html=True)
    mostrar_proveedores()
    st.stop()

# ── Catálogos (sin proyecto) ─────────────────────────────────────────────────
if modulo_activo == "materiales":
    from modulo_equipos import mostrar_equipos
    from modulo_materiales import mostrar_materiales
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>🔩 MATERIALES</div>
        <div class='hero-sub'>CATÁLOGO FV / RETIE</div>
    </div>""", unsafe_allow_html=True)
    mostrar_materiales()
    st.stop()

if modulo_activo == "equipos":
    from modulo_equipos import mostrar_equipos
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>🔧 EQUIPOS Y HERRAMIENTAS</div>
        <div class='hero-sub'>CATÁLOGO DE RECURSOS</div>
    </div>""", unsafe_allow_html=True)
    mostrar_equipos()
    st.stop()

if modulo_activo == "personal":
    from modulo_equipos import mostrar_personal
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>👷 PERSONAL IDÓNEO</div>
        <div class='hero-sub'>PERFILES FV / RETIE</div>
    </div>""", unsafe_allow_html=True)
    mostrar_personal()
    st.stop()

# ── Módulos que requieren proyecto ────────────────────────────────────────────
if modulo_activo == "simulador":
    from modulo_simulador import mostrar_simulador
    mostrar_simulador(proyecto_id, st.session_state)
    st.stop()

if modulo_activo in ("cotizacion", "presupuesto"):
    from modulo_cotizacion import mostrar_cotizacion
    mostrar_cotizacion(proyecto_id, st.session_state)
    st.stop()

# ── Módulo PUESTA A TIERRA ────────────────────────────────────────────────────
if modulo_activo == "tierra":
    from modulo_tierra import mostrar_tierra
    mostrar_tierra(proyecto_id, st.session_state)
    st.stop()

# ── Módulo ON-GRID ────────────────────────────────────────────────────────────
if modulo_activo == "ongrid" or tipo_sistema_activo == "ON-GRID":
    from modulo_ongrid import mostrar_ongrid
    mostrar_ongrid(proyecto_id, st.session_state)
    st.stop()

# ── Módulo HÍBRIDO ────────────────────────────────────────────────────────────
if modulo_activo == "hibrido" or tipo_sistema_activo == "HIBRIDO":
    from modulo_hibrido import mostrar_hibrido
    mostrar_hibrido(proyecto_id, st.session_state)
    st.stop()

# ── Gestión de usuarios ───────────────────────────────────────────────────────
if modulo_activo == "usuarios":
    mostrar_gestion_usuarios()
    st.stop()

# ── Perfil / cambio de contraseña ─────────────────────────────────────────────
if modulo_activo == "perfil":
    st.markdown("""
    <div class='hero-header'>
        <div class='hero-title'>👤 MI PERFIL</div>
        <div class='hero-sub'>CONFIGURACIÓN DE CUENTA · SEGURIDAD</div>
    </div>
    """, unsafe_allow_html=True)
    mostrar_cambio_password()
    st.stop()

# ── Módulo principal: Dimensionamiento OFF-GRID ───────────────────────────────
st.markdown("""
<div class='hero-header'>
    <div class='hero-title'>☀ SOLARCALC PRO
        <span style='display:inline-block;background:#FFB300;color:#0A0E1A;
            font-family:Rajdhani,sans-serif;font-weight:700;font-size:0.75rem;
            padding:2px 10px;border-radius:20px;letter-spacing:1px;margin-left:8px;
            vertical-align:middle;'>OFF-GRID</span>
    </div>
    <div class='hero-sub'>DIMENSIONAMIENTO — SISTEMA FOTOVOLTAICO AISLADO (OFF-GRID)</div>
</div>
""", unsafe_allow_html=True)

# Tabs principales
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12 = st.tabs([
    "⚡ 1 · Cargas",
    "🧾 2 · Recibo Luz",
    "🔋 3 · Tensión DC",
    "🌞 4 · Hora Solar",
    "🔆 5 · Panel Solar",
    "📊 6 · Potencia",
    "🔋 7 · Baterías",
    "🎛 8 · Controlador",
    "🔲 9 · Plano Paneles",
    "📐 10 · Plano General",
    "💹 11 · Económico",
    "🔌 12 · Cableado",
])

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — ANÁLISIS DE CARGAS
# ════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>1</span> ANÁLISIS DE CARGA ELÉCTRICA</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        Consumo día (Wh) = Cantidad × Potencia (W) × Horas uso/día<br>
        Potencia Inversor = Potencia × Factor (×4 para motorizados) | Total con 20% factor seguridad
    </div>
    """, unsafe_allow_html=True)

    # ── CATÁLOGO DE ELECTRODOMÉSTICOS ─────────────────────────────────────
    CATALOGO = {
        "💡 Iluminación": [
            ("Bombillo LED 5W",          5,    False),
            ("Bombillo LED 9W",          9,    False),
            ("Bombillo LED 15W",        15,    False),
            ("Tubo LED 18W",            18,    False),
            ("Tubo fluorescente 36W",   36,    False),
            ("Lámpara halógena 50W",    50,    False),
            ("Reflector LED 30W",       30,    False),
            ("Reflector LED 50W",       50,    False),
            ("Luminaria exterior 80W",  80,    False),
            ("Downlight LED 12W",       12,    False),
        ],
        "📺 Entretenimiento y Comunicación": [
            ("TV LED 32\"",             60,    False),
            ("TV LED 43\"",            100,    False),
            ("TV LED 55\"",            150,    False),
            ("TV LED 65\"",            200,    False),
            ("Decodificador / STB",     25,    False),
            ("Equipo de sonido",       100,    False),
            ("Subwoofer activo",       150,    False),
            ("Computador portátil",     65,    False),
            ("Computador de escritorio",250,   False),
            ("Monitor PC 24\"",         30,    False),
            ("Tablet / iPad",           10,    False),
            ("Router WiFi",             12,    False),
            ("Impresora inkjet",        15,    False),
            ("Impresora láser",        400,    False),
            ("Cámara de vigilancia",    10,    False),
            ("DVR 4 canales",           20,    False),
        ],
        "🍳 Cocina": [
            ("Nevera 200L",            150,    True),
            ("Nevera 300L",            200,    True),
            ("Nevera 400L",            250,    True),
            ("Nevera No Frost 400L",   300,    True),
            ("Congelador 200L",        150,    True),
            ("Microondas 700W",        700,    False),
            ("Microondas 1000W",      1000,    False),
            ("Licuadora doméstica",    300,    True),
            ("Licuadora industrial",   750,    True),
            ("Batidora de mano",       300,    False),
            ("Procesadora de alimentos",400,   True),
            ("Cafetera eléctrica",     900,    False),
            ("Hervidor eléctrico",    1500,    False),
            ("Tostadora 2 panes",      800,    False),
            ("Horno tostador",        1200,    False),
            ("Horno eléctrico 60L",   2000,    False),
            ("Plancha de cocina",     1500,    False),
            ("Extractor de jugos",     400,    True),
            ("Freidora de aire 3.5L",  900,    False),
            ("Freidora de aire 5.5L", 1500,    False),
        ],
        "🧺 Lavado y Limpieza": [
            ("Lavadora 7 kg",          500,    True),
            ("Lavadora 10 kg",         700,    True),
            ("Lavadora 14 kg",        1000,    True),
            ("Secadora ropa 7 kg",    2000,    False),
            ("Lavadora / Secadora",   2200,    True),
            ("Aspiradora doméstica",   800,    True),
            ("Aspiradora industrial", 1200,    True),
            ("Plancha de ropa 1000W", 1000,    False),
            ("Plancha de ropa 2000W", 2000,    False),
            ("Vaporizadora de ropa",  1500,    False),
        ],
        "❄ Climatización": [
            ("Ventilador de mesa",      45,    True),
            ("Ventilador de pie",       60,    True),
            ("Ventilador de techo",     75,    True),
            ("Aire acondicionado 9000 BTU",  900,  True),
            ("Aire acondicionado 12000 BTU",1200,  True),
            ("Aire acondicionado 18000 BTU",1800,  True),
            ("Aire acondicionado 24000 BTU",2400,  True),
            ("Mini Split 9000 BTU",     900,   True),
            ("Mini Split 12000 BTU",   1200,   True),
            ("Calefactor eléctrico",   1500,   False),
            ("Calefactor cerámico",    2000,   False),
        ],
        "🚿 Agua y Sanitarios": [
            ("Bomba agua 0.5 HP",       373,   True),
            ("Bomba agua 1 HP",         746,   True),
            ("Bomba agua 2 HP",        1492,   True),
            ("Bomba sumergible 0.5 HP", 373,   True),
            ("Calentador eléctrico paso",3500,  False),
            ("Calentador eléctrico tanque 40L",1500,False),
            ("Jacuzzi / Spa",          1500,   True),
            ("Bomba piscina 1 HP",      746,   True),
        ],
        "🔧 Taller y Herramientas": [
            ("Taladro eléctrico",       700,   True),
            ("Amoladora 4.5\"",         800,   True),
            ("Sierra circular",        1400,   True),
            ("Compresor 1 HP",          746,   True),
            ("Compresor 2 HP",         1492,   True),
            ("Soldadora eléctrica",    4000,   False),
            ("Pulidora industrial",    1000,   True),
        ],
        "🌾 Agropecuario": [
            ("Motor bomba riego 1 HP",  746,   True),
            ("Motor bomba riego 2 HP", 1492,   True),
            ("Incubadora 60 huevos",    60,    False),
            ("Incubadora 200 huevos",  150,    False),
            ("Picadora de pasto",      750,    True),
            ("Molino de granos",      1500,    True),
            ("Ordeñadora 1 unidad",    370,    True),
            ("Cerca eléctrica",         10,    False),
            ("Cámara frigorífica 2m³", 500,    True),
        ],
        "🏥 Salud y Cuidado Personal": [
            ("Secador de cabello 1200W",1200,  False),
            ("Secador de cabello 1800W",1800,  False),
            ("Plancha de cabello",      120,   False),
            ("Afeitadora eléctrica",     15,   False),
            ("CPAP",                     30,   False),
            ("Concentrador de oxígeno", 300,   False),
            ("Nebulizador",              50,   False),
        ],
        "🔋 Carga de Dispositivos": [
            ("Cargador celular",         10,   False),
            ("Cargador laptop 65W",      65,   False),
            ("Cargador laptop 90W",      90,   False),
            ("Cargador tablet",          18,   False),
            ("Cargador moto eléctrica", 500,   False),
            ("Cargador bici eléctrica", 200,   False),
        ],
    }

    # Estado de selección múltiple del catálogo
    if "catalog_sel" not in st.session_state:
        st.session_state.catalog_sel = {}

    # ── PANEL CATÁLOGO ────────────────────────────────────────────────────
    with st.expander("📋  CATÁLOGO DE ELECTRODOMÉSTICOS — Potencias Estándar", expanded=True):
        st.markdown("""
        <div class='info-note' style='margin-bottom:1rem;'>
            Selecciona los equipos del catálogo, ajusta cantidad y horas de uso, luego haz clic en
            <b>➕ Agregar Seleccionados</b>. También puedes agregar equipos personalizados en el formulario de abajo.
        </div>
        """, unsafe_allow_html=True)

        # Filtro de búsqueda
        buscar = st.text_input("🔍 Buscar equipo en catálogo...", placeholder="Ej: nevera, bombillo, bomba...", key="cat_buscar")

        for categoria, equipos in CATALOGO.items():
            # Filtrar si hay búsqueda
            if buscar:
                equipos_filtrados = [e for e in equipos if buscar.lower() in e[0].lower()]
                if not equipos_filtrados:
                    continue
            else:
                equipos_filtrados = equipos

            st.markdown(f"<div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; font-size:0.95rem; margin:0.8rem 0 0.4rem; letter-spacing:1px;'>{categoria}</div>", unsafe_allow_html=True)

            # Filas de 2 columnas por categoría
            for i in range(0, len(equipos_filtrados), 2):
                cols_cat = st.columns(2)
                for j, col_c in enumerate(cols_cat):
                    if i + j >= len(equipos_filtrados):
                        break
                    nombre_e, pot_e, motor_e = equipos_filtrados[i + j]
                    key_e = f"cat_{nombre_e}"
                    with col_c:
                        c1, c2, c3, c4 = st.columns([3, 1.2, 1.2, 0.7])
                        with c1:
                            seleccionado = st.checkbox(
                                f"{'⚡ ' if motor_e else ''}{nombre_e}",
                                key=key_e,
                                help=f"{'Equipo motorizado (×4 inversor)' if motor_e else 'Equipo estándar'}"
                            )
                        with c2:
                            cant_e = st.number_input("Cant", min_value=1, max_value=50, value=1,
                                                      key=f"cant_{nombre_e}", label_visibility="collapsed")
                        with c3:
                            pot_custom = st.number_input("W", min_value=1, max_value=10000, value=int(pot_e),
                                                          key=f"pot_{nombre_e}", label_visibility="collapsed",
                                                          help="Potencia estándar (editable)")
                        with c4:
                            hrs_e = st.number_input("h", min_value=0.5, max_value=24.0, value=4.0, step=0.5,
                                                     key=f"hrs_{nombre_e}", label_visibility="collapsed",
                                                     help="Horas/día de uso")
                        if seleccionado:
                            st.session_state.catalog_sel[nombre_e] = {
                                "potencia": pot_custom,
                                "cantidad": cant_e,
                                "horas": hrs_e,
                                "motor": motor_e
                            }
                        elif nombre_e in st.session_state.catalog_sel:
                            del st.session_state.catalog_sel[nombre_e]

        # Resumen seleccionados
        n_sel = len(st.session_state.catalog_sel)
        if n_sel > 0:
            # ── Cálculo detallado por equipo ──────────────────────────────
            filas_prev = []
            for nombre_v, datos_v in st.session_state.catalog_sel.items():
                cant_v   = datos_v["cantidad"]
                pot_v    = datos_v["potencia"]
                hrs_v    = datos_v["horas"]
                pot_tot  = cant_v * pot_v
                consumo  = pot_tot * hrs_v          # Wh/día = W × cant × h
                filas_prev.append({
                    "Equipo":          nombre_v,
                    "Cant.":           cant_v,
                    "Potencia (W)":    pot_v,
                    "Pot. Total (W)":  pot_tot,
                    "Horas/día":       hrs_v,
                    "Consumo (Wh/día)": consumo,
                    "Motor":           "⚡" if datos_v["motor"] else "—",
                })

            df_prev = pd.DataFrame(filas_prev)
            total_preview    = df_prev["Consumo (Wh/día)"].sum()
            total_pot        = df_prev["Pot. Total (W)"].sum()
            total_preview_fs = total_preview * 1.20

            # Tabla detalle
            st.markdown(f"""
            <div style='font-family:Rajdhani,sans-serif; font-size:0.95rem; font-weight:600;
                        color:#FFB300; letter-spacing:1px; margin:1rem 0 0.4rem;'>
                ✓ {n_sel} equipo(s) seleccionado(s) — Desglose de consumo estimado
            </div>
            """, unsafe_allow_html=True)
            st.dataframe(
                df_prev.set_index("Equipo"),
                use_container_width=True,
                hide_index=False,
            )

            # Totales destacados
            st.markdown(f"""
            <div style='display:grid; grid-template-columns:repeat(3,1fr); gap:0.8rem; margin:0.8rem 0;'>
                <div style='background:#1E2A3F; border:1px solid #2A3A55; border-radius:8px;
                            padding:0.8rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.4rem; color:#FFD54F;'>{total_pot:,.0f}</div>
                    <div style='font-size:0.72rem; color:#8A9BBD; margin-top:0.2rem;
                                text-transform:uppercase; letter-spacing:1px;'>W — Pot. total instalada</div>
                </div>
                <div style='background:#1E2A3F; border:1px solid rgba(255,179,0,0.45); border-radius:8px;
                            padding:0.8rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.4rem; color:#FFB300;'>{total_preview:,.0f}</div>
                    <div style='font-size:0.72rem; color:#8A9BBD; margin-top:0.2rem;
                                text-transform:uppercase; letter-spacing:1px;'>Wh/día — Consumo base</div>
                </div>
                <div style='background:#1E2A3F; border:1px solid rgba(0,230,118,0.4); border-radius:8px;
                            padding:0.8rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.4rem; color:#00E676;'>{total_preview_fs:,.0f}</div>
                    <div style='font-size:0.72rem; color:#8A9BBD; margin-top:0.2rem;
                                text-transform:uppercase; letter-spacing:1px;'>Wh/día — Con 20% FS</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button(f"➕ Agregar {n_sel} equipo(s) seleccionado(s) al proyecto", use_container_width=True, type="primary"):
                conn = get_conn()
                # Obtener nombres ya existentes en el proyecto (case-insensitive)
                existentes = pd.read_sql(
                    "SELECT LOWER(electrodomestico) as nom FROM cargas WHERE proyecto_id=?",
                    conn, params=(proyecto_id,))["nom"].tolist()
                added, skipped = 0, []
                for nombre_add, datos_add in st.session_state.catalog_sel.items():
                    if nombre_add.lower() in existentes:
                        skipped.append(nombre_add)
                        continue
                    conn.execute("""
                        INSERT INTO cargas(proyecto_id, electrodomestico, cantidad, potencia_w, horas_dia, es_motor)
                        VALUES(?,?,?,?,?,?)
                    """, (proyecto_id, nombre_add, datos_add["cantidad"], datos_add["potencia"],
                          datos_add["horas"], 1 if datos_add["motor"] else 0))
                    added += 1
                conn.commit()
                conn.close()
                # Limpiar selección
                for k in list(st.session_state.keys()):
                    if k.startswith("cat_") or k.startswith("cant_") or k.startswith("pot_") or k.startswith("hrs_"):
                        del st.session_state[k]
                st.session_state.catalog_sel = {}
                if added:
                    st.success(f"✓ {added} equipo(s) agregados al proyecto")
                if skipped:
                    st.warning(f"⚠ {len(skipped)} equipo(s) omitido(s) por duplicado: {', '.join(skipped)}")
                st.rerun()

    st.markdown("<hr class='sep'>", unsafe_allow_html=True)

    # ── FORMULARIO MANUAL ─────────────────────────────────────────────────
    st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
    st.markdown("**✏ Agregar equipo personalizado (no está en el catálogo)**")
    col1, col2, col3, col4 = st.columns([3, 1, 1.5, 1.5])
    with col1:
        elec_nombre = st.text_input("Electrodoméstico", placeholder="Ej: Equipo especial, Motor 3HP...", key="elec_n")
    with col2:
        elec_cant = st.number_input("Cantidad", min_value=1, max_value=100, value=1, key="elec_c")
    with col3:
        elec_pot = st.number_input("Potencia (W)", min_value=1, max_value=50000, value=100, key="elec_p")
    with col4:
        elec_hrs = st.number_input("Horas/día", min_value=0.1, max_value=24.0, value=4.0, step=0.5, key="elec_h")

    motor_detected = es_motorizado(elec_nombre) if elec_nombre else False
    if motor_detected:
        st.markdown(f"""
        <div class='info-note'>⚡ <b>{elec_nombre}</b> detectado como equipo motorizado — se aplicará factor de arranque (×2 a ×5 según tipo) en el cálculo del inversor</div>
        """, unsafe_allow_html=True)

    if st.button("➕ Agregar Carga Personalizada", use_container_width=True):
        if elec_nombre.strip() and elec_pot > 0:
            conn = get_conn()
            existe = conn.execute(
                "SELECT COUNT(*) FROM cargas WHERE proyecto_id=? AND LOWER(electrodomestico)=LOWER(?)",
                (proyecto_id, elec_nombre.strip())).fetchone()[0]
            if existe:
                st.error(f"⚠ '{elec_nombre.strip()}' ya está registrado en este proyecto. Edítalo en la tabla de cargas.")
            else:
                conn.execute("""
                    INSERT INTO cargas(proyecto_id, electrodomestico, cantidad, potencia_w, horas_dia, es_motor)
                    VALUES(?,?,?,?,?,?)
                """, (proyecto_id, elec_nombre.strip(), elec_cant, elec_pot, elec_hrs,
                      1 if motor_detected else 0))
                conn.commit()
                st.success("Carga agregada ✓")
                st.rerun()
            conn.close()
        else:
            st.error("Completa nombre y potencia")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── TABLA DE CARGAS REGISTRADAS ──────────────────────────────────────────
    conn = get_conn()
    cargas = pd.read_sql("SELECT * FROM cargas WHERE proyecto_id=?", conn, params=(proyecto_id,))
    conn.close()

    if cargas.empty:
        st.markdown("""
        <div class='info-note' style='text-align:center; padding:1.5rem;'>
            No hay cargas registradas. Agrega los electrodomésticos del proyecto.
        </div>
        """, unsafe_allow_html=True)
    else:
        # Calcular columnas derivadas
        cargas["pot_total_w"]    = cargas["cantidad"] * cargas["potencia_w"]
        cargas["consumo_dia_wh"] = cargas["pot_total_w"] * cargas["horas_dia"]

        # Tabla display
        display = cargas[["id","electrodomestico","cantidad","potencia_w","pot_total_w",
                           "horas_dia","consumo_dia_wh","es_motor"]].copy()
        display.columns = ["ID","Electrodoméstico","Cant","Potencia(W)","Pot.Total(W)",
                            "Horas/día","Consumo día(Wh)","Motor"]
        display["Motor"] = display["Motor"].map({1:"⚡ Motor", 0:"—"})
        st.dataframe(display.set_index("ID"), use_container_width=True)

        # Totales
        consumo_total      = cargas["consumo_dia_wh"].sum()
        consumo_fs         = consumo_total * 1.20
        _vdc_t1 = st.session_state.get("calc_vdc", 24)
        _inv_t1 = calcular_inversor(cargas, fs=0.80, fm=1.25, vdc=_vdc_t1)

        st.markdown(f"""
        <div class='metric-grid'>
            <div class='metric-box'>
                <div class='metric-val'>{consumo_total:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>Consumo Base</div>
            </div>
            <div class='metric-box' style='border-color:rgba(255,179,0,0.5);'>
                <div class='metric-val'>{consumo_fs:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>+ 20% Factor Seguridad</div>
            </div>
            <div class='metric-box'>
                <div class='metric-val'>{_inv_t1['pot_instalada']:,.0f}</div>
                <div class='metric-unit'>W</div>
                <div class='metric-label'>Pot. Instalada Total</div>
            </div>
            <div class='metric-box' style='border-color:rgba(255,179,0,0.5);'>
                <div class='metric-val'>{_inv_t1['inv_w']:,.0f}</div>
                <div class='metric-unit'>W</div>
                <div class='metric-label'>Inversor Recomendado</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class='result-highlight'>
            <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase; letter-spacing:1px;'>
                Inversor Recomendado — Metodología técnica (FS={int(_inv_t1['fs']*100)}% · FM={int(_inv_t1['fm']*100)}%)</div>
            <div class='val'>{_inv_t1['inv_kw']:.1f} kW &nbsp;<span style='font-size:1.2rem;color:#FFD54F;'>({_inv_t1['inv_w']:,.0f} W)</span></div>
        </div>
        <div class='formula-box' style='margin-top:0.5rem;font-size:0.78rem;'>
            P_inst={_inv_t1['pot_instalada']:,.0f}W · FS {int(_inv_t1['fs']*100)}%
            → P_sim={_inv_t1['pot_simultanea']:,.0f}W
            + Arr.motor={_inv_t1['pot_arranque']:,.0f}W
            → P_req={_inv_t1['pot_requerida']:,.0f}W
            × FM {int(_inv_t1['fm']*100)}%
            → {_inv_t1['pot_inv_minima']:,.0f}W
            → <b>Comercial: {_inv_t1['inv_kw']:.1f} kW</b>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr class='sep'>", unsafe_allow_html=True)

        # ── EDITAR / ELIMINAR ──────────────────────────────────────────────
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
            ✏ EDITAR / ELIMINAR EQUIPOS
        </div>
        """, unsafe_allow_html=True)

        # Selector de fila a editar/eliminar
        opciones_eq = {f"{int(r['id'])} — {r['electrodomestico']}": int(r["id"])
                       for _, r in cargas.iterrows()}
        sel_eq = st.selectbox("Seleccionar equipo:", list(opciones_eq.keys()), key="sel_eq_edit")
        eq_id_sel = opciones_eq[sel_eq]
        fila_sel  = cargas[cargas["id"] == eq_id_sel].iloc[0]

        with st.expander(f"📝  Editando: {fila_sel['electrodomestico']}", expanded=True):
            col_e1, col_e2, col_e3, col_e4, col_e5 = st.columns([3, 1, 1.5, 1.5, 1])
            with col_e1:
                edit_nombre = st.text_input("Nombre del equipo", value=fila_sel["electrodomestico"], key="edit_nom")
            with col_e2:
                edit_cant   = st.number_input("Cantidad", min_value=1, max_value=100,
                                               value=int(fila_sel["cantidad"]), key="edit_cant")
            with col_e3:
                edit_pot    = st.number_input("Potencia (W)", min_value=1, max_value=50000,
                                               value=int(fila_sel["potencia_w"]), key="edit_pot")
            with col_e4:
                edit_hrs    = st.number_input("Horas/día", min_value=0.1, max_value=24.0,
                                               value=float(fila_sel["horas_dia"]), step=0.5, key="edit_hrs")
            with col_e5:
                edit_motor  = st.checkbox("Motor ⚡", value=bool(fila_sel["es_motor"]), key="edit_mot",
                                           help="Marcar si es equipo motorizado — factor de arranque ×2 a ×5 según tipo")

            # Previsualización inline
            pot_t_prev = edit_cant * edit_pot
            cons_prev  = pot_t_prev * edit_hrs
            inv_color  = "FF5252" if edit_motor else "FFD54F"
            inv_label  = "(motor — arranque ×2-×5)" if edit_motor else ""
            st.markdown(f"""
            <div style='background:#161D30; border:1px dashed #2A3A55; border-radius:8px;
                        padding:0.6rem 1rem; font-size:0.82rem; color:#8A9BBD; margin:0.5rem 0;
                        display:flex; gap:2rem; flex-wrap:wrap;'>
                <span>Pot. total: <b style='color:#FFD54F;'>{pot_t_prev:,} W</b></span>
                <span>Consumo/día: <b style='color:#FFD54F;'>{cons_prev:,.1f} Wh</b></span>
                <span>Tipo: <b style='color:#{inv_color};'>{inv_label if edit_motor else "Estándar"}</b></span>
            </div>
            """, unsafe_allow_html=True)

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("💾 Guardar Cambios", use_container_width=True, key="btn_edit_save"):
                    conn = get_conn()
                    # Verificar duplicado solo si cambió el nombre
                    nombre_nuevo = edit_nombre.strip()
                    if nombre_nuevo.lower() != fila_sel["electrodomestico"].lower():
                        existe = conn.execute(
                            "SELECT COUNT(*) FROM cargas WHERE proyecto_id=? AND LOWER(electrodomestico)=LOWER(?) AND id!=?",
                            (proyecto_id, nombre_nuevo, eq_id_sel)).fetchone()[0]
                        if existe:
                            st.error(f"⚠ Ya existe un equipo con el nombre '{nombre_nuevo}' en este proyecto.")
                            conn.close()
                            st.stop()
                    conn.execute("""
                        UPDATE cargas SET electrodomestico=?, cantidad=?, potencia_w=?, horas_dia=?, es_motor=?
                        WHERE id=? AND proyecto_id=?
                    """, (nombre_nuevo, edit_cant, edit_pot, edit_hrs,
                          1 if edit_motor else 0, eq_id_sel, proyecto_id))
                    conn.commit()
                    conn.close()
                    st.success(f"✓ '{nombre_nuevo}' actualizado correctamente")
                    st.rerun()

            with col_btn2:
                if st.button("🗑 Eliminar este Equipo", use_container_width=True, key="btn_edit_del"):
                    conn = get_conn()
                    conn.execute("DELETE FROM cargas WHERE id=? AND proyecto_id=?", (eq_id_sel, proyecto_id))
                    conn.commit()
                    conn.close()
                    st.success(f"✓ '{fila_sel['electrodomestico']}' eliminado")
                    st.rerun()

        # ── DESCARGAS ──────────────────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
            📥 EXPORTAR INFORME DEL PROYECTO
        </div>
        """, unsafe_allow_html=True)

        conn = get_conn()
        pinfo = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
        conn.close()

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.markdown("""
            <div style='background:#1A2235; border:1px solid #2A3A55; border-radius:10px;
                        padding:1rem; text-align:center; margin-bottom:0.5rem;'>
                <div style='font-size:2rem;'>📊</div>
                <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600;'>EXCEL (.xlsx)</div>
                <div style='font-size:0.78rem; color:#8A9BBD; margin-top:0.3rem;'>
                    Hoja de cargas + hoja de dimensionamiento completo con formato
                </div>
            </div>
            """, unsafe_allow_html=True)
            try:
                xls_bytes = generar_excel(proyecto_id, pinfo)
                fname_xls = f"SolarCalc_{pinfo[1].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="⬇ Descargar Excel",
                    data=xls_bytes,
                    file_name=fname_xls,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_excel"
                )
            except Exception as ex:
                st.error(f"Error generando Excel: {ex}")

        with col_dl2:
            st.markdown("""
            <div style='background:#1A2235; border:1px solid #2A3A55; border-radius:10px;
                        padding:1rem; text-align:center; margin-bottom:0.5rem;'>
                <div style='font-size:2rem;'>📄</div>
                <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600;'>PDF (.pdf)</div>
                <div style='font-size:0.78rem; color:#8A9BBD; margin-top:0.3rem;'>
                    Informe ejecutivo completo: cargas + dimensionamiento en orientación horizontal
                </div>
            </div>
            """, unsafe_allow_html=True)
            try:
                pdf_bytes = generar_pdf(proyecto_id, pinfo)
                fname_pdf = f"SolarCalc_{pinfo[1].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
                st.download_button(
                    label="⬇ Descargar PDF",
                    data=pdf_bytes,
                    file_name=fname_pdf,
                    mime="application/pdf",
                    use_container_width=True,
                    key="dl_pdf"
                )
            except Exception as ex:
                st.error(f"Error generando PDF: {ex}")

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — RECIBO DE LUZ
# ════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>2</span> CONSUMO DESDE EL RECIBO DE ENERGÍA</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        kWh/día = kWh del período ÷ Días del período<br>
        Wh/día  = kWh/día × 1.000 | Consumo total = Cargas inventariadas + Consumo real del recibo
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='info-note' style='margin-bottom:1.2rem;'>
        📋 Ingresa los datos del recibo de energía (factura de electricidad). Puedes registrar varios períodos
        para ver la evolución del consumo y comparar contra el inventario de cargas del Módulo 1.
    </div>
    """, unsafe_allow_html=True)

    # ── FORMULARIO NUEVO RECIBO ────────────────────────────────────────────
    with st.expander("➕ Registrar recibo de energía", expanded=True):
        st.markdown("<div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; margin-bottom:0.8rem;'>DATOS DEL RECIBO</div>", unsafe_allow_html=True)

        col_r1, col_r2, col_r3 = st.columns(3)
        with col_r1:
            rec_periodo = st.text_input("Período facturado", placeholder="Ej: Enero 2025, Feb-Mar 2025",
                                         key="rec_periodo")
            rec_estrato = st.selectbox("Estrato / Categoría", [
                "Estrato 1", "Estrato 2", "Estrato 3", "Estrato 4",
                "Estrato 5", "Estrato 6", "Comercial", "Industrial", "Otro"
            ], key="rec_estrato")
        with col_r2:
            rec_kwh = st.number_input("kWh consumidos en el período", min_value=0.0,
                                       max_value=100000.0, value=0.0, step=0.5,
                                       key="rec_kwh",
                                       help="Valor que aparece en el recibo como 'Energía activa' o 'Consumo kWh'")
            rec_dias = st.number_input("Días del período", min_value=1, max_value=365,
                                        value=30, key="rec_dias",
                                        help="Número de días que cubre la factura (normalmente 30)")
        with col_r3:
            rec_tarifa = st.number_input("Tarifa ($ / kWh)", min_value=0.0, max_value=5000.0,
                                          value=0.0, step=1.0, key="rec_tarifa",
                                          help="Precio unitario por kWh según el recibo (opcional)")
            rec_valor = st.number_input("Valor total factura ($)", min_value=0.0,
                                         max_value=99999999.0, value=0.0, step=100.0,
                                         key="rec_valor",
                                         help="Monto total pagado (opcional)")

        rec_obs = st.text_input("Observaciones (opcional)",
                                 placeholder="Ej: Incluye temporada de verano, mes con bomba adicional...",
                                 key="rec_obs")

        # Preview cálculos
        if rec_kwh > 0 and rec_dias > 0:
            kwh_dia_prev  = rec_kwh / rec_dias
            wh_dia_prev   = kwh_dia_prev * 1000
            wh_dia_fs     = wh_dia_prev * 1.20
            tarifa_calc   = rec_valor / rec_kwh if rec_kwh > 0 and rec_valor > 0 else rec_tarifa
            costo_dia     = kwh_dia_prev * tarifa_calc if tarifa_calc > 0 else 0

            costo_span = (f'<span style="font-size:0.82rem; color:#8A9BBD;">Costo/día: '
                          f'<b style="color:#00BCD4; font-family:Share Tech Mono;">'
                          f'${costo_dia:,.0f}</b></span>') if costo_dia > 0 else ''
            st.markdown(f"""
            <div style='background:#161D30; border:1px dashed #FFB300; border-radius:8px;
                        padding:0.8rem 1.2rem; margin:0.8rem 0; display:flex; gap:2.5rem; flex-wrap:wrap;'>
                <span style='font-size:0.82rem; color:#8A9BBD;'>kWh/día:
                    <b style='color:#FFD54F; font-family:Share Tech Mono;'>{kwh_dia_prev:.2f}</b></span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Wh/día:
                    <b style='color:#FFB300; font-family:Share Tech Mono;'>{wh_dia_prev:,.0f}</b></span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Wh/día +20% FS:
                    <b style='color:#00E676; font-family:Share Tech Mono;'>{wh_dia_fs:,.0f}</b></span>
                {costo_span}
            </div>
            """, unsafe_allow_html=True)

        col_btn_r1, col_btn_r2 = st.columns([1, 2])
        with col_btn_r1:
            if st.button("💾 Guardar Recibo", use_container_width=True, key="btn_guardar_recibo"):
                if rec_kwh <= 0:
                    st.error("⚠ Ingresa los kWh del período")
                elif not rec_periodo.strip():
                    st.error("⚠ Ingresa el período del recibo")
                else:
                    conn = get_conn()
                    conn.execute("""
                        INSERT INTO recibos(proyecto_id, periodo, kwh_periodo, dias_periodo,
                            estrato, tarifa_kwh, valor_total, observaciones)
                        VALUES(?,?,?,?,?,?,?,?)
                    """, (proyecto_id, rec_periodo.strip(), rec_kwh, rec_dias,
                          rec_estrato, rec_tarifa if rec_tarifa > 0 else None,
                          rec_valor if rec_valor > 0 else None,
                          rec_obs.strip() if rec_obs.strip() else None))
                    conn.commit()
                    conn.close()
                    st.success("✓ Recibo registrado correctamente")
                    st.rerun()

    # ── HISTORIAL DE RECIBOS ───────────────────────────────────────────────
    conn = get_conn()
    recibos_df = pd.read_sql(
        "SELECT * FROM recibos WHERE proyecto_id=? ORDER BY id DESC", conn, params=(proyecto_id,))
    cargas_rec = pd.read_sql(
        "SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?", conn, params=(proyecto_id,))
    conn.close()

    if recibos_df.empty:
        st.markdown("""
        <div class='info-note' style='text-align:center; padding:1.2rem; margin-top:1rem;'>
            No hay recibos registrados. Agrega el primer recibo arriba.
        </div>
        """, unsafe_allow_html=True)
    else:
        # Calcular columnas derivadas
        recibos_df["kwh_dia"]   = recibos_df["kwh_periodo"] / recibos_df["dias_periodo"]
        recibos_df["wh_dia"]    = recibos_df["kwh_dia"] * 1000
        recibos_df["wh_dia_fs"] = recibos_df["wh_dia"] * 1.20
        recibos_df["tarifa_calc"] = recibos_df.apply(
            lambda r: r["valor_total"] / r["kwh_periodo"]
            if (r["valor_total"] and r["kwh_periodo"] > 0) else (r["tarifa_kwh"] or 0), axis=1)

        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
            📋 HISTORIAL DE RECIBOS
        </div>
        """, unsafe_allow_html=True)

        # Tabla resumen
        disp_rec = recibos_df[["id","periodo","estrato","kwh_periodo","dias_periodo",
                                "kwh_dia","wh_dia","wh_dia_fs","valor_total"]].copy()
        disp_rec.columns = ["ID","Período","Estrato/Cat.","kWh período","Días",
                             "kWh/día","Wh/día","Wh/día +20%FS","Valor ($)"]
        disp_rec["kWh/día"]     = disp_rec["kWh/día"].round(2)
        disp_rec["Wh/día"]      = disp_rec["Wh/día"].round(0).astype(int)
        disp_rec["Wh/día +20%FS"] = disp_rec["Wh/día +20%FS"].round(0).astype(int)
        st.dataframe(disp_rec.set_index("ID"), use_container_width=True)

        # Recibo más reciente como referencia
        recibo_ref = recibos_df.iloc[0]
        kwh_dia_ref  = recibo_ref["kwh_dia"]
        wh_dia_ref   = recibo_ref["wh_dia"]
        wh_dia_fs_ref= recibo_ref["wh_dia_fs"]

        # Consumo inventariado (Módulo 1)
        consumo_inv = 0.0
        if not cargas_rec.empty:
            consumo_inv = (cargas_rec["cantidad"] * cargas_rec["potencia_w"] * cargas_rec["horas_dia"]).sum()
        consumo_inv_fs = consumo_inv * 1.20

        # Diferencia
        diferencia    = wh_dia_ref - consumo_inv
        diferencia_fs = wh_dia_fs_ref - consumo_inv_fs
        pct_dif       = (diferencia / wh_dia_ref * 100) if wh_dia_ref > 0 else 0

        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.5rem;'>
            📊 REFERENCIA: RECIBO REAL vs INVENTARIO DE CARGAS
        </div>
        <div style='background:rgba(0,188,212,0.08);border:1px solid rgba(0,188,212,0.25);
                    border-radius:8px;padding:0.7rem 1rem;font-size:0.82rem;
                    color:#8A9BBD;margin-bottom:1rem;'>
            ℹ El recibo de energía refleja el <b style='color:#E8EDF5;'>consumo real total</b>
            del inmueble — incluye pérdidas, equipos en standby y cargas pasivas.
            Es la fuente más confiable para dimensionar.
            El inventario de cargas se muestra solo como referencia.
            <b style='color:#FFD54F;'>El dimensionamiento se basa en el valor que selecciones abajo.</b>
        </div>
        <div style='font-size:0.8rem; color:#8A9BBD; margin-bottom:1rem;'>
            Período de referencia: <b style='color:#FFD54F;'>{recibo_ref['periodo']}</b>
            — {recibo_ref['kwh_periodo']} kWh en {int(recibo_ref['dias_periodo'])} días
        </div>
        """, unsafe_allow_html=True)

        col_cmp1, col_cmp2, col_cmp3, col_cmp4 = st.columns(4)
        with col_cmp1:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#00BCD4;'>
                <div style='font-size:1.2rem; margin-bottom:0.3rem;'>🧾</div>
                <div class='metric-val' style='color:#00BCD4;'>{wh_dia_ref:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>RECIBO REAL</div>
            </div>""", unsafe_allow_html=True)
        with col_cmp2:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#00E676;'>
                <div style='font-size:1.2rem; margin-bottom:0.3rem;'>🧾</div>
                <div class='metric-val' style='color:#00E676;'>{wh_dia_fs_ref:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>RECIBO + 20% FS</div>
            </div>""", unsafe_allow_html=True)
        with col_cmp3:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#FFB300;'>
                <div style='font-size:1.2rem; margin-bottom:0.3rem;'>⚡</div>
                <div class='metric-val'>{consumo_inv:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>INVENTARIO CARGAS</div>
            </div>""", unsafe_allow_html=True)
        with col_cmp4:
            # Cobertura informativa — nunca un error
            if consumo_inv > 0 and wh_dia_ref > 0:
                pct_cob   = consumo_inv / wh_dia_ref * 100
                cob_color = "#00E676" if pct_cob >= 80 else "#FFD54F"
                st.markdown(f"""
                <div class='metric-box' style='border-color:{cob_color};'>
                    <div style='font-size:1.2rem; margin-bottom:0.3rem;'>📋</div>
                    <div class='metric-val' style='color:{cob_color};'>{pct_cob:.0f}%</div>
                    <div class='metric-unit'>del recibo</div>
                    <div class='metric-label'>COBERTURA INVENTARIO</div>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class='metric-box' style='border-color:#2A3A55;'>
                    <div style='font-size:1.2rem; margin-bottom:0.3rem;'>📋</div>
                    <div class='metric-val' style='color:#8A9BBD;'>—</div>
                    <div class='metric-unit'>sin inventario</div>
                    <div class='metric-label'>INVENTARIO VACÍO</div>
                </div>""", unsafe_allow_html=True)

        # Nota suave solo si el inventario cubre muy poco del recibo
        if consumo_inv > 0 and wh_dia_ref > 0:
            pct_cob2 = consumo_inv / wh_dia_ref * 100
            if pct_cob2 < 85:
                st.markdown(f"""
                <div style='background:rgba(255,211,78,0.07);border:1px solid rgba(255,211,78,0.2);
                            border-radius:8px;padding:0.6rem 1rem;margin-top:0.8rem;
                            font-size:0.8rem;color:#8A9BBD;'>
                    💡 El inventario cubre el {pct_cob2:.0f}% del recibo — es normal,
                    ya que el recibo incluye pérdidas eléctricas, equipos en standby
                    y consumos que no siempre se registran en el inventario.
                    <b style='color:#00BCD4;'>El recibo es la base recomendada para dimensionar.</b>
                </div>
                """, unsafe_allow_html=True)
            elif consumo_inv > wh_dia_ref:
                st.markdown(f"""
                <div class='result-ok' style='margin-top:0.8rem;'>
                    ✓ El inventario ({consumo_inv:,.0f} Wh/día) supera el recibo
                    ({wh_dia_ref:,.0f} Wh/día). Dimensionar desde el inventario
                    será conservador.
                </div>
                """, unsafe_allow_html=True)

        # ── SELECTOR: qué consumo usar para dimensionar ────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
            🎯 CONSUMO BASE PARA DIMENSIONAMIENTO
        </div>
        """, unsafe_allow_html=True)

        modo_consumo = st.radio(
            "Selecciona qué valor usar como base para el cálculo del sistema:",
            options=[
                f"📋 Inventario de cargas — {consumo_inv:,.0f} Wh/día",
                f"🧾 Recibo de energía (período: {recibo_ref['periodo']}) — {wh_dia_ref:,.0f} Wh/día",
                f"📊 Mayor de los dos — {max(consumo_inv, wh_dia_ref):,.0f} Wh/día (recomendado)",
                "✏ Valor personalizado",
            ],
            key="modo_consumo_radio",
            horizontal=False
        )

        consumo_custom_val = 0.0
        if "personalizado" in modo_consumo:
            consumo_custom_val = st.number_input(
                "Consumo base personalizado (Wh/día)", min_value=100.0,
                max_value=500000.0, value=float(max(consumo_inv, wh_dia_ref)),
                step=100.0, key="consumo_custom")

        if "Inventario" in modo_consumo:
            consumo_base_sel = consumo_inv
        elif "Recibo" in modo_consumo:
            consumo_base_sel = wh_dia_ref
        elif "Mayor" in modo_consumo:
            consumo_base_sel = max(consumo_inv, wh_dia_ref)
        else:
            consumo_base_sel = consumo_custom_val

        consumo_base_fs_sel = consumo_base_sel * 1.20

        st.markdown(f"""
        <div class='result-highlight'>
            <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase; letter-spacing:1px;'>
                Consumo base seleccionado → con Factor de Seguridad 20%
            </div>
            <div class='val'>{consumo_base_sel:,.0f} Wh/día
                <span style='font-size:1rem; color:#8A9BBD;'>→</span>
                <span style='color:#00E676;'>{consumo_base_fs_sel:,.0f} Wh/día</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── AUTONOMÍA DESEADA ──────────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif;font-size:1.1rem;font-weight:600;
                    color:#FFB300;letter-spacing:1px;margin-bottom:0.8rem;'>
            🔋 HORAS DE AUTONOMÍA DESEADA (OFF-GRID / HÍBRIDO)
        </div>
        """, unsafe_allow_html=True)

        col_aut1, col_aut2 = st.columns([1, 1])
        with col_aut1:
            horas_autonomia_sel = st.number_input(
                "Horas de autonomía deseada (h)",
                min_value=1.0, max_value=96.0,
                value=float(st.session_state.get("horas_autonomia_deseada", 24.0)),
                step=1.0,
                help="Tiempo que el banco de baterías debe cubrir sin sol ni red. "
                     "Ejemplos: 8h (nocturno) · 24h (1 día) · 48h (2 días) · 72h (3 días)",
                key="horas_aut_tab3")
            dias_autonomia_calc = horas_autonomia_sel / 24.0
        with col_aut2:
            st.markdown(f"""
            <div class='sol-card' style='padding:0.7rem 1rem;'>
                <div style='color:#8A9BBD;font-size:0.75rem;text-transform:uppercase;
                            letter-spacing:1px;margin-bottom:0.5rem;'>Referencia</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                    <tr><td style='color:#8A9BBD;padding:0.2rem 0;'>Nocturno (sin sol)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>8 h</td></tr>
                    <tr><td style='color:#8A9BBD;padding:0.2rem 0;'>1 día completo</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>24 h</td></tr>
                    <tr><td style='color:#8A9BBD;padding:0.2rem 0;'>2 días (fin de semana)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>48 h</td></tr>
                    <tr><td style='color:#8A9BBD;padding:0.2rem 0;'>3 días (nublado)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>72 h</td></tr>
                </table>
                <div style='margin-top:0.5rem;font-family:Share Tech Mono;color:#00E676;font-size:0.85rem;'>
                    {horas_autonomia_sel:.0f} h = {dias_autonomia_calc:.2f} días de autonomía
                </div>
            </div>
            """, unsafe_allow_html=True)

        # Guardar selección en session_state para que otros tabs la usen
        st.session_state["consumo_recibo_wh"]       = consumo_base_sel
        st.session_state["consumo_recibo_fs_wh"]    = consumo_base_fs_sel
        st.session_state["recibo_ref_periodo"]      = recibo_ref["periodo"]
        st.session_state["horas_autonomia_deseada"] = horas_autonomia_sel
        st.session_state["dias_autonomia_deseada"]  = dias_autonomia_calc

        # ── ANÁLISIS DE TENDENCIA ─────────────────────────────────────────
        if len(recibos_df) >= 2:
            st.markdown("<hr class='sep'>", unsafe_allow_html=True)
            st.markdown("""
            <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                        color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
                📈 TENDENCIA DE CONSUMO
            </div>
            """, unsafe_allow_html=True)
            rec_tend = recibos_df[["periodo","kwh_dia","wh_dia","valor_total"]].copy().iloc[::-1]
            rec_tend.columns = ["Período","kWh/día","Wh/día","Valor ($)"]
            kwh_promedio = rec_tend["kWh/día"].mean()
            kwh_max      = rec_tend["kWh/día"].max()
            kwh_min      = rec_tend["kWh/día"].min()

            col_t1, col_t2, col_t3 = st.columns(3)
            with col_t1:
                st.markdown(f"""
                <div class='metric-box'>
                    <div class='metric-val'>{kwh_promedio:.2f}</div>
                    <div class='metric-unit'>kWh/día</div>
                    <div class='metric-label'>PROMEDIO</div>
                </div>""", unsafe_allow_html=True)
            with col_t2:
                st.markdown(f"""
                <div class='metric-box' style='border-color:#FF5252;'>
                    <div class='metric-val' style='color:#FF5252;'>{kwh_max:.2f}</div>
                    <div class='metric-unit'>kWh/día</div>
                    <div class='metric-label'>MÁXIMO</div>
                </div>""", unsafe_allow_html=True)
            with col_t3:
                st.markdown(f"""
                <div class='metric-box' style='border-color:#00E676;'>
                    <div class='metric-val' style='color:#00E676;'>{kwh_min:.2f}</div>
                    <div class='metric-unit'>kWh/día</div>
                    <div class='metric-label'>MÍNIMO</div>
                </div>""", unsafe_allow_html=True)

            st.dataframe(rec_tend.set_index("Período"), use_container_width=True)

        # ── ELIMINAR RECIBO ────────────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        with st.expander("🗑 Eliminar un recibo registrado"):
            opciones_rec = {f"{int(r['id'])} — {r['periodo']} ({r['kwh_periodo']} kWh)": int(r["id"])
                            for _, r in recibos_df.iterrows()}
            sel_rec_del = st.selectbox("Seleccionar recibo a eliminar:", list(opciones_rec.keys()),
                                        key="sel_rec_del")
            if st.button("🗑 Confirmar eliminación", use_container_width=True, key="btn_del_recibo"):
                conn = get_conn()
                conn.execute("DELETE FROM recibos WHERE id=? AND proyecto_id=?",
                              (opciones_rec[sel_rec_del], proyecto_id))
                conn.commit()
                conn.close()
                st.success("Recibo eliminado ✓")
                st.rerun()


# ════════════════════════════════════════════════════════════════════════════
# TAB 3 — TENSIÓN DC
# ════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>3</span> ESTÁNDAR DE TENSIÓN DC DEL SISTEMA</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        0 – 1.999 Wh/día → 12 V DC<br>
        2.000 – 3.999 Wh/día → 24 V DC<br>
        ≥ 4.000 Wh/día → 48 V DC
    </div>
    """, unsafe_allow_html=True)

    conn = get_conn()
    cargas2 = pd.read_sql("SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?",
                          conn, params=(proyecto_id,))
    conn.close()

    # Consumo desde inventario
    consumo_inv2 = (cargas2["cantidad"] * cargas2["potencia_w"] * cargas2["horas_dia"]).sum() if not cargas2.empty else 0.0
    # Consumo desde recibo (si existe en session_state)
    consumo_rec2 = st.session_state.get("consumo_recibo_wh", 0.0)
    consumo_rec_fs2 = st.session_state.get("consumo_recibo_fs_wh", 0.0)
    periodo_rec2 = st.session_state.get("recibo_ref_periodo", "")

    # Selector de fuente
    opciones_f2 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec2 > 0:
        opciones_f2.append(f"🧾 Recibo de energía ({periodo_rec2})")
        opciones_f2.append("📊 Mayor de los dos (recomendado)")

    fuente2 = st.radio("Base de consumo para determinar tensión DC:",
                        opciones_f2, horizontal=True, key="fuente_vdc")

    if "Recibo" in fuente2:
        consumo_base2 = consumo_rec2
    elif "Mayor" in fuente2:
        consumo_base2 = max(consumo_inv2, consumo_rec2)
    else:
        consumo_base2 = consumo_inv2

    if consumo_base2 == 0:
        st.markdown("<div class='warn-box'>⚠ Registra cargas (Módulo 1) o un recibo (Módulo 2) para calcular la tensión DC</div>", unsafe_allow_html=True)
    else:
        consumo_fs2 = consumo_base2 * 1.20
        vdc = tension_dc(consumo_fs2)

        if consumo_rec2 > 0:
            st.markdown(f"""
            <div class='info-note' style='margin-bottom:1rem;'>
                Inventario de cargas: <b style='color:#FFD54F;'>{consumo_inv2:,.0f} Wh/día</b> &nbsp;|&nbsp;
                Recibo ({periodo_rec2}): <b style='color:#00BCD4;'>{consumo_rec2:,.0f} Wh/día</b> &nbsp;|&nbsp;
                <b>Usando: {consumo_base2:,.0f} Wh/día</b>
            </div>
            """, unsafe_allow_html=True)

        rangos = [
            ("0 – 1.999 Wh/día", "12 V", consumo_fs2 < 2000),
            ("2.000 – 3.999 Wh/día", "24 V", 2000 <= consumo_fs2 < 4000),
            ("≥ 4.000 Wh/día", "48 V", consumo_fs2 >= 4000),
        ]
        cols_r = st.columns(3)
        for i, (rango, volt, activo) in enumerate(rangos):
            with cols_r[i]:
                color = "#FFB300" if activo else "#2A3A55"
                bg = "rgba(255,179,0,0.12)" if activo else "transparent"
                st.markdown(f"""
                <div style='background:{bg}; border:2px solid {color}; border-radius:12px; padding:1.5rem; text-align:center;'>
                    <div style='font-size:0.75rem; color:#8A9BBD; text-transform:uppercase; letter-spacing:1px;'>{rango}</div>
                    <div style='font-family:Share Tech Mono; font-size:2.5rem; color:{color}; font-weight:700; margin:0.5rem 0;'>{volt}</div>
                    {'<div style="font-size:1.2rem;">✓ SELECCIONADO</div>' if activo else ''}
                </div>
                """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class='result-highlight' style='margin-top:1.5rem;'>
            <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase; letter-spacing:1px;'>Consumo diario (con FS 20%): {consumo_fs2:,.0f} Wh/día</div>
            <div class='val'>Tensión DC del Sistema: {vdc} V</div>
        </div>
        """, unsafe_allow_html=True)

        # Guardar en proyecto
        conn = get_conn()
        conn.execute("UPDATE proyectos SET tension_dc=? WHERE id=?", (vdc, proyecto_id))
        conn.commit()
        conn.close()

# ════════════════════════════════════════════════════════════════════════════
# TAB 4 — HORA SOLAR PICO
# ════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>4</span> HORA SOLAR PICO (HSP)</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        HSP = Irradiación mes menor (kWh/m²/mes) ÷ 30 días
    </div>
    """, unsafe_allow_html=True)

    col_pvg1, col_pvg2 = st.columns([1,1])
    with col_pvg1:
        st.markdown("""
        <div class='sol-card'>
            <div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600; margin-bottom:0.8rem;'>🌐 Herramienta PVGIS (Europa JRC)</div>
            <div style='color:#8A9BBD; font-size:0.85rem; margin-bottom:1rem;'>
                Accede a la herramienta oficial de la Comisión Europea para obtener la irradiación mensual de cualquier ubicación.
                Busca el mes con menor irradiación y anota el valor.
            </div>
            <div class='info-note'>
                📌 URL: <code>https://re.jrc.ec.europa.eu/pvg_tools/es/</code>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if st.button("🔗 Abrir PVGIS en nueva pestaña", use_container_width=True):
            st.markdown("""
            <script>window.open('https://re.jrc.ec.europa.eu/pvg_tools/es/', '_blank');</script>
            """, unsafe_allow_html=True)

        st.markdown("""
        <div class='sol-card' style='margin-top:1rem;'>
            <div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; margin-bottom:0.8rem;'>📋 Instrucciones PVGIS</div>
            <ol style='color:#8A9BBD; font-size:0.85rem; line-height:1.8;'>
                <li>Abre la herramienta PVGIS</li>
                <li>Selecciona la ubicación del proyecto en el mapa</li>
                <li>En "PV Power" o "Solar radiation", selecciona irradiación mensual</li>
                <li>Identifica el mes con el valor más bajo (ej: noviembre = 126.72 kWh/m²)</li>
                <li>Ingresa ese valor a la derecha</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)

    with col_pvg2:
        st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
        st.markdown("**Cálculo de HSP**")

        municipio_hsp = st.text_input("Municipio analizado", placeholder="Ej: Bogotá, Medellín, Cali...", key="municipio_hsp")
        mes_menor = st.selectbox("Mes con menor irradiación",
            ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"])
        irradiacion = st.number_input("Irradiación del mes menor (kWh/m²/mes)",
                                       min_value=10.0, max_value=300.0, value=126.72, step=0.01,
                                       help="Valor obtenido de PVGIS para el mes más desfavorable")

        hsp_calc = irradiacion / 30

        st.markdown(f"""
        <div class='result-highlight'>
            <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase;'>
                {irradiacion} kWh/m²/mes ÷ 30 días =
            </div>
            <div class='val'>{hsp_calc:.2f} h</div>
            <div style='color:#FFD54F; font-size:0.85rem; margin-top:0.3rem;'>Horas Solar Pico ({mes_menor})</div>
        </div>
        """, unsafe_allow_html=True)

        dias_autonomia = st.slider("Días de autonomía (opcional)", 1, 5, 1)
        hsp_ajustada = hsp_calc / dias_autonomia if dias_autonomia > 1 else hsp_calc
        if dias_autonomia > 1:
            st.markdown(f"""
            <div class='result-ok'>
                <div style='color:#8A9BBD; font-size:0.8rem;'>HSP ajustada ({dias_autonomia} días)</div>
                <div class='val'>{hsp_ajustada:.2f} h</div>
            </div>
            """, unsafe_allow_html=True)

        if st.button("💾 Guardar HSP al Proyecto", use_container_width=True):
            conn = get_conn()
            conn.execute("UPDATE proyectos SET hsp=?, municipio=? WHERE id=?",
                         (round(hsp_calc, 2), municipio_hsp or None, proyecto_id))
            conn.commit()
            conn.close()
            st.success(f"HSP = {hsp_calc:.2f} h guardado ✓")
            st.rerun()

        # Leer HSP guardado
        conn = get_conn()
        p3 = conn.execute("SELECT hsp, municipio FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
        conn.close()
        if p3 and p3[0]:
            st.markdown(f"""
            <div class='info-note' style='margin-top:0.5rem;'>
                ✓ HSP guardado en proyecto: <b>{p3[0]} h</b> — {p3[1] or ""}
            </div>
            """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 5 — PANEL SOLAR
# ════════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>5</span> SELECCIÓN DEL PANEL SOLAR</div>
    """, unsafe_allow_html=True)

    # Leer panel guardado
    conn = get_conn()
    panel_existente = pd.read_sql(
        "SELECT * FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1", conn, params=(proyecto_id,))
    conn.close()

    # ── CATÁLOGO DE PANELES SOLARES ────────────────────────────────────────
    CATALOGO_PANELES = {
        "JinkoSolar Tiger Neo 550":  {"wp": 550, "voc": 49.8, "isc": 13.96},
        "Trina Solar Vertex 550":    {"wp": 550, "voc": 49.5, "isc": 14.01},
        "Canadian Solar HiKu6 550":  {"wp": 550, "voc": 49.7, "isc": 13.98},
        "LONGi Hi-MO 6 550":         {"wp": 550, "voc": 50.2, "isc": 13.95},
        "JA Solar JAM72D30 550":     {"wp": 550, "voc": 49.6, "isc": 14.02},
        "Risen RSM132-8-650M":       {"wp": 650, "voc": 56.4, "isc": 14.71},
        "Bifacial 440W":             {"wp": 440, "voc": 41.8, "isc": 13.45},
        "Personalizado":             {"wp": 550, "voc": 49.9, "isc": 14.0},
    }

    with st.expander("📋 Catálogo de Paneles Solares", expanded=False):
        st.markdown("""
        <div class='info-note' style='margin-bottom:0.8rem;'>
            Selecciona un panel del catálogo para cargar sus parámetros automáticamente.
        </div>
        """, unsafe_allow_html=True)
        cat_cols = st.columns(4)
        for i, (nombre_panel, params) in enumerate(CATALOGO_PANELES.items()):
            if nombre_panel == "Personalizado":
                continue
            with cat_cols[i % 4]:
                st.markdown(f"""
                <div style='background:#1A2235;border:1px solid #2A3A55;border-radius:8px;
                            padding:0.6rem;text-align:center;margin-bottom:0.4rem;'>
                    <div style='font-size:0.72rem;color:#FFB300;font-weight:600;'>{nombre_panel}</div>
                    <div style='font-family:Share Tech Mono;font-size:0.85rem;color:#FFD54F;'>{params['wp']}Wp</div>
                    <div style='font-size:0.7rem;color:#8A9BBD;'>Voc={params['voc']}V | Isc={params['isc']}A</div>
                </div>
                """, unsafe_allow_html=True)
                if st.button(f"Usar", key=f"cat_panel_{nombre_panel}_{i}", use_container_width=True):
                    st.session_state["panel_cat_sel"] = nombre_panel
                    st.rerun()

    # Detectar si hay selección del catálogo
    _cat_sel = st.session_state.get("panel_cat_sel", "")
    _cat_params = CATALOGO_PANELES.get(_cat_sel, {})

    col_p1, col_p2 = st.columns([1,1])
    with col_p1:
        st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
        st.markdown("**Parámetros del panel seleccionado**")
        if _cat_sel:
            st.markdown(f"<div class='info-note' style='margin-bottom:0.5rem;'>✓ Panel del catálogo: <b>{_cat_sel}</b></div>", unsafe_allow_html=True)
        panel_modelo = st.text_input("Modelo / Referencia",
                                      value=_cat_params.get("wp") and _cat_sel or (panel_existente["modelo"].values[0] if not panel_existente.empty else ""),
                                      placeholder="Ej: Canadian Solar CS6W-550T")
        if _cat_sel and not panel_existente.empty == False:
            panel_modelo = _cat_sel
        panel_wp = st.number_input("Potencia pico (Wp)", min_value=10, max_value=1000,
                                    value=int(_cat_params.get("wp", int(panel_existente["potencia_wp"].values[0]) if not panel_existente.empty else 550)))
        panel_voc = st.number_input("Tensión Voc (V)", min_value=5.0, max_value=100.0,
                                     value=float(_cat_params.get("voc", float(panel_existente["voc"].values[0]) if not panel_existente.empty else 49.9)), step=0.1)
        panel_isc = st.number_input("Corriente Isc (A)", min_value=0.1, max_value=30.0,
                                     value=float(_cat_params.get("isc", float(panel_existente["isc"].values[0]) if not panel_existente.empty else 14.0)), step=0.1)

        # ── CÁLCULO DE STRINGS MPPT ──────────────────────────────────────────
        st.markdown("<hr class='sep' style='margin:0.8rem 0;'>", unsafe_allow_html=True)
        st.markdown("**🔗 Cálculo Automático de Strings MPPT**")
        mppt_max_v5 = st.number_input("Tensión máxima MPPT del controlador (V)",
                                       min_value=100.0, max_value=1500.0, value=600.0, step=10.0,
                                       help="Tensión máxima de entrada del controlador MPPT (ej: 600V)")

        if panel_voc > 0 and mppt_max_v5 > 0:
            paneles_por_string5 = int(mppt_max_v5 / panel_voc)
            st.markdown(f"""
            <div class='formula-box'>
                {mppt_max_v5:.0f}V ÷ {panel_voc}V (Voc) = <b style='color:#FFB300;'>{paneles_por_string5} paneles/string</b>
            </div>
            """, unsafe_allow_html=True)
            st.session_state["_paneles_por_string"] = paneles_por_string5

        if st.button("💾 Guardar Especificaciones Panel", use_container_width=True):
            conn = get_conn()
            conn.execute("DELETE FROM paneles WHERE proyecto_id=?", (proyecto_id,))
            conn.execute("INSERT INTO paneles(proyecto_id,modelo,potencia_wp,voc,isc) VALUES(?,?,?,?,?)",
                         (proyecto_id, panel_modelo, panel_wp, panel_voc, panel_isc))
            conn.commit()
            conn.close()
            st.session_state.pop("panel_cat_sel", None)
            st.success("Panel guardado ✓")
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with col_p2:
        _paneles_por_string5 = st.session_state.get("_paneles_por_string", int(600 / panel_voc) if panel_voc > 0 else 12)
        _fusible_dc5 = panel_isc * 1.25
        _fusible_std5 = next((f for f in [10,15,20,25,30,40] if f >= _fusible_dc5), 40)
        st.markdown(f"""
        <div class='sol-card'>
            <div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; margin-bottom:1rem; font-size:1.1rem;'>
                FICHA TECNICA
            </div>
            <div style='display:grid; grid-template-columns:1fr 1fr; gap:0.8rem;'>
                <div style='background:#161D30; border-radius:8px; padding:1rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.8rem; color:#FFB300;'>{panel_wp}</div>
                    <div style='font-size:0.75rem; color:#8A9BBD; margin-top:0.3rem;'>Wp - POTENCIA PICO</div>
                </div>
                <div style='background:#161D30; border-radius:8px; padding:1rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.8rem; color:#00BCD4;'>{panel_voc}</div>
                    <div style='font-size:0.75rem; color:#8A9BBD; margin-top:0.3rem;'>V - Voc</div>
                </div>
                <div style='background:#161D30; border-radius:8px; padding:1rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.8rem; color:#00E676;'>{panel_isc}</div>
                    <div style='font-size:0.75rem; color:#8A9BBD; margin-top:0.3rem;'>A - Isc</div>
                </div>
                <div style='background:#161D30; border-radius:8px; padding:1rem; text-align:center;'>
                    <div style='font-family:Share Tech Mono; font-size:1.8rem; color:#FFD54F;'>{(panel_wp/panel_voc):.1f}</div>
                    <div style='font-size:0.75rem; color:#8A9BBD; margin-top:0.3rem;'>A - Impp APROX</div>
                </div>
            </div>
            <div style='margin-top:1rem; color:#8A9BBD; font-size:0.8rem;'>
                Modelo: <span style='color:#E8EDF5;'>{panel_modelo or "Sin especificar"}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Tabla de strings
        st.markdown("""
        <div style='color:#FFB300;font-family:Rajdhani,sans-serif;font-weight:600;
                    font-size:0.9rem;letter-spacing:1px;margin:0.8rem 0 0.4rem;'>
            DISTRIBUCION DE STRINGS (ejemplos)
        </div>""", unsafe_allow_html=True)
        _ejemplos_pan = [5, 10, 15, 20, 30, 40]
        _filas_str = []
        for _np in _ejemplos_pan:
            _nstr = math.ceil(_np / _paneles_por_string5) if _paneles_por_string5 > 0 else 1
            _pp_str = math.ceil(_np / _nstr) if _nstr > 0 else _np
            _filas_str.append({"Paneles": _np, "Strings": _nstr, "Pan/String": _pp_str})
        st.dataframe(pd.DataFrame(_filas_str).set_index("Paneles"), use_container_width=True, hide_index=False)

        # Protecciones DC
        st.markdown(f"""
        <div class='sol-card' style='margin-top:0.8rem;'>
            <div style='color:#FFB300;font-family:Rajdhani,sans-serif;font-weight:600;
                        margin-bottom:0.6rem;'>PROTECCIONES DC (calculo automatico)</div>
            <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD;padding:0.35rem 0;'>Fusible string (Isc x 1.25)</td>
                    <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                        {panel_isc}A x 1.25 = {_fusible_dc5:.1f}A seleccionar: <b style='color:#FFB300;'>{_fusible_std5}A DC</b></td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD;padding:0.35rem 0;'>DPS DC</td>
                    <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                        Tipo II - 1000 VDC - 40 kA</td>
                </tr>
                <tr>
                    <td style='color:#8A9BBD;padding:0.35rem 0;'>Seccionador DC</td>
                    <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                        mayor o igual Voc string</td>
                </tr>
            </table>
        </div>
        """, unsafe_allow_html=True)

    # Catalogo de baterias LiFePO4
    st.markdown("<hr class='sep'>", unsafe_allow_html=True)
    st.markdown("""
    <div style='color:#FFB300;font-family:Rajdhani,sans-serif;font-weight:700;
                font-size:1.1rem;letter-spacing:1px;margin-bottom:0.8rem;'>
        CATALOGO DE BATERIAS LiFePO4
    </div>""", unsafe_allow_html=True)
    CATALOGO_BATERIAS = [
        {"Marca": "Pylontech", "Modelo": "US5000",       "kWh": 4.8,  "V": 48,   "Ah": 100},
        {"Marca": "Dyness",    "Modelo": "BX51100",      "kWh": 5.12, "V": 51.2, "Ah": 100},
        {"Marca": "BYD",       "Modelo": "Battery Box",  "kWh": 5.1,  "V": 51.2, "Ah": 100},
        {"Marca": "CATL",      "Modelo": "EnerOne Plus", "kWh": 5.0,  "V": 48,   "Ah": 104},
        {"Marca": "Sunsynk",   "Modelo": "LBSA016",      "kWh": 5.12, "V": 51.2, "Ah": 100},
        {"Marca": "Hubble",    "Modelo": "AM-5",         "kWh": 5.5,  "V": 51.2, "Ah": 107},
    ]
    df_bats_cat = pd.DataFrame(CATALOGO_BATERIAS)
    st.dataframe(df_bats_cat.set_index("Marca"), use_container_width=True, hide_index=False)
    st.markdown("""
    <div class='info-note'>
        LiFePO4 DoD recomendado 80-90%, ciclos de vida 3500-6000, BMS integrado.
    </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 6 — POTENCIA INSTALADA
# ════════════════════════════════════════════════════════════════════════════
with tab6:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>6</span> CÁLCULO DE POTENCIA INSTALADA EN PANELES</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        Potencia Instalada (Wp) = Consumo día con FS 20% (Wh) ÷ (HSP × PR)<br>
        Número de paneles = Potencia Instalada ÷ Potencia del panel (Wp)<br>
        <span style='font-size:0.85em;color:#8A9BBD;'>
        PR (Performance Ratio): pérdidas reales del sistema
        (temperatura ~5%, inversor ~3%, sombreado ~2%, suciedad ~2%, cableado ~2%)
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Leer datos necesarios
    conn = get_conn()
    cargas5 = pd.read_sql("SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?",
                           conn, params=(proyecto_id,))
    p5 = conn.execute("SELECT hsp FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    panel5 = conn.execute("SELECT potencia_wp FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                           (proyecto_id,)).fetchone()
    conn.close()

    hsp5 = p5[0] if p5 and p5[0] else None
    pot_panel5 = panel5[0] if panel5 else None

    consumo_inv5 = (cargas5["cantidad"] * cargas5["potencia_w"] * cargas5["horas_dia"]).sum() if not cargas5.empty else 0.0
    consumo_rec5 = st.session_state.get("consumo_recibo_wh", 0.0)
    periodo_rec5 = st.session_state.get("recibo_ref_periodo", "")

    opciones_f5 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec5 > 0:
        opciones_f5.append(f"🧾 Recibo de energía ({periodo_rec5})")
        opciones_f5.append("📊 Mayor de los dos (recomendado)")

    fuente5 = st.radio("Base de consumo para dimensionar paneles:",
                        opciones_f5, horizontal=True, key="fuente_pot5")

    if "Recibo" in fuente5:
        consumo5 = consumo_rec5
    elif "Mayor" in fuente5:
        consumo5 = max(consumo_inv5, consumo_rec5)
    else:
        consumo5 = consumo_inv5

    if consumo5 == 0:
        st.markdown("<div class='warn-box'>⚠ Ingresa las cargas en el Módulo 1 o un recibo en el Módulo 2</div>", unsafe_allow_html=True)
    else:
        consumo5_fs = consumo5 * 1.20

        if consumo_rec5 > 0:
            st.markdown(f"""
            <div class='info-note' style='margin-bottom:0.8rem;'>
                Inventario: <b style='color:#FFD54F;'>{consumo_inv5:,.0f} Wh/día</b> &nbsp;|&nbsp;
                Recibo: <b style='color:#00BCD4;'>{consumo_rec5:,.0f} Wh/día</b> &nbsp;|&nbsp;
                <b>Usando: {consumo5:,.0f} Wh/día → {consumo5_fs:,.0f} Wh/día con FS 20%</b>
            </div>
            """, unsafe_allow_html=True)

        col5a, col5b = st.columns([1,1])
        with col5a:
            st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
            st.markdown("**Parámetros de cálculo**")

            hsp_input5 = st.number_input("HSP (h/día)",
                                          min_value=0.5, max_value=12.0,
                                          value=float(hsp5) if hsp5 else 4.2, step=0.01,
                                          help="Horas Solar Pico del Módulo 4")
            pr_input5 = st.slider(
                "Performance Ratio — PR (%)", 65, 90, 75,
                help="PR típico OFF-GRID: 70–80%. Incluye pérdidas de temperatura, "
                     "inversor, sombreado, suciedad y cableado.",
                key="pr_offgrid")
            factor_bat5 = st.slider(
                "Factor baterías (OFF-GRID)", 120, 150, 130,
                help="Compensa pérdidas de carga/descarga. Típico: 120-150%. "
                     "P_FV = Consumo/(HSP×PR) × Factor_bat",
                key="factor_bat_offgrid")
            pot_panel_input5 = st.number_input("Potencia panel (Wp)",
                                                min_value=10, max_value=1000,
                                                value=int(pot_panel5) if pot_panel5 else 550,
                                                help="Potencia del panel del Módulo 5")

            if not hsp5:
                st.markdown("<div class='warn-box'>⚠ HSP no guardada. Usa el valor arriba o guarda desde Módulo 4</div>",
                            unsafe_allow_html=True)
            if not pot_panel5:
                st.markdown("<div class='warn-box'>⚠ Panel no configurado. Usa el valor arriba o guarda desde Módulo 5</div>",
                            unsafe_allow_html=True)

            pr_dec5             = pr_input5 / 100.0
            factor_bat5_dec     = factor_bat5 / 100.0
            consumo5_kwh        = consumo5 / 1000.0         # base sin FS
            hsp_efectiva5       = hsp_input5 * pr_dec5      # para mostrar en tabla
            # OFF-GRID: Pot_FV = (Consumo_kWh / (HSP × PR)) × Factor_baterías
            pot_base5_kw        = consumo5_kwh / (hsp_input5 * pr_dec5) if (hsp_input5 * pr_dec5) > 0 else 0
            potencia_instalada5 = pot_base5_kw * factor_bat5_dec * 1000
            num_pan5            = num_paneles(potencia_instalada5, pot_panel_input5)
            pot_real5           = num_pan5 * pot_panel_input5
            gen_dia5_kwh        = (pot_real5 / 1000) * hsp_input5 * pr_dec5

            # ── Guardar en session_state para Tabs 9, 10 y 11 ─────────────
            st.session_state["calc_num_paneles"]       = num_pan5
            st.session_state["calc_pot_panel_wp"]      = pot_panel_input5
            st.session_state["calc_hsp"]               = hsp_input5
            st.session_state["calc_pr"]                = pr_dec5
            st.session_state["calc_consumo_fs_wh"]     = consumo5_fs
            st.session_state["calc_potencia_inst_wp"]  = potencia_instalada5
            st.session_state["calc_pot_real_wp"]       = pot_real5
            st.session_state["calc_gen_dia_kwh"]       = gen_dia5_kwh
            st.session_state["calc_serie"]             = 1   # OFF-GRID no usa strings en serie
            st.session_state["calc_paralelo"]          = num_pan5

            st.markdown(f"""
            <hr class='sep'>
            <div style='color:#8A9BBD; font-size:0.8rem; margin-bottom:0.5rem;'>
                ({consumo5_kwh:.3f} kWh ÷ ({hsp_input5}h × {pr_input5}%)) × {factor_bat5}% =
                <b style='color:#FFD54F;'>{potencia_instalada5:,.0f} Wp mínimos</b>
            </div>
            """, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with col5b:
            st.markdown(f"""
            <div class='result-highlight'>
                <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase; letter-spacing:1px;'>
                    Pot. Mínima OFF-GRID (PR {pr_input5}% × Fact.bat {factor_bat5}%)</div>
                <div class='val'>{potencia_instalada5:,.0f} Wp</div>
            </div>
            <div class='metric-grid'>
                <div class='metric-box'>
                    <div class='metric-val'>{num_pan5}</div>
                    <div class='metric-unit'>paneles</div>
                    <div class='metric-label'>CANTIDAD REQUERIDA</div>
                </div>
                <div class='metric-box'>
                    <div class='metric-val'>{pot_panel_input5}</div>
                    <div class='metric-unit'>Wp/panel</div>
                    <div class='metric-label'>POTENCIA PANEL</div>
                </div>
                <div class='metric-box' style='border-color:rgba(0,230,118,0.4);'>
                    <div class='metric-val' style='color:#00E676;'>{pot_real5:,}</div>
                    <div class='metric-unit'>Wp</div>
                    <div class='metric-label'>POTENCIA REAL INSTALADA</div>
                </div>
                <div class='metric-box' style='border-color:rgba(0,188,212,0.4);'>
                    <div class='metric-val' style='color:#00BCD4;'>{gen_dia5_kwh:.2f}</div>
                    <div class='metric-unit'>kWh/día</div>
                    <div class='metric-label'>GENERACIÓN ESTIMADA</div>
                </div>
            </div>
            <div class='sol-card' style='margin-top:0.8rem;'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;font-weight:600;
                            margin-bottom:0.5rem;font-size:0.9rem;'>
                    DESGLOSE PÉRDIDAS (PR {pr_input5}%)</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.3rem 0;'>HSP bruta (PVGIS)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>{hsp_input5:.2f} h/día</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.3rem 0;'>Performance Ratio</td>
                        <td style='font-family:Share Tech Mono;color:#FF6B35;text-align:right;'>{pr_input5}%</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;background:#1A2235;'>
                        <td style='color:#FFB300;padding:0.3rem 0;font-weight:600;'>HSP efectiva</td>
                        <td style='font-family:Share Tech Mono;color:#FFB300;text-align:right;font-weight:700;'>{hsp_efectiva5:.2f} h/día</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.3rem 0;'>Consumo con FS 20%</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>{consumo5_fs:,.0f} Wh/día</td>
                    </tr>
                    <tr>
                        <td style='color:#8A9BBD;padding:0.3rem 0;'>Generación anual est.</td>
                        <td style='font-family:Share Tech Mono;color:#00E676;text-align:right;'>{gen_dia5_kwh*365:,.0f} kWh/año</td>
                    </tr>
                </table>
            </div>
            """, unsafe_allow_html=True)

            # ── Inversor — metodología técnica (igual a la usada en el resto del sistema)
            conn = get_conn()
            _cargas_t6 = pd.read_sql(
                "SELECT electrodomestico,cantidad,potencia_w,horas_dia,es_motor FROM cargas WHERE proyecto_id=?",
                conn, params=(proyecto_id,))
            conn.close()
            vdc5 = st.session_state.get("calc_vdc", tension_dc(consumo5_fs))
            _inv5 = calcular_inversor(_cargas_t6 if not _cargas_t6.empty else None,
                                      fs=0.80, fm=1.25, vdc=vdc5)
            if _inv5["inv_w"] == 0:
                _inv5_w_fb = consumo5_fs * 1.25
                _inv5["inv_w"]  = float(next(
                    (k*1000 for k in _KW_COMERCIALES if k*1000 >= _inv5_w_fb),
                    math.ceil(_inv5_w_fb/1000)*1000))
                _inv5["inv_kw"] = _inv5["inv_w"] / 1000
                _inv5["corr_dc"] = _inv5["inv_w"] / vdc5 if vdc5 > 0 else 0
            st.session_state["calc_inv_kw"]  = _inv5["inv_kw"]
            st.session_state["calc_inv_w"]   = _inv5["inv_w"]

            # Baterías estimadas (para la tabla resumen técnica)
            bats5 = calcular_baterias(consumo5_fs, vdc5, **_bat_params(st.session_state))

            st.markdown(f"""
            <div class='formula-box' style='font-size:0.78rem;margin-top:0.4rem;'>
                🔌 Inversor: P_inst={_inv5['pot_instalada']:,.0f}W × FS {int(_inv5['fs']*100)}%
                → P_sim={_inv5['pot_simultanea']:,.0f}W + Arr.motor={_inv5['pot_arranque']:,.0f}W
                → P_req={_inv5['pot_requerida']:,.0f}W × FM {int(_inv5['fm']*100)}%
                → <b>Comercial: {_inv5['inv_kw']:.1f} kW</b>
            </div>
            """, unsafe_allow_html=True)

            st.markdown(f"""
            <div style='display:grid;grid-template-columns:1fr 1fr;gap:0.6rem;margin-top:0.8rem;'>
              <div class='sol-card'>
                <div style='color:#00BCD4;font-family:Rajdhani,sans-serif;font-weight:600;margin-bottom:0.5rem;'>ARRAY FV</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>Paneles en paralelo</td><td style='font-family:Share Tech Mono;color:#FFB300;text-align:right;'>{num_pan5}</td></tr>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>Pot. real instalada</td><td style='font-family:Share Tech Mono;color:#00E676;text-align:right;'>{pot_real5:,} Wp</td></tr>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>Tensión DC sistema</td><td style='font-family:Share Tech Mono;color:#FFB300;text-align:right;'>{vdc5} V</td></tr>
                  <tr><td style='color:#8A9BBD;'>Gen. anual estimada</td><td style='font-family:Share Tech Mono;color:#00E676;text-align:right;'>{gen_dia5_kwh*365:,.0f} kWh/año</td></tr>
                </table>
              </div>
              <div class='sol-card'>
                <div style='color:#A78BFA;font-family:Rajdhani,sans-serif;font-weight:600;margin-bottom:0.5rem;'>BATERÍAS (estimado)</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>N° unidades (100Ah)</td><td style='font-family:Share Tech Mono;color:#A78BFA;text-align:right;'>{bats5['num_baterias']}</td></tr>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>Cap. banco</td><td style='font-family:Share Tech Mono;color:#A78BFA;text-align:right;'>{bats5['energia_kwh']:.2f} kWh</td></tr>
                  <tr style='border-bottom:1px solid #2A3A55;'><td style='color:#8A9BBD;'>Inversor recomendado</td><td style='font-family:Share Tech Mono;color:#00E676;text-align:right;'>{_inv5['inv_kw']:.1f} kW</td></tr>
                  <tr><td style='color:#8A9BBD;'>Corriente DC inversor</td><td style='font-family:Share Tech Mono;color:#FFB300;text-align:right;'>{_inv5['corr_dc']:.1f} A</td></tr>
                </table>
              </div>
            </div>
            """, unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 7 — BATERÍAS
# ════════════════════════════════════════════════════════════════════════════
with tab7:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>7</span> CÁLCULO DE BATERÍAS LITIO 100 Ah</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        Ah = (Consumo día Wh × Días autonomía) ÷ (Tensión DC (V) × DoD (%))<br>
        N° baterías = Ah requeridos ÷ Capacidad batería (Ah)
    </div>
    """, unsafe_allow_html=True)

    conn = get_conn()
    cargas6 = pd.read_sql("SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?",
                           conn, params=(proyecto_id,))
    p6 = conn.execute("SELECT tension_dc FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    conn.close()

    consumo_inv6 = (cargas6["cantidad"] * cargas6["potencia_w"] * cargas6["horas_dia"]).sum() if not cargas6.empty else 0.0
    consumo_rec6 = st.session_state.get("consumo_recibo_wh", 0.0)
    periodo_rec6 = st.session_state.get("recibo_ref_periodo", "")

    opciones_f6 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec6 > 0:
        opciones_f6.append(f"🧾 Recibo de energía ({periodo_rec6})")
        opciones_f6.append("📊 Mayor de los dos (recomendado)")

    fuente6 = st.radio("Base de consumo para dimensionar baterías:",
                        opciones_f6, horizontal=True, key="fuente_bat6")

    if "Recibo" in fuente6:
        consumo6 = consumo_rec6
    elif "Mayor" in fuente6:
        consumo6 = max(consumo_inv6, consumo_rec6)
    else:
        consumo6 = consumo_inv6

    if consumo6 == 0:
        st.markdown("<div class='warn-box'>⚠ Ingresa las cargas en el Módulo 1 o un recibo en el Módulo 2</div>", unsafe_allow_html=True)
    else:
        consumo6_fs = consumo6 * 1.20
        vdc6 = p6[0] if p6 and p6[0] else tension_dc(consumo6_fs)

        if consumo_rec6 > 0:
            st.markdown(f"""
            <div class='info-note' style='margin-bottom:0.8rem;'>
                Inventario: <b style='color:#FFD54F;'>{consumo_inv6:,.0f} Wh/día</b> &nbsp;|&nbsp;
                Recibo: <b style='color:#00BCD4;'>{consumo_rec6:,.0f} Wh/día</b> &nbsp;|&nbsp;
                <b>Usando: {consumo6:,.0f} Wh/día → {consumo6_fs:,.0f} Wh/día con FS 20%</b>
            </div>
            """, unsafe_allow_html=True)

        col6a, col6b = st.columns([1,1])
        with col6a:
            st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
            vdc_input6 = st.selectbox("Tensión DC del sistema (V)", [12, 24, 48],
                                       index=[12,24,48].index(vdc6) if vdc6 in [12,24,48] else 2)
            dod_input6 = st.slider("Profundidad de descarga — DoD (%)", 50, 100, 80,
                                    help="80% recomendado para baterías LiFePO4 | 50% para baterías AGM/GEL")
            eff_input6 = st.slider("Eficiencia del sistema — η (%)", 70, 98, 90,
                                    help="Incluye pérdidas por inversor (~5%), cableado (~2%), autodescarga (~3%). "
                                         "Típico: 85-92% para sistemas bien diseñados.")
            bat_cap6 = st.number_input("Capacidad por batería (Ah)", min_value=10, max_value=500, value=100)

            dod_dec   = dod_input6 / 100.0
            eff_dec   = eff_input6 / 100.0

            # Horas de autonomía — desde Tab 3 (recibo) si está configurado
            _horas_default6 = st.session_state.get("horas_autonomia_deseada", 24.0)
            horas_aut6 = st.number_input(
                "Horas de autonomía deseada (h)",
                min_value=1.0, max_value=120.0,
                value=float(_horas_default6),
                step=1.0,
                help="8h = solo nocturno | 24h = 1 día | 48h = 2 días | 72h = 3 días",
                key="horas_aut_bat6")
            dias_aut6 = horas_aut6 / 24.0

            # Fórmula 5 parámetros:
            # E_total = consumo × días / η
            # Ah_req  = E_total / (V × DoD)
            e_total6  = (consumo6_fs * dias_aut6) / max(eff_dec, 0.01)
            ah_req6   = e_total6 / (vdc_input6 * dod_dec)
            num_bat6  = math.ceil(ah_req6 / bat_cap6)
            if num_bat6 > 1 and num_bat6 % 2 != 0:
                num_bat6 += 1
            cap_real6      = num_bat6 * bat_cap6
            energia_real6  = cap_real6 * vdc_input6 / 1000
            autonomia_real6 = (cap_real6 * vdc_input6 * dod_dec * eff_dec) / max(consumo6_fs / 24, 0.001)

            # ── Guardar en session_state para Tabs 8, 9 y 10 ───────────────
            st.session_state["calc_num_baterias"]     = num_bat6
            st.session_state["calc_bat_cap_ah"]       = bat_cap6
            st.session_state["calc_ah_final"]         = ah_req6
            st.session_state["calc_vdc"]              = vdc_input6
            st.session_state["calc_dod_pct"]          = dod_input6
            st.session_state["calc_eff_pct"]          = eff_input6
            st.session_state["calc_horas_autonomia"]  = horas_aut6
            st.session_state["calc_dias_autonomia"]   = dias_aut6

            st.markdown("</div>", unsafe_allow_html=True)

        with col6b:
            st.markdown(f"""
            <div class='sol-card'>
            <div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; margin-bottom:1rem;'>PASO A PASO — 5 PARÁMETROS</div>
            <table style='width:100%; font-size:0.82rem; border-collapse:collapse;'>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>① Consumo diario (con FS 20%)</td>
                    <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{consumo6_fs:,.0f} Wh/día</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>② Días de autonomía</td>
                    <td style='font-family:Share Tech Mono; color:#00BCD4; text-align:right;'>{dias_aut6:.3f} días ({horas_aut6:.0f} h)</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>③ Voltaje banco</td>
                    <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{vdc_input6} V DC</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>④ DoD</td>
                    <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{dod_input6}% ({dod_dec:.2f})</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>⑤ Eficiencia sistema (η)</td>
                    <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{eff_input6}% ({eff_dec:.2f})</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55; background:#1A2235;'>
                    <td style='color:#FFB300; padding:0.35rem 0; font-weight:600;'>E_total = {consumo6_fs:,.0f} × {dias_aut6:.3f} ÷ {eff_dec:.2f}</td>
                    <td style='font-family:Share Tech Mono; color:#00E676; text-align:right; font-weight:700;'>{e_total6:,.1f} Wh</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55; background:#1A2235;'>
                    <td style='color:#FFB300; padding:0.35rem 0; font-weight:600;'>Ah = {e_total6:,.1f} ÷ ({vdc_input6} × {dod_dec:.2f})</td>
                    <td style='font-family:Share Tech Mono; color:#00E676; text-align:right; font-weight:700;'>{ah_req6:.1f} Ah</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>Capacidad por batería</td>
                    <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{bat_cap6} Ah</td>
                </tr>
                <tr style='border-bottom:1px solid #2A3A55;'>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>{ah_req6:.1f} ÷ {bat_cap6} → par más cercano</td>
                    <td style='font-family:Share Tech Mono; color:#00E676; text-align:right;'><b>{num_bat6} baterías</b></td>
                </tr>
                <tr>
                    <td style='color:#8A9BBD; padding:0.35rem 0;'>Autonomía real</td>
                    <td style='font-family:Share Tech Mono; color:#00BCD4; text-align:right;'>{autonomia_real6:.1f} h ({autonomia_real6/24:.2f} días)</td>
                </tr>
            </table>
            </div>
            """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class='result-highlight'>
            <div style='color:#8A9BBD; font-size:0.8rem; text-transform:uppercase; letter-spacing:1px;'>Ah requeridos (con DoD {dod_input6}% y η {eff_input6}%)</div>
            <div class='val'>{ah_req6:.0f} Ah @ {vdc_input6} V</div>
        </div>
        <div class='metric-grid'>
            <div class='metric-box' style='border-color:rgba(0,230,118,0.4);'>
                <div class='metric-val' style='color:#00E676;'>{num_bat6}</div>
                <div class='metric-unit'>unidades</div>
                <div class='metric-label'>BATERÍAS {bat_cap6} Ah</div>
            </div>
            <div class='metric-box'>
                <div class='metric-val'>{cap_real6:,}</div>
                <div class='metric-unit'>Ah</div>
                <div class='metric-label'>CAPACIDAD REAL INSTALADA</div>
            </div>
            <div class='metric-box'>
                <div class='metric-val'>{vdc_input6}</div>
                <div class='metric-unit'>V DC</div>
                <div class='metric-label'>TENSIÓN DEL BANCO</div>
            </div>
            <div class='metric-box'>
                <div class='metric-val'>{energia_real6:.2f}</div>
                <div class='metric-unit'>kWh</div>
                <div class='metric-label'>ENERGÍA BRUTA</div>
            </div>
            <div class='metric-box' style='border-color:rgba(0,188,212,0.4);'>
                <div class='metric-val' style='color:#00BCD4;'>{energia_real6 * dod_dec * eff_dec:.2f}</div>
                <div class='metric-unit'>kWh útiles</div>
                <div class='metric-label'>ENERGÍA UTILIZABLE</div>
            </div>
            <div class='metric-box' style='border-color:rgba(167,139,250,0.4);'>
                <div class='metric-val' style='color:#A78BFA;'>{autonomia_real6:.1f}</div>
                <div class='metric-unit'>horas ({autonomia_real6/24:.2f} días)</div>
                <div class='metric-label'>AUTONOMÍA REAL</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 8 — CONTROLADOR MPPT
# ════════════════════════════════════════════════════════════════════════════
with tab8:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>8</span> SELECCIÓN CONTROLADOR DE CARGA MPPT</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        Corriente controlador (A) = Isc (panel) × N° de paneles<br>
        Seleccionar el controlador estándar inmediatamente superior a ese valor
    </div>
    """, unsafe_allow_html=True)

    # Recopilar todos los datos calculados
    conn = get_conn()
    cargas7 = pd.read_sql("SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?",
                           conn, params=(proyecto_id,))
    p7 = conn.execute("SELECT hsp, tension_dc FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    panel7 = conn.execute("SELECT potencia_wp, voc, isc FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                           (proyecto_id,)).fetchone()
    recibo_res = pd.read_sql(
        "SELECT periodo, kwh_periodo, dias_periodo FROM recibos WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
        conn, params=(proyecto_id,))
    conn.close()

    consumo_inv7 = (cargas7["cantidad"] * cargas7["potencia_w"] * cargas7["horas_dia"]).sum() if not cargas7.empty else 0.0
    consumo_rec7 = st.session_state.get("consumo_recibo_wh", 0.0)
    periodo_rec7 = st.session_state.get("recibo_ref_periodo", "")

    opciones_f7 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec7 > 0:
        opciones_f7.append(f"🧾 Recibo de energía ({periodo_rec7})")
        opciones_f7.append("📊 Mayor de los dos (recomendado)")

    fuente7 = st.radio("Base de consumo para el resumen ejecutivo:",
                        opciones_f7, horizontal=True, key="fuente_ctrl7")

    if "Recibo" in fuente7:
        consumo7 = consumo_rec7
    elif "Mayor" in fuente7:
        consumo7 = max(consumo_inv7, consumo_rec7)
    else:
        consumo7 = consumo_inv7

    if consumo7 == 0:
        st.markdown("<div class='warn-box'>⚠ Ingresa las cargas en el Módulo 1 o un recibo en el Módulo 2</div>", unsafe_allow_html=True)
    else:
        consumo7_fs = consumo7 * 1.20
        vdc7 = p7[1] if p7 and p7[1] else tension_dc(consumo7_fs)
        hsp7 = p7[0] if p7 and p7[0] else 4.2
        pot_panel7 = panel7[0] if panel7 else 550

        if consumo_rec7 > 0:
            st.markdown(f"""
            <div class='info-note' style='margin-bottom:0.8rem;'>
                Inventario: <b style='color:#FFD54F;'>{consumo_inv7:,.0f} Wh/día</b> &nbsp;|&nbsp;
                Recibo: <b style='color:#00BCD4;'>{consumo_rec7:,.0f} Wh/día</b> &nbsp;|&nbsp;
                <b>Usando: {consumo7:,.0f} Wh/día → {consumo7_fs:,.0f} Wh/día con FS 20%</b>
            </div>
            """, unsafe_allow_html=True)

        # ── Usar datos ya calculados en Tab 6 (Potencia) si existen ────────
        # Esto evita recalcular con HSP/PR/factor genéricos y respeta los
        # ajustes que el usuario hizo en el Módulo 6, sin importar si la
        # fuente de consumo fue inventario de cargas o recibo de luz.
        _tab6_disponible = "calc_num_paneles" in st.session_state
        if _tab6_disponible:
            num_pan7         = st.session_state.get("calc_num_paneles", 0)
            pot_panel7       = st.session_state.get("calc_pot_panel_wp", pot_panel7)
            pot_real_paneles = st.session_state.get("calc_pot_real_wp", num_pan7 * pot_panel7)
            hsp7             = st.session_state.get("calc_hsp", hsp7)
            consumo7_fs      = st.session_state.get("calc_consumo_fs_wh", consumo7_fs)
            st.markdown("""
            <div class='info-note' style='margin-bottom:0.8rem;border-color:rgba(0,230,118,0.4);'>
                ✅ Usando paneles y potencia ya dimensionados en el <b>Módulo 6 · Potencia</b>
                (incluye tu ajuste de PR y factor de baterías)
            </div>
            """, unsafe_allow_html=True)
        else:
            # Fallback: recalcular si el usuario no ha pasado por Tab 6
            potencia_inst7   = consumo7_fs / hsp7
            num_pan7         = num_paneles(potencia_inst7, pot_panel7)
            pot_real_paneles = num_pan7 * pot_panel7
            st.markdown("""
            <div class='warn-box' style='margin-bottom:0.8rem;'>
                ⚠ No se encontraron datos del Módulo 6 · Potencia en esta sesión — se recalculó
                con HSP/PR por defecto. Visita el Módulo 6 para un dimensionamiento más preciso.
            </div>
            """, unsafe_allow_html=True)

        bats7 = calcular_baterias(consumo7_fs, vdc7, **_bat_params(st.session_state))
        # Isc del panel: priorizar el de la BD (Módulo 5), luego session_state
        isc7 = (panel7[2] if panel7 and len(panel7) > 2 and panel7[2] else
                st.session_state.get("calc_isc", st.session_state.get("panel_isc", 8.02)))
        ah_banco7 = bats7["num_baterias"] * 100
        corriente_mppt7 = isc7 * num_pan7   # Isc × N° paneles

        col7a, col7b = st.columns([1,1])
        with col7a:
            st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
            st.markdown("**Ajuste parámetros MPPT**")
            vdc_7 = st.selectbox("Tensión sistema (V)", [12, 24, 48],
                                   index=[12,24,48].index(vdc7) if vdc7 in [12,24,48] else 2, key="vdc7")
            isc_7 = st.number_input("Corriente Isc del panel (A)", min_value=1.0, max_value=30.0,
                                     value=float(isc7), step=0.01, key="isc7_ctrl",
                                     help="Corriente de cortocircuito del panel (del Módulo 5)")
            n_pan_7 = st.number_input("Número de paneles (N°)", min_value=1, max_value=500,
                                       value=int(num_pan7), step=1, key="npan7_ctrl")
            pot_paneles_7 = st.number_input("Potencia total paneles (Wp)", min_value=100, max_value=50000,
                                             value=int(pot_real_paneles), step=100)

            # 4. Corriente controlador = Isc × N_paneles
            corriente_ctrl = isc_7 * n_pan_7

            st.markdown(f"""
            <div class='formula-box' style='margin-top:1rem;'>
                Isc ({isc_7} A) × {n_pan_7} paneles = {corriente_ctrl:.1f} A → controlador estándar superior
            </div>
            """, unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)

        with col7b:
            # Determinar modelo MPPT sugerido
            if corriente_ctrl <= 40:
                mppt_modelo = "MPPT 40A"
                mppt_color = "#00E676"
            elif corriente_ctrl <= 60:
                mppt_modelo = "MPPT 60A"
                mppt_color = "#00BCD4"
            elif corriente_ctrl <= 100:
                mppt_modelo = "MPPT 100A"
                mppt_color = "#FFB300"
            else:
                mppt_modelo = f"MPPT {math.ceil(corriente_ctrl/50)*50}A"
                mppt_color = "#FF5252"
            # Guardar para el simulador
            st.session_state["calc_corr_mppt"]  = corriente_ctrl
            st.session_state["calc_mppt_modelo"] = mppt_modelo
            st.session_state["calc_isc"]         = isc_7

            st.markdown(f"""
            <div style='background:var(--card); border:1px solid var(--border); border-radius:12px; padding:1.5rem;'>
                <div style='color:#FFB300; font-family:Rajdhani,sans-serif; font-weight:600; font-size:1.1rem; margin-bottom:1rem;'>
                    ESPECIFICACIONES MPPT REQUERIDAS
                </div>
                <table style='width:100%; font-size:0.85rem; border-collapse:collapse;'>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD; padding:0.5rem 0;'>Tensión del sistema</td>
                        <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{vdc_7} V</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD; padding:0.5rem 0;'>Isc panel</td>
                        <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{isc_7} A</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD; padding:0.5rem 0;'>N° paneles</td>
                        <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{n_pan_7}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD; padding:0.5rem 0;'>Potencia mínima paneles</td>
                        <td style='font-family:Share Tech Mono; color:#FFD54F; text-align:right;'>{pot_paneles_7:,} Wp</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD; padding:0.5rem 0;'>Isc ({isc_7} A) × {n_pan_7} =</td>
                        <td style='font-family:Share Tech Mono; color:{mppt_color}; text-align:right; font-size:1.1rem; font-weight:700;'>{corriente_ctrl:.1f} A</td>
                    </tr>
                </table>
                <div style='margin-top:1rem; background:rgba(255,179,0,0.08); border:1px solid rgba(255,179,0,0.3); border-radius:8px; padding:1rem; text-align:center;'>
                    <div style='color:#8A9BBD; font-size:0.75rem; text-transform:uppercase; letter-spacing:1px;'>Controlador Recomendado</div>
                    <div style='font-family:Rajdhani,sans-serif; font-size:2rem; color:{mppt_color}; font-weight:700; margin:0.3rem 0;'>{mppt_modelo}</div>
                    <div style='font-size:0.8rem; color:#8A9BBD;'>
                        {vdc_7} V | {corriente_ctrl:.1f} A | ≥{pot_paneles_7:,} Wp
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        # ─── PROTECCIONES DC / AC ────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif;font-size:1.2rem;font-weight:700;
                    color:#FFB300;letter-spacing:1px;margin-bottom:0.8rem;'>
            🛡 PROTECCIONES ELÉCTRICAS DEL SISTEMA
        </div>""", unsafe_allow_html=True)

        _isc_prot = isc_7
        _n_str_prot = max(1, math.ceil(n_pan_7 / max(1, int(600 / voc_7) if (hasattr(st.session_state,"_paneles_por_string") and 0) else (int(600 / 49.8)))))
        try:
            _voc_prot = float(conn7b_voc) if False else float(panel7[1]) if panel7 and len(panel7) > 1 else 49.8
        except Exception:
            _voc_prot = 49.8

        _fus_str = _isc_prot * 1.25
        _fus_std = next((f for f in [10,15,20,25,30,40,50] if f >= _fus_str), 50)
        _corr_total_dc = _isc_prot * _n_str_prot
        _breaker_dc = _corr_total_dc * 1.25
        _breaker_dc_std = next((f for f in [20,25,32,40,50,63,80,100,125] if f >= _breaker_dc), 125)
        _pot_ac_w = pot_paneles_7
        _corr_ac = _pot_ac_w / 220 if _pot_ac_w > 0 else 20
        _breaker_ac = _corr_ac * 1.25
        _breaker_ac_std = next((f for f in [16,20,25,32,40,50,63] if f >= _breaker_ac), 63)

        col_prot1, col_prot2 = st.columns(2)
        with col_prot1:
            st.markdown(f"""
            <div class='sol-card'>
                <div style='color:#00BCD4;font-family:Rajdhani,sans-serif;font-weight:600;
                            margin-bottom:0.6rem;'>PROTECCIONES DC</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Fusible string (Isc x 1.25)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            {_isc_prot}A x 1.25 = {_fus_str:.1f}A<br>
                            <b style='color:#FFB300;'>Seleccionar: {_fus_std}A DC</b></td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Breaker DC general (Isc x N_str x 1.25)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            {_isc_prot}A x {_n_str_prot} str x 1.25 = {_breaker_dc:.1f}A<br>
                            <b style='color:#FFB300;'>Seleccionar: {_breaker_dc_std}A DC</b></td>
                    </tr>
                    <tr>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>DPS DC</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            Tipo II - 1000 VDC - 40 kA</td>
                    </tr>
                </table>
            </div>""", unsafe_allow_html=True)
        with col_prot2:
            st.markdown(f"""
            <div class='sol-card'>
                <div style='color:#00E676;font-family:Rajdhani,sans-serif;font-weight:600;
                            margin-bottom:0.6rem;'>PROTECCIONES AC</div>
                <table style='width:100%;font-size:0.8rem;border-collapse:collapse;'>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Breaker AC (P/220V x 1.25)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            {_pot_ac_w:,.0f}W / 220V = {_corr_ac:.1f}A x 1.25 = {_breaker_ac:.1f}A<br>
                            <b style='color:#00E676;'>Seleccionar: Breaker 2P {_breaker_ac_std}A</b></td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>DPS AC</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            Tipo II - 275V - 40 kA</td>
                    </tr>
                    <tr>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Puesta a tierra</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>
                            TT - conductor verde/amarillo</td>
                    </tr>
                </table>
            </div>""", unsafe_allow_html=True)

        # ─── RESUMEN FINAL ───────────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.4rem; font-weight:700; color:#FFB300; letter-spacing:2px; margin-bottom:1rem;'>
            ⚡ RESUMEN EJECUTIVO DEL SISTEMA FV
        </div>
        """, unsafe_allow_html=True)

        # ── Inversor desde metodología técnica ─────────────────────────────
        conn = get_conn()
        _cargas_r8 = pd.read_sql(
            "SELECT electrodomestico,cantidad,potencia_w,horas_dia,es_motor FROM cargas WHERE proyecto_id=?",
            conn, params=(proyecto_id,))
        conn.close()
        _inv8 = calcular_inversor(_cargas_r8 if not _cargas_r8.empty else None,
                                  fs=0.80, fm=1.25, vdc=vdc7)
        if _inv8["inv_w"] == 0:
            # Fallback cuando no hay inventario de cargas: usar consumo×FM
            _inv8_w_fallback = consumo7_fs * 1.25
            _inv8["inv_w"]  = float(next(
                (k*1000 for k in _KW_COMERCIALES if k*1000 >= _inv8_w_fallback),
                math.ceil(_inv8_w_fallback/1000)*1000))
            _inv8["inv_kw"] = _inv8["inv_w"] / 1000
            _inv8["corr_dc"] = _inv8["inv_w"] / vdc7 if vdc7 > 0 else 0
            _inv8["_fallback"] = True
        else:
            _inv8["_fallback"] = False

        # Recibo info banner
        if not recibo_res.empty:
            r0 = recibo_res.iloc[0]
            kwh_dia_r0 = r0["kwh_periodo"] / r0["dias_periodo"]
            st.markdown(f"""
            <div style='background:rgba(0,188,212,0.08); border:1px solid rgba(0,188,212,0.3);
                        border-radius:10px; padding:0.8rem 1.2rem; margin-bottom:1rem;
                        display:flex; gap:2.5rem; flex-wrap:wrap; align-items:center;'>
                <span style='font-size:1.4rem;'>🧾</span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Último recibo:
                    <b style='color:#00BCD4;'>{r0['periodo']}</b></span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Total período:
                    <b style='color:#FFD54F; font-family:Share Tech Mono;'>{r0['kwh_periodo']:.1f} kWh</b></span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Promedio diario:
                    <b style='color:#FFD54F; font-family:Share Tech Mono;'>{kwh_dia_r0:.2f} kWh/día
                    ({kwh_dia_r0*1000:,.0f} Wh/día)</b></span>
                <span style='font-size:0.82rem; color:#8A9BBD;'>Base usada:
                    <b style='color:#00E676;'>{consumo7:,.0f} Wh/día</b></span>
            </div>
            """, unsafe_allow_html=True)

        col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
        with col_r1:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#FFB300;'>
                <div style='font-size:1.5rem; margin-bottom:0.3rem;'>⚡</div>
                <div class='metric-val'>{consumo7_fs:,.0f}</div>
                <div class='metric-unit'>Wh/día</div>
                <div class='metric-label'>CONSUMO + 20% FS</div>
            </div>
            """, unsafe_allow_html=True)
        with col_r2:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#00BCD4;'>
                <div style='font-size:1.5rem; margin-bottom:0.3rem;'>🔆</div>
                <div class='metric-val'>{num_pan7}</div>
                <div class='metric-unit'>paneles {pot_panel7}Wp</div>
                <div class='metric-label'>CAMPO FOTOVOLTAICO</div>
            </div>
            """, unsafe_allow_html=True)
        with col_r3:
            st.markdown(f"""
            <div class='metric-box' style='border-color:#00E676;'>
                <div style='font-size:1.5rem; margin-bottom:0.3rem;'>🔋</div>
                <div class='metric-val'>{bats7["num_baterias"]}</div>
                <div class='metric-unit'>bat. 100Ah @ {vdc7}V</div>
                <div class='metric-label'>BANCO DE BATERÍAS</div>
            </div>
            """, unsafe_allow_html=True)
        with col_r4:
            st.markdown(f"""
            <div class='metric-box' style='border-color:{mppt_color};'>
                <div style='font-size:1.5rem; margin-bottom:0.3rem;'>🎛</div>
                <div class='metric-val'>{corriente_ctrl:.0f}</div>
                <div class='metric-unit'>A — {mppt_modelo}</div>
                <div class='metric-label'>CONTROLADOR MPPT</div>
            </div>
            """, unsafe_allow_html=True)
        with col_r5:
            st.markdown(f"""
            <div class='metric-box' style='border-color:rgba(255,107,53,0.6);'>
                <div style='font-size:1.5rem; margin-bottom:0.3rem;'>🔌</div>
                <div class='metric-val' style='color:#FF6B35;'>{_inv8['inv_kw']:.1f}</div>
                <div class='metric-unit'>kW / {_inv8['inv_w']:,.0f} W</div>
                <div class='metric-label'>INVERSOR REC.</div>
            </div>
            """, unsafe_allow_html=True)

        # Desglose técnico del inversor
        if _inv8["_fallback"]:
            _inv8_desc = f"Consumo {consumo7_fs:,.0f} Wh/día × FM 125% (sin inventario de cargas)"
        else:
            _inv8_desc = (
                f"P_inst={_inv8['pot_instalada']:,.0f}W · FS {int(_inv8['fs']*100)}%"
                f" → P_sim={_inv8['pot_simultanea']:,.0f}W"
                f" + Arr.motor={_inv8['pot_arranque']:,.0f}W"
                f" → P_req={_inv8['pot_requerida']:,.0f}W"
                f" × FM {int(_inv8['fm']*100)}%"
                f" → {_inv8['pot_inv_minima']:,.0f}W"
                f" → <b>Comercial: {_inv8['inv_kw']:.1f} kW</b>"
            )
        st.markdown(f"""
        <div class='formula-box' style='margin-top:0.6rem;font-size:0.78rem;'>
            🔌 Inversor (metodología técnica): {_inv8_desc}
            &nbsp;| Corriente DC @ {vdc7}V: <b>{_inv8['corr_dc']:.1f} A</b>
        </div>
        """, unsafe_allow_html=True)

        # Guardar resultado
        if st.button("💾 Guardar Resultados del Proyecto", use_container_width=True):
            conn = get_conn()
            conn.execute("""
                INSERT INTO resultados(proyecto_id, consumo_dia_wh, consumo_con_fs, tension_dc,
                    hsp, potencia_instalada_w, num_paneles, capacidad_baterias_ah, num_baterias, corriente_mppt)
                VALUES(?,?,?,?,?,?,?,?,?,?)
            """, (proyecto_id, consumo7, consumo7_fs, vdc7, hsp7, potencia_inst7,
                  num_pan7, bats7["ah_final"], bats7["num_baterias"], corriente_ctrl))
            conn.commit()
            conn.close()
            st.success("✓ Resultados guardados exitosamente")
            st.rerun()

        # Histórico
        conn = get_conn()
        hist = pd.read_sql(
            "SELECT generado, consumo_con_fs, tension_dc, hsp, num_paneles, num_baterias, corriente_mppt FROM resultados WHERE proyecto_id=? ORDER BY id DESC LIMIT 5",
            conn, params=(proyecto_id,))
        pinfo7 = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
        conn.close()
        if not hist.empty:
            st.markdown("**Últimos resultados guardados:**")
            hist.columns = ["Fecha","Consumo(Wh)","VDC","HSP(h)","Paneles","Baterías","MPPT(A)"]
            st.dataframe(hist, use_container_width=True, hide_index=True)

        # ── Descargas desde resumen ────────────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown("""
        <div style='font-family:Rajdhani,sans-serif; font-size:1.1rem; font-weight:600;
                    color:#FFB300; letter-spacing:1px; margin-bottom:0.8rem;'>
            📥 EXPORTAR INFORME COMPLETO
        </div>
        """, unsafe_allow_html=True)
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            try:
                xls7 = generar_excel(proyecto_id, pinfo7)
                st.download_button("⬇ Descargar Excel (.xlsx)",
                    data=xls7,
                    file_name=f"SolarCalc_{pinfo7[1].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_excel_t7")
            except Exception as ex:
                st.error(f"Error Excel: {ex}")
        with col_d2:
            try:
                pdf7 = generar_pdf(proyecto_id, pinfo7)
                st.download_button("⬇ Descargar PDF (.pdf)",
                    data=pdf7,
                    file_name=f"SolarCalc_{pinfo7[1].replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True, key="dl_pdf_t7")
            except Exception as ex:
                st.error(f"Error PDF: {ex}")

# ════════════════════════════════════════════════════════════════════════════
# TAB 9 — PLANO DE INSTALACIÓN DE PANELES
# ════════════════════════════════════════════════════════════════════════════
with tab9:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>9</span> PLANO DE INSTALACIÓN DE PANELES SOLARES</div>
    """, unsafe_allow_html=True)

    # ── Leer datos del proyecto ───────────────────────────────────────────────
    conn = get_conn()
    p9   = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    pan9 = conn.execute("SELECT * FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                         (proyecto_id,)).fetchone()
    cargas9 = pd.read_sql("SELECT cantidad,potencia_w,horas_dia FROM cargas WHERE proyecto_id=?",
                           conn, params=(proyecto_id,))
    recibos9 = pd.read_sql(
        "SELECT kwh_periodo,dias_periodo FROM recibos WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
        conn, params=(proyecto_id,))
    conn.close()

    # ── Calcular consumos de cada fuente ──────────────────────────────────────
    _hsp9_base  = p9[4] if p9 and p9[4] else 4.2
    _vdc9_base  = p9[3] if p9 and p9[3] else 48
    _pp9        = pan9[3] if pan9 else 550
    voc9        = pan9[4] if pan9 else 49.9
    isc9        = pan9[5] if pan9 else 14.0
    modelo9     = pan9[2] if pan9 else "Panel 550Wp"

    consumo_inv9 = (cargas9["cantidad"]*cargas9["potencia_w"]*cargas9["horas_dia"]).sum() \
                   if not cargas9.empty else 0.0
    if not recibos9.empty:
        r9 = recibos9.iloc[0]
        consumo_rec9 = r9["kwh_periodo"] / r9["dias_periodo"] * 1000
    else:
        consumo_rec9 = st.session_state.get("consumo_recibo_wh", 0.0)

    consumo_mayor9 = max(consumo_inv9, consumo_rec9)

    # Valores guardados por Tab 6 (pueden estar o no)
    calc_n_pan9  = st.session_state.get("calc_num_paneles")
    calc_pp9     = st.session_state.get("calc_pot_panel_wp",  _pp9)
    calc_hsp9    = st.session_state.get("calc_hsp",           _hsp9_base)
    calc_vdc9    = int(st.session_state.get("calc_vdc",       _vdc9_base))
    calc_fs9     = st.session_state.get("calc_consumo_fs_wh", 0.0)
    calc_pr9     = float(st.session_state.get("calc_pr", 0.75))   # PR de Tab 6

    # ── Selector de fuente para el plano ──────────────────────────────────────
    st.markdown("""
    <div style='font-family:Rajdhani,sans-serif;font-size:1rem;font-weight:600;
                color:#FFB300;letter-spacing:1px;margin-bottom:0.5rem;'>
        📐 BASE DE CÁLCULO PARA EL PLANO
    </div>""", unsafe_allow_html=True)

    fuentes9 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec9 > 0:
        fuentes9.append("🧾 Recibo de energía (Módulo 2)")
        fuentes9.append("📊 Mayor de los dos (recomendado)")
    if calc_n_pan9:
        fuentes9.append(f"🔆 Módulo 6 — Potencia ({calc_n_pan9} paneles ya calculados)")

    fuente9_sel = st.radio("Fuente:", fuentes9, horizontal=True, key="p9_fuente")

    # Calcular consumo y paneles según fuente seleccionada
    if "Módulo 6" in fuente9_sel and calc_n_pan9:
        consumo_base9 = calc_fs9 / 1.20 if calc_fs9 > 0 else consumo_mayor9
        consumo_fs9   = calc_fs9 if calc_fs9 > 0 else consumo_base9 * 1.20
        hsp9          = calc_hsp9
        vdc9          = calc_vdc9
        pot_panel9    = calc_pp9
        n_paneles9    = calc_n_pan9
        fuente_label9 = f"Módulo 6 · {calc_n_pan9} paneles de {int(calc_pp9)}Wp"
        fuente_color9 = "#00E676"
    elif "Recibo" in fuente9_sel and consumo_rec9 > 0:
        consumo_base9 = consumo_rec9
        consumo_fs9   = consumo_rec9 * 1.20
        hsp9          = calc_hsp9
        vdc9          = tension_dc(consumo_fs9)
        pot_panel9    = calc_pp9
        n_paneles9    = num_paneles(consumo_fs9 / (hsp9 * calc_pr9), pot_panel9) if hsp9 * calc_pr9 > 0 else 1
        fuente_label9 = f"Recibo · {consumo_rec9:,.0f} Wh/día → {n_paneles9} paneles"
        fuente_color9 = "#00BCD4"
    elif "Mayor" in fuente9_sel:
        consumo_base9 = consumo_mayor9
        consumo_fs9   = consumo_mayor9 * 1.20
        hsp9          = calc_hsp9
        vdc9          = tension_dc(consumo_fs9)
        pot_panel9    = calc_pp9
        n_paneles9    = num_paneles(consumo_fs9 / (hsp9 * calc_pr9), pot_panel9) if hsp9 * calc_pr9 > 0 else 1
        fuente_label9 = f"Mayor de los dos · {consumo_mayor9:,.0f} Wh/día → {n_paneles9} paneles"
        fuente_color9 = "#FFB300"
    else:  # Inventario
        consumo_base9 = consumo_inv9
        consumo_fs9   = consumo_inv9 * 1.20
        hsp9          = calc_hsp9
        vdc9          = tension_dc(consumo_fs9)
        pot_panel9    = calc_pp9
        n_paneles9    = num_paneles(consumo_fs9 / (hsp9 * calc_pr9), pot_panel9) if hsp9 * calc_pr9 > 0 else 1
        fuente_label9 = f"Inventario · {consumo_inv9:,.0f} Wh/día → {n_paneles9} paneles"
        fuente_color9 = "#FFB300"

    # Garantizar valores mínimos válidos para widgets
    n_paneles9 = max(1, int(n_paneles9))
    hsp9       = max(0.5, float(hsp9))
    vdc9       = int(vdc9) if vdc9 in (12, 24, 48) else 48
    pot_panel9 = max(10, float(pot_panel9))

    # Banner de fuente activa
    st.markdown(f"""
    <div style='background:rgba(0,0,0,0.2);border:1px solid {fuente_color9}55;
                border-left:4px solid {fuente_color9};border-radius:8px;
                padding:0.7rem 1rem;margin-bottom:1rem;
                display:flex;gap:2rem;flex-wrap:wrap;align-items:center;'>
        <span style='color:{fuente_color9};font-family:Rajdhani,sans-serif;
                     font-weight:700;font-size:0.95rem;'>{fuente_label9}</span>
        <span style='font-size:0.8rem;color:#8A9BBD;'>
            Consumo base: <b style='color:#FFD54F;'>{consumo_base9:,.0f} Wh/día</b>
            → con FS 20%: <b style='color:#FFD54F;'>{consumo_fs9:,.0f} Wh/día</b>
            | HSP: <b style='color:#00BCD4;'>{hsp9} h</b>
            | PR: <b style='color:#FF6B35;'>{int(calc_pr9*100)}%</b>
            | VDC: <b style='color:#00BCD4;'>{vdc9} V</b>
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Comparativa de fuentes disponibles
    with st.expander("📊 Ver comparativa de fuentes disponibles"):
        src_cols = st.columns(3)
        fuentes_comp = [
            ("⚡ Inventario", consumo_inv9,  "#FFB300"),
            ("🧾 Recibo",     consumo_rec9,  "#00BCD4"),
            ("📊 Mayor",      consumo_mayor9,"#00E676"),
        ]
        for col_c, (lbl_c, cons_c, clr_c) in zip(src_cols, fuentes_comp):
            if cons_c > 0:
                n_c = num_paneles(cons_c * 1.20 / (hsp9 * calc_pr9), pot_panel9) if hsp9 * calc_pr9 > 0 else 0
                col_c.markdown(f"""
                <div style='background:#1A2235;border:1px solid {clr_c}44;
                            border-radius:8px;padding:0.8rem;text-align:center;'>
                    <div style='font-size:0.75rem;color:#8A9BBD;'>{lbl_c}</div>
                    <div style='font-family:Share Tech Mono,monospace;
                                font-size:1.3rem;color:{clr_c};'>{cons_c:,.0f}</div>
                    <div style='font-size:0.7rem;color:#8A9BBD;'>Wh/día</div>
                    <div style='font-family:Share Tech Mono,monospace;
                                font-size:1rem;color:#FFD54F;margin-top:0.3rem;'>
                        {n_c} paneles</div>
                </div>""", unsafe_allow_html=True)
            else:
                col_c.markdown(f"""
                <div style='background:#161D30;border:1px dashed #2A3A55;
                            border-radius:8px;padding:0.8rem;text-align:center;
                            color:#2A3A55;font-size:0.8rem;'>{lbl_c}<br>Sin datos</div>""",
                    unsafe_allow_html=True)

    st.markdown("""
    <div class='info-note' style='margin-bottom:1.2rem;'>
        📐 Configura la distribución física de los paneles en la cubierta. El plano se genera
        automáticamente con la disposición en serie/paralelo requerida para el sistema.
    </div>
    """, unsafe_allow_html=True)

    # ── Parámetros de configuración ────────────────────────────────────────
    col9a, col9b, col9c = st.columns(3)
    with col9a:
        st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
        st.markdown("**🔆 Configuración del Array**")
        pan_ancho  = st.number_input("Ancho del panel (mm)", 1000, 2500, 1134, 10, key="p9_w")
        pan_alto   = st.number_input("Alto del panel (mm)",  900, 3000, 2278, 10, key="p9_h")
        separacion = st.number_input("Separación entre paneles (mm)", 20, 300, 50, 10, key="p9_sep")
        orientacion= st.radio("Orientación del panel", ["Portrait (vertical)", "Landscape (horizontal)"],
                               key="p9_ori", horizontal=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col9b:
        st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
        st.markdown("**⚡ Serie / Paralelo**")

        vmp_aprox  = voc9 * 0.80
        serie_auto = max(1, round(vdc9 / vmp_aprox)) if vmp_aprox > 0 else 1
        serie_auto = max(1, min(int(serie_auto), 20))   # clamp dentro de [1, 20]
        paralelo_auto = math.ceil(n_paneles9 / serie_auto) if n_paneles9 > 0 else 1
        paralelo_auto = max(1, min(int(paralelo_auto), 50))  # clamp dentro de [1, 50]

        st.markdown(f"""
        <div style='background:#161D30; border:1px solid rgba(255,179,0,0.3); border-radius:6px;
                    padding:0.5rem 0.8rem; margin-bottom:0.8rem; font-size:0.8rem;'>
            <span style='color:#8A9BBD;'>Paneles requeridos: </span>
            <b style='color:#FFB300; font-family:Share Tech Mono,monospace;'>{n_paneles9} × {int(pot_panel9)}Wp</b>
            <span style='color:#8A9BBD;'> | Vdc: </span>
            <b style='color:#00BCD4; font-family:Share Tech Mono,monospace;'>{vdc9}V</b>
        </div>
        """, unsafe_allow_html=True)

        serie_sel    = st.number_input("Paneles en serie (Ns)", 1, 20, serie_auto, key="p9_ns",
                                        help=f"Auto: {serie_auto} (Vdc={vdc9}V ÷ Vmp≈{vmp_aprox:.1f}V)")
        paralelo_sel = st.number_input("Strings en paralelo (Np)", 1, 50, paralelo_auto, key="p9_np",
                                        help=f"Auto: {paralelo_auto} ({n_paneles9} paneles ÷ {serie_auto} serie)")
        total_array  = serie_sel * paralelo_sel

        v_array   = round(voc9 * serie_sel, 1)
        i_array   = round(isc9 * paralelo_sel, 1)
        pot_array = round(pot_panel9 * total_array / 1000, 2)

        if total_array != n_paneles9:
            diff = total_array - n_paneles9
            diff_label = f"+{diff}" if diff > 0 else str(diff)
            diff_color = "#FF5252" if diff < 0 else "#FFD54F"
            st.markdown(f"""
            <div style='background:rgba(255,179,0,0.06); border:1px dashed rgba(255,179,0,0.3);
                        border-radius:6px; padding:0.4rem 0.8rem; font-size:0.78rem; color:#8A9BBD;'>
                Array: <b style='color:{diff_color};'>{total_array} paneles</b>
                vs requeridos: <b style='color:#FFB300;'>{n_paneles9}</b>
                (<span style='color:{diff_color};'>{diff_label}</span>)
            </div>
            """, unsafe_allow_html=True)

        st.markdown(f"""
        <div style='margin-top:0.6rem;'>
            <div class='formula-box'>
                Ns={serie_sel} series × Np={paralelo_sel} strings = <b>{total_array} paneles</b><br>
                Vtotal = {voc9}V × {serie_sel} = <b>{v_array}V</b><br>
                Itotal = {isc9}A × {paralelo_sel} = <b>{i_array}A</b><br>
                Pot. array = <b>{pot_array} kWp</b>
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with col9c:
        st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
        st.markdown("**🏠 Datos de la Cubierta**")
        techo_tipo   = st.selectbox("Tipo de cubierta", ["Techo plano", "Techo inclinado",
                                     "Estructura metálica", "Suelo / Campo abierto"], key="p9_techo")
        techo_inclin = st.number_input("Inclinación (°)", 0, 90, 15, key="p9_incl")
        techo_azimut = st.selectbox("Orientación (azimut)", ["Norte (0°)", "Noreste (45°)",
                                     "Este (90°)", "Sureste (135°)", "Sur (180°)",
                                     "Suroeste (225°)", "Oeste (270°)", "Noroeste (315°)"],
                                     index=4, key="p9_azim")
        techo_color  = st.color_picker("Color cubierta", "#2A3A55", key="p9_color")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<hr class='sep'>", unsafe_allow_html=True)

    # ── GENERACIÓN DEL PLANO SVG ────────────────────────────────────────────
    if orientacion == "Portrait (vertical)":
        pw_disp, ph_disp = pan_ancho, pan_alto
    else:
        pw_disp, ph_disp = pan_alto, pan_ancho

    ESCALA    = 0.09        # mm → px  (1134mm → ~102px)
    PW        = pw_disp  * ESCALA
    PH        = ph_disp  * ESCALA
    SEP       = separacion * ESCALA
    MARGIN    = 60
    COLS      = serie_sel
    ROWS      = paralelo_sel

    svg_w = int(COLS * PW + (COLS - 1) * SEP + 2 * MARGIN + 180)
    svg_h = int(ROWS * PH + (ROWS - 1) * SEP + 2 * MARGIN + 120)
    svg_h = max(svg_h, 420)

    # Paleta
    C_BG     = "#0A0E1A"
    C_ROOF   = techo_color
    C_PANEL  = "#1565C0"
    C_PANEL2 = "#1976D2"
    C_CELL   = "#0D47A1"
    C_WIRE   = "#FFB300"
    C_WIRE2  = "#FF5252"
    C_TEXT   = "#E8EDF5"
    C_DIM    = "#8A9BBD"
    C_FRAME  = "#FFB300"

    # Dimensiones cubierta con holgura
    roof_x = MARGIN - 20
    roof_y = MARGIN - 20
    roof_w = COLS * PW + (COLS - 1) * SEP + 40
    roof_h = ROWS * PH + (ROWS - 1) * SEP + 40

    # Build panel cell pattern (3 cols × 4 rows inside each panel)
    def panel_svg(x, y, pw, ph, idx):
        cx, cy = pw / 16, ph / 12
        cells  = ""
        for ci in range(4):
            for ri in range(3):
                cx0 = x + 4 + ci * (pw - 8) / 4
                cy0 = y + 4 + ri * (ph - 8) / 3
                cw0 = (pw - 8) / 4 - 1.5
                ch0 = (ph - 8) / 3 - 1.5
                cells += (f'<rect x="{cx0:.1f}" y="{cy0:.1f}" width="{cw0:.1f}" '
                          f'height="{ch0:.1f}" rx="1" fill="{C_CELL}" opacity="0.85"/>')
        # Junction box
        jbx = x + pw/2 - 6; jby = y + ph - 12
        label_num = f'<text x="{x+pw/2:.1f}" y="{y+ph/2+4:.1f}" text-anchor="middle" '
        label_num += f'font-family="Share Tech Mono,monospace" font-size="7" fill="#90CAF9" font-weight="bold">{idx}</text>'
        return f'''
        <g class="panel">
          <rect x="{x:.1f}" y="{y:.1f}" width="{pw:.1f}" height="{ph:.1f}"
                rx="3" fill="{C_PANEL}" stroke="{C_FRAME}" stroke-width="1.2"/>
          <rect x="{x+2:.1f}" y="{y+2:.1f}" width="{pw-4:.1f}" height="{ph-4:.1f}"
                rx="2" fill="none" stroke="#42A5F5" stroke-width="0.5" opacity="0.5"/>
          {cells}
          <rect x="{jbx:.1f}" y="{jby:.1f}" width="12" height="6" rx="1"
                fill="#263238" stroke="#FFB300" stroke-width="0.8"/>
          {label_num}
        </g>'''

    # Build all panels + wiring
    panels_svg  = ""
    wires_svg   = ""
    labels_svg  = ""
    panel_count = 0
    jbox_points = []   # (x_center, y_bottom) for each panel

    for row in range(ROWS):
        row_points = []
        for col in range(COLS):
            panel_count += 1
            px = MARGIN + col * (PW + SEP)
            py = MARGIN + row * (PH + SEP)
            panels_svg += panel_svg(px, py, PW, PH, panel_count)
            cx_j = px + PW / 2
            cy_j = py + PH
            row_points.append((cx_j, cy_j, px, py))

        # Wire in series inside string (horizontal bottom)
        for k in range(len(row_points) - 1):
            x1 = row_points[k][0];  y1 = row_points[k][1] - 3
            x2 = row_points[k+1][0]; y2 = y1
            wires_svg += (f'<line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" '
                          f'stroke="{C_WIRE}" stroke-width="2" stroke-dasharray="4,2"/>')

        # String label
        lx = MARGIN + COLS * (PW + SEP) + 6
        ly = MARGIN + row * (PH + SEP) + PH / 2
        labels_svg += (f'<text x="{lx:.1f}" y="{ly:.1f}" font-family="Rajdhani,sans-serif" '
                       f'font-size="9" fill="{C_DIM}">String {row+1}</text>')
        jbox_points.append((row_points[0][0], row_points[0][2]))   # leftmost panel of string

    # Wire strings in parallel (vertical left side)
    if len(jbox_points) > 1:
        vx = MARGIN - 18
        for idx, (_, px) in enumerate(jbox_points):
            py_top  = MARGIN + idx * (PH + SEP) + PH * 0.3
            py_bot  = MARGIN + idx * (PH + SEP) + PH * 0.7
            wires_svg += (f'<line x1="{px:.1f}" y1="{py_top:.1f}" x2="{vx:.1f}" y2="{py_top:.1f}" '
                          f'stroke="{C_WIRE2}" stroke-width="1.5"/>')
            wires_svg += (f'<line x1="{px:.1f}" y1="{py_bot:.1f}" x2="{vx:.1f}" y2="{py_bot:.1f}" '
                          f'stroke="{C_WIRE}" stroke-width="1.5"/>')
        # Vertical bus
        py_first_top = MARGIN + 0 * (PH + SEP) + PH * 0.3
        py_last_top  = MARGIN + (ROWS - 1) * (PH + SEP) + PH * 0.3
        py_first_bot = MARGIN + 0 * (PH + SEP) + PH * 0.7
        py_last_bot  = MARGIN + (ROWS - 1) * (PH + SEP) + PH * 0.7
        wires_svg += (f'<line x1="{vx}" y1="{py_first_top:.1f}" x2="{vx}" y2="{py_last_top:.1f}" '
                      f'stroke="{C_WIRE2}" stroke-width="2.5"/>')
        wires_svg += (f'<line x1="{vx-5}" y1="{py_first_bot:.1f}" x2="{vx-5}" y2="{py_last_bot:.1f}" '
                      f'stroke="{C_WIRE}" stroke-width="2.5"/>')
        # Bus to MPPT box
        mppt_x = vx - 22; mppt_y = int((py_first_top + py_last_top) / 2) - 24
        wires_svg += (f'<line x1="{vx}" y1="{(py_first_top+py_last_top)/2:.1f}" '
                      f'x2="{mppt_x+22:.1f}" y2="{mppt_y+12:.1f}" '
                      f'stroke="{C_WIRE2}" stroke-width="2"/>') 
        wires_svg += (f'<line x1="{vx-5}" y1="{(py_first_bot+py_last_bot)/2:.1f}" '
                      f'x2="{mppt_x+22:.1f}" y2="{mppt_y+24:.1f}" '
                      f'stroke="{C_WIRE}" stroke-width="2"/>')
        # MPPT box
        wires_svg += f'''
        <g>
          <rect x="{mppt_x-30}" y="{mppt_y}" width="52" height="36" rx="4"
                fill="#1A2235" stroke="{C_WIRE}" stroke-width="1.5"/>
          <text x="{mppt_x-4}" y="{mppt_y+14}" text-anchor="middle"
                font-family="Rajdhani,sans-serif" font-size="8" fill="{C_WIRE}" font-weight="700">MPPT</text>
          <text x="{mppt_x-4}" y="{mppt_y+26}" text-anchor="middle"
                font-family="Share Tech Mono,monospace" font-size="7" fill="{C_DIM}">{v_array}V/{i_array}A</text>
        </g>'''

    # Dimension arrows
    dim_y   = MARGIN + ROWS * (PH + SEP) + 16
    dim_x   = MARGIN + COLS * (PW + SEP) + 14
    total_w = COLS * PW + (COLS - 1) * SEP
    total_h = ROWS * PH + (ROWS - 1) * SEP
    dims_svg = f'''
    <line x1="{MARGIN:.1f}" y1="{dim_y:.1f}" x2="{MARGIN+total_w:.1f}" y2="{dim_y:.1f}"
          stroke="{C_DIM}" stroke-width="0.8" marker-end="url(#arr)" marker-start="url(#arr)"/>
    <text x="{MARGIN+total_w/2:.1f}" y="{dim_y+11:.1f}" text-anchor="middle"
          font-family="Share Tech Mono,monospace" font-size="8.5" fill="{C_DIM}">
        {total_w/ESCALA/1000:.2f} m</text>
    <line x1="{dim_x:.1f}" y1="{MARGIN:.1f}" x2="{dim_x:.1f}" y2="{MARGIN+total_h:.1f}"
          stroke="{C_DIM}" stroke-width="0.8" marker-end="url(#arr)" marker-start="url(#arr)"/>
    <text x="{dim_x+4:.1f}" y="{MARGIN+total_h/2:.1f}" text-anchor="start"
          font-family="Share Tech Mono,monospace" font-size="8.5" fill="{C_DIM}">{total_h/ESCALA/1000:.2f} m</text>
    '''

    # Compass rose
    cx_comp = svg_w - 44; cy_comp = 38
    compass_svg = f'''
    <g>
      <circle cx="{cx_comp}" cy="{cy_comp}" r="18" fill="#161D30" stroke="#2A3A55" stroke-width="1"/>
      <polygon points="{cx_comp},{cy_comp-16} {cx_comp-5},{cy_comp+4} {cx_comp},{cy_comp-2} {cx_comp+5},{cy_comp+4}"
               fill="{C_WIRE2}"/>
      <polygon points="{cx_comp},{cy_comp+16} {cx_comp-5},{cy_comp-4} {cx_comp},{cy_comp+2} {cx_comp+5},{cy_comp-4}"
               fill="{C_DIM}" opacity="0.7"/>
      <text x="{cx_comp}" y="{cy_comp-20}" text-anchor="middle"
            font-family="Rajdhani,sans-serif" font-size="8" fill="{C_WIRE2}" font-weight="700">N</text>
    </g>'''

    # Title block
    title_y = svg_h - 68
    def _svg_safe(val, default="—"):
        return str(val if val is not None else default).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    proyecto_nombre  = _svg_safe(p9[1] if p9 else None, "Proyecto")
    municipio_nombre = _svg_safe(p9[2] if p9 else None, "—")
    modelo9_safe     = _svg_safe(modelo9, "Panel")
    title_svg = f'''
    <rect x="0" y="{title_y}" width="{svg_w}" height="68" fill="#0F1525" stroke="#2A3A55" stroke-width="0.8"/>
    <line x1="0" y1="{title_y}" x2="{svg_w}" y2="{title_y}" stroke="{C_WIRE}" stroke-width="1.5"/>
    <text x="12" y="{title_y+18}" font-family="Rajdhani,sans-serif" font-size="13"
          fill="{C_WIRE}" font-weight="700">☀ PLANO DE INSTALACIÓN DE PANELES SOLARES</text>
    <text x="12" y="{title_y+33}" font-family="Rajdhani,sans-serif" font-size="9" fill="{C_TEXT}">
        Proyecto: {proyecto_nombre}  |  Ubicación: {municipio_nombre}  |  {techo_tipo} — {techo_inclin}° — {techo_azimut}</text>
    <text x="12" y="{title_y+47}" font-family="Share Tech Mono,monospace" font-size="8.5" fill="{C_DIM}">
        Panel: {modelo9_safe}  {pot_panel9}Wp  |  Array: {serie_sel}S × {paralelo_sel}P = {total_array} paneles  |
        Varray={v_array}V  Iarray={i_array}A  |  Pot.={pot_array}kWp</text>
    <text x="12" y="{title_y+60}" font-family="Share Tech Mono,monospace" font-size="8" fill="#2A3A55">
        Escala aprox. 1:{int(1/ESCALA*10)}  |  SolarCalc Pro  |  {datetime.now().strftime("%d/%m/%Y")}</text>
    <text x="{svg_w-12}" y="{title_y+18}" text-anchor="end"
          font-family="Share Tech Mono,monospace" font-size="9" fill="{C_DIM}">
        Plano N° 01  Rev.A</text>
    '''

    svg_code = f'''<svg viewBox="0 0 {svg_w} {svg_h}" xmlns="http://www.w3.org/2000/svg"
         style="background:{C_BG}; border-radius:12px; width:100%; max-height:700px;">
      <defs>
        <marker id="arr" markerWidth="6" markerHeight="6" refX="3" refY="3" orient="auto">
          <path d="M0,0 L6,3 L0,6 Z" fill="{C_DIM}"/>
        </marker>
        <filter id="glow">
          <feGaussianBlur stdDeviation="2" result="blur"/>
          <feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge>
        </filter>
      </defs>
      <!-- Background grid -->
      <defs>
        <pattern id="grid" width="20" height="20" patternUnits="userSpaceOnUse">
          <path d="M 20 0 L 0 0 0 20" fill="none" stroke="#1A2235" stroke-width="0.4"/>
        </pattern>
      </defs>
      <rect width="{svg_w}" height="{svg_h-68}" fill="url(#grid)"/>
      <!-- Roof/Surface -->
      <rect x="{roof_x:.1f}" y="{roof_y:.1f}" width="{roof_w:.1f}" height="{roof_h:.1f}"
            rx="6" fill="{C_ROOF}" opacity="0.25" stroke="{C_ROOF}" stroke-width="1.5"
            stroke-dasharray="6,3"/>
      <text x="{roof_x+8:.1f}" y="{roof_y+14:.1f}" font-family="Rajdhani,sans-serif"
            font-size="9" fill="{C_ROOF}" opacity="0.8">{techo_tipo.upper()}</text>
      <!-- Wiring -->
      {wires_svg}
      <!-- Panels -->
      {panels_svg}
      <!-- Labels -->
      {labels_svg}
      <!-- Dimensions -->
      {dims_svg}
      <!-- Compass -->
      {compass_svg}
      <!-- Title block -->
      {title_svg}
    </svg>'''

    render_svg(svg_code, height=max(svg_h, 480))

    # ── Leyenda y descarga ────────────────────────────────────────────────
    st.markdown("<hr class='sep'>", unsafe_allow_html=True)
    col9_leg1, col9_leg2, col9_leg3 = st.columns(3)
    with col9_leg1:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>LEYENDA</div>
            <div style='font-size:0.82rem; line-height:2;'>
                <span style='color:#1976D2;'>█</span> Panel solar {pot_panel9}Wp<br>
                <span style='color:{C_WIRE};'>━ ━</span> Cable positivo (+) serie<br>
                <span style='color:{C_WIRE2};'>━━</span> Cable negativo (−) string<br>
                <span style='color:{C_DIM};'>- - -</span> Límite cubierta<br>
                <span style='color:#FFB300;'>□</span> Caja de conexiones MPPT
            </div>
        </div>""", unsafe_allow_html=True)
    with col9_leg2:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>RESUMEN DEL ARRAY</div>
            <div style='font-size:0.82rem; line-height:2; font-family:Share Tech Mono,monospace;'>
                Paneles totales: <b style='color:#FFD54F;'>{total_array}</b><br>
                Configuración: <b style='color:#FFD54F;'>{serie_sel}S × {paralelo_sel}P</b><br>
                Vtotal array: <b style='color:#FFD54F;'>{v_array} V</b><br>
                Itotal array: <b style='color:#FFD54F;'>{i_array} A</b><br>
                Potencia total: <b style='color:#00E676;'>{pot_array} kWp</b>
            </div>
        </div>""", unsafe_allow_html=True)
    with col9_leg3:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>DIMENSIONES FÍSICAS</div>
            <div style='font-size:0.82rem; line-height:2; font-family:Share Tech Mono,monospace;'>
                Panel: <b style='color:#FFD54F;'>{pw_disp}×{ph_disp} mm</b><br>
                Ancho array: <b style='color:#FFD54F;'>{total_w/ESCALA/1000:.2f} m</b><br>
                Alto array: <b style='color:#FFD54F;'>{total_h/ESCALA/1000:.2f} m</b><br>
                Área neta: <b style='color:#FFD54F;'>{(total_w/ESCALA/1000)*(total_h/ESCALA/1000):.1f} m²</b><br>
                Inclinación: <b style='color:#00E676;'>{techo_inclin}° — {techo_azimut}</b>
            </div>
        </div>""", unsafe_allow_html=True)

    # Download SVG
    fname9 = f"Plano_Paneles_{(p9[1] if p9 else 'Proyecto').replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.svg"
    st.download_button("⬇ Descargar Plano de Paneles (SVG)",
                       data=svg_code.encode(),
                       file_name=fname9,
                       mime="image/svg+xml",
                       use_container_width=True, key="dl_svg_paneles")


# ════════════════════════════════════════════════════════════════════════════
# TAB 10 — PLANO GENERAL DEL PROYECTO
# ════════════════════════════════════════════════════════════════════════════
with tab10:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>10</span> PLANO GENERAL DEL SISTEMA FOTOVOLTAICO</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='info-note' style='margin-bottom:1.2rem;'>
        📐 Diagrama unifilar y de bloques del sistema completo: paneles → controlador MPPT →
        banco de baterías → inversor → cargas AC/DC. Incluye simbología eléctrica estándar.
    </div>
    """, unsafe_allow_html=True)

    # ── Leer datos del proyecto ───────────────────────────────────────────────
    conn = get_conn()
    p10   = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    pan10 = conn.execute("SELECT * FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
                          (proyecto_id,)).fetchone()
    cargas10 = pd.read_sql(
        "SELECT electrodomestico,cantidad,potencia_w,horas_dia,es_motor FROM cargas WHERE proyecto_id=?",
        conn, params=(proyecto_id,))
    recibos10 = pd.read_sql(
        "SELECT kwh_periodo,dias_periodo FROM recibos WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
        conn, params=(proyecto_id,))
    conn.close()

    # ── Calcular consumos de cada fuente ──────────────────────────────────────
    _hsp10_base = p10[4] if p10 and p10[4] else 4.2
    _vdc10_base = p10[3] if p10 and p10[3] else 48
    _pp10       = pan10[3] if pan10 else 550
    voc10       = pan10[4] if pan10 else 49.9
    isc10       = pan10[5] if pan10 else 14.0
    modelo10    = pan10[2] if pan10 else "Panel 550Wp"
    bat_cap10   = int(st.session_state.get("calc_bat_cap_ah", 100))

    consumo_inv10 = (cargas10["cantidad"]*cargas10["potencia_w"]*cargas10["horas_dia"]).sum() \
                    if not cargas10.empty else 0.0
    if not recibos10.empty:
        r10 = recibos10.iloc[0]
        consumo_rec10 = r10["kwh_periodo"] / r10["dias_periodo"] * 1000
    else:
        consumo_rec10 = st.session_state.get("consumo_recibo_wh", 0.0)

    consumo_mayor10 = max(consumo_inv10, consumo_rec10)

    # Valores guardados por módulos 6 y 7
    calc_n_pan10  = st.session_state.get("calc_num_paneles")
    calc_pp10     = st.session_state.get("calc_pot_panel_wp",  _pp10)
    calc_hsp10    = st.session_state.get("calc_hsp",           _hsp10_base)
    calc_vdc10    = int(st.session_state.get("calc_vdc",       _vdc10_base))
    calc_fs10     = st.session_state.get("calc_consumo_fs_wh", 0.0)
    calc_n_bat10  = st.session_state.get("calc_num_baterias")
    calc_pr10     = float(st.session_state.get("calc_pr", 0.75))   # PR de Tab 6

    # ── Selector de fuente ────────────────────────────────────────────────────
    st.markdown("""
    <div style='font-family:Rajdhani,sans-serif;font-size:1rem;font-weight:600;
                color:#FFB300;letter-spacing:1px;margin-bottom:0.5rem;'>
        📐 BASE DE CÁLCULO PARA EL PLANO GENERAL
    </div>""", unsafe_allow_html=True)

    fuentes10 = ["⚡ Inventario de cargas (Módulo 1)"]
    if consumo_rec10 > 0:
        fuentes10.append("🧾 Recibo de energía (Módulo 2)")
        fuentes10.append("📊 Mayor de los dos (recomendado)")
    if calc_n_pan10:
        fuentes10.append(f"🔆 Módulos 6+7 — ya calculados ({calc_n_pan10} paneles)")

    fuente10_sel = st.radio("Fuente:", fuentes10, horizontal=True, key="p10_fuente")

    # Resolver valores según fuente
    if "Módulos 6" in fuente10_sel and calc_n_pan10:
        consumo_base10 = calc_fs10 / 1.20 if calc_fs10 > 0 else consumo_mayor10
        consumo10_fs   = calc_fs10 if calc_fs10 > 0 else consumo_base10 * 1.20
        hsp10          = calc_hsp10
        vdc10          = calc_vdc10
        pot_panel10    = calc_pp10
        n_pan10        = calc_n_pan10
        n_bat10        = calc_n_bat10 if calc_n_bat10 else calcular_baterias(consumo10_fs, vdc10, **_bat_params(st.session_state))["num_baterias"]
        fuente_label10 = f"Módulos 6+7 · {n_pan10} paneles · {n_bat10} baterías"
        fuente_color10 = "#00E676"
    elif "Recibo" in fuente10_sel and consumo_rec10 > 0:
        consumo_base10 = consumo_rec10
        consumo10_fs   = consumo_rec10 * 1.20
        hsp10          = calc_hsp10
        vdc10          = tension_dc(consumo10_fs)
        pot_panel10    = calc_pp10
        n_pan10        = num_paneles(consumo10_fs / (hsp10 * calc_pr10), pot_panel10) if hsp10 * calc_pr10 > 0 else 1
        bats_r10       = calcular_baterias(consumo10_fs, vdc10, **_bat_params(st.session_state))
        n_bat10        = bats_r10["num_baterias"]
        fuente_label10 = f"Recibo · {consumo_rec10:,.0f} Wh/día → {n_pan10} paneles · {n_bat10} baterías"
        fuente_color10 = "#00BCD4"
    elif "Mayor" in fuente10_sel:
        consumo_base10 = consumo_mayor10
        consumo10_fs   = consumo_mayor10 * 1.20
        hsp10          = calc_hsp10
        vdc10          = tension_dc(consumo10_fs)
        pot_panel10    = calc_pp10
        n_pan10        = num_paneles(consumo10_fs / (hsp10 * calc_pr10), pot_panel10) if hsp10 * calc_pr10 > 0 else 1
        bats_m10       = calcular_baterias(consumo10_fs, vdc10, **_bat_params(st.session_state))
        n_bat10        = bats_m10["num_baterias"]
        fuente_label10 = f"Mayor de los dos · {consumo_mayor10:,.0f} Wh/día → {n_pan10} paneles · {n_bat10} baterías"
        fuente_color10 = "#FFB300"
    else:  # Inventario
        consumo_base10 = consumo_inv10
        consumo10_fs   = consumo_inv10 * 1.20
        hsp10          = calc_hsp10
        vdc10          = tension_dc(consumo10_fs)
        pot_panel10    = calc_pp10
        n_pan10        = num_paneles(consumo10_fs / (hsp10 * calc_pr10), pot_panel10) if hsp10 * calc_pr10 > 0 else 1
        bats_i10       = calcular_baterias(consumo10_fs, vdc10, **_bat_params(st.session_state))
        n_bat10        = bats_i10["num_baterias"]
        fuente_label10 = f"Inventario · {consumo_inv10:,.0f} Wh/día → {n_pan10} paneles · {n_bat10} baterías"
        fuente_color10 = "#FFB300"

    # Garantizar valores mínimos válidos
    n_pan10    = max(1, int(n_pan10))
    n_bat10    = max(1, int(n_bat10))
    hsp10      = max(0.5, float(hsp10))
    vdc10      = int(vdc10) if vdc10 in (12, 24, 48) else 48
    pot_panel10 = max(10, float(pot_panel10))
    ah_banco10 = n_bat10 * bat_cap10

    # Banner de fuente activa
    st.markdown(f"""
    <div style='background:rgba(0,0,0,0.2);border:1px solid {fuente_color10}55;
                border-left:4px solid {fuente_color10};border-radius:8px;
                padding:0.7rem 1rem;margin-bottom:1rem;
                display:flex;gap:2rem;flex-wrap:wrap;align-items:center;'>
        <span style='color:{fuente_color10};font-family:Rajdhani,sans-serif;
                     font-weight:700;font-size:0.95rem;'>{fuente_label10}</span>
        <span style='font-size:0.8rem;color:#8A9BBD;'>
            Consumo: <b style='color:#FFD54F;'>{consumo_base10:,.0f} Wh/día</b>
            | PR: <b style='color:#FF6B35;'>{int(calc_pr10*100)}%</b>
            → FS: <b style='color:#FFD54F;'>{consumo10_fs:,.0f} Wh/día</b>
            | HSP: <b style='color:#00BCD4;'>{hsp10} h</b>
            | VDC: <b style='color:#00BCD4;'>{vdc10} V</b>
            | Banco: <b style='color:#00BCD4;'>{ah_banco10} Ah</b>
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Comparativa de fuentes
    with st.expander("📊 Ver comparativa de fuentes disponibles"):
        src10_cols = st.columns(3)
        fuentes10_comp = [
            ("⚡ Inventario", consumo_inv10,   "#FFB300"),
            ("🧾 Recibo",     consumo_rec10,   "#00BCD4"),
            ("📊 Mayor",      consumo_mayor10, "#00E676"),
        ]
        for col_c10, (lbl_c10, cons_c10, clr_c10) in zip(src10_cols, fuentes10_comp):
            if cons_c10 > 0:
                fs_c10  = cons_c10 * 1.20
                vdc_c10 = tension_dc(fs_c10)
                n_c10   = num_paneles(fs_c10 / (hsp10 * calc_pr10), pot_panel10) if hsp10 * calc_pr10 > 0 else 0
                nb_c10  = calcular_baterias(fs_c10, vdc_c10, **_bat_params(st.session_state))["num_baterias"]
                col_c10.markdown(f"""
                <div style='background:#1A2235;border:1px solid {clr_c10}44;
                            border-radius:8px;padding:0.8rem;text-align:center;'>
                    <div style='font-size:0.75rem;color:#8A9BBD;'>{lbl_c10}</div>
                    <div style='font-family:Share Tech Mono,monospace;
                                font-size:1.1rem;color:{clr_c10};'>{cons_c10:,.0f}</div>
                    <div style='font-size:0.7rem;color:#8A9BBD;'>Wh/día</div>
                    <div style='font-size:0.78rem;color:#FFD54F;margin-top:0.3rem;'>
                        {n_c10} paneles · {nb_c10} bat. · {vdc_c10}V
                    </div>
                </div>""", unsafe_allow_html=True)
            else:
                col_c10.markdown(f"""
                <div style='background:#161D30;border:1px dashed #2A3A55;
                            border-radius:8px;padding:0.8rem;text-align:center;
                            color:#2A3A55;font-size:0.8rem;'>{lbl_c10}<br>Sin datos</div>""",
                    unsafe_allow_html=True)

    # ── Derivados para el SVG ─────────────────────────────────────────────────
    pot_inst10    = consumo10_fs / hsp10 if hsp10 > 0 else 0
    # 4. Controlador: Isc × N_paneles
    corr_mppt10   = isc10 * n_pan10
    # 5. Inversor — metodología técnica
    _inv10 = calcular_inversor(cargas10 if not cargas10.empty else None,
                               fs=0.80, fm=1.25, vdc=vdc10)
    if _inv10["inv_w"] == 0:
        # Fallback: sin inventario de cargas → usar consumo seleccionado × FM
        _inv10_fallback_w = consumo10_fs * 1.25
        _inv10["inv_w"]   = float(next(
            (k*1000 for k in _KW_COMERCIALES if k*1000 >= _inv10_fallback_w),
            math.ceil(_inv10_fallback_w/1000)*1000))
        _inv10["inv_kw"]  = _inv10["inv_w"] / 1000
        _inv10["corr_dc"] = _inv10["inv_w"] / vdc10 if vdc10 > 0 else 0
    pot_inv10_w = _inv10["inv_w"]
    vmp10    = round(voc10 * 0.80, 1)
    serie10  = max(1, round(vdc10 / vmp10)) if vmp10 > 0 else 1
    par10    = math.ceil(n_pan10 / serie10)
    v_array10  = round(voc10 * serie10, 1)
    i_array10  = round(isc10 * par10, 1)

    if corr_mppt10 <= 40:   mppt_label10 = "MPPT 40A"
    elif corr_mppt10 <= 60: mppt_label10 = "MPPT 60A"
    elif corr_mppt10 <=100: mppt_label10 = "MPPT 100A"
    else: mppt_label10 = f"MPPT {math.ceil(corr_mppt10/50)*50}A"
    W10 = 1100; H10 = 720

    C_BG10   = "#0A0E1A"
    C_CARD10 = "#1A2235"
    C_SOL10  = "#FFB300"
    C_WIRE10 = "#FFB300"
    C_NEG10  = "#FF5252"
    C_AC10   = "#00E676"
    C_DC10   = "#00BCD4"
    C_TEXT10 = "#E8EDF5"
    C_DIM10  = "#8A9BBD"
    C_GND10  = "#8A9BBD"

    def _svg_safe10(val, default="—"):
        return str(val if val is not None else default).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

    nombre10      = _svg_safe10(p10[1] if p10 else None, "Proyecto")
    municipio10   = _svg_safe10(p10[2] if p10 else None, "—")
    modelo10_safe = _svg_safe10(modelo10, "Panel")

    def box(x,y,w,h,fill,stroke,rx=6,opacity=1,sw=1.5):
        return f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="{rx}" fill="{fill}" stroke="{stroke}" stroke-width="{sw}" opacity="{opacity}"/>'
    def txt(x,y,text,size=10,fill=C_TEXT10,anchor="middle",weight="normal",family="Rajdhani,sans-serif"):
        return f'<text x="{x}" y="{y}" text-anchor="{anchor}" font-family="{family}" font-size="{size}" fill="{fill}" font-weight="{weight}">{text}</text>'
    def line(x1,y1,x2,y2,stroke=C_WIRE10,sw=2.5,dash=""):
        da = f' stroke-dasharray="{dash}"' if dash else ""
        return f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="{stroke}" stroke-width="{sw}"{da}/>'
    def wire_label(x,y,label,color=C_WIRE10):
        return (f'<rect x="{x-2}" y="{y-9}" width="{len(label)*6+8}" height="13" rx="3" '
                f'fill="#0F1525" stroke="{color}" stroke-width="0.8"/>'
                f'<text x="{x+2}" y="{y}" font-family="Share Tech Mono,monospace" font-size="7.5" '
                f'fill="{color}">{label}</text>')

    # ── Component blocks ──────────────────────────────────────────────────
    # 1. Solar array block (left)
    ARR_X=38; ARR_Y=100; ARR_W=160; ARR_H=300

    # Draw mini panels inside array box (3 rows × 2 cols)
    mini_panels = ""
    mp_cols=2; mp_rows=min(par10,4); mp_w=58; mp_h=54; mp_sep=6
    mp_ox = ARR_X+10; mp_oy = ARR_Y+18
    for mr in range(mp_rows):
        for mc in range(mp_cols):
            mpx = mp_ox + mc*(mp_w+mp_sep)
            mpy = mp_oy + mr*(mp_h+mp_sep)
            mini_panels += box(mpx,mpy,mp_w,mp_h,"#1565C0","#42A5F5",3)
            # cells 2×3
            cw=(mp_w-8)/2; ch=(mp_h-8)/3
            for cr in range(3):
                for cc in range(2):
                    mini_panels += box(mpx+4+cc*(cw+1.5),mpy+4+cr*(ch+1.5),cw-1,ch-1,"#0D47A1","none",1)
    if par10 > 4:
        mini_panels += txt(ARR_X+ARR_W/2, ARR_Y+mp_rows*(mp_h+mp_sep)+34, f"+{par10-4} strings más",7,C_DIM10)

    arr_block = f'''
    <g id="array">
      {box(ARR_X,ARR_Y,ARR_W,ARR_H,"#0F1525","#1976D2",8)}
      {box(ARR_X,ARR_Y,ARR_W,22,"#1565C0","#1976D2",8,1,0)}
      {txt(ARR_X+ARR_W/2,ARR_Y+14,"ARRAY FV",8,C_SOL10,"middle","700")}
      {mini_panels}
      {txt(ARR_X+ARR_W/2,ARR_Y+ARR_H-42,f"{serie10}S × {par10}P = {n_pan10} paneles",7.5,C_DIM10)}
      {txt(ARR_X+ARR_W/2,ARR_Y+ARR_H-28,f"Voc={v_array10}V  Isc={i_array10}A",7.5,"#42A5F5","middle","normal","Share Tech Mono,monospace")}
      {txt(ARR_X+ARR_W/2,ARR_Y+ARR_H-14,f"{modelo10_safe[:22]}",7,C_DIM10)}
    </g>'''

    # 2. MPPT Controller
    MPPT_X=260; MPPT_Y=170; MPPT_W=130; MPPT_H=120
    mppt_block = f'''
    <g id="mppt">
      {box(MPPT_X,MPPT_Y,MPPT_W,MPPT_H,"#0F1525",C_SOL10,8)}
      {box(MPPT_X,MPPT_Y,MPPT_W,22,"#1A2235",C_SOL10,8,1,0)}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+14,"CONTROLADOR MPPT",7.5,C_SOL10,"middle","700")}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+40,mppt_label10,12,C_SOL10,"middle","700")}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+57,f"Vin: {v_array10}V  Iin: {i_array10}A",7.5,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+70,f"Vout: {vdc10}V",7.5,C_DC10,"middle","normal","Share Tech Mono,monospace")}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+84,f"Iout: {corr_mppt10:.0f}A (max)",7.5,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {txt(MPPT_X+MPPT_W/2,MPPT_Y+100,f"Bat: {ah_banco10}Ah @ {vdc10}V",7.5,C_DC10,"middle","normal","Share Tech Mono,monospace")}
    </g>'''

    # 3. Battery bank
    BAT_X=260; BAT_Y=340; BAT_W=130; BAT_H=160
    bat_icons = ""
    bat_cols_d=min(n_bat10,4); bat_rows_d=math.ceil(n_bat10/4)
    bw_d=26; bh_d=34; bsep_d=4
    boff_x=(BAT_W-(bat_cols_d*bw_d+(bat_cols_d-1)*bsep_d))/2
    boff_y=28
    for bri in range(bat_rows_d):
        for bci in range(min(bat_cols_d,n_bat10-bri*bat_cols_d)):
            bx=BAT_X+boff_x+bci*(bw_d+bsep_d)
            by=BAT_Y+boff_y+bri*(bh_d+bsep_d)
            bat_icons += box(bx,by,bw_d,bh_d,"#1E2A3F","#00BCD4",3)
            bat_icons += box(bx+bw_d*0.3,by-4,bw_d*0.4,5,"#00BCD4","#00BCD4",1)  # terminal
            bat_icons += txt(bx+bw_d/2,by+bh_d/2+4,"Li",7,C_DC10)
    if n_bat10 > 16:
        bat_icons += txt(BAT_X+BAT_W/2,BAT_Y+boff_y+(bat_rows_d)*(bh_d+bsep_d)+4,
                         f"+{n_bat10-16} más",7,C_DIM10)
    bat_block = f'''
    <g id="baterias">
      {box(BAT_X,BAT_Y,BAT_W,BAT_H,"#0F1525",C_DC10,8)}
      {box(BAT_X,BAT_Y,BAT_W,22,"#1A2235",C_DC10,8,1,0)}
      {txt(BAT_X+BAT_W/2,BAT_Y+14,"BANCO BATERÍAS",7.5,C_DC10,"middle","700")}
      {bat_icons}
      {txt(BAT_X+BAT_W/2,BAT_Y+BAT_H-28,f"{n_bat10} × {bat_cap10}Ah Li  @{vdc10}V",7.5,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {txt(BAT_X+BAT_W/2,BAT_Y+BAT_H-14,f"Cap: {ah_banco10}Ah  {ah_banco10*vdc10/1000:.1f}kWh",7.5,C_DC10,"middle","normal","Share Tech Mono,monospace")}
    </g>'''

    # 4. Inverter
    INV_X=460; INV_Y=190; INV_W=130; INV_H=110
    inv_va   = int(pot_inv10_w)
    inv_kva  = round(pot_inv10_w/1000,1)
    inv_block = f'''
    <g id="inversor">
      {box(INV_X,INV_Y,INV_W,INV_H,"#0F1525","#00E676",8)}
      {box(INV_X,INV_Y,INV_W,22,"#1A2235","#00E676",8,1,0)}
      {txt(INV_X+INV_W/2,INV_Y+14,"INVERSOR",7.5,C_AC10,"middle","700")}
      <text x="{INV_X+INV_W/2}" y="{INV_Y+48}" text-anchor="middle"
            font-family="Rajdhani,sans-serif" font-size="22" fill="{C_AC10}" font-weight="700">~</text>
      {txt(INV_X+INV_W/2,INV_Y+65,f"Vin: {vdc10}V DC",7.5,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {txt(INV_X+INV_W/2,INV_Y+78,f"Vout: 120/220V AC",7.5,C_AC10,"middle","normal","Share Tech Mono,monospace")}
      {txt(INV_X+INV_W/2,INV_Y+91,f"Cap: {inv_kva} kVA",7.5,C_AC10,"middle","normal","Share Tech Mono,monospace")}
    </g>'''

    # 5. AC Loads panel
    LOAD_X=660; LOAD_Y=130; LOAD_W=150; LOAD_H=230
    n_loads = min(len(cargas10), 8) if not cargas10.empty else 3
    load_items = ""
    load_sample = cargas10.head(n_loads) if not cargas10.empty else pd.DataFrame()
    for li, (_, lr) in enumerate(load_sample.iterrows()):
        ly_l = LOAD_Y+32+li*22
        icon = "⚡" if int(lr["es_motor"]) else "💡"
        name = str(lr["electrodomestico"])[:18].replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
        load_items += txt(LOAD_X+16,ly_l+10,f"{icon} {name}",7.5,C_TEXT10,"start","normal","Barlow,sans-serif")
        load_items += txt(LOAD_X+LOAD_W-8,ly_l+10,f"{int(lr['potencia_w'])}W",7.5,C_DIM10,"end","normal","Share Tech Mono,monospace")
        load_items += line(LOAD_X+6,ly_l+14,LOAD_X+LOAD_W-6,ly_l+14,C_CARD10,0.5)
    extra = len(cargas10)-n_loads if not cargas10.empty and len(cargas10)>n_loads else 0
    if extra>0:
        load_items += txt(LOAD_X+LOAD_W/2,LOAD_Y+32+n_loads*22+8,f"+ {extra} equipos más",7.5,C_DIM10)
    total_loads_w = int((cargas10["cantidad"]*cargas10["potencia_w"]).sum()) if not cargas10.empty else 0
    loads_block = f'''
    <g id="cargas">
      {box(LOAD_X,LOAD_Y,LOAD_W,LOAD_H,"#0F1525",C_AC10,8)}
      {box(LOAD_X,LOAD_Y,LOAD_W,22,"#1A2235",C_AC10,8,1,0)}
      {txt(LOAD_X+LOAD_W/2,LOAD_Y+14,"CARGAS AC",7.5,C_AC10,"middle","700")}
      {load_items}
      {txt(LOAD_X+LOAD_W/2,LOAD_Y+LOAD_H-14,f"Total: {total_loads_w:,} W",8,C_AC10,"middle","700","Share Tech Mono,monospace")}
    </g>'''

    # 6. DC Loads (small box)
    DCL_X=660; DCL_Y=390; DCL_W=150; DCL_H=70
    dcl_block = f'''
    <g id="cargas_dc">
      {box(DCL_X,DCL_Y,DCL_W,DCL_H,"#0F1525",C_DC10,8)}
      {box(DCL_X,DCL_Y,DCL_W,22,"#1A2235",C_DC10,8,1,0)}
      {txt(DCL_X+DCL_W/2,DCL_Y+14,"CARGAS DC",7.5,C_DC10,"middle","700")}
      {txt(DCL_X+DCL_W/2,DCL_Y+38,f"Sistema {vdc10}V DC",8,C_TEXT10,"middle","normal","Share Tech Mono,monospace")}
      {txt(DCL_X+DCL_W/2,DCL_Y+54,"(LED, 12/24/48V)",7.5,C_DIM10)}
    </g>'''

    # 7. Protection / switchgear boxes
    PROT_X=840; PROT_Y=150; PROT_W=110; PROT_H=280
    prot_block = f'''
    <g id="protecciones">
      {box(PROT_X,PROT_Y,PROT_W,PROT_H,"#0F1525","#8A9BBD",8)}
      {box(PROT_X,PROT_Y,PROT_W,22,"#1A2235","#8A9BBD",8,1,0)}
      {txt(PROT_X+PROT_W/2,PROT_Y+14,"PROTECCIONES",7.5,C_DIM10,"middle","700")}
      {box(PROT_X+10,PROT_Y+32,PROT_W-20,28,"#1E2A3F","#FFB300",4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+44,"Fusible DC Array",7.5,C_SOL10)}
      {txt(PROT_X+PROT_W/2,PROT_Y+56,f"Ij={i_array10}A / {int(v_array10)}V",7,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {box(PROT_X+10,PROT_Y+72,PROT_W-20,28,"#1E2A3F","#00BCD4",4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+84,"Fusible Batería DC",7.5,C_DC10)}
      {txt(PROT_X+PROT_W/2,PROT_Y+96,f"Ij={corr_mppt10:.0f}A / {vdc10}V",7,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {box(PROT_X+10,PROT_Y+112,PROT_W-20,28,"#1E2A3F","#00E676",4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+124,"Interruptor AC",7.5,C_AC10)}
      {txt(PROT_X+PROT_W/2,PROT_Y+136,f"220V / {int(pot_inv10_w/220)}A",7,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
      {box(PROT_X+10,PROT_Y+152,PROT_W-20,28,"#1E2A3F","#FF5252",4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+164,"Descargador SPD",7.5,"#FF5252")}
      {txt(PROT_X+PROT_W/2,PROT_Y+176,"Sobretensión",7,C_DIM10)}
      {box(PROT_X+10,PROT_Y+192,PROT_W-20,28,"#1E2A3F","#8A9BBD",4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+204,"Puesta a Tierra",7.5,C_GND10)}
      {txt(PROT_X+PROT_W/2,PROT_Y+216,"Sistema FV",7,C_DIM10)}
      {box(PROT_X+10,PROT_Y+232,PROT_W-20,28,"#1E2A3F",C_SOL10,4)}
      {txt(PROT_X+PROT_W/2,PROT_Y+244,"Seccionador DC",7.5,C_SOL10)}
      {txt(PROT_X+PROT_W/2,PROT_Y+256,f"≥{int(v_array10)}V",7,C_DIM10,"middle","normal","Share Tech Mono,monospace")}
    </g>'''

    # Ground symbol at bottom
    gnd_x=530; gnd_y=530
    gnd = f'''
    <g id="tierra">
      {line(gnd_x,gnd_y,gnd_x,gnd_y+18,C_GND10,1.5)}
      {line(gnd_x-16,gnd_y+18,gnd_x+16,gnd_y+18,C_GND10,2)}
      {line(gnd_x-10,gnd_y+24,gnd_x+10,gnd_y+24,C_GND10,1.5)}
      {line(gnd_x-5, gnd_y+30,gnd_x+5, gnd_y+30,C_GND10,1)}
      {txt(gnd_x,gnd_y+42,"GND",7,C_GND10)}
    </g>'''

    # ── Wiring between blocks ──────────────────────────────────────────────
    wires10 = f'''
    <!-- Array → MPPT (DC+) -->
    {line(ARR_X+ARR_W, ARR_Y+ARR_H*0.38, MPPT_X, MPPT_Y+MPPT_H*0.38, C_WIRE10, 2.5)}
    <!-- Array → MPPT (DC−) -->
    {line(ARR_X+ARR_W, ARR_Y+ARR_H*0.62, MPPT_X, MPPT_Y+MPPT_H*0.62, C_NEG10, 2.5)}
    {wire_label(310,ARR_Y+ARR_H*0.38-4,f"+{v_array10}V DC",C_WIRE10)}
    {wire_label(310,ARR_Y+ARR_H*0.62+8,"− GND",C_NEG10)}
    <!-- MPPT → Battery (DC bus) -->
    {line(MPPT_X+MPPT_W/2, MPPT_Y+MPPT_H, MPPT_X+MPPT_W/2, BAT_Y, C_DC10, 2.5)}
    {wire_label(MPPT_X+MPPT_W/2+4, (MPPT_Y+MPPT_H+BAT_Y)//2, f"{vdc10}V DC Bus",C_DC10)}
    <!-- MPPT → Inverter (DC) -->
    {line(MPPT_X+MPPT_W, MPPT_Y+MPPT_H//2, INV_X, INV_Y+INV_H//2, C_DC10, 2.5)}
    {wire_label((MPPT_X+MPPT_W+INV_X)//2-10, MPPT_Y+MPPT_H//2-6, f"{vdc10}V DC",C_DC10)}
    <!-- Battery ↔ Inverter -->
    {line(BAT_X+BAT_W, BAT_Y+BAT_H//2, INV_X, INV_Y+INV_H*0.75, C_DC10, 2, "6,3")}
    <!-- Inverter → AC Loads -->
    {line(INV_X+INV_W, INV_Y+INV_H//2, LOAD_X, LOAD_Y+LOAD_H//2, C_AC10, 2.5)}
    {wire_label((INV_X+INV_W+LOAD_X)//2-10, INV_Y+INV_H//2-7, "220V AC 60Hz",C_AC10)}
    <!-- Battery → DC Loads -->
    {line(BAT_X+BAT_W, BAT_Y+BAT_H*0.75, DCL_X, DCL_Y+DCL_H//2, C_DC10, 2)}
    {wire_label((BAT_X+BAT_W+DCL_X)//2-10, BAT_Y+BAT_H*0.75-6, f"{vdc10}V DC",C_DC10)}
    <!-- Inverter → Protection -->
    {line(LOAD_X+LOAD_W, LOAD_Y+LOAD_H//2, PROT_X, PROT_Y+PROT_H//2, C_AC10, 1.5, "4,3")}
    <!-- Ground connections -->
    {line(gnd_x, INV_Y+INV_H, gnd_x, gnd_y, C_GND10, 1, "3,3")}
    {line(INV_X+INV_W//2, INV_Y+INV_H, gnd_x, INV_Y+INV_H, C_GND10, 1, "3,3")}
    '''

    # Title block
    tb_y10 = H10 - 72
    proyecto10 = _svg_safe10(p10[1] if p10 else None, "Proyecto")
    mun10      = _svg_safe10(p10[2] if p10 else None, "—")
    title10 = f'''
    <rect x="0" y="{tb_y10}" width="{W10}" height="72" fill="#0F1525" stroke="#2A3A55" stroke-width="0.8"/>
    <line x1="0" y1="{tb_y10}" x2="{W10}" y2="{tb_y10}" stroke="{C_SOL10}" stroke-width="1.5"/>
    <text x="12" y="{tb_y10+18}" font-family="Rajdhani,sans-serif" font-size="13"
          fill="{C_SOL10}" font-weight="700">☀ PLANO GENERAL — SISTEMA FOTOVOLTAICO AISLADO (OFF-GRID)</text>
    <text x="12" y="{tb_y10+33}" font-family="Rajdhani,sans-serif" font-size="9" fill="{C_TEXT10}">
        Proyecto: {proyecto10}  |  Ubicación: {mun10}  |  VDC={vdc10}V  |  HSP={hsp10}h/día</text>
    <text x="12" y="{tb_y10+47}" font-family="Share Tech Mono,monospace" font-size="8.5" fill="{C_DIM10}">
        Array: {n_pan10} paneles {pot_panel10}Wp ({serie10}S×{par10}P)  |
        Baterías: {n_bat10}×{bat_cap10}Ah@{vdc10}V = {ah_banco10*vdc10/1000:.1f}kWh  |
        MPPT: {mppt_label10}  |  Inversor: {inv_kva}kVA</text>
    <text x="12" y="{tb_y10+61}" font-family="Share Tech Mono,monospace" font-size="7.5" fill="#2A3A55">
        Consumo: {consumo10_fs:,.0f}Wh/día (+20%FS)  |  SolarCalc Pro  |  {datetime.now().strftime("%d/%m/%Y")}  |  Diagrama Unifilar — Plano N°02</text>
    <text x="{W10-12}" y="{tb_y10+18}" text-anchor="end"
          font-family="Share Tech Mono,monospace" font-size="9" fill="{C_DIM10}">Plano N° 02  Rev.A</text>
    '''

    # Legend box
    leg10_x=840; leg10_y=450
    legend10 = f'''
    <g>
      {box(leg10_x,leg10_y,160,110,"#0F1525","#2A3A55",6)}
      {txt(leg10_x+80,leg10_y+14,"REFERENCIAS",7.5,C_DIM10,"middle","700")}
      {line(leg10_x+10,leg10_y+28,leg10_x+40,leg10_y+28,C_WIRE10,2.5)}
      {txt(leg10_x+46,leg10_y+32,"Cable DC Positivo",7.5,C_TEXT10,"start")}
      {line(leg10_x+10,leg10_y+44,leg10_x+40,leg10_y+44,C_NEG10,2.5)}
      {txt(leg10_x+46,leg10_y+48,"Cable DC Negativo",7.5,C_TEXT10,"start")}
      {line(leg10_x+10,leg10_y+60,leg10_x+40,leg10_y+60,C_AC10,2.5)}
      {txt(leg10_x+46,leg10_y+64,"Cable AC 220V",7.5,C_TEXT10,"start")}
      {line(leg10_x+10,leg10_y+76,leg10_x+40,leg10_y+76,C_DC10,2.5)}
      {txt(leg10_x+46,leg10_y+80,"Bus DC Baterías",7.5,C_TEXT10,"start")}
      {line(leg10_x+10,leg10_y+92,leg10_x+40,leg10_y+92,C_GND10,1.5,"3,3")}
      {txt(leg10_x+46,leg10_y+96,"Puesta a Tierra",7.5,C_TEXT10,"start")}
    </g>'''

    svg10 = f'''<svg viewBox="0 0 {W10} {H10}" xmlns="http://www.w3.org/2000/svg"
         style="background:{C_BG10}; border-radius:12px; width:100%; max-height:720px;">
      <defs>
        <pattern id="grid10" width="20" height="20" patternUnits="userSpaceOnUse">
          <path d="M 20 0 L 0 0 0 20" fill="none" stroke="#1A2235" stroke-width="0.3"/>
        </pattern>
        <filter id="glow10"><feGaussianBlur stdDeviation="1.5" result="b"/>
          <feMerge><feMergeNode in="b"/><feMergeNode in="SourceGraphic"/></feMerge></filter>
      </defs>
      <rect width="{W10}" height="{H10-72}" fill="url(#grid10)"/>
      {wires10}
      {arr_block}
      {mppt_block}
      {bat_block}
      {inv_block}
      {loads_block}
      {dcl_block}
      {prot_block}
      {gnd}
      {legend10}
      {title10}
    </svg>'''

    render_svg(svg10, height=740)

    st.markdown("<hr class='sep'>", unsafe_allow_html=True)

    # ── Leyenda técnica ───────────────────────────────────────────────────
    col10a, col10b, col10c = st.columns(3)
    with col10a:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>COMPONENTES DEL SISTEMA</div>
            <div style='font-size:0.82rem; line-height:1.9;'>
                🔆 Array FV: <b style='color:#FFD54F;'>{n_pan10} × {pot_panel10}Wp</b><br>
                🎛 MPPT: <b style='color:#FFD54F;'>{mppt_label10}</b><br>
                🔋 Baterías: <b style='color:#FFD54F;'>{n_bat10} × {bat_cap10}Ah @ {vdc10}V</b><br>
                ⚡ Inversor: <b style='color:#FFD54F;'>{inv_kva} kVA DC/AC</b><br>
                🛡 Protecciones: <b style='color:#FFD54F;'>Fusibles + SPD + Tierra</b>
            </div>
        </div>""", unsafe_allow_html=True)
    with col10b:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>PARÁMETROS ELÉCTRICOS</div>
            <div style='font-size:0.82rem; line-height:1.9; font-family:Share Tech Mono,monospace;'>
                Vtensión DC sistema: <b style='color:#00BCD4;'>{vdc10} V</b><br>
                Vtensión array: <b style='color:#00BCD4;'>{v_array10} V</b><br>
                Corriente array: <b style='color:#00BCD4;'>{i_array10} A</b><br>
                Corriente MPPT: <b style='color:#00BCD4;'>{corr_mppt10:.0f} A</b><br>
                Energía banco: <b style='color:#00E676;'>{ah_banco10*vdc10/1000:.1f} kWh</b>
            </div>
        </div>""", unsafe_allow_html=True)
    with col10c:
        st.markdown(f"""
        <div class='sol-card'>
            <div style='font-family:Rajdhani,sans-serif; color:#FFB300; font-weight:600; margin-bottom:0.6rem;'>ENERGÉTICO</div>
            <div style='font-size:0.82rem; line-height:1.9; font-family:Share Tech Mono,monospace;'>
                Consumo base: <b style='color:#FFD54F;'>{consumo_base10:,.0f} Wh/día</b><br>
                Consumo + 20%FS: <b style='color:#FFD54F;'>{consumo10_fs:,.0f} Wh/día</b><br>
                HSP: <b style='color:#FFD54F;'>{hsp10} h/día</b><br>
                Potencia instalada: <b style='color:#FFD54F;'>{pot_inst10:,.0f} Wp</b><br>
                Municipio: <b style='color:#00E676;'>{p10[2] if p10 and p10[2] else "—"}</b>
            </div>
        </div>""", unsafe_allow_html=True)

    # Downloads
    fname10_svg = f"Plano_General_{(p10[1] if p10 else 'Proyecto').replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.svg"
    st.download_button("⬇ Descargar Plano General (SVG)",
                       data=svg10.encode(),
                       file_name=fname10_svg,
                       mime="image/svg+xml",
                       use_container_width=True, key="dl_svg_general")


# ════════════════════════════════════════════════════════════════════════════
# TAB 11 — ANÁLISIS ECONÓMICO Y AMBIENTAL (OFF-GRID)
# ════════════════════════════════════════════════════════════════════════════
with tab11:
    st.markdown("""
    <div class='sol-card-title'><span class='step-badge'>11</span>
    ANÁLISIS ECONÓMICO Y AMBIENTAL — SISTEMA OFF-GRID</div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class='formula-box'>
        Ahorro mensual = Consumo cubierto (kWh/mes) × Tarifa ($/kWh)<br>
        Payback simple = Inversión total ÷ Ahorro anual<br>
        CO₂ evitado = Generación anual (kWh) × 0.126 kgCO₂/kWh (factor Colombia)
    </div>
    """, unsafe_allow_html=True)

    # ── Leer datos del proyecto desde BD ─────────────────────────────────────
    conn = get_conn()
    p_eco = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    cargas_eco = pd.read_sql(
        "SELECT cantidad, potencia_w, horas_dia FROM cargas WHERE proyecto_id=?",
        conn, params=(proyecto_id,))
    panel_eco = conn.execute(
        "SELECT potencia_wp, voc, isc FROM paneles WHERE proyecto_id=? ORDER BY id DESC LIMIT 1",
        (proyecto_id,)).fetchone()
    conn.close()

    consumo_inv_eco = (cargas_eco["cantidad"] * cargas_eco["potencia_w"] * cargas_eco["horas_dia"]).sum()                       if not cargas_eco.empty else 0.0
    consumo_rec_eco = st.session_state.get("consumo_recibo_wh", 0.0)
    hsp_eco         = p_eco[4] if p_eco and p_eco[4] else None
    # Forzar tipos correctos desde SQLite (puede venir None, str o int)
    try:
        vdc_eco_bd = int(p_eco[5]) if p_eco and len(p_eco) > 5 and p_eco[5] not in (None, "") else None
    except (ValueError, TypeError):
        vdc_eco_bd = None
    pot_panel_eco   = int(panel_eco[0]) if panel_eco else 550

    consumo_base_eco = max(float(consumo_inv_eco), float(consumo_rec_eco))                        if consumo_rec_eco > 0 else float(consumo_inv_eco)
    consumo_fs_eco   = consumo_base_eco * 1.20

    if consumo_base_eco == 0:
        st.markdown("""
        <div class='warn-box'>⚠ No hay consumo registrado. Ingresa las cargas en el Tab 1
        o un recibo en el Tab 2 para habilitar el análisis económico.</div>
        """, unsafe_allow_html=True)
    elif not hsp_eco:
        st.markdown("""
        <div class='warn-box'>⚠ No hay HSP guardada. Completa el Tab 4 (Hora Solar) y
        guarda el valor para habilitar el análisis económico.</div>
        """, unsafe_allow_html=True)
    else:
        # ── Derivar VDC: BD → auto-detectar si falta ─────────────────────────
        if vdc_eco_bd and int(vdc_eco_bd) in (12, 24, 48):
            vdc_eco = int(vdc_eco_bd)
        else:
            vdc_eco = int(tension_dc(consumo_fs_eco))   # fallback automático

        # ── Derivar dimensionamiento actual (usa PR del Tab 6) ───────────────
        pr_eco      = float(st.session_state.get("calc_pr", 0.75))
        hsp_ef_eco  = float(hsp_eco) * pr_eco
        n_pan_eco   = num_paneles(consumo_fs_eco / hsp_ef_eco, pot_panel_eco) \
                      if hsp_ef_eco > 0 else 1
        n_pan_eco   = max(1, int(n_pan_eco))
        bats_eco    = calcular_baterias(float(consumo_fs_eco), int(vdc_eco), **_bat_params(st.session_state))
        n_bat_eco   = int(bats_eco["num_baterias"])
        bat_cap_eco = 100   # capacidad estándar por batería (Ah)
        ah_eco      = n_bat_eco * bat_cap_eco
        gen_dia_eco = (pot_panel_eco * n_pan_eco / 1000) * float(hsp_eco) * pr_eco
        gen_mes_eco = gen_dia_eco * 30
        gen_anio_eco= gen_dia_eco * 365
        co2_eco     = gen_anio_eco * 0.126
        arboles_eco = co2_eco / 21

        col_eco1, col_eco2 = st.columns(2)

        with col_eco1:
            st.markdown("<div class='sol-card'>", unsafe_allow_html=True)
            st.markdown(
                "<div style='color:#FFB300;font-family:Rajdhani,sans-serif;"
                "font-weight:600;margin-bottom:0.8rem;font-size:1rem;'>"
                "💰 PARÁMETROS ECONÓMICOS</div>",
                unsafe_allow_html=True)

            tarifa_eco   = st.number_input(
                "Tarifa energía ($/kWh)", 100.0, 5000.0, 700.0, 50.0,
                help="Tarifa promedio Colombia: $600–$900/kWh", key="eco_tarifa")
            ppanel_eco   = st.number_input(
                "Precio panel solar ($/unidad)", 50000.0, 2000000.0, 320000.0, 10000.0,
                key="eco_ppanel")
            pbat_eco     = st.number_input(
                "Precio batería ($/unidad)", 50000.0, 5000000.0, 450000.0, 50000.0,
                key="eco_pbat")
            pcontrol_eco = st.number_input(
                "Precio controlador MPPT ($)", 50000.0, 3000000.0, 600000.0, 50000.0,
                key="eco_pcontrol")
            pinv_eco     = st.number_input(
                "Precio inversor ($)", 100000.0, 10000000.0, 1200000.0, 100000.0,
                key="eco_pinv")
            potros_eco   = st.number_input(
                "Otros costos (estructura, cable, MO) ($)",
                0.0, 50000000.0, 1500000.0, 200000.0, key="eco_otros")

            st.markdown("</div>", unsafe_allow_html=True)

        with col_eco2:
            # ── Cálculos económicos ──────────────────────────────────────────
            inv_pan_eco   = n_pan_eco  * ppanel_eco
            inv_bat_eco   = n_bat_eco  * pbat_eco
            inv_total_eco = inv_pan_eco + inv_bat_eco + pcontrol_eco + pinv_eco + potros_eco

            ahorro_mes_eco  = (consumo_base_eco / 1000) * 30 * tarifa_eco
            ahorro_anio_eco = ahorro_mes_eco * 12
            payback_eco     = inv_total_eco / ahorro_anio_eco if ahorro_anio_eco > 0 else 99
            tir_eco         = (ahorro_anio_eco / inv_total_eco) * 100 if inv_total_eco > 0 else 0

            # Indicadores resumen
            st.markdown(f"""
            <div class='result-highlight' style='border-color:rgba(0,230,118,0.5);
                 background:linear-gradient(135deg,rgba(0,230,118,0.08),rgba(0,230,118,0.02));'>
                <div style='color:#8A9BBD;font-size:0.8rem;text-transform:uppercase;'>
                    Ahorro mensual estimado</div>
                <div class='val' style='color:#00E676;'>${ahorro_mes_eco:,.0f} / mes</div>
            </div>
            <div class='metric-grid' style='margin-top:0.8rem;'>
                <div class='metric-box' style='border-color:rgba(255,179,0,0.5);'>
                    <div class='metric-val'>${inv_total_eco/1000000:.1f}M</div>
                    <div class='metric-unit'>COP</div>
                    <div class='metric-label'>INVERSIÓN TOTAL</div>
                </div>
                <div class='metric-box' style='border-color:rgba(0,230,118,0.5);'>
                    <div class='metric-val' style='color:#00E676;'>${ahorro_anio_eco:,.0f}</div>
                    <div class='metric-unit'>$/año</div>
                    <div class='metric-label'>AHORRO ANUAL</div>
                </div>
                <div class='metric-box' style='border-color:rgba(0,188,212,0.5);'>
                    <div class='metric-val' style='color:#00BCD4;'>{payback_eco:.1f}</div>
                    <div class='metric-unit'>años</div>
                    <div class='metric-label'>PAYBACK</div>
                </div>
                <div class='metric-box' style='border-color:rgba(0,188,212,0.5);'>
                    <div class='metric-val' style='color:#00BCD4;'>{tir_eco:.1f}%</div>
                    <div class='metric-unit'>/ año</div>
                    <div class='metric-label'>TIR SIMPLE</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # ── Tabla financiera detallada ────────────────────────────────────
            st.markdown(f"""
            <div class='sol-card' style='margin-top:0.8rem;'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;
                            font-weight:600;margin-bottom:0.8rem;'>DESGLOSE DE INVERSIÓN</div>
                <table style='width:100%;font-size:0.82rem;border-collapse:collapse;'>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Paneles ({n_pan_eco} × ${ppanel_eco:,.0f})</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>${inv_pan_eco:,.0f}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Baterías ({n_bat_eco} × ${pbat_eco:,.0f})</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>${inv_bat_eco:,.0f}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Controlador MPPT</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>${pcontrol_eco:,.0f}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Inversor</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>${pinv_eco:,.0f}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Otros (estructura + cable + MO)</td>
                        <td style='font-family:Share Tech Mono;color:#FFD54F;text-align:right;'>${potros_eco:,.0f}</td>
                    </tr>
                    <tr style='background:#1A2235;border-bottom:1px solid #2A3A55;'>
                        <td style='color:#FFB300;padding:0.4rem 0;font-weight:700;'>INVERSIÓN TOTAL</td>
                        <td style='font-family:Share Tech Mono;color:#FFB300;text-align:right;
                                   font-weight:700;font-size:0.95rem;'>${inv_total_eco:,.0f}</td>
                    </tr>
                    <tr style='border-bottom:1px solid #2A3A55;'>
                        <td style='color:#8A9BBD;padding:0.35rem 0;'>Ahorro mensual</td>
                        <td style='font-family:Share Tech Mono;color:#00E676;text-align:right;'>${ahorro_mes_eco:,.0f}</td>
                    </tr>
                    <tr style='background:#1A2235;'>
                        <td style='color:#00E676;padding:0.4rem 0;font-weight:700;'>AHORRO ANUAL</td>
                        <td style='font-family:Share Tech Mono;color:#00E676;text-align:right;
                                   font-weight:700;font-size:0.95rem;'>${ahorro_anio_eco:,.0f}</td>
                    </tr>
                </table>
            </div>
            """, unsafe_allow_html=True)

        # ── Fila ambiental + indicadores adicionales ──────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        col_amb1, col_amb2, col_amb3, col_amb4 = st.columns(4)

        for c_amb, icon_a, lbl_a, val_a, unit_a, col_a in [
            (col_amb1, "🌿", "CO₂ evitado/año",   f"{co2_eco:,.0f}",   "kg CO₂",    "#00BCD4"),
            (col_amb2, "🌳", "Árboles equiv./año", f"{arboles_eco:.0f}", "árboles",   "#00E676"),
            (col_amb3, "☀", "Generación anual",   f"{gen_anio_eco:,.0f}","kWh/año",  "#FFB300"),
            (col_amb4, "📅", "Vida útil sistema",  "20–25",              "años",      "#8A9BBD"),
        ]:
            with c_amb:
                st.markdown(f"""
                <div class='metric-box' style='text-align:center;border-top:2px solid {col_a};'>
                    <div style='font-size:1.6rem;margin-bottom:0.3rem;'>{icon_a}</div>
                    <div class='metric-val' style='color:{col_a};font-size:1.4rem;'>{val_a}</div>
                    <div class='metric-unit'>{unit_a}</div>
                    <div class='metric-label'>{lbl_a}</div>
                </div>""", unsafe_allow_html=True)

        # ── Resumen del sistema para contexto ─────────────────────────────────
        st.markdown("<hr class='sep'>", unsafe_allow_html=True)
        st.markdown(
            "<div style='color:#FFB300;font-family:Rajdhani,sans-serif;"
            "font-weight:600;font-size:1rem;margin-bottom:0.8rem;'>"
            "📋 SISTEMA DIMENSIONADO (BASE DE CÁLCULO)</div>",
            unsafe_allow_html=True)

        col_sum1, col_sum2, col_sum3 = st.columns(3)
        with col_sum1:
            st.markdown(f"""
            <div class='sol-card'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;
                            font-weight:600;margin-bottom:0.6rem;'>ARRAY FV</div>
                <div style='font-size:0.82rem;line-height:2;font-family:Share Tech Mono,monospace;'>
                    Paneles: <b style='color:#FFD54F;'>{n_pan_eco} × {pot_panel_eco} Wp</b><br>
                    Pot. inst.: <b style='color:#FFD54F;'>{n_pan_eco*pot_panel_eco/1000:.2f} kWp</b><br>
                    Gen. día: <b style='color:#00E676;'>{gen_dia_eco:.2f} kWh/día</b><br>
                    HSP bruta: <b style='color:#00BCD4;'>{hsp_eco} h/día</b><br>
                    PR aplicado: <b style='color:#FF6B35;'>{int(pr_eco*100)}%</b><br>
                    HSP efectiva: <b style='color:#FFB300;'>{hsp_ef_eco:.2f} h/día</b>
                </div>
            </div>""", unsafe_allow_html=True)
        with col_sum2:
            st.markdown(f"""
            <div class='sol-card'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;
                            font-weight:600;margin-bottom:0.6rem;'>BANCO BATERÍAS</div>
                <div style='font-size:0.82rem;line-height:2;font-family:Share Tech Mono,monospace;'>
                    N° baterías: <b style='color:#FFD54F;'>{n_bat_eco} unidades</b><br>
                    Cap. banco: <b style='color:#FFD54F;'>{ah_eco} Ah @ {vdc_eco}V</b><br>
                    Energía: <b style='color:#00E676;'>{ah_eco*vdc_eco/1000:.2f} kWh</b><br>
                    Tensión DC: <b style='color:#00BCD4;'>{vdc_eco} V</b>
                </div>
            </div>""", unsafe_allow_html=True)
        with col_sum3:
            st.markdown(f"""
            <div class='sol-card'>
                <div style='color:#FFB300;font-family:Rajdhani,sans-serif;
                            font-weight:600;margin-bottom:0.6rem;'>CONSUMO</div>
                <div style='font-size:0.82rem;line-height:2;font-family:Share Tech Mono,monospace;'>
                    Base: <b style='color:#FFD54F;'>{consumo_base_eco:,.0f} Wh/día</b><br>
                    Con 25% FS: <b style='color:#FFD54F;'>{consumo_fs_eco:,.0f} Wh/día</b><br>
                    Mensual: <b style='color:#00E676;'>{consumo_base_eco*30/1000:.1f} kWh/mes</b><br>
                    Anual: <b style='color:#00BCD4;'>{consumo_base_eco*365/1000:.0f} kWh/año</b>
                </div>
            </div>""", unsafe_allow_html=True)

        # ── Nota normativa ─────────────────────────────────────────────────
        st.markdown("""
        <div class='info-note' style='margin-top:1rem;'>
            ℹ <b>Notas metodológicas:</b>
            El ahorro mensual corresponde al costo evitado de comprar esa energía a la red eléctrica
            (o al generador diésel en zonas aisladas). Para sistemas OFF-GRID totalmente aislados,
            el análisis compara contra el costo de alternativas (generadores, velas, pilas).
            El factor CO₂ de 0.126 kgCO₂/kWh corresponde al factor de emisión de la red colombiana
            (UPME). Vida útil: paneles 25–30 años, baterías AGM/GEL 5–8 años, LiFePO4 10–15 años.
            <b>Normas aplicables: RETIE, NTC 2050, CREG 030-2018.</b>
        </div>
        """, unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# TAB 12 — CABLEADO
# ════════════════════════════════════════════════════════════════════════════
with tab12:
    if not proyecto_id:
        st.markdown("<div class='warn-box'>⚠ Selecciona o crea un proyecto en el menú lateral para calcular el cableado.</div>",
                    unsafe_allow_html=True)
    else:
        mostrar_cableado(proyecto_id, st.session_state)

# ─── FOOTER ─────────────────────────────────────────────────────────────────
st.markdown("""
<div style='text-align:center; padding:2rem 0 1rem; color:#2A3A55; font-size:0.75rem; letter-spacing:2px;'>
    SOLARCALC PRO · DIMENSIONAMIENTO FOTOVOLTAICO OFF-GRID · SQLITE3 + STREAMLIT
</div>
""", unsafe_allow_html=True)
