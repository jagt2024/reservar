"""
modulo_presupuesto.py — Presupuesto de Obra del Proyecto FV
SolarCalc Pro · Módulo externo
"""
import streamlit as st
import sqlite3
import pandas as pd
import math
import io
from datetime import datetime

from db_utils import get_conn, init_modulos_db

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─── Colores ─────────────────────────────────────────────────────────────────
SOL   = "#FFB300"; DARK  = "#0A0E1A"; CARD  = "#1A2235"
CARD2 = "#1E2A3F"; BRD   = "#2A3A55"; TEXT  = "#E8EDF5"
TEXT2 = "#8A9BBD"; GREEN = "#00E676"; CYAN  = "#00BCD4"
MONO  = "#FFD54F"; RED   = "#FF5252"

# ─── Helpers ─────────────────────────────────────────────────────────────────
def _cap_options(proyecto_id):
    conn = get_conn()
    caps = pd.read_sql(
        "SELECT id, orden, nombre FROM presupuesto_capitulos WHERE proyecto_id=? ORDER BY orden",
        conn, params=(proyecto_id,))
    conn.close()
    return caps

def _get_items(capitulo_id):
    conn = get_conn()
    items = pd.read_sql(
        "SELECT * FROM presupuesto_items WHERE capitulo_id=? ORDER BY item",
        conn, params=(capitulo_id,))
    conn.close()
    return items

def _total_proyecto(proyecto_id):
    conn = get_conn()
    r = conn.execute(
        "SELECT COALESCE(SUM(cantidad*valor_unitario),0) FROM presupuesto_items WHERE proyecto_id=?",
        (proyecto_id,)).fetchone()[0]
    conn.close()
    return r

def _materiales_lista():
    conn = get_conn()
    mats = pd.read_sql("SELECT id,categoria,descripcion,unidad,precio_ref FROM materiales WHERE activo=1 ORDER BY categoria,descripcion",
                        conn)
    conn.close()
    return mats

def _equipos_lista():
    conn = get_conn()
    eq = pd.read_sql("SELECT id,tipo,categoria,descripcion,unidad,precio_ref FROM equipos_herramientas WHERE activo=1 ORDER BY tipo,categoria",
                      conn)
    conn.close()
    return eq

def _personal_lista():
    conn = get_conn()
    per = pd.read_sql("SELECT id,cargo,perfil,certificacion,salario_dia FROM personal WHERE activo=1 ORDER BY cargo",
                       conn)
    conn.close()
    return per

# ─── Exportar Excel ───────────────────────────────────────────────────────────
def exportar_excel_presupuesto(proyecto_id, proyecto_nombre):
    conn = get_conn()
    caps = pd.read_sql(
        "SELECT * FROM presupuesto_capitulos WHERE proyecto_id=? ORDER BY orden",
        conn, params=(proyecto_id,))
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Presupuesto"
    ws.sheet_view.showGridLines = False

    col_ws = {"A":6,"B":8,"C":42,"D":8,"E":14,"F":14,"G":18}
    for col,w in col_ws.items(): ws.column_dimensions[col].width = w

    def hf(hex_c): return PatternFill("solid", fgColor=hex_c)
    def fn(bold=False,color="E8EDF5",size=9): return Font(bold=bold,color=color,size=size,name="Calibri")
    def ctr(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def lft(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
    def rgt(): return Alignment(horizontal="right",vertical="center")
    def brd():
        s=Side(style="thin",color="2A3A55"); return Border(left=s,right=s,top=s,bottom=s)

    # Encabezado
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"☀ PRESUPUESTO — {proyecto_nombre.upper()}"
    ws["A1"].fill=hf("0A0E1A"); ws["A1"].font=Font(bold=True,color="FFB300",size=14,name="Calibri")
    ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=30

    ws.merge_cells("A2:G2")
    ws["A2"].value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Todos los valores en COP $"
    ws["A2"].fill=hf("1A2235"); ws["A2"].font=fn(color="8A9BBD",size=8)
    ws["A2"].alignment=ctr(); ws.row_dimensions[2].height=14

    # Encabezado columnas
    hdrs=[("A","ID"),("B","ÍTEM"),("C","DESCRIPCIÓN ACTIVIDAD"),
          ("D","UND"),("E","CANTIDAD"),("F","VLR UNITARIO"),("G","VLR PARCIAL")]
    for col,h in hdrs:
        ws[f"{col}4"].value=h; ws[f"{col}4"].fill=hf("FFB300")
        ws[f"{col}4"].font=Font(bold=True,color="0A0E1A",size=9,name="Calibri")
        ws[f"{col}4"].alignment=ctr(); ws[f"{col}4"].border=brd()
    ws.row_dimensions[4].height=20

    row = 5; gran_total = 0
    for _, cap in caps.iterrows():
        # Fila de capítulo
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = f"CAPÍTULO {cap['orden']}: {cap['nombre'].upper()}"
        ws[f"A{row}"].fill=hf("1A2235"); ws[f"A{row}"].font=Font(bold=True,color="FFB300",size=9,name="Calibri")
        ws[f"A{row}"].alignment=lft(); ws.row_dimensions[row].height=18; row+=1

        items = _get_items(int(cap["id"]))
        cap_total = 0; alt=False
        for _, it in items.iterrows():
            vp = it["cantidad"] * it["valor_unitario"]
            cap_total += vp; gran_total += vp
            bg="161D30" if alt else "1E2A3F"; alt=not alt
            vals=[("A",str(int(it["id"])),ctr()),("B",it["item"],ctr()),
                  ("C",it["descripcion"],lft()),("D",it["unidad"] or "",ctr()),
                  ("E",it["cantidad"],rgt()),("F",it["valor_unitario"],rgt()),
                  ("G",vp,rgt())]
            for col,v,al in vals:
                c=ws[f"{col}{row}"]; c.value=v; c.fill=hf(bg)
                c.font=fn(color="E8EDF5"); c.alignment=al; c.border=brd()
                if col in ("E","F","G") and isinstance(v,(int,float)): c.number_format="#,##0"
            ws.row_dimensions[row].height=16; row+=1

        # Subtotal capítulo
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value=f"SUBTOTAL {cap['nombre'].upper()}"
        ws[f"A{row}"].fill=hf("0F1525"); ws[f"A{row}"].font=Font(bold=True,color="FFD54F",size=9,name="Calibri")
        ws[f"A{row}"].alignment=lft()
        ws[f"G{row}"].value=cap_total; ws[f"G{row}"].fill=hf("0F1525")
        ws[f"G{row}"].font=Font(bold=True,color="FFD54F",size=9,name="Calibri")
        ws[f"G{row}"].alignment=rgt(); ws[f"G{row}"].number_format="#,##0"
        ws.row_dimensions[row].height=18; row+=2

    # Gran total
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value="GRAN TOTAL DEL PRESUPUESTO"
    ws[f"A{row}"].fill=hf("FFB300"); ws[f"A{row}"].font=Font(bold=True,color="0A0E1A",size=10,name="Calibri")
    ws[f"A{row}"].alignment=lft()
    ws[f"G{row}"].value=gran_total; ws[f"G{row}"].fill=hf("FFB300")
    ws[f"G{row}"].font=Font(bold=True,color="0A0E1A",size=11,name="Calibri")
    ws[f"G{row}"].alignment=rgt(); ws[f"G{row}"].number_format="#,##0"
    ws.row_dimensions[row].height=24

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════
def mostrar_presupuesto(proyecto_id: int, ss: dict):
    init_modulos_db()

    st.markdown("""
    <div style='background:linear-gradient(135deg,#0A0E1A,#1A2235);border:1px solid #2A3A55;
     border-radius:12px;padding:1.2rem 1.5rem;margin-bottom:1.5rem;'>
        <div style='font-family:Rajdhani,sans-serif;font-size:1.6rem;font-weight:700;
         color:#FFB300;letter-spacing:2px;'>💰 PRESUPUESTO DEL PROYECTO</div>
        <div style='color:#8A9BBD;font-size:0.8rem;letter-spacing:2px;margin-top:0.2rem;'>
            CAPÍTULOS · ACTIVIDADES · MATERIALES · EQUIPOS · PERSONAL</div>
    </div>
    """, unsafe_allow_html=True)

    conn = get_conn()
    p = conn.execute("SELECT * FROM proyectos WHERE id=?", (proyecto_id,)).fetchone()
    conn.close()
    proyecto_nombre = p[1] if p else "Proyecto"

    # Tabs internos del módulo
    pt1, pt2, pt3, pt4, pt5 = st.tabs([
        "📋 Capítulos e Ítems",
        "🔩 Materiales",
        "🔧 Equipos y Herramientas",
        "👷 Personal",
        "📊 Resumen y Descarga",
    ])

    # ══ TAB 1 — Capítulos e Ítems ════════════════════════════════════════════
    with pt1:
        st.markdown("##### ➕ Nuevo capítulo")
        c1,c2,c3 = st.columns([1,3,1])
        with c1: cap_orden = st.number_input("Orden", 1, 99, 1, key="cap_ord")
        with c2: cap_nom   = st.text_input("Nombre del capítulo",
                                            placeholder="Ej: Suministro de materiales, Mano de obra...",
                                            key="cap_nom")
        with c3:
            st.markdown("<div style='margin-top:1.6rem;'>", unsafe_allow_html=True)
            if st.button("➕ Crear capítulo", use_container_width=True):
                if cap_nom.strip():
                    conn = get_conn()
                    conn.execute("INSERT INTO presupuesto_capitulos(proyecto_id,orden,nombre) VALUES(?,?,?)",
                                  (proyecto_id, cap_orden, cap_nom.strip()))
                    conn.commit(); conn.close()
                    st.success("Capítulo creado ✓"); st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        caps = _cap_options(proyecto_id)
        if caps.empty:
            st.info("No hay capítulos. Crea el primero arriba."); return

        # Selector de capítulo activo
        st.markdown("<hr style='border-color:#2A3A55;'>", unsafe_allow_html=True)
        cap_opts = {f"{int(r['orden'])}. {r['nombre']}": int(r["id"]) for _, r in caps.iterrows()}
        cap_sel_nom = st.selectbox("Capítulo activo:", list(cap_opts.keys()), key="cap_sel")
        cap_sel_id  = cap_opts[cap_sel_nom]

        # Fuentes de recursos
        mats  = _materiales_lista()
        eqs   = _equipos_lista()
        pers  = _personal_lista()

        fuente_opts = ["— Ingreso manual —"]
        if not mats.empty:
            fuente_opts += [f"[MAT] {r['descripcion']} ({r['unidad']}) — ${r['precio_ref']:,.0f}"
                            for _, r in mats.iterrows()]
        if not eqs.empty:
            fuente_opts += [f"[{r['tipo'][:3].upper()}] {r['descripcion']} ({r['unidad']}) — ${r['precio_ref']:,.0f}"
                            for _, r in eqs.iterrows()]
        if not pers.empty:
            fuente_opts += [f"[PER] {r['cargo']} — ${r['salario_dia']:,.0f}/día"
                            for _, r in pers.iterrows()]

        st.markdown(f"##### ✏ Agregar ítem al capítulo: *{cap_sel_nom}*")
        fuente_sel = st.selectbox("Cargar desde catálogo (opcional):", fuente_opts, key="it_fuente")

        # Auto-rellenar desde fuente
        def_desc=""; def_und=""; def_vu=0.0; def_tipo="Otro"; def_rid=None
        if fuente_sel != "— Ingreso manual —":
            tag = fuente_sel[:5]
            if tag == "[MAT]":
                idx_f = fuente_opts.index(fuente_sel) - 1
                if 0 <= idx_f < len(mats):
                    rm = mats.iloc[idx_f]
                    def_desc = rm["descripcion"]; def_und  = rm["unidad"]
                    def_vu   = float(rm["precio_ref"]); def_tipo = "Material"
                    def_rid  = int(rm["id"])
            elif tag in ("[EQU]","[HER]"):
                idx_f = fuente_opts.index(fuente_sel) - 1 - len(mats)
                if 0 <= idx_f < len(eqs):
                    re = eqs.iloc[idx_f]
                    def_desc = re["descripcion"]; def_und  = re["unidad"]
                    def_vu   = float(re["precio_ref"]); def_tipo = re["tipo"]
                    def_rid  = int(re["id"])
            elif tag == "[PER]":
                idx_f = fuente_opts.index(fuente_sel) - 1 - len(mats) - len(eqs)
                if 0 <= idx_f < len(pers):
                    rp = pers.iloc[idx_f]
                    def_desc = rp["cargo"]; def_und  = "día"
                    def_vu   = float(rp["salario_dia"]); def_tipo = "Personal"
                    def_rid  = int(rp["id"])

        fi1,fi2,fi3,fi4 = st.columns([1,3,1.2,1.5])
        with fi1: it_item = st.text_input("Ítem", placeholder="1.1", key="it_item")
        with fi2: it_desc = st.text_input("Descripción actividad",
                                           value=def_desc, key="it_desc")
        with fi3: it_und  = st.text_input("Unidad", value=def_und,
                                           placeholder="UND,ML,M²,GL...", key="it_und")
        fi5,fi6,fi7 = st.columns([1.5,1.8,1])
        with fi5: it_cant = st.number_input("Cantidad", 0.0, 999999.0, 1.0, 0.5, key="it_cant")
        with fi6: it_vu   = st.number_input("Valor unitario ($)", 0.0, 999999999.0,
                                             def_vu, 1000.0, key="it_vu")
        with fi7:
            vp_preview = it_cant * it_vu
            st.markdown(f"""
            <div style='background:#1E2A3F;border:1px solid #2A3A55;border-radius:8px;
             padding:0.6rem;text-align:center;margin-top:1.55rem;'>
                <div style='font-size:0.7rem;color:#8A9BBD;'>Valor parcial</div>
                <div style='font-family:Share Tech Mono,monospace;color:#FFB300;font-size:1rem;'>
                    ${vp_preview:,.0f}</div>
            </div>""", unsafe_allow_html=True)

        if st.button("➕ Agregar ítem al presupuesto", use_container_width=True):
            if it_item.strip() and it_desc.strip():
                conn = get_conn()
                conn.execute("""
                    INSERT INTO presupuesto_items
                        (capitulo_id,proyecto_id,item,descripcion,unidad,cantidad,
                         valor_unitario,tipo_recurso,recurso_id)
                    VALUES(?,?,?,?,?,?,?,?,?)
                """, (cap_sel_id, proyecto_id, it_item.strip(), it_desc.strip(),
                      it_und.strip(), it_cant, it_vu, def_tipo, def_rid))
                conn.commit(); conn.close()
                st.success("Ítem agregado ✓"); st.rerun()
            else:
                st.error("Completa Ítem y Descripción")

        # Tabla del capítulo activo
        st.markdown("<hr style='border-color:#2A3A55;'>", unsafe_allow_html=True)
        items_df = _get_items(cap_sel_id)
        if not items_df.empty:
            items_df["valor_parcial"] = items_df["cantidad"] * items_df["valor_unitario"]
            show = items_df[["id","item","descripcion","unidad","cantidad",
                              "valor_unitario","valor_parcial"]].copy()
            show.columns = ["ID","Ítem","Descripción","Und","Cant","VLR Unitario","VLR Parcial"]
            st.dataframe(show.set_index("ID"), use_container_width=True)

            sub = items_df["valor_parcial"].sum()
            st.markdown(f"""
            <div style='background:#0F1525;border:1px solid rgba(255,179,0,0.4);border-radius:8px;
             padding:0.7rem 1.2rem;text-align:right;margin-top:0.5rem;'>
                <span style='color:#8A9BBD;'>Subtotal {cap_sel_nom}:</span>
                <b style='font-family:Share Tech Mono,monospace;color:#FFB300;font-size:1.1rem;
                 margin-left:1rem;'>$ {sub:,.0f}</b>
            </div>""", unsafe_allow_html=True)

            # Editar / Eliminar
            with st.expander("✏ Editar / Eliminar ítem"):
                it_sel_opts = {f"{int(r['id'])} — {r['item']} {r['descripcion'][:30]}": int(r["id"])
                               for _, r in items_df.iterrows()}
                it_edit_k = st.selectbox("Seleccionar:", list(it_sel_opts.keys()), key="it_edit_sel")
                it_edit_id = it_sel_opts[it_edit_k]
                it_row = items_df[items_df["id"]==it_edit_id].iloc[0]

                ec1,ec2,ec3,ec4 = st.columns([2,1.2,1.5,1])
                with ec1: e_desc = st.text_input("Descripción", value=it_row["descripcion"], key="e_desc")
                with ec2: e_und  = st.text_input("Unidad", value=it_row["unidad"] or "", key="e_und")
                with ec3: e_cant = st.number_input("Cantidad", 0.0, 999999.0,
                                                    float(it_row["cantidad"]), 0.5, key="e_cant")
                with ec4: e_vu   = st.number_input("VLR Unitario", 0.0, 999999999.0,
                                                    float(it_row["valor_unitario"]), 1000.0, key="e_vu")
                eb1,eb2 = st.columns(2)
                with eb1:
                    if st.button("💾 Guardar cambios", use_container_width=True, key="it_save"):
                        conn = get_conn()
                        conn.execute("UPDATE presupuesto_items SET descripcion=?,unidad=?,cantidad=?,valor_unitario=? WHERE id=?",
                                      (e_desc, e_und, e_cant, e_vu, it_edit_id))
                        conn.commit(); conn.close()
                        st.success("Actualizado ✓"); st.rerun()
                with eb2:
                    if st.button("🗑 Eliminar ítem", use_container_width=True, key="it_del"):
                        conn = get_conn()
                        conn.execute("DELETE FROM presupuesto_items WHERE id=?", (it_edit_id,))
                        conn.commit(); conn.close()
                        st.success("Eliminado ✓"); st.rerun()
        else:
            st.info("No hay ítems en este capítulo aún.")

    # ══ TAB 2 — Materiales ════════════════════════════════════════════════════
    with pt2:
        from modulo_materiales import mostrar_materiales
        mostrar_materiales()

    # ══ TAB 3 — Equipos y Herramientas ═══════════════════════════════════════
    with pt3:
        from modulo_equipos import mostrar_equipos
        mostrar_equipos()

    # ══ TAB 4 — Personal ═════════════════════════════════════════════════════
    with pt4:
        from modulo_equipos import mostrar_personal
        mostrar_personal()

    # ══ TAB 5 — Resumen y Descarga ════════════════════════════════════════════
    with pt5:
        caps_r = _cap_options(proyecto_id)
        if caps_r.empty:
            st.info("No hay capítulos registrados."); return

        gran_total_r = _total_proyecto(proyecto_id)
        aiu_pct = st.number_input("AIU (%)", 0.0, 50.0, 25.0, 0.5, key="aiu_pct",
                                   help="Administración, Imprevistos y Utilidad")
        aiu_val  = gran_total_r * aiu_pct / 100
        total_c_aiu = gran_total_r + aiu_val

        st.markdown(f"""
        <div style='background:#0F1525;border:1px solid #2A3A55;border-radius:12px;padding:1.5rem;'>
            <div style='font-family:Rajdhani,sans-serif;font-size:1.1rem;font-weight:700;
             color:#FFB300;letter-spacing:1px;margin-bottom:1rem;'>📊 RESUMEN GENERAL</div>
        """, unsafe_allow_html=True)

        total_caps = 0
        for _, cap_r in caps_r.iterrows():
            items_r = _get_items(int(cap_r["id"]))
            sub_r = (items_r["cantidad"] * items_r["valor_unitario"]).sum() if not items_r.empty else 0
            total_caps += sub_r
            st.markdown(f"""
            <div style='display:flex;justify-content:space-between;padding:0.4rem 0;
             border-bottom:1px solid #1A2235;'>
                <span style='color:#8A9BBD;font-size:0.85rem;'>
                    CAP {int(cap_r['orden'])}. {cap_r['nombre']}</span>
                <b style='font-family:Share Tech Mono,monospace;color:#FFD54F;'>
                    $ {sub_r:,.0f}</b>
            </div>""", unsafe_allow_html=True)

        st.markdown(f"""
            <div style='display:flex;justify-content:space-between;padding:0.6rem 0;margin-top:0.5rem;'>
                <span style='color:#8A9BBD;'>SUBTOTAL DIRECTO</span>
                <b style='font-family:Share Tech Mono,monospace;color:#FFD54F;'>$ {gran_total_r:,.0f}</b>
            </div>
            <div style='display:flex;justify-content:space-between;padding:0.4rem 0;'>
                <span style='color:#8A9BBD;'>AIU ({aiu_pct:.1f}%)</span>
                <b style='font-family:Share Tech Mono,monospace;color:#FFD54F;'>$ {aiu_val:,.0f}</b>
            </div>
            <div style='display:flex;justify-content:space-between;padding:0.8rem 0;
             border-top:2px solid #FFB300;margin-top:0.5rem;'>
                <span style='color:#FFB300;font-family:Rajdhani,sans-serif;font-size:1.1rem;
                 font-weight:700;'>TOTAL PRESUPUESTO</span>
                <b style='font-family:Share Tech Mono,monospace;color:#FFB300;font-size:1.3rem;'>
                    $ {total_c_aiu:,.0f}</b>
            </div>
        </div>""", unsafe_allow_html=True)

        # Guardar total en session_state para simulador
        ss["presup_total"] = total_c_aiu

        st.markdown("<br>", unsafe_allow_html=True)
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            xls_bytes = exportar_excel_presupuesto(proyecto_id, proyecto_nombre)
            st.download_button("⬇ Descargar Excel (.xlsx)",
                               data=xls_bytes,
                               file_name=f"Presupuesto_{proyecto_nombre.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True, key="dl_presup_xls")
