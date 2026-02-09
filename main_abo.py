import streamlit as st
from streamlit_option_menu import option_menu
import numpy as np
from inicio_abo import Inicio
from crear_reserva_abo import CrearReserva
from modificar_reserva_abo import ModificarReservaAbo
from eliminar_reserva_abo import EliminarReserva
from servicios_abo import Servicios
from informacion_abo import Informacion
from generar_excel_abo import GenerarExcel
from generaQR.generar_qr_abo import GenerarQr, GenerarQr_standalone
from consulta_st_excel import ConsultarAgenda, ConsultarAgenda_standalone
from descargar_agenda_abo import download_and_process_data
from user_management import user_management_system, logout
from buscar_info import streamlit_app
from facturacion_servicios_abo import generar_factura
from estadisticas_reservas_abo import main_reservas_abo
from estadisticas_facturacion_abo import main_factura
from whatsapp_sender_abo import whatsapp_sender
from ticket_support_app import soporte
from Reposistorio_Prompts_Juridicos.legal_prompts_app import legal_prompts_main
import datetime as dt
from openpyxl import load_workbook
from calendar import month_name
from datetime import date
import os
import datetime
import time
import pytz
import toml
from PIL import Image
import sys
import logging

# Configurar la página y ocultar la opción "Manage app" del menú
st.set_page_config(
    page_title="JURIDICO",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    }
)

# Ocultar elementos de la interfaz de Streamlit usando CSS personalizado
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}  /* Oculta el menú hamburguesa */
            footer {visibility: hidden;}  /* Oculta el footer "Made with Streamlit" */
            header {visibility: hidden;}  /* Oculta la cabecera */
            .stDeployButton {display:none;}  /* Oculta el botón de deploy */
            .css-1rs6os {visibility: hidden;}  /* Oculta el menú de configuración */
            .css-14xtw13 {visibility: hidden;}  /* Para algunas versiones de Streamlit */
            .css-1avcm0n {visibility: hidden;}  /* Para algunas versiones de Streamlit (menú hamburguesa) */
            
            /* En algunas versiones más recientes se usan diferentes clases CSS */
            /* Puedes identificar las clases específicas usando inspeccionar elemento en tu navegador */
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

st.cache_data.clear()
st.cache_resource.clear()

def clear_session_state():
    for key in list(st.session_state.keys()):
        del st.session_state[key]
 
def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))

logging.basicConfig(level=logging.DEBUG, filename='main_abo.log', filemode='w',
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Cargar configuraciones desde config.toml
with open("./.streamlit/config.toml", "r") as f:
    config = toml.load(f)

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "5"
os.environ["REQUESTS_READ_TIMEOUT"] = "5"

logo = Image.open("./assets/logoJAGT.ico")

datos_book = load_workbook("archivos/parametros_abogados.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]

    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)   
      data.append(_row[0])
    return data

fecha_hasta = int('20261228')

fecha = dt.datetime.now()
fecha_hoy = int(fecha.strftime("%Y%m%d"))

persona = dataBook("sw")
result_per = np.setdiff1d(persona,'')

sw_persona = result_per
  
page_title = 'Agenda Actividad' 
page_icon= "assets/plantilla-logo-bufete-abogados.avif" 
title="Resevas"
layout = 'centered'

# Inicializar el estado de sesión si no existe
if 'current_page' not in st.session_state:
    st.session_state.current_page = 'Inicio'

class Model:
  
  menuTitle = "Reserve y Agende en Linea"
  option1  = 'Inicio'
  option10 = 'Descargar Agenda'
  option9  = 'Consultar Agenda'
  option5  = 'Nuestros Servicios'
  option6  = 'Mas Informacion'
  option7  = 'Generar Archivos'
  option8  = 'Generar Codigo QR'
  option17 = 'Repositorio de Prompts Juridicos'
  option11 = 'Buscar Informacion'
  option15 = 'Enviar Whatsapp'
  option12 = 'Facturacion'
  option13 = 'Estadisticas de Resservas'
  option14 = 'Estadisticas de Facturacion'
  option16 = 'Soporte - PQRS'
    
  def add_app(self,title, function):
      self.apps.append({
         "title":title,
         "function":function
       })
      
  def css_load(css_file):
    try:
      with open(css_file, "r") as file:
        st.markdown(f"<style>{file.read()}</style>",unsafe_allow_html=True)
    except IOError  as e:
      if "not readable" in str(e):
        print("Error no readable check mode")
      else:
        print(f"IOError {e}")
      
  css_load(r"style/main.css")

if fecha_hasta < fecha_hoy:
  
   sw_persona == ['False']
      
   st.warning('Ha caducado el tiempo autorizado para su uso favor comuniquese con el administrador')

else:  

  # Función para cambiar entre páginas
  def cambiar_pagina(page_name):
    st.session_state.current_page = page_name
    st.rerun()
  
  def view(model):
    
    try:
      current_page = st.session_state.current_page
      
      # SIDEBAR PRINCIPAL (página de Inicio)
      if current_page == 'Inicio':
          with st.sidebar:
              st.markdown("""
              <style>
                  .stButton>button {
                      width: 100%;
                      border-radius: 10px;
                      height: 3em;
                      background-color: #FF6B6B;
                      color: white;
                  }
                  .stButton>button:hover {
                      background-color: #FF5252;
                      color: white;
                  }
              </style>
              """, unsafe_allow_html=True)
              
              app = option_menu(
                  menu_title=model.menuTitle,
                  options=[
                      model.option1,
                      model.option10,
                      model.option9,
                      model.option5,
                      model.option6,
                      model.option7,
                      model.option8,
                      model.option17,
                      model.option11,
                      model.option15,
                      model.option12,
                      model.option13,
                      model.option14,
                      model.option16
                  ],
                  icons=[
                      'house-fill',
                      'cloud-download-fill',
                      'calendar-check-fill',
                      'clipboard-check-fill',
                      'info-circle-fill',
                      'file-earmark-excel-fill',
                      'qr-code',
                      'book-fill',
                      'search',
                      'whatsapp',
                      'receipt',
                      'bar-chart-fill',
                      'pie-chart-fill',
                      'headset'
                  ],
                  menu_icon='calendar-date',
                  default_index=0,
                  orientation='vertical',
                  styles={
                      "container": {"padding": "5px", "background-color": '#262626'},
                      "icon": {"color": "white", "font-size": "23px"},
                      "nav-link": {"color":"white","font-size": "16px", "text-align": "left", "margin":"0px", "--hover-color": "#FF6B6B"},
                      "nav-link-selected": {"background-color": "#FF6B6B"}
                  }
              )

          st.markdown("""
          <style>
            [data-testid=stSidebar] {
                background-color: #262626;
            }
          </style>
          """, unsafe_allow_html=True)

          # Actualizarif app == model.option:
          if app == model.option10:
              cambiar_pagina('download_and_process_data')
          elif app == model.option9:
              cambiar_pagina('ConsultarAgenda_standalone')
          elif app == model.option5:
              cambiar_pagina('Servicios')
          elif app == model.option6:
              cambiar_pagina('Informacion')
          elif app == model.option7:
              cambiar_pagina('GenerarExcel')
          elif app == model.option8:
              cambiar_pagina('GenerarQr_standalone')
          elif app == model.option17:
              cambiar_pagina('legal_prompts_main')
          elif app == model.option11:
              cambiar_pagina('streamlit_app')
          elif app == model.option15:
              cambiar_pagina('whatsapp_sender')
          elif app == model.option12:
              cambiar_pagina('generar_factura')
          elif app == model.option13:
              cambiar_pagina('main_reservas_abo')
          elif app == model.option14:
              cambiar_pagina('main_factura')
          elif app == model.option16:
              cambiar_pagina('soporte')

      # SIDEBARS ESPECÍFICOS PARA CADA PÁGINA
      elif current_page == 'download_and_process_data':
         with st.sidebar:
              st.header("🏢 Descargar Agenda")
              if st.button("🏠 Volver al Inicio", key="btn_back_download"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'ConsultarAgenda_standalone':
         with st.sidebar:
              st.header("🏢 Consultar Clientes")
              if st.button("🏠 Volver al Inicio", key="btn_back_consultar"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'Servicios':
         with st.sidebar:
              st.header("🏢 Nuestros Servicios")
              if st.button("🏠 Volver al Inicio", key="btn_back_servicios"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'Informacion':
         with st.sidebar:
              st.header("🏢 Información")
              if st.button("🏠 Volver al Inicio", key="btn_back_info"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'GenerarExcel':
         with st.sidebar:
              st.header("🏢 Generar Excel")
              if st.button("🏠 Volver al Inicio", key="btn_back_excel"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'GenerarQr_standalone':
         with st.sidebar:
              st.header("🏢 Generar Código QR")
              if st.button("🏠 Volver al Inicio", key="btn_back_qr"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'legal_prompts_main':
         with st.sidebar:
              st.header("🏢 Repositorio Prompts")
              if st.button("🏠 Volver al Inicio", key="btn_back_prompts"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'streamlit_app':
          with st.sidebar:
              st.header("🏢 Buscar Información")
              if st.button("🏠 Volver al Inicio", key="btn_back_buscar"):
                  cambiar_pagina('Inicio')
              st.divider()
                  
      elif current_page == 'generar_factura':
         with st.sidebar:
              st.header("🏢 Generar Factura")
              if st.button("🏠 Volver al Inicio", key="btn_back_factura"):
                  cambiar_pagina('Inicio')
              st.divider() 
                    
      elif current_page == 'main_reservas_abo':
         with st.sidebar:
              st.header("🏢 Estadísticas de Reservas")
              if st.button("🏠 Volver al Inicio", key="btn_back_estadisticas"):
                  cambiar_pagina('Inicio')
              st.divider()  
             
      elif current_page == 'main_factura':
         with st.sidebar:
              st.header("🏢 Estadísticas de Facturación")
              if st.button("🏠 Volver al Inicio", key="btn_back_facturacion"):
                  cambiar_pagina('Inicio')
              st.divider()  
          
      elif current_page == 'whatsapp_sender':
         with st.sidebar:
              st.header("🏢 Enviar WhatsApp")
              if st.button("🏠 Volver al Inicio", key="btn_back_whatsapp"):
                  cambiar_pagina('Inicio')
              st.divider()  
        
      elif current_page == 'soporte':
          with st.sidebar:
              st.header("🏢 Soporte - PQRS")
              if st.button("🏠 Volver al Inicio", key="btn_back_soporte"):
                  cambiar_pagina('Inicio')
              st.divider() 
      
      # Crear columnas para el diseño (solo en página de Inicio)
      if current_page == 'Inicio':
          col1, col2, col3, col4 = st.columns(4)

          # Columna 1: Reloj Digital
          with col1:
                clock_placeholder = st.empty()
                    
          with col2:
                st.image(logo, width=150)
            
          # Columna 2: Calendario
          with col3:
                calendar_placeholder = st.empty()
              
          with col4:
              with st.form(key='myform4',clear_on_submit=True):
                  limpiar = st.form_submit_button("Limpiar Opcion")
                  if limpiar:
                    clear_session_state()
                    st.rerun()
        
          # Crear los tabs con los estilos personalizados
          tabs = st.tabs(["Inicio", "Crear Reserva", "Modificar Reserva", "Eliminar Reserva", "Informacion", "Servicios"])
        
          with tabs[0]:
                Inicio().view(Inicio.Model())
                 
          with tabs[1]:
                CrearReserva().view(CrearReserva.Model())
        
          with tabs[2]:
                ModificarReservaAbo().view(ModificarReservaAbo.Model())
        
          with tabs[3]:
                EliminarReserva().view(EliminarReserva.Model())
          
          with tabs[4]:
                Informacion().view(Informacion.Model())
              
          with tabs[5]:
                Servicios().view(Servicios.Model())
          
          def update_clock_and_calendar():
              while True:
                  now = datetime.datetime.now(pytz.timezone('America/Bogota'))
                  clock_placeholder.markdown(f'<p class="clock">Hora: {now.strftime("%H:%M:%S")}</p>', unsafe_allow_html=True)

                  today = date.today()
                  meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
                  mes_es = meses_es[today.month - 1]
                  dias_es = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
                  dia_es = dias_es[today.weekday()]
            
                  calendar_placeholder.markdown(f'<p class="calendar">Día: {dia_es.capitalize()} {today.day} de {mes_es} de {today.year}<br></p>', unsafe_allow_html=True)
            
                  time.sleep(5)
      
      # EJECUTAR LAS FUNCIONES SEGÚN LA PÁGINA SELECCIONADA
      else:
          # Mostrar el contenido de la página seleccionada en el área principal
          if current_page == 'download_and_process_data':
              download_and_process_data('./.streamlit/secrets.toml')
              
          elif current_page == 'ConsultarAgenda_standalone':
              ConsultarAgenda_standalone()
              
          elif current_page == 'Servicios':
              Servicios().view(Servicios.Model())
              
          elif current_page == 'Informacion':
              Informacion().view(Informacion.Model())
              
          elif current_page == 'GenerarExcel':
              GenerarExcel().view(GenerarExcel.Model())
              
          elif current_page == 'GenerarQr_standalone':
              GenerarQr_standalone()
              
          elif current_page == 'legal_prompts_main':
              legal_prompts_main()
              
          elif current_page == 'streamlit_app':
              streamlit_app()
              
          elif current_page == 'generar_factura':
              generar_factura()
              
          elif current_page == 'main_reservas_abo':
              main_reservas_abo()
              
          elif current_page == 'main_factura':
              main_factura()
              
          elif current_page == 'whatsapp_sender':
              whatsapp_sender()
              
          elif current_page == 'soporte':
              soporte()

      with st.sidebar:
            st.markdown("---")
            st.text("Version: 0.0.1")
            st.text("Ano: 2024")
            st.text("Autor: JAGT")
            st.markdown("---")

            st.markdown("""
            <script>
              window.onerror = function(message, source, lineno, colno, error) {
              fetch('/log_error', {
              method: 'POST',
              headers: {'Content-Type': 'main_abo/json'},
              body: JSON.stringify({message, source, lineno, colno, error: error.stack})
                });
              };
            </script>
            """, unsafe_allow_html=True)
  
    except SystemError as err:
      raise Exception(f'A ocurrido un error en main_abo.py: {err}')
    except Exception as e:
        st.error(f"Ocurrió un error en main_abo.py: {e}")
    
    sys.excepthook = global_exception_handler
        
  view(Model())
