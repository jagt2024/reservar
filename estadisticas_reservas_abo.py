# Función para cargar las credenciales y conectar con Google Sheets
import streamlit as st
import pandas as pd
import plotly.express as px
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import toml
from openpyxl import load_workbook

# secrets = toml.load('./.stream/secrets.toml')
datos_book = load_workbook("archivos/parametros_abogados.xlsx", read_only=False)

# secrets = toml.load('./.stream/secrets.toml')
def dataBookSheetUrl(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row)
      url = _row[1]
    return url

def load_credentials():
    with open('./.streamlit/secrets.toml', 'r') as toml_file:
        config = toml.load(toml_file)
        creds = config['sheetsemp']['credentials_sheet']
    return creds

sheetUrl = dataBookSheetUrl("sw")    

def get_google_sheet_data(creds):
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_info(creds, scopes=scope)
    client = gspread.authorize(credentials)

    sheet_url = str(sheetUrl)
    sheet = client.open_by_url(sheet_url)
    worksheet = sheet.worksheet('reservas')
    data = worksheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

# Cargar y procesar datos
@st.cache_data
def load_data():
    creds = load_credentials()
    df = get_google_sheet_data(creds)
    
    # Intentar diferentes formatos de fecha
    date_formats = ['%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']
    
    for date_format in date_formats:
        try:
            df['FECHA'] = pd.to_datetime(df['FECHA'], format=date_format)
            break
        except ValueError:
            continue
    
    if not pd.api.types.is_datetime64_any_dtype(df['FECHA']):
        # Si ningún formato funciona, intentar con parse_dates
        df['FECHA'] = pd.to_datetime(df['FECHA'], infer_datetime_format=True, errors='coerce')
    
    # Eliminar filas con fechas inválidas
    df = df.dropna(subset=['FECHA'])
    
    return df

# Aplicación Streamlit
def main():
    st.title("Estadísticas de Reservas - Servicios y Encargados")

    try:
        # Cargar datos
        df = load_data()

        # Filtro de fechas
        st.sidebar.header("Filtros")
        date_range = st.sidebar.date_input("Seleccione rango de fechas",
                                           [df['FECHA'].min(), df['FECHA'].max()],
                                           min_value=df['FECHA'].min(),
                                           max_value=df['FECHA'].max())
        
        start_date, end_date = date_range
        filtered_df = df[(df['FECHA'] >= pd.to_datetime(start_date)) & 
                         (df['FECHA'] <= pd.to_datetime(end_date))]

        # Métricas generales
        st.header("Métricas Generales")
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Reservas", len(filtered_df))
        col2.metric("Servicios Únicos", filtered_df['SERVICIOS'].nunique())
        col3.metric("Encargados Únicos", filtered_df['ENCARGADO'].nunique())

        # Gráfico de barras: Servicios más solicitados
        st.header("Servicios Más Solicitados")
        services_count = filtered_df['SERVICIOS'].value_counts().head(10)
        fig = px.bar(services_count, x=services_count.index, y=services_count.values,
                     labels={'x': 'Servicio', 'y': 'Cantidad'})
        st.plotly_chart(fig)

        # Gráfico circular: Distribución de servicios por encargado
        st.header("Distribución de Servicios por Encargado")
        encargado_servicios = filtered_df.groupby('ENCARGADO')['SERVICIOS'].count()
        fig = px.pie(encargado_servicios, values='SERVICIOS', names=encargado_servicios.index,
                     title='Distribución de Servicios por Encargado')
        st.plotly_chart(fig)

        # Gráfico de líneas: Reservas por día
        st.header("Reservas por Día")
        daily_reservations = filtered_df.groupby('FECHA').size().reset_index(name='count')
        fig = px.line(daily_reservations, x='FECHA', y='count',
                      labels={'count': 'Número de Reservas', 'FECHA': 'Fecha'})
        st.plotly_chart(fig)

        # Tabla de resumen de encargados
        st.header("Resumen de Encargados")
        encargado_summary = filtered_df.groupby('ENCARGADO').agg({
            'SERVICIOS': ['count', 'nunique']
        })
        encargado_summary.columns = ['Total Servicios', 'Servicios Únicos']
        encargado_summary = encargado_summary.reset_index()
        st.dataframe(encargado_summary)

        # Servicios más populares por encargado
        st.header("Servicios Más Populares por Encargado")
        encargado_select = st.selectbox("Seleccione un Encargado", 
                                        options=['Todos'] + list(filtered_df['ENCARGADO'].unique()))
        
        if encargado_select == 'Todos':
            encargado_services = filtered_df
        else:
            encargado_services = filtered_df[filtered_df['ENCARGADO'] == encargado_select]
        
        top_services = encargado_services['SERVICIOS'].value_counts().head(5)
        fig = px.bar(top_services, x=top_services.index, y=top_services.values,
                     labels={'x': 'Servicio', 'y': 'Cantidad'})
        st.plotly_chart(fig)

    except Exception as e:
        st.error(f"Se produjo un error al cargar o procesar los datos: {str(e)}")
        st.error("Por favor, verifique sus credenciales y la conexión a Internet.")

#if __name__ == "__main__":
#    main()