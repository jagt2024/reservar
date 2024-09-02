import streamlit as st
import pandas as pd
import base64
import io
from openpyxl import load_workbook

def load_excel(file):
    xls = pd.ExcelFile(file)
    return xls

def get_table_download_link(original_file, edited_sheet, edited_sheet_name, filename="datos_editados.xlsx"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Copiar todas las hojas del archivo original
        book = load_workbook(original_file)
        for sheet_name in book.sheetnames:
            if sheet_name != edited_sheet_name:
                # Si no es la hoja editada, copiar la hoja original
                df = pd.read_excel(original_file, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Si es la hoja editada, usar los datos editados
                edited_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
    
    excel_data = output.getvalue()
    b64 = base64.b64encode(excel_data).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Descargar datos editados (Excel)</a>'
    return href

def main():
    st.title("Editor de Excel con Streamlit")
    
    # Cargar archivo
    uploaded_file = st.file_uploader("Escoge un archivo Excel", type=['xlsx', 'xls'])
    
    if uploaded_file is not None:
        xls = load_excel(uploaded_file)
        st.success("Archivo cargado exitosamente!")
        
        # Seleccionar hoja
        sheet_name = st.selectbox("Selecciona una hoja para editar", xls.sheet_names)
        df = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Seleccionar columnas
        all_columns = df.columns.tolist()
        selected_columns = st.multiselect("Selecciona las columnas a mostrar", all_columns, default=all_columns)
        
        # Número de registros
        num_records = st.number_input("Número de registros a mostrar", min_value=1, max_value=len(df), value=min(10, len(df)))
        
        # Ordenar por columna
        sort_column = st.selectbox("Ordenar por", ["Sin ordenar"] + all_columns)
        
        # Aplicar filtros y ordenamiento
        if selected_columns:
            df_view = df[selected_columns]
        else:
            df_view = df
        
        if sort_column != "Sin ordenar":
            df_view = df_view.sort_values(by=sort_column, ascending=False)  # Ordenar de forma descendente
        
        df_view = df_view.head(num_records)
        
        # Mostrar datos en una tabla editable
        edited_df = st.data_editor(df_view)
        
        # Opciones para añadir, modificar o eliminar filas
        st.subheader("Opciones de edición")
        edit_option = st.radio("Seleccione una opción:", ["Añadir fila", "Eliminar fila"])
        
        if edit_option == "Añadir fila":
            new_row = {}
            for col in selected_columns:
                new_row[col] = st.text_input(f"Valor para {col}")
            
            if st.button("Añadir nueva fila"):
                edited_df = pd.concat([edited_df, pd.DataFrame([new_row])], ignore_index=True)
                st.success("Nueva fila añadida!")
        
        elif edit_option == "Eliminar fila":
            row_to_delete = st.number_input("Número de fila a eliminar", min_value=0, max_value=len(edited_df)-1)
            
            if st.button("Eliminar fila"):
                edited_df = edited_df.drop(edited_df.index[row_to_delete])
                st.success(f"Fila {row_to_delete} eliminada!")
        
        # Mostrar datos actualizados
        st.subheader("Datos actualizados")
        st.write(edited_df)
        
        # Descargar datos editados
        st.markdown(get_table_download_link(uploaded_file, edited_df, sheet_name), unsafe_allow_html=True)

if __name__ == "__main__":
    main()
