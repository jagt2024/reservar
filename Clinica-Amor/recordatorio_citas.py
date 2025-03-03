import streamlit as st
import pandas as pd
import smtplib
import time
import toml
import numpy as np
import gspread
import urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
import os
#import pywhatkit
from openpyxl import load_workbook
import schedule
import threading
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError

# Configuración de la página
#st.set_page_config(
#    page_title="Sistema de Recordatorio de Citas AMO",
#    page_icon="📅",
#    layout="wide"
#)

# Constantes para reintentos
MAX_RETRIES = 5
INITIAL_RETRY_DELAY = 1

datos_book = load_workbook("./archivos-amo/parametros_empresa.xlsx", read_only=False)

# Configuración de logs
def log_activity(message, error=False):
    """Registra actividad del sistema con marca de tiempo"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {'ERROR: ' if error else ''}{message}"
    
    if 'system_logs' not in st.session_state:
        st.session_state['system_logs'] = []
    
    st.session_state['system_logs'].insert(0, log_entry)
    
    # Mantener solo los últimos 100 registros
    if len(st.session_state['system_logs']) > 100:
        st.session_state['system_logs'] = st.session_state['system_logs'][:100]
    
    # Imprimir en consola para debugging
    print(log_entry)

# Función para manejar llamadas a la API
def api_call_handler(func):
    try:
        return func()
    except gspread.exceptions.APIError as e:
        if e.response.status_code == 429:
            log_activity("Límite de cuota excedida. Esperando...", error=True)
            time.sleep(10)
            return func()
        else:
            log_activity(f"Error de API: {str(e)}", error=True)
            raise e

def dataBookTelEnc(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
        #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[2]
       break
    return data

# Función para obtener teléfono del encargado desde Google Sheets
def obtener_telefono_encargado(worksheet, nombre_encargado):
    try:
        # Primero intentamos obtener la hoja de encargados
        sheet = worksheet.spreadsheet
        try:
            ws_encargados = sheet.worksheet('encargado')
        except:
            log_activity(f"No se encontró la hoja 'encargado'. Verificar el nombre de la hoja.", error=True)
            return None
        
        # Buscar el encargado por nombre
        try:
            cell = ws_encargados.find(nombre_encargado)
            if cell:
                # Asumimos que el teléfono está en la tercera columna (ajustar según estructura)
                row = ws_encargados.row_values(cell.row)
                if len(row) >= 3:
                    return row[2]  # Índice 2 (tercera columna)
            
            log_activity(f"No se encontró el teléfono para el encargado: {nombre_encargado}", error=True)
            return None
        except Exception as e:
            log_activity(f"Error al buscar encargado: {str(e)}", error=True)
            return None
    except Exception as e:
        log_activity(f"Error al obtener teléfono del encargado: {str(e)}", error=True)
        return None

# Función para cargar credenciales desde el archivo secrets.toml
def load_credentials_from_secrets():
    try:
        # Cargar desde la ubicación estándar de secretos de Streamlit
        return st.secrets["sheetsemp"]["credentials_sheet"]
    except Exception as e:
        log_activity(f"Error al cargar las credenciales desde .streamlit/secrets.toml: {str(e)}", error=True)
        st.info("Verifique que el archivo secrets.toml existe en la carpeta .streamlit y tiene el formato correcto.")
        return None

# Función para obtener datos de Google Sheets
def get_google_sheet_data(creds):
    for intento in range(MAX_RETRIES):
        try:
            with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
                scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
                credentials = Credentials.from_service_account_info(creds, scopes=scope)
                client = gspread.authorize(credentials)
                sheet = client.open('gestion-reservas-amo')
                worksheet = sheet.worksheet('reservas')
                data = api_call_handler(lambda: worksheet.get_all_values())
                
                if not data:
                    log_activity("No se encontraron datos en la hoja de cálculo.", error=True)
                    return None, None
                
                df = pd.DataFrame(data[1:], columns=data[0])
                
                if df.empty:
                    log_activity("El DataFrame está vacío después de cargar los datos.", error=True)
                    return None, None
                    
                # Convertir la columna de fecha a datetime
                df['FECHA'] = pd.to_datetime(df['FECHA'])
                
                log_activity(f"Datos cargados correctamente. {len(df)} citas encontradas.")
                return df, worksheet
                
        except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    log_activity(f"Límite de cuota excedida. Esperando {delay} segundos...", error=True)
                    time.sleep(delay)
                    continue
                else:
                    log_activity("Se excedió el límite de intentos. Por favor, intenta más tarde.", error=True)
            else:
                log_activity(f"Error de la API: {str(error)}", error=True)
            return None, None

        except Exception as e:
            log_activity(f"Error al cargar los datos: {str(e)}", error=True)
            return None, None
    
    return None, None

# Función para enviar correo electrónico
def enviar_correo(destinatario, asunto, mensaje, remitente, password):
    try:
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = destinatario
        msg['Subject'] = asunto
        
        msg.attach(MIMEText(mensaje, 'plain'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remitente, password)
        server.send_message(msg)
        server.quit()
        
        log_activity(f"Correo enviado correctamente a {destinatario}")
        return True
    except Exception as e:
        log_activity(f"Error al enviar correo a {destinatario}: {str(e)}", error=True)
        return False

# MEJORADO: Función para enviar mensaje por WhatsApp usando alternativa sin navegador
def enviar_whatsapp(numero, mensaje):
    try:
        # Formatear número de teléfono (eliminar el + si existe)
        if numero.startswith('+'):
            numero = numero[1:]
        
        # Intento de usar método alternativo sin abrir navegador
        # Nota: Esto requiere que WhatsApp Web ya esté abierto y conectado
        try:
            # Esto intenta enviar un mensaje sin abrir el navegador y sin espera
            pass
            #pywhatkit.sendwhatmsg_instantly(
            #    f"+{numero}", 
            #    mensaje, 
            #    tab_close=True,
            #    close_time=10,  # Cerrar después de 5 segundos
            #    wait_time=20  # Esperar solo 10 segundos para cargar
            #)
            #log_activity(f"WhatsApp enviado correctamente a +{numero}")
            #return True
        except Exception as e1:
            # Si falla, intentar con el método tradicional
            log_activity(f"Error con método rápido de WhatsApp: {str(e1)}", error=True)
            
            # Obtener hora actual + 1 minuto para programar el envío
            ahora = datetime.now()
            hora = ahora.hour
            minuto = ahora.minute + 1
            
            # Ajustar si el minuto es mayor a 59
            if minuto >= 60:
                hora += 1
                minuto -= 60
            
            # Usar pywhatkit tradicional
            #pywhatkit.sendwhatmsg(
            #    f"+{numero}", 
            #    mensaje, 
            #    hora, 
            #    minuto, 
            #    wait_time=60,  # Aumentar tiempo de espera
            #    tab_close=True
            #)
            #log_activity(f"WhatsApp enviado con método tradicional a +{numero}")
            #return True
    except Exception as e:
        log_activity(f"Error al enviar WhatsApp a +{numero}: {str(e)}", error=True)
        return False

def generate_whatsapp_link(telefono, mensaje):
    # Genera un enlace para iniciar un chat de WhatsApp con un número y mensaje predefinido
    encoded_message = mensaje.replace(' ', '%20')
    
    #mensaje_codificado = urllib.parse.quote(mensaje)
    # Asegurarse de que el teléfono no tenga el signo +
    #if isinstance(telefono, str) and telefono.startswith('+'):
    #    telefono = telefono[1:]

    return f"https://wa.me/{telefono}?text={encoded_message}"
    #return f"https://wa.me/{telefono}?text={mensaje_codificado}"

# Nueva función para generar y almacenar enlaces de WhatsApp para todas las citas
def generar_enlaces_whatsapp(df_citas, config):
    if 'whatsapp_links' not in st.session_state:
        st.session_state['whatsapp_links'] = {}
    
    ahora = datetime.now()
        
    for idx, cita in df_citas.iterrows():
        try:
            fecha_cita = cita['FECHA'].date() if isinstance(cita['FECHA'], datetime) else datetime.strptime(cita['FECHA'], '%Y-%m-%d').date()
            hora_cita = cita['HORA']
            
            # Obtener teléfono del encargado
            nombre_encargado = cita['ENCARGADO']
            result_telenc = dataBookTelEnc("encargado", nombre_encargado)
            #print(f'result_telefono : {result_telenc}')
            
            # Fix: Better type handling for telephone number
            tel_encargado = None
            
            # Handle different types that might be returned from dataBookTelEnc
            if isinstance(result_telenc, np.ndarray):
                # If it's a numpy array, get the first non-empty value
                non_empty_values = [val for val in result_telenc if val]
                if non_empty_values:
                    tel_encargado = str(non_empty_values[0])
            elif isinstance(result_telenc, (list, tuple)):
                # If it's a list or tuple, get the first non-empty value
                non_empty_values = [val for val in result_telenc if val]
                if non_empty_values:
                    tel_encargado = str(non_empty_values[0])
            elif result_telenc:  # This handles strings, integers, etc.
                # If it's a scalar value and not empty/None
                tel_encargado = str(result_telenc)
            
            # Check if we got a valid phone number
            if not tel_encargado:
                log_activity(f"No se encontró teléfono válido para el encargado {nombre_encargado}", error=True)
                continue
                
            # Remove any non-numeric characters from the phone number
            tel_encargado = ''.join(filter(str.isdigit, tel_encargado))[:12]
            #print(f'tel_encargado1 : {tel_encargado}' )
            
            if not tel_encargado:
                log_activity(f"Teléfono sin dígitos válidos para el encargado {nombre_encargado}", error=True)
                continue
                
            # Generar identificador único para esta cita
            cita_id = f"{cita['NOMBRE']}_{fecha_cita}_{hora_cita}_{nombre_encargado}"
            
            # Citas para mañana
            if fecha_cita == (ahora + timedelta(days=1)).date():
                mensaje = f"Recordatorio: Tiene una cita mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el paciente {cita['NOMBRE']} para el servicio de {cita['SERVICIOS']}. {config['nombre_clinica']}"
                enlace = generate_whatsapp_link(tel_encargado, mensaje)
                #st.markdown(f"Click en el enlace si desea Enviar Whatsapp {enlace}")
                
                st.session_state['whatsapp_links'][cita_id] = {
                    'enlace': enlace,
                    'encargado': nombre_encargado,
                    'telefono': tel_encargado,
                    'tipo': 'mañana',
                    'mensaje': mensaje,
                    'nombre_paciente': cita['NOMBRE'],
                    'fecha': fecha_cita.strftime('%d/%m/%Y'),
                    'hora': hora_cita,
                    'servicio': cita['SERVICIOS']
                }
                            
            # Citas para hoy
            elif fecha_cita == ahora.date():
                mensaje = f"Recordatorio: Tiene una cita HOY {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el paciente {cita['NOMBRE']} para el servicio de {cita['SERVICIOS']}. {config['nombre_clinica']}"
                enlace = generate_whatsapp_link(tel_encargado, mensaje)
                #st.markdown(f"Click en el enlace si desea Enviar Whatsapp {enlace}")
                st.session_state['whatsapp_links'][cita_id] = {
                    'enlace': enlace,
                    'encargado': nombre_encargado,
                    'telefono': tel_encargado,
                    'tipo': 'hoy',
                    'mensaje': mensaje,
                    'nombre_paciente': cita['NOMBRE'],
                    'fecha': fecha_cita.strftime('%d/%m/%Y'),
                    'hora': hora_cita,
                    'servicio': cita['SERVICIOS']
                }
        
        except Exception as e:
            log_activity(f"Error al generar enlace WhatsApp para {cita.get('NOMBRE', 'desconocido')}: {str(e)}", error=True)


# Función para verificar las citas y enviar recordatorios/
def verificar_citas(df_citas, worksheet, config):
    if df_citas is None or worksheet is None:
        log_activity("No se pueden verificar citas: datos no disponibles", error=True)
        return
        
    ahora = datetime.now()
    log_activity(f"Iniciando verificación de citas a las {ahora.strftime('%H:%M:%S')}")
    
    mensajes_enviados = 0
    errores = 0

    # Generar enlaces de WhatsApp para todas las citas
    generar_enlaces_whatsapp(df_citas, config)
        
    #tel_encargado = dataBookTelEnc("encargado", nombre_encargado)
    # Filtrar citas para hoy y mañana
    for _, cita in df_citas.iterrows():
        try:
            fecha_cita = cita['FECHA'].date() if isinstance(cita['FECHA'], datetime) else datetime.strptime(cita['FECHA'], '%Y-%m-%d').date()
            hora_cita = cita['HORA']
            
            # Obtener teléfono del encargado
            nombre_encargado = cita['ENCARGADO']
            
            #tel_encargado = obtener_telefono_encargado(worksheet, nombre_encargado)
                       
            # Citas para mañana - enviar recordatorio hoy
            if fecha_cita == (ahora + timedelta(days=1)).date():
                log_activity(f"Procesando recordatorio para mañana: {cita['NOMBRE']} - {fecha_cita}")
                
                # Recordatorio al paciente por correo
                if config['correo_activo'] and cita['EMAIL']:
                    mensaje_paciente = f"""
                    Estimado/a {cita['NOMBRE']},
                    
                    Le recordamos que tiene una cita programada para mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} para el servicio de {cita['SERVICIOS']} con el/la encargado {cita['ENCARGADO']}.
                    
                    Por favor, confirme su asistencia o comuníquese si necesita reprogramar.
                    
                    Saludos cordiales,
                    {config['nombre_clinica']}
                    """
                    
                    exito = enviar_correo(
                        cita['EMAIL'], 
                        f"Recordatorio de Cita - {config['nombre_clinica']}", 
                        mensaje_paciente,
                        config['correo_remitente'],
                        config['password_correo']
                    )
                    
                    if exito:
                        mensajes_enviados += 1
                    else:
                        errores += 1
                
                    # Recordatorio al encargado por correo
                    mensaje_encargado = f"""
                    Estimado/a {cita['ENCARGADO']},
                
                    Le recordamos que tiene una cita programada para mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el/la paciente {cita['NOMBRE']}.
                
                    Servicio: {cita['SERVICIOS']}
                
                    Saludos cordiales,
                    {config['nombre_clinica']}
                    """
                
                    exito = enviar_correo(
                        cita['CORREO ENCARGADO'], 
                        f"Recordatorio de Cita - {config['nombre_clinica']}", 
                        mensaje_encargado,
                        config['correo_remitente'],
                        config['password_correo']
                    )

                    if exito:
                        mensajes_enviados += 1
                    else:
                        errores += 1

                # Recordatorio por WhatsApp
                #if config['whatsapp_activo']: #and cita['TELEFONO']:
                    # Mensaje al paciente
                #    mensaje_whatsapp_paciente = f"Recordatorio: Tiene una cita mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} para el servicio de {cita['SERVICIOS']} con el encargado {cita['ENCARGADO']}. {config['nombre_clinica']}"
                #    exito = enviar_whatsapp(cita['TELEFONO'], mensaje_whatsapp_paciente)
                    
                #    if exito:
                #        mensajes_enviados += 1
                #    else:
                #        errores += 1
                    
                    # Mensaje al encargado
                #    if tel_encargado:
                #mensaje_whatsapp_encargado = f"Recordatorio: Tiene una cita mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el paciente {cita['NOMBRE']} para el servicio de {cita['SERVICIOS']}. {config['nombre_clinica']}"

                 #       whatsapp_link = generate_whatsapp_link(str(tel_encargado), mensaje_whatsapp_encargado)

                        #st.markdown(f"Click en enlace para Enviar Whatsapp {whatsapp_link}")

                 #      log_activity(f"Enlace generado: {whatsapp_link}")
                         
                        #exito = generate_whatsapp_link(tel_encargado, mensaje_whatsapp_encargado)
                        
                        #if exito:
                        #    mensajes_enviados += 1
                        #else:
                        #    errores += 1
            
                    #else: 
                    #    log_activity("No hay un Telefono para el Encargado")
                    #    st.info("No hay un Telefono para el Encargado")

            # Citas para hoy - enviar recordatorio hoy temprano
            elif fecha_cita == ahora.date() and ahora.hour < 19:  # Antes de las 10 PM
                log_activity(f"Procesando recordatorio para HOY: {cita['NOMBRE']} - {fecha_cita}")
                
                # Recordatorio al paciente por correo
                if config['correo_activo'] and cita['EMAIL']:
                    mensaje_paciente = f"""
                    Estimado/a {cita['NOMBRE']},
                    
                    Le recordamos que tiene una cita programada para HOY {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} para el servicio de {cita['SERVICIOS']} con el/la encargado {cita['ENCARGADO']}.
                    
                    Saludos cordiales,
                    {config['nombre_clinica']}
                    """
                    
                    exito = enviar_correo(
                        cita['EMAIL'], 
                        f"Recordatorio de Cita HOY - {config['nombre_clinica']}", 
                        mensaje_paciente,
                        config['correo_remitente'],
                        config['password_correo']
                    )
                    
                    if exito:
                        mensajes_enviados += 1
                    else:
                        errores += 1
                
                    # Recordatorio al encargado por correo
                    mensaje_encargado = f"""
                    Estimado/a {cita['ENCARGADO']},
                
                    Le recordamos que tiene una cita programada para mañana {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el/la paciente {cita['NOMBRE']}.
                
                    Servicio: {cita['SERVICIOS']}
                
                    Saludos cordiales,
                    {config['nombre_clinica']}
                    """
                
                    exito = enviar_correo(
                        cita['CORREO ENCARGADO'], 
                        f"Recordatorio de Cita - {config['nombre_clinica']}", 
                        mensaje_encargado,
                        config['correo_remitente'],
                        config['password_correo']
                    )

                    if exito:
                        mensajes_enviados += 1
                    else:
                        errores += 1

                # Recordatorio por WhatsApp
               # if config['whatsapp_activo']: #and cita['TELEFONO']:
                #    # Mensaje al paciente
                #    mensaje_whatsapp_paciente = f"Recordatorio: Tiene una cita HOY {fecha_cita.#strftime('%d/%m/%Y')} a las {hora_cita} para el servicio de {cita#['SERVICIOS']} con el encargado {cita['ENCARGADO']}. {config['nombre_clinica']}"
                #    exito = enviar_whatsapp(cita['TELEFONO'], mensaje_whatsapp_paciente)
                    
                #    if exito:
                #        mensajes_enviados += 1
                #    else:
                #        errores += 1
                    
                    # Mensaje al encargado
                #    if tel_encargado:
                #        mensaje_whatsapp_encargado = f"Recordatorio: Tiene una cita HOY {fecha_cita.strftime('%d/%m/%Y')} a las {hora_cita} con el paciente {cita['NOMBRE']} para el servicio de {cita['SERVICIOS']}. {config['nombre_clinica']}"
                        
                 #       whatsapp_link = generate_whatsapp_link(str(tel_encargado), mensaje_whatsapp_encargado)
                        
                        #st.markdown(f"Click en enlace para Enviar Whatsapp {whatsapp_link}")
                        
                 #       log_activity(f"Enlace generado: {whatsapp_link}")

                        #exito = generate_whatsapp_link(tel_encargado, #mensaje_whatsapp_encargado)
                                                
                        #if exito:
                        #    mensajes_enviados += 1
                        #else:
                        #    errores += 1
                    #else: 
                    #    log_activity("No hay un Telefono para el Encargado")
                    #    st.info("No hay un Telefono para el Encargado")

                #        exito = enviar_whatsapp(tel_encargado, mensaje_whatsapp_encargado)
                        
                #        if exito:
                #            mensajes_enviados += 1
                #        else:
                #            errores += 1
        except Exception as e:
            log_activity(f"Error al procesar cita para {cita.get('NOMBRE', 'desconocido')}: {str(e)}", error=True)
            errores += 1
    
    log_activity(f"Verificación completada. Mensajes enviados: {mensajes_enviados}, Errores: {errores}")

# Función para ejecutar los recordatorios en segundo plano
def ejecutar_recordatorios(creds, config):
    # CORREGIDO: Programar la tarea para que se ejecute cada día a las 8 AM y también a las 12:30 PM
    schedule.every().day.at("08:00").do(ejecutar_verificacion_citas, creds=creds, config=config)
    #schedule.every().day.at("15:00").do(ejecutar_verificacion_citas, creds=creds, config=config)
    
    # Ejecutar inmediatamente al iniciar el sistema
    ejecutar_verificacion_citas(creds=creds, config=config)
    
    while True:
        try:
            schedule.run_pending()
            time.sleep(60)  # Revisar cada minuto
        except Exception as e:
            log_activity(f"Error en el hilo de recordatorios: {str(e)}", error=True)
            time.sleep(300)  # Esperar 5 minutos en caso de error

def ejecutar_verificacion_citas(creds, config):
    try:
        # Obtener datos actualizados de Google Sheets
        df_citas, worksheet = get_google_sheet_data(creds)
        if df_citas is not None and worksheet is not None:
            verificar_citas(df_citas, worksheet, config)
        else:
            log_activity("No se pudieron obtener datos para verificar citas", error=True)
    except Exception as e:
        log_activity(f"Error al ejecutar verificación de citas: {str(e)}", error=True)

# Función para iniciar el hilo de recordatorios
def iniciar_sistema_recordatorios(creds, config):
    if 'thread_recordatorios_activo' in st.session_state and st.session_state['thread_recordatorios_activo']:
        log_activity("El sistema de recordatorios ya está activo.")
        return None
        
    log_activity("Iniciando hilo de recordatorios...")
    thread = threading.Thread(target=ejecutar_recordatorios, args=(creds, config))
    thread.daemon = True  # El hilo se cerrará cuando el programa principal termine
    thread.start()
    
    # Marcar como activo
    st.session_state['thread_recordatorios_activo'] = True
    
    return thread

# Función para detener el sistema de recordatorios (para pruebas)
def detener_sistema_recordatorios():
    st.session_state['thread_recordatorios_activo'] = False
    log_activity("Sistema de recordatorios marcado para detener. Se detendrá en la próxima iteración.")

# Función para preparar datos para visualización
def preparar_datos_citas(df):
    # Asegurarse de que las columnas necesarias estén presentes
    columnas_requeridas = [
        'NOMBRE', 'EMAIL', 'TELEFONO', 'FECHA', 'HORA', 
        'SERVICIOS', 'ENCARGADO', 'CORREO ENCARGADO'
    ]
    
    # Verificar si faltan columnas
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        log_activity(f"Faltan las siguientes columnas en la hoja de cálculo: {', '.join(columnas_faltantes)}", error=True)
        return None
    
    # Seleccionar solo las columnas necesarias
    df_citas = df[columnas_requeridas].copy()
    
    # Formatear la fecha para mejor visualización
    df_citas['FECHA'] = df_citas['FECHA'].dt.strftime('%Y-%m-%d')
    
    return df_citas

# Función para probar envío de recordatorios
def probar_recordatorio(creds, config, email, telefono):
    try:
        mensaje_prueba = f"""
        Esto es un mensaje de prueba del Sistema de Recordatorio de Citas.
        
        Si recibe este mensaje, la configuración es correcta.
        
        Saludos,
        {config['nombre_clinica']}
        """
        
        resultados = []
        
        if config['correo_activo'] and email:
            exito_correo = enviar_correo(
                email,
                "PRUEBA - Sistema de Recordatorio de Citas",
                mensaje_prueba,
                config['correo_remitente'],
                config['password_correo']
            )
            resultados.append(f"Correo: {'✅ Enviado' if exito_correo else '❌ Error'}")
        
        if config['whatsapp_activo'] and telefono:
            enlace_whatsapp = generate_whatsapp_link(
                telefono,
                f"PRUEBA - Sistema de Recordatorio de Citas: {config['nombre_clinica']}"
            )
            resultados.append(f"WhatsApp: ✅ Enlace generado: {enlace_whatsapp}")
        
        return resultados
    except Exception as e:
        log_activity(f"Error en prueba de recordatorio: {str(e)}", error=True)
        return [f"Error general: {str(e)}"]

# Interfaz principal de Streamlit
def recordatorio():
    st.title("🏥 Sistema de Recordatorio de Citas")
    
    # Inicializar logs si no existen
    if 'system_logs' not in st.session_state:
        st.session_state['system_logs'] = []
    
    # Inicializar estado del hilo
    if 'thread_recordatorios_activo' not in st.session_state:
        st.session_state['thread_recordatorios_activo'] = False
    
    if 'whatsapp_links' not in st.session_state:
        st.session_state['whatsapp_links'] = {}

    # Cargar credenciales automáticamente desde secrets.toml
    creds = load_credentials_from_secrets()
    
    if creds is not None:
        
        # Configuración en la pantalla principal de forma horizontal
        st.header("⚙️ Configuración del Sistema")
        
        # Organizar en columnas
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Datos Generales")
            nombre_clinica = st.text_input("Nombre de la Clínica/Consultorio", "CLINICA DEL AMOR")
            
            st.subheader("Configuración de Correo")
            correo_activo = st.checkbox("Activar notificaciones por correo", value=True)
            correo_remitente = "josegarjagt@gmail.com"
            password_correo = "ulbiiydnlgkllkub"
        
        with col2:
            st.subheader("Configuración de WhatsApp")
            whatsapp_activo = st.checkbox("Activar notificaciones por WhatsApp", value=True)
            
            # Mostrar información de estado del sistema
            st.subheader("Estado del Sistema")
            estado = "✅ Activo" if st.session_state['thread_recordatorios_activo'] else "⚠️ Inactivo"
            st.info(f"Estado actual: {estado}")
        
        # Botones de acción en línea horizontal
        col_btn1, col_btn2, col_btn3 = st.columns(3)
        
        with col_btn1:
            if st.session_state['thread_recordatorios_activo']:
                if st.button("🛑 Detener Sistema"):
                    detener_sistema_recordatorios()
                    st.success("Sistema marcado para detener.")
            else:
                if st.button("▶️ Iniciar Sistema de Recordatorios"):
                    config = {
                        'nombre_clinica': nombre_clinica,
                        'correo_activo': correo_activo,
                        'correo_remitente': correo_remitente,
                        'password_correo': password_correo,
                        'whatsapp_activo': whatsapp_activo
                    }
                    
                    if correo_activo and (not correo_remitente or not password_correo):
                        st.error("Para activar notificaciones por correo, debe proporcionar un correo remitente y contraseña.")
                    else:
                        st.session_state['thread_recordatorios'] = iniciar_sistema_recordatorios(creds, config)
                        st.success("Sistema de recordatorios iniciado correctamente!")
        
        with col_btn2:
            if st.button("🔄 Actualizar Datos"):
                st.rerun()
        
        # Pestañas para diferentes secciones
        tab1, tab2, tab3 = st.tabs(["📋 Gestión de Citas", "📊 Envio Mensajes", "ℹ️ Ayuda"])
        
        with tab1:
            st.header("Gestión de Citas")
            
            # Cargar datos de Google Sheets
            with st.spinner("Cargando datos de Google Sheets..."):
                df, worksheet = get_google_sheet_data(creds)
                
                if df is not None:
                    # Preparar datos para visualización
                    df_citas = preparar_datos_citas(df)
                    
                    if df_citas is not None:
                        # Mostrar tabla de citas
                        st.subheader("Citas Programadas")
                        st.dataframe(df_citas)
                        
                        # Opción para actualizar datos
                        if st.button("Actualizar Datos"):
                            st.rerun()

                        # Generar enlaces de WhatsApp para todas las citas
                        config = {
                            'nombre_clinica': nombre_clinica,
                            'correo_activo': correo_activo,
                            'correo_remitente': correo_remitente,
                            'password_correo': password_correo,
                            'whatsapp_activo': whatsapp_activo
                        }
                        generar_enlaces_whatsapp(df_citas, config)
                else:
                    st.error("No se pudieron cargar los datos. Verifica la conexión y las credenciales.")
        
        with tab2:
            st.header("📱 Enviar Menseje WhatsApp (Opcional)")
            
            # Crear dos secciones para hoy y mañana
            col_hoy, col_manana = st.columns(2)
            
            # Obtener los enlaces de WhatsApp
            enlaces = st.session_state['whatsapp_links']
            
            # Filtrar por tipo
            enlaces_hoy = {k: v for k, v in enlaces.items() if v['tipo'] == 'hoy'}
            enlaces_manana = {k: v for k, v in enlaces.items() if v['tipo'] == 'mañana'}
            
            with col_hoy:
                st.subheader("Citas para HOY")
                if not enlaces_hoy:
                    st.info("No hay citas programadas para hoy.")
                else:
                    for cita_id, info in enlaces_hoy.items():
                        with st.expander(f"{info['nombre_paciente']} - {info['hora']}"):
                            st.write(f"**Paciente:** {info['nombre_paciente']}")
                            st.write(f"**Encargado:** {info['encargado']}")
                            st.write(f"**Servicio:** {info['servicio']}")
                            st.write(f"**Hora:** {info['hora']}")
                            st.write(f"**Teléfono del encargado:** {info['telefono']}")
                            st.write(f"**Mensaje:**\n{info['mensaje']}")
                            st.markdown(f"[📱 Enviar WhatsApp al Encargado]({info['enlace']})", unsafe_allow_html=True)
                            
            with col_manana:
                st.subheader("Citas para MAÑANA")
                if not enlaces_manana:
                    st.info("No hay citas programadas para mañana.")
                else:
                    for cita_id, info in enlaces_manana.items():
                        with st.expander(f"{info['nombre_paciente']} - {info['hora']}"):
                            st.write(f"**Paciente:** {info['nombre_paciente']}")
                            st.write(f"**Encargado:** {info['encargado']}")
                            st.write(f"**Servicio:** {info['servicio']}")
                            st.write(f"**Hora:** {info['hora']}")
                            st.write(f"**Teléfono del encargado:** {info['telefono']}")
                            st.write(f"**Mensaje:**\n{info['mensaje']}")
                            st.markdown(f"[📱 Enviar WhatsApp al Encargado]({info['enlace']})", unsafe_allow_html=True)

            # Log de actividad del sistema
            #st.subheader("Log de Actividad")
            
            # Botón para limpiar log
            #if st.button("Limpiar Log"):
            #    st.session_state['system_logs'] = []
            #    st.success("Log limpiado correctamente.")
            
            # Mostrar los últimos logs
            #for log_entry in st.session_state['system_logs']:
            #    if "ERROR" in log_entry:
            #        st.error(log_entry)
            #    else:
            #        st.text(log_entry)
        
            if st.session_state['thread_recordatorios_activo']:
                if st.button("Detener Sistema", key='det_sys'):
                    detener_sistema_recordatorios()
                    st.success("Sistema marcado para detener.")
    
        with tab3:
            st.header("Ayuda y Guía de Uso")
            
            st.subheader("Configuración Inicial")
            st.markdown("""
            1. **Configuración de Correo**:
               - Active la opción "Activar notificaciones por correo"
               - Ingrese su correo de Gmail
                           
            2. **Configuración de WhatsApp**:
               - Active la opción "Activar notificaciones por WhatsApp"
               - Se abrirá automáticamente WhatsApp Web para enviar mensajes
            """)
            
            st.subheader("Funcionamiento")
            st.markdown("""
            - El sistema enviará recordatorios:
              - Un día antes de la cita (recordatorio previo)
              - El mismo día de la cita (recordatorio el día de la cita)
            - Los recordatorios se enviarán tanto al paciente como al encargado
            - Para que el sistema funcione, debe mantener esta aplicación ejecutándose
            """)
            
            st.subheader("Problemas Comunes")
            st.markdown("""
            - **Error al cargar el archivo secrets.toml**: Verifique que exista en la carpeta .streamlit y tenga el formato correcto
            - **Error de conexión con Google Sheets**: Verifique las credenciales y permisos
            - **Error al enviar correos**: Verifique que la contraseña de aplicación sea correcta
            - **Error con WhatsApp**: Asegúrese de tener WhatsApp Web disponible
            """)
    else:
        st.error("No se pudieron cargar las credenciales desde .streamlit/secrets.toml")
        
if __name__ == "__main__":
    recordatorio()