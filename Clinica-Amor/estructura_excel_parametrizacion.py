import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os

# Crear un nuevo archivo Excel
wb = Workbook()

# Configurar la primera hoja para Formatos
ws1 = wb.active
ws1.title = "Formatos"

# Añadir encabezados con formato
headers_formatos = ["TipoTerapia", "ColorFondo", "ColorTexto", "Plantilla", "Logo", "ConfiguracionAdicional"]
for col_num, header in enumerate(headers_formatos, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

# Añadir datos de ejemplo para Formatos
formatos_data = [
    ["Fisioterapia", "#E6F3FF", "#000000", "plantilla_fisio.html", "logo_fisio.png", 
     json.dumps({"mostrarCalendario": True, "permitirImprimir": True, "maxSesiones": 20})],
    ["Psicología", "#FFF4E6", "#333333", "plantilla_psico.html", "logo_psico.png", 
     json.dumps({"mostrarCalendario": True, "permitirImprimir": True, "duracionSesion": 60})],
    ["Ocupacional", "#E6FFEA", "#003300", "plantilla_ocupacional.html", "logo_ocupacional.png", 
     json.dumps({"mostrarCalendario": True, "permitirImprimir": True, "requiereInforme": True})],
    ["Fonoaudiología", "#F0E6FF", "#330033", "plantilla_fono.html", "logo_fono.png", 
     json.dumps({"mostrarCalendario": True, "permitirImprimir": True, "permitirAudios": True})],
    ["Hidroterapia", "#E6F9FF", "#003366", "plantilla_hidro.html", "logo_hidro.png", 
     json.dumps({"mostrarCalendario": True, "permitirImprimir": True, "temperaturaAgua": 35})]
]

for row_num, row_data in enumerate(formatos_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = Alignment(vertical='center')
        if col_num == 6:  # JSON column
            cell.alignment = Alignment(wrap_text=True)

# Ajustar ancho de columnas
for col in ws1.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = (max_length + 2) 
    ws1.column_dimensions[column].width = adjusted_width

# Crear una segunda hoja para Campos Obligatorios
ws2 = wb.create_sheet(title="CamposObligatorios")

# Añadir encabezados con formato
headers_campos = ["Normativa", "Campo", "Descripcion", "Tipo", "Validacion"]
for col_num, header in enumerate(headers_campos, 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

# Añadir datos de ejemplo para Campos Obligatorios
campos_data = [
    ["ResolucionMin001", "IdentificacionPaciente", "Número de documento de identidad del paciente", "texto", "^[0-9]{6,12}$"],
    ["ResolucionMin001", "NombreCompleto", "Nombre completo del paciente", "texto", "^[a-zA-Z ]{2,100}$"],
    ["ResolucionMin001", "FechaNacimiento", "Fecha de nacimiento del paciente", "fecha", ""],
    ["ResolucionMin001", "DiagnosticoPrincipal", "Diagnóstico principal según CIE-10", "texto", "^[A-Z][0-9]{2}(\\.[0-9])?$"],
    ["ResolucionMin002", "FechaInicio", "Fecha de inicio del tratamiento", "fecha", ""],
    ["ResolucionMin002", "ProfesionalResponsable", "Nombre del profesional responsable", "texto", "^[a-zA-Z ]{2,100}$"],
    ["ResolucionMin002", "NumeroTarjetaProfesional", "Número de tarjeta profesional", "texto", "^[0-9]{6,12}$"],
    ["ResolucionMin002", "PlanTratamiento", "Descripción del plan de tratamiento", "texto", ""],
    ["NormaISO9001", "Consentimiento", "Firma del consentimiento informado", "booleano", ""],
    ["NormaISO9001", "HistoriaClinica", "Número de historia clínica", "texto", "^HC[0-9]{6}$"],
    ["NormativaLocal", "EPS", "Entidad prestadora de salud", "seleccion", "EPS_Sura,Sanitas,Nueva EPS,Compensar,Otras"],
    ["NormativaLocal", "MunicipioProcedencia", "Municipio de procedencia", "texto", ""],
    ["NormativaLocal", "Telefono", "Número de teléfono de contacto", "texto", "^[0-9]{7,10}$"]
]

for row_num, row_data in enumerate(campos_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = Alignment(vertical='center')

# Ajustar ancho de columnas para la segunda hoja
for col in ws2.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = (max_length + 2)
    ws2.column_dimensions[column].width = adjusted_width

# Crear una tercera hoja para Formularios de Evaluación
ws3 = wb.create_sheet(title="FormulariosEvaluacion")

# Añadir encabezados con formato
headers_formularios = ["IDFormulario", "Nombre", "Descripcion", "TipoTerapia", "Seccion", "CamposSeccion"]
for col_num, header in enumerate(headers_formularios, 1):
    cell = ws3.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

# Crear campos para formularios en formato JSON
campos_evaluacion_inicial = json.dumps([
    {"id": "motivo_consulta", "nombre": "Motivo de Consulta", "tipo": "texto", "obligatorio": True, "placeholder": "Describa el motivo de consulta..."},
    {"id": "antecedentes", "nombre": "Antecedentes", "tipo": "texto", "obligatorio": True, "placeholder": "Describa los antecedentes relevantes..."},
    {"id": "expectativas", "nombre": "Expectativas del Tratamiento", "tipo": "texto", "obligatorio": True, "placeholder": "Describa las expectativas..."}
])

campos_evaluacion_dolor = json.dumps([
    {"id": "nivel_dolor", "nombre": "Nivel de Dolor (1-10)", "tipo": "numero", "obligatorio": True, "min": 1, "max": 10},
    {"id": "ubicacion_dolor", "nombre": "Ubicación del Dolor", "tipo": "texto", "obligatorio": True},
    {"id": "factores_dolor", "nombre": "Factores que Aumentan el Dolor", "tipo": "texto", "obligatorio": False}
])

campos_evaluacion_movilidad = json.dumps([
    {"id": "rango_flexion", "nombre": "Rango de Flexión (°)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 180},
    {"id": "rango_extension", "nombre": "Rango de Extensión (°)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 180},
    {"id": "rango_abduccion", "nombre": "Rango de Abducción (°)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 180}
])

campos_evaluacion_fuerza = json.dumps([
    {"id": "fuerza_mmss", "nombre": "Fuerza Miembros Superiores (0-5)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 5},
    {"id": "fuerza_mmii", "nombre": "Fuerza Miembros Inferiores (0-5)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 5},
    {"id": "fuerza_tronco", "nombre": "Fuerza Tronco (0-5)", "tipo": "numero", "obligatorio": True, "min": 0, "max": 5}
])

campos_evaluacion_emocional = json.dumps([
    {"id": "estado_animo", "nombre": "Estado de Ánimo", "tipo": "seleccion", "obligatorio": True, "opciones": "Muy bajo,Bajo,Normal,Elevado,Muy elevado"},
    {"id": "ansiedad", "nombre": "Nivel de Ansiedad", "tipo": "seleccion", "obligatorio": True, "opciones": "Ausente,Leve,Moderada,Grave,Muy grave"},
    {"id": "depresion", "nombre": "Síntomas de Depresión", "tipo": "seleccion", "obligatorio": True, "opciones": "Ausente,Leve,Moderada,Grave,Muy grave"}
])

campos_evaluacion_cognitiva = json.dumps([
    {"id": "orientacion", "nombre": "Orientación", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Alterada"},
    {"id": "memoria", "nombre": "Memoria", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Levemente alterada,Moderadamente alterada,Gravemente alterada"},
    {"id": "atencion", "nombre": "Atención", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Alterada"}
])

campos_evaluacion_habla = json.dumps([
    {"id": "articulacion", "nombre": "Articulación", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Alterada"},
    {"id": "fluidez", "nombre": "Fluidez", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Alterada"},
    {"id": "comprension", "nombre": "Comprensión", "tipo": "seleccion", "obligatorio": True, "opciones": "Normal,Alterada"}
])

campos_evaluacion_seguimiento = json.dumps([
    {"id": "progreso", "nombre": "Progreso Observado", "tipo": "texto", "obligatorio": True},
    {"id": "complicaciones", "nombre": "Complicaciones", "tipo": "texto", "obligatorio": False},
    {"id": "ajustes_tratamiento", "nombre": "Ajustes al Tratamiento", "tipo": "texto", "obligatorio": False}
])

campos_plan_tratamiento = json.dumps([
    {"id": "objetivos", "nombre": "Objetivos Terapéuticos", "tipo": "texto", "obligatorio": True},
    {"id": "intervenciones", "nombre": "Intervenciones Propuestas", "tipo": "texto", "obligatorio": True},
    {"id": "frecuencia", "nombre": "Frecuencia de Sesiones", "tipo": "numero", "obligatorio": True, "min": 1, "max": 7}
])

campos_alta = json.dumps([
    {"id": "resultado_tratamiento", "nombre": "Resultado del Tratamiento", "tipo": "seleccion", "obligatorio": True, "opciones": "Sin mejora,Mejora leve,Mejora moderada,Mejora significativa,Resolución completa"},
    {"id": "recomendaciones", "nombre": "Recomendaciones", "tipo": "texto", "obligatorio": True},
    {"id": "necesidad_seguimiento", "nombre": "Necesidad de Seguimiento", "tipo": "booleano", "obligatorio": True}
])

# Añadir datos de ejemplo para Formularios
formularios_data = [
    ["FORM001", "Evaluación Inicial Fisioterapia", "Formulario de evaluación inicial para pacientes de fisioterapia", "Fisioterapia", "Datos Generales", campos_evaluacion_inicial],
    ["FORM001", "Evaluación Inicial Fisioterapia", "Formulario de evaluación inicial para pacientes de fisioterapia", "Fisioterapia", "Evaluación del Dolor", campos_evaluacion_dolor],
    ["FORM001", "Evaluación Inicial Fisioterapia", "Formulario de evaluación inicial para pacientes de fisioterapia", "Fisioterapia", "Evaluación de Movilidad", campos_evaluacion_movilidad],
    ["FORM001", "Evaluación Inicial Fisioterapia", "Formulario de evaluación inicial para pacientes de fisioterapia", "Fisioterapia", "Evaluación de Fuerza", campos_evaluacion_fuerza],
    ["FORM001", "Evaluación Inicial Fisioterapia", "Formulario de evaluación inicial para pacientes de fisioterapia", "Fisioterapia", "Plan de Tratamiento", campos_plan_tratamiento],
    
    ["FORM002", "Evaluación Inicial Psicología", "Formulario de evaluación inicial para pacientes de psicología", "Psicología", "Datos Generales", campos_evaluacion_inicial],
    ["FORM002", "Evaluación Inicial Psicología", "Formulario de evaluación inicial para pacientes de psicología", "Psicología", "Evaluación Emocional", campos_evaluacion_emocional],
    ["FORM002", "Evaluación Inicial Psicología", "Formulario de evaluación inicial para pacientes de psicología", "Psicología", "Evaluación Cognitiva", campos_evaluacion_cognitiva],
    ["FORM002", "Evaluación Inicial Psicología", "Formulario de evaluación inicial para pacientes de psicología", "Psicología", "Plan de Tratamiento", campos_plan_tratamiento],
    
    ["FORM003", "Evaluación Inicial Fonoaudiología", "Formulario de evaluación inicial para pacientes de fonoaudiología", "Fonoaudiología", "Datos Generales", campos_evaluacion_inicial],
    ["FORM003", "Evaluación Inicial Fonoaudiología", "Formulario de evaluación inicial para pacientes de fonoaudiología", "Fonoaudiología", "Evaluación del Habla", campos_evaluacion_habla],
    ["FORM003", "Evaluación Inicial Fonoaudiología", "Formulario de evaluación inicial para pacientes de fonoaudiología", "Fonoaudiología", "Plan de Tratamiento", campos_plan_tratamiento],
    
    ["FORM004", "Seguimiento Fisioterapia", "Formulario de seguimiento para pacientes de fisioterapia", "Fisioterapia", "Evaluación del Dolor", campos_evaluacion_dolor],
    ["FORM004", "Seguimiento Fisioterapia", "Formulario de seguimiento para pacientes de fisioterapia", "Fisioterapia", "Evaluación de Movilidad", campos_evaluacion_movilidad],
    ["FORM004", "Seguimiento Fisioterapia", "Formulario de seguimiento para pacientes de fisioterapia", "Fisioterapia", "Seguimiento", campos_evaluacion_seguimiento],
    
    ["FORM005", "Alta Fisioterapia", "Formulario de alta para pacientes de fisioterapia", "Fisioterapia", "Evaluación Final", campos_evaluacion_movilidad],
    ["FORM005", "Alta Fisioterapia", "Formulario de alta para pacientes de fisioterapia", "Fisioterapia", "Resultados", campos_alta],
    
    ["FORM006", "Alta Psicología", "Formulario de alta para pacientes de psicología", "Psicología", "Evaluación Final", campos_evaluacion_emocional],
    ["FORM006", "Alta Psicología", "Formulario de alta para pacientes de psicología", "Psicología", "Resultados", campos_alta]
]

for row_num, row_data in enumerate(formularios_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = Alignment(vertical='center')
        if col_num == 6:  # JSON column
            cell.alignment = Alignment(wrap_text=True)

# Ajustar ancho de columnas para la tercera hoja
for col in ws3.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
    adjusted_width = min(max_length + 2, 100)  # Limitar ancho máximo
    ws3.column_dimensions[column].width = adjusted_width

# Guardar el archivo Excel
excel_filename = "ParametrizacionTerapias.xlsx"
wb.save(excel_filename)

print(f"Archivo Excel '{excel_filename}' creado correctamente.")

# Mostrar información sobre el archivo generado
print("\nEstructura del archivo Excel generado:")
print("1. Hoja 'Formatos': Configuración visual por tipo de terapia")
print("2. Hoja 'CamposObligatorios': Campos requeridos según normativa")
print("3. Hoja 'FormulariosEvaluacion': Formularios personalizados por tipo de terapia")

print("\nEl archivo está listo para ser cargado en la aplicación Streamlit.")
