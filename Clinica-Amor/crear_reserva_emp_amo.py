import streamlit as st
from google_sheets_emp import GoogleSheet
#from google_calendar_emp import GoogleCalendar
#from validacion_reservas_dp import validar_existencia_reserva, formatear_fecha, formatear_hora
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import sqlite3
from sqlite3 import Error
import os 
import sys
import logging
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
from googleapiclient.errors import HttpError
from googleapiclient.discovery import build
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json

st.cache_data.clear()
st.cache_resource.clear()

# Constantes
MAX_RETRIES = 3
INITIAL_RETRY_DELAY = 1
#SPREADSHEET_ID = 'TU-ID-DE-SPREADSHEET'  # Reemplaza con tu ID
#SCOPE = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

# Configuración de caché
class Cache:
    def __init__(self, ttl_minutes=5):
        self.data = None
        self.last_fetch = None
        self.ttl = timedelta(minutes=ttl_minutes)

    def is_valid(self):
        if self.last_fetch is None or self.data is None:
            return False
        return datetime.now() - self.last_fetch < self.ttl

    def set_data(self, data):
        self.data = data
        self.last_fetch = datetime.now()

    def get_data(self):
        return self.data

# Inicializar caché en session state
if 'cache' not in st.session_state:
    st.session_state.cache = Cache()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='crear_reserva_emp_amo.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Cargar el archivo Excel una sola vez
datos_book = load_workbook("./archivos-amo/parametros_empresa.xlsx", read_only=False)

def api_call_handler(func):
  # Number of retries
  for i in range(0, 10):
    try:
      return func()
    except Exception as e:
      print(e)
      time.sleep(2 ** i)
  print("The program couldn't connect to the Google Spreadsheet API for 10 times. Give up and check it manually.")
  raise SystemError

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       data = _row[2]
       break
  return data

def dataBookZonaEnc(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[5]
       break
  return data

def dataBookTelEnc(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[2]
       break
  return data

def dataBookProducto(hoja,producto):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == producto:
       data = _row[0]
       break
       #print(f'su correo es {_row[1]}')
  return data

def dataBookPrecio(hoja,producto):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == producto:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data

def get_conductores_por_zona(zona):
    
    if not zona:
        return []
    
    # Mapeo de zonas a nombres de hojas
    mapping = {
        'Psicología': 'encargado_psicologia',
        'Consultoría Social': 'encargado_consultoria',
        'Terapia': 'encargado_terapia',
        'Inscripciones': 'encargado_inscripciones',
        'Cursos': 'encargado_cursos'
    }
    
    try:
        hoja = mapping.get(zona)
        if hoja:
            encargado = dataBook(hoja)
            return [c for c in encargado if c != 'X' and c is not None]
        return []
    except Exception as e:
        st.error(f"Error al obtener conductores: {str(e)}")
        return []

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=30)).time()
    return new_time.strftime("%H:%M")

def add_hour_and_half2(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")
  
def calcular_diferencia_tiempo_futuro(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Calcular la fecha y hora futura (1 hora y 30 minutos después)
        tiempo_futuro = fecha_hora + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos
        diferencia = tiempo_futuro - fecha_hora
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def calcular_diferencia_tiempo(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora_parametro = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Obtener fecha y hora actual
        fecha_hora_actual = datetime.now()
        
        # Calcular la fecha y hora futura sumando 1:30h a la fecha del parámetro
        tiempo_futuro = fecha_hora_parametro + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos entre el tiempo futuro y la hora actual
        diferencia = tiempo_futuro - fecha_hora_actual
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

#class CrearReservaEmp:
  
#  class Model:
#    pageTitle = "***Generar Reserva***"
  
 # def view(self,model):
 #   st.title(model.pageTitle)

def create_connection():
    """Create a database connection to the SQLite database"""
    conn = None
    try:
        conn = sqlite3.connect('reservas_dp.db')
        return conn
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def create_tables(conn):
    """Create the necessary tables in the database"""
    try:
        cursor = conn.cursor()
        
        # Create reservas table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS reservas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                email TEXT,
                fecha DATE NOT NULL,
                hora TEXT NOT NULL,
                servicio TEXT NOT NULL,
                precio TEXT,
                encargado TEXT NOT NULL,
                email_encargado TEXT,
                zona TEXT,
                producto TEXT,
                direccion TEXT,
                notas TEXT,
                uid TEXT UNIQUE,
                whatsapp BOOLEAN,
                telefono TEXT,
                whatsapp_web TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
    except Error as e:
        print(f"Error creating tables: {e}")

def insert_reserva(conn, reserva_data):
    
    sql = '''INSERT INTO reservas(
                nombre, email, fecha, hora, servicio, precio, 
                encargado, email_encargado, zona, producto, direccion, 
                notas, uid, whatsapp, telefono, whatsapp_web)
             VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, reserva_data)
        conn.commit()
        return cursor.lastrowid
    except Error as e:
        print(f"Error inserting reserva: {e}")
        return None

def check_existing_reserva(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing reserva: {e}")
        return False 

def check_existing_encargado(conn, encargado, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE encargado = ? AND fecha = ? AND hora = ?
            '''
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (encargado, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing encargado: {e}")
        return False 

def limpiar_campos_formulario():
    
    try:
        # Lista de campos a limpiar
         valores_default = {
            'nuevo_nombre': '',
            'selection_option': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
            'productos_str': '',
            'whatsapp': ''
            #'fecha',
            #'hora',
            #'servicio_selector'
         }
        
         # Actualizar el session state con los valores por defecto
         for campo, valor in valores_default.items():
            if campo in st.session_state:
                # Eliminar la entrada actual del session state
                del st.session_state[campo]
        
         # Forzar la recarga de la página para reiniciar los widgets
         st.rerun()
        
         return True
        
    except Exception as e:
        st.error(f"Error al limpiar los campos del formulario: {str(e)}")
        logging.error(f"Error en limpiar_campos_formulario: {str(e)}")
        return False

def inicializar_valores_default():
    
    valores_default = {
            'nuevo_nombre': '',
            'selection_option': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
            'productos_str': '',
            'whatsapp': ''
    }
    
    for campo, valor in valores_default.items():
        if campo not in st.session_state:
            st.session_state[campo] = valor


def poblar_campos_formulario(email,  direccion,  telefono ):
    
    try:
        # Lista de campos a limpiar
         valores_default = {
            'email': 'email',
            'direccion': 'direccion',
            'telefono': 'telfono'            
         }
        
         # Actualizar el session state con los valores por defecto
         for campo, valor in valores_default.items():
            if campo not in st.session_state:
                st.session_state[campo] = valor
        
         # Forzar la recarga de la página para reiniciar los widgets
         st.rerun()
        
         return True
        
    except Exception as e:
        st.error(f"Error al poblar los campos del formulario: {str(e)}")
        logging.error(f"Error en poblar_campos_formulario: {str(e)}")
        return False
        
def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheetsemp']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def add_new_client(creds, nombre):
  """Añade un nuevo cliente a la hoja de Google Sheets"""
  for intento in range(MAX_RETRIES):
    try:
     with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-amo')
        worksheet = sheet.worksheet('clientes')
        
        # Obtiene el último ID y añade 1
        last_row = len(worksheet.get_all_values())
        new_id = last_row  # Asumiendo que la primera fila es encabezado
        
        # Añade la nueva fila
        worksheet.append_row([new_id, nombre])
        
        st.success(f"Cliente '{nombre}' añadido exitosamente!")
        return True
    
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error al añadir el cliente: {str(e)}")
        return False

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def consultar_reserva(nombre, fecha, hora):
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        
        workbook = gc.open('gestion-reservas-amo')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = api_call_handler(lambda:worksheet.get_all_records())

        # Verificar si hay registros antes de crear el DataFrame
        if not registros:
            return False  # No hay datos en la hoja
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)

        # Verificar si las columnas necesarias existen
        required_columns = ['NOMBRE', 'FECHA', 'HORA']
        if not all(col in df.columns for col in required_columns):
            st.warning("La hoja no contiene todas las columnas necesarias")
            return False
        
        try:

            # Realizar la búsqueda
            reserva = df[
                (df['NOMBRE'].str.lower() == nombre.lower()) &
                (df['FECHA'] == fecha) &
                (df['HORA'] == hora)
            ]

        except AttributeError:
            # En caso de que alguna columna no sea del tipo esperado
            st.warning("Error en el formato de los datos")
            return False
        
        #return not reserva.empty
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:

            #st.warning("Solicitud de Cliente No Existe")
            return False #, None
            
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def get_google_sheet_data(creds):
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-amo')
        worksheet = sheet.worksheet('clientes')  # Cambiado a 'clientes' en lugar de 'reservas'
        data = api_call_handler(lambda: worksheet.get_all_values())
        
        if not data:
            st.error("No se encontraron datos en la hoja de cálculo.")
            return None
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        if df.empty:
            st.error("El DataFrame está vacío después de cargar los datos.")
            return None
        
        return df
    
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error al cargar los datos: {str(e)}")
        return None

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def consultar_encargado(encargado, fecha, hora):
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-amo')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = api_call_handler(lambda:worksheet.get_all_records())
        
        # Verificar si hay registros antes de crear el DataFrame
        if not registros:
            return False  # No hay datos en la hoja
        
        # Convertir a DataFrame solo si hay registros
        df = pd.DataFrame(registros)
        
        # Verificar si las columnas necesarias existen
        required_columns = ['ENCARGADO', 'FECHA', 'HORA']
        if not all(col in df.columns for col in required_columns):
            st.warning("La hoja no contiene todas las columnas necesarias")
            return False
        
        # Realizar la búsqueda asegurándose de que no haya valores nulos
        try:
            encargado_registro = df[
                (df['ENCARGADO'].fillna('').str.lower() == encargado.lower()) &
                (df['FECHA'].fillna('') == fecha) &
                (df['HORA'].fillna('') == hora)
            ]
        except AttributeError:
            # En caso de que alguna columna no sea del tipo esperado
            st.warning("Error en el formato de los datos")
            return False
        
        if not encargado_registro.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()

            #datos_reserva = {
            #'EMAIL': encargado_registro['EMAIL'].iloc[0]
            #}

            return True #, datos_reserva

            #return True #, detalles_reserva
        else:
            #st.warning("Solicitud de Cliente No Existe")
            return False

        #return not encargado_registro.empty

    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False
            
    except Exception as e:
        st.error(f"Error al consultar encargado: {str(e)}")
        return False

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def consultar_otros(nombre):
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-amo')
        worksheet = workbook.worksheet('clientes')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()

        if not registros:
            return False  # No hay datos en la hoja
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)

        # Verificar si las columnas necesarias existen
        required_columns = ['NOMBRE']
        if not all(col in df.columns for col in required_columns):
            st.warning("La hoja no contiene todas las columnas necesarias")
            return False
        
        try:
        
            # Realizar la búsqueda
            reserva = df[
                (df['NOMBRE'].str.lower() == nombre.lower()) ]

        except AttributeError:
            # En caso de que alguna columna no sea del tipo esperado
            st.warning("Error en el formato de los datos")
            return False
        
        # Verificar si se encontró la reserva
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
                        # Extraer los campos solicitados
            datos_reserva = {
                'EMAIL': reserva['EMAIL'].iloc[0],
                'TELEFONO': reserva['TELEFONO'].iloc[0],
                'DIRECCION': reserva['DIRECCION'].iloc[0]
            }

            return True, datos_reserva

        else:
            #st.warning("Solicitud de Cliente No Existe")
            return False #, None

    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error en la aplicación: {str(e)}")

def generate_whatsapp_link(phone_number, message):
    encoded_message = message.replace(' ', '%20')
    return f"https://wa.me/{phone_number}?text={encoded_message}"

#@st.cache_data(ttl=300)  # Cache de 5 minutos
def crea_reserva():
  #for intento in range(MAX_RETRIES):
    try:
      #with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        st.title('Generar Reserva del Servicio')
        st.write("---")
        
        # Load initial data
        horas = dataBook("horario")
        servicios = dataBook("servicio")
        producto = dataBook("precios")
        
        # Inicializar valores por defecto
        inicializar_valores_default()
        
        # Inicializar lista de productos si no existe
        if 'productos_seleccionados' not in st.session_state:
            st.session_state.productos_seleccionados = []
        
        # Crear columnas para organizar la interfaz
        col1, col2 = st.columns([1, 1])
        
        with col1:
          #for intento in range(MAX_RETRIES):
            try:
             #with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
                # Cargar credenciales
                creds = load_credentials_from_toml()
        
                # Carga inicial de datos
                df = get_google_sheet_data(creds)
        
                if df is not None and not df.empty:
                    # Lista de nombres existentes
                    nombres_existentes = df['NOMBRE'].tolist() if 'NOMBRE' in df.columns else []
            
                    # Crea el selectbox con opción para añadir nuevo
                    selected_option = st.selectbox(
                         "Seleccione un cliente",
                        options=['-- Añadir Nuevo Cliente --'] + nombres_existentes
                    )
            
                    # Si se selecciona añadir nuevo
                    if selected_option == '-- Añadir Nuevo Cliente --':
                        #with st.form("nuevo_cliente"):
                            nuevo_nombre = st.text_input("Ingrese el nombre del nuevo cliente")
                            #submitted = st.form_submit_button("Añadir Cliente")
                    
                            if nuevo_nombre:
                                if nuevo_nombre in nombres_existentes:
                                    st.warning("Este cliente ya existe en la lista.")
                                else:
                                    pass                                 
                                    #if add_new_client(creds, nuevo_nombre):
                                    #    st.rerun()  # Recarga la página para  actualizar la lista
                    else:
                        
                        st.write(f"Cliente seleccionado: {selected_option}")
                        
                        #valida, result = consultar_otros(selected_option)

                        #if valida:
                            
                        #    email_e = result['EMAIL']
                        #    telefono_e = result['TELEFONO']
                        #    direccion_e = result['DIRECCION']   

                        #    poblar_campos_formulario(email_e,  direccion_e,  telefono_e)

                        #else:
                            # Si hay error, result será un diccionario
                        #    st.warning("Cliente No Existe")
                            #print(f"Error: {result['message']}")

                else:
                    st.error("No se pudieron cargar los datos. Por favor, verifica la conexión.")

            except Exception as e:
                st.error(f"Error en la aplicación: {str(e)}")

            #nombre = st.text_input('Nombre Solicitante*: ', placeholder='Nombre', key='nombre',  value=st.session_state.nombre)
            
            # Selector de servicio
            conductor_seleccionado = None
            
            email = st.text_input('Email Solicitante:', placeholder='Email', 
                                  key='email', value=st.session_state.email)
                         
            # Manejo de zonas para ciertos servicios
            zona_seleccionada = None

            direccion = st.text_input('Direccion Ubicacion solicitante:', placeholder='Direccion', key='direccion', value=st.session_state.direccion)
            fecha = st.date_input('Fecha Servicio*: ')
            hora = st.selectbox('Hora Servicio: ', horas, key="hora_new")
          
            whatsapp = st.checkbox('Envio a WhatsApp Si/No (Opcional)')
            telefono = st.text_input('Nro. Telefono', key='telefono', value=st.session_state.telefono)

            notas = st.text_area('Nota de Consulta o Motivo Especial(Opcional)', 
                                 key='notas', value=st.session_state.notas)

        with col2:

            servicio_seleccionado = st.selectbox(
                'Seleccione el servicio:',
                servicios,
                key='servicio_selector'
            )
            if servicio_seleccionado in ['Psicología', 'Consultoría Social', 'Terapia', 'Cursos']:
                zonas = ['Psicología', 'Consultoría Social', 'Terapia', 'Inscripciones', 'Cursos']
                zona_seleccionada = st.info(f'Area de Servicio: {servicio_seleccionado}'
                    #skey='zona_selector'
                    #label_visibility='hidden'
                )

                zona_seleccionada = servicio_seleccionado
                encargado = get_conductores_por_zona(zona_seleccionada)
            else:
                encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]

            # Mostrar selector de conductor si hay conductores disponibles
            if encargado:
                conductor_seleccionado = st.selectbox(
                    'Encargado :',
                    encargado,
                    key='conductor_selector'
                )
            else:
                st.warning("Elija el encargado para este servicio")
                #return

            # Sección para agregar productos
            st.write("### Agregar Servicio")
            
            # Selector de producto
            producto_seleccionado = st.selectbox(
                'Seleccione la especialidad:',
                producto,
                key='producto_selector'
            )
            
            # Cantidad e información del producto
            cantidad = st.number_input('Cantidad', min_value=1, key='cant')
            precio = dataBookPrecio("precios", producto_seleccionado)
            
            # Botón para agregar producto
            if st.button('Aceptar seleccion', type="primary"):
                producto_info = {
                    'producto': producto_seleccionado,
                    'cantidad': cantidad,
                    'precio': precio
                }
                st.session_state.productos_seleccionados.append(producto_info)
            
            # Mostrar productos seleccionados
            if st.session_state.productos_seleccionados:
                st.write("### Especilidad Seleccionada:")
                total_productos = 0
                for idx, prod in enumerate(st.session_state.productos_seleccionados):
                    subtotal = prod['cantidad'] * prod['precio']
                    total_productos += subtotal
                    st.write(f"{idx+1}. {prod['producto']} - Cantidad: {prod['cantidad']} - Subtotal: ${subtotal:,.0f}")
                
                st.write(f"**Total: ${total_productos:,.0f}**")
                
                # Opción de eliminar productos
                if st.button('Limpiar seleccion', type="primary"):
                    st.session_state.productos_seleccionados = []
        
            # Solo continuar si tenemos un conductor seleccionado
            existe_db2 = None
            if conductor_seleccionado:
                # Validaciones existentes de disponibilidad
                existe_db2 = consultar_encargado(conductor_seleccionado, str(fecha), hora)

            if existe_db2:
                resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')
                if resultado > 0 and resultado <= 60:
                    st.warning("Encargado se encuentra atendiendo un servicio")
                elif resultado >= 60:
                    st.warning("Encargado ya tiene agenda para esa fecha y hora")
                elif resultado <= -170:
                    st.warning("No puede agendarse con una fecha y hora vencida")
                else:
                    st.success("La Hora de solicitud está disponible")
            
            else:
                
                resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')
                if resultado <= -170:
                    st.warning("No puede agendarse con una fecha y hora vencida")
                else:
                    st.success("La Hora de solicitud está disponible")
        
        existe_cliente = consultar_reserva(selected_option, str(fecha), hora)

        if existe_cliente:
          st.warning("El Cliente ya tiene agenda para esa fecha y hora")
          return
        else:
          st.write("---")
          st.write("### Resumen de Solicitud:")
        
          # Create columns for horizontal summary
          if st.session_state.productos_seleccionados:
            num_productos = len(st.session_state.productos_seleccionados)
            cols = st.columns(min(num_productos, 4))  # Max 4 products per row
            
            total_pedido = 0
            for idx, producto in enumerate(st.session_state.productos_seleccionados):
                col_idx = idx % 4
                with cols[col_idx]:
                    st.markdown(f"""
                    🛍️ **Espeialidad {idx + 1}**
                    - Especialidad: {producto['producto']}
                    - Cantidad: {producto['cantidad']}
                    - Costo: ${producto['precio']:,.0f}
                    - Subtotal: ${(producto['cantidad'] * producto['precio']):,.0f}
                    """)
                total_pedido += producto['cantidad'] * producto['precio']
            
            # Show service details after products
            st.write("---")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.markdown(f"""
                **Detalles del Servicio:**
                - 🚗 Encargado: {conductor_seleccionado}
                - 🎯 Servicio: {servicio_seleccionado}
                """)
            
            with col2:
                st.markdown(f"""
                **Detalles de Reserva:**
                - 📅 Fecha: {fecha}
                - 🕒 Hora: {hora}
                """)
            
            with col3:
                st.markdown(f"""
                **Resumen de la Solicitud:**
                - 📦 Total Solicitado: {num_productos}
                - 💰 Total a Pagar: ${total_pedido:,.0f}
                """)

            if servicio_seleccionado in ['Psicología', 'Consultoría Social', 'Terapia', 'Cursos']:
                st.info(f"📍 Area de Trabajo: {zona_seleccionada}")


        st.write("---")
        
        with st.form(key='myform0', clear_on_submit=True):
            enviar = st.form_submit_button(" Solicitar Servicio", type="primary")
            
            if enviar:
                with st.spinner('Cargando...'):
                 if servicio_seleccionado != 'Consulta esn General':
                    # Validaciones
                    if  not email or not direccion:
                        st.warning('Se requiere completar los campos de  Email y Direccion son obligatorios')
                    elif not validate_email(email):
                        st.warning('El email no es valido')
                    elif whatsapp and not telefono:
                        st.warning('Se requiere el numero del Celular')
                    elif not st.session_state.productos_seleccionados:
                        st.warning('Debe seleccionar al menos una especialidad')
                    else:
                        # Preparar información de productos para guardar
                        productos_str = "; ".join([
                            f"Servicio.-{p['producto']}: Cant.-{p['cantidad']}: Costo-{p['precio']}" 
                            for p in st.session_state.productos_seleccionados
                        ])
                        
                        # Calcular precio total
                        precio_total = sum(p['cantidad'] * p['precio'] for p in st.session_state.productos_seleccionados)
                        
                        # Obtener email del encargado
                        emailencargado = dataBookEncEmail("encargado", conductor_seleccionado)
                        #result_email = np.setdiff1d(emailencargado,'')

                        tel_encargado = dataBookTelEnc("encargado", conductor_seleccionado)

                        if selected_option == '-- Añadir Nuevo Cliente --':        
                         
                          try:
                           
                            values2 = [(nuevo_nombre, email, direccion, telefono, zona_seleccionada, productos_str,  str(datetime.now()))] 
                            
                            gs = GoogleSheet(st.secrets['sheetsemp']['credentials_sheet'], 'gestion-reservas-amo', 'clientes')
                            range2 = gs.get_last_row_range()
                            gs.write_data(range2, values2)

                            # Generar UID
                            uid = generate_uid()
                        
                            # Preparar valores para guardar
                            values = [(
                                nuevo_nombre, email, str(fecha), hora, servicio_seleccionado, 
                                precio_total, conductor_seleccionado, str(emailencargado), 
                                zona_seleccionada, productos_str, len(st.session_state.productos_seleccionados), 
                                direccion, notas, uid, whatsapp, str(57)+telefono, 
                                f"web.whatsapp.com/send?phone=&text=Reserva para { selected_option}", 
                                '=ArrayFormula(SI(M3=VERDADERO;HIPERVINCULO(O3;"Enviar");"No Enviar"))','Reservado'
                            )]

                            # Guardar en Google Sheets
                            gs = GoogleSheet(st.secrets['sheetsemp']['credentials_sheet'], 'gestion-reservas-amo', 'reservas')
                            range2 = gs.get_last_row_range()
                            gs.write_data(range2, values)

                            values3 = [(
                                nuevo_nombre, email, str(fecha), hora, servicio_seleccionado, 
                                conductor_seleccionado, zona_seleccionada, productos_str  )]

                            gs2 = GoogleSheet(st.secrets['sheetsemp']['credentials_sheet'], 'gestion-reservas-amo', 'asistencia')
                            range2 = gs2.get_last_row_range()
                            gs2.write_data(range2, values3)

                            st.success('Su solicitud ha sido reservada de forma exitosa, la confirmación fue enviada al correo')
                            
                            # Enviar emails
                            send_email2(email,  nuevo_nombre, fecha, hora, servicio_seleccionado, productos_str, precio_total, conductor_seleccionado, notas)
                            
                            send_email_emp(email,  nuevo_nombre, fecha, hora, servicio_seleccionado, productos_str, precio_total, conductor_seleccionado, notas, str(emailencargado))
                            
                            # Envío por WhatsApp (si aplica)
                            if whatsapp == True or whatsapp == 'Verdadero':
                                contact = str(57)+telefono
                                message = f'Cordial saludo: Sr(a): {nuevo_nombre} La Reserva se creó con éxito para el día: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para el servicio: {servicio_seleccionado}. Especialidad: {productos_str}. Cordialmente, aplicación de Servicios.'

                                contact2 =  str(57)+str(tel_encargado)
                                whatsapp_link = generate_whatsapp_link(contact2, message)
                                st.markdown(f"Click en enlace si desea Enviar Whatsapp {whatsapp_link}")
                                
                                time.sleep(10)                       
                                     
                          except Exception as e:
                             st.error(f"Error al guardar la reserva: {str(e)}")   

                        else:

                            # Generar UID
                            uid = generate_uid()
                        
                            # Preparar valores para guardar
                            values = [(
                                selected_option, email, str(fecha), hora, servicio_seleccionado, 
                                precio_total, conductor_seleccionado, str(emailencargado), 
                                zona_seleccionada, productos_str, len(st.session_state.productos_seleccionados), 
                                direccion, notas, uid, whatsapp, str(57)+telefono, 
                                f"web.whatsapp.com/send?phone=&text=Reserva para { selected_option}", 
                                '=ArrayFormula(SI(M3=VERDADERO;HIPERVINCULO(O3;"Enviar");"No Enviar"))','Reservado'
                            )]
                            
                            try:           
                        
                                # Guardar en Google Sheets
                                gs = GoogleSheet(st.secrets['sheetsemp']['credentials_sheet'], 'gestion-reservas-amo', 'reservas')
                                range2 = gs.get_last_row_range()
                                gs.write_data(range2, values)
                                
                                values3 = [(
                                selected_option, email, str(fecha), hora, servicio_seleccionado, conductor_seleccionado, zona_seleccionada, productos_str  )]

                                gs2 = GoogleSheet(st.secrets['sheetsemp']['credentials_sheet'], 'gestion-reservas-amo', 'asistencia')
                                range2 = gs2.get_last_row_range()
                                gs2.write_data(range2, values3)
                            
                                st.success('Su solicitud ha sido reservada de forma exitosa, la confirmación fue enviada al correo')
                            
                                # Enviar emails
                                send_email2(email,  selected_option, fecha, hora, servicio_seleccionado, productos_str, precio_total, conductor_seleccionado, notas)
                            
                                send_email_emp(email,  selected_option, fecha, hora, servicio_seleccionado, productos_str, precio_total, conductor_seleccionado, notas, str(emailencargado))
                            
                                # Envío por WhatsApp (si aplica)
                                if whatsapp == True or whatsapp == 'Verdadero':
                                    contact = str(57)+telefono
                                    message = f'Cordial saludo: Sr(a): { selected_option} La Reserva se creó con éxito para el día: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para el servicio: {servicio_seleccionado}. Especialidad: {productos_str}. Cordialmente, Clinica del Amor.'
                                
                                    whatsapp_link = generate_whatsapp_link(contact, message)
                                    st.markdown(f"Click en el enlace si desea Enviar Whatsapp {whatsapp_link}")
                                    
                                    contact2 =  str(57)+str(tel_encargado)
                                    whatsapp_link = generate_whatsapp_link(contact, message)
                                    (f'{whatsapp_link}')

                                    time.sleep(10)
                        
                            except Exception as e:
                               st.error(f"Error al guardar la reserva: {str(e)}")
                               # Limpiar campos
                 else:
                    if selected_option == '-- Añadir Nuevo Cliente --':
                        send_email_emp(email,  nuevo_nombre, fecha, hora, servicio_seleccionado, '', '',  '', notas, '')

                        st.success('Su Consulta ha sido enviada de forma exitosa')
                    else:
                        send_email_emp(email,  selected_option, fecha, hora, servicio_seleccionado, '', '',  '', notas, '')

                        st.success('Su Consulta ha sido enviada de forma exitosa')

                if limpiar_campos_formulario():
                   st.session_state.productos_seleccionados = []
                   st.success('Campos limpiados exitosamente')


    #except HttpError as error:
    #    if error.resp.status == 429:  # Error de cuota excedida
    #       if intento < MAX_RETRIES - 1:
    #          delay = INITIAL_RETRY_DELAY * (2 ** intento)
    #          st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
    #          time.sleep(delay)
    #          continue
    #       else:
    #          st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
    #    else:
    #        st.error(f"Error de la API: {str(error)}")

    except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error(f"Error crítico en la aplicación. Por favor, contacte al administrador. {str(e)}")

sys.excepthook = global_exception_handler

#if __name__ == "__main__":
#   crea_reserva()