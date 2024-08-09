import qrcode
import streamlit as st
import numpy as np
import os
#import cv2
#from pyzbar.pyzbar import decode
from googleapiclient.errors import HttpError
#import pandas as pd
from openpyxl import load_workbook

datos_book = load_workbook("./archivos/parametros_abogados.xlsx", read_only=False)

#data = pd.read_csv('BASE.csv')
#print(data.head())

class GenerarQr:
  
  class Model:

    pageTitle ='*** Generacion de Codigos QR ***'

    #ws2 = datos_book["encargado"]
 
  def view(self,model):
    st.title(model.pageTitle)
    
    def dataBookServicio(hoja):
      ws1 = datos_book[hoja]
      data = []
      for row in ws1.iter_rows(min_row=2, min_col=0):
        resultado = [col.value for col in row]
        data.append(resultado[0:5])
        #print(f'data {data}')
      return data
    
    def dataBookQR(hoja):
      ws1 = datos_book[hoja]
      data = []
      for row in range(1, ws1.max_row):
        _row=[]
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
          _row.append(col[row].value)
          data.append(_row[0])
          #print(f'data {data}')
      return data
    
    def dataBookQR2(hoja,acargo):
      ws1 = datos_book[hoja]
      data = []
      for row in range(1,ws1.max_row):
        _row=[]
        for col in ws1.iter_cols(1,ws1.max_column):
          _row.append(col[row].value)
          data.append(_row) 
          #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
          if _row[0] == acargo:
            nombre = _row[0]
            break
      return nombre
    
    def dataBookQR3(hoja,acargo):
      ws1 = datos_book[hoja]
      data = []
      for row in range(1,ws1.max_row):
        _row=[]
        for col in ws1.iter_cols(1,ws1.max_column):
          _row.append(col[row].value)
          data.append(_row) 
          #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
          if _row[0] == acargo:
            codigo = _row[3]
            break
      return codigo
    
    #def read_qr_code(file):
      
    #  image = cv2.imread(file)
    #  gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    #  decoded_objects = decode(gray_image)
      
    #  if decoded_objects:
    #    for obj in decoded_objects:
    #     data = obj.data.decode('utf-8')
    #      st.success(f'El dato del codigo QR es : {data} ')
      
    #  else:
    #    st.warning('No se encontro el  codigo QR para el nombre dado')

  
    with st.form(key='myform',clear_on_submit=True):
      opciones = ["Generar Todos", "Generar Uno Especifico", "Leer QR" ]
      
      col1, col2 = st.columns(2)
      opcion = col1.selectbox('Tipo Generacion de Archivo*:', opciones)
      
      codigoqr = dataBookQR("encargado")
      result_qr = np.setdiff1d(codigoqr,'X')
      acargo = col2.selectbox('A cargo de:',result_qr)                
      #print(f"Result QR: {acargo}")
        
      if opcion == "Generar Todos":
          
         generar = st.form_submit_button(" Generar ")
  
         if generar:   
            with st.spinner('Cargando...'):
              try:   
                os.chdir("generaQR")
                encargado = dataBookServicio("encargado")
                cedula = encargado[0][3]
  
                for i in range(len(encargado)):
   
                  if encargado[i][4] != "Generado":
            
                    cod_proveedor = encargado[i][3]
                    nombre_proveedor = encargado[i][0]
                    telefono = encargado[i][2]
                    codigoqr = encargado[i][4]
                    
                    #os.chdir("./generaQR")
                    img = qrcode.make(cod_proveedor)
                    img.save(f"img/{nombre_proveedor}.png")
      
                    ws2 = encargado
                    ws2[i][4] = "Generado"
      
                #datos_book.save("./archivos/parametros_abogados.xlsx")

                image = qrcode.make("https://app-abogados.streamlit.app/")
                image.save('img/qr_app.png')
                
                os.chdir("..")
                st.success('Archivos generados exitosamente')
                st.balloons()
                 
              except HttpError as err:
                st.warning(f'se presento un  Errror {err} ')
                raise Exception(f'A ocurrido un error en genera CodigoQR: {err}')

      elif opcion == "Generar uno Especifico":

         generar = st.form_submit_button(" Generar ")
  
         if generar:   
            with st.spinner('Cargando...'):
              try:   
                os.chdir("generaQR")
                encargado = dataBookServicio("encargado")
                cedula = encargado[0][3]
  
                for i in range(len(encargado)):
   
                  if encargado[i][4] != "Generado":
            
                    cod_proveedor = encargado[i][3]
                    nombre_proveedor = encargado[i][0]
                    telefono = encargado[i][2]
                    codigoqr = encargado[i][4]
                    
                    #os.chdir("./generaQR")
                    img = qrcode.make(cod_proveedor)
                    img.save(f"img/{nombre_proveedor}.png")
      
                    ws2 = encargado
                    ws2[i][4] = "Generado"
      
                #datos_book.save("./archivos/parametros_abogados.xlsx")
                
                os.chdir("..")
                st.success('Archivos generados exitosamente')
                st.balloons()
                 
              except HttpError as err:
                st.warning(f'se presento un  Errror {err} ')
                raise Exception(f'A ocurrido un error en genera CodigoQR: {err}')

      elif opcion == "Leer QR":
                  
          nombre = dataBookQR2("encargado", acargo)
          codigo = dataBookQR3("encargado", acargo)
            
          if nombre == acargo:
                        
            generar = st.form_submit_button(" Generar ")
              
            if generar:
              
              with st.spinner('Cargando...'): 
              
                try: 
                    #os.chdir("generaQR")
                    #file = (f"C:/Users/hp  pc/Desktop/Programas practica Python/App - Reservas/generaQR/img/{nombre}.png") #=  datos
                    #print("nombre")
          
                    #leer = read_qr_code(file)
                    st.success(f' El dato del codigo QR es : {codigo} y corresponde a : {nombre} ')
                    
                    #os.chdir("..")
                    st.success(f'Codigo QR se leyo exitosamente')
                    st.balloons()
                
                except HttpError as err:
                  st.warning(f'se presento un  Errror {err} ')
                  raise Exception(f'A ocurrido un error al leer el Codigo QR : {err}') 