import qrcode
import streamlit as st
import numpy as np
import os
from googleapiclient.errors import HttpError
from openpyxl import load_workbook

# Cargar el libro de Excel con los parámetros
datos_book = load_workbook("./archivos/parametros_abogados.xlsx", read_only=False)

class GenerarQr:
    """Clase para generar y leer códigos QR de la agenda"""
  
    class Model:
        pageTitle = '*** Generación de Códigos QR ***'
 
    def view(self, model):
        """Método principal que muestra la interfaz de generación de QR"""
        st.title(model.pageTitle)
        
        def dataBookServicio(hoja):
            """Obtiene los datos de servicios desde el Excel"""
            ws1 = datos_book[hoja]
            data = []
            for row in ws1.iter_rows(min_row=2, min_col=0):
                resultado = [col.value for col in row]
                data.append(resultado[0:5])
            return data
        
        def dataBookQR(hoja):
            """Obtiene la lista de encargados para el QR"""
            ws1 = datos_book[hoja]
            data = []
            for row in range(1, ws1.max_row):
                _row = []
                for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
                    _row.append(col[row].value)
                    data.append(_row[0])
            return data
        
        def dataBookQR2(hoja, acargo):
            """Busca un encargado específico en el Excel"""
            ws1 = datos_book[hoja]
            data = []
            nombre = None
            for row in range(1, ws1.max_row):
                _row = []
                for col in ws1.iter_cols(1, ws1.max_column):
                    _row.append(col[row].value)
                    data.append(_row) 
                    if _row[0] == acargo:
                        nombre = _row[0]
                        break
            return nombre
        
        def dataBookQR3(hoja, acargo):
            """Obtiene el código de un encargado específico"""
            ws1 = datos_book[hoja]
            data = []
            codigo = None
            for row in range(1, ws1.max_row):
                _row = []
                for col in ws1.iter_cols(min_row=0, min_col=0, max_col=4):
                    _row.append(col[row].value)
                    if _row[0] == acargo:
                        codigo = _row[3:4]
            return codigo

        # Crear el formulario principal
        with st.form(key='myform_qr', clear_on_submit=True):
            
            st.markdown("### 🔧 Configuración de Generación")
            
            opciones = ["Generar Todos", "Generar Uno Específico", "Leer QR"]
            
            col1, col2 = st.columns(2)
            
            with col1:
                opcion = st.selectbox(
                    '📋 Tipo Generación de Archivo*:', 
                    opciones,
                    help="Selecciona si deseas generar todos los QR, uno específico o leer un QR existente"
                )
            
            with col2:
                codigoqr = dataBookQR("encargado")
                result_qr = np.setdiff1d(codigoqr, 'X')
                acargo = st.selectbox(
                    '👤 A cargo de:',
                    result_qr,
                    help="Selecciona el encargado para el QR"
                )
            
            # Botón de generar
            generar = st.form_submit_button("🚀 Generar", use_container_width=True)
            
            # OPCIÓN 1: Generar Todos
            if opcion == "Generar Todos" and generar:
                with st.spinner('⏳ Generando todos los códigos QR...'):
                    try:   
                        # Cambiar al directorio de generación
                        os.chdir("generaQR")
                        encargado = dataBookServicio("encargado")
                        
                        # Crear el directorio img si no existe
                        if not os.path.exists("img"):
                            os.makedirs("img")
                        
                        contador = 0
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        for i in range(len(encargado)):
                            if encargado[i][4] != "Generado":
                                cod_proveedor = encargado[i][3]
                                nombre_proveedor = encargado[i][0]
                                
                                # Generar el QR
                                img = qrcode.make(cod_proveedor)
                                img.save(f"img/{nombre_proveedor}.png")
                                
                                # Actualizar estado
                                ws2 = encargado
                                ws2[i][4] = "Generado"
                                contador += 1
                                
                                # Actualizar progreso
                                progress = (i + 1) / len(encargado)
                                progress_bar.progress(progress)
                                status_text.text(f"Generando: {nombre_proveedor}")
                        
                        # Generar QR de la aplicación
                        image = qrcode.make("https://app-abogados.streamlit.app/")
                        image.save('img/qr_app.png')
                        
                        os.chdir("..")
                        
                        st.success(f'✅ {contador} archivos generados exitosamente')
                        st.balloons()
                        
                    except Exception as err:
                        os.chdir("..")
                        st.error(f'❌ Se presentó un error: {err}')
                        raise Exception(f'Ocurrió un error en genera CodigoQR: {err}')

            # OPCIÓN 2: Generar Uno Específico
            elif opcion == "Generar Uno Específico" and generar:
                with st.spinner(f'⏳ Generando código QR para {acargo}...'):
                    try:   
                        os.chdir("generaQR")
                        
                        # Crear el directorio img si no existe
                        if not os.path.exists("img"):
                            os.makedirs("img")
                        
                        encargado = dataBookServicio("encargado")
                        qr_generado = False
                        
                        for i in range(len(encargado)):
                            cod_proveedor = encargado[i][3]
                            nombre_proveedor = encargado[i][0]
                            
                            if nombre_proveedor == acargo:           
                                # Generar el QR
                                img = qrcode.make(cod_proveedor)
                                img.save(f"img/{nombre_proveedor}.png")
                                qr_generado = True
                                
                                # Mostrar vista previa
                                st.image(f"img/{nombre_proveedor}.png", 
                                        caption=f"QR generado para {nombre_proveedor}",
                                        width=300)
                                break
                        
                        os.chdir("..")
                        
                        if qr_generado:
                            st.success(f'✅ Código QR generado exitosamente para {acargo}')
                            st.balloons()
                        else:
                            st.warning(f'⚠️ No se encontró información para {acargo}')
                        
                    except Exception as err:
                        os.chdir("..")
                        st.error(f'❌ Se presentó un error: {err}')
                        raise Exception(f'Ocurrió un error en genera CodigoQR: {err}')

            # OPCIÓN 3: Leer QR
            elif opcion == "Leer QR" and generar:
                nombre = dataBookQR2("encargado", acargo)
                
                if nombre == acargo:
                    with st.spinner(f'⏳ Leyendo código QR de {acargo}...'):
                        try: 
                            codigo1 = dataBookQR3("encargado", acargo)
                            
                            # Mostrar información
                            st.info(f'📋 **Encargado:** {nombre}')
                            st.success(f'🔢 **Código:** {codigo1}')
                            
                            # Intentar mostrar la imagen del QR si existe
                            qr_path = f"generaQR/img/{nombre}.png"
                            if os.path.exists(qr_path):
                                st.image(qr_path, 
                                        caption=f"Código QR de {nombre}",
                                        width=300)
                            else:
                                st.warning("⚠️ Imagen del código QR no encontrada. Genera el QR primero.")
                            
                            st.balloons()
                        
                        except Exception as err:
                            st.error(f'❌ Se presentó un error: {err}')
                            raise Exception(f'Ocurrió un error al leer el Código QR: {err}')
                else:
                    st.warning(f'⚠️ No se encontró información para {acargo}')
        
        # Información adicional
        with st.expander("ℹ️ Información sobre códigos QR"):
            st.markdown("""
            ### ¿Cómo funciona esta herramienta?
            
            **Generar Todos**: Crea códigos QR para todos los encargados que aún no tienen uno generado.
            
            **Generar Uno Específico**: Crea un código QR solo para el encargado seleccionado.
            
            **Leer QR**: Muestra la información asociada a un código QR específico.
            
            #### Ubicación de los archivos:
            Los códigos QR se guardan en: `generaQR/img/`
            
            #### Nota:
            - Los códigos QR se generan automáticamente con la información del encargado
            - Cada QR es único y corresponde al código del encargado en el sistema
            """)


def GenerarQr_standalone():
    """
    Función independiente para ejecutar GenerarQr sin necesidad de la clase
    Útil cuando se llama directamente desde el menú principal
    """
    generador = GenerarQr()
    generador.view(GenerarQr.Model())


if __name__ == "__main__":
    # Si se ejecuta directamente, mostrar la interfaz
    GenerarQr_standalone()
