import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
import gspread
import toml
import time

st.cache_data.clear()
st.cache_resource.clear()

# Constantes
MAX_RETRIES = 3
INITIAL_RETRY_DELAY = 1

# Configuración de caché
class Cache:
    def __init__(self, ttl_minutes=5):
        self.data = None
        self.last_fetch = None
        self.ttl = timedelta(minutes=ttl_minutes)

    def is_valid(self):
        if self.last_fetch is None or self.data is None:
            return False
        return datetime.now() - self.last_fetch < self.ttl

    def set_data(self, data):
        self.data = data
        self.last_fetch = datetime.now()

    def get_data(self):
        return self.data

# Inicializar caché en session state
if 'cache' not in st.session_state:
    st.session_state.cache = Cache()

# Inicializar el estado para el formulario de edición
if 'editing' not in st.session_state:
    st.session_state.editing = False
if 'edit_index' not in st.session_state:
    st.session_state.edit_index = None

datos_book = load_workbook("./archivos-dlv/parametros_empresa.xlsx", read_only=False)

def api_call_handler(func):
  # Number of retries
  for i in range(0, 10):
    try:
      return func()
    except Exception as e:
      print(e)
      time.sleep(2 ** i)
      
  print("The program couldn't connect to the Google Spreadsheet API for 10 times. Give up and check it manually.")

  raise SystemError

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def load_credentials_from_toml(file_path):
    with open(file_path, 'r') as toml_file:
        config = toml.load(toml_file)
        credentials = config['sheetsemp']['credentials_sheet']
    return credentials

def get_google_sheet_data(creds):
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-dlv')
        worksheet = sheet.worksheet('devoluciones')
        data = api_call_handler(lambda:worksheet.get_all_values())
        
        if not data:
            st.error("No se encontraron datos en la hoja de cálculo.")
            return None
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        if df.empty:
            st.error("El DataFrame está vacío después de cargar los datos.")
            return None
            
        # Convertir la columna de fecha a datetime
        df['FECHA'] = pd.to_datetime(df['FECHA'])
        
        return df, worksheet
        
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return None, None

    except Exception as e:
        st.error(f"Error al cargar los datos: {str(e)}")
        return None, None

def update_google_sheet(creds, row_idx, column_idx, value):
    try:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-dlv')
        worksheet = sheet.worksheet('devoluciones')
        
        # Actualizar la celda específica (row_idx + 2 porque hay encabezados y la API usa índices basados en 1)
        api_call_handler(lambda: worksheet.update_cell(row_idx + 2, column_idx, value))
        
        # Actualizar la fecha de actualización en una columna dedicada (asumiendo que existe una columna FECHA CAMBIO)
        fecha_cambio = worksheet.find("FECHA CAMBIO").col
        api_call_handler(lambda: worksheet.update_cell(row_idx + 2, fecha_cambio, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        
        return True
    except Exception as e:
        st.error(f"Error al actualizar los datos: {str(e)}")
        return False

def consulta_recibido():

    st.title('Opcion de Entrega-Cambio o Devolucion')
    
    # Cargar credenciales y datos
    try:
        creds = load_credentials_from_toml('./.streamlit/secrets.toml')
        result = get_google_sheet_data(creds)
        if result is None:
            return
        df, worksheet = result
    except Exception as e:
        st.error(f"Error al cargar las credenciales: {str(e)}")
        return
    
    if df is not None:
        # Comprobar si existe la columna FECHA CAMBIO, si no, verificar si se puede añadir
        if 'FECHA CAMBIO' not in df.columns:
            st.warning("La columna FECHA CAMBIO no existe en la hoja. Se utilizará para registrar cuando se modifican los datos.")
            try:
                # Añadir la columna al dataframe
                df['FECHA CAMBIO'] = ""
                # Intentar añadir la columna a la hoja de cálculo
                headers = worksheet.row_values(1)
                if 'FECHA CAMBIO' not in headers:
                    worksheet.append_col(['FECHA CAMBIO'] + [""] * (len(df)))
            except Exception as e:
                st.error(f"No se pudo añadir la columna FECHA CAMBIO: {str(e)}")
        
        # Mostrar la interfaz de búsqueda
        st.header('Filtros de búsqueda')
        
        producto = dataBook("precios")
        
        # Crear tres columnas para los filtros de selección múltiple
        col1, col2 = st.columns(2)
        
        with col1:
            servicios = st.multiselect(
                'Servicios',
                options=sorted(df['SERVICIOS'].unique())
            )
        
        with col2:
            encargados = st.multiselect(
                'Encargados',
                options=sorted(df['ENCARGADO'].unique())
            )
    
        # Crear dos columnas para el rango de fechas
        st.subheader('Rango de Fechas')
        fecha_min = df['FECHA'].min()
        fecha_max = df['FECHA'].max()
        
        col1, col2 = st.columns(2)
        
        with col1:
            start_date = st.date_input(
                'Fecha inicial',
                fecha_min,
                min_value=fecha_min,
                max_value=fecha_max
            )
        
        with col2:
            end_date = st.date_input(
                'Fecha final',
                fecha_max,
                min_value=fecha_min,
                max_value=fecha_max
            )
        
        # Convertir las fechas seleccionadas a datetime
        start_date = pd.to_datetime(start_date)
        end_date = pd.to_datetime(end_date)
        
        # Aplicar filtros
        mask = (df['FECHA'] >= start_date) & (df['FECHA'] <= end_date)
        
        if servicios:
            mask &= df['SERVICIOS'].isin(servicios)
        if encargados:
            mask &= df['ENCARGADO'].isin(encargados)
                    
        filtered_df = df[mask]
        
        # Mostrar resultados
        st.markdown("---")  # Línea divisoria
        st.subheader('Resultados de la búsqueda')
        
        if not filtered_df.empty:
            # Métricas en una fila
            met1, met2, met3 = st.columns(3)
            with met1:
                st.metric("Total Servicios", len(filtered_df['SERVICIOS'].unique()))
            with met2:
                st.metric("Total Encargados", len(filtered_df['ENCARGADO'].unique()))
            
            st.write(f'Se encontraron {len(filtered_df)} registros')
            
            # Mostrar el dataframe con un botón de edición
            edited_df = filtered_df.reset_index()  # Resetear índice para tener una referencia consistente
            
            # Mostrar cada fila con un botón de edición
            for i, row in edited_df.iterrows():
                col1, col2 = st.columns([0.9, 0.1])
                with col1:
                    # Crear un contenedor para cada fila
                    expander = st.expander(f"{row['FECHA'].strftime('%Y-%m-%d')} - {row['SERVICIOS']} - {row['ENCARGADO']}")
                    with expander:
                        # Mostrar detalles de la fila
                        for col in filtered_df.columns:
                            if col != 'index':  # Evitar mostrar el índice creado por reset_index()
                                st.write(f"**{col}:** {row[col]}")
                
                with col2:
                    # Botón de edición
                    edit_button_clicked = st.button(f"Editar", key=f"edit_{i}")
                    if edit_button_clicked:
                        st.session_state.editing = True
                        st.session_state.edit_index = row['index']
                
                # Mostrar el formulario de edición justo debajo del registro que se está editando
                if st.session_state.editing and st.session_state.edit_index == row['index']:
                    with st.container():
                        st.markdown("#### Editar Registro")
                        row_index = st.session_state.edit_index
                        row_data = df.iloc[row_index].copy()
                        
                        # Mostrar información no editable para contexto
                        st.write(f"Fecha: {row_data['FECHA'].strftime('%Y-%m-%d')}")
                        st.write(f"Servicio: {row_data['SERVICIOS']}")
                        st.write(f"Encargado: {row_data['ENCARGADO']}")
                        
                        tipos= ['Entregado', 'Cambiado', 'Devuelto']
                        # Campos editables
                        new_producto_cambio = st.selectbox(
                            'Seleccione el producto de Cambio:',
                            producto,
                            key=f"id_{i}"
                            #key='producto_selector'
                        )
                        #new_producto_cambio = st.text_input("PRODUCTO CAMBIO", value=row_data.get('PRODUCTO CAMBIO', ''), key=f"id_{i}")

                        new_encargado = st.text_input("NOMBRE ENCARGADO(Entrega, Cambia o Devuelve)  CAMBIO" #, value=row_data.get('ENCARGADO CAMBIO', '')
                                                      , key=f"encargado cambio_{i}")

                        new_tipo =  st.selectbox(
                            'Tipo Operacion:',
                            tipos,
                            key=f"tipo_{i}"
                            #key='producto_selector'
                        )
                                                 
                        new_recibido_por = st.text_input("RECIBIDO POR", #value=row_data.get('RECIBIDO POR', ''), 
                                                         key=f"asist_{i}")

                        new_observacion = st.text_area("OBSERVACIONES", #value=row_data.get('OBSERVACIONES', ''), 
                                                       key=f"obs_{i}")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("Guardar Cambios", key=f"save_{i}"):
                                # Actualizar los valores en la hoja de cálculo
                                cols_to_update = {
                                    'PRODUCTO CAMBIO': new_producto_cambio,
                                    'TIPO OPERACION': new_tipo,
                                    'RECIBIDO POR': new_recibido_por,
                                    'ENCARGADO  CAMBIO': new_encargado,
                                    'OBSERVACIONES': new_observacion
                                }
                                
                                success = True
                                for col_name, new_value in cols_to_update.items():
                                    if col_name in df.columns:
                                        col_idx = df.columns.get_loc(col_name) + 1  # +1 porque gspread usa índices basados en 1
                                        update_success = update_google_sheet(creds, row_index, col_idx, new_value)
                                        if not update_success:
                                            success = False
                                            break
                                
                                if success:
                                    st.success("¡Datos actualizados correctamente!")
                                    # Recargar los datos para mostrar los cambios
                                    result = get_google_sheet_data(creds)
                                    if result is not None:
                                        df, worksheet = result
                                    # Salir del modo edición
                                    st.session_state.editing = False
                                    st.session_state.edit_index = None
                                    st.rerun()
                        with col2:
                            if st.button("Cancelar", key=f"cancel_{i}"):
                                st.session_state.editing = False
                                st.session_state.edit_index = None
                                st.rerun()

            # Botón de descarga
            csv = filtered_df.to_csv(index=False)
            st.download_button(
                label="Descargar resultados como CSV",
                data=csv,
                file_name="resultados_devoluciones.csv",
                mime="text/csv"
            )

        else:
            st.warning('No se encontraron registros con los filtros seleccionados')

if __name__ == '__main__':
   consulta_recibido()