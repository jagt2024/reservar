import streamlit as st
from google_sheets_abo import GoogleSheet
from google_calendar_abo import GoogleCalendar
from sendemail_abo import send_email2
from sendemail_empresa_abo import send_email_emp
import numpy as np
import datetime as dt
import datetime
import re
from openpyxl import load_workbook
import os 
import sys
import logging
from typing import List, Optional
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='eliminar_reserva_abo.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

datos_book = load_workbook("archivos/parametros_abogados.xlsx", read_only=False) 

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       idcalendar = _row[2]
       break
  return idcalendar

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       precio = _row[1]
       #print(f'su correo es {_row[1]}')
  return precio
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       emailenc = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return emailenc

def validate_email(email):
  pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
  if re.match(pattern, email):
    return True
  else:
    return False
  
def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheets']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def consultar_otros(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-abo')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['PROCESO'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            return False, None
            
        # Extraer los campos solicitados
        datos_reserva = {
            'ENCARGADO': reserva['ENCARGADO'].iloc[0],
            'SERVICIOS': reserva['SERVICIOS'].iloc[0],
            'PRECIO': reserva['PRECIO'].iloc[0],
            'ACCION': reserva['ACCION'].iloc[0]
        }

        return True, datos_reserva
      
    except Exception as e:
        st.error(f"Error al consultar otros: {str(e)}")    
        return True, datos_reserva

def consultar_reserva(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-abo')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['PROCESO'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False #, None

def eliminar_reserva_sheet(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-abo')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['PROCESO'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            st.warning("No se encontró la reserva a eliminar.")
            return False
        
        # Obtener el índice de la fila a eliminar
        row_index = int(reserva.index[0]) + 2 
                
        # Eliminar la fila usando delete_rows (inicio, fin)
        worksheet.delete_rows(row_index, row_index)
        
        st.success("Reserva eliminada exitosamente.")
        return True
        
    except Exception as e:
        st.error(f"Error al eliminar la reserva: {str(e)}")
        return False

class EliminarReserva:
  
  class Model:
    pageTitle = "***Eliminar Agenda***"
  try:

    def view(self,model):
      st.title(model.pageTitle)
    
      horas = dataBook("horario")
      servicio = dataBook("servicio")
      #result_serv = np.setdiff1d(servicio,'')       
      encargado = dataBook("encargado")
      #result_estil = np.setdiff1d(encargado,'X')
      #estados = dataBook("estado")
      #result_estado = np.setdiff1d(estados,'')
    
      document='gestion-reservas-abo'
      sheet = 'reservas'
      credentials = st.secrets['sheets']['credentials_sheet']
      time_zone = 'GMT-05:00' # 'South America'
       
      st.subheader('Eliminar Reserva')
    
      #result_hours = np.setdiff1d(horas, "00:00") 
        
      c1, c2 = st.columns(2)
      
      with c1:
        nombre = st.text_input('Numero de Proceso*: ', placeholder='Numero Proceso') # label_visibility='hidden')
        #servicios = st.selectbox('servicio: ',servicio)           
        #encargado = st.selectbox('Encargado',encargado)
        fecha  = st.date_input('Fecha*: ')
        
      with c2:
        #estados = st.selectbox('estado: ',result_estado)
        
        hora = st.selectbox('Hora: ', horas)
        email  = st.text_input('Email*:', placeholder='Email')
      
      if nombre and hora !=  dt.datetime.utcnow().strftime("%H%M"): 
 
        with st.form(key='myform3',clear_on_submit=True):
      
         eliminar = st.form_submit_button('Eliminar')

         if eliminar:
            
          with st.spinner('Cargando...'):
        
           if not nombre or not email:
             st.warning('Se Require completar los campos para cosulta y Modificcacion')

           elif not validate_email(email):
             st.warning('El email no es valido')

           else:                            
           
             existe_db2 = consultar_reserva(nombre, str(fecha), hora)

             if existe_db2:

                valida, result = consultar_otros(nombre, str(fecha), hora)

                if valida:       
                  encargado = result['ENCARGADO']
                  servicios = result['SERVICIOS']
                  precio = result['PRECIO']
                  accion = result['ACCION']

                else:
                  # Si hay error, result será un diccionario
                  print(f"Error: {result['message']}") 
                  
                eliminar_reserva_sheet(nombre, str(fecha), hora)
                  
                #gs = GoogleSheet(credentials, document, sheet)
                #range = gs.write_data_by_uid(uid, values)

                #calendar.delete_event()
                                          
                send_email2(email, nombre, fecha, hora, servicios, precio, encargado, 'De acuerdo con su solicitud se cancelo la reserva. Gracias por su atencion.')
                send_email_emp(email, nombre, fecha, hora, servicios, precio, encargado, 'De acuerdo con su solicitud se cancelo la reserva. Gracias por su atencion.')
                
                st.success('Su solicitud ha sido cacelada de forrma exitosa')
             else:
                st.success('Reserva no existe para el cliente en esa Fecha y Hora por favor verifique')

  except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")
        print(f'Error crítico en la aplicación: {str(e)}')

sys.excepthook = global_exception_handler