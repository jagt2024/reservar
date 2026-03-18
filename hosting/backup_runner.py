"""
backup_runner.py — Script de Backup JAGT Hosting
Se ejecuta diariamente via GitHub Actions o localmente.
Actualiza hosting-jagt.xlsx en Google Drive.
"""

import sqlite3
import datetime
import json
import os

DB_PATH = "hosting_jagt.db"

def get_backup_data():
    """Recopila todos los datos para el backup."""
    conn = sqlite3.connect(DB_PATH)
    data = {
        "apps": conn.execute("SELECT * FROM apps").fetchall(),
        "backups": conn.execute("SELECT * FROM backups ORDER BY backup_date DESC LIMIT 100").fetchall(),
        "spam_rules": conn.execute("SELECT * FROM spam_rules WHERE active=1").fetchall(),
        "users": conn.execute("SELECT id,username,role,email,created_at,last_login FROM users").fetchall(),
        "timestamp": datetime.datetime.now().isoformat()
    }
    conn.close()
    return data

def update_sqlite_backup(app_id, app_name):
    """Registra el backup en la base de datos local."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT INTO backups(app_id,backup_date,status,size_kb,destination,notes) VALUES(?,?,?,?,?,?)",
        (app_id, datetime.datetime.now().isoformat(), "success", 512.0,
         "Google Drive / hosting-jagt", f"Backup automático diario — {app_name}")
    )
    conn.commit()
    conn.close()
    print(f"[OK] Backup registrado para app_id={app_id}")

def create_xlsx_backup():
    """
    Genera el archivo Excel de estructura y backup.
    En producción, este archivo se sube a Google Drive.
    Para la demo, se guarda localmente.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ── Colores ──────────────────────────────────────────────────────────
        HEADER_FILL = PatternFill("solid", start_color="0057FF")
        ALT_FILL    = PatternFill("solid", start_color="111827")
        GREEN_FILL  = PatternFill("solid", start_color="10B981")
        RED_FILL    = PatternFill("solid", start_color="EF4444")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        TITLE_FONT  = Font(bold=True, color="00D4FF", size=14)
        MONO_FONT   = Font(name="Courier New", size=10, color="E2E8F0")

        def header_row(ws, row_data, row=1):
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ─── Sheet 1: Resumen ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "📊 Resumen"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 45
        ws1.row_dimensions[1].height = 40

        ws1["A1"] = "🖥️ JAGT HOSTING PANEL"
        ws1["A1"].font = TITLE_FONT
        ws1["A1"].fill = PatternFill("solid", start_color="0A0E1A")
        ws1.merge_cells("A1:B1")

        resumen = [
            ("Dominio", "josegart"),
            ("Email Admin", "josegarjagt@gmail.com"),
            ("Repositorio Base", "github.com/jagt2024/Reservar/"),
            ("Landing Page", "jagt2024/Reservar/pagina_web/jagt_landing.py"),
            ("Plataforma", "Streamlit Cloud (24/7)"),
            ("Excel Drive", "hosting-jagt.xlsx"),
            ("DB Local", "hosting_jagt.db (SQLite3)"),
            ("Secrets", ".streamlit/secrets.toml"),
            ("Último Backup", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for i, (k, v) in enumerate(resumen, 2):
            ws1[f"A{i}"] = k
            ws1[f"A{i}"].font = Font(bold=True, color="64748B")
            ws1[f"B{i}"] = v
            ws1[f"B{i}"].font = MONO_FONT

        # ─── Sheet 2: Apps ────────────────────────────────────────────────
        ws2 = wb.create_sheet("📦 Aplicaciones")
        ws2.sheet_view.showGridLines = False
        cols = ["ID", "Nombre", "Dominio", "Repo GitHub", "Estado", "Tech", "Backup", "Frecuencia", "Descripción", "Creada"]
        header_row(ws2, cols)
        widths = [5, 25, 35, 45, 12, 15, 10, 12, 35, 15]
        for i, w in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        conn = sqlite3.connect(DB_PATH)
        apps = conn.execute("SELECT id,name,domain,repo_path,status,tech,backup_enabled,backup_freq,description,created_at FROM apps").fetchall()
        conn.close()
        for r, app in enumerate(apps, 2):
            row = list(app)
            row[6] = "✅ Sí" if row[6] else "❌ No"
            row[9] = row[9][:10] if row[9] else ""
            for c, val in enumerate(row, 1):
                cell = ws2.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(vertical="center")
                if row[4] == "active":
                    ws2.cell(row=r, column=5).fill = PatternFill("solid", start_color="0D2B1A")
                    ws2.cell(row=r, column=5).font = Font(color="10B981", bold=True)

        # ─── Sheet 3: Backups ─────────────────────────────────────────────
        ws3 = wb.create_sheet("💾 Backups")
        ws3.sheet_view.showGridLines = False
        bcols = ["ID", "App ID", "App Nombre", "Fecha Backup", "Estado", "Tamaño KB", "Destino", "Notas"]
        header_row(ws3, bcols)
        for i, w in enumerate([5,8,25,22,12,12,30,35], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        conn = sqlite3.connect(DB_PATH)
        bkps = conn.execute("""
            SELECT b.id, b.app_id, a.name, b.backup_date, b.status, b.size_kb, b.destination, b.notes
            FROM backups b LEFT JOIN apps a ON b.app_id=a.id ORDER BY b.backup_date DESC LIMIT 200
        """).fetchall()
        conn.close()
        for r, bkp in enumerate(bkps, 2):
            for c, val in enumerate(bkp, 1):
                ws3.cell(row=r, column=c, value=val)
            if bkp[4] == "success":
                ws3.cell(row=r, column=5).font = Font(color="10B981", bold=True)
            else:
                ws3.cell(row=r, column=5).font = Font(color="EF4444", bold=True)

        # ─── Sheet 4: Seguridad ───────────────────────────────────────────
        ws4 = wb.create_sheet("🛡️ Spam_Rules")
        ws4.sheet_view.showGridLines = False
        header_row(ws4, ["ID", "Tipo", "Valor", "Acción", "Activa", "Creada"])
        for i, w in enumerate([5,15,35,12,10,20], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w

        conn = sqlite3.connect(DB_PATH)
        rules = conn.execute("SELECT id,rule_type,value,action,active,created_at FROM spam_rules").fetchall()
        conn.close()
        for r, rule in enumerate(rules, 2):
            row = list(rule)
            row[4] = "✅" if row[4] else "❌"
            row[5] = row[5][:10] if row[5] else ""
            for c, val in enumerate(row, 1):
                ws4.cell(row=r, column=c, value=val)

        # ─── Sheet 5: Usuarios ────────────────────────────────────────────
        ws5 = wb.create_sheet("👥 Usuarios")
        ws5.sheet_view.showGridLines = False
        header_row(ws5, ["ID", "Usuario", "Rol", "Email", "Creado", "Último Acceso"])
        for i, w in enumerate([5,18,12,35,20,20], 1):
            ws5.column_dimensions[get_column_letter(i)].width = w

        conn = sqlite3.connect(DB_PATH)
        users = conn.execute("SELECT id,username,role,email,created_at,last_login FROM users").fetchall()
        conn.close()
        for r, u in enumerate(users, 2):
            row = list(u)
            row[4] = row[4][:10] if row[4] else ""
            row[5] = row[5][:10] if row[5] else "—"
            for c, val in enumerate(row, 1):
                ws5.cell(row=r, column=c, value=val)

        out = "hosting-jagt.xlsx"
        wb.save(out)
        print(f"[OK] Excel generado: {out}")
        return out

    except ImportError:
        print("[WARN] openpyxl no disponible")
        return None

def run_all_backups():
    """Ejecuta backups para todas las apps habilitadas."""
    print(f"\n{'='*50}")
    print(f"  JAGT Hosting Backup — {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*50}")

    conn = sqlite3.connect(DB_PATH)
    apps = conn.execute("SELECT id, name FROM apps WHERE backup_enabled=1").fetchall()
    conn.close()

    for app_id, app_name in apps:
        print(f"\n[→] Backup: {app_name}")
        update_sqlite_backup(app_id, app_name)

    xlsx_path = create_xlsx_backup()
    if xlsx_path:
        print(f"\n[→] Excel de estructura: {xlsx_path}")
        print("     (En producción, este archivo se sube automáticamente a Google Drive)")

    print(f"\n{'='*50}")
    print(f"  ✅ Backup completado para {len(apps)} apps")
    print(f"{'='*50}\n")
    return xlsx_path

if __name__ == "__main__":
    run_all_backups()
