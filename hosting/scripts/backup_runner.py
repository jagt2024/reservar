"""
backup_runner.py — JAGT Hosting Backup Completo
================================================
Ejecuta el backup diario completo:
  1. Inicializa la DB si no existe
  2. Genera hosting-jagt.xlsx con toda la estructura del hosting
  3. Sube / actualiza hosting-jagt.xlsx en Google Drive
  4. Lista TODOS los archivos .xlsx en la carpeta JAGT-Hosting
  5. Crea copia de respaldo de cada uno con sufijo _backup_FECHA
  6. Registra cada backup en hosting_jagt.db

Carpeta Drive  : JAGT-Hosting
Folder ID      : 1ueiD9ERvgFF9fZ7aMQUBtCuSGW1R4nSn
Cuenta servicio: servicio-cuenta-reserva-empres@appreservasem p.iam.gserviceaccount.com

Uso:
  python backup_runner.py            # backup completo
  python backup_runner.py --dry-run  # solo listar archivos sin hacer backup
"""

import sqlite3
import datetime
import json
import os
import sys
import io
from typing import Optional

# ── Configuración ─────────────────────────────────────────────────────────────

def clean_sheet_id(raw: str) -> str:
    """
    Extrae el Sheet ID puro desde cualquier formato:
      - URL completa : https://docs.google.com/spreadsheets/d/ID/edit#gid=0
      - Solo ID/edit : 1ABC.../edit
      - Solo ID      : 1ABC...
    Retorna solo el ID (cadena alfanumérica sin slashes ni parámetros).
    """
    s = raw.strip()
    # Caso 1: URL completa con /spreadsheets/d/
    if "spreadsheets/d/" in s:
        s = s.split("spreadsheets/d/")[1]
    # Eliminar todo lo que venga después del ID:
    # el ID termina antes de /, ?, # o espacios
    import re
    match = re.match(r'([A-Za-z0-9_\-]+)', s)
    if match:
        return match.group(1)
    return s

DB_PATH      = "hosting_jagt.db"
XLSX_NAME    = "hosting-jagt.xlsx"
DRIVE_FOLDER = "1ueiD9ERvgFF9fZ7aMQUBtCuSGW1R4nSn"
MIME_XLSX    = ("application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet")

# ── Credenciales ──────────────────────────────────────────────────────────────
def get_google_credentials() -> Optional[dict]:
    """
    Carga credenciales de cuenta de servicio Google.
    Orden: st.secrets → secrets.toml manual → variable de entorno
    """
    def _parse(raw) -> Optional[dict]:
        creds = json.loads(raw) if isinstance(raw, str) else dict(raw)
        if "private_key" not in creds or "client_email" not in creds:
            return None
        # Corregir saltos de línea del private_key
        creds["private_key"] = (
            creds["private_key"]
            .replace("\\\\n", "\n")
            .replace("\\n", "\n")
        )
        return creds

    # 1. st.secrets
    try:
        import streamlit as st
        raw = st.secrets["google"]["GOOGLE_CREDENTIALS"]
        c = _parse(raw)
        if c:
            print("[OK] Credenciales desde st.secrets")
            return c
    except Exception:
        pass

    # 2. secrets.toml manual
    candidates = [
        ".streamlit/secrets.toml",
        os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     ".streamlit", "secrets.toml"),
        os.path.expanduser("~/.streamlit/secrets.toml"),
    ]
    for toml_path in candidates:
        if not os.path.isfile(toml_path):
            continue
        try:
            import toml
            cfg = toml.load(toml_path)
            raw = cfg.get("google", {}).get("GOOGLE_CREDENTIALS")
            if not raw:
                continue
            c = _parse(raw)
            if c:
                print(f"[OK] Credenciales desde {toml_path}")
                return c
        except ImportError:
            try:
                try:
                    import tomllib
                except ImportError:
                    import tomli as tomllib
                with open(toml_path, "rb") as f:
                    cfg = tomllib.load(f)
                raw = cfg.get("google", {}).get("GOOGLE_CREDENTIALS")
                if raw:
                    c = _parse(raw)
                    if c:
                        print(f"[OK] Credenciales desde {toml_path} (tomllib)")
                        return c
            except Exception:
                continue
        except Exception:
            continue

    # 3. Variable de entorno
    env_raw = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if env_raw:
        try:
            c = _parse(env_raw)
            if c:
                print("[OK] Credenciales desde variable de entorno")
                return c
        except Exception as e:
            print(f"[ERROR] Variable GOOGLE_CREDENTIALS_JSON inválida: {e}")

    print("[WARN] No se encontraron credenciales de Google Drive.")
    return None


# ── Cliente Drive ─────────────────────────────────────────────────────────────
def get_drive_service(creds_dict: dict):
    """Crea y retorna el cliente autenticado de Google Drive API."""
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        credentials = service_account.Credentials.from_service_account_info(
            creds_dict, scopes=["https://www.googleapis.com/auth/drive"]
        )
        return build("drive", "v3", credentials=credentials,
                     cache_discovery=False)
    except Exception as e:
        print(f"[ERROR] No se pudo autenticar con Google Drive: {e}")
        return None


# ── Listar archivos xlsx en Drive ─────────────────────────────────────────────
def list_drive_xlsx(service, folder_id: str) -> list:
    """
    Lista todos los archivos .xlsx en la carpeta JAGT-Hosting.
    Retorna lista de dicts: {id, name, size, modifiedTime}
    """
    try:
        query = (
            f"'{folder_id}' in parents "
            f"and mimeType='{MIME_XLSX}' "
            f"and trashed=false"
        )
        results = service.files().list(
            q=query,
            fields="files(id, name, size, modifiedTime)",
            orderBy="name",
            pageSize=200
        ).execute()
        return results.get("files", [])
    except Exception as e:
        print(f"[ERROR] No se pudo listar archivos de Drive: {e}")
        return []


# ── Subir o actualizar un archivo en Drive ────────────────────────────────────
def upload_file_to_drive(service, local_path: str, file_name: str,
                         folder_id: str, update_existing: bool = True) -> Optional[str]:
    """
    Sube un archivo local a Drive.
    Si update_existing=True y ya existe un archivo con ese nombre, lo actualiza.
    Retorna el file_id del archivo creado/actualizado, o None si falla.
    """
    try:
        from googleapiclient.http import MediaFileUpload

        media = MediaFileUpload(local_path, mimetype=MIME_XLSX, resumable=True)

        if update_existing:
            query = (f"name='{file_name}' and '{folder_id}' in parents "
                     f"and trashed=false")
            results = service.files().list(
                q=query, fields="files(id, name)"
            ).execute()
            existing = results.get("files", [])

            if existing:
                file_id = existing[0]["id"]
                service.files().update(
                    fileId=file_id, media_body=media
                ).execute()
                print(f"[OK] Actualizado en Drive  : {file_name}  (ID: {file_id})")
                return file_id

        # Crear nuevo
        file_meta = {"name": file_name, "parents": [folder_id]}
        created = service.files().create(
            body=file_meta, media_body=media, fields="id"
        ).execute()
        file_id = created["id"]
        print(f"[OK] Creado en Drive       : {file_name}  (ID: {file_id})")
        return file_id

    except Exception as e:
        print(f"[ERROR] No se pudo subir '{file_name}': {e}")
        return None


# ── Exportar Google Sheet → .xlsx y subir a Drive ────────────────────────────
def export_sheet_to_drive(service, sheet_id: str, app_name: str,
                          folder_id: str, app_id: int,
                          now_str: str) -> bool:
    """
    Exporta un Google Sheet como .xlsx usando la Drive API (exportMedia),
    lo sube a la carpeta JAGT-Hosting con dos archivos:
      - app_name.xlsx            (versión actual, se sobreescribe)
      - app_name_backup_FECHA.xlsx (copia histórica)
    Registra ambas operaciones en hosting_jagt.db.
    """
    from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

    safe_name   = app_name.replace(" ", "_").replace("/", "-")
    current_name = f"{safe_name}.xlsx"
    backup_name  = f"{safe_name}_backup_{now_str}.xlsx"
    mime_export  = ("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet")

    print(f"\n  [→] Exportando Sheet: {app_name}")
    print(f"      Sheet ID  : {sheet_id}")
    print(f"      Archivo   : {current_name}")
    print(f"      Backup    : {backup_name}")

    # ── Paso 1: Exportar el Sheet como xlsx ──────────────────────────────────
    try:
        request  = service.files().export_media(
            fileId=sheet_id, mimeType=mime_export
        )
        buf  = io.BytesIO()
        dl   = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = dl.next_chunk()
        file_bytes = buf.getvalue()
        size_kb    = len(file_bytes) / 1024
        print(f"      Exportado : {size_kb:.1f} KB")
    except Exception as e:
        msg = str(e)
        print(f"      [ERROR] No se pudo exportar: {msg}")
        _register_backup_db(app_id, app_name, "error", 0,
                            f"Google Drive / JAGT-Hosting / {backup_name}",
                            f"Error exportando Sheet {sheet_id}: {msg[:200]}")
        return False

    # ── Paso 2: Subir versión actual (sobreescribir si existe) ───────────────
    def _upload(name: str, data: bytes, update: bool) -> Optional[str]:
        media = MediaIoBaseUpload(
            io.BytesIO(data), mimetype=mime_export, resumable=False
        )
        if update:
            q = (f"name='{name}' and '{folder_id}' in parents "
                 f"and trashed=false")
            res = service.files().list(q=q, fields="files(id)").execute()
            existing = res.get("files", [])
            if existing:
                fid = existing[0]["id"]
                service.files().update(fileId=fid, media_body=media).execute()
                print(f"      [OK] Actualizado en Drive : {name}")
                return fid
        meta    = {"name": name, "parents": [folder_id]}
        created = service.files().create(
            body=meta, media_body=media, fields="id"
        ).execute()
        print(f"      [OK] Creado en Drive      : {name}")
        return created["id"]

    try:
        _upload(current_name, file_bytes, update=True)
        _upload(backup_name,  file_bytes, update=False)
    except Exception as e:
        print(f"      [ERROR] No se pudo subir a Drive: {e}")
        _register_backup_db(app_id, app_name, "error", size_kb,
                            f"Google Drive / JAGT-Hosting / {backup_name}",
                            f"Error subiendo a Drive: {str(e)[:200]}")
        return False

    # ── Paso 3: Registrar ambos en la DB ─────────────────────────────────────
    _register_backup_db(
        app_id, app_name, "success", size_kb,
        f"Google Drive / JAGT-Hosting / {current_name}",
        f"Versión actual exportada desde Google Sheets ({sheet_id})"
    )
    _register_backup_db(
        app_id, app_name, "success", size_kb,
        f"Google Drive / JAGT-Hosting / {backup_name}",
        f"Backup histórico exportado desde Google Sheets ({sheet_id})"
    )
    return True


def _register_backup_db(app_id: int, app_name: str, status: str,
                        size_kb: float, destination: str, notes: str):
    """Inserta un registro en la tabla backups."""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute(
            """INSERT INTO backups
               (app_id, backup_date, status, size_kb, destination, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (app_id, datetime.datetime.now().isoformat(),
             status, round(size_kb, 2), destination, notes)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"      [WARN] No se pudo registrar en DB: {e}")


# ── Copiar archivo dentro de Drive (backup) ───────────────────────────────────
def backup_xlsx_in_drive(service, file_id: str, file_name: str,
                         folder_id: str, app_id: int,
                         now_str: str) -> bool:
    """
    Descarga un .xlsx de Drive y lo sube como copia de respaldo
    con sufijo _backup_FECHA en la misma carpeta.
    Registra el resultado en hosting_jagt.db.
    """
    from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

    base_name   = file_name.replace(".xlsx", "")
    backup_name = f"{base_name}_backup_{now_str}.xlsx"
    status      = "error"
    size_kb     = 0.0
    notes       = ""

    try:
        # 1. Descargar archivo original
        request = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        file_bytes = buf.getvalue()
        size_kb    = len(file_bytes) / 1024

        # 2. Subir copia con nombre de backup
        media = MediaIoBaseUpload(
            io.BytesIO(file_bytes), mimetype=MIME_XLSX, resumable=False
        )
        file_meta = {"name": backup_name, "parents": [folder_id]}
        created   = service.files().create(
            body=file_meta, media_body=media, fields="id"
        ).execute()

        status = "success"
        notes  = f"Backup automático de '{file_name}' → '{backup_name}'"
        print(f"[OK] Backup creado         : {backup_name}  "
              f"({size_kb:.1f} KB)")

    except Exception as e:
        notes = f"Error en backup de '{file_name}': {str(e)[:200]}"
        print(f"[ERROR] {notes}")

    # 3. Registrar en DB
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.execute(
            """INSERT INTO backups
               (app_id, backup_date, status, size_kb, destination, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                app_id,
                datetime.datetime.now().isoformat(),
                status,
                round(size_kb, 2),
                f"Google Drive / JAGT-Hosting / {backup_name}",
                notes,
            )
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] No se pudo registrar en DB: {e}")

    return status == "success"


# ── Registrar backup general en SQLite ────────────────────────────────────────
def register_backup(app_id: int, app_name: str, destination: str,
                    size_kb: float = 0.0, notes: str = ""):
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """INSERT INTO backups
           (app_id, backup_date, status, size_kb, destination, notes)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (app_id, datetime.datetime.now().isoformat(), "success",
         round(size_kb, 2), destination,
         notes or f"Backup automático diario — {app_name}")
    )
    conn.commit()
    conn.close()


# ── Generar Excel de estructura ───────────────────────────────────────────────
def create_xlsx_backup() -> Optional[str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl no disponible — omitiendo Excel")
        return None

    wb          = openpyxl.Workbook()
    HEADER_FILL = PatternFill("solid", start_color="0057FF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, color="00D4FF", size=14)
    MONO_FONT   = Font(name="Courier New", size=10, color="333333")

    def hrow(ws, cols, row=1):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja 1: Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 48
    ws1.row_dimensions[1].height = 36
    ws1["A1"] = "JAGT HOSTING PANEL"
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].fill = PatternFill("solid", start_color="0A0E1A")
    ws1.merge_cells("A1:B1")
    info = [
        ("Dominio",         "josegart"),
        ("Email Admin",     "josegarjagt@gmail.com"),
        ("Repositorio",     "github.com/jagt2024/reservar/"),
        ("Plataforma",      "Streamlit Cloud (24/7)"),
        ("Carpeta Drive",   "JAGT-Hosting"),
        ("Folder ID",       DRIVE_FOLDER),
        ("Excel Hosting",   XLSX_NAME),
        ("DB Local",        "hosting_jagt.db (SQLite3)"),
        ("Último Backup",   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(info, 2):
        ws1[f"A{i}"] = k
        ws1[f"A{i}"].font = Font(bold=True, color="555555")
        ws1[f"B{i}"] = v
        ws1[f"B{i}"].font = MONO_FONT

    conn = sqlite3.connect(DB_PATH)

    # Hoja 2: Aplicaciones
    ws2 = wb.create_sheet("Aplicaciones")
    ws2.sheet_view.showGridLines = False
    hrow(ws2, ["ID","Nombre","Dominio","Repo GitHub","Estado","Tech",
               "Backup","Frecuencia","Descripción","Creada"])
    set_widths(ws2, [5,25,35,45,12,15,10,12,35,15])
    for r, row in enumerate(conn.execute(
        "SELECT id,name,domain,repo_path,status,tech,backup_enabled,"
        "backup_freq,description,created_at FROM apps"
    ).fetchall(), 2):
        row = list(row)
        row[6] = "Si" if row[6] else "No"
        row[9] = (row[9] or "")[:10]
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val).alignment = Alignment(
                vertical="center")
        fc = "10B981" if row[4] == "active" else "EF4444"
        ws2.cell(row=r, column=5).font = Font(color=fc, bold=True)

    # Hoja 3: Backups
    ws3 = wb.create_sheet("Backups")
    ws3.sheet_view.showGridLines = False
    hrow(ws3, ["ID","App ID","App Nombre","Fecha Backup","Estado",
               "Tamaño KB","Destino","Notas"])
    set_widths(ws3, [5,8,25,22,12,12,35,45])
    for r, row in enumerate(conn.execute("""
        SELECT b.id, b.app_id, a.name, b.backup_date, b.status,
               b.size_kb, b.destination, b.notes
        FROM backups b LEFT JOIN apps a ON b.app_id=a.id
        ORDER BY b.backup_date DESC LIMIT 500
    """).fetchall(), 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
        fc = "10B981" if row[4] == "success" else "EF4444"
        ws3.cell(row=r, column=5).font = Font(color=fc, bold=True)

    # Hoja 4: Spam Rules
    ws4 = wb.create_sheet("Spam_Rules")
    ws4.sheet_view.showGridLines = False
    hrow(ws4, ["ID","Tipo","Valor","Acción","Activa","Creada"])
    set_widths(ws4, [5,15,35,12,10,20])
    for r, row in enumerate(conn.execute(
        "SELECT id,rule_type,value,action,active,created_at FROM spam_rules"
    ).fetchall(), 2):
        row = list(row)
        row[4] = "Si" if row[4] else "No"
        row[5] = (row[5] or "")[:10]
        for c, val in enumerate(row, 1):
            ws4.cell(row=r, column=c, value=val)

    # Hoja 5: Usuarios
    ws5 = wb.create_sheet("Usuarios")
    ws5.sheet_view.showGridLines = False
    hrow(ws5, ["ID","Usuario","Rol","Email","Creado","Último Acceso"])
    set_widths(ws5, [5,18,12,38,20,20])
    for r, row in enumerate(conn.execute(
        "SELECT id,username,role,email,created_at,last_login FROM users"
    ).fetchall(), 2):
        row = list(row)
        row[4] = (row[4] or "")[:10]
        row[5] = (row[5] or "—")[:10]
        for c, val in enumerate(row, 1):
            ws5.cell(row=r, column=c, value=val)

    conn.close()
    wb.save(XLSX_NAME)
    size_kb = os.path.getsize(XLSX_NAME) / 1024
    print(f"[OK] Excel generado        : {XLSX_NAME}  ({size_kb:.1f} KB)")
    return XLSX_NAME


# ── Inicializar DB ────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS apps (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        name           TEXT NOT NULL,
        domain         TEXT,
        repo_path      TEXT,
        status         TEXT DEFAULT 'active',
        tech           TEXT DEFAULT 'streamlit',
        description    TEXT,
        created_at     TEXT,
        backup_enabled INTEGER DEFAULT 1,
        backup_freq    TEXT DEFAULT 'daily',
        sheet_id       TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS users (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        username      TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role          TEXT DEFAULT 'viewer',
        email         TEXT,
        created_at    TEXT,
        last_login    TEXT
    );
    CREATE TABLE IF NOT EXISTS backups (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        app_id      INTEGER,
        backup_date TEXT,
        status      TEXT,
        size_kb     REAL,
        destination TEXT,
        notes       TEXT,
        FOREIGN KEY(app_id) REFERENCES apps(id)
    );
    CREATE TABLE IF NOT EXISTS spam_rules (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        rule_type  TEXT,
        value      TEXT,
        action     TEXT DEFAULT 'block',
        created_at TEXT,
        active     INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS access_log (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        ip        TEXT, path TEXT, action TEXT,
        timestamp TEXT, blocked INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS secrets_config (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        app_name   TEXT, key_name TEXT,
        key_value  TEXT, updated_at TEXT
    );
    """)
    conn.commit()

    import hashlib
    if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        pw = hashlib.sha256("jagt2024!".encode()).hexdigest()
        conn.execute(
            "INSERT INTO users(username,password_hash,role,email,created_at)"
            " VALUES(?,?,?,?,?)",
            ("josegart", pw, "admin", "josegarjagt@gmail.com",
             datetime.datetime.now().isoformat())
        )
        conn.commit()
        print("[OK] Usuario admin creado")

    if conn.execute("SELECT COUNT(*) FROM apps").fetchone()[0] == 0:
        now  = datetime.datetime.now().isoformat()
        seed = [
            ("JAGT Landing Page",
             "josegart-landing.streamlit.app",
             "jagt2024/reservar/pagina_web/jagt_landing.py",
             "active","streamlit","Página web personal",now,1,"daily"),
            ("reservar App",
             "josegart-reservar.streamlit.app",
             "jagt2024/reservar/",
             "active","streamlit","Sistema de reservas",now,1,"daily"),
            ("Panel Hosting",
             "josegart-hosting.streamlit.app",
             "jagt2024/reservar/hosting/app.py",
             "active","streamlit","Panel de control",now,1,"daily"),
        ]
        conn.executemany(
            "INSERT INTO apps(name,domain,repo_path,status,tech,description,"
            "created_at,backup_enabled,backup_freq) VALUES(?,?,?,?,?,?,?,?,?)",
            seed
        )
        conn.commit()
        print(f"[OK] {len(seed)} apps iniciales registradas")

    # Migración: agregar sheet_id si no existe
    try:
        conn2 = sqlite3.connect(DB_PATH)
        conn2.execute("ALTER TABLE apps ADD COLUMN sheet_id TEXT DEFAULT ''")
        conn2.commit()
        conn2.close()
    except Exception:
        pass  # Columna ya existe
    conn.close()
    print(f"[OK] DB lista              : {os.path.abspath(DB_PATH)}")


# ── Runner principal ──────────────────────────────────────────────────────────
def run_all_backups(dry_run: bool = False):
    sep      = "=" * 58
    now_str  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    now_disp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"\n{sep}")
    print(f"  JAGT Hosting — Backup Completo")
    print(f"  {now_disp}{'  [DRY RUN]' if dry_run else ''}")
    print(f"  Carpeta Drive: JAGT-Hosting ({DRIVE_FOLDER})")
    print(f"{sep}")

    # ── PASO 0: Inicializar DB ────────────────────────────────────────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 0 · Verificar base de datos")
    print(f"{'─'*40}")
    init_db()

    # ── PASO 1: Generar Excel de estructura ───────────────────────────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 1 · Generar {XLSX_NAME}")
    print(f"{'─'*40}")
    xlsx_path = create_xlsx_backup()
    if not xlsx_path:
        print("[ERROR] No se pudo generar el Excel. Abortando.")
        return

    # ── PASO 2: Conectar a Google Drive ───────────────────────────────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 2 · Conectar a Google Drive")
    print(f"{'─'*40}")
    creds   = get_google_credentials()
    service = get_drive_service(creds) if creds else None

    if not service:
        print("[WARN] Sin conexión a Drive — los backups quedan solo en local.")
        drive_available = False
    else:
        print("[OK] Conexión a Google Drive establecida")
        drive_available = True

    # ── PASO 3: Subir hosting-jagt.xlsx (actualizar) ──────────────────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 3 · Subir {XLSX_NAME} a Drive")
    print(f"{'─'*40}")
    hosting_xlsx_ok = False
    if drive_available and not dry_run:
        fid = upload_file_to_drive(
            service, xlsx_path, XLSX_NAME, DRIVE_FOLDER, update_existing=True
        )
        hosting_xlsx_ok = fid is not None
    elif dry_run:
        print(f"[DRY] Omitiría subir: {XLSX_NAME}")
        hosting_xlsx_ok = True

    # ── PASO 4: Listar todos los .xlsx en Drive ───────────────────────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 4 · Listar archivos .xlsx en JAGT-Hosting")
    print(f"{'─'*40}")
    drive_files = []
    if drive_available:
        drive_files = list_drive_xlsx(service, DRIVE_FOLDER)
        print(f"[OK] {len(drive_files)} archivo(s) .xlsx encontrado(s):")
        for f in drive_files:
            size_kb = int(f.get("size", 0)) / 1024
            print(f"     📊 {f['name']:<45} {size_kb:>7.1f} KB"
                  f"  mod: {f.get('modifiedTime','')[:10]}")
    else:
        print("[WARN] Sin conexión — no se puede listar archivos de Drive")

    # ── PASO 5: Exportar Google Sheets + Backup de .xlsx en Drive ────────────
    print(f"\n{'─'*40}")
    print(f"  PASO 5 · Exportar Sheets y hacer backup de archivos")
    print(f"{'─'*40}")

    # Obtener apps con sus sheet_id desde la DB
    conn    = sqlite3.connect(DB_PATH)
    apps_db = conn.execute(
        "SELECT id, name, sheet_id FROM apps WHERE backup_enabled=1"
    ).fetchall()
    conn.close()

    results_backup = []   # (nombre, status, app_name)

    # ── 5A: Exportar cada Google Sheet registrado ─────────────────────────────
    apps_with_sheet    = [(aid, aname, sid) for aid, aname, sid in apps_db if sid and sid.strip()]
    apps_without_sheet = [(aid, aname) for aid, aname, sid in apps_db if not sid or not sid.strip()]

    if not drive_available:
        print("[WARN] Sin conexión a Drive — omitiendo exportación de Sheets")
    elif not apps_with_sheet:
        print("[WARN] Ninguna app tiene Sheet ID registrado.")
        print("       Agrega el Sheet ID de cada app en:")
        print("       Panel → Aplicaciones → Lista de Apps → (expander) → Sheet ID")
    else:
        print(f"\n  Exportando {len(apps_with_sheet)} Google Sheet(s)...")
        for app_id, app_name, sheet_id in apps_with_sheet:
            sid_clean = clean_sheet_id(sheet_id)
            print(f"  App        : {app_name}")
            print(f"  Sheet raw  : {sheet_id.strip()}")
            print(f"  Sheet ID   : {sid_clean}")

            if dry_run:
                print(f"  [DRY] Omitiría exportar Sheet: {app_name} ({sid_clean})")
                results_backup.append((f"{app_name}.xlsx", "dry_run", app_name))
                continue

            ok = export_sheet_to_drive(
                service, sid_clean, app_name, DRIVE_FOLDER, app_id, now_str
            )
            results_backup.append(
                (f"{app_name}.xlsx", "ok" if ok else "error", app_name)
            )

    if apps_without_sheet:
        print(f"\n  Apps sin Sheet ID ({len(apps_without_sheet)}) — omitidas:")
        for _, aname in apps_without_sheet:
            print(f"    ⚠️  {aname}")

    # ── 5B: Backup de archivos .xlsx YA existentes en Drive ───────────────────
    # (archivos que no son exports de Sheets sino archivos subidos manualmente)
    if drive_files and drive_available:
        print(f"\n  Backup de archivos .xlsx existentes en Drive...")
        apps_map = {
            a[1].lower().replace(" ", ""): (a[0], a[1])
            for a in apps_db
        }
        for f in drive_files:
            fid   = f["id"]
            fname = f["name"]

            # Omitir archivos que ya son backups o exports de apps
            if "_backup_" in fname:
                print(f"  [SKIP] {fname}  ← ya es backup")
                continue
            # Omitir si ya fue generado en 5A (mismo nombre que un app export)
            app_exports = {
                f"{a[1].replace(' ','_').replace('/','_')}.xlsx"
                for a in apps_db
            }
            if fname in app_exports:
                print(f"  [SKIP] {fname}  ← generado en exportación de Sheet")
                continue

            matched_id, matched_name = 0, "General"
            fname_norm = fname.lower().replace(" ","").replace(".xlsx","")
            for key, (aid, aname) in apps_map.items():
                if key in fname_norm or fname_norm in key:
                    matched_id, matched_name = aid, aname
                    break

            if dry_run:
                print(f"  [DRY] Omitiría backup: {fname}")
                results_backup.append((fname, "dry_run", matched_name))
                continue

            print(f"  [→] Backup Drive file: {fname}  (app: {matched_name})")
            ok = backup_xlsx_in_drive(
                service, fid, fname, DRIVE_FOLDER, matched_id, now_str
            )
            results_backup.append((fname, "ok" if ok else "error", matched_name))

    # Registrar backup del Excel de estructura del hosting
    if hosting_xlsx_ok and not dry_run:
        for app_id, app_name, _ in apps_db:
            register_backup(
                app_id, app_name,
                f"Google Drive / JAGT-Hosting / {XLSX_NAME}",
                size_kb=os.path.getsize(xlsx_path) / 1024,
                notes="Backup automático diario — estructura hosting"
            )

    # ── Resumen ───────────────────────────────────────────────────────────────
    ok_count  = sum(1 for _, s, _ in results_backup if s == "ok")
    err_count = sum(1 for _, s, _ in results_backup if s == "error")
    skipped   = sum(1 for _, s, _ in results_backup if s == "dry_run")

    print(f"\n{sep}")
    print(f"  RESUMEN DEL BACKUP — {now_disp}")
    print(f"{sep}")
    print(f"  Excel hosting subido : {'✅' if hosting_xlsx_ok else '❌'} {XLSX_NAME}")
    print(f"  Archivos en Drive    : {len(drive_files)}")
    print(f"  Backups exitosos     : {ok_count}")
    if err_count:
        print(f"  Backups con error    : {err_count}")
    if skipped:
        print(f"  Omitidos (dry-run)   : {skipped}")
    apps_con_sheet = sum(1 for _,_,s in apps_db if s and s.strip())
    print(f"  Apps respaldadas     : {len(apps_db)}")
    print(f"  Apps con Sheet ID    : {apps_con_sheet}")
    print(f"  URL Drive            : https://drive.google.com/drive/folders/{DRIVE_FOLDER}")
    print(f"{sep}\n")

    if results_backup:
        print("  Detalle de backups:")
        for fname, status, aname in results_backup:
            icon = {"ok":"✅","error":"❌","dry_run":"🔵"}.get(status,"❓")
            print(f"    {icon} {fname:<45} → {aname}")
        print()


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    run_all_backups(dry_run=dry)
