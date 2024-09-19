import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import os
import toml
import base64
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
import datetime
import time
import io
from openpyxl import load_workbook

datos_book = load_workbook("archivos/parametros_empresa.xlsx", read_only=False)

def dataBookSheetUrl(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row)
      url = _row[1]
    return url

def load_credentials_from_toml(file_path):
    with open(file_path, 'r') as toml_file:
        config = toml.load(toml_file)
        credentials = config['sheetsemp']['credentials_sheet']
    return credentials
    #config['credentials_sheet']

sheetUrl = dataBookSheetUrl("sw")

def get_google_sheet_data(creds):
  try:
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_info(creds, scopes=scope)
    client = gspread.authorize(credentials)

    sheet_url = str(sheetUrl)
    sheet = client.open_by_url(sheet_url)
    worksheet = sheet.worksheet('reservas')
    data = worksheet.get_all_values()
    
    if not data:
            st.error("No se encontraron datos en la hoja de cálculo.")
            return None
    
    df = pd.DataFrame(data[1:], columns=data[0])
    
    if df.empty:
       st.error("El DataFrame está vacío después de cargar los datos.")
       return None
    
    df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
    
    # Informar sobre filas con fechas inválidas
    invalid_dates = df[df['FECHA'].isna()]
    if not invalid_dates.empty:
        st.warning(f"Se encontraron {len(invalid_dates)} filas con fechas inválidas. Estas filas serán excluidas del análisis.")
        st.write("Primeras 5 filas con fechas inválidas:")
        st.write(invalid_dates.head())
    
    if df.empty:
        st.error("El DataFrame está vacío después de eliminar las fechas inválidas.")
        return None
    # Eliminar filas con fechas inválidas
    
    df = df.dropna(subset=['FECHA'])
    
    return df

  except Exception as e:
     st.error(f"Error al obtener datos de Google Sheets: {str(e)}")
     return None

def get_binary_file_downloader_html(bin_file, file_label='File'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">Descargar {file_label}</a>'
    return href

def filter_data_by_last_days(df, num_days=8):
    if df is None or df.empty:
        st.error("No hay datos para filtrar.")
        return None

    today = pd.Timestamp.today().date()
    start_date = today - pd.Timedelta(days=num_days-1)
    
    # Filtrar los datos, ignorando las fechas nulas
    filtered_df = df[df['FECHA'].notna() & (df['FECHA'].dt.date >= start_date)]
    
    if filtered_df.empty:
        st.warning("No hay datos en los últimos 8 días.")
        return None
    
    return filtered_df

def process_and_display_data(df):
  try:
      
    if df is None or df.empty:
       st.error("No hay datos para procesar.")
       return

    df = filter_data_by_last_days(df)
        
    if df is None or df.empty:
       st.error("No hay datos después de filtrar por los últimos días.")
       return
    
    temp_file_path = "./archivos/temp_gestion-reservas-emp.xlsx"
    df.to_excel(temp_file_path, index=False)
    
    # Add a download button for the generated file
    with open(temp_file_path, "rb") as file:
        btn = st.download_button(
            label="Descargar archivo Excel",
            data=file,
            file_name="gestion-reservas-emp.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    st.write(f"Se han procesado {len(df)} registros válidos.")
    #st.write("Primeros 5 registros de la hoja:")
    #st.dataframe(df.head())

    #st.write("Últimos 5 registros de la hoja:")
    #st.dataframe(df.tail())
    
    #st.markdown(get_binary_file_downloader_html(temp_file_path, 'Excel'), #unsafe_allow_html=True)

    #st.write("Estadísticas básicas:")
    #st.write(df.describe())

    #if 'fecha' in df.columns and 'monto' in df.columns:
    #    st.write("Gráfico de Reservas por Fecha:")
    #    chart_data = df.groupby('fecha')['monto'].sum().reset_index()
    #    st.bar_chart(chart_data.set_index('fecha'))

  except Exception as e:
        st.error(f"Error al procesar los datos: {str(e)}")
        #st.write("Detalles del DataFrame:")
        #st.write(df.dtypes)
        #st.write(df.head()) 

def download_and_process_data(creds_path):
    try:
        creds = load_credentials_from_toml(creds_path)
        st.success("Credenciales cargadas correctamente")
    except Exception as e:
        st.error(f"Error al cargar las credenciales: {str(e)}")
        return

    try:
        with st.spinner('Descargando datos...'):
           df = get_google_sheet_data(creds)
        if df is not None and not df.empty:
           st.success('Datos descargados correctamente!')
           process_and_display_data(df)
        else:
            st.error("No se pudieron obtener datos válidos de Google Sheets.")
    except Exception as e:
        st.error(f"Error al procesar los datos: {str(e)}")

# Esta función main() puede ser eliminada si no se va a usar este script directamente
#def main():
#    st.title('Descarga de Google Sheets')
#    if st.button('Descargar Datos'):
#        download_and_process_data('path/to/your/credentials.toml')

#if __name__ == "__main__":
#    main()