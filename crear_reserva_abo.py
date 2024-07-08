import streamlit as st
from google_sheets_abo import GoogleSheet
from google_calendar_abo import GoogleCalendar
from sendemail_abo import send_email2
from sendemail_empresa_abo import send_email_emp
import numpy as np
import datetime
import datetime as dt
import re
import uuid
from openpyxl import load_workbook
#import pywhatkit
#import pyautogui, webbrowser
#from time import sleep

datos_book = load_workbook("archivos/parametros_abogados.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       idcalendar = _row[2]
       break
  return idcalendar

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       precio = _row[1]
       #print(f'su correo es {_row[1]}')
  return precio
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       emailenc = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return emailenc
 
def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")

def generate_uid():
    return str(uuid.uuid4())

class CrearReserva:
  
  class Model:
    pageTitle = "***Generar Agenda***"
  
  def view(self,model):
    st.title(model.pageTitle)

    with st.form(key='myform1',clear_on_submit=True):
  
      horas = dataBook("horario")
      
      estados = dataBook("estado")
      result_estado = np.setdiff1d(estados,'')
      
      jurisdicciones = dataBook("jurisdiccion")
      result_juris = np.setdiff1d(jurisdicciones,'')
      
      servicio = dataBook("servicio")
      result_serv = np.setdiff1d(servicio,'') 
    
      encargado = dataBook("encargado")
      result_estil = np.setdiff1d(encargado,'X') 
    
      partes = dataBook("parte-procesal")
      result_partes = np.setdiff1d(partes,'')
      
      document='gestion-reservas-abo'
      sheet = 'reservas'
      credentials = st.secrets['sheets']['credentials_sheet']
      time_zone = 'GMT-05:00' # 'South America'
     
      c1, c2 = st.columns(2)
      nombre = c1.text_input('Numero del Proceso o Nombre Ciente *: ', placeholder='Numero - Nombre') # label_visibility='hidden')
      estados = c2.selectbox('Estado*: ',result_estado)
      servicios = c1.selectbox('Servicio Juridico*: ',result_serv)
      partes = c2.selectbox('Partes Procesales', result_partes)
      jurisdicciones = c1.selectbox('Jurisdiccion*: ',result_juris)
      fecha  = c2.date_input('Fecha Agenda*: ')
      email  = c1.text_input('Email Empresa o Personal:', placeholder='Email')
          
      idcalendarserv = dataBookServicioId("servicio", servicios)
      #print(f'idcalendarserv = {idcalendarserv}')
      result_id = np.setdiff1d(idcalendarserv,'')
  
      #idcalendar = "josegarjagt@gmail.com"
                 
      if fecha:
        id = ""
        if servicios == servicio:
          id = result_id
        
      calendar = GoogleCalendar(id) #credentials, idcalendar
        
      encargado = c1.selectbox('Abogado Encargado:',result_estil)
      #hora = c2.selectbox('Hora: ',horas)
            
      emailencargado = dataBookEncEmail("encargado",encargado)
      result_email = np.setdiff1d(emailencargado,'X')
           
      hours_blocked = calendar.list_upcoming_events()
      result_hours = np.setdiff1d(horas, hours_blocked) 
      hora = c2.selectbox('Hora: ',result_hours)
      acciones = c2.text_area('Solicitud, Accion o Medio Control')
      hechos = c1.text_area('Hechos (Opcional)')
      causas = c2.text_area('Causas (Opcional)')
      precio = dataBookPrecio("servicio", servicios)
        
      whatsapp = c1.checkbox('Envio a WhatsApp Si/No (Opcional)')
      telefono = c1.text_input('Nro. Telefono')

      enviar = st.form_submit_button(" Reservar ")

      #Backend
      if enviar:
        with st.spinner('Cargando...'):
          if not nombre or not servicio or not encargado:
            st.warning('Se Require completar los campos con * son obligatorios')
        
          elif not validate_email(email):
            st.warning('El email no es valido')
        
          else:
          
            parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
            hours1 = parsed_time.hour
            minutes1 =  parsed_time.minute
                    
            end_hours = add_hour_and_half(hora)
          
            parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
            hours2 = parsed_time2.hour
            minutes2 =  parsed_time2.minute

            start_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours1,minutes1).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
                             
            end_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours2,minutes2).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
            gs = GoogleSheet(credentials, document, sheet)
            existe = False
            
            if len(gs.sheet.get_all_values()) +1 > 2:
                 
              last_row = len(gs.sheet.get_all_values()) +1
              print(f'last_row {last_row}')
              data = gs.sheet.get_values()
              print(f'data {data}')
              data2 = data[1:]
              #print(f'data2 {data2}')
              range_start = f"A{last_row}"
              #print(f'range_start {range_start}')
              range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
              #print(f'range_end {range_end}')
          
              for row in data2:
        
                nom = [row[0]]
                serv = [row[4]]
                fech = str(row[2])
                hora2 = str(row[3])
                uid1 = str(row[11])

                if nom != ['DATA']:
              
                  fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
                  fech1 = int(fech2.strftime("%Y%m%d"))
                  fechacalendarint = int(fecha.strftime("%Y%m%d"))
                  hora3 = datetime.datetime.strptime(hora2,'%H:%M')
                  fechahora_ini = int(hora3.strftime('%H%M'))
                  horacalendar = datetime.datetime.strptime(hora,'%H:%M')
                  horacalendarint = int(horacalendar.strftime('%H%M'))
                  #horawhat = int(horacalendar.strftime('%H'))
                  #minuto = int(horacalendar.strftime('%M'))
              
                  #print(f'Fechas y horas {fech1}, {fechacalendarint}, {fechahora_ini}, { horacalendarint}')
              
                  if nom == [nombre] and fech1 == fechacalendarint and fechahora_ini == horacalendarint:
                    existe = True
                    st.warning("El cliente ya tiene agenda para esa fecha y hora")
                    break
            
            if existe == False:
                       
                hora_actual = dt.datetime.now()
                hora_actual_int = int(hora_actual.strftime("%H%M"))
              
                hora_calendar = datetime.datetime.strptime(hora,'%H:%M')
                #print(hora_actual)
                hora_calendar_int = int(hora_calendar.strftime('%H%M'))
          
                hoy = dt.datetime.now()
                fechoy = int(hoy.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))
          
                if fechacalendarint >= fechoy:
                   
                  #if fechacalendarint == fechoy and hora_calendar_int < hora_actual_int:
            
                  #  st.warning('La hora seleccionda es invalida para hoy')
                  #  print('La hora seleccionda es invalida para hoy')
                   
                  #else:
                  whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el encargado: {encargado} para el servicio de : {servicios} con la accion {acciones}")
                  
                  uid = generate_uid()
                  values = [(nombre,email,str(fecha),hora,servicios,precio, encargado, partes, acciones, hechos, causas, uid, whatsapp, str(57)+telefono, whatsappweb)]
                  gs = GoogleSheet(credentials, document, sheet)
          
                  range = gs.get_last_row_range()
                  gs.write_data(range,values)
                    
                  calendar.create_event(servicios+". Proceso #: "+nombre, start_time, end_time, time_zone, attendees=result_email)

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): {nombre}, su soliciud de reserva se ha realizado de forma exitosa y se agendo el servicio {servicio}, para el dia {fecha}, a las {hora}. Gracias por su cofianza. Cordialmente aplicacion de Reservas y Agendamiento.'
                      # Para abrir en buscador Edge                  
                  #  enviarwhats = webbrowser.open('https://web.whatsapp.com/send?phone='+contact+"&text="+message)
                      # Para abrir en buscador chrome
                      #chrome_path = 'C:/Program Files(x86)/Google/Chrome/Application/chrome.exe %s'
                      #webbrowser.get(chrome_path).open('https://web.whatsapp.com/send?phone='+contact+"&text="+message)
                  
                  #  sleep(130)
                  #  pyautogui.click(1230.964)
                  #  sleep(5)
                    #pyautogui.typewrite(message)
                  #  pyautogui.press('enter')
                  #  sleep(2)
                  #  pyautogui.hotkey('ctrl','w')
                  #  sleep(1)
                
                      #pywhatkit.sendwhatmsg('+'+str(57)+contact, message, horawhat, minuto)
                  
                  st.success('Su solicitud ha sido reservada de forrma exitosa')
                  send_email2(email, nombre, fecha, hora, servicios, precio, encargado,  partes, acciones, hechos, causas)
                  send_email_emp(email, nombre, fecha, hora, servicios, precio, encargado, partes, acciones, hechos, causas, attendees=result_email)

