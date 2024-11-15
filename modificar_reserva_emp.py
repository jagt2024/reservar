import streamlit as st
from google_sheets_emp import GoogleSheet
from google_calendar_emp import GoogleCalendar
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
import datetime as dt
import  re
import datetime 
from  openpyxl import load_workbook
import os
import sys
import logging
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json
#import pywhatkit
#import pyautogui, webbrowser
#from time import sleep

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "5"
os.environ["REQUESTS_READ_TIMEOUT"] = "5"

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='modificar_reserva_emp.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

datos_book = load_workbook("archivos/parametros_empresa.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       idcalendar = _row[2]
       break
  return idcalendar

def dataBookPrecio(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       precio = _row[1]
       #print(f'su correo es {_row[1]}')
  return precio
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       emailenc = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return emailenc

def validate_email(email):
  pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
  if re.match(pattern, email):
    return True
  else:
    return False
  
def add_hour_and_half(time):
  parsed_time = dt.datetime.strptime(time, "%H:%M").time()
  new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
  return new_time.strftime("%H:%M")

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheets']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def consultar_reserva(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-emp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Normalizar el formato de fecha y hora para la búsqueda
        #try:
        #    fecha_busqueda = datetime.strptime(fecha, '%d-%m-%Y').strftime#('%d-%m-%Y')
        #    hora_busqueda = datetime.strptime(hora, '%H:%M').strftime('%H:%M')
        #except ValueError:
        #    st.error("Formato de fecha u hora inválido")
        #    return False, None
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:
            return False #, None
            
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False, None

def consultar_uuid(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-emp')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            return False, None
            
        # Extraer los campos solicitados
        datos_reserva = {
            'UID': reserva['UID'].iloc[0]         
        }
        
        return True, datos_reserva
        
    except Exception as e:
        st.error(f"Error al consultar UID: {str(e)}")
        return False,(f"Error al consultar UID: {str(e)}")

class ModificarReservaEmp:
 try: 
  class Model:
    pageTitle = "***Modificar Reserva***"
  
  def view(self,model):
    st.title(model.pageTitle)
      
    horas = dataBook('horario')
    horas_ant = dataBook('horario')
    
    servicio = dataBook('servicio')
    result_serv = np.setdiff1d(servicio,'')
    servicio_ant = dataBook('servicio')
    result_serv = np.setdiff1d(servicio_ant,'')
    
    encargado = dataBook("encargado")
    result_estil = np.setdiff1d(encargado,'X')
      
    encargado_ant = dataBook('encargado')
    result_est = np.setdiff1d(encargado_ant,'')
   
    document='gestion-reservas-emp'
    sheet = 'reservas'
    credentials = st.secrets['sheetsemp']['credentials_sheet']
    time_zone = 'GMT-05:00' # 'South America'
  
    st.subheader('Ingrese los datos de la agenda a Modificar') 
    
    c1, c2 = st.columns(2)
    with c1:
        nombre = st.text_input('Nombre entidad o persona*: ', placeholder='Nombre') # label_visibility='hidden')
        fecha_ant  = st.date_input('Fecha Agendada*: ')
        servicio_ant = st.selectbox('Servicio Agendado', result_serv)
        
    with c2:
        email  = st.text_input('Su Email*:', placeholder='Email')
        encargado_ant = st.selectbox('encargado Agndado',result_est)
        #hora_ant = c2.selectbox('Hora Agendada: ',horas)
        result_hours_ant = np.setdiff1d(horas, "00:00") 
        hora_ant = st.selectbox('Hora Ant: ',result_hours_ant)
      
    hora_act = dt.datetime.now()
    hora_actual= hora_act.strftime("%H:%M")
         
    if nombre and hora_actual != hora_ant:
     
        existe_db = consultar_reserva(nombre, str(fecha_ant), hora_ant)
      
        if existe_db:
          existe = True
        
          with st.container():
            st.write("---")
            st.subheader('Ingrese los datos para la Nueva Agenda')
      
            valida, result = consultar_uuid(nombre, str(fecha_ant), hora_ant)
      
            a1, a2 = st.columns(2)
          
            fecha  = a1.date_input('Fecha*: ')
            servicios = a1.selectbox('Servicios', result_serv)
        
            precio = dataBookPrecio("servicio", servicios)
            result_precio = np.setdiff1d(precio,'')
                        
            encargado = a2.selectbox('Encargado',result_est)
            emailencargado = dataBookEncEmail("encargado",encargado)
            result_email = np.setdiff1d(emailencargado,'X')
        
            idcalendarserv = dataBookServicioId("servicio", servicios)
            #print(f'idcalendarserv = {idcalendarserv}')
            result_id = np.setdiff1d(idcalendarserv,'')
  
            #idcalendar = "josegarjagt@gmail.com"
                 
            if fecha:
              id = ""
              if servicios == servicio:
                id = result_id
         
            #calendar = GoogleCalendar(id) #credentials, idcalendar
                    
            #hours_blocked = calendar.list_upcoming_events()
            #result_hours = np.setdiff1d(horas, hours_blocked) 
            
            hora = a2.selectbox('Hora: ', horas)
    
            notas = a1.text_area('Nota o Mensaje(Opcional)')
            whatsapp = a2.checkbox('Envio a WhatsApp Si/No (Opcional)')
            telefono = a2.text_input('Nro. Telefono')

            with st.form(key='myform',clear_on_submit=True):
             actualizar = st.form_submit_button('Actualizar')

             if actualizar:
              with st.spinner('Cargando...'):
                if not nombre or not email or not servicio or not encargado:
                  st.warning('Se Require completar los campos con * son obligatorios')
        
                elif not validate_email(email):
                  st.warning('El email no es valido')
        
                else:
                  parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
                  hours1 = parsed_time.hour
                  minutes1 =  parsed_time.minute
                    
                  end_hours = add_hour_and_half(hora)
          
                  parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
                  hours2 = parsed_time2.hour
                  minutes2 =  parsed_time2.minute

                  start_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours1,minutes1).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
                  end_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours2,minutes2).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
                  gs = GoogleSheet(credentials, document, sheet)
                   
                  #last_row = len(gs.sheet.get_all_values()) +1
                  #print(f'last_row {last_row}')
                  #data = gs.sheet.get_values()
                  #print(f'data {data}')
                  #data2 = data[1:]
                  #print(f'data2 {data2}')
                  #range_start = f"A{last_row}"
                  #print(f'range_start {range_start}')
                  #range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
                  #print(f'range_end {range_end}')
          
                  #for row in data2:
        
                    ##nom = [row[0]]
                    #serv = [row[4]]
                    #fech = str(row[2])
                    #hora2 = str(row[3])
                    #nota = [row[6]]
                    #uid1 = str(row[8])

                    #if nom != ['DATA']:
              
                    #fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
                    #fech1 = int(fech2.strftime("%Y%m%d"))
                    #fechacalendarint = int(fecha_ant.strftime("%Y%m%d"))
                    #hora3 = datetime.datetime.strptime(hora2,'%H:%M')
                    #fechahora_ini = int(hora3.strftime('%H%M'))
                    #horacalendar = datetime.datetime.strptime(hora_ant,'%H:%M')
                    #horacalendarint = int(horacalendar.strftime('%H%M'))
              
                    #print(f'Esto es elregistro anterior  {fech1}, {hora3}, {fechahora_ini}, {horacalendarint}')
                            
                    #if nom == [nombre] and fech1 == fechacalendarint and serv == [servicio_ant] and fechahora_ini == horacalendarint and nota != ["Agenda Cancelada"]: 
                
                    #fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
                    #print(f'Esto es fech2 {fech2}')
                
                    #fech1 = int(fech2.strftime("%Y%m%d"))
                
                    #hora3 = datetime.datetime.strptime(hora2,'%H:%M')
                
                    #hora_actual = dt.datetime.now()
                    ##hora_actual_int = int(hora_actual.strftime("%H%M"))
                
                    #hora_calendar = datetime.datetime.strptime(hora,'%H:%M')
                    #print(hora_actual)
                    ##hora_calendar_int = int(hora_calendar.strftime('%H%M'))
                
                    ##fechahora_ini = int(hora3.strftime('%H%M'))
                 
                    #parsed_time = dt.datetime.strptime(hora2, "%H:%M").time()
                    ##hours3 = parsed_time.hour
                    #minutes3 =  parsed_time.minute
                  
                    #end_hours = add_hour_and_half(hora2)
          
                    #parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
                    #hours4 = parsed_time2.hour
                    #minutes4 =  parsed_time2.minute
                              
                    #hoy = dt.datetime.now()
                    #fechoy = int(hoy.strftime("%Y%m%d"))
                
                    #horahoy = dt.datetime.now()
                    #horahoy2 = int(horahoy.strftime("%H%M"))
                
                    #if fech1 >= fechoy:  #and fechahora_ini >= horahoy2:
                  
                      #if fech1 == fechoy and hora_actual_int < #hora_calendar_int:
                    
                        #calendar.update_event(servicios+". "+nombre, start_time, end_time, time_zone,attendees=result_email)

                  if valida:
                
                        uid1 = result['UID']


                        whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Resserva se realizo con exito para el dia: {fecha} a las: {hora} con el encargado: {encargado} para el servicio de : {servicios}")
                                    
                        uid = uid1
                        values = [(nombre,email,str(fecha),hora,servicios, precio,encargado, notas, uid, whatsapp,str(57)+telefono, whatsappweb)]
                        gs = GoogleSheet(credentials, document, sheet)

                        range = gs.write_data_by_uid(uid, values)
                        
                        st.success('Su solicitud ha sido actualizada de forrma exitosa')

                        send_email2(email, nombre, fecha, hora, servicios, precio, encargado,  notas='De acuerdo con su solicitud su reserva se reprogramo. Gracias por su atencion.')
                        send_email_emp(email, nombre, fecha, hora, servicios, precio, encargado, notas='De acuerdo con su solicitud su reserva se reprogramo. Gracias por su atencion.')          
                          
                      #else:
                      #  st.warning('La hora seleccionda es invalida para hoy')
                      #  #print('La hora seleccionda es invalida para hoy')
                      
                  else:
                      st.warning('No se enconro el UID valido ')
                      #print('La hora seleccionda es invalida para hoy')                     
        else:     
          st.warning('El cliente No tiene agenda o esta cancelada verifique la informacion.')
           
 except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")
        print(f'Error crítico en la aplicación: {str(e)}')

sys.excepthook = global_exception_handler