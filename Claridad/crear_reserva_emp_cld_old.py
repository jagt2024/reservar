import streamlit as st
from google_sheets_emp import GoogleSheet
#from google_calendar_emp import GoogleCalendar
#from validacion_reservas_dp import validar_existencia_reserva, formatear_fecha, formatear_hora
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
from datetime import datetime, timedelta
import datetime as dt
import re
import uuid
from time import ctime
import requests
import time
import sqlite3
from sqlite3 import Error
import os 
import sys
import logging
#import ntplib
#from ntplib import NTPClient
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import toml
import json

st.cache_data.clear()
st.cache_resource.clear()

# Constantes
MAX_RETRIES = 3
INITIAL_RETRY_DELAY = 1
#SPREADSHEET_ID = 'TU-ID-DE-SPREADSHEET'  # Reemplaza con tu ID
#SCOPE = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

# Configuración de caché
class Cache:
    def __init__(self, ttl_minutes=5):
        self.data = None
        self.last_fetch = None
        self.ttl = timedelta(minutes=ttl_minutes)

    def is_valid(self):
        if self.last_fetch is None or self.data is None:
            return False
        return datetime.now() - self.last_fetch < self.ttl

    def set_data(self, data):
        self.data = data
        self.last_fetch = datetime.now()

    def get_data(self):
        return self.data
# Inicializar caché en session state
if 'cache' not in st.session_state:
    st.session_state.cache = Cache()
        
def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))
  
logging.basicConfig(level=logging.DEBUG, filename='crear_reserva_emp_cld.log', filemode='w',
format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Cargar el archivo Excel una sola vez
datos_book = load_workbook("./archivos-cld/parametros_empresa.xlsx", read_only=False)

def api_call_handler(func):
  # Number of retries
  for i in range(0, 10):
    try:
      return func()
    except Exception as e:
      print(e)
      time.sleep(2 ** i)
  print("The program couldn't connect to the Google Spreadsheet API for 10 times. Give up and check it manually.")
  raise SystemError

def dataBook(hoja):
    
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
        _row = []
        for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
            _row.append(col[row].value)
        data.append(_row[0])
    return data

def dataBookServicio(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in ws1.iter_rows(min_row=2, min_col=1):
      resultado = [col.value for col in row]
      data.append(resultado[0:2])
      #print(f'data {data}')
    return data
  
def dataBookServicioId(hoja,servicio):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == servicio:
       serv = _row[0]
       data = _row[2]
       break
  return data

def dataBookZonaEnc(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[5]
       break
  return data

def dataBookProducto(hoja,producto):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == producto:
       data = _row[0]
       break
       #print(f'su correo es {_row[1]}')
  return data

def dataBookPrecio(hoja,producto):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == producto:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data
 
def dataBookEncEmail(hoja, encargado):
  ws1 = datos_book[hoja]
  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
        data.append(_row) 
    #print(f'El encargado es {_row[0]}, su correo es {_row[1]}')
    if _row[0] == encargado:
       data = _row[1]
       break
       #print(f'su correo es {_row[1]}')
  return data

def get_conductores_por_zona(zona):
    
    if not zona:
        return []
    
    # Mapeo de zonas a nombres de hojas
    mapping = {
        'Norte': 'encargado_norte',
        'Sur': 'encargado_sur',
        'Oriente': 'encargado_oriente',
        'Occidente': 'encargado_occidente',
        'Oficina': 'encargado_oficina'
    }
    
    try:
        hoja = mapping.get(zona)
        if hoja:
            encargado = dataBook(hoja)
            return [c for c in encargado if c != 'X' and c is not None]
        return []
    except Exception as e:
        st.error(f"Error al obtener conductores: {str(e)}")
        return []

def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=30)).time()
    return new_time.strftime("%H:%M")

def add_hour_and_half2(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")
  
def calcular_diferencia_tiempo_futuro(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Calcular la fecha y hora futura (1 hora y 30 minutos después)
        tiempo_futuro = fecha_hora + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos
        diferencia = tiempo_futuro - fecha_hora
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def calcular_diferencia_tiempo(fecha_hora_str):
    try:
        # Convertir el string de fecha y hora a objeto datetime
        fecha_hora_parametro = datetime.strptime(fecha_hora_str, "%Y-%m-%d %H:%M")
        
        # Obtener fecha y hora actual
        fecha_hora_actual = datetime.now()
        
        # Calcular la fecha y hora futura sumando 1:30h a la fecha del parámetro
        tiempo_futuro = fecha_hora_parametro + timedelta(hours=1, minutes=30)
        
        # Calcular la diferencia en minutos entre el tiempo futuro y la hora actual
        diferencia = tiempo_futuro - fecha_hora_actual
        minutos_totales = diferencia.total_seconds() / 60
        
        return int(minutos_totales)
    except ValueError:
        return "Error: Formato de fecha y hora incorrecto. Use el formato 'YYYY-MM-DD HH:MM'"

def generate_uid():
    return str(uuid.uuid4())
  
def sendMessage(numero, mensaje):
  url = 'http://localhost:3001/lead'
  
  data = {
    "message": mensaje,
    "phone": numero 
  }
  headers = {
    'Content-Type': 'application/json'
  }
  print(data)
  response = requests.post(url, json=data, headers=headers)
  time.sleep(2)
  return response

#class CrearReservaEmp:
  
#  class Model:
#    pageTitle = "***Generar Reserva***"
  
 # def view(self,model):
 #   st.title(model.pageTitle)

def create_connection():
    """Create a database connection to the SQLite database"""
    conn = None
    try:
        conn = sqlite3.connect('reservas_dp.db')
        return conn
    except Error as e:
        print(f"Error connecting to database: {e}")
        return None

def create_tables(conn):
    """Create the necessary tables in the database"""
    try:
        cursor = conn.cursor()
        
        # Create reservas table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS reservas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                email TEXT,
                fecha DATE NOT NULL,
                hora TEXT NOT NULL,
                servicio TEXT NOT NULL,
                precio TEXT,
                encargado TEXT NOT NULL,
                email_encargado TEXT,
                zona TEXT,
                producto TEXT,
                direccion TEXT,
                notas TEXT,
                uid TEXT UNIQUE,
                whatsapp BOOLEAN,
                telefono TEXT,
                whatsapp_web TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
    except Error as e:
        print(f"Error creating tables: {e}")

def insert_reserva(conn, reserva_data):
    
    sql = '''INSERT INTO reservas(
                nombre, email, fecha, hora, servicio, precio, 
                encargado, email_encargado, zona, producto, direccion, 
                notas, uid, whatsapp, telefono, whatsapp_web)
             VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, reserva_data)
        conn.commit()
        return cursor.lastrowid
    except Error as e:
        print(f"Error inserting reserva: {e}")
        return None

def check_existing_reserva(conn, nombre, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE nombre = ? AND fecha = ? AND hora = ?
            '''    
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (nombre, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing reserva: {e}")
        return False 

def check_existing_encargado(conn, encargado, fecha, hora):
    """Check if a reserva already exists for the given name, date and time"""
    sql = '''SELECT COUNT(*) FROM reservas 
            WHERE encargado = ? AND fecha = ? AND hora = ?
            '''
    try:
        cursor = conn.cursor()
        cursor.execute(sql, (encargado, fecha, hora))
        count = cursor.fetchone()[0]
        return count > 0
    except Error as e:
        print(f"Error checking existing encargado: {e}")
        return False 

def limpiar_campos_formulario():
    
    try:
        # Lista de campos a limpiar
         valores_default = {
            'nombre': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
            #'fecha',
            #'hora',
            #'servicio_selector'
         }
        
         # Actualizar el session state con los valores por defecto
         for campo, valor in valores_default.items():
            if campo in st.session_state:
                # Eliminar la entrada actual del session state
                del st.session_state[campo]
        
         # Forzar la recarga de la página para reiniciar los widgets
         st.rerun()
        
         return True
        
    except Exception as e:
        st.error(f"Error al limpiar los campos del formulario: {str(e)}")
        logging.error(f"Error en limpiar_campos_formulario: {str(e)}")
        return False

def inicializar_valores_default():
    
    valores_default = {
            'nombre': '',
            'email': '',
            'direccion': '',
            'telefono': '',
            'notas': '',
    }
    
    for campo, valor in valores_default.items():
        if campo not in st.session_state:
            st.session_state[campo] = valor

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheetsemp']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

@st.cache_data(ttl=300)  # Cache de 5 minutos
def add_new_client(creds, nombre):
    """Añade un nuevo cliente a la hoja de Google Sheets"""
  for intento in range(MAX_RETRIES):  
    try:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-cld')
        worksheet = sheet.worksheet('clientes')
        
        # Obtiene el último ID y añade 1
        last_row = len(worksheet.get_all_values())
        new_id = last_row  # Asumiendo que la primera fila es encabezado
        
        # Añade la nueva fila
        worksheet.append_row([new_id, nombre])
        
        st.success(f"Cliente '{nombre}' añadido exitosamente!")
        return True
    
     except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error al añadir el cliente: {str(e)}")
        return False

@st.cache_data(ttl=300)  # Cache de 5 minutos
def consultar_reserva(nombre, fecha, hora):
  for intento in range(MAX_RETRIES):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        
        workbook = gc.open('gestion-reservas-cld')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = api_call_handler(lambda:worksheet.get_all_records())

        # Verificar si hay registros antes de crear el DataFrame
        if not registros:
            return False  # No hay datos en la hoja
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)

        # Verificar si las columnas necesarias existen
        required_columns = ['NOMBRE', 'FECHA', 'HORA']
        if not all(col in df.columns for col in required_columns):
            st.warning("La hoja no contiene todas las columnas necesarias")
            return False
        
        try:

            # Realizar la búsqueda
            reserva = df[
                (df['NOMBRE'].str.lower() == nombre.lower()) &
                (df['FECHA'] == fecha) &
                (df['HORA'] == hora)
            ]

        except AttributeError:
            # En caso de que alguna columna no sea del tipo esperado
            st.warning("Error en el formato de los datos")
            return False
        
        #return not reserva.empty
        
        if not reserva.empty:
            # Si encuentra la reserva, devuelve True y los detalles
            #detalles_reserva = reserva.iloc[0].to_dict()
            return True #, detalles_reserva
        else:

            #st.warning("Solicitud de Cliente No Existe")
            return False #, None
    
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False

    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False

@st.cache_data(ttl=300)  # Cache de 5 minutos@st.cache_data(ttl=300)  # Cache de 5 minutos
def consultar_encargado(encargado, fecha, hora):
  for intento in range(MAX_RETRIES):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-cld')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = api_call_handler(lambda:worksheet.get_all_records())
        
        # Verificar si hay registros antes de crear el DataFrame
        if not registros:
            return False  # No hay datos en la hoja
        
        # Convertir a DataFrame solo si hay registros
        df = pd.DataFrame(registros)
        
        # Verificar si las columnas necesarias existen
        required_columns = ['ENCARGADO', 'FECHA', 'HORA']
        if not all(col in df.columns for col in required_columns):
            st.warning("La hoja no contiene todas las columnas necesarias")
            return False
        
        # Realizar la búsqueda asegurándose de que no haya valores nulos
        try:
            encargado_registro = df[
                (df['ENCARGADO'].fillna('').str.lower() == encargado.lower()) &
                (df['FECHA'].fillna('') == fecha) &
                (df['HORA'].fillna('') == hora)
            ]
        except AttributeError:
            # En caso de que alguna columna no sea del tipo esperado
            st.warning("Error en el formato de los datos")
            return False
        
        return not encargado_registro.empty
    
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return False
            
    except Exception as e:
        st.error(f"Error al consultar encargado: {str(e)}")
        return False

def consultar_otros(nombre, fecha, hora):
    try:
        # Cargar credenciales
        creds = load_credentials_from_toml()
        if not creds:
            st.error("Error al cargar las credenciales")
            return False, None

        # Configurar el alcance y autenticación
        scope = ['https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive']
        
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        gc = gspread.authorize(credentials)
        
        # Abrir el archivo y la hoja específica
        workbook = gc.open('gestion-reservas-cld')
        worksheet = workbook.worksheet('reservas')
        
        # Obtener todos los registros
        registros = worksheet.get_all_records()
        
        # Convertir a DataFrame para facilitar la búsqueda
        df = pd.DataFrame(registros)
        
        # Realizar la búsqueda
        reserva = df[
            (df['NOMBRE'].str.lower() == nombre.lower()) &
            (df['FECHA'] == fecha) &
            (df['HORA'] == hora)
        ]
        
        # Verificar si se encontró la reserva
        if reserva.empty:
            return False, None
            
        # Extraer los campos solicitados
        datos_reserva = {
            'NOMBRE': reserva['NOMBRE'].iloc[0],
            'ENCARGADO': reserva['ENCARGADO'].iloc[0],
            'ZONA': reserva['ZONA'].iloc[0],
            'FECHA': reserva['FECHA'].iloc[0],
            'HORA': reserva['HORA'].iloc[0]
        }
        
        return True, datos_reserva
        
    except Exception as e:
        st.error(f"Error al consultar la reserva: {str(e)}")
        return False,(f"Error al consultar la reserva: {str(e)}")

def generate_whatsapp_link(phone_number, message):
    encoded_message = message.replace(' ', '%20')
    return f"https://wa.me/{phone_number}?text={encoded_message}"

def crea_reserva():
    
  try:
     
    st.title('Generar Reserva del Servicio')
    
    st.write("---")
       
    try:
                
        horas = dataBook("horario")
        #print(f'horas {horas}')

        #zonas = dataBook("zonas")
        #result_zonas = np.setdiff1d(zonas,'')
        #print(f'zona {zona} {result_zonas}')
      
        servicios = dataBook("servicio")
        result_serv = np.setdiff1d(servicios,'')

        producto = dataBook("precios")
        result_prod =  np.setdiff1d(producto,'')

        precio = dataBookPrecio("precios",producto)
        result_precio = np.setdiff1d(precio,'')

        #encargado = dataBook("encargado")
        #encargado = np.setdiff1d(nencargado,'')
        
        #zona_enc = dataBookZonaEnc("zonas", encargado)
        #result_zonaencc = np.setdiff1d(zona_enc,'')
        #print(f'servicio2 {servicio2}')
            
        #servicio2 = dataBookServicio2("servicio", 'Hacia el Aeropuerto')
        #result_serv2 = np.setdiff1d(servicio2,'')
        #print(f'servicio2 {servicio2}')
           
        servicioprecio = dataBookServicio("servicio")
        #print(f'servicio Precio {servicioprecio}')
        muestra = (f'servicio precio; {servicioprecio}')
        #print(f'(muestra= {muestra})')
        #result_servpre = np.setdiff1d(servicioprecio,'')
    
        #print(f'encargado {encargado}, {result_estil}') 
                            
        document='gestion-reservas-cld'
        sheet = 'reservas'
        credentials = st.secrets['sheetsemp']['credentials_sheet']
        time_zone = 'America/Bogota' #'GMT-05:00' # 'South America'
        
        #calendar = GoogleCalendar() #credentials, idcalendar
                       
        # Inicializar valores por defecto
        inicializar_valores_default()
        
        # Crear columnas para organizar la interfaz
        col1, col2 = st.columns([1, 1])
        
        with col1:
            # Cargar credenciales
            creds = load_credentials_from_toml('./.streamlit/secrets.toml')
        
            # Carga inicial de datos
            df = get_google_sheet_data(creds)
        
            if df is not None and not df.empty:
              # Lista de nombres existentes
              nombres_existentes = df['NOMBRE'].tolist() if 'NOMBRE' in df.columns else []
            
              # Crea el selectbox con opción para añadir nuevo
              selected_option = st.selectbox(
                  "Seleccione un cliente",
                  options=['-- Añadir Nuevo Cliente --'] + nombres_existentes
              )
            
              # Si se selecciona añadir nuevo
              if selected_option == '-- Añadir Nuevo Cliente --':
                 #with st.form("nuevo_cliente"):
                 nuevo_nombre = st.text_input("Ingrese el nombre del nuevo cliente")
                 #submitted = st.form_submit_button("Añadir Cliente", key='b2')
                   
                 #if submitted and nuevo_nombre:
                  if nuevo_nombre in nombres_existentes:
                     st.warning("Este cliente ya existe en la lista.")
                  else:
                    pass
                    # if add_new_client(creds, nuevo_nombre):
                        #st.rerun()  # Recarga la página para  actualizar la lista
              else:
                 st.write(f"Cliente seleccionado: {selected_option}")

            else:
                st.error("No se pudieron cargar los datos. Por favor, verifica la conexión.")
    
         
            #nombre = st.text_input('Nombre Solicitante*: ', placeholder='Nombre', #key='nombre',value=st.session_state.nombre
            #        ) # label_visibility='hidden')
                        
            # Lista de servicios disponibles
            #servicios = selectbox('Servicios*: ',result_serv)
            #productos = selectbox('Productos:', result_prod)

            #producto = dataBookProducto("precios", producto)
            # Selector de servicio
            servicio_seleccionado = st.selectbox(
                'Seleccione el servicio:',
                servicios,
                key='servicio_selector'
            )
            
            # Si es hacia el aeropuerto, mostrar selector de zona
            if servicio_seleccionado == 'Entrega' or servicio_seleccionado == 'Cambio' or servicio_seleccionado == 'Pedido':
                #precio_serv ='35.000'
                cantidad = st.number_input('Cantidad', key='cant')
                zonas = ['Norte', 'Sur', 'Oriente', 'Occidente', 'Oficina']
                zona_seleccionada = st.selectbox(
                    'Seleccione la zona:',
                    zonas,
                    key='zona_selector'
                )
                
                # Obtener conductores según la zona
                encargado = get_conductores_por_zona(zona_seleccionada)
                            
            elif servicio_seleccionado == 'Consulta':
                #precio_serv ='30.000'
                              
                # Para otros servicios, mostrar lista general de conductores
                encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
                
            hora = st.selectbox('Hora Servicio: ', horas, key="hora_new")
            direccion = st.text_input('Direccion Ubicacion solicitante :', placeholder='Direccion', key='direccion',value=st.session_state.direccion)
            
            notas = st.text_area('Nota de Consulta u Observacion(Opcional)', key='notas'
                                 ,value=st.session_state.notas)
        with col2:

            email  = st.text_input('Email Solicitante:', placeholder='Email', key='email',value=st.session_state.email
            )
            # Selector de producto
            producto_seleccionado = st.selectbox(
                'Seleccione el producto:',
                producto,
                key='producto_selector'
            )
            fecha  = st.date_input('Fecha Servicio*: ')             
                        
            # Mostrar selector de conductor si hay conductores disponibles
            if encargado:
                conductor_seleccionado = st.selectbox(
                    'Encargado Entrega:',
                    encargado,
                    key='conductor_selector'
                )
            
            existe_db2 = consultar_encargado(conductor_seleccionado, str(fecha), hora)

            if existe_db2:
               resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')
               #print(f'fecha y hora {fecha} {hora}')
               if resultado > 0 and resultado <= 90:
                  st.warning("Encargado se encuetra atendiedo un servicio")
               elif resultado >= 60:
                  st.warning("Encargado ya tiene agenda para esa fecha y hora")
               elif resultado <= -170:
                  st.warning("No pude agendarse con una fecha y/o  hora vencida")
               else:
                  st.success("La Hora de solicitud está disponible")
            else:
             # print(f'fecha y hora reserva {fecha} {hora}, fecha_hora_actual {datetime.now()}')
              resultado = calcular_diferencia_tiempo(f'{fecha} {hora}')

              #print(f"Resultado para la Hora actual : {resultado}")

              #print(f'resultado {resultado}')
              if resultado <= -170:
                st.warning("No pude agendarse con una fecha y/u  hora vencida")
              else:
                st.success("La Hora de solicitud está disponible")

            whatsapp = st.checkbox('Envio a WhatsApp Si/No (Opcional)')
            telefono = st.text_input('Nro. Telefono', key='telefono',value=st.session_state.telefono)

            precio = dataBookPrecio("precios", producto_seleccionado)
            # Mostrar resumen de la selección
            st.write("---")
            st.write("### Resumen de Solicitud:")
            
            info = {
                    "🚗 Encargado entrega": conductor_seleccionado,   "🎯 Servicio": servicio_seleccionado, "Producto": producto_seleccionado, "Cantidad": cantidad,"Precio Unitario": precio, "Fecha": fecha, "Hora":  hora
                }
            
            if servicio_seleccionado == 'Entrega' or servicio_seleccionado == 'Cambio' or servicio_seleccionado == 'Pedido':

                info["📍 Zona"] = zona_seleccionada
                
                for key, value in info.items():
                    st.write(f"{key}: **{value}**")

            else:
               encargado = [c for c in dataBook("encargado") if c != 'X' and c is not None]
               
               zona_enc = dataBookZonaEnc("encargado", conductor_seleccionado)
               
               info["📍 Zona"] = zona_enc
                
               for key, value in info.items():
                  st.write(f"{key}: **{value}**")

               #st.warning("No hay conductores disponibles para la selección actual.")
    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            

    except Exception as e:
       st.error(f"Error en la aplicación: {str(e)}")
       st.error("Por favor, verifica que el archivo Excel y las hojas existan.")

    st.write("---")
    
    with st.form(key='myform0',clear_on_submit=True):
    
     enviar = st.form_submit_button(" Reservar ", type="primary")

     #Backend
     if enviar:
        with st.spinner('Cargando...'):
         if not nombre or not servicio_seleccionado or not encargado or not email or not direccion:
            st.warning('Se Require completar los campos con * son obligatorios')
        
         elif not validate_email(email):
            st.warning('El email no es valido')
         elif whatsapp == True and not telefono:
            st.warning('Se Require el numero del Celular')
        
         else:
            # Create database connection
            conn = create_connection()
            if conn is None:
                st.error("Error: No se pudo conectar a la base de datos")
                return
                
            # Create tables if they don't exist
            #create_tables(conn)
                         
            # Check if reservation already exists in database
            existe_db = consultar_reserva(nombre, str(fecha), hora)

            if existe_db:
               existe = True
               st.warning("Ciente Ya tiene agenda para esa fecha y hora")
            else:
               #gs = GoogleSheet(credentials, document, sheet)
               existe = False
              
            if existe == False:
                
                precio = dataBookPrecio("precios", producto_seleccionado)
                result_precio = np.setdiff1d(precio,'')
                #print(f'Precio = {precio} result_precio = {result_precio}')
               
                emailencargado = dataBookEncEmail("encargado",conductor_seleccionado)
                #result_email = np.setdiff1d(emailencargado,'') 
                #print(f'Emailencargado = {emailencargado}, result_email ={result_email}')
      
                #st.text(muestra)
                       
                whatsappweb = (f"web.whatsapp.com/send?phone=&text= Sr(a). {nombre} La Resserva se realizo con exito para el dia: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para el servicio de : {servicio_seleccionado} con el producto: {producto_seleccionado}")
                 
                boton = '=ArrayFormula(SI(M3=VERDADERO;HIPERVINCULO(O3;"Enviar");"No Enviar"))'
                 
                if servicio_seleccionado != 'Consulta':
                    
                    uid = generate_uid()
                    values = [(nombre,email,str(fecha),hora, servicio_seleccionado, precio, conductor_seleccionado, str(emailencargado), zona_seleccionada, producto_seleccionado, cantidad,direccion, notas, uid, whatsapp,str(57)+telefono, whatsappweb, boton)]
                  
                    try:
                        #reserva_data = (
                        #nombre, email, fecha, hora, servicio_seleccionado,precio_serv, #conductor_seleccionado, str(emailencargado), zona_seleccionada, #direccion, notas, uid, whatsapp, str(57)+telefono, whatsappweb
                        #)
                     
                        #insert_reserva(conn, reserva_data)
                        
                        gs = GoogleSheet(credentials, document, sheet)
          
                        range = gs.get_last_row_range()
                        gs.write_data(range,values)

                        send_email2(email, nombre, fecha, hora, servicio_seleccionado, producto_seleccionado,precio*cantidad, conductor_seleccionado,  notas)
                
                        send_email_emp(email, nombre, fecha, hora, servicio_seleccionado, producto_seleccionado,precio*cantidad, conductor_seleccionado, notas, str(emailencargado))
                                             
                        st.success('Su solicitud ha sido reservada de forrma exitosa, la confirmacion fue enviada al correo')
                        
                        if whatsapp == True:
                           contact = str(57)+telefono
                           message = f'Cordial saludo: Sr(a): {nombre} La Reserva se creo con exito para el dia: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para realizar el servcio: {servicio_seleccionado}") para el producto:{producto_seleccionado}. Cordialmente aplicacion de Reservas y Agendamiento.'

                           phone_number = contact
                           mensaje = message 
                           whatsapp_link = generate_whatsapp_link(phone_number, mensaje)
                           st.markdown(f"Click si desea Enviar a su Whatsapp {whatsapp_link}")
                           time.sleep(10)
                        
                    except HttpError as error:
                        if error.resp.status == 429:  # Error de cuota excedida
                            if intento < MAX_RETRIES - 1:
                                delay = INITIAL_RETRY_DELAY * (2 ** intento)
                                st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                                time.sleep(delay)
                                continue
                            else:
                                st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
                        else:
                            st.error(f"Error de la API: {str(error)}")

                    except Exception as e:
                        st.error(f"Error al guardar en la base de datos: {str(e)}")
                    #finally:
                    #    conn.close()
                
                else:
                    
                    uid = generate_uid()
                    
                    values = [(nombre,email,str(fecha),hora, servicio_seleccionado, precio, conductor_seleccionado, str(emailencargado), zona_seleccionada, producto_seleccionado, cantidad, direccion, notas, uid, whatsapp,str(57)+telefono, whatsappweb, boton)]
                  
                    try:
                        ##reserva_data = (
                        #nombre, email, fecha, hora, servicio_seleccionado,precio_serv,
                        #conductor_seleccionado, str(emailencargado), str(zona_enc), direccion, notas, uid, whatsapp, str(57)+telefono, whatsappweb
                        #)
                     
                        #insert_reserva(conn, reserva_data)
                    
                        gs = GoogleSheet(credentials, document, sheet)
          
                        range = gs.get_last_row_range()
                        gs.write_data(range,values)
                
                        send_email2(email, nombre, fecha, hora, servicio_seleccionado, producto_seleccionado, precio*cantidad, conductor_seleccionado,  notas)
                
                        send_email_emp(email, nombre, fecha, hora, servicio_seleccionado, producto_seleccionado,precio*cantidad, conductor_seleccionado, notas, str(emailencargado)) 
                     
                        st.success('Su solicitud ha sido reservada de forrma exitosa, la confirmacion fue enviada al correo')

                        if whatsapp == True:
                           contact = str(57)+telefono
                           message = f'Cordial saludo: Sr(a): {nombre} La Reserva se creo con exito para el dia: {fecha} a las: {hora} con el encargado: {conductor_seleccionado} para realizar el servcio: {servicio_seleccionado}") para el producto: {producto_seleccionado}. Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                           phone_number = contact
                           mensaje = message 
                           whatsapp_link = generate_whatsapp_link(phone_number, mensaje)
                           st.markdown(f"Click si desea Enviar a su Whatsapp {whatsapp_link}")
                           time.sleep(10)

                    except HttpError as error:
                        if error.resp.status == 429:  # Error de cuota excedida
                            if intento < MAX_RETRIES - 1:
                                delay = INITIAL_RETRY_DELAY * (2 ** intento)
                                st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                                time.sleep(delay)
                                continue
                            else:
                                st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
                        else:
                            st.error(f"Error de la API: {str(error)}")
                            
                    except Exception as e:
                        st.error(f"Error al guardar en la base de datos: {str(e)}")
                    #finally:
                    #    conn.close()

                if limpiar_campos_formulario():
                   st.success('Campos limpiaddos exitosamente')                 
                
                  #calendar.create_event(servicios+". "+nombre, 
                  #start_time, end_time, time_zone, attendees=result_email)   

                  #if whatsapp == True:
                  #  contact = str(57)+telefono
                  #  message = f'Cordial saludo: Sr(a): Proceso {nombre} La Agenda se creo con exito para el dia: {fecha} a las: {hora} con el abogado encargado: {encargado} para el servicio de : {servicios} para realizar {acciones}"). Cordialmente aplicacion de Reservas y Agendamiento.'
                                          
                  #  sendMessage(contact, message)
                  #  sendMessage(str(57)+str(telefonoencargado), message)

  except Exception as e:
        logging.error(f"Error crítico en la aplicación: {str(e)}")
        st.error("Error crítico en la aplicación. Por favor, contacte al administrador.")

sys.excepthook = global_exception_handler

#if __name__ == "__main__":
#   crea_reserva()


##Error al cargar los datos: APIError: [429]: Quota exceeded for quota metric 'Read #requests' and limit 'Read requests per minute per user' of service 'sheets.googleapis.#com' for consumer 'project_number:719348869159'.