import streamlit as st
from streamlit_option_menu import option_menu
import numpy as np
from inicio_emp_cld import InicioEmp
from crear_reserva_emp_cld import crea_reserva
from modificar_reserva_emp_cld import modificar_reserva
from eliminar_reserva_emp_cld import eliminar_reserva
from catalogo_claridad import catalogo
from carga_archivo_clientes import carga
#from servicios_emp_dp import ServiciosEmp
#from informacion_distrito_privado import info_dp
#from generar_excel_emp_cld import GenerarExcelEmp
from consultar_pagos import consulta_pagos
from generador_qr import codigoqr
from consultar_reservas_cld  import consulta_reserva
#from consulta_st_excel import ConsultarAgenda
#from descargar_agenda_emp import download_and_process_data
#from authentication_users import authenticate_user
#from user_management import user_management_system
#from buscar_info import streamlit_app
from facturacion_servicios_emp import generar_factura
from consultar_facturas import consulta_facturas
from backup_google_drive_cld import backup
#from estadisticas_reservas_emp import reservas
#from estadisticas_facturacion_emp import factura
#from whatsapp_sender_st import whatsapp_sender
from ticket_support_app import soporte
from interface_pago_reserva import pago
#from parametros_empresa import parametros
#from localizador_gps import show_gps_tracker
#from ingresos_gastos import control
#from control_bancario import bancos
#from mobile_gps_tracker_ip7 import run_gps_tracker
import datetime as dt
from openpyxl import load_workbook
import os
from datetime import date
import datetime
import time
import pytz
import toml
from PIL import Image
import sys
import logging

# Configurar la página y ocultar la opción "Manage app" del menú
st.set_page_config(
    page_title="Servicio Actividad",
    page_icon="./assets-cld/brillol.png",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': None,  #'https://www.ejemplo.com/ayuda',
        'Report a bug': None,  #'https://www.ejemplo.com/reportar-bug',
        'About': None  #'# Esta es mi increíble aplicación Streamlit!'
        # La opción 'Manage app' no está incluida, por lo que no aparecerá
    }
)

# Ocultar elementos de la interfaz de Streamlit usando CSS personalizado
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}  /* Oculta el menú hamburguesa */
            footer {visibility: hidden;}  /* Oculta el footer "Made with Streamlit" */
            header {visibility: hidden;}  /* Oculta la cabecera */
            .stDeployButton {display:none;}  /* Oculta el botón de deploy */
            .css-1rs6os {visibility: hidden;}  /* Oculta el menú de configuración */
            .css-14xtw13 {visibility: hidden;}  /* Para algunas versiones de Streamlit */
            .css-1avcm0n {visibility: hidden;}  /* Para algunas versiones de Streamlit (menú hamburguesa) */
            
            /* En algunas versiones más recientes se usan diferentes clases CSS */
            /* Puedes identificar las clases específicas usando inspeccionar elemento en tu navegador */
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

st.cache_data.clear()
st.cache_resource.clear()

def global_exception_handler(exc_type, exc_value, exc_traceback):
    st.error(f"Error no manejado: {exc_type.__name__}: {exc_value}")
    logging.error("Error no manejado", exc_info=(exc_type, exc_value, exc_traceback))

logging.basicConfig(level=logging.DEBUG, filename='main_emp_cld.log', filemode='w', format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# En diferentes partes de tu código:
#logging.debug('Entrando en función X')


def clear_session_state():
    for key in list(st.session_state.keys()):
        del st.session_state[key]

# Cargar configuraciones desde config.toml
with open("./.streamlit/config.toml", "r") as f:
    config = toml.load(f)

os.environ["REQUESTS_CONNECT_TIMEOUT"] = "5"
os.environ["REQUESTS_READ_TIMEOUT"] = "5"

logo = Image.open("./assets-cld/brillol.png")  

datos_book_emp = load_workbook("./archivos-cld/parametros_empresa.xlsx", read_only=False)

def cargar_configuracion():
    try:
        config = toml.load("./.streamlit/config.toml")
        return config["seguridad"]["clave_protegida"]
    except FileNotFoundError:
        st.error("Archivo de configuración no encontrado.")
        return None
    except KeyError:
        st.error("Clave no encontrada en el archivo de configuración.")
        return None

def limpiar_pantalla():
    for key in list(st.session_state.keys()):
        if key != 'menu_selection':
          del st.session_state[key]
        
def dataBook_emp(hoja):
    ws1 = datos_book_emp[hoja]
    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

fecha_hasta = int('20260630')
#print(f'fecha hasta: {feha_hasta}'

fecha = dt.datetime.now()
fecha_hoy = int(fecha.strftime("%Y%m%d"))
#print(f'fecha_hoy: {fecha_hoy}')

empresa = dataBook_emp("sw")
result_emp = np.setdiff1d(empresa,'')
#print(f'sw_empresa {result_emp}')

sw_empresa = result_emp
  
page_title = 'Servicio Actividad' 
page_icon= "./assets-cld/brillol.png" 
title="Solicitudes"
layout = 'centered'

#st.set_page_config(page_title=page_title, page_icon=page_icon,layout=layout)

class Model:
  
  menuTitle = "Programe y Solicite sus Productos"
  option1 = 'Inicio'
  #option10 = 'Descargar Agenda'
  option9  = 'Consultar Solicitudes'
  #option22 = 'Consulta Pagos'
  #option5  = 'Nuestros Servicios'
  #option6  = 'Mas Informacion'
  #option7  = 'Generar Archivos'
  option8  = 'Generar Codigo QR'
  #option11 = 'Buscar Informacion'
  #option15 = 'Eviar Whatsapp'
  option25 = 'Consultar Facturas'
  option12 = 'Facturacion'
  option22 = 'Cargar Pagos'
  option23 = 'Cosultar Pagos'
  #option19 = 'Control Ingresos y Gastos'
  #option21 = 'Control Bancario'
  option18 = 'Copia de Seguridad Solicitudes'
  option24 = 'Carga Archivo Nuevos Clientes'
  #option13 = 'Estadisticas de Resservas'
  #option14 = 'Estadisticas de Facturacion'
  #option16 = 'Soporte - PQRS'
  #option17 = 'Geolocalizacion'
  #option20 = 'Parametrizacion'
        
  #def __init__(self):
  #  self.apps=[]
    
  def add_app(self,title, function):
      self.apps.append({
         "title":title,
         "function":function
       })
      
  def css_load(css_file):
        
    try:
      with open(css_file, "r") as file:
        st.markdown(f"<style>{file.read()}</style>",unsafe_allow_html=True)
    except IOError  as e:
      if "not readable" in str(e):
        print("Error no readable check mode")
      else:
        print(f"IOError {e}")
      
  css_load(r"style/main.css")

if fecha_hasta < fecha_hoy:
  
   sw_empresa == ['FALSO'] or sw_empresa == ['False']
      
   st.warning('Ha caducado el tiempo autorizado para su uso favor comuniquese con el administrador')

else:  
  
  #if user_management_system():
      
  def view(model):
      try:
  
        with st.sidebar:
    
          app = option_menu(model.menuTitle,
                         [model.option1, model.option9,model.option8, model.option25, model.option12, model.option22, model.option23, model.option18, model.option24                                                #model.option10,model.option22,model.option6,model.option7,model.option8, model.option11, model.option15, model.option12, model.option18, model.option13, model.option14, model.option19, model.option16, model.option17, model.option20, model.option21],
                         ],
                         
                         icons=['bi bi-app-indicator',
                                'bi bi-calendar2-date', 
                                'bi bi-calendar2-date',
                                'bi bi-calendar2-date',
                                'bi bi-award',
                                'bi bi-clipboard-minus-fill'
                               ],
                         default_index=0,
                         styles= {
                           "container":{ "reservas": "5!important",
                           "background-color":'#acbee9'},
                           "icon":{"color":"white","font-size":"23px"},
                           "nav-lik":{"color":"white","font-size":"20px","text-aling":"left", "margin":"0px"},
                           "nav-lik-selected":{"backgroud-color":"#02ab21"},})
                       #orientation='horizontal')
        st.markdown(f"""
        <style>
            @import url('https://fonts.googleapis.com/css2?family={config['fonts']['clock_font']}:wght@400;500&display=swap');
            .clock {{
              font-family: '{config['fonts']['clock_font']}', sans-serif;
              font-size: {config['font_sizes']['clock_font_size']};
              color: {config['colors']['clock_color']};
              text-shadow: 0 0 10px rgba(255,255,0,0.7);
              margin-top: {config['layout']['top_margin']};
              margin-bottom: {config['layout']['bottom_margin']};
            }}
            .calendar {{
              font-family: '{config['fonts']['calendar_font']}', sans-serif;
              font-size: {config['font_sizes']['calendar_font_size']};
              color: {config['colors']['calendar_color']};
              margin-top: {config['layout']['top_margin']};
              margin-bottom: {config['layout']['bottom_margin']};
            }}
        </style>
        """, unsafe_allow_html=True)

        # Crear columnas para el diseño
        col1, col2, col3, col4 = st.columns(4)

        # Columna 1: Reloj Digital
        with col1:
          #st.header("Reloj Digital")
          clock_placeholder = st.empty()
                     
          #st.markdown('''<iframe src="https://www.msn.com/es-co/el-tiempo/pronostico/in-Ricaurte,Cundinamarca?loc=eyJsIjoiUmljYXVydGUiLCJyIjoiQ3VuZGluYW1hcmNhIiwicjIiOiJSaWNhdXJ0ZSIsImMiOiJDb2xvbWJpYSIsImkiOiJDTyIsImciOiJlcy1jbyIsIngiOi03NC43NzE4LCJ5Ijo0LjI3Nzd9&weadegreetype=C&ocid=msedgdhp&cvid=CCD5BAAE25564F08A5B252DCA0EEF333&content=TeaserTempRecord_wxnwtsrec" width="300" height="250" style="border:0;" allowfullscreen="" loading="lazy" referrerpolicy="no-referrer-when-downgrade"></iframe>''',unsafe_allow_html=True)
        
        with col2:
          st.image(logo, width=250)  # Ajusta el ancho según sea necesario
        
        # Columna 2: Calendario
        with col3:
          #st.header("Calendario")
          #calendar_placeholder = st.empty()
          pass
        with col4:
          calendar_placeholder = st.empty()
          with st.form(key='myform4',clear_on_submit=True):
            limpiar = st.form_submit_button("Limpiar Pantalla")
            if limpiar:
              #st.submit_button("Limpiar Opcion")
              clear_session_state()
              st.rerun()

        def update_clock_and_calendar():
           while True:
              now = datetime.datetime.now(pytz.timezone('America/Bogota'))
              clock_placeholder.markdown(f'<p class="clock">Hora: {now.strftime("%H:%M:%S")}</p>', unsafe_allow_html=True)

              today = date.today()
              meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
              mes_es = meses_es[today.month - 1]
              dias_es = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
              dia_es = dias_es[today.weekday()]
        
              calendar_placeholder.markdown(f'<p class="calendar">Día: {dia_es.capitalize()} {today.day} de {mes_es} de {today.year}<br></p>', unsafe_allow_html=True)
        
              time.sleep(1)
              #st.rerun()  
       
        #st.markdown("""
        #<style>
        #  /* Estilos para pantallas grandes */
        #  @media (min-width: 768px) {
        #  .stTabs [data-baseweb="tab-list"] {
        #    gap: 24px;
        #      padding: 0px 10px;
        #  }

        #  .stTabs [data-baseweb="tab"] {
        #      height: 50px;
        #      white-space: pre-wrap;
        #      background-color: #f0f2f6;
        #      border-radius: 5px;
        #      padding: 10px 20px;
        #      font-size: 16px;
        #      font-weight: 600;
        #      color: #31333F;
        #  }

        #  .stTabs [data-baseweb="tab"]:hover {
        #      background-color: #e0e2e6;
        #      color: #1f77b4;
        #  }

        #  .stTabs [data-baseweb="tab"][aria-selected="true"] {
        #      background-color: #1f77b4;
        #      color: white;
        #  }

        #  .stTabs [data-baseweb="tab"] div {
        #      font-size: 20px;
        #    }
         # }

         # /* Estilos para pantallas pequeñas */
         # @media (max-width: 767px) {
         # .stTabs [data-baseweb="tab-list"] {
         #   gap: 3px;
         #   padding: 0px 5px;
         # }

         # .stTabs [data-baseweb="tab"] {
         #     height: 40px;
         #     white-space: pre-wrap;
         #     background-color: #f0f2f6;
         #     border-radius: 5px;
         #     padding: 6px 10px;
         #     font-size: 10px;
         #     font-weight: 300;
         #     color: #31333F;
         # }

         # .stTabs [data-baseweb="tab"]:hover {
         #     background-color: #e0e2e6;
         #     color: #1f77b4;
         # }

         # .stTabs [data-baseweb="tab"][aria-selected="true"] {
         #     background-color: #1f77b4;
         #     color: white;
         # }

         # .stTabs [data-baseweb="tab"] div {
         #     font-size: 10px;
         #   }
        #  }
       # </style>
       # """, unsafe_allow_html=True)

        
        # Crear los tabs con los estilos personalizados
               
        tabs = st.tabs(["Inicio", "Crear Solicitud", "Modificar Solicitud","Eliminar Solicitud", "Catalogo Productos", "Soporte - PQRS"
        #, "Registrar Pago", "Soporte - PQRS"])
        ])
            
        #with tabs[0]:
        #  InicioEmp()
          #InicioEmp().view(InicioEmp.Model())
              
        with tabs[1]:
          crea_reserva()
    
        with tabs[2]:
          modificar_reserva()
    
        with tabs[3]:
          eliminar_reserva()
        
        with tabs[4]:
          catalogo()
        
        with tabs[5]:
          soporte()

        with st.sidebar:
          st.markdown("---")
          st.text("Version: 0.0.1")
          st.text("Ano: 2025")
          st.text("Autor: JAGT")
          st.markdown("---")
    
        if sw_empresa == ['VERDADERO'] or sw_empresa == ['True']:
          try:
                   
            clave_correcta = cargar_configuracion()
            if clave_correcta is None:
              st.error("No se pudo cargar la configuración. La opción protegida no estará disponible.")
              return 
      
            if app == model.option1:
              InicioEmp().view(InicioEmp.Model())
            #if app == model.option5:
            #  ServiciosEmp().view(ServiciosEmp.Model())


            #if app == model.option6:
            #  info_dp()
            #if app == model.option7:
            #  GenerarExcelEmp().view(GenerarExcelEmp.Model())
            if app == model.option22:
              pago()
            if app == model.option23:
              consulta_pagos()
            if app == model.option8:
              codigoqr()
            if app == model.option9:
              consulta_reserva()
            #if app == model.option10:
                
              #if 'prev_selection' not in st.session_state or st.session_state.prev_selection != #model.option10:
              #limpiar_pantalla()
              #st.session_state.prev_selection = model.option10
           
              #clave_ingresada = st.text_input("Ingresa la clave para acceder a la Opción Protegida:", type="password")
           
              #if clave_ingresada == clave_correcta:
                 
              #    st.session_state.prev_selection = model.option10
              #    st.success("Clave correcta. Acceso concedido.")
             # st.write(download_and_process_data('./.streamlit/secrets.toml'))
              
                #del st.session_state.clave_ingresada
                #limpiar_pantalla()
                #st.session_state.clave_ingresada = ""
                #st.rerun()
           
              #elif clave_ingresada:
              #  st.error("Clave incorrecta. Acceso denegado.")         

            #if app == model.option11:
            #  streamlit_app()
            if app == model.option25:
              consulta_facturas()

            if app == model.option12:
              generar_factura()

            if app == model.option18:
              backup()

            if app == model.option24:
              carga()
                            
            #if app == model.option13:
           
              #if authenticate_user(): 
            
             # reservas()
                            
            #if app == model.option14:
           
              #if authenticate_user():
          
            #  factura()

            #if app == model.option19:
           
              #if authenticate_user():
          
            #  control()
            
            #if app == model.option21:
            #  bancos()
              
            #if app == model.option15:
           
              #if authenticate_user(): 
            
            #  whatsapp_sender()
              
            #if app == model.option16:
            #  soporte()
               
            #if app == model.option17:
            #  show_gps_tracker()

            #if app == model.option20:
            #  parametros()
          
          except Exception as e:
            st.error(f"Ocurrió un error: {e}")
                                      
      except SystemError as err:
        raise Exception(f'A ocurrido un error en main_emp.py: {err}')
      except Exception as e:
        st.error(f"Ocurrió un error en main_emp.py: {e}")
        
      update_clock_and_calendar()
      
sys.excepthook = global_exception_handler

view(Model())