import streamlit as st
from google_sheets_emp import GoogleSheet
import pandas as pd
from datetime import datetime, timedelta
import random
import qrcode
import numpy as np
import uuid
import toml
from PIL import Image
import io
import base64
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
#from user_management import logout 
import os
import sqlite3
import json
from openpyxl import load_workbook
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import logging

st.cache_data.clear()
st.cache_resource.clear()

# Constantes
MAX_RETRIES = 3
INITIAL_RETRY_DELAY = 1

# Configuración de caché
class Cache:
    def __init__(self, ttl_minutes=5):
        self.data = None
        self.last_fetch = None
        self.ttl = timedelta(minutes=ttl_minutes)

    def is_valid(self):
        if self.last_fetch is None or self.data is None:
            return False
        return datetime.now() - self.last_fetch < self.ttl

    def set_data(self, data):
        self.data = data
        self.last_fetch = datetime.now()

    def get_data(self):
        return self.data

# Inicializar caché en session state
if 'cache' not in st.session_state:
    st.session_state.cache = Cache()

# Configurar logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

datos_book = load_workbook("./archivos-cld/parametros_empresa.xlsx", read_only=False)

def load_credentials_from_toml():
    try:
        with open('./.streamlit/secrets.toml', 'r') as toml_file:
            config = toml.load(toml_file)
            creds = config['sheetsemp']['credentials_sheet']
            if isinstance(creds, str):
                creds = json.loads(creds)
            return creds
    except Exception as e:
        st.error(f"Error al cargar credenciales: {str(e)}")
        return None

def dataBook(hoja):
    ws1 = datos_book[hoja]
    data = []
    for row in range(1, ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(min_row=0, min_col=1, max_col=ws1.max_column):
        _row.append(col[row].value)
      data.append(_row[0])
      #print(f'data {data}')
    return data

# Función para cargar datos del emisor desde Excel
def cargar_datos_emisor():
    try:
        df = pd.read_excel('./archivos-cld/parametros_empresa.xlsx', sheet_name='emisor')
        emisor = df.iloc[0]
        return {
            'nombre': emisor['NOMBRE'],
            'nit': emisor['NIT'],
            'direccion': emisor['DIRECCION'],
            'ciudad': emisor['CIUDAD']
        }
    except Exception as e:
        st.error(f"Error al cargar datos del emisor: {str(e)}")
        return None

def get_clientes_from_sheets():
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        creds = st.secrets['sheetsemp']['credentials_sheet']
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-cld')
    
        # Get reservations data
        reservas_ws = sheet.worksheet('reservas')
        reservas_data = reservas_ws.get_all_records()
        df_reservas = pd.DataFrame(reservas_data)
    
        # Get payments data
        pagos_ws = sheet.worksheet('pagos')
        pagos_data = pagos_ws.get_all_records()
        df_pagos = pd.DataFrame(pagos_data)
    
        # Filter clients not in payments sheet
        clientes_sin_pago = df_reservas[~df_reservas.apply(lambda row: any((df_pagos['Nombre'] == row['NOMBRE']) & (df_pagos['Producto'] == row['PRODUCTO'])), axis=1)]
    
        return clientes_sin_pago[['NOMBRE', 'DIRECCION', 'EMAIL']].drop_duplicates()

    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return None

    except Exception as e:
        st.error(f"Error al cargar los datos: {str(e)}")
        return None

def get_data_from_sheets():
  for intento in range(MAX_RETRIES):
    try:
      with st.spinner(f'Cargando datos... (Intento {intento + 1}/{MAX_RETRIES})'):
        creds = st.secrets['sheetsemp']['credentials_sheet']
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        credentials = Credentials.from_service_account_info(creds, scopes=scope)
        client = gspread.authorize(credentials)
        sheet = client.open('gestion-reservas-cld')
    
        # Get both sheets' data
        reservas_ws = sheet.worksheet('reservas')
        reservas_data = reservas_ws.get_all_records()
        df_reservas = pd.DataFrame(reservas_data)
    
        pagos_ws = sheet.worksheet('pagos')
        pagos_data = pagos_ws.get_all_records()
        df_pagos = pd.DataFrame(pagos_data)
    
        # Filter out records that exist in payments
        return df_reservas[~df_reservas.apply(lambda row: any((df_pagos['Nombre'] == row['NOMBRE']) & (df_pagos['Producto'] == row['PRODUCTO'])), axis=1)]

    except HttpError as error:
            if error.resp.status == 429:  # Error de cuota excedida
                if intento < MAX_RETRIES - 1:
                    delay = INITIAL_RETRY_DELAY * (2 ** intento)
                    st.warning(f"Límite de cuota excedida. Esperando {delay} segundos...")
                    time.sleep(delay)
                    continue
                else:
                    st.error("Se excedió el límite de intentos. Por favor, intenta más tarde.")
            else:
                st.error(f"Error de la API: {str(error)}")
            return None

    except Exception as e:
        st.error(f"Error al cargar los datos: {str(e)}")
        return None

def cargar_datos_cliente(nombre_seleccionado, df_clientes):
    cliente = df_clientes[df_clientes['NOMBRE'] == nombre_seleccionado].iloc[0]
    return {
        'direccion': cliente['DIRECCION'],
        'email': cliente['EMAIL']
    }

def get_productos_cliente(nombre_cliente, df):
    productos_cliente = df[df['NOMBRE'] == nombre_cliente]
    # Aseguramos que los valores numéricos sean válidos
    productos_cliente['CANTIDAD'] 
    productos_cliente['PRECIO'] 
    return productos_cliente[['PRODUCTO', 'CANTIDAD', 'PRECIO']].drop_duplicates()

def generate_uid():
    return str(uuid.uuid4())

# Función para inicializar la base de datos
def init_db():
    conn = sqlite3.connect('facturas.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS facturas
                 (numero_factura TEXT PRIMARY KEY, 
                  fecha_factura TEXT,
                  emisor_nombre TEXT,
                  emisor_nit TEXT,
                  emisor_ciudad TEXT,
                  cliente_nombre TEXT,
                  cliente_nit TEXT,
                  cliente_direccion TEXT,
                  cliente_email TEXT,
                  servicios TEXT,
                  producto TEXT,
                  subtotal REAL,
                  iva_total_producto REAL,
                  total_producto REAL)''')
    conn.commit()
    conn.close()

# Función para guardar la factura en la base de datos
def guardar_factura_en_db(numero_factura, fecha_factura, emisor_nombre, emisor_nit, emisor_ciudad, cliente_nombre, cliente_nit, cliente_direccion, cliente_email, productos, producto, subtotal, iva_total_producto, total_producto):
    conn = sqlite3.connect('facturas.db')
    c = conn.cursor()
    c.execute('''INSERT OR REPLACE INTO facturas VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (numero_factura, fecha_factura, emisor_nombre, emisor_nit, emisor_ciudad,
         cliente_nombre, cliente_nit, cliente_direccion, producto, cliente_email,json.dumps(productos), subtotal, iva_total_producto, total_producto))

         
    conn.commit()
    conn.close()
    
def enviar_factura_por_email(email_cliente, pdf_buffer, numero_factura):
    smtp_server = "smtp.gmail.com"
    port = 465
    sender_email = st.secrets['emails']['smtp_user']
    password = st.secrets['emails']['smtp_password']

    logger.info(f"Iniciando envío de correo a {email_cliente}")
    
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = email_cliente
    message["Subject"] = f"Factura {numero_factura}"

    body = f"Adjunto encontrará la factura {numero_factura}. Gracias por su preferencia."
    message.attach(MIMEText(body, "plain"))

    pdf_attachment = MIMEApplication(pdf_buffer.getvalue(), _subtype="pdf")
    pdf_attachment.add_header('Content-Disposition', f'attachment; filename=factura_{numero_factura}.pdf')
    message.attach(pdf_attachment)

    try:
        logger.debug("Intentando conexión SMTP_SSL")
        with smtplib.SMTP_SSL(smtp_server, port) as server:
            logger.debug("Conexión SMTP_SSL exitosa")
            logger.debug("Intentando login")
            server.login(sender_email, password)
            logger.debug("Login exitoso")
            logger.debug("Enviando mensaje")
            server.send_message(message)
            logger.info("Mensaje enviado exitosamente")
        return True
    except Exception as e:
        logger.error(f"Error al enviar el correo: {str(e)}", exc_info=True)
        return False

def generar_numero_factura():
    if 'numero_factura' not in st.session_state:
        st.session_state.numero_factura = f"FACT-{random.randint(1000, 9999)}"
    return st.session_state.numero_factura

def limpiar_campos():
    st.session_state.numero_factura = generar_numero_factura()
    if 'nombre_cliente' in st.session_state:
        #del st.session_state.nombre_cliente
        st.session_state.nombre_cliente = ''
    if 'direccion_cliente' in st.session_state:
        #del st.session_state.direccion_cliente
        st.session_state.direccion_cliente=''
    if 'email_cliente' in st.session_state:
        #del st.session_state.email_cliente
        st.session_state.email_cliente=''
    if 'nit_cliente' in st.session_state:
        del st.session_state.nit_cliente
        st.session_state.nit_cliente = ''
    for i in range(10):
        if f"desc_{i}" in st.session_state:
            #del st.session_state[f"desc_{i}"]
            st.session_state[f"desc_{i}"] = ''
        if f"cant_{i}" in st.session_state:
            #del st.session_state[f"cant_{i}"]
            st.session_state[f"cant_{i}"] = 0
        if f"precio_{i}" in st.session_state:
            #del st.session_state[f"precio_{i}"]
            st.session_state[f"precio_{i}"]= 0

def calcular_iva(precio):
    return precio * 0.19  # 19% IVA en Colombia

def calcular_precio_sin_iva(precio_con_iva):
    return precio_con_iva / 1.19

def calcular_totales(productos):
    subtotal = sum(p['subtotal'] for p in productos)
    iva_total_producto = sum(p['iva'] for p in productos)
    total_producto = sum(p['total_producto'] for p in productos)
    return subtotal, iva_total_producto, total_producto

def generar_qr(datos):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(datos)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    return img_to_bytes(img)

def img_to_bytes(img):
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()

def generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, productos, subtotal, iva_total_producto, total_producto, logo_bytes, qr_bytes, emisor_nombre, emisor_nit, emisor_ciudad):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    
    if logo_bytes:
        logo_image = ReportLabImage(io.BytesIO(logo_bytes), width=2*inch, height=1*inch)
        elements.append(logo_image)

    elements.append(Paragraph(f"Emisor: {emisor_nombre}", styles['Normal']))
    elements.append(Paragraph(f"NIT Emisor: {emisor_nit}", styles['Normal']))
    elements.append(Paragraph(f"Ciudad: {emisor_ciudad}", styles['Normal']))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Factura Nº: {numero_factura}", styles['Heading1']))
    elements.append(Paragraph(f"Fecha: {fecha_factura}", styles['Normal']))
    elements.append(Paragraph(f"Cliente: {nombre_cliente}", styles['Normal']))
    elements.append(Paragraph(f"NIT/Cédula: {nit_cliente}", styles['Normal']))
    elements.append(Paragraph(f"Dirección: {direccion_cliente}", styles['Normal']))
    elements.append(Paragraph(f"Email: {email_cliente}", styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Descripción', 'Cantidad', 'Precio Unitario (sin IVA)', 'IVA', 'Subtotal (con IVA)', 'Iva Total']]
    for servicio in productos:
        data.append([
            servicio['descripcion'],
            str(servicio['cantidad']),
            f"${servicio['precio_unitario_sin_iva']:,.2f}",
            f"${servicio['iva']:,.2f}",
            f"${servicio['subtotal']:,.2f}",
            f"${servicio['iva_total_producto']:,.2f}"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)

    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Subtotal (sin IVA): ${subtotal:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"IVA Total: ${iva_total_producto:,.2f}", styles['Normal']))
    elements.append(Paragraph(f"Total: ${total_producto:,.2f}", styles['Normal']))

    qr_image = ReportLabImage(io.BytesIO(qr_bytes), width=2*inch, height=2*inch)
    elements.append(qr_image)

    doc.build(elements)
    buffer.seek(0)
    return buffer

def generar_factura():
    st.title("Generación de Factura")

    document='gestion-reservas-cld'
    sheet = 'facturacion'
    credentials = st.secrets['sheetsemp']['credentials_sheet']
    time_zone = 'America/Bogota' #'GMT-05:00' # 'South America'
        
    #init_db()

    # Cargar datos del emisor
    emisor_data = cargar_datos_emisor()
    if not emisor_data:
        st.error("No se pudieron cargar los datos del emisor. Por favor, verifique el archivo Excel.")
        return

    logo_path = "./assets-cld/brillol.png"
    logo_bytes = None
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_bytes = f.read()
        st.image(logo_bytes, width=200, caption="Logo de la Empresa")
    else:
        st.warning("No se encontró el logo en ./assets-cld/brillol.png")

    st.subheader("Información del Emisor")
    st.write(f"**Nombre del Emisor:** {emisor_data['nombre']}")
    st.write(f"**NIT del Emisor:** {emisor_data['nit']}")
    st.write(f"**Dirección del Emisor:** {emisor_data['direccion']}")
    st.write(f"**Ciudad del Emisor:** {emisor_data['ciudad']}")

    # Fetch all data from sheets
    df_completo = get_data_from_sheets()
    df_clientes = get_clientes_from_sheets()
    nombres_clientes = df_clientes['NOMBRE'].tolist()
    
    st.subheader("Información del Cliente")
    nombre_cliente = st.selectbox("Seleccione el Cliente", nombres_clientes, key="nombre_cliente")
    
    # Auto-fill client information when selected
    if nombre_cliente:
        datos_cliente = cargar_datos_cliente(nombre_cliente, df_clientes)
        direccion_cliente = st.text_input("Dirección del Cliente", value=datos_cliente['direccion'], key="direccion_cliente")
        email_cliente = st.text_input("Correo Electrónico del Cliente", value=datos_cliente['email'], key="email_cliente")
        nit_cliente = st.text_input("NIT/Cédula del Cliente", key="nit_cliente")

        st.subheader("Número de Factura")
        numero_factura = st.text_input("Número de Factura", value=generar_numero_factura(), key="numero_factura")
        
        # Get products associated with the selected client
        productos_cliente = get_productos_cliente(nombre_cliente, df_completo)
        
        st.subheader("Productos")
        productos = []
        for i, (index, producto) in enumerate(productos_cliente.iterrows()):
            with st.expander(f"Producto {i + 1}", expanded=i == 0):
                descripcion = st.text_input("Descripción del Producto", 
                              value=producto['PRODUCTO'], 
                              key=f"desc_{i}", 
                              disabled=True)
    
                try:
                    cantidad = int(producto['CANTIDAD']) if pd.notna(producto['CANTIDAD']) else 0
                except (ValueError, TypeError):
                    cantidad = 0
        
                cantidad = st.number_input("Cantidad", 
                             value=cantidad,
                             key=f"cant_{i}", 
                             disabled=True)
    
                try:
                    precio = producto['PRECIO']*1000 if pd.notna(producto['PRECIO']) else 0.0
                except (ValueError, TypeError):
                    precio = 0.0
        
                precio_unitario_con_iva = st.number_input("Precio Unitario (con IVA)", 
                                             value=precio,
                                             format="%.2f",
                                             key=f"precio_{i}", 
                                             disabled=True)
    
                if descripcion and cantidad > 0 and precio_unitario_con_iva > 0:
                    precio_unitario_sin_iva = round(calcular_precio_sin_iva(precio_unitario_con_iva), 2)
                    iva = round(calcular_iva(precio_unitario_sin_iva), 2)
                    subtotal = round(precio_unitario_sin_iva, 2)
                    iva_total_producto = round(iva, 2)
                    total_producto = round(precio_unitario_con_iva, 2)
                
                    productos.append({
                        "descripcion": descripcion,
                        "cantidad": cantidad,
                        "precio_unitario_sin_iva": round(precio_unitario_sin_iva, 2),
                        "iva": round(iva, 2),
                        "iva_total_producto": round(iva_total_producto, 2),
                        "subtotal": round(subtotal, 2),
                        "total_producto": round(total_producto, 2)
                    })

                    # Calcular totales fuera del loop
                    if productos:
                        st.write(f"**Subtotal (sin IVA):** ${subtotal:,.2f} COP")
                        st.write(f"**IVA Total:** ${iva_total_producto:,.2f} COP")
                        st.write(f"**Total:** ${total_producto:,.2f} COP")
                    else:
                        subtotal = iva_total_producto = total_producto = 0.0
    
            # Calcular totales fuera del loop
            if productos:
                subtotal = round(sum(p['subtotal'] for p in productos), 2)
                iva_total_producto = round(sum(p['iva_total_producto'] for p in productos), 2)
                total_producto = round(sum(p['total_producto'] for p in productos), 2)
            else:
                subtotal = iva_total_producto = total_producto = 0.0

        if st.button("Generar Factura", key="generar_factura"):
            fecha_factura = datetime.now().strftime('%Y-%m-%d')
            
            st.subheader("Factura Generada")
            
            if logo_bytes:
                st.image(logo_bytes, width=200)

            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**Emisor:** {emisor_data['nombre']}")
                st.write(f"**NIT Emisor:** {emisor_data['nit']}")
                st.write(f"**Dirección Emisor:** {emisor_data['direccion']}")
                st.write(f"**Ciudad Emisor:** {emisor_data['ciudad']}")
                st.write(f"**Número de Factura:** {numero_factura}")
                st.write(f"**Fecha:** {fecha_factura}")
            with col2:
                st.write(f"**Cliente:** {nombre_cliente}")
                st.write(f"**NIT/Cédula:** {nit_cliente}")
                st.write(f"**Dirección:** {direccion_cliente}")
                st.write(f"**Email:** {email_cliente}")
                
            df_productos = pd.DataFrame(productos)
            st.table(df_productos)

            st.write(f"**Subtotal (sin IVA):** ${subtotal:,.2f} COP")
            st.write(f"**IVA Total:** ${iva_total_producto:,.2f} COP")
            st.write(f"**Total:** ${total_producto:,.2f} COP")

            datos_qr = f"Factura: {numero_factura}\nFecha: {fecha_factura}\nEmisor: {emisor_data['nombre']}\nCliente: {nombre_cliente}\nTotal: ${total_producto:,.2f} COP"
            qr_bytes = generar_qr(datos_qr)
            
            st.image(qr_bytes, caption='Código QR de la Factura', width=300)
            
            try:
               
                pdf_buffer = generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, productos, subtotal, iva_total_producto, total_producto, logo_bytes, qr_bytes, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'])

                st.download_button(
                label="Descargar Factura como PDF",
                data=pdf_buffer,
                file_name=f"factura_{numero_factura}.pdf",
                mime="application/pdf",
                )

                df_productos = pd.DataFrame(productos)
                csv = df_productos.to_csv(index=False)
                st.download_button(
                label="Descargar Factura como CSV",
                data=csv,
                file_name=f"factura_{numero_factura}.csv",
                mime="text/csv",
                )

                st.download_button(
                label="Descargar Código QR",
                data=qr_bytes,
                file_name=f"codigo_qr_factura_{numero_factura}.png",
                mime="image/png",
                )

            except Exception as e:
                st.error(f"Error al generar el PDF o guardar en la base de datos: {str(e)}")

    else:
        st.warning("Por favor seleccione un cliente para ver sus productos asociados.")
    #else:
    #    st.warning("Por favor, agregue al menos un servicio para generar la factura.")

    #Opción para enviar la factura por correo electrónico
    if st.button("Enviar Factura por Correo Electrónico"):
       if email_cliente:
          fecha_factura = datetime.now().strftime('%Y-%m-%d')
          datos_qr = f"Factura: {numero_factura}\nFecha: {fecha_factura}\nEmisor: {emisor_data['nombre']}\nCliente: {nombre_cliente}\nTotal: ${total_producto:,.2f} COP"
          qr_bytes = generar_qr(datos_qr)
          with st.spinner("Enviando factura por correo electrónico..."):
            logger.info(f"Generando PDF para factura {numero_factura}")
            pdf_buffer = generar_pdf_factura(numero_factura, fecha_factura, nombre_cliente, nit_cliente, direccion_cliente, email_cliente, productos, subtotal, iva_total_producto, total_producto, logo_bytes, qr_bytes, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'])
            logger.info("PDF generado correctamente")
                    
            logger.info(f"Intentando enviar correo a {email_cliente}")
            if enviar_factura_por_email(email_cliente, pdf_buffer, numero_factura):
               st.success(f"Factura enviada exitosamente a {email_cliente}")
               logger.info(f"Factura enviada exitosamente a {email_cliente}")
            else:
               st.error("No se pudo enviar la factura por correo electrónico. Por favor, revise los logs para más detalles.")
               logger.error("Fallo en el envío de la factura por correo electrónico")
       else:
          st.warning("Por favor, ingrese el correo electrónico del cliente para enviar la factura.")
          logger.warning("Intento de envío de factura sin dirección de correo electrónico")
    
    if st.button("Confirmar y Guardar Factura"):

       fecha_factura = datetime.now().strftime('%Y-%m-%d')
       nombre = emisor_data['nombre']
       nit = emisor_data['nit']
       ciudad = emisor_data['ciudad']
       subtotal = (f"{subtotal:,.2f}")
       iva_total_producto = (f"{iva_total_producto:,.2f}")
       total_producto = (f"{total_producto:,.2f}")
       
       uid = generate_uid()
                    
       values = [(numero_factura, str(fecha_factura), nombre, str(nit), ciudad, nombre_cliente, str(nit_cliente), direccion_cliente, str(email_cliente), json.dumps(productos), subtotal, iva_total_producto, total_producto, uid)]
                
       gs = GoogleSheet(credentials, document, sheet)
          
       rango = gs.get_last_row_range()

       gs.write_data(rango,values)

       #guardar_factura_en_db(numero_factura, fecha_factura, emisor_data['nombre'], emisor_data['nit'], emisor_data['ciudad'], nombre_cliente, nit_cliente, direccion_cliente, email_cliente, productos, subtotal, iva_total, total)

       st.success("Factura guardada en la base de datos exitosamente.")
    
    # Limpiar campos después de generar y descargar la factura
    #st.button("Limpiar campos y generar nueva factura", on_click=limpiar_campos)

#if __name__ == "__main__":
#    generar_factura()