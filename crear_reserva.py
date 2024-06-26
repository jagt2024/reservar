import streamlit as st
from google_sheets import GoogleSheet
from google_calendar import GoogleCalendar
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
import datetime
import datetime as dt
import re
import uuid
from openpyxl import load_workbook
#import pywhatkit
#import pyautogui, webbrowser
#from time import sleep

datos_book = load_workbook("archivos/parametros.xlsx", read_only=False)

def dataBook(hoja):
    ws1 = datos_book[hoja]

    data = []
    for row in range(1,ws1.max_row):
      _row=[]
      for col in ws1.iter_cols(1,ws1.max_column):
        _row.append(col[row].value)
    
      data.append(_row)
      #print(f'data {data}')
    return data
  
def validate_email(email):
    pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
    if re.match(pattern, email):
      return True
    else:
      return False
  
def add_hour_and_half(time):
    parsed_time = dt.datetime.strptime(time, "%H:%M").time()
    new_time = (dt.datetime.combine(dt.date.today(), parsed_time) + dt.timedelta(hours=1, minutes=0)).time()
    return new_time.strftime("%H:%M")

def generate_uid():
    return str(uuid.uuid4())

class CrearReserva:
  
  class Model:
    pageTitle = "***Generar Reserva***"
  
  def view(self,model):
    st.title(model.pageTitle)

    with st.form(key='myform1',clear_on_submit=True):
  
      horas = dataBook("horario")
      
      servicio = dataBook("servicio")
      result_serv = np.setdiff1d(servicio,'') 
    
      encargado = dataBook("encargado")
      result_estil = np.setdiff1d(encargado,'X') 
    
      document='gestion-reservas'
      sheet = 'reservas'
      credentials = st.secrets['sheets']['credentials_sheet']
      time_zone = 'GMT-05:00' # 'South America'
  
      idcalendar = "josegarjagt@gmail.com"
      idcalendar1 = "ba6000578facbb0df71389d7c4b76555afe42838fc641f417d7cc3a91bf86b7f@group.calendar.google.com"
      idcalendar2 = "ebd1ed404fd7c23d10dc5ee277f470bc28942fdd19879aec16a61add832897f2@group.  calendar.google.com"
      idcalendar3 = "cba31f328b4ec714869a2becbecd214ddf68c8bf74cc0aac984fc0f9c5ddb36f@group.calendar.google.com"
      idcalendar4 = "ed26cef01eaf4ba394d77252c097fc4da150a2624db87bf96b25f900483201bd@group.calendar.google.com"
      idcalendar5 = "afe8f37afe5e919952f780062fd693d57680899895c7dbad0878f2dbf8c91be2@group.calendar.google.com"
      idcalendar6 = "ac9164e4adb34328aafee5b87fe610a44a616df06944a585d64060f4b29ee8b5@group.calendar.google.com"
      idcalendar7 = "3e434060d0f95a7a15966031efc94bd591fd1b37d470228a3f66d9158ab9e83a@group.calendar.google.com"
      idcalendar8 = "9e5cb9aeb1d7ee6500e2b9a02118270b7c14b97adba9860492656d69ab5d9f54@group.calendar.google.com"
      idcalendar9 = "790891854a9130f51711b964bdd84024cbb5e6fdc0cbcb479cae3248dabdfc9b@group.calendar.google.com"
      idcalendar10 = "1e87a215fbe5b4350a345065b46f886015205618e5350077e847392f6f29af8b@group.calendar.google.com"
      idcalendar11 = "family03761753017923947946@group.calendar.google.com"
      idcalendar12 = "addressbook#contacts@group.v.calendar.google.com"
      idcalendar13 = "es.co#holiday@group.v.calendar.google.com"
      
      #st.subheader('Reservar')
      
      c1, c2 = st.columns(2)
      nombre = c1.text_input('Nombre entidad o persona*: ', placeholder='Nombre') # label_visibility='hidden')
      email  = c2.text_input('Email:', placeholder='Email')
      fecha  = c1.date_input('Fecha*: ')
      servicio = c1.selectbox('Servicio*: ',result_serv)
      notas = c1.text_area('Nota o Mensaje(Opcional)')
    
      if fecha:
        if servicio == "Peluqueria":
          id = idcalendar5
        elif servicio == "Mercado":
          id = idcalendar1
        elif servicio == "Cita Medica":
          id = idcalendar2
        elif servicio == "Banco":
          id = idcalendar3
        elif servicio == "Taller":
          id = idcalendar4
        elif servicio == "Ferreteria":
          id = idcalendar6
        elif servicio == "Juzgado":
          id = idcalendar7
        elif servicio == "Impuestos":
          id = idcalendar8
        elif servicio == "Iglesia":
          id = idcalendar9
        elif servicio == "Celebracion":
          id = idcalendar10
        elif servicio == "Familia":
          id = idcalendar11
        elif servicio == "Cumpleaños":
          id = idcalendar12
        elif servicio == "Festivos en Colombia":
          id = idcalendar12
        
      calendar = GoogleCalendar(id) #credentials, idcalendar
        
      encargado = c2.selectbox('Encargado:',result_estil)
      #hora = c2.selectbox('Hora: ',horas)
      
      if encargado == "Jose Alfaro":
        email_encargado = "josealfaro@gmail.com"
      elif encargado ==  "Andres Garcia":
        email_encargado = "andresgarcia@gmail.com"
      elif encargado ==  "Mario Vargas":
        email_encargado = "mariovargas@gmail.com"
      elif encargado ==  "Stella Gomez":
        email_encargado = "stellagomez@gmail.com"
      else:
        email_encargado = "otroemail@gmail.com"
    
      attendees = [email_encargado, email]
      
      hours_blocked = calendar.list_upcoming_events()
      result_hours = np.setdiff1d(horas, hours_blocked) 
      hora = c2.selectbox('Hora: ',result_hours)
    
      whatsapp = c2.checkbox('Envio a WhatsApp Si/No (Opcional)')
      telefono = c2.text_input('Nro. Telefono')

      enviar = st.form_submit_button(" Reservar ")

      #Backend
      if enviar:
        with st.spinner('Cargando...'):
          if not nombre or not servicio or not encargado:
            st.warning('Se Require completar los campos con * son obligatorios')
        
          elif not validate_email(email):
            st.warning('El email no es valido')
        
          else:
          
            parsed_time = dt.datetime.strptime(hora, "%H:%M").time()
            hours1 = parsed_time.hour
            minutes1 =  parsed_time.minute
                    
            end_hours = add_hour_and_half(hora)
          
            parsed_time2 = dt.datetime.strptime(end_hours, "%H:%M").time()
            hours2 = parsed_time2.hour
            minutes2 =  parsed_time2.minute

            start_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours1-5,minutes1).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
                             
            end_time = dt.datetime(fecha.year, fecha.month, fecha.day, hours2-5,minutes2).astimezone(dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S')
          
            gs = GoogleSheet(credentials, document, sheet)
                   
            last_row = len(gs.sheet.get_all_values()) +1
            #print(f'last_row {last_row}')
            data = gs.sheet.get_values()
            #print(f'data {data}')
            data2 = data[1:]
            #print(f'data2 {data2}')
            range_start = f"A{last_row}"
            #print(f'range_start {range_start}')
            range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
            #print(f'range_end {range_end}')
          
            for row in data2:
        
              nom = [row[0]]
              serv = [row[4]]
              fech = str(row[2])
              hora2 = str(row[3])
              uid1 = str(row[7])

              if nom != ['DATA']:
              
                fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
                fech1 = int(fech2.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))
                hora3 = datetime.datetime.strptime(hora2,'%H:%M')
                fechahora_ini = int(hora3.strftime('%H%M'))
                horacalendar = datetime.datetime.strptime(hora,'%H:%M')
                horacalendarint = int(horacalendar.strftime('%H%M'))
                #horawhat = int(horacalendar.strftime('%H'))
                #minuto = int(horacalendar.strftime('%M'))
              
                #print(f'Fechas y horas {fech1}, {fechacalendarint}, {fechahora_ini}, { horacalendarint}')
              
                if nom == [nombre] and fech1 == fechacalendarint and fechahora_ini == horacalendarint:
                  st.warning("El cliente ya tiene agenda para esa fecha y hora")
                  break
            
              if (nom == [nombre] or nom != [nombre]): #and serv != [servicio] and fech1 >=   fechacalendarint and fechahora_ini != horacalendarint:
                       
                hora_actual = dt.datetime.now()
                hora_actual_int = int(hora_actual.strftime("%H%M"))
              
                hora_calendar = datetime.datetime.strptime(hora,'%H:%M')
                #print(hora_actual)
                hora_calendar_int = int(hora_calendar.strftime('%H%M'))
          
                hoy = dt.datetime.now()
                fechoy = int(hoy.strftime("%Y%m%d"))
                fechacalendarint = int(fecha.strftime("%Y%m%d"))
          
                if fechacalendarint >= fechoy:
                   
                  if fechacalendarint == fechoy and hora_actual_int < hora_calendar_int:
            
                    st.warning('La hora seleccionda es invalida para hoy')
                    print('La hora seleccionda es invalida para hoy')
                   
                  else:
                      
                    uid = generate_uid()
                    values = [(nombre,email,str(fecha),hora,servicio,encargado, notas, uid, whatsapp, telefono)]
                    gs = GoogleSheet(credentials, document, sheet)
          
                    range = gs.get_last_row_range()
                    gs.write_data(range,values)
                     
                    calendar.create_event(servicio+". "+nombre, start_time, end_time, time_zone, attendees=attendees)

                    #if whatsapp == True:
                    #  contact = str(57)+telefono
                    ##  message = f'Cordial saludo: Sr(a): {nombre}, su soliciud de reserva se ha realizado de forma exitosa y se agendo el servicio {servicio}, para el dia {fecha}, a las {hora}. Gracias por su cofianza. Cordialmente aplicacion de Reservas y Agendamiento.'
                      # Para abrir en buscador Edge                  
                    #  enviarwhats = webbrowser.open('https://web.whatsapp.com/send?phone='+contact+"&text="+message)
                      # Para abrir en buscador chrome
                      #chrome_path = 'C:/Program Files(x86)/Google/Chrome/Application/chrome.exe %s'
                      #webbrowser.get(chrome_path).open('https://web.whatsapp.com/send?phone='+contact+"&text="+message)
                  
                    #  sleep(130)
                    #  pyautogui.click(1230.964)
                    #  sleep(5)
                      #pyautogui.typewrite(message)
                    #  pyautogui.press('enter')
                    #  sleep(2)
                    #  pyautogui.hotkey('ctrl','w')
                    #  sleep(1)
                
                      #pywhatkit.sendwhatmsg('+'+str(57)+contact, message, horawhat, minuto)
                  
                    st.success('Su solicitud ha sido reservada de forrma exitosa')
                    send_email2(email, nombre, fecha, hora, servicio, encargado,  notas)
                    #send_email_emp(email, nombre, fecha, hora, servicio, encargado,  notas)  
                    break
