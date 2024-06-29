import streamlit as st
from google_sheets import GoogleSheet
from google_calendar import GoogleCalendar
from sendemail import send_email2
from sendemail_empresa import send_email_emp
import numpy as np
import datetime as dt
import datetime
import re
from openpyxl import load_workbook

datos_book = load_workbook("archivos/parametros_empresa.xlsx", read_only=False) 

def dataBook(hoja):
  ws1 = datos_book[hoja]

  data = []
  for row in range(1,ws1.max_row):
    _row=[]
    for col in ws1.iter_cols(1,ws1.max_column):
      _row.append(col[row].value)
    
    data.append(_row)
  #print(f'data {data}')
  return data

def validate_email(email):
  pattern = re.compile('^[\w\.-]+@[\w\.-]+\.\w+$')
  if re.match(pattern, email):
    return True
  else:
    return False
  
class EliminarReservaEmp:
  
  class Model:
    pageTitle = "***Eliminar Reserva***"
  
  def view(self,model):
    st.title(model.pageTitle)
  
    with st.form(key='myform3',clear_on_submit=True):
  
      horas = dataBook("horario")
      servicio = dataBook("servicio")
      result_serv = np.setdiff1d(servicio,'')       
      encargado = dataBook("encargado")
      result_estil = np.setdiff1d(encargado,'X')
    
      document='gestion-reservas'
      sheet = 'reservas'
      credentials = st.secrets['sheets']['credentials_sheet']
      time_zone = 'GMT-05:00' # 'South America'
  
      idcalendar = "josegarjagt@gmail.com"
      idcalendar1 = "617d2384ffee7bf87962b771b740e372fc7d8b1e1bd5db386d3c38b0dbf0bce5@group.calendar.google.com"
      idcalendar2 = "28907ee879da61f67b82a7af31f161bddc00760d22f951a6691994451038b7d7@group.calendar.google.com"
      idcalendar3 = "228e2b2270fa33d0e9708f40ef5f23ac1315a7f7ddddaf262da22562c141d3e4@group.calendar.google.com"
      idcalendar4 = "171e4f700155cd208306f2de463285ccd9481949193e55b88db4b0241fedbb6d@group.calendar.google.com"
      
      st.subheader('Eliminar Reserva')
    
      result_hours = np.setdiff1d(horas, "00:00") 
        
      c1, c2 = st.columns(2)
      nombre = c1.text_input('Nombre entidad o persona*: ', placeholder='Nombre') # label_visibility='hidden')
      email  = c2.text_input('Su Email*:', placeholder='Email')
      fecha  = c1.date_input('Fecha*: ')
      servicio = c1.selectbox('servicio: ',result_serv)
      hora = c2.selectbox('Hora: ',result_hours)
        
      if fecha:
        if servicio == "Corte Hombre":
          id = idcalendar1
        elif servicio == "Corte Mujer":
          id = idcalendar2
        elif servicio == "Arreglo Barba":
          id = idcalendar3
        elif servicio == "Afeitar":
          id = idcalendar4
        elif servicio == "Otro":
          id = idcalendar4
            
      encargado = c2.selectbox('Encargado',result_estil)
      #hora = c2.selectbox('Hora: ',horas)
     
      calendar = GoogleCalendar(id) #credentials, idcalendar
      
      if encargado == "Jose Alfaro":
        email_encargado = "josealfaro@gmail.com"
      elif encargado ==  "Andres Garcia":
        email_encargado = "andresgarcia@gmail.com"
      elif encargado ==  "Mario Vargas":
        email_encargado = "mariovargas@gmail.com"
      elif encargado ==  "Stella Gomez":
        email_encargado = "stellagomez@gmail.com"
      else:
        email_encargado = "otroemail@gmail.com"
    
      attendees = [email_encargado, email]
         
      eliminar = st.form_submit_button('Eliminar')

      if eliminar:
            
       with st.spinner('Cargando...'):
        if not nombre or not email or not servicio or not encargado:
          st.warning('Se Require completar los campos con * son obligatorios')
        
        elif not validate_email(email):
          st.warning('El email no es valido')
        
          gs = GoogleSheet(credentials, document, sheet)
                   
          last_row = len(gs.sheet.get_all_values()) +1
          data = gs.sheet.get_values()
          data2 = data[1:]
          range_start = f"A{last_row}"
          range_end = f"{chr(ord('A')+len(data[0])-1)}{last_row}"
         
          for row in data2:
            nom = [row[0]]
            serv = [row[4]]
            fech = str(row[2])
            hora2 = str(row[3])
            nota = [row[6]]
            uid1 = str(row[7])

            if nom != ['DATA']:
              
              fech2 = datetime.datetime.strptime(fech,'%Y-%m-%d')
              fech1 = int(fech2.strftime("%Y%m%d"))
              fechacalendarint = int(fecha.strftime("%Y%m%d"))
              hora3 = datetime.datetime.strptime(hora2,'%H:%M')
              fechahora_ini = int(hora3.strftime('%H%M'))
              horacalendar = datetime.datetime.strptime(hora,'%H:%M')
              horacalendarint = int(horacalendar.strftime('%H%M'))
              
              #print(f'nombre fechas y servicio {nom}, {serv}, {fech1}, {fechacalendarint}, {fechahora_ini}, {horacalendarint}, {nota}')
              
              if nom == [nombre] and serv == [servicio] and fech1 == fechacalendarint and fechahora_ini == horacalendarint and nota != ["Agenda Cancelada"]:
             
                uid = str(uid1)
                values = [(nombre,email,str(fecha),str(horacalendarint),servicio,encargado, "Agenda Cancelada", uid,"False")]
                  
                gs = GoogleSheet(credentials, document, sheet)
                range = gs.write_data_by_uid(uid, values)

                calendar.delete_event()
                                          
                send_email2(email, nombre, fecha, hora3, servicio, encargado,  notas='De acuerdo con su solicitud se cancelo la reserva. Gracias por su atencion.')
                send_email_emp(email, nombre, fecha, hora, servicio, encargado, notas='De acuerdo con su solicitud se cancelo la reserva. Gracias por su atencion.')
                
                st.success('Su solicitud ha sido actualizada de forrma exitosa')
                                    
            if nom == [nombre] and serv == [servicio] and fech1 == fechacalendarint and (fechahora_ini != horacalendarint or nota == 'Agenda Cancelada'):  
                st.warning('El cliente No tiene agenda o esta vencida o cancelda verifique su correo.')
                print('El cliente No tiene agenda o esta vencida verifique su correo.')
                break
